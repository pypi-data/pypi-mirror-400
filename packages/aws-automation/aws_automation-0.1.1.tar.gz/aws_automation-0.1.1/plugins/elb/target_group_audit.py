"""
plugins/elb/target_group_audit.py - Target Group 미사용 분석

ELB에 연결되지 않은 Target Group 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticloadbalancing:DescribeTargetGroups",
        "elasticloadbalancing:DescribeTargetHealth",
    ],
}


class TargetGroupStatus(Enum):
    """Target Group 상태"""

    NORMAL = "normal"
    UNATTACHED = "unattached"  # LB에 연결 안 됨
    NO_TARGETS = "no_targets"  # 타겟 없음
    ALL_UNHEALTHY = "all_unhealthy"  # 모든 타겟 비정상


@dataclass
class TargetGroupInfo:
    """Target Group 정보"""

    account_id: str
    account_name: str
    region: str
    arn: str
    name: str
    protocol: str | None
    port: int | None
    target_type: str  # instance, ip, lambda, alb
    vpc_id: str | None
    load_balancer_arns: list[str] = field(default_factory=list)
    total_targets: int = 0
    healthy_targets: int = 0
    unhealthy_targets: int = 0

    @property
    def is_attached(self) -> bool:
        return len(self.load_balancer_arns) > 0


@dataclass
class TargetGroupFinding:
    """Target Group 분석 결과"""

    tg: TargetGroupInfo
    status: TargetGroupStatus
    recommendation: str


@dataclass
class TargetGroupAnalysisResult:
    """Target Group 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unattached_count: int = 0
    no_targets_count: int = 0
    unhealthy_count: int = 0
    normal_count: int = 0
    findings: list[TargetGroupFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_target_groups(session, account_id: str, account_name: str, region: str) -> list[TargetGroupInfo]:
    """Target Groups 수집"""
    from botocore.exceptions import ClientError

    elbv2 = get_client(session, "elbv2", region_name=region)
    target_groups = []

    paginator = elbv2.get_paginator("describe_target_groups")
    for page in paginator.paginate():
        for tg in page.get("TargetGroups", []):
            tg_info = TargetGroupInfo(
                account_id=account_id,
                account_name=account_name,
                region=region,
                arn=tg["TargetGroupArn"],
                name=tg["TargetGroupName"],
                protocol=tg.get("Protocol"),
                port=tg.get("Port"),
                target_type=tg.get("TargetType", "instance"),
                vpc_id=tg.get("VpcId"),
                load_balancer_arns=tg.get("LoadBalancerArns", []),
            )

            # 타겟 상태 조회
            try:
                health_resp = elbv2.describe_target_health(TargetGroupArn=tg["TargetGroupArn"])
                targets = health_resp.get("TargetHealthDescriptions", [])
                tg_info.total_targets = len(targets)
                tg_info.healthy_targets = sum(1 for t in targets if t.get("TargetHealth", {}).get("State") == "healthy")
                tg_info.unhealthy_targets = tg_info.total_targets - tg_info.healthy_targets
            except ClientError:
                pass

            target_groups.append(tg_info)

    return target_groups


# =============================================================================
# 분석
# =============================================================================


def analyze_target_groups(
    target_groups: list[TargetGroupInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> TargetGroupAnalysisResult:
    """Target Group 분석"""
    result = TargetGroupAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(target_groups),
    )

    for tg in target_groups:
        # LB에 연결 안 됨
        if not tg.is_attached:
            result.unattached_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.UNATTACHED,
                    recommendation="로드밸런서에 연결되지 않음 - 삭제 검토",
                )
            )
            continue

        # 타겟 없음
        if tg.total_targets == 0:
            result.no_targets_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.NO_TARGETS,
                    recommendation="등록된 타겟 없음 - 타겟 등록 또는 삭제 검토",
                )
            )
            continue

        # 모든 타겟 비정상
        if tg.healthy_targets == 0 and tg.total_targets > 0:
            result.unhealthy_count += 1
            result.findings.append(
                TargetGroupFinding(
                    tg=tg,
                    status=TargetGroupStatus.ALL_UNHEALTHY,
                    recommendation="모든 타겟 비정상 - 헬스체크 확인 필요",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            TargetGroupFinding(
                tg=tg,
                status=TargetGroupStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[TargetGroupAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Target Group 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = ["Account", "Region", "전체", "미연결", "타겟 없음", "비정상", "정상"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_count)
        ws.cell(row=row, column=4, value=r.unattached_count)
        ws.cell(row=row, column=5, value=r.no_targets_count)
        ws.cell(row=row, column=6, value=r.unhealthy_count)
        ws.cell(row=row, column=7, value=r.normal_count)

        if r.unattached_count > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.no_targets_count > 0:
            ws.cell(row=row, column=5).fill = yellow_fill
        if r.unhealthy_count > 0:
            ws.cell(row=row, column=6).fill = red_fill

    # 상세 시트
    ws_detail = wb.create_sheet("Target Groups")
    detail_headers = [
        "Account",
        "Region",
        "Name",
        "상태",
        "Type",
        "Protocol",
        "Port",
        "LB 연결",
        "Total Targets",
        "Healthy",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != TargetGroupStatus.NORMAL:
                detail_row += 1
                tg = f.tg
                ws_detail.cell(row=detail_row, column=1, value=tg.account_name)
                ws_detail.cell(row=detail_row, column=2, value=tg.region)
                ws_detail.cell(row=detail_row, column=3, value=tg.name)
                ws_detail.cell(row=detail_row, column=4, value=f.status.value)
                ws_detail.cell(row=detail_row, column=5, value=tg.target_type)
                ws_detail.cell(row=detail_row, column=6, value=tg.protocol or "-")
                ws_detail.cell(row=detail_row, column=7, value=tg.port or "-")
                ws_detail.cell(row=detail_row, column=8, value=len(tg.load_balancer_arns))
                ws_detail.cell(row=detail_row, column=9, value=tg.total_targets)
                ws_detail.cell(row=detail_row, column=10, value=tg.healthy_targets)
                ws_detail.cell(row=detail_row, column=11, value=f.recommendation)

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"TargetGroup_Audit_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> TargetGroupAnalysisResult | None:
    """단일 계정/리전의 Target Group 수집 및 분석 (병렬 실행용)"""
    target_groups = collect_target_groups(session, account_id, account_name, region)
    if not target_groups:
        return None
    return analyze_target_groups(target_groups, account_id, account_name, region)


def run(ctx) -> None:
    """Target Group 미사용 분석"""
    console.print("[bold]Target Group 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elbv2")
    results: list[TargetGroupAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_unattached = sum(r.unattached_count for r in results)
    total_no_targets = sum(r.no_targets_count for r in results)
    total_unhealthy = sum(r.unhealthy_count for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미연결: [red]{total_unattached}개[/red]")
    console.print(f"타겟 없음: [yellow]{total_no_targets}개[/yellow]")
    console.print(f"전체 비정상: [red]{total_unhealthy}개[/red]")

    # 보고서 생성
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("targetgroup-audit").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
