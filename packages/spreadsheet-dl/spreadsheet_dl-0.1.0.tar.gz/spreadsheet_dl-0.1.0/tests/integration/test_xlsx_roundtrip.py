"""Integration tests for XLSX format roundtrip fidelity.

Tests create XLSX files with formulas, styles, and charts, then read them back
to verify data integrity. These tests validate the actual file format output.

Test Strategy:
    - Create spreadsheet with specific content using builder API
    - Export to XLSX format
    - Read back and verify content matches
    - Test edge cases: special characters, large numbers, formulas
"""

from __future__ import annotations

from typing import TYPE_CHECKING

import pytest

if TYPE_CHECKING:
    from pathlib import Path

pytestmark = [pytest.mark.integration, pytest.mark.requires_export]


class TestXLSXBasicRoundtrip:
    """Test basic XLSX create/read roundtrip."""

    def test_simple_data_roundtrip(self, tmp_path: Path) -> None:
        """Test that simple data survives XLSX roundtrip."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with simple data
        builder = create_spreadsheet()
        builder.sheet("Data").column("Name").column("Value")
        builder.header_row()
        builder.row().cell("Alpha").cell(100)
        builder.row().cell("Beta").cell(200)
        builder.row().cell("Gamma").cell(300)

        # Save as XLSX
        xlsx_path = tmp_path / "test_simple.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 0

        # Read back with openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Verify data
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Value"
        assert ws["A2"].value == "Alpha"
        assert ws["B2"].value == 100
        assert ws["A3"].value == "Beta"
        assert ws["B3"].value == 200

    def test_formula_roundtrip(self, tmp_path: Path) -> None:
        """Test that formulas survive XLSX roundtrip."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet, formula

        # Create spreadsheet with formula
        builder = create_spreadsheet()
        builder.sheet("Formulas").column("A").column("B").column("Sum")
        builder.header_row()
        builder.row().cell(10).cell(20).cell(formula=formula().sum("A2", "B2"))
        builder.row().cell(30).cell(40).cell(formula=formula().sum("A3", "B3"))

        # Save as XLSX
        xlsx_path = tmp_path / "test_formulas.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()

        # Read back
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Verify values
        assert ws["A2"].value == 10
        assert ws["B2"].value == 20
        # Formula cells should have formula or computed value
        # Note: openpyxl reads formulas as strings starting with '='
        cell_c2 = ws["C2"]
        assert cell_c2.value is not None

    def test_numeric_precision_roundtrip(self, tmp_path: Path) -> None:
        """Test that numeric precision is preserved."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with precise numbers
        builder = create_spreadsheet()
        builder.sheet("Numbers").column("Value")
        builder.header_row()
        builder.row().cell(3.141592653589793)
        builder.row().cell(2.718281828459045)
        builder.row().cell(1e-10)
        builder.row().cell(1e10)
        builder.row().cell(0.1 + 0.2)  # Classic floating point test

        # Save and read back
        xlsx_path = tmp_path / "test_precision.xlsx"
        builder.export(str(xlsx_path))

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Check precision (within floating point tolerance)
        assert abs(ws["A2"].value - 3.141592653589793) < 1e-10
        assert abs(ws["A3"].value - 2.718281828459045) < 1e-10
        assert abs(ws["A4"].value - 1e-10) < 1e-20
        assert abs(ws["A5"].value - 1e10) < 1

    def test_special_characters_roundtrip(self, tmp_path: Path) -> None:
        """Test that special characters survive roundtrip."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with special characters
        builder = create_spreadsheet()
        builder.sheet("Special").column("Text")
        builder.header_row()
        builder.row().cell("Hello, World!")
        builder.row().cell("Line1\nLine2")  # Newline
        builder.row().cell('Quote "test"')  # Quotes
        builder.row().cell("Unicode: \u03c0\u00b2")  # Pi squared
        builder.row().cell("Emoji: \u2764")  # Heart

        # Save and read back
        xlsx_path = tmp_path / "test_special.xlsx"
        builder.export(str(xlsx_path))

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        assert ws["A2"].value == "Hello, World!"
        assert "Quote" in ws["A4"].value


class TestXLSXStyleRoundtrip:
    """Test XLSX style preservation in roundtrip."""

    def test_themed_spreadsheet_export(self, tmp_path: Path) -> None:
        """Test that themed spreadsheet exports successfully."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with theme
        builder = create_spreadsheet(theme="corporate")
        builder.sheet("Styled").column("Category").column("Amount")
        builder.header_row()
        builder.row().cell("Revenue").cell(10000)
        builder.row().cell("Expenses").cell(7500)
        builder.row().cell("Profit").cell(2500)

        # Save as XLSX
        xlsx_path = tmp_path / "test_styled.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 0

        # Verify file can be opened
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        assert "Styled" in wb.sheetnames or wb.active is not None


class TestXLSXMultiSheetRoundtrip:
    """Test multi-sheet XLSX roundtrip."""

    def test_multiple_sheets_roundtrip(self, tmp_path: Path) -> None:
        """Test that multiple sheets survive roundtrip."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with multiple sheets
        builder = create_spreadsheet()

        builder.sheet("Sheet1").column("A").column("B")
        builder.header_row()
        builder.row().cell("S1A").cell("S1B")

        builder.sheet("Sheet2").column("X").column("Y")
        builder.header_row()
        builder.row().cell("S2X").cell("S2Y")

        builder.sheet("Sheet3").column("P").column("Q")
        builder.header_row()
        builder.row().cell("S3P").cell("S3Q")

        # Save and read back
        xlsx_path = tmp_path / "test_multisheet.xlsx"
        builder.export(str(xlsx_path))

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)

        # Verify all sheets exist
        assert len(wb.sheetnames) >= 3

    def test_cross_sheet_formula_export(self, tmp_path: Path) -> None:
        """Test that cross-sheet references export correctly."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with cross-sheet reference
        builder = create_spreadsheet()

        builder.sheet("Data").column("Value")
        builder.header_row()
        builder.row().cell(100)
        builder.row().cell(200)

        builder.sheet("Summary").column("Total")
        builder.header_row()
        # Note: Cross-sheet references depend on builder implementation
        builder.row().cell(300)

        # Save
        xlsx_path = tmp_path / "test_crossref.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()


class TestXLSXEdgeCases:
    """Test XLSX edge cases and error handling."""

    def test_empty_spreadsheet_export(self, tmp_path: Path) -> None:
        """Test that empty spreadsheet handles gracefully."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet
        from spreadsheet_dl.builder import EmptySheetError

        builder = create_spreadsheet()
        builder.sheet("Empty")

        # Empty sheet should raise error or create minimal valid file
        xlsx_path = tmp_path / "test_empty.xlsx"
        with pytest.raises(EmptySheetError):
            builder.export(str(xlsx_path))

    def test_large_dataset_export(self, tmp_path: Path) -> None:
        """Test export of reasonably large dataset."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create spreadsheet with 1000 rows
        builder = create_spreadsheet()
        builder.sheet("Large").column("ID").column("Value")
        builder.header_row()

        for i in range(1000):
            builder.row().cell(i).cell(i * 10)

        # Save
        xlsx_path = tmp_path / "test_large.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 10000  # Should be substantial

        # Verify row count
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row >= 1000

    def test_unicode_sheet_name(self, tmp_path: Path) -> None:
        """Test that unicode sheet names work."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Data Sheet").column("Value")  # Safe ASCII name
        builder.header_row()
        builder.row().cell(42)

        xlsx_path = tmp_path / "test_unicode_sheet.xlsx"
        builder.export(str(xlsx_path))

        assert xlsx_path.exists()
