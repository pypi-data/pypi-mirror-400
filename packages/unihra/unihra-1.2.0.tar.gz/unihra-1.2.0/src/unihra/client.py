import json
import requests
import pandas as pd
from typing import List, Generator, Dict, Any, Literal, Optional
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
try:
    from tqdm.auto import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

from .exceptions import (
    UnihraError, UnihraApiError, UnihraConnectionError, 
    UnihraValidationError, UnihraDependencyError, raise_for_error_code
)

BASE_URL = "https://unihra.ru"
ACTION_MAP = {
    "Добавить": "add",
    "Увеличить": "increase",
    "Уменьшить": "decrease",
    "Ок": "ok",
    "Ничего не делать": "ok"
}

class UnihraClient:
    """
    Official Python Client for Unihra API.
    """

    def __init__(self, api_key: str, base_url: str = BASE_URL, max_retries: int = 0):
        self.base_url = base_url.rstrip('/')
        self.api_v1 = f"{self.base_url}/api/v1"
        self.session = requests.Session()
        
        self.session.headers.update({
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "UnihraPythonSDK/1.6.0"
        })

        if max_retries > 0:
            retry_strategy = Retry(
                total=max_retries,
                backoff_factor=2,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["POST", "GET"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            self.session.mount("https://", adapter)
            self.session.mount("http://", adapter)

    def health(self) -> Dict[str, Any]:
        try:
            resp = self.session.get(f"{self.api_v1}/health")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Health check failed: {e}")

    def get_page_structure(self, task_id: str) -> List[Dict[str, Any]]:
        """
        Fetch detailed page structure list (Own Page + Competitors).
        """
        try:
            resp = self.session.get(f"{self.api_v1}/report/structure/{task_id}")
            resp.raise_for_status()
            data = resp.json()
            
            # API returns a List of objects. Normalize each one.
            if isinstance(data, list):
                return [self._normalize_keys(item) for item in data]
            return []
        except requests.exceptions.RequestException as e:
            return []

    def analyze(
        self, 
        own_page: str, 
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: Literal['ru', 'en'] = 'ru',
        verbose: bool = False
    ) -> Dict[str, Any]:
        last_event = {}
        pbar = None
        if verbose:
            if TQDM_AVAILABLE:
                pbar = tqdm(total=100, desc="Analyzing SEO", unit="%")
            else:
                print("Note: Install 'tqdm' to see a visual progress bar.")

        try:
            for event in self.analyze_stream(own_page, competitors, queries, lang):
                last_event = event
                
                if pbar:
                    state = event.get("state")
                    progress = event.get("progress", 0)
                    
                    if isinstance(progress, (int, float)):
                        pbar.n = int(progress)
                        pbar.refresh()
                    
                    if state == "PROCESSING" or state == "PROGRESS":
                        msg = "Processing"
                        details = event.get("details", {})
                        if isinstance(details, dict) and "message" in details:
                            msg = details["message"][:40]
                        pbar.set_description(f"{msg}")
                    elif state == "SUCCESS":
                        pbar.set_description("Completed ✅")
                        pbar.n = 100
                        pbar.refresh()

                if event.get("state") == "SUCCESS":
                    return event.get("result", {})
                    
        except Exception as e:
            if pbar: 
                pbar.set_description("Failed ❌")
                pbar.close()
            raise e
        finally:
            if pbar: 
                pbar.close()
        
        return last_event

    def analyze_stream(
        self, 
        own_page: str, 
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: str = 'ru'
    ) -> Generator[Dict, None, None]:
        if not competitors:
            raise UnihraValidationError("Competitor list cannot be empty.")

        payload = {
            "own_page": own_page, 
            "competitor_urls": competitors,
            "queries": queries or [],
            "lang": lang
        }

        try:
            resp = self.session.post(f"{self.api_v1}/process", json=payload)
            
            if resp.status_code == 401:
                raise UnihraApiError("Invalid API Key or unauthorized access", code=401)
            resp.raise_for_status()
            
            task_id = resp.json().get("task_id")
            if not task_id:
                raise UnihraApiError("API response missing 'task_id'")

            stream_url = f"{self.api_v1}/process/status/{task_id}"
            
            with self.session.get(stream_url, stream=True) as s_resp:
                s_resp.raise_for_status()
                
                for line in s_resp.iter_lines():
                    if not line: 
                        continue
                    
                    if line.startswith(b'data: '):
                        try:
                            decoded_line = line[6:].decode('utf-8')
                            data = json.loads(decoded_line)
                            state = data.get("state")
                            
                            if state == "FAILURE":
                                error_obj = data.get("error")
                                if isinstance(error_obj, dict):
                                    code = error_obj.get("code", 9999)
                                    msg = error_obj.get("message", "Unknown error")
                                else:
                                    code = data.get("error_code", 9999)
                                    msg = data.get("message", "Unknown error")

                                raise_for_error_code(code, msg, data)
                            
                            if state == "SUCCESS":
                                raw_result = data.get("result", {})
                                normalized_result = self._normalize_keys(raw_result)
                                if lang == 'en':
                                    final_result = self._translate_action_values(normalized_result)
                                else:
                                    final_result = normalized_result

                                # Fetch list of structures
                                structure_data = self.get_page_structure(task_id)
                                if structure_data:
                                    final_result['page_structure'] = structure_data

                                data["result"] = final_result
                                yield data
                                break
                            
                            yield data
                                
                        except json.JSONDecodeError:
                            continue
                            
        except requests.exceptions.RetryError:
            raise UnihraConnectionError("Max retries exceeded. The service might be temporarily unavailable.")
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Network error: {e}")

    def _normalize_keys(self, data: Dict[str, Any]) -> Dict[str, Any]:
        new_data = {}
        if not isinstance(data, dict):
            return data
            
        for key, value in data.items():
            new_key = key.lower().replace(" ", "_").replace("-", "_")
            new_data[new_key] = value
        return new_data

    def _translate_action_values(self, result: Dict[str, Any]) -> Dict[str, Any]:
        if "block_comparison" in result and isinstance(result["block_comparison"], list):
            for item in result["block_comparison"]:
                if "action_needed" in item:
                    russian_action = item["action_needed"]
                    item["action_needed"] = ACTION_MAP.get(russian_action, russian_action)
        return result

    def _flatten_structure_list(self, structure_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Flatten list of structures for DataFrame (Table view).
        """
        flat_rows = []
        for item in structure_list:
            flat_item = {'url': item.get('url')}
            
            # Metrics
            if 'metrics' in item:
                for k, v in item['metrics'].items():
                    flat_item[k] = v
            # Content
            if 'content' in item:
                for k, v in item['content'].items():
                    flat_item[k] = v
            # Meta Tags
            if 'meta_tags' in item:
                for k, v in item['meta_tags'].items():
                    flat_item[k] = v
            
            flat_rows.append(flat_item)
        return flat_rows

    def get_dataframe(self, result: Dict[str, Any], section: str = "block_comparison") -> pd.DataFrame:
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is not installed. Run: pip install pandas")

        normalized_section = section.lower().replace(" ", "_").replace("-", "_")
        
        if normalized_section == "page_structure":
            data = result.get("page_structure", [])
            if not data:
                return pd.DataFrame()
            flat_list = self._flatten_structure_list(data)
            return pd.DataFrame(flat_list)

        data = result.get(normalized_section, [])
        return pd.DataFrame(data)

    def _reorder_tech_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        tech_cols = ['id', 'block_id', 'analysis_id']
        existing_tech = [c for c in df.columns if c in tech_cols]
        main_cols = [c for c in df.columns if c not in tech_cols]
        
        if not existing_tech:
            return df
            
        return df[main_cols + existing_tech]

    def save_report(self, result: Dict[str, Any], filename: str = "report.xlsx", style_output: bool = True):
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is required. Run: pip install pandas openpyxl")

        df_blocks = pd.DataFrame(result.get("block_comparison", []))
        
        ngrams_data = result.get("ngrams_analysis") or result.get("n_grams_analysis") or []
        df_ngrams = pd.DataFrame(ngrams_data)
        
        gaps_data = result.get("semantic_context_analysis") or result.get("semantic_context_gaps") or []
        df_gaps = pd.DataFrame(gaps_data)
        
        drmaxs_data = result.get("drmaxs", {})
        
        structure_data = result.get("page_structure", [])

        if filename.endswith(".csv"):
            if not df_blocks.empty:
                df_blocks = self._reorder_tech_columns(df_blocks)
            df_blocks.to_csv(filename, index=False, encoding='utf-8-sig')
        else:
            try:
                import openpyxl
            except ImportError:
                raise UnihraDependencyError("Library 'openpyxl' is required for Excel export.")

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 0. Page Structure
                if structure_data:
                    sheet = "Page Structure"
                    flat_struct = self._flatten_structure_list(structure_data)
                    df_struct = pd.DataFrame(flat_struct)
                    
                    # Reorder: URL first
                    cols = df_struct.columns.tolist()
                    if 'url' in cols:
                        cols.insert(0, cols.pop(cols.index('url')))
                        df_struct = df_struct[cols]

                    df_struct.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_struct, sheet_type="structure")

                # 1. Semantic Gaps
                if not df_gaps.empty:
                    sheet = "Semantic Gaps"
                    desired_cols = ['lemma', 'recommendation', 'context_snippet', 'gap', 'coverage_percent', 'competitor_avg_score', 'own_score']
                    existing_cols = [c for c in desired_cols if c in df_gaps.columns]
                    other_cols = [c for c in df_gaps.columns if c not in desired_cols]
                    
                    df_gaps_ordered = df_gaps[existing_cols + other_cols]
                    df_gaps_ordered = self._reorder_tech_columns(df_gaps_ordered)
                    
                    df_gaps_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_gaps_ordered, sheet_type="gaps")

                # 2. Word Analysis
                if not df_blocks.empty:
                    sheet = "Word Analysis"
                    df_blocks_ordered = self._reorder_tech_columns(df_blocks)
                    df_blocks_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_blocks_ordered, sheet_type="word_analysis")
                
                # 3. N-Grams
                if not df_ngrams.empty:
                    sheet = "N-Grams"
                    df_ngrams_ordered = self._reorder_tech_columns(df_ngrams)
                    df_ngrams_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_ngrams_ordered, sheet_type="ngrams")
                
                # 4. Vectors
                if drmaxs_data and isinstance(drmaxs_data, dict):
                    for subkey, subdata in drmaxs_data.items():
                        if subdata and isinstance(subdata, list):
                            df_dr = pd.DataFrame(subdata)
                            df_dr_ordered = self._reorder_tech_columns(df_dr)
                            
                            safe_name = subkey.replace("_", " ").title().replace("By", "")
                            sheet_name = f"Vectors {safe_name}"[:31]
                            df_dr_ordered.to_excel(writer, sheet_name=sheet_name, index=False)
                            if style_output: self._style_worksheet(writer.sheets[sheet_name], df_dr_ordered, sheet_type="vectors")

    def _style_worksheet(self, worksheet, df, sheet_type="generic"):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="363636", end_color="363636", fill_type="solid")
        tech_header_font = Font(bold=True, color="000000") 
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        tech_cols = ['id', 'block_id', 'analysis_id']

        # Format Headers
        for cell in worksheet[1]:
            col_name = str(cell.value) if cell.value else ""
            if col_name in tech_cols:
                cell.font = tech_header_font
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-width
        for idx, col in enumerate(df.columns):
            if col in tech_cols:
                worksheet.column_dimensions[get_column_letter(idx + 1)].hidden = True
                continue
            max_len = max([len(str(s)) for s in df[col].astype(str).values] + [len(col)])
            final_width = min(max_len + 2, 70) 
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = final_width

        col_map = {name: i + 1 for i, name in enumerate(df.columns)}

        if sheet_type == "structure":
            # Wrap text for content fields
            for col_name in ['url', 'h1_heading', 'meta_title', 'meta_description', 'heading_structure_raw']:
                if col_name in col_map:
                    idx = col_map[col_name]
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=idx).alignment = Alignment(wrap_text=True)

        elif sheet_type == "gaps":
            if 'own_score' in col_map and 'lemma' in col_map:
                score_idx = col_map['own_score']
                lemma_idx = col_map['lemma']
                for row in range(2, worksheet.max_row + 1):
                    score_val = worksheet.cell(row=row, column=score_idx).value
                    try:
                        is_missing = float(score_val) == 0 if score_val is not None else True
                    except (ValueError, TypeError):
                        is_missing = True
                    if is_missing:
                        worksheet.cell(row=row, column=lemma_idx).fill = red_fill
                    else:
                        worksheet.cell(row=row, column=lemma_idx).fill = green_fill

        else:
            # Word Analysis, Vectors, etc.
            target_cols = []
            if sheet_type == "word_analysis":
                target_names = ["word", "lemma"]
                target_cols = [col_map[c] for c in target_names if c in col_map]
            elif sheet_type == "ngrams":
                target_names = ["ngram"]
                target_cols = [col_map[c] for c in target_names if c in col_map]
            elif sheet_type == "vectors":
                target_names = ["word"]
                target_cols = [col_map[c] for c in target_names if c in col_map]

            bool_col = 'present_on_own_page'
            if bool_col not in col_map and 'present_in_own' in col_map:
                 bool_col = 'present_in_own'

            if bool_col in col_map and target_cols:
                bool_idx = col_map[bool_col]
                for row in range(2, worksheet.max_row + 1):
                    is_present = worksheet.cell(row=row, column=bool_idx).value
                    fill_color = green_fill if is_present is True else red_fill if is_present is False else None
                    if fill_color:
                        for t_idx in target_cols:
                            worksheet.cell(row=row, column=t_idx).fill = fill_color