"""
Excel写入操作类
提供Excel文件创建、修改和保存功能
"""
import io
import base64
from typing import Dict, Any, Optional, List
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import xlwt
except ImportError:
    xlwt = None

from .utils import (
    base64_to_excel,
    excel_to_base64,
    detect_file_format,
    validate_file_size,
    format_error_response,
    format_success_response,
    parse_cell_address,
    parse_range_address,
    validate_cell_address,
    validate_range_address,
    get_column_letter,
)


class ExcelWriter:
    """Excel文件写入器"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化写入器
        
        Args:
            config: 配置字典
        """
        self.config = config or {}
    
    def _workbook_to_base64(self, workbook) -> str:
        """
        将工作簿转换为Base64编码
        
        Args:
            workbook: openpyxl工作簿对象
            
        Returns:
            Base64编码字符串
        """
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return base64.b64encode(buffer.read()).decode("utf-8")
    
    def create_workbook(self, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        创建新工作簿
        
        Args:
            sheet_name: 默认工作表名称
            
        Returns:
            包含新工作簿Base64编码的响应字典
        """
        try:
            workbook = openpyxl.Workbook()
            workbook.active.title = sheet_name
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"创建工作簿失败：{str(e)}", "CREATE_ERROR")
    
    def create_sheet(
        self,
        file_base64: str,
        sheet_name: str,
        position: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        创建工作表
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet_name: 新工作表名称
            position: 插入位置，可选，默认末尾
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式不支持创建工作表", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet_name in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表已存在：{sheet_name}", "SHEET_EXISTS")
            
            new_sheet = workbook.create_sheet(title=sheet_name)
            
            if position is not None:
                # 移动工作表到指定位置
                workbook.move_sheet(new_sheet, offset=position - len(workbook.sheetnames) + 1)
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"创建工作表失败：{str(e)}", "CREATE_ERROR")
    
    def write_cell_data(
        self,
        file_base64: str,
        sheet: str,
        cell: str,
        value: Any,
        data_type: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        写入单元格数据
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            cell: 单元格地址
            value: 数据值
            data_type: 数据类型，可选：text, number, date, boolean
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式写入功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_cell_address(cell):
                workbook.close()
                return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
            
            try:
                row_idx, col_idx = parse_cell_address(cell)
            except ValueError as e:
                workbook.close()
                return format_error_response(str(e), "INVALID_CELL")
            
            cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
            
            # 根据data_type设置值
            if data_type == "number":
                cell_obj.value = float(value) if value is not None else None
            elif data_type == "date":
                if isinstance(value, str):
                    # 尝试解析日期字符串
                    try:
                        cell_obj.value = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    except:
                        cell_obj.value = value
                else:
                    cell_obj.value = value
            elif data_type == "boolean":
                cell_obj.value = bool(value) if value is not None else None
            else:  # text or None
                cell_obj.value = value
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"写入单元格数据失败：{str(e)}", "WRITE_ERROR")
    
    def write_cell_formula(
        self,
        file_base64: str,
        sheet: str,
        cell: str,
        formula: str
    ) -> Dict[str, Any]:
        """
        写入单元格公式
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            cell: 单元格地址
            formula: 公式文本（应以=开头）
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式写入公式功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_cell_address(cell):
                workbook.close()
                return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
            
            try:
                row_idx, col_idx = parse_cell_address(cell)
            except ValueError as e:
                workbook.close()
                return format_error_response(str(e), "INVALID_CELL")
            
            # 确保公式以=开头
            if not formula.startswith("="):
                formula = "=" + formula
            
            cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
            cell_obj.value = formula
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"写入单元格公式失败：{str(e)}", "WRITE_ERROR")
    
    def write_range_data(
        self,
        file_base64: str,
        sheet: str,
        start_cell: str,
        data: List[List[Any]]
    ) -> Dict[str, Any]:
        """
        批量写入范围数据
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            start_cell: 起始单元格地址
            data: 二维数据数组
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式批量写入功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_cell_address(start_cell):
                workbook.close()
                return format_error_response(f"无效的起始单元格地址：{start_cell}", "INVALID_CELL")
            
            try:
                start_row, start_col = parse_cell_address(start_cell)
            except ValueError as e:
                workbook.close()
                return format_error_response(str(e), "INVALID_CELL")
            
            # 批量写入数据
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    target_sheet.cell(
                        row=start_row + 1 + row_idx,
                        column=start_col + 1 + col_idx,
                        value=cell_value
                    )
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"批量写入数据失败：{str(e)}", "WRITE_ERROR")
    
    def merge_cells(
        self,
        file_base64: str,
        sheet: str,
        range: str
    ) -> Dict[str, Any]:
        """
        合并单元格
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            range: 合并范围，如"A1:B1"
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式合并单元格功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_range_address(range):
                workbook.close()
                return format_error_response(f"无效的范围地址：{range}", "INVALID_RANGE")
            
            target_sheet.merge_cells(range)
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"合并单元格失败：{str(e)}", "WRITE_ERROR")
    
    def unmerge_cells(
        self,
        file_base64: str,
        sheet: str,
        range: str
    ) -> Dict[str, Any]:
        """
        取消合并单元格
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            range: 取消合并的范围
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式取消合并单元格功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_range_address(range):
                workbook.close()
                return format_error_response(f"无效的范围地址：{range}", "INVALID_RANGE")
            
            target_sheet.unmerge_cells(range)
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"取消合并单元格失败：{str(e)}", "WRITE_ERROR")
    
    def create_chart(
        self,
        file_base64: str,
        sheet: str,
        chart_type: str,
        data_range: str,
        title: Optional[str] = None,
        position: Optional[Dict[str, int]] = None
    ) -> Dict[str, Any]:
        """
        创建图表
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            chart_type: 图表类型：bar, line, pie, scatter, area
            data_range: 数据源范围
            title: 图表标题，可选
            position: 图表位置，可选，包含x, y, width, height
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式创建图表功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            # 创建图表对象
            chart_map = {
                "bar": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "scatter": ScatterChart,
                "area": AreaChart
            }
            
            if chart_type not in chart_map:
                workbook.close()
                return format_error_response(
                    f"不支持的图表类型：{chart_type}，支持的类型：{','.join(chart_map.keys())}",
                    "INVALID_CHART_TYPE"
                )
            
            chart = chart_map[chart_type]()
            
            # 设置数据源
            from openpyxl.chart.reference import Reference
            # 解析数据范围
            (start_row, start_col), (end_row, end_col) = parse_range_address(data_range)
            data = Reference(
                target_sheet,
                min_col=start_col + 1,
                min_row=start_row + 1,
                max_col=end_col + 1 if end_col is not None else start_col + 1,
                max_row=end_row + 1 if end_row is not None else start_row + 1
            )
            chart.add_data(data, titles_from_data=True)
            
            # 设置标题
            if title:
                chart.title = title
            
            # 设置位置
            if position:
                chart.anchor = target_sheet.cell(
                    row=position.get("y", 1),
                    column=position.get("x", 1)
                )
            
            # 添加图表到工作表
            target_sheet.add_chart(chart)
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"创建图表失败：{str(e)}", "WRITE_ERROR")
    
    def update_chart(
        self,
        file_base64: str,
        sheet: str,
        chart_id: str,
        chart_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        更新图表
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            chart_id: 图表ID或名称
            chart_config: 图表配置
            
        Returns:
            包含更新后文件的响应字典
        """
        # 简化实现：重新创建图表
        return format_error_response("更新图表功能待完善", "FEATURE_NOT_IMPLEMENTED")
    
    def delete_chart(
        self,
        file_base64: str,
        sheet: str,
        chart_id: str
    ) -> Dict[str, Any]:
        """
        删除图表
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            chart_id: 图表ID或名称
            
        Returns:
            包含更新后文件的响应字典
        """
        # 简化实现
        return format_error_response("删除图表功能待完善", "FEATURE_NOT_IMPLEMENTED")
    
    def create_table(
        self,
        file_base64: str,
        sheet: str,
        range: str,
        table_style: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        创建表格
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            range: 表格范围
            table_style: 表格样式，可选
            
        Returns:
            包含更新后文件的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            return format_error_response(".xls格式创建表格功能待实现", "FORMAT_NOT_SUPPORTED")
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_range_address(range):
                workbook.close()
                return format_error_response(f"无效的范围地址：{range}", "INVALID_RANGE")
            
            # 创建表格
            table = Table(displayName=f"Table{len(target_sheet.tables) + 1}", ref=range)
            
            if table_style:
                style = TableStyleInfo(name=table_style, showFirstColumn=False,
                                     showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                table.tableStyleInfo = style
            
            target_sheet.add_table(table)
            
            file_base64 = self._workbook_to_base64(workbook)
            workbook.close()
            
            return format_success_response({
                "file": file_base64
            })
        except Exception as e:
            return format_error_response(f"创建表格失败：{str(e)}", "WRITE_ERROR")
    
    def save_workbook(
        self,
        file_base64: str,
        options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        保存工作簿
        
        Args:
            file_base64: Excel文件的Base64编码
            options: 保存选项，可选
            
        Returns:
            包含保存后文件的响应字典
        """
        # 实际上文件已经在每次写入操作时保存了
        # 这个方法主要用于应用保存选项（如压缩等）
        return format_success_response({
            "file": file_base64
        })

