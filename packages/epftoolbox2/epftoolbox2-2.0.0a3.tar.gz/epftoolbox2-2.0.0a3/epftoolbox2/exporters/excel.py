from typing import List
from pathlib import Path
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from .base import Exporter
from ..results.report import EvaluationReport


class ExcelExporter(Exporter):
    def __init__(self, path: str, sheets: List[str] = None):
        """
        Args:
            path: Path to the output Excel file.
            sheets: Which sheets to include. Options: "summary", "hour", "horizon",
                    "hour_horizon", "year", "year_horizon", "details". Defaults to all.
        """
        self.path = Path(path)
        self.sheets = sheets or ["summary", "hour", "horizon", "hour_horizon", "year", "year_horizon", "details"]

    def export(self, report: EvaluationReport) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            if "summary" in self.sheets:
                self._write_summary(writer, report)
            if "hour" in self.sheets:
                self._write_pivot_sheet(writer, report, "hour", report.by_hour(), "hour")
            if "horizon" in self.sheets:
                self._write_pivot_sheet(writer, report, "horizon", report.by_horizon(), "horizon")
            if "hour_horizon" in self.sheets:
                self._write_matrix_sheet(writer, report, "HourHorizon", report.by_hour_horizon(), "hour", "horizon")
            if "year" in self.sheets:
                self._write_pivot_sheet(writer, report, "year", report.by_year(), "year")
            if "year_horizon" in self.sheets:
                self._write_matrix_sheet(writer, report, "YearHorizon", report.by_year_horizon(), "year", "horizon")
            if "details" in self.sheets:
                self._write_details(writer, report)

    def _write_summary(self, writer: pd.ExcelWriter, report: EvaluationReport) -> None:
        summary = report.summary()
        sheet_name = "Summary"
        summary.to_excel(writer, sheet_name=sheet_name, index=False)
        self._apply_color_scale(writer, sheet_name, 2, 2, len(summary), len(summary.columns) - 1)

    def _write_pivot_sheet(self, writer: pd.ExcelWriter, report: EvaluationReport, sheet_prefix: str, data: pd.DataFrame, col_col: str) -> None:
        metrics = [ev.name for ev in report.evaluators]

        for metric in metrics:
            pivot = data.pivot(index="model", columns=col_col, values=metric)
            pivot.columns = [f"{col_col}_{c}" for c in pivot.columns]
            pivot.index.name = metric

            sheet_name = f"{sheet_prefix}_{metric}"[:31]
            pivot.to_excel(writer, sheet_name=sheet_name)

            per_column = col_col == "horizon"
            self._apply_color_scale(writer, sheet_name, 2, 2, len(pivot), len(pivot.columns), per_column=per_column)

    def _write_matrix_sheet(self, writer: pd.ExcelWriter, report: EvaluationReport, sheet_prefix: str, data: pd.DataFrame, row_col: str, col_col: str) -> None:
        metrics = [ev.name for ev in report.evaluators]
        models = list(data["model"].unique())

        for metric in metrics:
            sheet_name = f"{sheet_prefix}_{metric}"[:31]
            ws = writer.book.create_sheet(sheet_name)

            col_offset = 1
            for model in models:
                model_df = data[data["model"] == model]
                pivot = model_df.pivot(index=row_col, columns=col_col, values=metric)
                ws.cell(row=1, column=col_offset, value=model)
                for c_idx, col_val in enumerate(pivot.columns, start=1):
                    ws.cell(row=2, column=col_offset + c_idx, value=f"{col_col}_{col_val}")
                for r_idx, row_val in enumerate(pivot.index, start=1):
                    ws.cell(row=2 + r_idx, column=col_offset, value=f"{row_col}_{row_val}")
                    for c_idx, col_val in enumerate(pivot.columns, start=1):
                        ws.cell(row=2 + r_idx, column=col_offset + c_idx, value=pivot.loc[row_val, col_val])

                self._apply_color_scale(writer, sheet_name, 3, col_offset + 1, len(pivot), len(pivot.columns))
                col_offset += len(pivot.columns) + 2

    def _apply_color_scale(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        start_row: int,
        start_col: int,
        num_rows: int,
        num_cols: int,
        per_column: bool = False,
    ) -> None:
        ws = writer.sheets[sheet_name]
        end_row = start_row + num_rows - 1
        end_col = start_col + num_cols - 1

        if num_rows <= 0 or num_cols <= 0:
            return

        rule = ColorScaleRule(
            start_type="min",
            start_color="63BE7B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="F8696B",
        )

        if per_column:
            for col in range(start_col, end_col + 1):
                start_cell = f"{get_column_letter(col)}{start_row}"
                end_cell = f"{get_column_letter(col)}{end_row}"
                ws.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)
        else:
            start_cell = f"{get_column_letter(start_col)}{start_row}"
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            ws.conditional_formatting.add(f"{start_cell}:{end_cell}", rule)

    def _write_details(self, writer: pd.ExcelWriter, report: EvaluationReport) -> None:
        if not report.results:
            return

        model_names = list(report.results.keys())
        base_df = report.results[model_names[0]].copy()

        base_cols = ["run_date", "target_date", "hour", "horizon", "day_in_test", "actual"]
        details_df = base_df[base_cols].copy()

        for model_name in model_names:
            model_df = report.results[model_name]
            details_df[f"prediction_{model_name}"] = model_df["prediction"].values

        details_df = details_df.sort_values(by=["target_date", "hour", "horizon"]).reset_index(drop=True)

        sheet_name = "Details"
        details_df.to_excel(writer, sheet_name=sheet_name, index=False)
