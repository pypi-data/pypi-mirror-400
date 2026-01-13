import calendar
import io
import math
import os
import random
import re
import zipfile
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import (
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    IntegerField,
    Q,
    Sum,
    Value,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import translation
from django.utils.encoding import escape_uri_path
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from fnschool import count_chinese_characters
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..forms import (
    CategoryForm,
    ConsumptionForm,
    IngredientForm,
    PurchasedIngredientsWorkBookForm,
)
from ..models import Category, Consumption, Ingredient, MealType
from ..views import decimal_prec
from .spreadsheet import MealTypeWorkbook, get_CNY_TEXT


def get_workbook_zip(request, month):

    year, month = [int(v) for v in month.split("-")]
    first_date_of_year = datetime(year, 1, 1).date()
    last_date_of_year = datetime(year, 12, 31).date()

    meal_types = MealType.objects.filter(
        Q(user=request.user) & Q(is_disabled=False)
    ).all()

    categories = (
        Category.objects.annotate(ingredients_count=Count("ingredients"))
        .filter(
            Q(user=request.user)
            & Q(is_disabled=False)
            & Q(ingredients_count__gt=0)
        )
        .order_by("priority")
        .all()
    )

    user_ingredients = (
        Ingredient.objects.annotate(
            total_consumed=Coalesce(
                Sum("consumptions__amount_used"), 0, output_field=IntegerField()
            )
        )
        .filter(
            Q(is_disabled=False)
            & (
                Q(
                    consumptions__date_of_using__range=(
                        first_date_of_year,
                        last_date_of_year,
                    )
                )
                | Q(quantity__gt=F("total_consumed"))
                | (
                    Q(storage_date__gte=first_date_of_year)
                    & Q(storage_date__lte=last_date_of_year)
                )
            )
            & Q(user=request.user)
            & Q(category__is_disabled=False)
            & Q(meal_type__is_disabled=False)
        )
        .select_related("meal_type", "category")
        .prefetch_related("consumptions")
        .distinct()
    )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for meal_type in meal_types:
            filename = (
                _(
                    "Canteen {meal_type} Daybook WorkBook ({month}) of {affiliation}"
                ).format(
                    meal_type=meal_type,
                    month=f"{year}{month:0>2}",
                    affiliation=request.user.affiliation,
                )
                + ".xlsx"
            )
            __ingredients = user_ingredients.filter(meal_type=meal_type).all()
            if len(__ingredients) < 1:
                continue

            wb = MealTypeWorkbook(
                request,
                year=year,
                month=month,
                ingredients=__ingredients,
                meal_type=meal_type,
                categories=categories,
            ).fill_in()
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            zip_file.writestr(filename, excel_buffer.getvalue())

    zip_buffer.seek(0)
    return zip_buffer


# The end.
