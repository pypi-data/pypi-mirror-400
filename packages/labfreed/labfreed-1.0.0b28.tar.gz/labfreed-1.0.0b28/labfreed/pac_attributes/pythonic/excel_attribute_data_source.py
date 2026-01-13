from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Optional, Tuple, List, Dict
from urllib.parse import urlparse, urlsplit, urlunsplit, parse_qsl, urlencode

from cachetools import TTLCache, cached

from labfreed.pac_attributes.api_data_models.response import AttributeGroup
from labfreed.pac_attributes.pythonic.py_attributes import pyAttribute, pyAttributes
from labfreed.pac_attributes.server.server import AttributeGroupDataSource
from labfreed.pac_cat.pac_cat import PAC_CAT

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("Please install labfreed with the [extended] extra: pip install labfreed[extended]")

# ---------------------------------------------------------------------
# Cache (shared by all instances). TTL can be overridden per instance.
# ---------------------------------------------------------------------
_cache = TTLCache(maxsize=128, ttl=0)


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def _is_sharepoint_url(s: str) -> bool:
    try:
        parsed = urlparse(s)
        if parsed.scheme not in {"http", "https"}:
            return False
        host = (parsed.netloc or "").lower()
        return ("sharepoint.com" in host) or ("1drv.ms" in host)
    except Exception:
        return False


def _is_local_path(s: str) -> bool:
    return os.path.exists(s) if not _is_sharepoint_url(s) else False


def _ensure_download_query(u: str) -> str:
    parts = urlsplit(u)
    q = dict(parse_qsl(parts.query, keep_blank_values=True))
    # Preserve all params (guest tokens, etc.), just force file response.
    q["download"] = "1"
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(q, doseq=True), parts.fragment))


def _get_row_by_first_cell(sheet_rows: List[tuple], match_value: str, base_url: str) -> Optional[Dict[str, object]]:
    if not sheet_rows:
        return None
    headers = sheet_rows[0]
    for row in sheet_rows[1:]:
        if not row:
            continue
        first = str(row[0]).strip() if row[0] is not None else ""
        if first == match_value:
            return {
                base_url + str(headers[i]).strip(): row[i]
                for i in range(1, len(headers))
                if headers[i] is not None
            }
    return None


# ---------------------------------------------------------------------
# Base class
# ---------------------------------------------------------------------
class _BaseExcelAttributeDataSource(AttributeGroupDataSource):
    """
    Common mapping logic from Excel rows to AttributeGroup.
    Subclasses implement `_read_rows_and_last_changed()`.
    """

    def __init__(self, *, base_url: str = "", cache_duration_seconds: int = 0, uses_pac_cat_short_form:bool=True, pac_to_key=None, header_mappings=None, **kwargs):
        self._base_url = base_url
        self._uses_pac_cat_short_form = uses_pac_cat_short_form
        self._pac_to_key = pac_to_key
        self._header_mappings = header_mappings or dict()
        # allow instance-level TTL override
        try:
            _cache.ttl = int(cache_duration_seconds)
        except Exception:
            pass
        super().__init__(**kwargs)

    def is_static(self) -> bool:
        return False

    def _read_rows_and_last_changed(self) -> Tuple[List[tuple], Optional[datetime]]:
        raise NotImplementedError

    @property
    def provides_attributes(self) -> List[str]:
        rows, _ = self._read_rows_and_last_changed()
        if not rows:
            return []
        return [self._base_url + r for r in rows[0][1:]]

    def attributes(self, pac_url:str) -> Optional[AttributeGroup]:
        try:
            p = PAC_CAT.from_url(pac_url)
            pac_url = p.to_url(use_short_notation=self._uses_pac_cat_short_form, include_extensions=self._include_extensions)
            print(f'Lookup in Excel of {pac_url}')
        except:
            ... # might as well try to match the original input
            
        if f:= self._pac_to_key:
            key = f(pac_url)
        else:
            key = pac_url
            
        rows, last_changed = self._read_rows_and_last_changed()
        d = _get_row_by_first_cell(rows, key, self._base_url)
        if not d:
            return None
        attributes = [pyAttribute(key= self._header_mappings.get(k, k), value=v) for k, v in d.items() if v is not None]
        return AttributeGroup(
            group_key=self._attribute_group_key,
            attributes=pyAttributes(attributes).to_payload_attributes()
        )


# ---------------------------------------------------------------------
# Local file implementation
# ---------------------------------------------------------------------
class LocalExcelAttributeDataSource(_BaseExcelAttributeDataSource):
    def __init__(self, file_path: str, **kwargs):
        self._file_path = file_path
        super().__init__(**kwargs)

    @cached(_cache)
    def _read_rows_and_last_changed(self) -> Tuple[List[tuple], Optional[datetime]]:
        logging.info(f"Attempting to load workbook: {self._file_path!r}")

        try:
            wb = load_workbook(
                filename=self._file_path,
                read_only=True,
                data_only=True
            )
            ws = wb.active
            logging.info(f"Workbook opened successfully. Active sheet: {ws.title!r}")

            rows = list(ws.iter_rows(values_only=True))
            logging.info(f"Read {len(rows)} rows from {self._file_path!r}")

            last_changed = wb.properties.modified
            logging.info(f"Workbook 'modified' property: {last_changed}")

            wb.close()
            return rows, last_changed

        except FileNotFoundError:
            logging.error(f"Workbook not found at: {self._file_path!r}", exc_info=True)
            raise
        except PermissionError:
            logging.error(f"Permission denied when accessing: {self._file_path!r}", exc_info=True)
            raise
        except Exception as e:
            logging.exception(f"Unexpected error reading workbook {self._file_path!r}: {e}")
            raise


# # ---------------------------------------------------------------------
# # SharePoint/OneDrive (anonymous only)
# # ---------------------------------------------------------------------
# class SharePointExcelAttributeDataSource(_BaseExcelAttributeDataSource):
#     """
#     Anonymous 'anyone with the link' reader for SharePoint/OneDrive Excel.
#     No MSAL, no AAD app permissions.
#     """

#     def __init__(self, url: str, *, timeout: int = 30, **kwargs):
#         self._url = url
#         self._timeout = timeout
#         super().__init__(**kwargs)

#     @cached(_cache)
#     def _read_rows_and_last_changed(self) -> Tuple[List[tuple], Optional[datetime]]:
#         content, last_changed = self._download_bytes_anon(self._url, timeout=self._timeout)
#         with io.BytesIO(content) as fh:
#             wb = load_workbook(filename=fh, read_only=True, data_only=True)
#             ws = wb.active
#             rows = list(ws.iter_rows(values_only=True))
#             wb_last = wb.properties.modified  # often None for streamed files
#             wb.close()
#         return rows, (wb_last or last_changed)

#     def _download_bytes_anon(self, url: str, *, timeout: int) -> Tuple[bytes, Optional[datetime]]:
#         u = _ensure_download_query(url)
#         headers = {"User-Agent": "python-requests/anon-sharepoint-downloader"}
#         resp = requests.get(u, headers=headers, timeout=timeout, allow_redirects=True)
#         resp.raise_for_status()

#         last = None
#         lm_hdr = resp.headers.get("Last-Modified")
#         if lm_hdr:
#             try:
#                 last = parsedate_to_datetime(lm_hdr)
#                 if last.tzinfo is None:
#                     last = last.replace(tzinfo=timezone.utc)
#                 else:
#                     last = last.astimezone(timezone.utc)
#             except Exception:
#                 last = None
#         return resp.content, last




__all__ = [
    "LocalExcelAttributeDataSource",
]
