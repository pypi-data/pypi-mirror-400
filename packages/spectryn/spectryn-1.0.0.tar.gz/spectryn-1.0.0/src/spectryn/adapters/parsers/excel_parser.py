"""
Excel Parser - Parse Excel files into domain entities.

Implements the DocumentParserPort interface for Excel spreadsheets.

Requires openpyxl library for .xlsx files.
"""

import logging
from pathlib import Path
from typing import Any

from spectryn.core.domain.entities import Comment, Epic, Subtask, UserStory
from spectryn.core.domain.enums import Priority, Status
from spectryn.core.domain.value_objects import (
    AcceptanceCriteria,
    Description,
    IssueKey,
    StoryId,
)
from spectryn.core.ports.document_parser import DocumentParserPort, ParserError


class ExcelParser(DocumentParserPort):
    """
    Parser for Excel (.xlsx, .xls) story specification files.

    Parses spreadsheet data where each row represents a story and
    columns represent fields. Similar to CSV but with Excel-specific features.

    Expected Excel format (first sheet is used by default):

    | id        | title        | description                        | story_points | priority | status   | acceptance_criteria | subtasks    |
    |-----------|--------------|------------------------------------|--------------| ---------|----------|---------------------|-------------|
    | STORY-001 | Story Title  | As a user, I want..., so that...   | 5            | high     | planned  | AC1; AC2; AC3       | Task1;Task2 |
    | STORY-002 | Another Story| Simple description                 | 3            | medium   | done     | AC1                 |             |

    Note: Story IDs can use any PREFIX-NUMBER format (US-001, PROJ-123, FEAT-001, etc.)

    Column mappings (case-insensitive, flexible):
    - id, story_id → Story ID
    - title, name, story → Title
    - description, desc → Description
    - story_points, points, sp → Story Points
    - priority → Priority
    - status → Status
    - acceptance_criteria, ac → Acceptance Criteria (semicolon-separated)
    - subtasks, tasks → Subtasks (semicolon-separated)
    - technical_notes, notes → Technical Notes
    - links, related → Links (semicolon-separated, format: "type:target")
    - comments → Comments (semicolon-separated)

    Features:
    - Supports multiple sheets (each sheet becomes a separate epic)
    - Auto-detects header row
    - Handles merged cells
    - Supports formulas (uses calculated values)
    """

    # Column name mappings (same as CSV)
    COLUMN_MAPPINGS = {
        "id": ["id", "story_id", "story id", "story-id", "issue_id", "issue id"],
        "title": ["title", "name", "story", "summary", "story_title", "story title"],
        "description": ["description", "desc", "user_story", "user story", "as_a"],
        "story_points": ["story_points", "story points", "points", "sp", "estimate"],
        "priority": ["priority", "prio", "p"],
        "status": ["status", "state", "stage"],
        "acceptance_criteria": ["acceptance_criteria", "acceptance criteria", "ac", "criteria"],
        "subtasks": ["subtasks", "tasks", "sub_tasks", "sub-tasks"],
        "technical_notes": ["technical_notes", "technical notes", "notes", "tech_notes"],
        "links": ["links", "related", "dependencies", "relations"],
        "comments": ["comments", "discussion"],
        "assignee": ["assignee", "assigned", "owner", "assigned_to", "assigned to"],
        "labels": ["labels", "tags"],
        "epic": ["epic", "epic_key", "epic key", "parent"],
    }

    def __init__(self, sheet_name: str | None = None) -> None:
        """
        Initialize the Excel parser.

        Args:
            sheet_name: Optional specific sheet name to parse (default: first sheet)
        """
        self.logger = logging.getLogger("ExcelParser")
        self.sheet_name = sheet_name
        self._openpyxl = None

    def _ensure_openpyxl(self) -> Any:
        """Lazy load openpyxl library."""
        if self._openpyxl is None:
            try:
                import openpyxl

                self._openpyxl = openpyxl
            except ImportError:
                raise ParserError(
                    "openpyxl library required for Excel parsing. "
                    "Install with: pip install openpyxl"
                )
        return self._openpyxl

    # -------------------------------------------------------------------------
    # DocumentParserPort Implementation
    # -------------------------------------------------------------------------

    @property
    def name(self) -> str:
        return "Excel"

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm", ".xls"]

    def can_parse(self, source: str | Path) -> bool:
        """Check if source is a valid Excel file."""
        if isinstance(source, Path):
            return source.suffix.lower() in self.supported_extensions

        # Can't parse Excel from string content
        return False

    def parse_stories(self, source: str | Path) -> list[UserStory]:
        """Parse user stories from Excel source."""
        if isinstance(source, str):
            source = Path(source)

        rows = self._load_excel(source)

        if not rows:
            return []

        stories = []
        for i, row in enumerate(rows):
            try:
                story = self._parse_row(row, i)
                if story:
                    stories.append(story)
            except Exception as e:
                story_id = row.get("id", f"row-{i}")
                self.logger.warning(f"Failed to parse story {story_id}: {e}")

        return stories

    def parse_epic(self, source: str | Path) -> Epic | None:
        """Parse an epic with its stories from Excel source."""
        if isinstance(source, str):
            source = Path(source)

        stories = self.parse_stories(source)

        if not stories:
            return None

        # Use filename or sheet name as epic title
        epic_title = self.sheet_name or source.stem

        return Epic(
            key=IssueKey("EPIC-0"),
            title=epic_title,
            stories=stories,
        )

    def parse_all_sheets(self, source: str | Path) -> list[Epic]:
        """
        Parse all sheets as separate epics.

        Args:
            source: Path to Excel file

        Returns:
            List of Epic entities, one per sheet
        """
        if isinstance(source, str):
            source = Path(source)

        openpyxl = self._ensure_openpyxl()

        try:
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        except Exception as e:
            raise ParserError(f"Failed to open Excel file: {e}")

        epics = []

        for sheet_name in workbook.sheetnames:
            self.sheet_name = sheet_name
            rows = self._load_sheet(workbook[sheet_name])

            if not rows:
                continue

            stories = []
            for i, row in enumerate(rows):
                try:
                    story = self._parse_row(row, i)
                    if story:
                        stories.append(story)
                except Exception as e:
                    self.logger.warning(f"Failed to parse row {i} in sheet {sheet_name}: {e}")

            if stories:
                epics.append(
                    Epic(
                        key=IssueKey("EPIC-0"),
                        title=sheet_name,
                        stories=stories,
                    )
                )

        workbook.close()
        return epics

    def validate(self, source: str | Path) -> list[str]:
        """Validate Excel source without full parsing."""
        errors: list[str] = []

        if isinstance(source, str):
            source = Path(source)

        if not source.exists():
            errors.append(f"File not found: {source}")
            return errors

        try:
            rows = self._load_excel(source)
        except ParserError as e:
            return [str(e)]

        if not rows:
            errors.append("No data rows found in Excel file")
            return errors

        for i, row in enumerate(rows):
            if not row.get("id") and not row.get("title"):
                errors.append(f"Row {i + 1}: missing both 'id' and 'title'")

        return errors

    # -------------------------------------------------------------------------
    # Private Methods
    # -------------------------------------------------------------------------

    def _load_excel(self, source: Path) -> list[dict[str, str]]:
        """Load Excel content from file."""
        openpyxl = self._ensure_openpyxl()

        try:
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        except Exception as e:
            raise ParserError(f"Failed to open Excel file: {e}")

        try:
            # Get specified sheet or first sheet
            if self.sheet_name and self.sheet_name in workbook.sheetnames:
                sheet = workbook[self.sheet_name]
            else:
                sheet = workbook.active or workbook.worksheets[0]

            return self._load_sheet(sheet)
        finally:
            workbook.close()

    def _load_sheet(self, sheet: Any) -> list[dict[str, str]]:
        """Load data from a worksheet."""
        rows = []

        # Get all rows as list of values
        all_rows = list(sheet.iter_rows(values_only=True))

        if not all_rows:
            return rows

        # Find header row (first row with recognizable column names)
        header_row_idx = 0
        headers = []

        for i, row in enumerate(all_rows[:5]):  # Check first 5 rows for headers
            if row and any(cell for cell in row if cell):
                potential_headers = [str(cell).lower().strip() if cell else "" for cell in row]
                # Check if this looks like a header row
                expected = ["id", "title", "name", "story", "description"]
                if any(h in potential_headers for h in expected):
                    headers = potential_headers
                    header_row_idx = i
                    break

        if not headers:
            # Use first non-empty row as headers
            for i, row in enumerate(all_rows):
                if row and any(cell for cell in row if cell):
                    headers = [
                        str(cell).lower().strip() if cell else f"col_{j}"
                        for j, cell in enumerate(row)
                    ]
                    header_row_idx = i
                    break

        if not headers:
            return rows

        # Parse data rows
        for row in all_rows[header_row_idx + 1 :]:
            if not row or not any(cell for cell in row if cell):
                continue

            row_dict = {}
            for j, cell in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = str(cell).strip() if cell is not None else ""

            normalized = self._normalize_row(row_dict)
            if normalized.get("id") or normalized.get("title"):
                rows.append(normalized)

        return rows

    def _normalize_row(self, row: dict[str, str]) -> dict[str, str]:
        """Normalize column names to standard field names."""
        normalized: dict[str, str] = {}

        for standard_name, aliases in self.COLUMN_MAPPINGS.items():
            for key, value in row.items():
                if key.lower().strip() in aliases:
                    normalized[standard_name] = value.strip() if value else ""
                    break

        return normalized

    def _parse_row(self, row: dict[str, str], index: int) -> UserStory | None:
        """Parse a single row into a UserStory.

        Accepts any PREFIX-NUMBER format for story IDs (e.g., US-001, EU-042, PROJ-123).
        """
        story_id = row.get("id", "")
        if not story_id:
            story_id = f"STORY-{index + 1:03d}"

        title = row.get("title", "").strip()
        if not title:
            return None

        description = self._parse_description(row.get("description", ""))
        story_points = self._parse_int(row.get("story_points", "0"))
        priority = Priority.from_string(row.get("priority", "medium"))
        status = Status.from_string(row.get("status", "planned"))

        acceptance = self._parse_semicolon_list(row.get("acceptance_criteria", ""))
        subtasks = self._parse_subtasks(row.get("subtasks", ""))
        links = self._parse_links(row.get("links", ""))
        comments = self._parse_comments(row.get("comments", ""))

        tech_notes = row.get("technical_notes", "")
        assignee = row.get("assignee") or None
        labels = [l.strip() for l in row.get("labels", "").split(";") if l.strip()]

        return UserStory(
            id=StoryId(story_id),
            title=title,
            description=description,
            acceptance_criteria=AcceptanceCriteria.from_list(acceptance, [False] * len(acceptance)),
            technical_notes=tech_notes,
            story_points=story_points,
            priority=priority,
            status=status,
            subtasks=subtasks,
            commits=[],
            links=links,
            comments=comments,
            assignee=assignee,
            labels=labels,
        )

    def _parse_description(self, text: str) -> Description | None:
        """Parse description from text."""
        if not text:
            return None

        import re

        pattern = r"As a[n]?\s+(.+?),?\s+I want\s+(.+?),?\s+so that\s+(.+)"
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return Description(
                role=match.group(1).strip(),
                want=match.group(2).strip(),
                benefit=match.group(3).strip(),
            )

        return Description(role="", want=text, benefit="")

    def _parse_semicolon_list(self, text: str) -> list[str]:
        """Parse semicolon-separated list."""
        if not text:
            return []
        return [item.strip() for item in text.split(";") if item.strip()]

    def _parse_subtasks(self, text: str) -> list[Subtask]:
        """Parse semicolon-separated subtasks."""
        if not text:
            return []

        subtasks = []
        items = self._parse_semicolon_list(text)

        for i, item in enumerate(items):
            subtasks.append(
                Subtask(
                    number=i + 1,
                    name=item,
                    description="",
                    story_points=1,
                    status=Status.PLANNED,
                )
            )

        return subtasks

    def _parse_links(self, text: str) -> list[tuple[str, str]]:
        """Parse semicolon-separated links."""
        if not text:
            return []

        links = []
        items = self._parse_semicolon_list(text)

        for item in items:
            if ":" in item:
                parts = item.split(":", 1)
                link_type = parts[0].strip().lower().replace("_", " ")
                target = parts[1].strip()
                if target:
                    links.append((link_type, target))

        return links

    def _parse_comments(self, text: str) -> list[Comment]:
        """Parse semicolon-separated comments."""
        if not text:
            return []

        comments = []
        items = self._parse_semicolon_list(text)

        for item in items:
            comments.append(
                Comment(
                    body=item,
                    author=None,
                    created_at=None,
                    comment_type="text",
                )
            )

        return comments

    def _parse_int(self, value: str) -> int:
        """Parse integer from string."""
        if not value:
            return 0

        import re

        cleaned = re.sub(r"[^\d-]", "", str(value))

        try:
            return int(cleaned) if cleaned else 0
        except ValueError:
            return 0
