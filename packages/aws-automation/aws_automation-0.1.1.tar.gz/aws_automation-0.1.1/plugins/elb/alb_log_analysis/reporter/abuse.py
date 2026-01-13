"""Abuse IP sheets writer."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

from .base import BaseSheetWriter, get_column_letter
from .config import HEADERS, SHEET_NAMES, STATUS_CODE_TYPES, SheetConfig

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class AbuseIPSheetWriter(BaseSheetWriter):
    """Creates abuse IP list sheet."""

    def write(self) -> None:
        """Create abuse IP list sheet."""
        try:
            abuse_ips_list, abuse_ip_details = self.get_normalized_abuse_ips()
            if not abuse_ips_list:
                return

            # Get matching abuse IPs (those actually in logs)
            actual_client_ips = set(self.data.get("client_ip_counts", {}).keys())
            matching_ips = actual_client_ips.intersection(set(abuse_ips_list))

            if not matching_ips:
                return

            ws = self.create_sheet(SHEET_NAMES.ABUSE_IP_LIST)
            headers = list(HEADERS.ABUSE_IP)
            self.write_header_row(ws, headers)

            # Sort by request count
            client_ip_counts = self.data.get("client_ip_counts", {})
            sorted_ips = sorted(
                matching_ips,
                key=lambda ip: client_ip_counts.get(ip, 0),
                reverse=True,
            )

            self._write_abuse_ip_rows(ws, sorted_ips, abuse_ip_details)
            self._finalize_abuse_sheet(ws, headers, len(matching_ips))

        except Exception as e:
            logger.error(f"Abuse IP 시트 생성 중 오류: {e}")

    def _write_abuse_ip_rows(
        self,
        ws: Worksheet,
        sorted_ips: list[str],
        abuse_ip_details: dict[str, Any],
    ) -> None:
        """Write abuse IP rows."""
        client_ip_counts = self.data.get("client_ip_counts", {})
        country_mapping = self.data.get("ip_country_mapping", {})
        border = self.styles.thin_border

        for row_idx, ip in enumerate(sorted_ips, start=2):
            details = abuse_ip_details.get(ip, {}) if isinstance(abuse_ip_details, dict) else {}
            request_count = client_ip_counts.get(ip, 0)

            # Count (A)
            cell = ws.cell(row=row_idx, column=1, value=request_count)
            cell.border = border
            cell.alignment = self.styles.align_right
            cell.number_format = "#,##0"

            # IP (B)
            cell = ws.cell(row=row_idx, column=2, value=ip)
            cell.border = border
            cell.alignment = self.styles.align_center

            # Country (C)
            country = country_mapping.get(ip, "N/A")
            cell = ws.cell(row=row_idx, column=3, value=country)
            cell.border = border
            cell.alignment = self.styles.align_center

            # ASN (D)
            cell = ws.cell(row=row_idx, column=4, value=details.get("asn", "N/A"))
            cell.border = border
            cell.alignment = self.styles.align_center

            # ISP (E)
            cell = ws.cell(row=row_idx, column=5, value=details.get("isp", "N/A"))
            cell.border = border
            cell.alignment = self.styles.align_center

    def _finalize_abuse_sheet(
        self,
        ws: Worksheet,
        headers: list[str],
        data_count: int,
    ) -> None:
        """Finalize abuse IP sheet."""
        ws.row_dimensions[1].height = SheetConfig.HEADER_ROW_HEIGHT

        if data_count > 0:
            for row_idx in range(2, 2 + data_count):
                ws.row_dimensions[row_idx].height = SheetConfig.DATA_ROW_HEIGHT

            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{data_count + 1}"

        self._apply_column_widths(ws, headers)
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE


class AbuseRequestsSheetWriter(BaseSheetWriter):
    """Creates abuse IP requests analysis sheet."""

    def write(self) -> None:
        """Create abuse IP requests sheet."""
        try:
            matching_ips = self.get_matching_abuse_ips()
            if not matching_ips:
                return

            # Collect all logs from abuse IPs
            abuse_requests = self._collect_abuse_requests(matching_ips)
            if not abuse_requests:
                return

            ws = self.create_sheet(SHEET_NAMES.ABUSE_REQUESTS)
            headers = list(HEADERS.ABUSE_REQUESTS)
            self.write_header_row(ws, headers)

            # Sort by timestamp and write
            sorted_requests = sorted(abuse_requests, key=self._safe_timestamp_key)
            self._write_abuse_request_rows(ws, sorted_requests, headers)

            # Finalize
            if sorted_requests:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A1:{last_col}{len(sorted_requests) + 1}"

            self.apply_wrap_text(ws, headers)
            self._apply_column_widths(ws, headers)
            ws.row_dimensions[1].height = SheetConfig.HEADER_ROW_HEIGHT
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE

        except Exception as e:
            logger.error(f"악성 IP 요청 분석 시트 생성 중 오류: {e}")

    def _collect_abuse_requests(self, abuse_ips: set[str]) -> list[dict[str, Any]]:
        """Collect all requests from abuse IPs."""
        all_logs = []

        for status_type in STATUS_CODE_TYPES:
            if status_type in self.data and isinstance(self.data[status_type], dict):
                full_logs = self.data[status_type].get("full_logs", [])
                if isinstance(full_logs, list):
                    all_logs.extend(full_logs)

        # Filter for abuse IPs
        return [log for log in all_logs if log.get("client_ip", "N/A") in abuse_ips]

    def _safe_timestamp_key(self, entry: dict[str, Any]) -> datetime:
        """Get timestamp for sorting."""
        ts = entry.get("timestamp")
        if isinstance(ts, datetime):
            return ts
        if isinstance(ts, str):
            try:
                return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        return datetime.min

    def _write_abuse_request_rows(
        self,
        ws: Worksheet,
        requests: list[dict[str, Any]],
        headers: list[str],
    ) -> None:
        """Write abuse request rows."""
        border = self.styles.thin_border

        for row_idx, log in enumerate(requests, start=2):
            client_ip = log.get("client_ip", "N/A")
            country = self.get_country_code(client_ip)

            target = log.get("target", "")
            target_field = "" if not target or target == "-" else target
            target_group = log.get("target_group_name", "") or ""

            timestamp_str = self.format_timestamp(log.get("timestamp"))
            method = log.get("http_method", "").replace("-", "")
            request = log.get("request", "N/A")
            user_agent = log.get("user_agent", "N/A")
            elb_status = self.convert_status_code(log.get("elb_status_code", "N/A"))
            backend_status = self.convert_status_code(log.get("target_status_code", "N/A"))

            values = [
                timestamp_str,
                client_ip,
                country,
                target_field,
                target_group,
                method,
                request,
                user_agent,
                elb_status,
                backend_status,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                # Alignment
                if col_idx in (1, 2, 3, 4, 5, 6):  # Timestamp through Method
                    cell.alignment = self.styles.align_center
                elif col_idx in (7, 8):  # Request, User Agent
                    cell.alignment = self.styles.align_left
                elif col_idx in (9, 10):  # Status codes
                    cell.alignment = self.styles.align_right
                    if isinstance(value, int):
                        cell.number_format = "0"
