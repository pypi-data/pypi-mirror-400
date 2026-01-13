"""Multi-format export module for SpreadsheetDL.

Provides export functionality to multiple formats including Excel (XLSX),
CSV, and PDF while preserving formatting where possible.

Features:
    - Export to Excel (.xlsx) with formatting
    - Export to CSV for data portability
    - Export to PDF for reports
    - Preserve themes and formatting where possible
    - Batch export to multiple formats
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any

from spreadsheet_dl.exceptions import (
    FileError,
    SpreadsheetDLError,
)
from spreadsheet_dl.progress import BatchProgress


class MultiExportFormat(Enum):
    """Supported formats for multi-format export operations."""

    XLSX = "xlsx"
    CSV = "csv"
    PDF = "pdf"
    JSON = "json"


class MultiExportError(SpreadsheetDLError):
    """Base exception for multi-format export errors."""

    error_code = "FT-MXP-1300"


class FormatNotSupportedError(MultiExportError):
    """Raised when export format is not supported."""

    error_code = "FT-MXP-1301"

    def __init__(
        self,
        format_name: str,
        available_formats: list[str] | None = None,
        **kwargs: Any,
    ) -> None:
        """Initialize the instance."""
        self.format_name = format_name
        self.available_formats = available_formats or [
            f.value for f in MultiExportFormat
        ]
        super().__init__(
            f"Export format not supported: {format_name}",
            details=f"Available formats: {', '.join(self.available_formats)}",
            suggestion="Use one of the supported export formats.",
            **kwargs,
        )


class ExportDependencyError(MultiExportError):
    """Raised when required library for export is not installed."""

    error_code = "FT-MXP-1302"

    def __init__(
        self,
        format_name: str,
        library: str,
        install_cmd: str,
        **kwargs: Any,
    ) -> None:
        """Initialize the instance."""
        self.format_name = format_name
        self.library = library
        super().__init__(
            f"Library required for {format_name} export: {library}",
            suggestion=f"Install with: {install_cmd}",
            **kwargs,
        )


@dataclass
class ExportOptions:
    """Options for export operations."""

    # General options
    include_headers: bool = True
    include_formulas: bool = False  # Export formula results, not formulas
    preserve_formatting: bool = True

    # Sheet selection
    sheet_names: list[str] | None = None  # None = all sheets
    active_sheet_only: bool = False

    # CSV-specific options
    csv_delimiter: str = ","
    csv_quoting: int = csv.QUOTE_MINIMAL
    csv_encoding: str = "utf-8"

    # Excel-specific options
    xlsx_date_format: str = "YYYY-MM-DD"
    xlsx_number_format: str = "#,##0.00"
    xlsx_currency_format: str = '"$"#,##0.00'

    # PDF-specific options
    pdf_page_size: str = "letter"  # letter, a4, legal
    pdf_orientation: str = "portrait"  # portrait, landscape
    pdf_title: str = ""
    pdf_author: str = "SpreadsheetDL"
    pdf_include_summary: bool = True


@dataclass
class SheetData:
    """Data structure for a spreadsheet sheet."""

    name: str
    rows: list[list[Any]] = field(default_factory=list)
    column_widths: list[int] = field(default_factory=list)
    headers: list[str] = field(default_factory=list)
    styles: dict[str, Any] = field(default_factory=dict)

    @property
    def row_count(self) -> int:
        """Get number of rows."""
        return len(self.rows)

    @property
    def column_count(self) -> int:
        """Get number of columns."""
        if self.rows:
            return max(len(row) for row in self.rows)
        return 0


class MultiFormatExporter:
    """Export ODS files to multiple formats.

    Supports Excel (XLSX), CSV, and PDF export with formatting preservation.

    Example:
        >>> exporter = MultiFormatExporter()
        >>> exporter.options is not None
        True
        >>> exporter.options.include_headers
        True
    """

    def __init__(self, options: ExportOptions | None = None) -> None:
        """Initialize exporter.

        Args:
            options: Export options. If None, uses defaults.
        """
        self.options = options or ExportOptions()

    def export(
        self,
        ods_path: str | Path,
        output_path: str | Path,
        format: MultiExportFormat | str,
    ) -> Path:
        """Export ODS file to specified format.

        Args:
            ods_path: Path to source ODS file.
            output_path: Path for output file.
            format: Export format (xlsx, csv, pdf, json).

        Returns:
            Path to exported file.

        Raises:
            FileError: If source file doesn't exist.
            FormatNotSupportedError: If format is not supported.
            MultiExportError: If export fails.
        """
        ods_path = Path(ods_path)
        output_path = Path(output_path)

        if not ods_path.exists():
            raise FileError(f"Source file not found: {ods_path}")

        # Parse format
        format_obj: MultiExportFormat
        if isinstance(format, str):
            try:
                format_obj = MultiExportFormat(format.lower())
            except ValueError as exc:
                raise FormatNotSupportedError(format) from exc
        else:
            format_obj = format

        # Load ODS data
        sheet_data = self._load_ods(ods_path)

        # Export to target format
        if format_obj == MultiExportFormat.XLSX:
            return self._export_xlsx(sheet_data, output_path)
        elif format_obj == MultiExportFormat.CSV:
            return self._export_csv(sheet_data, output_path)
        elif format_obj == MultiExportFormat.PDF:
            return self._export_pdf(sheet_data, output_path)
        elif format_obj == MultiExportFormat.JSON:
            return self._export_json(sheet_data, output_path)
        else:
            raise FormatNotSupportedError(format_obj.value)

    def export_batch(
        self,
        ods_path: str | Path,
        output_dir: str | Path,
        formats: list[MultiExportFormat | str],
    ) -> dict[str, Path | None]:
        """Export ODS file to multiple formats.

        Args:
            ods_path: Path to source ODS file.
            output_dir: Directory for output files.
            formats: List of export formats.

        Returns:
            Dictionary mapping format names to output paths.
        """
        ods_path = Path(ods_path)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        stem = ods_path.stem
        results: dict[str, Path | None] = {}

        for fmt in formats:
            if isinstance(fmt, str):
                fmt = MultiExportFormat(fmt.lower())

            output_name = f"{stem}.{fmt.value}"
            output_path = output_dir / output_name

            try:
                results[fmt.value] = self.export(ods_path, output_path, fmt)
            except (OSError, ValueError, MultiExportError):
                results[fmt.value] = None
                # Continue with other formats

        return results

    def _load_ods(self, ods_path: Path) -> list[SheetData]:
        """Load data from ODS file."""
        try:
            from odf import text as odf_text
            from odf.opendocument import load
            from odf.table import Table, TableCell, TableRow
        except ImportError as exc:
            raise ExportDependencyError(
                "ODS",
                "odfpy",
                "pip install odfpy",
            ) from exc

        doc = load(str(ods_path))
        sheets: list[SheetData] = []

        for table in doc.spreadsheet.getElementsByType(Table):
            sheet_name = table.getAttribute("name") or "Sheet"

            # Skip sheets if filtering
            if self.options.sheet_names and sheet_name not in self.options.sheet_names:
                continue

            sheet = SheetData(name=sheet_name)
            first_row = True

            for row in table.getElementsByType(TableRow):
                row_data: list[Any] = []

                for cell in row.getElementsByType(TableCell):
                    # Handle repeated columns
                    repeat = cell.getAttribute("numbercolumnsrepeated")
                    repeat_count = int(repeat) if repeat else 1

                    value = self._extract_cell_value(cell, odf_text)

                    # Add value (or repeat it)
                    # Limit empty cell repeats to 100 (ODS often pads with many empties)
                    # but allow up to 16384 for non-empty values (Excel max columns)
                    max_repeat = 100 if (value is None or value == "") else 16384
                    for _ in range(min(repeat_count, max_repeat)):
                        row_data.append(value)

                # Skip empty rows at the end
                if any(v is not None and v != "" for v in row_data):
                    if first_row and self.options.include_headers:
                        sheet.headers = [str(v) if v else "" for v in row_data]
                        first_row = False
                    sheet.rows.append(row_data)

            if sheet.rows:
                sheets.append(sheet)

            if self.options.active_sheet_only:
                break

        return sheets

    def _extract_cell_value(self, cell: Any, odf_text: Any) -> Any:
        """Extract value from ODS cell."""
        value_type = cell.getAttribute("valuetype")

        if value_type in ("float", "currency", "percentage"):
            val = cell.getAttribute("value")
            return Decimal(val) if val else None
        elif value_type == "date":
            date_val = cell.getAttribute("datevalue")
            if date_val:
                try:
                    return datetime.fromisoformat(date_val).date()
                except ValueError:
                    return date_val
            return None
        elif value_type == "boolean":
            return cell.getAttribute("booleanvalue") == "true"
        else:
            # Text or empty
            try:
                text_content = []
                for p in cell.getElementsByType(odf_text.P):
                    text_content.append(str(p))
                return " ".join(text_content) if text_content else None
            except (AttributeError, TypeError):
                return None

    def _export_xlsx(self, sheets: list[SheetData], output_path: Path) -> Path:
        """Export to Excel XLSX format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ExportDependencyError(
                "XLSX",
                "openpyxl",
                "pip install openpyxl",
            ) from exc

        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.name[:31])  # Excel sheet name limit

            # Style definitions
            _header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font_white = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Use progress bar for large sheets (>100 rows)
            total_rows = len(sheet_data.rows)
            use_progress = total_rows > 100

            if use_progress:
                with BatchProgress(
                    total_rows, f"Exporting {sheet_data.name} to XLSX"
                ) as progress:
                    for row_idx, row_data in enumerate(sheet_data.rows, 1):
                        self._export_xlsx_row(
                            ws,
                            row_idx,
                            row_data,
                            header_font_white,
                            header_fill,
                            thin_border,
                        )
                        progress.update()
            else:
                for row_idx, row_data in enumerate(sheet_data.rows, 1):
                    self._export_xlsx_row(
                        ws,
                        row_idx,
                        row_data,
                        header_font_white,
                        header_fill,
                        thin_border,
                    )

            # Auto-adjust column widths
            for col_idx in range(1, sheet_data.column_count + 1):
                column = get_column_letter(col_idx)
                max_length = 0
                for row in sheet_data.rows:
                    if col_idx <= len(row) and row[col_idx - 1] is not None:
                        max_length = max(max_length, len(str(row[col_idx - 1])))
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    def _export_xlsx_row(
        self,
        ws: Any,
        row_idx: int,
        row_data: list[Any],
        header_font_white: Any,
        header_fill: Any,
        thin_border: Any,
    ) -> None:
        """Export a single row to XLSX worksheet."""
        from openpyxl.styles import Alignment

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Set value
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = self.options.xlsx_currency_format
            elif isinstance(value, date):
                cell.value = value
                cell.number_format = self.options.xlsx_date_format
            else:
                cell.value = value

            # Apply formatting
            if self.options.preserve_formatting:
                cell.border = thin_border

                # Header styling (first row)
                if row_idx == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

    def _export_csv(self, sheets: list[SheetData], output_path: Path) -> Path:
        """Export to CSV format."""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # For multiple sheets, create separate files or combine
        if len(sheets) == 1:
            self._write_csv_sheet(sheets[0], output_path)
        else:
            # Create a directory with separate CSV files per sheet
            csv_dir = output_path.parent / output_path.stem
            csv_dir.mkdir(exist_ok=True)

            for sheet in sheets:
                sheet_path = csv_dir / f"{sheet.name}.csv"
                self._write_csv_sheet(sheet, sheet_path)

            # Also create combined file
            self._write_csv_combined(sheets, output_path)

        return output_path

    def _write_csv_sheet(self, sheet: SheetData, output_path: Path) -> None:
        """Write a single sheet to CSV."""
        with open(
            output_path, "w", newline="", encoding=self.options.csv_encoding
        ) as f:
            writer = csv.writer(
                f,
                delimiter=self.options.csv_delimiter,
                quoting=csv.QUOTE_MINIMAL
                if self.options.csv_quoting == 1
                else csv.QUOTE_ALL,
            )

            # Use progress for large sheets
            total_rows = len(sheet.rows)
            use_progress = total_rows > 100

            if use_progress:
                with BatchProgress(
                    total_rows, f"Exporting {sheet.name} to CSV"
                ) as progress:
                    for row in sheet.rows:
                        csv_row = self._convert_row_to_csv(row)
                        writer.writerow(csv_row)
                        progress.update()
            else:
                for row in sheet.rows:
                    csv_row = self._convert_row_to_csv(row)
                    writer.writerow(csv_row)

    def _convert_row_to_csv(self, row: list[Any]) -> list[Any]:
        """Convert a row's values to CSV-compatible format."""
        csv_row: list[Any] = []
        for value in row:
            if isinstance(value, Decimal):
                csv_row.append(float(value))
            elif isinstance(value, date):
                csv_row.append(value.isoformat())
            elif value is None:
                csv_row.append("")
            else:
                csv_row.append(value)
        return csv_row

    def _write_csv_combined(self, sheets: list[SheetData], output_path: Path) -> None:
        """Write all sheets to a single CSV with sheet separators."""
        with open(
            output_path, "w", newline="", encoding=self.options.csv_encoding
        ) as f:
            writer = csv.writer(
                f,
                delimiter=self.options.csv_delimiter,
                quoting=csv.QUOTE_MINIMAL
                if self.options.csv_quoting == 1
                else csv.QUOTE_ALL,
            )

            for i, sheet in enumerate(sheets):
                if i > 0:
                    writer.writerow([])  # Blank row separator
                writer.writerow([f"=== {sheet.name} ==="])

                for row in sheet.rows:
                    csv_row: list[Any] = []
                    for value in row:
                        if isinstance(value, Decimal):
                            csv_row.append(float(value))
                        elif isinstance(value, date):
                            csv_row.append(value.isoformat())
                        elif value is None:
                            csv_row.append("")
                        else:
                            csv_row.append(value)
                    writer.writerow(csv_row)

    def _export_pdf(self, sheets: list[SheetData], output_path: Path) -> Path:
        """Export to PDF format."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, LETTER, landscape, portrait
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError as exc:
            raise ExportDependencyError(
                "PDF",
                "reportlab",
                "pip install reportlab",
            ) from exc

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Page size
        page_sizes = {"letter": LETTER, "a4": A4, "legal": (8.5 * inch, 14 * inch)}
        page_size = page_sizes.get(self.options.pdf_page_size.lower(), LETTER)

        if self.options.pdf_orientation == "landscape":
            page_size = landscape(page_size)
        else:
            page_size = portrait(page_size)

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=page_size,
            title=self.options.pdf_title or "Finance Report",
            author=self.options.pdf_author,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        if self.options.pdf_title:
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                spaceAfter=30,
            )
            elements.append(Paragraph(self.options.pdf_title, title_style))

        # Add each sheet
        for sheet in sheets:
            # Sheet header
            elements.append(Paragraph(sheet.name, styles["Heading2"]))
            elements.append(Spacer(1, 12))

            if not sheet.rows:
                continue

            # Prepare table data
            table_data = []
            for row in sheet.rows:
                table_row = []
                for value in row:
                    if isinstance(value, Decimal):
                        table_row.append(f"${float(value):,.2f}")
                    elif isinstance(value, date):
                        table_row.append(value.strftime("%Y-%m-%d"))
                    elif value is None:
                        table_row.append("")
                    else:
                        table_row.append(str(value))
                table_data.append(table_row)

            if not table_data:
                continue

            # Calculate column widths
            available_width = page_size[0] - 2 * inch
            col_count = max(len(row) for row in table_data)
            col_width = available_width / col_count

            # Create table
            table = Table(table_data, colWidths=[col_width] * col_count)

            # Table style
            table_style = TableStyle(
                [
                    # Header row
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),  # First column left
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),  # Others right
                    # Grid
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    # Alternating rows
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.lightgrey],
                    ),
                    # Padding
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
            table.setStyle(table_style)

            elements.append(table)
            elements.append(Spacer(1, 20))

        # Summary section
        if self.options.pdf_include_summary:
            elements.append(Paragraph("Summary", styles["Heading2"]))
            summary_text = self._generate_pdf_summary(sheets)
            elements.append(Paragraph(summary_text, styles["Normal"]))

        # Build PDF
        doc.build(elements)
        return output_path

    def _generate_pdf_summary(self, sheets: list[SheetData]) -> str:
        """Generate summary text for PDF."""
        total_rows = sum(s.row_count for s in sheets)
        sheet_names = ", ".join(s.name for s in sheets)

        return (
            f"This report contains {len(sheets)} sheet(s): {sheet_names}. "
            f"Total of {total_rows} rows of data. "
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}."
        )

    def _export_json(self, sheets: list[SheetData], output_path: Path) -> Path:
        """Export to JSON format."""
        import json

        output_path.parent.mkdir(parents=True, exist_ok=True)

        export_data: dict[str, Any] = {
            "export_time": datetime.now().isoformat(),
            "sheets": [],
        }

        for sheet in sheets:
            sheet_dict: dict[str, Any] = {
                "name": sheet.name,
                "headers": sheet.headers,
                "row_count": sheet.row_count,
                "column_count": sheet.column_count,
                "data": [],
            }

            for row in sheet.rows:
                row_dict: dict[str, Any] = {}
                for i, value in enumerate(row):
                    header = sheet.headers[i] if i < len(sheet.headers) else f"col_{i}"
                    if isinstance(value, Decimal):
                        row_dict[header] = float(value)
                    elif isinstance(value, date):
                        row_dict[header] = value.isoformat()
                    else:
                        row_dict[header] = value
                sheet_dict["data"].append(row_dict)

            export_data["sheets"].append(sheet_dict)

        with open(output_path, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

        return output_path


def export_to_xlsx(
    ods_path: str | Path,
    output_path: str | Path,
    options: ExportOptions | None = None,
) -> Path:
    """Convenience function to export ODS to Excel.

    Args:
        ods_path: Path to ODS file.
        output_path: Path for XLSX output.
        options: Export options.

    Returns:
        Path to exported file.
    """
    exporter = MultiFormatExporter(options)
    return exporter.export(ods_path, output_path, MultiExportFormat.XLSX)


def export_to_csv(
    ods_path: str | Path,
    output_path: str | Path,
    options: ExportOptions | None = None,
) -> Path:
    """Convenience function to export ODS to CSV.

    Args:
        ods_path: Path to ODS file.
        output_path: Path for CSV output.
        options: Export options.

    Returns:
        Path to exported file.
    """
    exporter = MultiFormatExporter(options)
    return exporter.export(ods_path, output_path, MultiExportFormat.CSV)


def export_to_pdf(
    ods_path: str | Path,
    output_path: str | Path,
    options: ExportOptions | None = None,
) -> Path:
    """Convenience function to export ODS to PDF.

    Args:
        ods_path: Path to ODS file.
        output_path: Path for PDF output.
        options: Export options.

    Returns:
        Path to exported file.
    """
    exporter = MultiFormatExporter(options)
    return exporter.export(ods_path, output_path, MultiExportFormat.PDF)
