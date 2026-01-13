"""
Tests for multi-format export module.

Tests the MultiFormatExporter class and related functionality for
exporting ODS files to XLSX, CSV, PDF, and JSON formats.
"""

from __future__ import annotations

import csv
import json
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

import pytest

from spreadsheet_dl.exceptions import FileError
from spreadsheet_dl.export import (
    ExportDependencyError,
    ExportOptions,
    FormatNotSupportedError,
    MultiExportError,
    MultiExportFormat,
    MultiFormatExporter,
    SheetData,
    export_to_csv,
    export_to_pdf,
    export_to_xlsx,
)

if TYPE_CHECKING:
    from collections.abc import Generator

# Check if openpyxl is available
try:
    import openpyxl  # noqa: F401

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

pytestmark = [
    pytest.mark.integration,
    pytest.mark.requires_files,
    pytest.mark.requires_export,
    pytest.mark.rendering,
]


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    """Create a temporary directory for testing."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_sheet_data() -> SheetData:
    """Create sample sheet data for testing."""
    return SheetData(
        name="Budget",
        headers=["Category", "Budget", "Spent", "Remaining"],
        rows=[
            ["Category", "Budget", "Spent", "Remaining"],
            ["Housing", Decimal("1500.00"), Decimal("1450.00"), Decimal("50.00")],
            ["Food", Decimal("500.00"), Decimal("520.00"), Decimal("-20.00")],
            ["Transport", Decimal("300.00"), Decimal("250.00"), Decimal("50.00")],
        ],
    )


class TestMultiExportFormat:
    """Tests for MultiExportFormat enum."""

    def test_all_formats_have_values(self) -> None:
        """Test all export formats have string values."""
        for fmt in MultiExportFormat:
            assert isinstance(fmt.value, str)
            assert len(fmt.value) > 0

    def test_format_values(self) -> None:
        """Test specific format values."""
        assert MultiExportFormat.XLSX.value == "xlsx"
        assert MultiExportFormat.CSV.value == "csv"
        assert MultiExportFormat.PDF.value == "pdf"
        assert MultiExportFormat.JSON.value == "json"


class TestExportOptions:
    """Tests for ExportOptions class."""

    def test_default_options(self) -> None:
        """Test default option values."""
        options = ExportOptions()

        assert options.include_headers is True
        assert options.include_formulas is False
        assert options.preserve_formatting is True
        assert options.csv_delimiter == ","
        assert options.csv_encoding == "utf-8"

    def test_custom_options(self) -> None:
        """Test custom option values."""
        options = ExportOptions(
            include_headers=False,
            csv_delimiter=";",
            pdf_page_size="a4",
        )

        assert options.include_headers is False
        assert options.csv_delimiter == ";"
        assert options.pdf_page_size == "a4"


class TestSheetData:
    """Tests for SheetData class."""

    def test_row_count(self, sample_sheet_data: SheetData) -> None:
        """Test row count property."""
        assert sample_sheet_data.row_count == 4

    def test_column_count(self, sample_sheet_data: SheetData) -> None:
        """Test column count property."""
        assert sample_sheet_data.column_count == 4

    def test_empty_sheet(self) -> None:
        """Test empty sheet properties."""
        sheet = SheetData(name="Empty")
        assert sheet.row_count == 0
        assert sheet.column_count == 0


class TestMultiFormatExporter:
    """Tests for MultiFormatExporter class."""

    def test_init_defaults(self) -> None:
        """Test default initialization."""
        exporter = MultiFormatExporter()
        assert exporter.options.include_headers is True

    def test_init_custom_options(self) -> None:
        """Test initialization with custom options."""
        options = ExportOptions(include_headers=False)
        exporter = MultiFormatExporter(options)
        assert exporter.options.include_headers is False

    def test_export_file_not_found(self, temp_dir: Path) -> None:
        """Test export with non-existent file."""
        exporter = MultiFormatExporter()

        with pytest.raises(FileError):
            exporter.export(
                temp_dir / "nonexistent.ods",
                temp_dir / "output.xlsx",
                MultiExportFormat.XLSX,
            )

    def test_export_unsupported_format(self, temp_dir: Path) -> None:
        """Test export with unsupported format."""
        exporter = MultiFormatExporter()

        # Create a dummy file
        ods_file = temp_dir / "test.ods"
        ods_file.write_text("dummy")

        with pytest.raises(FormatNotSupportedError):
            exporter.export(ods_file, temp_dir / "output.xyz", "xyz")

    def test_export_format_string(self, temp_dir: Path) -> None:
        """Test export accepts format as string."""
        exporter = MultiFormatExporter()

        # This should not raise - it validates format parsing
        # Actual export would fail without odfpy
        with pytest.raises((FileError, ExportDependencyError)):
            exporter.export(
                temp_dir / "test.ods",
                temp_dir / "output.xlsx",
                "xlsx",  # String instead of enum
            )


class TestExportXLSX:
    """Tests for XLSX export functionality."""

    def test_export_xlsx_requires_openpyxl(
        self, temp_dir: Path, sample_sheet_data: SheetData
    ) -> None:
        """Test that XLSX export requires openpyxl."""
        exporter = MultiFormatExporter()

        try:
            import openpyxl  # noqa: F401

            # If openpyxl is available, export should work
            output_path = temp_dir / "output.xlsx"
            result = exporter._export_xlsx([sample_sheet_data], output_path)
            assert result.exists()
        except ImportError:
            # Expected if openpyxl not installed
            with pytest.raises(ExportDependencyError) as exc_info:
                exporter._export_xlsx([sample_sheet_data], temp_dir / "output.xlsx")
            assert "openpyxl" in str(exc_info.value)

    @pytest.mark.skipif(
        not HAS_OPENPYXL,
        reason="openpyxl required",
    )
    def test_export_xlsx_content(
        self, temp_dir: Path, sample_sheet_data: SheetData
    ) -> None:
        """Test XLSX export content."""
        from openpyxl import load_workbook

        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.xlsx"

        exporter._export_xlsx([sample_sheet_data], output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        assert ws.title == "Budget"
        assert ws.cell(1, 1).value == "Category"


class TestExportCSV:
    """Tests for CSV export functionality."""

    def test_export_csv_single_sheet(
        self, temp_dir: Path, sample_sheet_data: SheetData
    ) -> None:
        """Test CSV export for single sheet."""
        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.csv"

        result = exporter._export_csv([sample_sheet_data], output_path)

        assert result.exists()

        # Read and verify content
        with open(result) as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) == 4
        assert rows[0][0] == "Category"
        assert float(rows[1][1]) == 1500.00

    def test_export_csv_custom_delimiter(
        self, temp_dir: Path, sample_sheet_data: SheetData
    ) -> None:
        """Test CSV export with custom delimiter."""
        options = ExportOptions(csv_delimiter=";")
        exporter = MultiFormatExporter(options)
        output_path = temp_dir / "output.csv"

        result = exporter._export_csv([sample_sheet_data], output_path)

        with open(result) as f:
            content = f.read()

        assert ";" in content

    def test_export_csv_multiple_sheets(self, temp_dir: Path) -> None:
        """Test CSV export for multiple sheets."""
        sheet1 = SheetData(name="Sheet1", rows=[["A", "B"], ["1", "2"]])
        sheet2 = SheetData(name="Sheet2", rows=[["C", "D"], ["3", "4"]])

        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.csv"

        result = exporter._export_csv([sheet1, sheet2], output_path)

        # Combined file should exist
        assert result.exists()

        # Individual sheet files should exist in subdirectory
        csv_dir = temp_dir / "output"
        assert csv_dir.exists()
        assert (csv_dir / "Sheet1.csv").exists()
        assert (csv_dir / "Sheet2.csv").exists()


class TestExportPDF:
    """Tests for PDF export functionality."""

    def test_export_pdf_requires_reportlab(
        self, temp_dir: Path, sample_sheet_data: SheetData
    ) -> None:
        """Test that PDF export requires reportlab."""
        exporter = MultiFormatExporter()

        try:
            import reportlab  # noqa: F401

            # If reportlab is available, export should work
            output_path = temp_dir / "output.pdf"
            result = exporter._export_pdf([sample_sheet_data], output_path)
            assert result.exists()
        except ImportError:
            # Expected if reportlab not installed
            with pytest.raises(ExportDependencyError) as exc_info:
                exporter._export_pdf([sample_sheet_data], temp_dir / "output.pdf")
            assert "reportlab" in str(exc_info.value)

    def test_export_pdf_options(self, temp_dir: Path) -> None:
        """Test PDF export with custom options."""
        options = ExportOptions(
            pdf_page_size="a4",
            pdf_orientation="landscape",
            pdf_title="Test Report",
            pdf_author="Test Author",
        )
        exporter = MultiFormatExporter(options)

        assert exporter.options.pdf_page_size == "a4"
        assert exporter.options.pdf_orientation == "landscape"


class TestExportJSON:
    """Tests for JSON export functionality."""

    def test_export_json(self, temp_dir: Path, sample_sheet_data: SheetData) -> None:
        """Test JSON export."""
        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.json"

        result = exporter._export_json([sample_sheet_data], output_path)

        assert result.exists()

        with open(result) as f:
            data = json.load(f)

        assert "export_time" in data
        assert "sheets" in data
        assert len(data["sheets"]) == 1
        assert data["sheets"][0]["name"] == "Budget"

    def test_export_json_decimal_serialization(self, temp_dir: Path) -> None:
        """Test that Decimal values are serialized correctly."""
        sheet = SheetData(
            name="Test",
            headers=["Amount"],
            rows=[["Amount"], [Decimal("123.45")]],
        )

        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.json"

        exporter._export_json([sheet], output_path)

        with open(output_path) as f:
            data = json.load(f)

        # Decimal should be converted to float
        assert data["sheets"][0]["data"][1]["Amount"] == 123.45

    def test_export_json_date_serialization(self, temp_dir: Path) -> None:
        """Test that date values are serialized correctly."""
        sheet = SheetData(
            name="Test",
            headers=["Date"],
            rows=[["Date"], [date(2024, 12, 28)]],
        )

        exporter = MultiFormatExporter()
        output_path = temp_dir / "output.json"

        exporter._export_json([sheet], output_path)

        with open(output_path) as f:
            data = json.load(f)

        assert data["sheets"][0]["data"][1]["Date"] == "2024-12-28"


class TestExportBatch:
    """Tests for batch export functionality."""

    def test_export_batch(self, temp_dir: Path, sample_budget_file: Path) -> None:
        """Test batch export to multiple formats."""
        exporter = MultiFormatExporter()

        # Test batch export with CSV and JSON formats
        results = exporter.export_batch(
            sample_budget_file,
            temp_dir / "output",
            [MultiExportFormat.CSV, MultiExportFormat.JSON],
        )

        # Both formats should be in results
        assert MultiExportFormat.CSV.value in results
        assert MultiExportFormat.JSON.value in results
        # Exported files should exist
        assert results[MultiExportFormat.CSV.value] is not None
        assert results[MultiExportFormat.JSON.value] is not None


class TestConvenienceFunctions:
    """Tests for convenience export functions."""

    def test_export_to_xlsx_not_found(self, temp_dir: Path) -> None:
        """Test export_to_xlsx with non-existent file."""
        with pytest.raises(FileError):
            export_to_xlsx(
                temp_dir / "nonexistent.ods",
                temp_dir / "output.xlsx",
            )

    def test_export_to_csv_not_found(self, temp_dir: Path) -> None:
        """Test export_to_csv with non-existent file."""
        with pytest.raises(FileError):
            export_to_csv(
                temp_dir / "nonexistent.ods",
                temp_dir / "output.csv",
            )

    def test_export_to_pdf_not_found(self, temp_dir: Path) -> None:
        """Test export_to_pdf with non-existent file."""
        with pytest.raises(FileError):
            export_to_pdf(
                temp_dir / "nonexistent.ods",
                temp_dir / "output.pdf",
            )


class TestExportExceptions:
    """Tests for export exceptions."""

    def test_multi_export_error_base(self) -> None:
        """Test MultiExportError base class."""
        error = MultiExportError("Test error")
        assert "FT-MXP-1300" in error.error_code

    def test_format_not_supported_error(self) -> None:
        """Test FormatNotSupportedError."""
        error = FormatNotSupportedError("xyz")
        assert "FT-MXP-1301" in error.error_code
        assert "xyz" in str(error)

    def test_export_dependency_error(self) -> None:
        """Test ExportDependencyError."""
        error = ExportDependencyError("XLSX", "openpyxl", "pip install openpyxl")
        assert "FT-MXP-1302" in error.error_code
        assert "openpyxl" in str(error)
