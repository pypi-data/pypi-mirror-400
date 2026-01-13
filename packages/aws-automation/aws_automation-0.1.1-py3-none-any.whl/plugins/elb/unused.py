"""
plugins/elb/unused.py - 미사용 ELB 분석

타겟이 없거나 비정상인 Load Balancer 탐지

분석 기준:
- ALB/NLB: 타겟 그룹 없음 또는 등록된 타겟 없음 또는 모든 타겟 unhealthy
- CLB: 등록된 인스턴스 없음 또는 모든 인스턴스 unhealthy

월간 비용:
- ALB/NLB: ~$16.43/월 (고정) + LCU/NLCU
- CLB: ~$18.25/월 (고정) + 데이터 처리
- GWLB: ~$9.13/월 (고정)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_elb_monthly_cost

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticloadbalancing:DescribeLoadBalancers",
        "elasticloadbalancing:DescribeTargetGroups",
        "elasticloadbalancing:DescribeTargetHealth",
        "elasticloadbalancing:DescribeInstanceHealth",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용 (타겟 없음)
    UNHEALTHY = "unhealthy"  # 모든 타겟 비정상
    NORMAL = "normal"  # 정상 사용


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class TargetGroupInfo:
    """타겟 그룹 정보"""

    arn: str
    name: str
    target_type: str
    total_targets: int
    healthy_targets: int
    unhealthy_targets: int


@dataclass
class LoadBalancerInfo:
    """Load Balancer 정보"""

    arn: str
    name: str
    dns_name: str
    lb_type: str  # application, network, gateway, classic
    scheme: str  # internet-facing, internal
    state: str
    vpc_id: str
    availability_zones: list[str]
    created_time: datetime | None
    tags: dict[str, str]

    # 타겟 그룹 (ALB/NLB/GWLB)
    target_groups: list[TargetGroupInfo] = field(default_factory=list)

    # CLB 전용
    registered_instances: int = 0
    healthy_instances: int = 0

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def total_targets(self) -> int:
        """전체 타겟 수"""
        if self.lb_type == "classic":
            return self.registered_instances
        return sum(tg.total_targets for tg in self.target_groups)

    @property
    def healthy_targets(self) -> int:
        """정상 타겟 수"""
        if self.lb_type == "classic":
            return self.healthy_instances
        return sum(tg.healthy_targets for tg in self.target_groups)


@dataclass
class LBFinding:
    """LB 분석 결과"""

    lb: LoadBalancerInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class LBAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[LBFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    unhealthy_count: int = 0
    normal_count: int = 0

    # 비용
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집 - ALB/NLB/GWLB (elbv2)
# =============================================================================


def collect_v2_load_balancers(session, account_id: str, account_name: str, region: str) -> list[LoadBalancerInfo]:
    """ALB/NLB/GWLB 목록 수집"""
    from botocore.exceptions import ClientError

    load_balancers = []

    try:
        elbv2 = get_client(session, "elbv2", region_name=region)

        # Load Balancers 조회
        paginator = elbv2.get_paginator("describe_load_balancers")
        for page in paginator.paginate():
            for data in page.get("LoadBalancers", []):
                lb_arn = data.get("LoadBalancerArn", "")
                lb_type = data.get("Type", "application")

                # 태그 조회
                tags = {}
                try:
                    tag_response = elbv2.describe_tags(ResourceArns=[lb_arn])
                    for tag_desc in tag_response.get("TagDescriptions", []):
                        for t in tag_desc.get("Tags", []):
                            key = t.get("Key", "")
                            if not key.startswith("aws:"):
                                tags[key] = t.get("Value", "")
                except ClientError:
                    pass

                lb = LoadBalancerInfo(
                    arn=lb_arn,
                    name=data.get("LoadBalancerName", ""),
                    dns_name=data.get("DNSName", ""),
                    lb_type=lb_type,
                    scheme=data.get("Scheme", ""),
                    state=data.get("State", {}).get("Code", ""),
                    vpc_id=data.get("VpcId", ""),
                    availability_zones=[az.get("ZoneName", "") for az in data.get("AvailabilityZones", [])],
                    created_time=data.get("CreatedTime"),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=get_elb_monthly_cost(region, lb_type),
                )

                # 타겟 그룹 조회
                lb.target_groups = _get_target_groups(elbv2, lb_arn)
                load_balancers.append(lb)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} ELBv2 수집 오류: {error_code}[/yellow]")

    return load_balancers


def _get_target_groups(elbv2, lb_arn: str) -> list[TargetGroupInfo]:
    """LB에 연결된 타겟 그룹 조회"""
    from botocore.exceptions import ClientError

    target_groups = []

    try:
        response = elbv2.describe_target_groups(LoadBalancerArn=lb_arn)

        for tg in response.get("TargetGroups", []):
            tg_arn = tg.get("TargetGroupArn", "")

            # 타겟 헬스 조회
            healthy = 0
            unhealthy = 0
            total = 0

            try:
                health_response = elbv2.describe_target_health(TargetGroupArn=tg_arn)
                for target in health_response.get("TargetHealthDescriptions", []):
                    total += 1
                    state = target.get("TargetHealth", {}).get("State", "")
                    if state == "healthy":
                        healthy += 1
                    else:
                        unhealthy += 1
            except ClientError:
                pass

            target_groups.append(
                TargetGroupInfo(
                    arn=tg_arn,
                    name=tg.get("TargetGroupName", ""),
                    target_type=tg.get("TargetType", ""),
                    total_targets=total,
                    healthy_targets=healthy,
                    unhealthy_targets=unhealthy,
                )
            )

    except ClientError:
        pass

    return target_groups


# =============================================================================
# 수집 - CLB (elb)
# =============================================================================


def collect_classic_load_balancers(session, account_id: str, account_name: str, region: str) -> list[LoadBalancerInfo]:
    """Classic Load Balancer 목록 수집"""
    from botocore.exceptions import ClientError

    load_balancers = []

    try:
        elb = get_client(session, "elb", region_name=region)

        response = elb.describe_load_balancers()

        for data in response.get("LoadBalancerDescriptions", []):
            lb_name = data.get("LoadBalancerName", "")

            # 태그 조회
            tags = {}
            try:
                tag_response = elb.describe_tags(LoadBalancerNames=[lb_name])
                for tag_desc in tag_response.get("TagDescriptions", []):
                    for t in tag_desc.get("Tags", []):
                        key = t.get("Key", "")
                        if not key.startswith("aws:"):
                            tags[key] = t.get("Value", "")
            except ClientError:
                pass

            # 인스턴스 헬스 조회
            instances = data.get("Instances", [])
            healthy = 0
            try:
                if instances:
                    health_response = elb.describe_instance_health(LoadBalancerName=lb_name)
                    for state in health_response.get("InstanceStates", []):
                        if state.get("State") == "InService":
                            healthy += 1
            except ClientError:
                pass

            lb = LoadBalancerInfo(
                arn=f"arn:aws:elasticloadbalancing:{region}:{account_id}:loadbalancer/{lb_name}",
                name=lb_name,
                dns_name=data.get("DNSName", ""),
                lb_type="classic",
                scheme=data.get("Scheme", ""),
                state="active",  # CLB는 state 없음
                vpc_id=data.get("VPCId", ""),
                availability_zones=data.get("AvailabilityZones", []),
                created_time=data.get("CreatedTime"),
                tags=tags,
                registered_instances=len(instances),
                healthy_instances=healthy,
                account_id=account_id,
                account_name=account_name,
                region=region,
                monthly_cost=get_elb_monthly_cost(region, "classic"),
            )
            load_balancers.append(lb)

    except ClientError as e:
        # CLB가 없는 리전도 있음
        if "not available" not in str(e).lower():
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"    [yellow]{account_name}/{region} CLB 수집 오류: {error_code}[/yellow]")

    return load_balancers


# =============================================================================
# 분석
# =============================================================================


def analyze_load_balancers(
    load_balancers: list[LoadBalancerInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> LBAnalysisResult:
    """Load Balancer 미사용 분석"""
    result = LBAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for lb in load_balancers:
        finding = _analyze_single_lb(lb)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_monthly_cost += lb.monthly_cost
        elif finding.usage_status == UsageStatus.UNHEALTHY:
            result.unhealthy_count += 1
            result.unused_monthly_cost += lb.monthly_cost  # unhealthy도 낭비
        else:
            result.normal_count += 1

    result.total_count = len(load_balancers)
    return result


def _analyze_single_lb(lb: LoadBalancerInfo) -> LBFinding:
    """개별 LB 분석"""

    # 비활성 상태
    if lb.state not in ("active", ""):
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.LOW,
            description=f"비활성 상태 ({lb.state})",
            recommendation="상태 확인 또는 삭제 검토",
        )

    total = lb.total_targets
    healthy = lb.healthy_targets

    # 타겟 없음
    if total == 0:
        # 타겟 그룹 자체가 없는 경우 (ALB/NLB)
        if lb.lb_type != "classic" and not lb.target_groups:
            return LBFinding(
                lb=lb,
                usage_status=UsageStatus.UNUSED,
                severity=Severity.HIGH,
                description=f"타겟 그룹 없음 (${lb.monthly_cost:.2f}/월)",
                recommendation="타겟 그룹 연결 또는 삭제 검토",
            )

        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.HIGH,
            description=f"등록된 타겟 없음 (${lb.monthly_cost:.2f}/월)",
            recommendation="타겟 등록 또는 삭제 검토",
        )

    # 모든 타겟 unhealthy
    if healthy == 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.UNHEALTHY,
            severity=Severity.HIGH,
            description=f"모든 타겟 비정상 ({total}개 unhealthy)",
            recommendation="타겟 헬스체크 확인",
        )

    # 일부 unhealthy
    unhealthy = total - healthy
    if unhealthy > 0:
        return LBFinding(
            lb=lb,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.MEDIUM,
            description=f"일부 타겟 비정상 ({healthy}/{total} healthy)",
            recommendation="비정상 타겟 확인",
        )

    # 정상
    return LBFinding(
        lb=lb,
        usage_status=UsageStatus.NORMAL,
        severity=Severity.INFO,
        description=f"정상 ({healthy}개 healthy)",
        recommendation="정상 운영 중",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[LBAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.UNHEALTHY: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Load Balancer 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "unhealthy": sum(r.unhealthy_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    stats = [
        ("항목", "값"),
        ("전체 LB", totals["total"]),
        ("미사용", totals["unused"]),
        ("Unhealthy", totals["unhealthy"]),
        ("정상", totals["normal"]),
        ("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}"),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "Name",
        "Type",
        "Scheme",
        "Usage",
        "Severity",
        "Targets",
        "Healthy",
        "Monthly Cost ($)",
        "DNS Name",
        "Description",
        "Recommendation",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # 미사용/unhealthy만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.UNHEALTHY):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.lb.monthly_cost, reverse=True)

    for f in all_findings:
        lb = f.lb
        ws2.append(
            [
                lb.account_name,
                lb.region,
                lb.name,
                lb.lb_type.upper(),
                lb.scheme,
                f.usage_status.value,
                f.severity.value,
                lb.total_targets,
                lb.healthy_targets,
                round(lb.monthly_cost, 2),
                lb.dns_name,
                f.description,
                f.recommendation,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=6).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"ELB_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> LBAnalysisResult:
    """단일 계정/리전의 ELB 수집 및 분석 (병렬 실행용)"""
    v2_lbs = collect_v2_load_balancers(session, account_id, account_name, region)
    classic_lbs = collect_classic_load_balancers(session, account_id, account_name, region)
    all_lbs = v2_lbs + classic_lbs
    return analyze_load_balancers(all_lbs, account_id, account_name, region)


def run(ctx) -> None:
    """미사용 ELB 분석 실행"""
    console.print("[bold]미사용 ELB 분석 시작...[/bold]")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="elasticloadbalancing")

    all_results: list[LBAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not all_results:
        console.print("[yellow]분석할 ELB 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "unhealthy": sum(r.unhealthy_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "unused_cost": sum(r.unused_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 ELB: {totals['total']}개[/bold]")
    if totals["unused"] > 0:
        console.print(f"  [red bold]미사용: {totals['unused']}개[/red bold]")
    if totals["unhealthy"] > 0:
        console.print(f"  [yellow]Unhealthy: {totals['unhealthy']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")

    if totals["unused_cost"] > 0:
        console.print(f"\n  [red]미사용 월 비용: ${totals['unused_cost']:.2f}[/red]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("elb-unused").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
