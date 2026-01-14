import csv
import json
import os
import shutil
import unittest
from unittest.mock import patch, Mock, MagicMock

from ebi_eva_common_pyutils.config import WritableConfig
from openpyxl.reader.excel import load_workbook
from requests import HTTPError

from eva_sub_cli import SUB_CLI_CONFIG_FILE
from eva_sub_cli.exceptions.invalid_file_type_exception import InvalidFileTypeError
from eva_sub_cli.exceptions.metadata_template_version_exception import MetadataTemplateVersionException, \
    MetadataTemplateVersionNotFoundException
from eva_sub_cli.exceptions.submission_not_found_exception import SubmissionNotFoundException
from eva_sub_cli.exceptions.submission_status_exception import SubmissionStatusException
from eva_sub_cli.file_utils import is_vcf_file
from eva_sub_cli.metadata import EvaMetadataJson
from eva_sub_cli.orchestrator import orchestrate_process, VALIDATE, SUBMIT, DOCKER, check_validation_required, \
    verify_and_get_metadata_xlsx_version, get_metadata_xlsx_template_link, get_sub_cli_github_tags, \
    remove_non_vcf_files_from_metadata
from eva_sub_cli.submit import SUB_CLI_CONFIG_KEY_SUBMISSION_ID, SUB_CLI_CONFIG_KEY_SUBMISSION_UPLOAD_URL
from eva_sub_cli.validators.validator import READY_FOR_SUBMISSION_TO_EVA, ALL_VALIDATION_TASKS
from tests.test_utils import touch


class TestOrchestrator(unittest.TestCase):
    project_title = 'Example Project'
    resource_dir = os.path.join(os.path.dirname(__file__), 'resources')
    test_sub_dir = os.path.join(resource_dir, 'test_sub_dir')
    config_file = os.path.join(test_sub_dir, SUB_CLI_CONFIG_FILE)

    mapping_file = os.path.join(test_sub_dir, 'vcf_mapping_file.csv')
    metadata_json = os.path.join(test_sub_dir, 'sub_metadata.json')
    metadata_xlsx = os.path.join(test_sub_dir, 'sub_metadata.xlsx')
    metadata_xlsx_version = '3.0.0'
    metadata_xlsx_with_project_accession = os.path.join(test_sub_dir, 'EVA_Submission_test_with_project_accession.xlsx')
    metadata_xlsx_version_v2 = '2.0.1'
    metadata_xlsx_v2 = os.path.join(test_sub_dir, 'EVA_Submission_test_V2.xlsx')
    metadata_xlsx_version_missing = os.path.join(test_sub_dir, 'sub_metadata_version_missing.xlsx')
    metadata_json_with_non_vcf_files = os.path.join(test_sub_dir, 'EVA_Submission_test_with_non_vcf_files.json')
    metadata_xlsx_with_non_vcf_files = os.path.join(test_sub_dir, 'EVA_Submission_test_with_non_vcf_files.xlsx')

    def setUp(self) -> None:
        if os.path.exists(self.test_sub_dir):
            shutil.rmtree(self.test_sub_dir)
        os.makedirs(self.test_sub_dir)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test.json'), self.metadata_json)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test.xlsx'), self.metadata_xlsx)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_V2.xlsx'), self.metadata_xlsx_v2)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_with_project_accession.xlsx'),
                    self.metadata_xlsx_with_project_accession)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_version_missing.xlsx'),
                    self.metadata_xlsx_version_missing)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_with_non_vcf_files.json'),
                    self.metadata_json_with_non_vcf_files)
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_with_non_vcf_files.xlsx'),
                    self.metadata_xlsx_with_non_vcf_files)
        for file_name in ['example1.vcf.gz', 'example2.vcf', 'example3.vcf', 'GCA_000001405.27_fasta.fa']:
            touch(os.path.join(self.test_sub_dir, file_name))
        self.curr_wd = os.getcwd()
        os.chdir(self.test_sub_dir)

    def tearDown(self) -> None:
        os.chdir(self.curr_wd)
        if os.path.exists(self.test_sub_dir):
            shutil.rmtree(self.test_sub_dir)

    def test_remove_non_vcf_files_from_metadata_json(self):
        # assert non vcf files exist in metadata
        metadata = EvaMetadataJson(self.metadata_json_with_non_vcf_files)
        assert any(not is_vcf_file(f['fileName']) for f in metadata.files)

        # remove non vcf files from metadata
        remove_non_vcf_files_from_metadata(self.metadata_json_with_non_vcf_files, None)

        # assert non vcf files are removed from metadata
        metadata = EvaMetadataJson(self.metadata_json_with_non_vcf_files)
        assert all(is_vcf_file(f['fileName']) for f in metadata.files)

    def test_remove_non_vcf_files_from_metadata_xlsx(self):
        # assert non vcf files exist in metadata
        workbook = load_workbook(self.metadata_xlsx_with_non_vcf_files)
        files_sheet = workbook['Files']
        files_headers = {cell.value: cell.column - 1 for cell in files_sheet[1]}
        assert any(not is_vcf_file(row[files_headers['File Name']]) for row in files_sheet.iter_rows(min_row=2, values_only=True) if row[files_headers['File Name']] is not None)

        # remove non vcf files from metadata
        remove_non_vcf_files_from_metadata(None, self.metadata_xlsx_with_non_vcf_files)

        # assert non vcf files are removed from metadata
        workbook = load_workbook(self.metadata_xlsx_with_non_vcf_files)
        files_sheet = workbook['Files']
        files_headers = {cell.value: cell.column - 1 for cell in files_sheet[1]}
        assert all(is_vcf_file(row[files_headers['File Name']]) for row in files_sheet.iter_rows(min_row=2, values_only=True) if row[files_headers['File Name']] is not None)

    def test_check_validation_required(self):
        tasks = ['submit']

        sub_config = WritableConfig(self.test_sub_dir, 'config.yaml')
        sub_config.set(READY_FOR_SUBMISSION_TO_EVA, value=False)
        self.assertTrue(check_validation_required(tasks, sub_config))

        sub_config.set(READY_FOR_SUBMISSION_TO_EVA, value=True)

        self.assertFalse(check_validation_required(tasks, sub_config))

        with patch('eva_sub_cli.submission_ws.SubmissionWSClient.get_submission_status') as get_submission_status_mock:
            sub_config.set(READY_FOR_SUBMISSION_TO_EVA, value=True)
            sub_config.set(SUB_CLI_CONFIG_KEY_SUBMISSION_ID, value='test123')

            get_submission_status_mock.return_value = 'OPEN'
            self.assertFalse(check_validation_required(tasks, sub_config))

            get_submission_status_mock.return_value = 'FAILED'
            self.assertTrue(check_validation_required(tasks, sub_config))
            # A FAILED submission status reset the submission ID and submission URL
            self.assertEqual(sub_config.get(SUB_CLI_CONFIG_KEY_SUBMISSION_ID), None)
            self.assertEqual(sub_config.get(SUB_CLI_CONFIG_KEY_SUBMISSION_UPLOAD_URL), None)

            sub_config.set(READY_FOR_SUBMISSION_TO_EVA, value=True)
            sub_config.set(SUB_CLI_CONFIG_KEY_SUBMISSION_ID, value='test123')

            mock_response = Mock(status_code=404)
            get_submission_status_mock.side_effect = HTTPError(response=mock_response)
            with self.assertRaises(SubmissionNotFoundException):
                check_validation_required(tasks, sub_config)

            mock_response = Mock(status_code=500)
            get_submission_status_mock.side_effect = HTTPError(response=mock_response)
            with self.assertRaises(SubmissionStatusException):
                check_validation_required(tasks, sub_config)

    def test_orchestrate_validate(self):
        with patch('eva_sub_cli.orchestrator.get_vcf_files') as m_get_vcf, \
                patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch(
                    'eva_sub_cli.orchestrator.get_project_title_and_create_vcf_files_mapping') as m_get_project_title_and_create_vcf_files_mapping, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator:
            m_get_project_title_and_create_vcf_files_mapping.return_value = self.project_title, self.mapping_file
            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx, tasks=[VALIDATE], executor=DOCKER)
            m_get_project_title_and_create_vcf_files_mapping.assert_called_once_with(self.test_sub_dir,
                                                                                     None,
                                                                                     self.metadata_xlsx,
                                                                                     self.metadata_xlsx_version)
            m_get_vcf.assert_called_once_with(self.mapping_file)
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, None, self.metadata_xlsx, self.metadata_xlsx_version,
                validation_tasks=ALL_VALIDATION_TASKS, submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_validate_submit(self):
        with patch('eva_sub_cli.orchestrator.get_vcf_files') as m_get_vcf, \
                patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch(
                    'eva_sub_cli.orchestrator.get_project_title_and_create_vcf_files_mapping') as m_get_project_title_and_create_vcf_files_mapping, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.StudySubmitter') as m_submitter, \
                patch('eva_sub_cli.orchestrator.check_validation_required', return_value=True):
            # Empty config
            config = WritableConfig()
            m_config.return_value = config
            m_get_project_title_and_create_vcf_files_mapping.return_value = self.project_title, self.mapping_file

            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx, tasks=[SUBMIT], executor=DOCKER)
            m_get_vcf.assert_called_once_with(self.mapping_file)
            # Validate was run because the config show it was not run successfully before
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, None, self.metadata_xlsx,
                self.metadata_xlsx_version, validation_tasks=ALL_VALIDATION_TASKS,
                submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

            # Submit was created
            m_submitter.assert_any_call(self.test_sub_dir, submission_config=m_config.return_value,
                                        username=None, password=None)
            with m_submitter() as submitter:
                submitter.submit.assert_called_once_with()

    def test_orchestrate_submit_no_validate(self):
        with patch('eva_sub_cli.orchestrator.get_vcf_files') as m_get_vcf, \
                patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch(
                    'eva_sub_cli.orchestrator.get_project_title_and_create_vcf_files_mapping') as m_get_project_title_and_create_vcf_files_mapping, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.StudySubmitter') as m_submitter:
            # Empty config
            m_config.return_value = {READY_FOR_SUBMISSION_TO_EVA: True}
            m_get_project_title_and_create_vcf_files_mapping.return_value = self.project_title, self.mapping_file

            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx, tasks=[SUBMIT], executor=DOCKER)
            m_get_vcf.assert_called_once_with(self.mapping_file)
            # Validate was not run because the config showed it was run successfully before
            assert m_docker_validator.call_count == 0

            # Submit was created
            m_submitter.assert_any_call(self.test_sub_dir, submission_config=m_config.return_value,
                                        username=None, password=None)
            with m_submitter() as submitter:
                submitter.submit.assert_called_once_with()

    def test_orchestrate_with_metadata_json_without_asm_report(self):
        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator:
            orchestrate_process(self.test_sub_dir, self.metadata_json,
                                None, tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_json
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert row['report'] == ''
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, self.metadata_json, None, None,
                validation_tasks=ALL_VALIDATION_TASKS, submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_with_metadata_json_with_asm_report(self):
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_with_asm_report.json'), self.metadata_json)

        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.os.path.isfile'):
            orchestrate_process(self.test_sub_dir, self.metadata_json, None,
                                tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_json
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert 'GCA_000001405.27_report.txt' in row['report']
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, self.metadata_json, None, None,
                validation_tasks=ALL_VALIDATION_TASKS, submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_with_metadata_json_with_project_accession(self):
        shutil.copy(os.path.join(self.resource_dir, 'EVA_Submission_test_with_project_accession.json'), self.metadata_json)

        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.os.path.isfile'), \
                patch('eva_sub_cli.orchestrator.get_project_title_from_ena') as get_project_title_from_ena:
            get_project_title_from_ena.return_value = self.project_title
            orchestrate_process(self.test_sub_dir, self.metadata_json, None,
                                tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_json
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert 'GCA_000001405.27_report.txt' in row['report']
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, self.metadata_json, None, None,
                validation_tasks=ALL_VALIDATION_TASKS, submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_non_vcf_files_filtered_out(self):
        metadata_with_non_vcf_file = os.path.join(self.test_sub_dir, 'updated_metadata.json')
        with open(self.metadata_json) as f:
            metadata = json.load(f)
        metadata['files'].append({
            'analysisAlias': 'VD1',
            'fileName': 'test.vcf.gz.csi'
        })
        with open(metadata_with_non_vcf_file, 'w') as f:
            json.dump(metadata, f)
        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.os.path.isfile'):
            orchestrate_process(self.test_sub_dir, metadata_with_non_vcf_file,
                                None, tasks=[VALIDATE], executor=DOCKER, resume=False)
            # Mapping file was created from the metadata_json
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                vcf_files = set()
                for row in reader:
                    vcf_files.add(row['vcf'])
                assert len(vcf_files) == 3
                assert 'test.vcf.gz.csi' not in vcf_files
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, metadata_with_non_vcf_file, None, None,
                validation_tasks=ALL_VALIDATION_TASKS, submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_with_metadata_xlsx(self):
        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator:
            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx,
                                tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_xlsx
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert row['report'] == ''
            m_docker_validator.assert_any_call(
                self.mapping_file, self.test_sub_dir, self.project_title, None, self.metadata_xlsx,
                self.metadata_xlsx_version, validation_tasks=ALL_VALIDATION_TASKS,
                submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_with_metadata_xlsx_v2(self):
        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator:
            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx_v2,
                                tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_xlsx
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert row['report'] == ''
            m_docker_validator.assert_any_call(self.mapping_file, self.test_sub_dir, self.project_title, None,
                                               self.metadata_xlsx_v2, self.metadata_xlsx_version_v2,
                                               validation_tasks=ALL_VALIDATION_TASKS,
                                               submission_config=m_config.return_value, shallow_validation=False
                                               )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_orchestrate_with_metadata_xlsx_having_project_accession(self):
        with patch('eva_sub_cli.orchestrator.WritableConfig') as m_config, \
                patch('eva_sub_cli.orchestrator.DockerValidator') as m_docker_validator, \
                patch('eva_sub_cli.orchestrator.get_project_title_from_ena') as get_project_title_from_ena:
            get_project_title_from_ena.return_value = self.project_title

            orchestrate_process(self.test_sub_dir, None, self.metadata_xlsx_with_project_accession,
                                tasks=[VALIDATE], executor=DOCKER)
            # Mapping file was created from the metadata_xlsx
            assert os.path.exists(self.mapping_file)
            with open(self.mapping_file) as open_file:
                reader = csv.DictReader(open_file, delimiter=',')
                for row in reader:
                    assert 'example' in row['vcf']
                    assert row['report'] == ''
            m_docker_validator.assert_any_call(self.mapping_file, self.test_sub_dir, self.project_title, None,
                self.metadata_xlsx_with_project_accession, self.metadata_xlsx_version, validation_tasks=ALL_VALIDATION_TASKS,
                submission_config=m_config.return_value, shallow_validation=False
            )
            m_docker_validator().validate_and_report.assert_called_once_with()

    def test_metadata_file_does_not_exist_error(self):
        with self.assertRaises(FileNotFoundError) as context:
            orchestrate_process(self.test_sub_dir, None, 'Non_existing_metadata.xlsx',
                                tasks=[VALIDATE], executor=DOCKER)
        self.assertRegex(
            str(context.exception),
            r"The provided metadata file .*/resources/test_sub_dir/Non_existing_metadata.xlsx does not exist"
        )

    def test_fasta_file_compressed(self):
        metadata_with_compressed_fasta = os.path.join(self.test_sub_dir, 'updated_metadata.json')
        with open(self.metadata_json) as f:
            metadata = json.load(f)
        for analysis in metadata['analysis']:
            analysis['referenceFasta'] = os.path.join(self.test_sub_dir, 'genome.fa.gz')
        with open(metadata_with_compressed_fasta, 'w') as f:
            json.dump(metadata, f)
        with patch('eva_sub_cli.orchestrator.os.path.isfile'):
            with self.assertRaises(InvalidFileTypeError):
                orchestrate_process(self.test_sub_dir, metadata_with_compressed_fasta,
                                    None, tasks=[VALIDATE], executor=DOCKER)

    def test_get_sub_cli_github_tags(self):
        with patch("requests.get") as mock_req:
            mock_response = MagicMock()
            mock_response.status_code = 500
            mock_response.raise_for_status.side_effect = HTTPError("Internal Server Error")
            mock_req.return_value = mock_response

            assert get_sub_cli_github_tags() == []

    def test_get_metadata_xlsx_template_link(self):
        with patch('eva_sub_cli.orchestrator.get_sub_cli_version') as sub_cli_version, \
                patch('eva_sub_cli.orchestrator.get_sub_cli_github_tags') as sub_cli_tags:
            sub_cli_version.return_value = '1.1.6'
            sub_cli_tags.return_value = ['1.1.6']
            assert get_metadata_xlsx_template_link() == 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/refs/tags/v1.1.6/eva_sub_cli/etc/EVA_Submission_template.xlsx'

            sub_cli_version.return_value = '1.1.5'
            sub_cli_tags.return_value = ['1.1.6']
            assert get_metadata_xlsx_template_link() == 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/main/eva_sub_cli/etc/EVA_Submission_template.xlsx'

            sub_cli_version.return_value = '1.1.5'
            sub_cli_tags.return_value = []
            assert get_metadata_xlsx_template_link() == 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/main/eva_sub_cli/etc/EVA_Submission_template.xlsx'

    def test_metadata_xlsx_version_should_pass_as_version_is_equal_to_min_required(self):
        verify_and_get_metadata_xlsx_version(self.metadata_xlsx, '1.1.6')

    def test_metadata_xlsx_version_should_pass_as_version_is_greater_than_min_required(self):
        verify_and_get_metadata_xlsx_version(self.metadata_xlsx, '1.1.5')

    def test_metadata_xlsx_version_should_fail_as_version_is_lower_than_min_required(self):
        with patch('eva_sub_cli.orchestrator.get_metadata_xlsx_template_link') as template_link:
            template_link.return_value = 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/v0.4.4/eva-sub-cli/eva_sub_cli/etc/EVA_Submission_template.xlsx'
            try:
                verify_and_get_metadata_xlsx_version(self.metadata_xlsx, '1.1.8')
            except MetadataTemplateVersionException as mte:
                assert mte.message == ("Metadata template version 1.1.6 is lower than min required 1.1.8. "
                                       "Please download the correct template from EVA github project "
                                       "https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/v0.4.4/eva-sub-cli/eva_sub_cli/etc/EVA_Submission_template.xlsx")

    def test_metadata_xlsx_version_should_fail_as_version_is_not_found(self):
        with patch('eva_sub_cli.orchestrator.get_metadata_xlsx_template_link') as template_link:
            template_link.return_value = 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/v0.4.4/eva-sub-cli/eva_sub_cli/etc/EVA_Submission_template.xlsx'
            try:
                verify_and_get_metadata_xlsx_version(self.metadata_xlsx_version_missing, '1.1.8')
            except MetadataTemplateVersionNotFoundException as mte:
                assert mte.message == (
                    f"No version information found in metadata xlsx sheet {self.metadata_xlsx_version_missing}. "
                    f"Please download the correct template from EVA github project "
                    f"https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/v0.4.4/eva-sub-cli/eva_sub_cli/etc/EVA_Submission_template.xlsx")
