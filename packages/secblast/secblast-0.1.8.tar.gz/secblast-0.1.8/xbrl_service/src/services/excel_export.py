"""
Excel export service for financial data.
Generates .xlsx files with financial statements organized by presentation hierarchy.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Dict, Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exports financial data to Excel format."""

    # Styling
    HEADER_FONT = Font(bold=True, size=12)
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(bold=True, size=12, color="4472C4")
    ABSTRACT_FONT = Font(bold=True, size=11)
    TOTAL_FONT = Font(bold=True)
    LINK_FONT = Font(color="0563C1", underline="single")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
    LIGHT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="000000")
    )

    def export_filing(
        self,
        data: Dict[str, Any],
        cik: int,
        accession_number: str,
        filing_date: date,
        form_type: str,
        company_name: str = "",
    ) -> BytesIO:
        """
        Export a filing's financial data to Excel.
        Creates separate sheets for each statement type.
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        metadata = data.get("metadata", {})
        presentation = data.get("presentation_trees", {})
        all_facts = data.get("all_facts", {})
        labels = data.get("labels", {})

        # Create cover sheet first
        self._create_cover_sheet(wb, cik, accession_number, filing_date, form_type, company_name, metadata)

        # Create sheets for each statement type (3 core statements only)
        statement_types = [
            ("Balance Sheet", "balance_sheet"),
            ("Income Statement", "income_statement"),
            ("Cash Flow", "cash_flow"),
        ]

        for sheet_name, tree_key in statement_types:
            trees = presentation.get(tree_key, [])
            if trees:
                self._create_statement_sheet(
                    wb, sheet_name, trees, all_facts, labels, filing_date, form_type, company_name
                )

        # Create calculations sheet
        calculations = data.get("calculation_relationships", {})
        if calculations:
            self._create_calculations_sheet(wb, calculations, all_facts, labels)

        # Ensure at least one sheet exists
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet("Info")
            ws["A1"] = "No financial data available"

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_statement_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        trees: List[Dict],
        all_facts: Dict[str, List[Dict]],
        labels: Dict[str, str],
        filing_date: date,
        form_type: str,
        company_name: str = "",
    ) -> None:
        """Create a sheet for a financial statement."""
        ws = wb.create_sheet(sheet_name)

        # Company name header
        if company_name:
            ws["A1"] = company_name
            ws["A1"].font = self.TITLE_FONT
            ws["A2"] = sheet_name
            ws["A2"].font = self.SUBTITLE_FONT
            ws["A3"] = f"As of {filing_date.strftime('%B %d, %Y')} ({form_type})"
            ws["A3"].font = Font(italic=True)
            header_row = 5
        else:
            ws["A1"] = sheet_name
            ws["A1"].font = self.TITLE_FONT
            ws["A2"] = f"As of {filing_date.strftime('%B %d, %Y')} ({form_type})"
            ws["A2"].font = Font(italic=True)
            header_row = 4

        # Column headers
        ws.cell(row=header_row, column=1, value="Line Item").font = self.HEADER_FONT_WHITE
        ws.cell(row=header_row, column=2, value="Value").font = self.HEADER_FONT_WHITE
        ws.cell(row=header_row, column=3, value="Unit").font = self.HEADER_FONT_WHITE
        ws.cell(row=header_row, column=1).fill = self.HEADER_FILL
        ws.cell(row=header_row, column=2).fill = self.HEADER_FILL
        ws.cell(row=header_row, column=3).fill = self.HEADER_FILL

        # Write statement data
        row = header_row + 1
        for tree in trees:
            row = self._write_tree_node(ws, tree, all_facts, labels, row, level=0)

        # Adjust column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10

    def _write_tree_node(
        self,
        ws,
        node: Dict,
        all_facts: Dict[str, List[Dict]],
        labels: Dict[str, str],
        row: int,
        level: int = 0,
    ) -> int:
        """Recursively write a presentation tree node to the worksheet."""
        concept = node.get("concept", "")
        label = node.get("label", labels.get(concept, concept))
        is_abstract = node.get("is_abstract", False)
        children = node.get("children", [])

        # Get fact value
        value = None
        unit = ""
        if not is_abstract and concept in all_facts:
            facts = all_facts[concept]
            # Get first fact without dimensions
            no_dim = [f for f in facts if not f.get("dimensions")]
            if no_dim:
                value = no_dim[0].get("value")
                unit = no_dim[0].get("unit", "")
            elif facts:
                value = facts[0].get("value")
                unit = facts[0].get("unit", "")

        # Write row
        indent = "  " * level
        ws.cell(row=row, column=1, value=f"{indent}{label}")

        if is_abstract:
            ws.cell(row=row, column=1).font = self.ABSTRACT_FONT
        elif "total" in label.lower():
            ws.cell(row=row, column=1).font = self.TOTAL_FONT
            ws.cell(row=row, column=2).font = self.TOTAL_FONT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

        if value is not None:
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3, value=unit)

        row += 1

        # Write children
        for child in children:
            row = self._write_tree_node(ws, child, all_facts, labels, row, level + 1)

        return row

    def _create_calculations_sheet(
        self,
        wb: Workbook,
        calculations: Dict[str, Any],
        all_facts: Dict[str, List[Dict]],
        labels: Dict[str, str],
    ) -> None:
        """Create a sheet showing calculation relationships."""
        ws = wb.create_sheet("Calculations")

        ws["A1"] = "Calculation Relationships"
        ws["A1"].font = self.TITLE_FONT
        ws["A2"] = "Shows how financial line items are calculated from their components"
        ws["A2"].font = Font(italic=True)

        row = 4

        for parent_concept, calc_data in calculations.items():
            parent_label = calc_data.get("parent_label", parent_concept)
            children = calc_data.get("children", [])

            if not children:
                continue

            # Parent row (result)
            ws.cell(row=row, column=1, value=parent_label)
            ws.cell(row=row, column=1).font = self.TOTAL_FONT

            # Get parent value
            if parent_concept in all_facts:
                facts = all_facts[parent_concept]
                no_dim = [f for f in facts if not f.get("dimensions")]
                if no_dim:
                    ws.cell(row=row, column=2, value=no_dim[0].get("value"))
                    ws.cell(row=row, column=2).number_format = '#,##0'

            row += 1

            # Child rows (components)
            for child in children:
                child_concept = child.get("concept", "")
                child_label = child.get("label", labels.get(child_concept, child_concept))
                weight = child.get("weight", 1.0)

                # Get the fact value
                child_value = None
                if child_concept in all_facts:
                    facts = all_facts[child_concept]
                    no_dim = [f for f in facts if not f.get("dimensions")]
                    if no_dim:
                        child_value = no_dim[0].get("value")

                # Calculate effective contribution (weight * value)
                if child_value is not None:
                    effective_value = weight * child_value
                    sign = "+" if effective_value >= 0 else "-"
                else:
                    effective_value = None
                    sign = "+" if weight >= 0 else "-"

                ws.cell(row=row, column=1, value=f"  {sign} {child_label}")

                if effective_value is not None:
                    ws.cell(row=row, column=2, value=effective_value)
                    # Use accounting format that shows negatives in parentheses
                    ws.cell(row=row, column=2).number_format = '#,##0;(#,##0)'

                row += 1

            row += 1  # Blank row between calculations

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20

    def _create_cover_sheet(
        self,
        wb: Workbook,
        cik: int,
        accession_number: str,
        filing_date: date,
        form_type: str,
        company_name: str,
        metadata: Dict[str, Any],
    ) -> None:
        """Create a cover sheet with branding, links, and filing metadata."""
        ws = wb.create_sheet("Summary", 0)  # Insert at first position

        # SecBlast branding
        ws["A1"] = "SecBlast Financial Data Export"
        ws["A1"].font = Font(bold=True, size=16, color="4472C4")

        ws["A2"] = "www.secblast.com"
        ws["A2"].font = self.LINK_FONT
        ws["A2"].hyperlink = "https://www.secblast.com"

        # Company info with link
        row = 4
        if company_name:
            ws.cell(row=row, column=1, value="Company")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=company_name)
            ws.cell(row=row, column=2).font = self.LINK_FONT
            ws.cell(row=row, column=2).hyperlink = f"https://www.secblast.com/cik/{cik}"
            row += 1

        # Filing info
        info_rows = [
            ("CIK", str(cik)),
            ("Accession Number", accession_number),
            ("Form Type", form_type),
            ("Filing Date", filing_date.strftime("%Y-%m-%d")),
            ("Fiscal Year", str(metadata.get("fiscal_year", "N/A"))),
            ("Fiscal Period", metadata.get("fiscal_period", "N/A")),
        ]

        period_end = metadata.get("document_period_end")
        if period_end:
            if isinstance(period_end, date):
                info_rows.append(("Period End", period_end.strftime("%Y-%m-%d")))
            else:
                info_rows.append(("Period End", str(period_end)))

        for label, value in info_rows:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # SEC filing link
        row += 1
        ws.cell(row=row, column=1, value="SEC Filing")
        ws.cell(row=row, column=1).font = Font(bold=True)
        sec_url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type={form_type}&dateb=&owner=include&count=40"
        ws.cell(row=row, column=2, value="View on SEC.gov")
        ws.cell(row=row, column=2).font = self.LINK_FONT
        ws.cell(row=row, column=2).hyperlink = sec_url

        # Sheet contents - only show sheets that have data
        row += 2
        ws.cell(row=row, column=1, value="Contents")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        # List actual sheets (will be created if data exists)
        for content in ["Balance Sheet", "Income Statement", "Cash Flow", "Calculations"]:
            ws.cell(row=row, column=1, value=f"  • {content}")
            row += 1

        # Export timestamp
        row += 1
        ws.cell(row=row, column=1, value="Exported")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC"))

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40

    def export_history(
        self,
        cik: int,
        history: Dict[str, List[Dict]],
        company_name: str = "",
    ) -> BytesIO:
        """
        Export historical data for selected metrics to Excel.
        One row per period, columns for each metric.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Historical Data"

        # Title
        ws["A1"] = f"Historical Financial Data - CIK {cik}"
        if company_name:
            ws["A1"] = f"Historical Financial Data - {company_name}"
        ws["A1"].font = self.TITLE_FONT

        # Headers
        headers = ["Period End", "Filing Date", "Fiscal Year", "Form Type"]
        concept_names = list(history.keys())
        headers.extend(concept_names)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL

        # Collect all periods
        all_periods = {}
        for concept, data_points in history.items():
            for dp in data_points:
                period_key = str(dp.get("period_end"))
                if period_key not in all_periods:
                    all_periods[period_key] = {
                        "period_end": dp.get("period_end"),
                        "filing_date": dp.get("filing_date"),
                        "fiscal_year": dp.get("fiscal_year"),
                        "form_type": dp.get("form_type"),
                    }
                all_periods[period_key][concept] = dp.get("value")

        # Sort by period end descending
        sorted_periods = sorted(all_periods.values(), key=lambda x: str(x.get("period_end", "")), reverse=True)

        # Write data rows
        row = 4
        for period_data in sorted_periods:
            ws.cell(row=row, column=1, value=str(period_data.get("period_end", "")))
            ws.cell(row=row, column=2, value=str(period_data.get("filing_date", "")))
            ws.cell(row=row, column=3, value=period_data.get("fiscal_year", ""))
            ws.cell(row=row, column=4, value=period_data.get("form_type", ""))

            for col_idx, concept in enumerate(concept_names, 5):
                value = period_data.get(concept)
                if value is not None:
                    ws.cell(row=row, column=col_idx, value=value)
                    ws.cell(row=row, column=col_idx).number_format = '#,##0'

            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        for col_idx in range(5, 5 + len(concept_names)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
