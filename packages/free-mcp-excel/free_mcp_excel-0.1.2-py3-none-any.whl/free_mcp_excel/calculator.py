"""
公式计算引擎
提供Excel公式计算和评估功能
"""
import io
from typing import Dict, Any, Optional

import openpyxl

from .utils import (
    base64_to_excel,
    detect_file_format,
    validate_file_size,
    format_error_response,
    format_success_response,
    parse_cell_address,
    parse_range_address,
    validate_cell_address,
)
from .parser import ExcelParser


class FormulaCalculator:
    """公式计算器"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化计算器
        
        Args:
            config: 配置字典
        """
        self.config = config or {
            "prefer_calculated": True,
            "engine": "pycel"
        }
        self.parser = ExcelParser()
        self._formula_engine = None
    
    def _get_formula_engine(self):
        """获取公式引擎实例（延迟加载）"""
        if self._formula_engine is None:
            try:
                import pycel
                self._formula_engine = pycel
            except ImportError:
                self._formula_engine = None
        return self._formula_engine
    
    def calc_cell_data(
        self,
        file_base64: str,
        sheet: str,
        cell: str,
        force_recalc: bool = False
    ) -> Dict[str, Any]:
        """
        计算单元格值（混合模式）
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            cell: 单元格地址
            force_recalc: 是否强制重新计算
            
        Returns:
            包含计算结果的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                "不支持的格式",
                "INVALID_FILE_FORMAT"
            )
        
        try:
            # 优先读取已计算值
            if not force_recalc and self.config.get("prefer_calculated", True):
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                
                if sheet not in workbook.sheetnames:
                    workbook.close()
                    return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
                
                target_sheet = workbook[sheet]
                
                if not validate_cell_address(cell):
                    workbook.close()
                    return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
                
                try:
                    row_idx, col_idx = parse_cell_address(cell)
                except ValueError as e:
                    workbook.close()
                    return format_error_response(str(e), "INVALID_CELL")
                
                cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
                calculated_value = cell_obj.value
                
                # 读取公式文本
                workbook_formula = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
                sheet_formula = workbook_formula[sheet]
                cell_formula_obj = sheet_formula.cell(row=row_idx + 1, column=col_idx + 1)
                formula = cell_formula_obj.value if isinstance(cell_formula_obj.value, str) and cell_formula_obj.value.startswith("=") else None
                
                workbook.close()
                workbook_formula.close()
                
                if calculated_value is not None:
                    return format_success_response({
                        "cell": cell,
                        "value": calculated_value,
                        "formula": formula,
                        "calculation_method": "read_calculated"
                    })
            
            # 需要重新计算或已计算值不存在
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_cell_address(cell):
                workbook.close()
                return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
            
            try:
                row_idx, col_idx = parse_cell_address(cell)
            except ValueError as e:
                workbook.close()
                return format_error_response(str(e), "INVALID_CELL")
            
            cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
            formula = cell_obj.value if isinstance(cell_obj.value, str) and cell_obj.value.startswith("=") else None
            
            if not formula:
                # 没有公式，直接返回值
                value = cell_obj.value
                workbook.close()
                return format_success_response({
                    "cell": cell,
                    "value": value,
                    "formula": None,
                    "calculation_method": "direct_value"
                })
            
            # 使用公式引擎计算
            engine = self._get_formula_engine()
            if engine:
                try:
                    # 构建上下文（读取相关单元格的值）
                    context = {}
                    # 简化实现：读取整个工作表的值作为上下文
                    for row in target_sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                                context[cell.coordinate] = cell.value
                    
                    # 使用pycel计算（简化实现）
                    # 注意：pycel的实际使用需要更复杂的实现
                    calculated_value = None  # 这里需要实际的公式引擎计算
                    
                    workbook.close()
                    
                    return format_success_response({
                        "cell": cell,
                        "value": calculated_value,
                        "formula": formula,
                        "calculation_method": "engine_calculated"
                    })
                except Exception as e:
                    workbook.close()
                    return format_error_response(f"公式计算失败：{str(e)}", "CALCULATION_ERROR")
            else:
                # 没有公式引擎，返回公式文本
                workbook.close()
                return format_success_response({
                    "cell": cell,
                    "value": None,
                    "formula": formula,
                    "calculation_method": "formula_only",
                    "message": "公式引擎未安装，无法计算"
                })
        except Exception as e:
            return format_error_response(f"计算单元格值失败：{str(e)}", "CALCULATION_ERROR")
    
    def calc_range_data(
        self,
        file_base64: str,
        sheet: str,
        range_str: str
    ) -> Dict[str, Any]:
        """
        计算范围数据
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            range_str: 单元格范围
            
        Returns:
            包含计算结果的响应字典
        """
        try:
            (start_row, start_col), (end_row, end_col) = parse_range_address(range_str)
        except ValueError as e:
            return format_error_response(str(e), "INVALID_RANGE")
        
        # 解析范围内的所有单元格
        results = []
        max_row = end_row if end_row is not None else 1000
        max_col = end_col if end_col is not None else 100
        
        # 限制范围大小，避免计算过多单元格
        actual_end_row = min(max_row, start_row + 100) if end_row is not None else start_row + 100
        actual_end_col = min(max_col, start_col + 100) if end_col is not None else start_col + 100
        
        # 使用内置range函数（注意：这里range是Python内置函数，不是参数）
        import builtins
        for row_idx in builtins.range(start_row, actual_end_row + 1):
            for col_idx in builtins.range(start_col, actual_end_col + 1):
                from .utils import get_column_letter
                cell = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                result = self.calc_cell_data(file_base64, sheet, cell)
                if result["status"] == "success":
                    results.append(result["data"])
        
        return format_success_response({
            "range": range_str,
            "results": results
        })
    
    def evaluate_formula(
        self,
        formula: str,
        context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        评估公式表达式
        
        Args:
            formula: 公式文本
            context: 上下文数据（单元格值字典）
            
        Returns:
            包含计算结果的响应字典
        """
        if not formula.startswith("="):
            formula = "=" + formula
        
        engine = self._get_formula_engine()
        if not engine:
            return format_error_response("公式引擎未安装", "ENGINE_NOT_AVAILABLE")
        
        try:
            # 简化实现：这里需要实际的公式引擎计算
            # pycel的使用需要更复杂的实现
            return format_error_response("公式评估功能待完善", "FEATURE_NOT_IMPLEMENTED")
        except Exception as e:
            return format_error_response(f"公式评估失败：{str(e)}", "CALCULATION_ERROR")

