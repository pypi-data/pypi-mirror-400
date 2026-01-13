"""
plugins/rds/unused.py - RDS 유휴 인스턴스 분석

유휴/저사용 RDS 인스턴스 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 미사용 기준: 7일간 연결 수 평균 0
UNUSED_DAYS_THRESHOLD = 7
# 저사용 기준: CPU 평균 5% 미만
LOW_USAGE_CPU_THRESHOLD = 5.0

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "rds:DescribeDBInstances",
        "cloudwatch:GetMetricStatistics",
    ],
}


class InstanceStatus(Enum):
    """인스턴스 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    LOW_USAGE = "low_usage"
    STOPPED = "stopped"


@dataclass
class RDSInstanceInfo:
    """RDS 인스턴스 정보"""

    account_id: str
    account_name: str
    region: str
    db_instance_id: str
    db_instance_class: str
    engine: str
    engine_version: str
    status: str
    multi_az: bool
    storage_type: str
    allocated_storage: int  # GB
    created_at: datetime | None
    # CloudWatch 지표
    avg_connections: float = 0.0
    avg_cpu: float = 0.0
    avg_read_iops: float = 0.0
    avg_write_iops: float = 0.0

    @property
    def estimated_monthly_cost(self) -> float:
        """대략적인 월간 비용 추정"""
        # 간단한 가격 맵 (db.t3 기준, Multi-AZ 2배)
        price_map = {
            "db.t3.micro": 0.017,
            "db.t3.small": 0.034,
            "db.t3.medium": 0.068,
            "db.t3.large": 0.136,
            "db.t4g.micro": 0.016,
            "db.t4g.small": 0.032,
            "db.t4g.medium": 0.065,
            "db.r5.large": 0.24,
            "db.r5.xlarge": 0.48,
            "db.r6g.large": 0.218,
            "db.m5.large": 0.171,
            "db.m6g.large": 0.155,
        }
        hourly = price_map.get(self.db_instance_class, 0.15)
        if self.multi_az:
            hourly *= 2
        instance_cost = hourly * 730

        # 스토리지 비용 (gp2 기준 약 $0.115/GB)
        storage_cost = self.allocated_storage * 0.115

        return instance_cost + storage_cost


@dataclass
class InstanceFinding:
    """인스턴스 분석 결과"""

    instance: RDSInstanceInfo
    status: InstanceStatus
    recommendation: str


@dataclass
class RDSAnalysisResult:
    """RDS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_instances: int = 0
    unused_instances: int = 0
    low_usage_instances: int = 0
    stopped_instances: int = 0
    normal_instances: int = 0
    unused_monthly_cost: float = 0.0
    low_usage_monthly_cost: float = 0.0
    findings: list[InstanceFinding] = field(default_factory=list)


def collect_rds_instances(session, account_id: str, account_name: str, region: str) -> list[RDSInstanceInfo]:
    """RDS 인스턴스 수집"""
    from botocore.exceptions import ClientError

    rds = get_client(session, "rds", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    instances = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = rds.get_paginator("describe_db_instances")
        for page in paginator.paginate():
            for db in page.get("DBInstances", []):
                db_id = db.get("DBInstanceIdentifier", "")

                instance = RDSInstanceInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    db_instance_id=db_id,
                    db_instance_class=db.get("DBInstanceClass", ""),
                    engine=db.get("Engine", ""),
                    engine_version=db.get("EngineVersion", ""),
                    status=db.get("DBInstanceStatus", ""),
                    multi_az=db.get("MultiAZ", False),
                    storage_type=db.get("StorageType", ""),
                    allocated_storage=db.get("AllocatedStorage", 0),
                    created_at=db.get("InstanceCreateTime"),
                )

                # 정지된 인스턴스는 CloudWatch 지표 조회 불필요
                if instance.status == "stopped":
                    instances.append(instance)
                    continue

                # CloudWatch 지표 조회
                try:
                    # DatabaseConnections
                    conn_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/RDS",
                        MetricName="DatabaseConnections",
                        Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Average"],
                    )
                    if conn_resp.get("Datapoints"):
                        instance.avg_connections = sum(d["Average"] for d in conn_resp["Datapoints"]) / len(
                            conn_resp["Datapoints"]
                        )

                    # CPUUtilization
                    cpu_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/RDS",
                        MetricName="CPUUtilization",
                        Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Average"],
                    )
                    if cpu_resp.get("Datapoints"):
                        instance.avg_cpu = sum(d["Average"] for d in cpu_resp["Datapoints"]) / len(
                            cpu_resp["Datapoints"]
                        )

                    # ReadIOPS
                    read_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/RDS",
                        MetricName="ReadIOPS",
                        Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Average"],
                    )
                    if read_resp.get("Datapoints"):
                        instance.avg_read_iops = sum(d["Average"] for d in read_resp["Datapoints"]) / len(
                            read_resp["Datapoints"]
                        )

                    # WriteIOPS
                    write_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/RDS",
                        MetricName="WriteIOPS",
                        Dimensions=[{"Name": "DBInstanceIdentifier", "Value": db_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Average"],
                    )
                    if write_resp.get("Datapoints"):
                        instance.avg_write_iops = sum(d["Average"] for d in write_resp["Datapoints"]) / len(
                            write_resp["Datapoints"]
                        )

                except ClientError:
                    pass

                instances.append(instance)
    except ClientError:
        pass

    return instances


def analyze_instances(
    instances: list[RDSInstanceInfo], account_id: str, account_name: str, region: str
) -> RDSAnalysisResult:
    """RDS 인스턴스 분석"""
    result = RDSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_instances=len(instances),
    )

    for instance in instances:
        # 정지된 인스턴스
        if instance.status == "stopped":
            result.stopped_instances += 1
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.STOPPED,
                    recommendation="정지됨 - 장기 미사용 시 스냅샷 후 삭제 검토",
                )
            )
            continue

        # 미사용: 연결 수 평균 0
        if instance.avg_connections == 0:
            result.unused_instances += 1
            result.unused_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.UNUSED,
                    recommendation=f"연결 없음 - 삭제 검토 (${instance.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 저사용: CPU 5% 미만
        if instance.avg_cpu < LOW_USAGE_CPU_THRESHOLD:
            result.low_usage_instances += 1
            result.low_usage_monthly_cost += instance.estimated_monthly_cost
            result.findings.append(
                InstanceFinding(
                    instance=instance,
                    status=InstanceStatus.LOW_USAGE,
                    recommendation=f"저사용 (CPU {instance.avg_cpu:.1f}%) - 다운사이징 검토",
                )
            )
            continue

        result.normal_instances += 1
        result.findings.append(
            InstanceFinding(
                instance=instance,
                status=InstanceStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[RDSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "RDS 유휴 인스턴스 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Account",
        "Region",
        "전체",
        "미사용",
        "저사용",
        "정지",
        "정상",
        "미사용 비용",
        "저사용 비용",
    ]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_instances)
        ws.cell(row=row, column=4, value=r.unused_instances)
        ws.cell(row=row, column=5, value=r.low_usage_instances)
        ws.cell(row=row, column=6, value=r.stopped_instances)
        ws.cell(row=row, column=7, value=r.normal_instances)
        ws.cell(row=row, column=8, value=f"${r.unused_monthly_cost:,.2f}")
        ws.cell(row=row, column=9, value=f"${r.low_usage_monthly_cost:,.2f}")
        if r.unused_instances > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.low_usage_instances > 0:
            ws.cell(row=row, column=5).fill = yellow_fill
        if r.stopped_instances > 0:
            ws.cell(row=row, column=6).fill = gray_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Instances")
    detail_headers = [
        "Account",
        "Region",
        "Instance ID",
        "Engine",
        "Class",
        "Storage",
        "Multi-AZ",
        "상태",
        "Avg Conn",
        "Avg CPU",
        "월간 비용",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != InstanceStatus.NORMAL:
                detail_row += 1
                inst = f.instance
                ws_detail.cell(row=detail_row, column=1, value=inst.account_name)
                ws_detail.cell(row=detail_row, column=2, value=inst.region)
                ws_detail.cell(row=detail_row, column=3, value=inst.db_instance_id)
                ws_detail.cell(row=detail_row, column=4, value=inst.engine)
                ws_detail.cell(row=detail_row, column=5, value=inst.db_instance_class)
                ws_detail.cell(row=detail_row, column=6, value=f"{inst.allocated_storage} GB")
                ws_detail.cell(row=detail_row, column=7, value="Yes" if inst.multi_az else "No")
                ws_detail.cell(row=detail_row, column=8, value=f.status.value)
                ws_detail.cell(row=detail_row, column=9, value=f"{inst.avg_connections:.1f}")
                ws_detail.cell(row=detail_row, column=10, value=f"{inst.avg_cpu:.1f}%")
                ws_detail.cell(
                    row=detail_row,
                    column=11,
                    value=f"${inst.estimated_monthly_cost:.2f}",
                )
                ws_detail.cell(row=detail_row, column=12, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"RDS_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RDSAnalysisResult | None:
    """단일 계정/리전의 RDS 인스턴스 수집 및 분석 (병렬 실행용)"""
    instances = collect_rds_instances(session, account_id, account_name, region)
    if not instances:
        return None
    return analyze_instances(instances, account_id, account_name, region)


def run(ctx) -> None:
    """RDS 유휴 인스턴스 분석"""
    console.print("[bold]RDS 유휴 인스턴스 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="rds")
    results: list[RDSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_instances for r in results)
    total_low = sum(r.low_usage_instances for r in results)
    total_stopped = sum(r.stopped_instances for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)
    low_cost = sum(r.low_usage_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월) / "
        f"저사용: [yellow]{total_low}개[/yellow] (${low_cost:,.2f}/월) / "
        f"정지: {total_stopped}개"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("rds-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
