"""
plugins/cost/unused_eni.py - ENI 미사용 분석

미사용 ENI (Elastic Network Interface) 탐지

분석 기준:
- Status가 "available"인 ENI (아무것도 연결되지 않음)
- AWS 관리형 ENI는 제외 (NAT Gateway, Lambda, VPC Endpoint 등)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeNetworkInterfaces",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용 (available 상태)
    NORMAL = "normal"  # 정상 사용 (in-use)
    PENDING = "pending"  # 확인 필요
    AWS_MANAGED = "aws_managed"  # AWS 관리형 (삭제 불가)


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class ENIInfo:
    """ENI 정보"""

    id: str
    description: str
    status: str
    vpc_id: str
    subnet_id: str
    availability_zone: str
    private_ip: str
    public_ip: str
    interface_type: str
    requester_id: str
    owner_id: str
    instance_id: str
    attachment_status: str
    security_groups: list[str]
    tags: dict[str, str]
    name: str

    # 메타
    account_id: str
    account_name: str
    region: str

    @property
    def is_attached(self) -> bool:
        """연결 여부"""
        return self.status == "in-use"

    @property
    def is_aws_managed(self) -> bool:
        """AWS 관리형 여부"""
        if self.requester_id and self.requester_id != self.owner_id:
            return True
        aws_managed_types = {
            "nat_gateway",
            "gateway_load_balancer",
            "gateway_load_balancer_endpoint",
            "vpc_endpoint",
            "efa",
            "trunk",
            "load_balancer",
            "lambda",
        }
        return self.interface_type in aws_managed_types


@dataclass
class ENIFinding:
    """ENI 분석 결과"""

    eni: ENIInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class ENIAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[ENIFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    normal_count: int = 0
    aws_managed_count: int = 0
    pending_count: int = 0


# =============================================================================
# 수집
# =============================================================================


def collect_enis(session, account_id: str, account_name: str, region: str) -> list[ENIInfo]:
    """ENI 목록 수집"""
    from botocore.exceptions import ClientError

    enis = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        paginator = ec2.get_paginator("describe_network_interfaces")

        for page in paginator.paginate():
            for data in page.get("NetworkInterfaces", []):
                attachment = data.get("Attachment")

                # 태그 파싱
                tags = {
                    t.get("Key", ""): t.get("Value", "")
                    for t in data.get("TagSet", [])
                    if not t.get("Key", "").startswith("aws:")
                }

                eni = ENIInfo(
                    id=data.get("NetworkInterfaceId", ""),
                    description=data.get("Description", ""),
                    status=data.get("Status", ""),
                    vpc_id=data.get("VpcId", ""),
                    subnet_id=data.get("SubnetId", ""),
                    availability_zone=data.get("AvailabilityZone", ""),
                    private_ip=data.get("PrivateIpAddress", ""),
                    public_ip=data.get("Association", {}).get("PublicIp", "") if data.get("Association") else "",
                    interface_type=data.get("InterfaceType", ""),
                    requester_id=data.get("RequesterId", ""),
                    owner_id=data.get("OwnerId", ""),
                    instance_id=attachment.get("InstanceId", "") if attachment else "",
                    attachment_status=attachment.get("Status", "") if attachment else "",
                    security_groups=[g.get("GroupId", "") for g in data.get("Groups", [])],
                    tags=tags,
                    name=tags.get("Name", ""),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )
                enis.append(eni)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} ENI 수집 오류: {error_code}[/yellow]")

    return enis


# =============================================================================
# 분석
# =============================================================================


def analyze_enis(enis: list[ENIInfo], account_id: str, account_name: str, region: str) -> ENIAnalysisResult:
    """ENI 미사용 분석"""
    result = ENIAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for eni in enis:
        finding = _analyze_single_eni(eni)
        result.findings.append(finding)

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
        elif finding.usage_status == UsageStatus.NORMAL:
            result.normal_count += 1
        elif finding.usage_status == UsageStatus.AWS_MANAGED:
            result.aws_managed_count += 1
        elif finding.usage_status == UsageStatus.PENDING:
            result.pending_count += 1

    result.total_count = len(enis)
    return result


def _analyze_single_eni(eni: ENIInfo) -> ENIFinding:
    """개별 ENI 분석"""

    # 1. AWS 관리형
    if eni.is_aws_managed:
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.AWS_MANAGED,
            severity=Severity.INFO,
            description=f"AWS 관리형 ({eni.interface_type})",
            recommendation="삭제하지 마세요.",
        )

    # 2. 연결됨 = 정상
    if eni.is_attached:
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"사용 중 ({eni.status})",
            recommendation="정상 사용 중",
        )

    # 3. Available = 미사용
    if eni.status == "available":
        desc_lower = eni.description.lower()

        # 특정 패턴은 주의
        if "efs" in desc_lower:
            return ENIFinding(
                eni=eni,
                usage_status=UsageStatus.PENDING,
                severity=Severity.LOW,
                description="EFS 관련 ENI",
                recommendation="EFS 마운트 타겟 확인",
            )

        if "elb" in desc_lower or "load" in desc_lower:
            return ENIFinding(
                eni=eni,
                usage_status=UsageStatus.PENDING,
                severity=Severity.LOW,
                description="Load Balancer 관련 ENI",
                recommendation="LB 상태 확인",
            )

        # 일반 미사용
        return ENIFinding(
            eni=eni,
            usage_status=UsageStatus.UNUSED,
            severity=Severity.HIGH,
            description="미사용 ENI (available, 연결 없음)",
            recommendation="삭제 검토",
        )

    # 4. 기타
    return ENIFinding(
        eni=eni,
        usage_status=UsageStatus.PENDING,
        severity=Severity.INFO,
        description=f"상태: {eni.status}",
        recommendation="상태 안정화 대기",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[ENIAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.PENDING: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
        UsageStatus.AWS_MANAGED: PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "ENI 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "aws_managed": sum(r.aws_managed_count for r in results),
        "pending": sum(r.pending_count for r in results),
    }

    stats = [
        ("항목", "값"),
        ("전체 ENI", totals["total"]),
        ("미사용", totals["unused"]),
        ("정상 사용", totals["normal"]),
        ("AWS 관리형", totals["aws_managed"]),
        ("확인 필요", totals["pending"]),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "ENI ID",
        "Name",
        "Status",
        "Usage",
        "Severity",
        "Description",
        "Recommendation",
        "VPC ID",
        "Subnet ID",
        "Private IP",
        "Type",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # 미사용/확인필요만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.PENDING):
                all_findings.append(f)

    # 심각도순
    severity_order = {
        Severity.HIGH: 0,
        Severity.MEDIUM: 1,
        Severity.LOW: 2,
        Severity.INFO: 3,
    }
    all_findings.sort(key=lambda x: severity_order.get(x.severity, 9))

    for f in all_findings:
        eni = f.eni
        ws2.append(
            [
                eni.account_name,
                eni.region,
                eni.id,
                eni.name,
                eni.status,
                f.usage_status.value,
                f.severity.value,
                f.description,
                f.recommendation,
                eni.vpc_id,
                eni.subnet_id,
                eni.private_ip,
                eni.interface_type,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=6).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 40)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"ENI_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ENIAnalysisResult:
    """단일 계정/리전의 ENI 수집 및 분석 (병렬 실행용)"""
    enis = collect_enis(session, account_id, account_name, region)
    return analyze_enis(enis, account_id, account_name, region)


def run(ctx) -> None:
    """ENI 미사용 분석 실행"""
    console.print("[bold]ENI 미사용 분석 시작...[/bold]")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ec2")

    all_results: list[ENIAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not all_results:
        console.print("[yellow]분석할 ENI 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "aws_managed": sum(r.aws_managed_count for r in all_results),
        "pending": sum(r.pending_count for r in all_results),
    }

    console.print(f"\n[bold]전체 ENI: {totals['total']}개[/bold]")
    if totals["unused"] > 0:
        console.print(f"  [red bold]미사용: {totals['unused']}개[/red bold]")
    if totals["pending"] > 0:
        console.print(f"  [yellow]확인 필요: {totals['pending']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")
    console.print(f"  [dim]AWS 관리형: {totals['aws_managed']}개[/dim]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("eni-audit").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
