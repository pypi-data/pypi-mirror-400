"""Style caching for high-performance Excel generation."""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.styles import Alignment, Border, Font, PatternFill
    from openpyxl.workbook import Workbook


class StyleCache:
    """Singleton cache for openpyxl styles."""

    _instance: StyleCache | None = None
    _initialized: bool = False

    def __new__(cls) -> StyleCache:
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    @classmethod
    def get(cls) -> StyleCache:
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def _ensure_init(self) -> None:
        if self._initialized:
            return

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin")
        self._border = Border(left=thin, right=thin, top=thin, bottom=thin)

        self._align_center = Alignment(horizontal="center", vertical="center")
        self._align_left = Alignment(horizontal="left", vertical="center")
        self._align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self._align_right = Alignment(horizontal="right", vertical="center")

        self._fill_abuse = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        self._fill_danger = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self._fill_warning = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        self._fill_title = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        self._fill_header = PatternFill(start_color="EBF1FA", end_color="EBF1FA", fill_type="solid")

        self._font_bold = Font(name="Consolas", bold=True)
        self._font_italic_gray = Font(name="Consolas", italic=True, color="808080")
        self._font_header = Font(name="Consolas", bold=True, color="FFFFFF")
        self._font_title = Font(name="Consolas", size=16, bold=True, color="1F4E79")
        self._font_section = Font(name="Consolas", size=12, bold=True, color="2F5597")
        self._font_label = Font(name="Consolas", size=11, bold=True)
        self._font_value = Font(name="Consolas", size=11)

        self._initialized = True

    @property
    def thin_border(self) -> Border:
        self._ensure_init()
        return self._border

    @property
    def align_center(self) -> Alignment:
        self._ensure_init()
        return self._align_center

    @property
    def align_left(self) -> Alignment:
        self._ensure_init()
        return self._align_left

    @property
    def align_left_wrap(self) -> Alignment:
        self._ensure_init()
        return self._align_left_wrap

    @property
    def align_right(self) -> Alignment:
        self._ensure_init()
        return self._align_right

    @property
    def fill_abuse(self) -> PatternFill:
        self._ensure_init()
        return self._fill_abuse

    @property
    def danger_fill(self) -> PatternFill:
        self._ensure_init()
        return self._fill_danger

    @property
    def warning_fill(self) -> PatternFill:
        self._ensure_init()
        return self._fill_warning

    @property
    def title_fill(self) -> PatternFill:
        self._ensure_init()
        return self._fill_title

    @property
    def header_fill(self) -> PatternFill:
        self._ensure_init()
        return self._fill_header

    @property
    def font_bold(self) -> Font:
        self._ensure_init()
        return self._font_bold

    @property
    def font_italic_gray(self) -> Font:
        self._ensure_init()
        return self._font_italic_gray

    @property
    def title_font(self) -> Font:
        self._ensure_init()
        return self._font_title

    @property
    def header_font(self) -> Font:
        self._ensure_init()
        return self._font_section

    @property
    def label_font(self) -> Font:
        self._ensure_init()
        return self._font_label

    @property
    def value_font(self) -> Font:
        self._ensure_init()
        return self._font_value

    def get_header_style(self) -> dict[str, Any]:
        """Get header style dict."""
        self._ensure_init()
        from openpyxl.styles import PatternFill

        return {
            "font": self._font_header,
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": self._align_center,
            "border": self._border,
        }

    def register_named_styles(self, wb: Workbook) -> None:
        """Register named styles with workbook."""
        from openpyxl.styles import NamedStyle

        self._ensure_init()

        styles = [
            ("data_cell", self._font_value, self._align_left, None),
            ("number_cell", self._font_value, self._align_right, "#,##0"),
            ("center_cell", self._font_value, self._align_center, None),
        ]

        for name, font, align, fmt in styles:
            try:
                style = NamedStyle(name=name)
                style.font = font
                style.alignment = align
                style.border = self._border
                if fmt:
                    style.number_format = fmt
                wb.add_named_style(style)
            except ValueError:
                pass  # Already exists


_cache: StyleCache | None = None


def get_style_cache() -> StyleCache:
    global _cache
    if _cache is None:
        _cache = StyleCache.get()
    return _cache
