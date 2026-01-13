"""
plugins/fn/versions.py - Lambda Version/Alias 정리

Lambda Version/Alias 관리:
- 오래된 버전 탐지
- 미사용 Alias 탐지
- 버전 정리 권고

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import contextlib
import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListVersionsByFunction",
        "lambda:ListAliases",
    ],
}


class VersionStatus(Enum):
    """버전 상태"""

    CURRENT = "current"  # 현재 사용 중
    ALIAS_TARGET = "alias_target"  # Alias가 가리킴
    OLD = "old"  # 오래된 버전
    UNUSED = "unused"  # 미사용
    LATEST = "latest"  # $LATEST


class AliasStatus(Enum):
    """Alias 상태"""

    ACTIVE = "active"  # 사용 중 (트래픽 있음)
    INACTIVE = "inactive"  # 미사용 (트래픽 없음)


@dataclass
class LambdaVersion:
    """Lambda 버전 정보"""

    function_name: str
    version: str
    description: str
    runtime: str
    code_size_bytes: int
    last_modified: datetime | None
    code_sha256: str


@dataclass
class LambdaAlias:
    """Lambda Alias 정보"""

    function_name: str
    alias_name: str
    function_version: str
    description: str
    routing_config: dict | None = None


@dataclass
class FunctionVersionInfo:
    """함수별 Version/Alias 정보"""

    function_name: str
    function_arn: str
    runtime: str
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 버전 목록
    versions: list[LambdaVersion] = field(default_factory=list)
    aliases: list[LambdaAlias] = field(default_factory=list)

    # 분석 결과
    old_versions: list[str] = field(default_factory=list)
    unused_versions: list[str] = field(default_factory=list)
    inactive_aliases: list[str] = field(default_factory=list)

    @property
    def version_count(self) -> int:
        return len(self.versions)

    @property
    def alias_count(self) -> int:
        return len(self.aliases)

    @property
    def issue_count(self) -> int:
        return len(self.old_versions) + len(self.unused_versions) + len(self.inactive_aliases)


@dataclass
class VersionAuditResult:
    """Version/Alias 감사 결과"""

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    total_versions: int = 0
    total_aliases: int = 0
    old_version_count: int = 0
    unused_version_count: int = 0
    inactive_alias_count: int = 0
    functions: list[FunctionVersionInfo] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_versions(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[FunctionVersionInfo]:
    """Lambda 버전 정보 수집"""
    from botocore.exceptions import ClientError

    result = []

    try:
        lambda_client = get_client(session, "lambda", region_name=region)

        paginator = lambda_client.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page.get("Functions", []):
                function_name = fn.get("FunctionName", "")

                info = FunctionVersionInfo(
                    function_name=function_name,
                    function_arn=fn.get("FunctionArn", ""),
                    runtime=fn.get("Runtime", ""),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )

                # 버전 목록
                try:
                    versions_paginator = lambda_client.get_paginator("list_versions_by_function")
                    for v_page in versions_paginator.paginate(FunctionName=function_name):
                        for v in v_page.get("Versions", []):
                            version = v.get("Version", "")
                            if version == "$LATEST":
                                continue  # $LATEST는 제외

                            last_modified = None
                            lm_str = v.get("LastModified")
                            if lm_str:
                                with contextlib.suppress(ValueError):
                                    last_modified = datetime.fromisoformat(lm_str.replace("Z", "+00:00"))

                            info.versions.append(
                                LambdaVersion(
                                    function_name=function_name,
                                    version=version,
                                    description=v.get("Description", ""),
                                    runtime=v.get("Runtime", ""),
                                    code_size_bytes=v.get("CodeSize", 0),
                                    last_modified=last_modified,
                                    code_sha256=v.get("CodeSha256", ""),
                                )
                            )
                except ClientError:
                    pass

                # Alias 목록
                try:
                    aliases_paginator = lambda_client.get_paginator("list_aliases")
                    for a_page in aliases_paginator.paginate(FunctionName=function_name):
                        for a in a_page.get("Aliases", []):
                            info.aliases.append(
                                LambdaAlias(
                                    function_name=function_name,
                                    alias_name=a.get("Name", ""),
                                    function_version=a.get("FunctionVersion", ""),
                                    description=a.get("Description", ""),
                                    routing_config=a.get("RoutingConfig"),
                                )
                            )
                except ClientError:
                    pass

                result.append(info)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return result


# =============================================================================
# 분석
# =============================================================================


def analyze_versions(
    functions: list[FunctionVersionInfo],
    account_id: str,
    account_name: str,
    region: str,
    old_days: int = 90,
) -> VersionAuditResult:
    """Version/Alias 분석"""
    result = VersionAuditResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    now = datetime.now(timezone.utc)
    cutoff_date = now - timedelta(days=old_days)

    for func in functions:
        result.total_versions += func.version_count
        result.total_aliases += func.alias_count

        # Alias가 가리키는 버전 목록
        aliased_versions: set[str] = set()
        for alias in func.aliases:
            aliased_versions.add(alias.function_version)
            # routing config에 있는 버전도 포함
            if alias.routing_config:
                for v in alias.routing_config.get("AdditionalVersionWeights", {}):
                    aliased_versions.add(v)

        # 버전 분석
        sorted_versions = sorted(
            func.versions,
            key=lambda v: int(v.version) if v.version.isdigit() else 0,
            reverse=True,
        )

        for i, version in enumerate(sorted_versions):
            # Alias가 가리키는 버전은 사용 중
            if version.version in aliased_versions:
                continue

            # 최신 버전 (상위 3개)은 유지
            if i < 3:
                continue

            # 오래된 버전
            if version.last_modified and version.last_modified < cutoff_date:
                func.old_versions.append(version.version)
                result.old_version_count += 1
            else:
                func.unused_versions.append(version.version)
                result.unused_version_count += 1

        # 미사용 Alias (여기서는 트래픽 확인이 어려우므로 패스)
        # CloudWatch 메트릭으로 확인 필요

        result.functions.append(func)

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[VersionAuditResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Lambda Version/Alias 감사"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = ["Account", "Region", "전체 함수", "총 버전", "총 Alias", "오래된 버전", "미사용 버전"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_functions)
        ws.cell(row=row, column=4, value=r.total_versions)
        ws.cell(row=row, column=5, value=r.total_aliases)
        ws.cell(row=row, column=6, value=r.old_version_count)
        ws.cell(row=row, column=7, value=r.unused_version_count)
        if r.old_version_count > 0:
            ws.cell(row=row, column=6).fill = yellow_fill

    # 총계
    total_versions = sum(r.total_versions for r in results)
    total_old = sum(r.old_version_count for r in results)
    total_unused = sum(r.unused_version_count for r in results)

    row += 2
    ws.cell(row=row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_versions).font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_old).font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_unused).font = Font(bold=True)

    row += 2
    ws.cell(row=row, column=1, value="참고").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="- 오래된 버전: 90일 이상 지난 버전 (Alias 미지정)")
    row += 1
    ws.cell(row=row, column=1, value="- 미사용 버전: Alias 미지정 버전 (최신 3개 제외)")
    row += 1
    ws.cell(row=row, column=1, value="- 삭제 시 코드 복구 불가하므로 신중히 결정하세요")

    # Functions with Issues
    ws_issues = wb.create_sheet("Cleanup Candidates")
    issue_headers = [
        "Account",
        "Region",
        "Function",
        "Runtime",
        "총 버전",
        "오래된 버전",
        "미사용 버전",
        "Alias 수",
        "정리 대상 버전",
    ]
    for col, h in enumerate(issue_headers, 1):
        ws_issues.cell(row=1, column=col, value=h).fill = header_fill
        ws_issues.cell(row=1, column=col).font = header_font

    issue_row = 1
    for r in results:
        for func in r.functions:
            if func.issue_count > 0:
                issue_row += 1
                cleanup_versions = func.old_versions + func.unused_versions
                ws_issues.cell(row=issue_row, column=1, value=func.account_name)
                ws_issues.cell(row=issue_row, column=2, value=func.region)
                ws_issues.cell(row=issue_row, column=3, value=func.function_name)
                ws_issues.cell(row=issue_row, column=4, value=func.runtime)
                ws_issues.cell(row=issue_row, column=5, value=func.version_count)
                ws_issues.cell(row=issue_row, column=6, value=len(func.old_versions))
                ws_issues.cell(row=issue_row, column=7, value=len(func.unused_versions))
                ws_issues.cell(row=issue_row, column=8, value=func.alias_count)
                ws_issues.cell(
                    row=issue_row,
                    column=9,
                    value=", ".join(cleanup_versions[:10]) + ("..." if len(cleanup_versions) > 10 else ""),
                )

    # All Versions
    ws_all = wb.create_sheet("All Versions")
    all_headers = [
        "Account",
        "Region",
        "Function",
        "Version",
        "Description",
        "Runtime",
        "Code Size",
        "Last Modified",
        "상태",
    ]
    for col, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col, value=h).fill = header_fill
        ws_all.cell(row=1, column=col).font = header_font

    all_row = 1
    for r in results:
        for func in r.functions:
            # Alias가 가리키는 버전
            aliased_versions: set[str] = set()
            for alias in func.aliases:
                aliased_versions.add(alias.function_version)

            sorted_versions = sorted(
                func.versions,
                key=lambda v: int(v.version) if v.version.isdigit() else 0,
                reverse=True,
            )

            for i, v in enumerate(sorted_versions):
                all_row += 1
                ws_all.cell(row=all_row, column=1, value=func.account_name)
                ws_all.cell(row=all_row, column=2, value=func.region)
                ws_all.cell(row=all_row, column=3, value=func.function_name)
                ws_all.cell(row=all_row, column=4, value=v.version)
                ws_all.cell(
                    row=all_row,
                    column=5,
                    value=v.description[:50] if v.description else "-",
                )
                ws_all.cell(row=all_row, column=6, value=v.runtime)
                ws_all.cell(
                    row=all_row,
                    column=7,
                    value=f"{v.code_size_bytes / 1024 / 1024:.2f} MB",
                )
                ws_all.cell(
                    row=all_row,
                    column=8,
                    value=v.last_modified.strftime("%Y-%m-%d") if v.last_modified else "-",
                )

                # 상태
                if v.version in aliased_versions:
                    status = "alias_target"
                elif v.version in func.old_versions:
                    status = "old"
                    ws_all.cell(row=all_row, column=9).fill = yellow_fill
                elif v.version in func.unused_versions:
                    status = "unused"
                elif i < 3:
                    status = "current"
                else:
                    status = "unused"
                ws_all.cell(row=all_row, column=9, value=status)

    # Aliases
    ws_alias = wb.create_sheet("Aliases")
    alias_headers = [
        "Account",
        "Region",
        "Function",
        "Alias",
        "Target Version",
        "Description",
        "Routing",
    ]
    for col, h in enumerate(alias_headers, 1):
        ws_alias.cell(row=1, column=col, value=h).fill = header_fill
        ws_alias.cell(row=1, column=col).font = header_font

    alias_row = 1
    for r in results:
        for func in r.functions:
            for alias in func.aliases:
                alias_row += 1
                ws_alias.cell(row=alias_row, column=1, value=func.account_name)
                ws_alias.cell(row=alias_row, column=2, value=func.region)
                ws_alias.cell(row=alias_row, column=3, value=func.function_name)
                ws_alias.cell(row=alias_row, column=4, value=alias.alias_name)
                ws_alias.cell(row=alias_row, column=5, value=alias.function_version)
                ws_alias.cell(row=alias_row, column=6, value=alias.description or "-")
                routing = "-"
                if alias.routing_config:
                    weights = alias.routing_config.get("AdditionalVersionWeights", {})
                    if weights:
                        routing = ", ".join(f"v{k}: {v * 100:.0f}%" for k, v in weights.items())
                ws_alias.cell(row=alias_row, column=7, value=routing)

    # 열 너비
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Lambda_Version_Audit_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> VersionAuditResult | None:
    """단일 계정/리전의 Lambda 버전 수집 및 분석 (병렬 실행용)"""
    functions = collect_versions(session, account_id, account_name, region)
    if not functions:
        return None
    return analyze_versions(functions, account_id, account_name, region)


def run(ctx) -> None:
    """Version/Alias 감사 실행"""
    console.print("[bold]Lambda Version/Alias 감사 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")
    results: list[VersionAuditResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_versions = sum(r.total_versions for r in results)
    total_aliases = sum(r.total_aliases for r in results)
    total_old = sum(r.old_version_count for r in results)
    total_unused = sum(r.unused_version_count for r in results)
    total_cleanup = total_old + total_unused

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"총 버전: {total_versions}개")
    console.print(f"총 Alias: {total_aliases}개")

    if total_cleanup > 0:
        console.print(f"[yellow]정리 대상: {total_cleanup}개[/yellow]")
        console.print(f"  - 오래된 버전 (90일+): {total_old}개")
        console.print(f"  - 미사용 버전: {total_unused}개")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("lambda-version").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
