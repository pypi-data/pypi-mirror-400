"""Tests for conversion to Excel module."""

from pathlib import Path

import openpyxl

import raesl.excel


def test_excel(datadir: Path, update, tmpdir):
    """Integral test for Excel module."""
    spec = datadir / "specs" / "pump_example.esl"

    ref = datadir / "excel" / "pump.xlsx"
    output = ref if update else tmpdir / "pump.xlsx"
    raesl.excel.convert(spec, output=output)

    if update:
        return

    wb = openpyxl.load_workbook(output)
    wb_ref = openpyxl.load_workbook(ref)

    # This is in no means comprehensive, but serves as a global output test.
    assert wb.sheetnames == wb_ref.sheetnames
    for ws, ws_ref in zip(wb.worksheets, wb_ref.worksheets):
        for col, col_ref in zip(ws.iter_cols(), ws_ref.iter_cols()):
            for cell, cell_ref in zip(col, col_ref):
                assert cell.value == cell_ref.value
                assert cell.style == cell_ref.style


def test_component_overview(tmpdir):
    """Integral test for component overview generation."""

    from raesl import datasets
    from raesl.excel import component_overview

    graph = datasets.get("rally-car")

    for comp in graph.get_nodes_by_kind("component"):
        output = tmpdir / f"overview_{comp.name}.xlsx"

        component_overview(graph, comp.name, output=output)
