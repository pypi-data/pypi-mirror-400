"""
plugins/ec2/ami_audit.py - AMI 미사용 분석

사용되지 않는 AMI 탐지

분석 기준:
- 어떤 EC2 인스턴스에서도 사용하지 않는 AMI
- 14일 이상 경과 (최근 생성은 제외)

비용:
- AMI 자체는 무료, 백업 스냅샷에서 $0.05/GB-월 발생

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import contextlib
import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum

from dateutil.parser import parse
from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_snapshot_monthly_cost

console = Console()


# =============================================================================
# 상수
# =============================================================================

# 최근 생성 기준 (일) - 이내는 제외
RECENT_DAYS = 14

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeImages",
        "ec2:DescribeInstances",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용
    NORMAL = "normal"  # 사용 중


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class AMIInfo:
    """AMI 정보"""

    id: str
    name: str
    description: str
    state: str
    architecture: str
    platform: str
    root_device_type: str
    creation_date: datetime | None
    owner_id: str
    public: bool
    tags: dict[str, str]

    # 스냅샷 정보
    snapshot_ids: list[str] = field(default_factory=list)
    total_size_gb: int = 0

    # 사용 정보
    used_by_instances: list[str] = field(default_factory=list)

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def age_days(self) -> int:
        """AMI 나이 (일)"""
        if not self.creation_date:
            return 0
        now = datetime.now(timezone.utc)
        delta = now - self.creation_date.replace(tzinfo=timezone.utc)
        return delta.days

    @property
    def is_used(self) -> bool:
        """사용 여부"""
        return len(self.used_by_instances) > 0

    @property
    def is_recent(self) -> bool:
        """최근 생성 여부"""
        return self.age_days < RECENT_DAYS


@dataclass
class AMIFinding:
    """AMI 분석 결과"""

    ami: AMIInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class AMIAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[AMIFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    normal_count: int = 0

    # 용량/비용
    total_size_gb: int = 0
    unused_size_gb: int = 0
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_amis(session, account_id: str, account_name: str, region: str) -> list[AMIInfo]:
    """AMI 목록 수집 (자체 소유만)"""
    from botocore.exceptions import ClientError

    amis = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)

        # 자체 소유 AMI 조회
        response = ec2.describe_images(Owners=["self"])

        for data in response.get("Images", []):
            if data.get("State") != "available":
                continue

            # 태그 파싱
            tags = {
                t.get("Key", ""): t.get("Value", "")
                for t in data.get("Tags", [])
                if not t.get("Key", "").startswith("aws:")
            }

            # 스냅샷 ID 및 크기 추출
            snapshot_ids = []
            total_size_gb = 0
            for bdm in data.get("BlockDeviceMappings", []):
                ebs = bdm.get("Ebs", {})
                snap_id = ebs.get("SnapshotId", "")
                if snap_id:
                    snapshot_ids.append(snap_id)
                    total_size_gb += ebs.get("VolumeSize", 0)

            # 생성일 파싱
            creation_date = None
            if data.get("CreationDate"):
                with contextlib.suppress(Exception):
                    creation_date = parse(data["CreationDate"])

            monthly_cost = get_snapshot_monthly_cost(region, total_size_gb)

            ami = AMIInfo(
                id=data.get("ImageId", ""),
                name=data.get("Name", ""),
                description=data.get("Description", ""),
                state=data.get("State", ""),
                architecture=data.get("Architecture", ""),
                platform=data.get("PlatformDetails", ""),
                root_device_type=data.get("RootDeviceType", ""),
                creation_date=creation_date,
                owner_id=data.get("OwnerId", ""),
                public=data.get("Public", False),
                tags=tags,
                snapshot_ids=snapshot_ids,
                total_size_gb=total_size_gb,
                account_id=account_id,
                account_name=account_name,
                region=region,
                monthly_cost=monthly_cost,
            )
            amis.append(ami)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} AMI 수집 오류: {error_code}[/yellow]")

    return amis


def get_used_ami_ids(session, region: str) -> set[str]:
    """EC2 인스턴스에서 사용 중인 AMI ID 목록"""
    from botocore.exceptions import ClientError

    used_ids = set()

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        paginator = ec2.get_paginator("describe_instances")

        for page in paginator.paginate():
            for reservation in page.get("Reservations", []):
                for instance in reservation.get("Instances", []):
                    ami_id = instance.get("ImageId", "")
                    if ami_id:
                        used_ids.add(ami_id)

    except ClientError:
        pass

    return used_ids


# =============================================================================
# 분석
# =============================================================================


def analyze_amis(
    amis: list[AMIInfo],
    used_ami_ids: set[str],
    account_id: str,
    account_name: str,
    region: str,
) -> AMIAnalysisResult:
    """AMI 미사용 분석"""
    result = AMIAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for ami in amis:
        # 사용 여부 설정
        if ami.id in used_ami_ids:
            ami.used_by_instances = ["in-use"]

        finding = _analyze_single_ami(ami)
        result.findings.append(finding)
        result.total_size_gb += ami.total_size_gb

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_size_gb += ami.total_size_gb
            result.unused_monthly_cost += ami.monthly_cost
        else:
            result.normal_count += 1

    result.total_count = len(amis)
    return result


def _analyze_single_ami(ami: AMIInfo) -> AMIFinding:
    """개별 AMI 분석"""

    # 사용 중
    if ami.is_used:
        return AMIFinding(
            ami=ami,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description="인스턴스에서 사용 중",
            recommendation="정상",
        )

    # 최근 생성 (14일 이내)
    if ami.is_recent:
        return AMIFinding(
            ami=ami,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"최근 생성 ({ami.age_days}일)",
            recommendation="모니터링",
        )

    # 미사용 + 오래됨
    if ami.total_size_gb >= 100:
        severity = Severity.HIGH
    elif ami.total_size_gb >= 50:
        severity = Severity.MEDIUM
    else:
        severity = Severity.LOW

    return AMIFinding(
        ami=ami,
        usage_status=UsageStatus.UNUSED,
        severity=severity,
        description=f"미사용 AMI ({ami.age_days}일, {ami.total_size_gb}GB, ${ami.monthly_cost:.2f}/월)",
        recommendation="삭제 검토 (스냅샷 비용 절감)",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[AMIAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "AMI 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "unused_size": sum(r.unused_size_gb for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    stats = [
        ("항목", "값"),
        ("전체 AMI", totals["total"]),
        ("미사용", totals["unused"]),
        ("사용 중", totals["normal"]),
        ("전체 스냅샷 용량 (GB)", totals["total_size"]),
        ("미사용 스냅샷 용량 (GB)", totals["unused_size"]),
        ("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}"),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "AMI ID",
        "Name",
        "Usage",
        "Severity",
        "Size (GB)",
        "Age (days)",
        "Monthly Cost ($)",
        "Architecture",
        "Created",
        "Description",
        "Recommendation",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 미사용만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status == UsageStatus.UNUSED:
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.ami.monthly_cost, reverse=True)

    for f in all_findings:
        ami = f.ami
        ws2.append(
            [
                ami.account_name,
                ami.region,
                ami.id,
                ami.name,
                f.usage_status.value,
                f.severity.value,
                ami.total_size_gb,
                ami.age_days,
                round(ami.monthly_cost, 2),
                ami.architecture,
                ami.creation_date.strftime("%Y-%m-%d") if ami.creation_date else "",
                f.description,
                f.recommendation,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=5).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"AMI_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> AMIAnalysisResult | None:
    """단일 계정/리전의 AMI 수집 및 분석 (병렬 실행용)"""
    # 수집
    amis = collect_amis(session, account_id, account_name, region)

    if not amis:
        return None

    # 사용 중인 AMI 조회
    used_ami_ids = get_used_ami_ids(session, region)

    # 분석
    return analyze_amis(amis, used_ami_ids, account_id, account_name, region)


def run(ctx) -> None:
    """AMI 미사용 분석 실행"""
    console.print("[bold]AMI 미사용 분석 시작...[/bold]")
    console.print(f"  [dim]기준: {RECENT_DAYS}일 이내 생성은 제외[/dim]")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ec2")

    # None 결과 필터링
    all_results: list[AMIAnalysisResult] = [r for r in result.get_data() if r is not None]

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not all_results:
        console.print("[yellow]분석할 AMI 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "unused_size": sum(r.unused_size_gb for r in all_results),
        "unused_cost": sum(r.unused_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 AMI: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["unused"] > 0:
        console.print(
            f"  [red bold]미사용: {totals['unused']}개 ({totals['unused_size']}GB, ${totals['unused_cost']:.2f}/월)[/red bold]"
        )
    console.print(f"  [green]사용 중: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("ami-audit").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
