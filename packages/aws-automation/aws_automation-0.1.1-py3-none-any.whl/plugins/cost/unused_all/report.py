"""
plugins/cost/unused_all/report.py - Excel 보고서 생성

미사용 리소스 종합 분석 Excel 보고서 생성
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .types import UnusedAllResult


def generate_report(result: UnusedAllResult, output_dir: str) -> str:
    """종합 Excel 보고서 생성"""
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # ===== Summary =====
    ws = wb.create_sheet("Summary")
    ws["A1"] = "미사용 리소스 종합 보고서"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    row = 4
    for col, h in enumerate(["리소스", "전체", "미사용", "월간 낭비"], 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    # 카테고리별 리소스 목록
    resources = [
        # Compute (EC2)
        ("AMI", "ami_total", "ami_unused", "ami_monthly_waste"),
        ("EBS", "ebs_total", "ebs_unused", "ebs_monthly_waste"),
        ("EBS Snapshot", "snap_total", "snap_unused", "snap_monthly_waste"),
        ("EIP", "eip_total", "eip_unused", "eip_monthly_waste"),
        ("ENI", "eni_total", "eni_unused", None),
        # Networking (VPC)
        ("NAT Gateway", "nat_total", "nat_unused", "nat_monthly_waste"),
        ("VPC Endpoint", "endpoint_total", "endpoint_unused", "endpoint_monthly_waste"),
        # Load Balancing
        ("ELB", "elb_total", "elb_unused", "elb_monthly_waste"),
        ("Target Group", "tg_total", "tg_issue", None),
        # Database
        ("DynamoDB", "dynamodb_total", "dynamodb_unused", "dynamodb_monthly_waste"),
        (
            "ElastiCache",
            "elasticache_total",
            "elasticache_unused",
            "elasticache_monthly_waste",
        ),
        (
            "RDS Instance",
            "rds_instance_total",
            "rds_instance_unused",
            "rds_instance_monthly_waste",
        ),
        ("RDS Snapshot", "rds_snap_total", "rds_snap_old", "rds_snap_monthly_waste"),
        # Storage
        ("ECR", "ecr_total", "ecr_issue", "ecr_monthly_waste"),
        ("EFS", "efs_total", "efs_unused", "efs_monthly_waste"),
        ("S3", "s3_total", "s3_empty", None),
        # Serverless
        ("API Gateway", "apigateway_total", "apigateway_unused", None),
        ("EventBridge", "eventbridge_total", "eventbridge_unused", None),
        ("Lambda", "lambda_total", "lambda_unused", "lambda_monthly_waste"),
        # Messaging
        ("SNS", "sns_total", "sns_unused", None),
        ("SQS", "sqs_total", "sqs_unused", None),
        # Security
        ("ACM", "acm_total", "acm_unused", None),
        ("KMS", "kms_total", "kms_unused", "kms_monthly_waste"),
        ("Secrets Manager", "secret_total", "secret_unused", "secret_monthly_waste"),
        # Monitoring
        ("CloudWatch Alarm", "cw_alarm_total", "cw_alarm_orphan", None),
        ("Log Group", "loggroup_total", "loggroup_issue", "loggroup_monthly_waste"),
        # DNS (Global)
        ("Route53", "route53_total", "route53_empty", "route53_monthly_waste"),
    ]

    for name, total_attr, unused_attr, waste_attr in resources:
        row += 1
        total = sum(getattr(s, total_attr, 0) for s in result.summaries)
        unused = sum(getattr(s, unused_attr, 0) for s in result.summaries)
        waste = sum(getattr(s, waste_attr, 0) for s in result.summaries) if waste_attr else 0
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=unused)
        ws.cell(row=row, column=4, value=f"${waste:,.2f}" if waste > 0 else "-")
        if unused > 0:
            ws.cell(row=row, column=3).fill = red_fill

    # 총 절감
    total_waste = sum(
        s.nat_monthly_waste
        + s.ebs_monthly_waste
        + s.eip_monthly_waste
        + s.elb_monthly_waste
        + s.snap_monthly_waste
        + s.ami_monthly_waste
        + s.rds_snap_monthly_waste
        + s.loggroup_monthly_waste
        + s.endpoint_monthly_waste
        + s.secret_monthly_waste
        + s.kms_monthly_waste
        + s.ecr_monthly_waste
        + s.route53_monthly_waste
        + s.lambda_monthly_waste
        + s.elasticache_monthly_waste
        + s.rds_instance_monthly_waste
        + s.efs_monthly_waste
        + s.dynamodb_monthly_waste
        for s in result.summaries
    )
    row += 2
    ws.cell(row=row, column=1, value="총 월간 절감 가능").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"${total_waste:,.2f}").font = Font(bold=True, color="FF0000")

    # ===== 상세 시트들 (카테고리별 정렬) =====
    # Compute (EC2)
    _create_ami_sheet(wb, result.ami_results, header_fill, header_font)
    _create_ebs_sheet(wb, result.ebs_results, header_fill, header_font)
    _create_snap_sheet(wb, result.snap_results, header_fill, header_font)
    _create_eip_sheet(wb, result.eip_results, header_fill, header_font)
    _create_eni_sheet(wb, result.eni_results, header_fill, header_font)
    # Networking (VPC)
    _create_nat_sheet(wb, result.nat_findings, header_fill, header_font)
    _create_endpoint_sheet(wb, result.endpoint_results, header_fill, header_font)
    # Load Balancing
    _create_elb_sheet(wb, result.elb_results, header_fill, header_font)
    _create_tg_sheet(wb, result.tg_results, header_fill, header_font)
    # Database
    _create_dynamodb_sheet(wb, result.dynamodb_results, header_fill, header_font)
    _create_elasticache_sheet(wb, result.elasticache_results, header_fill, header_font)
    _create_rds_instance_sheet(wb, result.rds_instance_results, header_fill, header_font)
    _create_rds_snap_sheet(wb, result.rds_snap_results, header_fill, header_font)
    # Storage
    _create_ecr_sheet(wb, result.ecr_results, header_fill, header_font)
    _create_efs_sheet(wb, result.efs_results, header_fill, header_font)
    _create_s3_sheet(wb, result.s3_results, header_fill, header_font)
    # Serverless
    _create_apigateway_sheet(wb, result.apigateway_results, header_fill, header_font)
    _create_eventbridge_sheet(wb, result.eventbridge_results, header_fill, header_font)
    _create_lambda_sheet(wb, result.lambda_results, header_fill, header_font)
    # Messaging
    _create_sns_sheet(wb, result.sns_results, header_fill, header_font)
    _create_sqs_sheet(wb, result.sqs_results, header_fill, header_font)
    # Security
    _create_acm_sheet(wb, result.acm_results, header_fill, header_font)
    _create_kms_sheet(wb, result.kms_results, header_fill, header_font)
    _create_secret_sheet(wb, result.secret_results, header_fill, header_font)
    # Monitoring
    _create_cw_alarm_sheet(wb, result.cw_alarm_results, header_fill, header_font)
    _create_loggroup_sheet(wb, result.loggroup_results, header_fill, header_font)
    # DNS (Global)
    _create_route53_sheet(wb, result.route53_results, header_fill, header_font)

    # 열 너비 조정
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            col_tuple = tuple(col_cells)
            max_len = max(len(str(c.value) if c.value else "") for c in col_tuple)
            col_idx = col_tuple[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Unused_Resources_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 개별 시트 생성 함수들
# =============================================================================


def _create_nat_sheet(wb, findings, header_fill, header_font):
    ws = wb.create_sheet("NAT Gateway")
    ws.append(
        [
            "Account",
            "Region",
            "NAT ID",
            "Name",
            "Usage",
            "Monthly Waste",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for nat_result in findings:
        for f in nat_result.findings:
            if f.usage_status.value in ("unused", "low_usage"):
                ws.append(
                    [
                        f.nat.account_name,
                        f.nat.region,
                        f.nat.nat_gateway_id,
                        f.nat.name,
                        f.usage_status.value,
                        f"${f.monthly_waste:,.2f}",
                        f.recommendation,
                    ]
                )


def _create_eni_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ENI")
    ws.append(["Account", "Region", "ENI ID", "Name", "Usage", "Type", "Recommendation"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "pending"):
                ws.append(
                    [
                        f.eni.account_name,
                        f.eni.region,
                        f.eni.id,
                        f.eni.name,
                        f.usage_status.value,
                        f.eni.interface_type,
                        f.recommendation,
                    ]
                )


def _create_ebs_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EBS")
    ws.append(
        [
            "Account",
            "Region",
            "Volume ID",
            "Name",
            "Type",
            "Size (GB)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "pending"):
                ws.append(
                    [
                        f.volume.account_name,
                        f.volume.region,
                        f.volume.id,
                        f.volume.name,
                        f.volume.volume_type,
                        f.volume.size_gb,
                        round(f.volume.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_eip_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EIP")
    ws.append(
        [
            "Account",
            "Region",
            "Allocation ID",
            "Public IP",
            "Name",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "unused":
                ws.append(
                    [
                        f.eip.account_name,
                        f.eip.region,
                        f.eip.allocation_id,
                        f.eip.public_ip,
                        f.eip.name,
                        round(f.eip.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_elb_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ELB")
    ws.append(
        [
            "Account",
            "Region",
            "Name",
            "Type",
            "Usage",
            "Targets",
            "Healthy",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("unused", "unhealthy"):
                ws.append(
                    [
                        f.lb.account_name,
                        f.lb.region,
                        f.lb.name,
                        f.lb.lb_type.upper(),
                        f.usage_status.value,
                        f.lb.total_targets,
                        f.lb.healthy_targets,
                        round(f.lb.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_snap_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("EBS Snapshot")
    ws.append(
        [
            "Account",
            "Region",
            "Snapshot ID",
            "Name",
            "Usage",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value in ("orphan", "old"):
                ws.append(
                    [
                        f.snapshot.account_name,
                        f.snapshot.region,
                        f.snapshot.id,
                        f.snapshot.name,
                        f.usage_status.value,
                        f.snapshot.volume_size_gb,
                        f.snapshot.age_days,
                        round(f.snapshot.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_ami_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("AMI")
    ws.append(
        [
            "Account",
            "Region",
            "AMI ID",
            "Name",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "unused":
                ws.append(
                    [
                        f.ami.account_name,
                        f.ami.region,
                        f.ami.id,
                        f.ami.name,
                        f.ami.total_size_gb,
                        f.ami.age_days,
                        round(f.ami.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_rds_snap_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("RDS Snapshot")
    ws.append(
        [
            "Account",
            "Region",
            "Snapshot ID",
            "DB Identifier",
            "Type",
            "Engine",
            "Size (GB)",
            "Age (days)",
            "Monthly Cost",
            "Recommendation",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.usage_status.value == "old":
                ws.append(
                    [
                        f.snapshot.account_name,
                        f.snapshot.region,
                        f.snapshot.id,
                        f.snapshot.db_identifier,
                        f.snapshot.snapshot_type.value.upper(),
                        f.snapshot.engine,
                        f.snapshot.allocated_storage_gb,
                        f.snapshot.age_days,
                        round(f.snapshot.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_loggroup_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Log Group")
    ws.append(
        [
            "Account",
            "Region",
            "Log Group",
            "상태",
            "저장 (GB)",
            "보존 기간",
            "마지막 Ingestion",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                lg = f.log_group
                ws.append(
                    [
                        lg.account_name,
                        lg.region,
                        lg.name,
                        f.status.value,
                        round(lg.stored_gb, 4),
                        f"{lg.retention_days}일" if lg.retention_days else "무기한",
                        lg.last_ingestion_time.strftime("%Y-%m-%d") if lg.last_ingestion_time else "-",
                        round(lg.monthly_cost, 4),
                        f.recommendation,
                    ]
                )


def _create_tg_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Target Group")
    ws.append(
        [
            "Account",
            "Region",
            "Name",
            "상태",
            "Type",
            "Protocol",
            "Port",
            "LB 연결",
            "Total Targets",
            "Healthy",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                tg = f.tg
                ws.append(
                    [
                        tg.account_name,
                        tg.region,
                        tg.name,
                        f.status.value,
                        tg.target_type,
                        tg.protocol or "-",
                        tg.port or "-",
                        len(tg.load_balancer_arns),
                        tg.total_targets,
                        tg.healthy_targets,
                        f.recommendation,
                    ]
                )


def _create_endpoint_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("VPC Endpoint")
    ws.append(
        [
            "Account",
            "Region",
            "Endpoint ID",
            "Type",
            "Service",
            "VPC",
            "State",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                ep = f.endpoint
                ws.append(
                    [
                        ep.account_name,
                        ep.region,
                        ep.endpoint_id,
                        ep.endpoint_type,
                        ep.service_name.split(".")[-1],
                        ep.vpc_id,
                        ep.state,
                        round(ep.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_secret_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Secrets Manager")
    ws.append(["Account", "Region", "Name", "상태", "마지막 액세스", "월간 비용", "권장 조치"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                sec = f.secret
                last_access = sec.last_accessed_date.strftime("%Y-%m-%d") if sec.last_accessed_date else "없음"
                ws.append(
                    [
                        sec.account_name,
                        sec.region,
                        sec.name,
                        f.status.value,
                        last_access,
                        round(sec.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_kms_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("KMS")
    ws.append(
        [
            "Account",
            "Region",
            "Key ID",
            "Description",
            "상태",
            "Manager",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                key = f.key
                ws.append(
                    [
                        key.account_name,
                        key.region,
                        key.key_id,
                        key.description[:50] if key.description else "-",
                        f.status.value,
                        key.key_manager,
                        round(key.monthly_cost, 2),
                        f.recommendation,
                    ]
                )


def _create_ecr_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("ECR")
    ws.append(
        [
            "Account",
            "Region",
            "Repository",
            "상태",
            "이미지 수",
            "오래된 이미지",
            "총 크기",
            "낭비 비용",
            "Lifecycle",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                repo = f.repo
                ws.append(
                    [
                        repo.account_name,
                        repo.region,
                        repo.name,
                        f.status.value,
                        repo.image_count,
                        repo.old_image_count,
                        f"{repo.total_size_gb:.2f} GB",
                        f"${repo.old_images_monthly_cost:.2f}",
                        "있음" if repo.has_lifecycle_policy else "없음",
                        f.recommendation,
                    ]
                )


def _create_route53_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Route53")
    ws.append(["Account", "Zone ID", "Domain", "Type", "상태", "레코드 수", "월간 비용", "권장 조치"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                zone = f.zone
                ws.append(
                    [
                        zone.account_name,
                        zone.zone_id,
                        zone.name,
                        "Private" if zone.is_private else "Public",
                        f.status.value,
                        zone.record_count,
                        f"${zone.monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_s3_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("S3")
    ws.append(
        [
            "Account",
            "Bucket",
            "Region",
            "상태",
            "객체 수",
            "크기",
            "버전관리",
            "Lifecycle",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                bucket = f.bucket
                ws.append(
                    [
                        bucket.account_name,
                        bucket.name,
                        bucket.region,
                        f.status.value,
                        bucket.object_count,
                        f"{bucket.total_size_mb:.2f} MB",
                        "Enabled" if bucket.versioning_enabled else "Disabled",
                        "있음" if bucket.has_lifecycle else "없음",
                        f.recommendation,
                    ]
                )


def _create_lambda_sheet(wb, results, header_fill, header_font):
    ws = wb.create_sheet("Lambda")
    ws.append(
        [
            "Account",
            "Region",
            "Function Name",
            "Runtime",
            "Memory (MB)",
            "상태",
            "월간 낭비",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                fn = f.function
                ws.append(
                    [
                        fn.account_name,
                        fn.region,
                        fn.function_name,
                        fn.runtime,
                        fn.memory_mb,
                        f.status.value,
                        f"${f.monthly_waste:.2f}" if f.monthly_waste > 0 else "-",
                        f.recommendation,
                    ]
                )


def _create_elasticache_sheet(wb, results, header_fill, header_font):
    """ElastiCache 상세 시트 생성"""
    ws = wb.create_sheet("ElastiCache")
    ws.append(
        [
            "Account",
            "Region",
            "Cluster ID",
            "Engine",
            "Node Type",
            "Nodes",
            "상태",
            "Avg Conn",
            "Avg CPU",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                c = f.cluster
                ws.append(
                    [
                        c.account_name,
                        c.region,
                        c.cluster_id,
                        c.engine,
                        c.node_type,
                        c.num_nodes,
                        f.status.value,
                        f"{c.avg_connections:.1f}",
                        f"{c.avg_cpu:.1f}%",
                        f"${c.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_rds_instance_sheet(wb, results, header_fill, header_font):
    """RDS Instance 상세 시트 생성"""
    ws = wb.create_sheet("RDS Instance")
    ws.append(
        [
            "Account",
            "Region",
            "Instance ID",
            "Engine",
            "Class",
            "Storage",
            "Multi-AZ",
            "상태",
            "Avg Conn",
            "Avg CPU",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                inst = f.instance
                ws.append(
                    [
                        inst.account_name,
                        inst.region,
                        inst.db_instance_id,
                        inst.engine,
                        inst.db_instance_class,
                        f"{inst.allocated_storage} GB",
                        "Yes" if inst.multi_az else "No",
                        f.status.value,
                        f"{inst.avg_connections:.1f}",
                        f"{inst.avg_cpu:.1f}%",
                        f"${inst.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_efs_sheet(wb, results, header_fill, header_font):
    """EFS 상세 시트 생성"""
    ws = wb.create_sheet("EFS")
    ws.append(
        [
            "Account",
            "Region",
            "ID",
            "Name",
            "Size",
            "Mount Targets",
            "Mode",
            "상태",
            "Avg Conn",
            "Total I/O",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                fs = f.efs
                ws.append(
                    [
                        fs.account_name,
                        fs.region,
                        fs.file_system_id,
                        fs.name or "-",
                        f"{fs.size_gb:.2f} GB",
                        fs.mount_target_count,
                        fs.throughput_mode,
                        f.status.value,
                        f"{fs.avg_client_connections:.1f}",
                        f"{fs.total_io_bytes / (1024**2):.1f} MB",
                        f"${fs.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )


def _create_sqs_sheet(wb, results, header_fill, header_font):
    """SQS 상세 시트 생성"""
    ws = wb.create_sheet("SQS")
    ws.append(
        [
            "Account",
            "Region",
            "Queue Name",
            "Type",
            "Messages",
            "상태",
            "Sent",
            "Received",
            "Deleted",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                q = f.queue
                queue_type = "FIFO" if q.is_fifo else "Standard"
                if q.is_dlq:
                    queue_type += " (DLQ)"
                ws.append(
                    [
                        q.account_name,
                        q.region,
                        q.queue_name,
                        queue_type,
                        q.approximate_messages,
                        f.status.value,
                        int(q.messages_sent),
                        int(q.messages_received),
                        int(q.messages_deleted),
                        f.recommendation,
                    ]
                )


def _create_sns_sheet(wb, results, header_fill, header_font):
    """SNS 상세 시트 생성"""
    ws = wb.create_sheet("SNS")
    ws.append(
        [
            "Account",
            "Region",
            "Topic Name",
            "Subscribers",
            "상태",
            "Published",
            "Delivered",
            "Failed",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                t = f.topic
                ws.append(
                    [
                        t.account_name,
                        t.region,
                        t.topic_name,
                        t.subscription_count,
                        f.status.value,
                        int(t.messages_published),
                        int(t.notifications_delivered),
                        int(t.notifications_failed),
                        f.recommendation,
                    ]
                )


def _create_acm_sheet(wb, results, header_fill, header_font):
    """ACM 상세 시트 생성"""
    ws = wb.create_sheet("ACM")
    ws.append(
        [
            "Account",
            "Region",
            "Domain",
            "Type",
            "Status",
            "Expiry",
            "Days Left",
            "In Use",
            "분석상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                c = f.cert
                ws.append(
                    [
                        c.account_name,
                        c.region,
                        c.domain_name,
                        c.cert_type,
                        c.status,
                        c.not_after.strftime("%Y-%m-%d") if c.not_after else "-",
                        c.days_until_expiry if c.days_until_expiry else "-",
                        len(c.in_use_by),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_apigateway_sheet(wb, results, header_fill, header_font):
    """API Gateway 상세 시트 생성"""
    ws = wb.create_sheet("API Gateway")
    ws.append(
        [
            "Account",
            "Region",
            "API Name",
            "Type",
            "Endpoint",
            "Stages",
            "Requests",
            "상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                a = f.api
                ws.append(
                    [
                        a.account_name,
                        a.region,
                        a.api_name,
                        a.api_type,
                        a.endpoint_type,
                        a.stage_count,
                        int(a.total_requests),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_eventbridge_sheet(wb, results, header_fill, header_font):
    """EventBridge 상세 시트 생성"""
    ws = wb.create_sheet("EventBridge")
    ws.append(
        [
            "Account",
            "Region",
            "Rule Name",
            "Event Bus",
            "State",
            "Schedule",
            "Targets",
            "Triggers",
            "상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                rule = f.rule
                ws.append(
                    [
                        rule.account_name,
                        rule.region,
                        rule.rule_name,
                        rule.event_bus_name,
                        rule.state,
                        rule.schedule_expression or "-",
                        rule.target_count,
                        int(rule.triggered_rules),
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_cw_alarm_sheet(wb, results, header_fill, header_font):
    """CloudWatch Alarm 상세 시트 생성"""
    ws = wb.create_sheet("CloudWatch Alarm")
    ws.append(
        [
            "Account",
            "Region",
            "Alarm Name",
            "Namespace",
            "Metric",
            "Dimensions",
            "State",
            "분석상태",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                a = f.alarm
                ws.append(
                    [
                        a.account_name,
                        a.region,
                        a.alarm_name,
                        a.namespace,
                        a.metric_name,
                        a.dimensions,
                        a.state,
                        f.status.value,
                        f.recommendation,
                    ]
                )


def _create_dynamodb_sheet(wb, results, header_fill, header_font):
    """DynamoDB 상세 시트 생성"""
    ws = wb.create_sheet("DynamoDB")
    ws.append(
        [
            "Account",
            "Region",
            "Table Name",
            "Billing Mode",
            "Items",
            "Size (MB)",
            "RCU",
            "WCU",
            "상태",
            "월간 비용",
            "권장 조치",
        ]
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for r in results:
        for f in r.findings:
            if f.status.value != "normal":
                t = f.table
                ws.append(
                    [
                        t.account_name,
                        t.region,
                        t.table_name,
                        t.billing_mode,
                        t.item_count,
                        f"{t.size_mb:.2f}",
                        t.provisioned_read,
                        t.provisioned_write,
                        f.status.value,
                        f"${t.estimated_monthly_cost:.2f}",
                        f.recommendation,
                    ]
                )
