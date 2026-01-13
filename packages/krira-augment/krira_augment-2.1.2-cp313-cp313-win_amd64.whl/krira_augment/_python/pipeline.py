"""
Pipeline Orchestrator for Krira Chunker V2.0.

Orchestrates the Clean -> Transform -> Chunk workflow.
This is the main entry point for users processing CSV and XLSX files.

Performance Requirements:
- Must process 1GB files in under 10 minutes on a 4-core CPU.
- Memory usage must remain constant O(1) regardless of file size.
- Uses streaming architecture (generators) throughout.
"""

import csv
import logging
import os
import sys
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Any, Dict, Generator, Iterator, List, Optional, Tuple, Union

# Import sibling modules
from .cleaning import CleaningConfig, DataCleaner
from .transformation import DataTransformer, TransformConfig


# Setup logger
LOGGER = logging.getLogger("krira_augment.pipeline")


# =============================================================================
# Configuration
# =============================================================================

@dataclass
class PipelineConfig:
    """
    Master configuration for the full Clean -> Transform -> Chunk pipeline.
    
    Combines CleaningConfig, TransformConfig, and chunk settings into a
    single configuration object.
    
    Attributes:
        cleaning_config: Configuration for the DataCleaner stage.
        transform_config: Configuration for the DataTransformer stage.
        chunk_config: Configuration for the chunking stage (optional, uses defaults).
        
        csv_batch_rows: Number of rows to process per batch for CSV files.
        xlsx_batch_rows: Number of rows to process per batch for XLSX files.
        log_progress_every: Log progress every N rows processed.
        encoding_fallbacks: List of encodings to try if UTF-8 fails.
        
    Example:
        >>> from krira_augment import PipelineConfig, CleaningConfig, TransformConfig
        >>> cfg = PipelineConfig(
        ...     cleaning_config=CleaningConfig(remove_headers=True),
        ...     transform_config=TransformConfig(output_format="markdown"),
        ... )
    """
    
    # Sub-configurations
    cleaning_config: CleaningConfig = field(default_factory=CleaningConfig)
    """Configuration for DataCleaner."""
    
    transform_config: TransformConfig = field(default_factory=TransformConfig)
    """Configuration for DataTransformer."""
    
    chunk_config: Optional[Any] = None
    """Configuration for chunking (ChunkConfig from Krira_Chunker). Uses defaults if None."""
    
    # Batch processing
    csv_batch_rows: int = 50_000
    """Number of rows to process per batch for CSV files."""
    
    xlsx_batch_rows: int = 25_000
    """Number of rows to process per batch for XLSX files."""
    
    # Progress logging
    log_progress_every: int = 100_000
    """Log progress status every N rows processed."""
    
    # Encoding fallbacks
    encoding_fallbacks: Tuple[str, ...] = ("utf-8", "latin-1", "cp1252", "utf-16")
    """List of encodings to try when reading files."""
    
    def __post_init__(self) -> None:
        """Validate configuration parameters."""
        if self.csv_batch_rows <= 0:
            raise ValueError(
                f"csv_batch_rows must be positive, got {self.csv_batch_rows}"
            )
        if self.xlsx_batch_rows <= 0:
            raise ValueError(
                f"xlsx_batch_rows must be positive, got {self.xlsx_batch_rows}"
            )


# =============================================================================
# Pipeline Exception Classes
# =============================================================================

class PipelineError(Exception):
    """Base exception for pipeline errors."""
    pass


class FileNotFoundPipelineError(PipelineError):
    """Raised when the input file is not found."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        super().__init__(f"File not found: {file_path}")


class PermissionPipelineError(PipelineError):
    """Raised when file permissions prevent access."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        super().__init__(
            f"Permission denied: {file_path}. "
            "Please check file permissions and try again."
        )


class UnsupportedFormatPipelineError(PipelineError):
    """Raised when the file format is not supported."""
    
    def __init__(self, file_path: str, extension: str):
        self.file_path = file_path
        self.extension = extension
        super().__init__(
            f"Unsupported file format: {extension}. "
            "Supported formats: .csv, .xlsx"
        )


# =============================================================================
# KriraPipeline
# =============================================================================

class KriraPipeline:
    """
    Orchestrates Clean -> Transform -> Chunk workflow.
    
    This is the main entry point for users. It chains DataCleaner,
    DataTransformer, and chunking in sequence.
    
    Features:
        - Streaming architecture for O(1) memory usage
        - Automatic file type detection (CSV, XLSX)
        - Multiple encoding fallback
        - Progress logging
        - Comprehensive error handling
        - Pipeline statistics
    
    Example:
        >>> from krira_augment import KriraPipeline, PipelineConfig
        >>> cfg = PipelineConfig()
        >>> pipeline = KriraPipeline(cfg)
        >>> for chunk in pipeline.process_file("data.csv"):
        ...     print(chunk["text"][:100])
    """
    
    def __init__(self, config: PipelineConfig) -> None:
        """
        Initialize all pipeline components.
        
        Args:
            config: Master configuration object.
            
        Implementation:
            - Instantiates DataCleaner with config.cleaning_config.
            - Instantiates DataTransformer with config.transform_config.
            - Validates that all configs are compatible.
            
        Raises:
            TypeError: If config is not a PipelineConfig instance.
        """
        if not isinstance(config, PipelineConfig):
            raise TypeError(
                f"config must be PipelineConfig, got {type(config).__name__}"
            )
        
        self.config = config
        
        # Initialize sub-components
        self.cleaner = DataCleaner(config.cleaning_config)
        self.transformer = DataTransformer(config.transform_config)
        
        # Lazy-load chunker (requires Krira_Chunker)
        self._chunker = None
        self._chunk_config = config.chunk_config
        
        # Pipeline statistics
        self._stats = {
            "rows_processed": 0,
            "chunks_created": 0,
            "bytes_cleaned": 0,
            "patterns_removed": 0,
            "files_processed": 0,
        }
        
        # Reduced batch size flag (for memory recovery)
        self._reduced_batch = False
    
    @property
    def chunker(self):
        """Lazy-load the chunker from Krira_Chunker."""
        if self._chunker is None:
            try:
                from Krira_Chunker import FastChunker, ChunkConfig
                
                if self._chunk_config is None:
                    self._chunk_config = ChunkConfig()
                
                self._chunker = FastChunker(self._chunk_config)
            except ImportError:
                LOGGER.warning(
                    "Krira_Chunker not available. "
                    "Using simple text chunking fallback."
                )
                self._chunker = None
        
        return self._chunker
    
    def _detect_separator(self, header_line: str) -> str:
        """
        Auto-detect CSV separator from header line.
        
        Args:
            header_line: First line of the CSV file.
            
        Returns:
            Detected separator character.
        """
        tab_count = header_line.count("\t")
        comma_count = header_line.count(",")
        semicolon_count = header_line.count(";")
        
        if tab_count > max(comma_count, semicolon_count):
            return "\t"
        elif semicolon_count > comma_count:
            return ";"
        return ","
    
    def _read_file_with_encoding(
        self, 
        file_path: str
    ) -> Tuple[StringIO, str]:
        """
        Read file with automatic encoding detection.
        
        Args:
            file_path: Path to the file.
            
        Returns:
            Tuple of (StringIO with content, detected encoding).
            
        Raises:
            UnicodeDecodeError: If no encoding works.
        """
        for encoding in self.config.encoding_fallbacks:
            try:
                with open(file_path, 'r', encoding=encoding, errors='strict') as f:
                    content = f.read()
                return StringIO(content), encoding
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        # Last resort: read with replace mode
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        return StringIO(content), 'utf-8-replace'
    
    def _process_csv_streaming(
        self, 
        file_path: str
    ) -> Generator[Dict[str, Any], None, None]:
        """
        Process CSV file in streaming mode.
        
        Args:
            file_path: Path to CSV file.
            
        Yields:
            Chunk dictionaries with 'text' and 'metadata'.
        """
        batch_size = self.config.csv_batch_rows
        
        if self._reduced_batch:
            batch_size = batch_size // 4
            LOGGER.info("Using reduced batch size: %d", batch_size)
        
        try:
            content_io, encoding = self._read_file_with_encoding(file_path)
            LOGGER.debug("Reading CSV with encoding: %s", encoding)
        except Exception as e:
            LOGGER.error("Failed to read file: %s", e)
            raise
        
        # Detect separator
        first_line = content_io.readline()
        separator = self._detect_separator(first_line)
        content_io.seek(0)
        
        try:
            reader = csv.reader(content_io, delimiter=separator)
        except csv.Error as e:
            LOGGER.error("CSV parsing error: %s", e)
            raise
        
        # Read header
        try:
            headers = next(reader)
        except StopIteration:
            LOGGER.warning("Empty CSV file: %s", file_path)
            return
        
        # Clean headers
        headers = [h.strip() or f"col_{i+1}" for i, h in enumerate(headers)]
        
        base_meta = {
            "source": os.path.basename(file_path),
            "source_path": os.path.abspath(file_path),
            "source_type": "csv",
            "encoding": encoding,
        }
        
        batch_texts: List[str] = []
        batch_row_ids: List[int] = []
        chunk_index = 0
        
        for row_num, row in enumerate(reader, start=1):
            try:
                # Step 1: Transform row to text
                row_dict = dict(zip(headers, row))
                row_text = self.transformer.transform_row(row_dict)
                
                if not row_text or not row_text.strip():
                    continue
                
                # Step 2: Clean the text
                cleaned_text = self.cleaner.clean_text(row_text)
                
                if not cleaned_text:
                    continue
                
                batch_texts.append(cleaned_text)
                batch_row_ids.append(row_num)
                self._stats["rows_processed"] += 1
                
                # Log progress
                if row_num % self.config.log_progress_every == 0:
                    LOGGER.info("Processed %d rows...", row_num)
                
                # Process batch
                if len(batch_texts) >= batch_size:
                    for chunk in self._chunk_batch(
                        batch_texts,
                        batch_row_ids,
                        base_meta,
                        chunk_index
                    ):
                        chunk_index = chunk["metadata"]["chunk_index"] + 1
                        self._stats["chunks_created"] += 1
                        yield chunk
                    
                    batch_texts = []
                    batch_row_ids = []
                    
            except csv.Error as e:
                LOGGER.warning("Skipping malformed row %d: %s", row_num, e)
                continue
        
        # Flush remaining batch
        if batch_texts:
            for chunk in self._chunk_batch(
                batch_texts,
                batch_row_ids,
                base_meta,
                chunk_index
            ):
                self._stats["chunks_created"] += 1
                yield chunk
        
        # Update cleaning stats
        cleaner_stats = self.cleaner.get_stats()
        self._stats["bytes_cleaned"] += cleaner_stats["bytes_cleaned"]
        self._stats["patterns_removed"] += cleaner_stats["patterns_removed"]
    
    def _process_xlsx_streaming(
        self, 
        file_path: str
    ) -> Generator[Dict[str, Any], None, None]:
        """
        Process XLSX file in streaming mode.
        
        Args:
            file_path: Path to XLSX file.
            
        Yields:
            Chunk dictionaries with 'text' and 'metadata'.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX processing. "
                "Install with: pip install openpyxl"
            )
        
        batch_size = self.config.xlsx_batch_rows
        
        if self._reduced_batch:
            batch_size = batch_size // 4
            LOGGER.info("Using reduced batch size: %d", batch_size)
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            LOGGER.error("Failed to open XLSX file: %s", e)
            raise
        
        chunk_index = 0
        
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                base_meta = {
                    "source": os.path.basename(file_path),
                    "source_path": os.path.abspath(file_path),
                    "source_type": "xlsx",
                    "sheet": sheet_name,
                }
                
                rows_iter = ws.iter_rows(values_only=True)
                
                # Get headers
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    LOGGER.debug("Empty sheet: %s", sheet_name)
                    continue
                
                headers = [
                    str(h).strip() if h is not None and str(h).strip() 
                    else f"col_{i+1}"
                    for i, h in enumerate(header_row)
                ]
                
                batch_texts: List[str] = []
                batch_row_ids: List[int] = []
                
                for row_num, row in enumerate(rows_iter, start=1):
                    try:
                        # Step 1: Transform row to text
                        row_text = self.transformer.excel_row_to_text(
                            headers, list(row)
                        )
                        
                        if not row_text or not row_text.strip():
                            continue
                        
                        # Step 2: Clean the text
                        cleaned_text = self.cleaner.clean_text(row_text)
                        
                        if not cleaned_text:
                            continue
                        
                        batch_texts.append(cleaned_text)
                        batch_row_ids.append(row_num)
                        self._stats["rows_processed"] += 1
                        
                        # Log progress
                        if row_num % self.config.log_progress_every == 0:
                            LOGGER.info(
                                "Processed %d rows from sheet '%s'...",
                                row_num, sheet_name
                            )
                        
                        # Process batch
                        if len(batch_texts) >= batch_size:
                            for chunk in self._chunk_batch(
                                batch_texts,
                                batch_row_ids,
                                base_meta,
                                chunk_index
                            ):
                                chunk_index = chunk["metadata"]["chunk_index"] + 1
                                self._stats["chunks_created"] += 1
                                yield chunk
                            
                            batch_texts = []
                            batch_row_ids = []
                            
                    except Exception as e:
                        LOGGER.warning(
                            "Skipping row %d in sheet '%s': %s",
                            row_num, sheet_name, e
                        )
                        continue
                
                # Flush remaining batch for this sheet
                if batch_texts:
                    for chunk in self._chunk_batch(
                        batch_texts,
                        batch_row_ids,
                        base_meta,
                        chunk_index
                    ):
                        chunk_index = chunk["metadata"]["chunk_index"] + 1
                        self._stats["chunks_created"] += 1
                        yield chunk
        
        finally:
            try:
                wb.close()
            except Exception:
                pass
        
        # Update cleaning stats
        cleaner_stats = self.cleaner.get_stats()
        self._stats["bytes_cleaned"] += cleaner_stats["bytes_cleaned"]
        self._stats["patterns_removed"] += cleaner_stats["patterns_removed"]
    
    def _chunk_batch(
        self,
        texts: List[str],
        row_ids: List[int],
        base_meta: Dict[str, Any],
        start_chunk_index: int
    ) -> Generator[Dict[str, Any], None, None]:
        """
        Chunk a batch of cleaned text rows.
        
        Args:
            texts: List of cleaned text strings.
            row_ids: Corresponding row IDs.
            base_meta: Base metadata for chunks.
            start_chunk_index: Starting chunk index.
            
        Yields:
            Chunk dictionaries.
        """
        if not texts:
            return
        
        # Use Krira_Chunker if available
        if self.chunker is not None:
            try:
                for chunk in self.chunker.chunk_units(
                    units=texts,
                    base_meta=base_meta,
                    joiner="\n",
                    locator=base_meta.get("source", "unknown"),
                    range_key="row",
                    range_vals=row_ids,
                    start_chunk_index=start_chunk_index,
                ):
                    yield chunk
                return
            except Exception as e:
                LOGGER.warning("Chunker failed, using fallback: %s", e)
        
        # Fallback: Simple chunking
        combined_text = "\n".join(texts)
        
        # Get chunk size from config or use default
        if self._chunk_config is not None:
            max_chars = getattr(self._chunk_config, 'max_chars', 2000)
        else:
            max_chars = 2000
        
        chunk_index = start_chunk_index
        start = 0
        
        while start < len(combined_text):
            end = min(start + max_chars, len(combined_text))
            
            # Try to break at newline
            if end < len(combined_text):
                newline_pos = combined_text.rfind('\n', start, end)
                if newline_pos > start:
                    end = newline_pos + 1
            
            chunk_text = combined_text[start:end].strip()
            
            if chunk_text:
                yield {
                    "text": chunk_text,
                    "metadata": {
                        **base_meta,
                        "chunk_index": chunk_index,
                        "char_start": start,
                        "char_end": end,
                    }
                }
                chunk_index += 1
            
            start = end
    
    def process_file(
        self, 
        file_path: str
    ) -> Generator[Dict[str, Any], None, None]:
        """
        Process a file through the full Clean -> Transform -> Chunk pipeline.
        
        Args:
            file_path: Path to input file (CSV or XLSX).
            
        Yields:
            Chunk dictionaries with keys: 'text', 'metadata'.
            
        Algorithm:
            1. Validate file exists and is accessible.
            2. Detect file type from extension.
            3. Open file in streaming mode.
            4. For each row batch:
               a. Transform row to text format.
               b. Pass through cleaner.clean_text().
               c. Pass through chunker.
               d. Yield chunks.
            5. Close file handle.
            
        Performance Requirements:
            - Uses generators throughout (no list accumulation).
            - Memory usage stays under 500MB for 10GB files.
            - Logs progress every log_progress_every rows.
            
        Raises:
            FileNotFoundPipelineError: If file does not exist.
            PermissionPipelineError: If file cannot be read.
            UnsupportedFormatPipelineError: If file format not supported.
        """
        # Validate file exists
        if not os.path.exists(file_path):
            raise FileNotFoundPipelineError(file_path)
        
        # Check permissions
        if not os.access(file_path, os.R_OK):
            raise PermissionPipelineError(file_path)
        
        # Detect file type
        path = Path(file_path)
        extension = path.suffix.lower()
        
        self._stats["files_processed"] += 1
        LOGGER.info("Processing file: %s", file_path)
        
        try:
            if extension == ".csv":
                yield from self._process_csv_streaming(file_path)
            elif extension == ".xlsx":
                yield from self._process_xlsx_streaming(file_path)
            else:
                raise UnsupportedFormatPipelineError(file_path, extension)
                
        except MemoryError:
            # Handle memory errors by reducing batch size
            if not self._reduced_batch:
                LOGGER.warning(
                    "MemoryError encountered. Reducing batch size and retrying..."
                )
                self._reduced_batch = True
                
                # Reset stats for retry
                self._stats["rows_processed"] = 0
                self._stats["chunks_created"] = 0
                
                if extension == ".csv":
                    yield from self._process_csv_streaming(file_path)
                elif extension == ".xlsx":
                    yield from self._process_xlsx_streaming(file_path)
            else:
                raise
        
        LOGGER.info(
            "Completed: %d rows -> %d chunks",
            self._stats["rows_processed"],
            self._stats["chunks_created"]
        )
    
    def process_text(self, text: str) -> Generator[Dict[str, Any], None, None]:
        """
        Process raw text through Clean -> Chunk pipeline.
        
        Args:
            text: Raw input text.
            
        Yields:
            Chunk dictionaries.
        """
        # Clean the text
        cleaned = self.cleaner.clean_text(text)
        
        if not cleaned:
            return
        
        base_meta = {
            "source": "text_input",
            "source_type": "text",
        }
        
        yield from self._chunk_batch(
            texts=[cleaned],
            row_ids=[0],
            base_meta=base_meta,
            start_chunk_index=0
        )
    
    def get_stats(self) -> Dict[str, int]:
        """
        Return pipeline statistics after processing.
        
        Returns:
            Dictionary with keys:
            - 'rows_processed': Total rows read.
            - 'chunks_created': Total chunks generated.
            - 'bytes_cleaned': Total text bytes cleaned.
            - 'patterns_removed': Count of regex matches removed.
            - 'files_processed': Number of files processed.
        """
        return dict(self._stats)
    
    def reset_stats(self) -> None:
        """Reset all statistics counters."""
        self._stats = {
            "rows_processed": 0,
            "chunks_created": 0,
            "bytes_cleaned": 0,
            "patterns_removed": 0,
            "files_processed": 0,
        }
        self.cleaner.reset_stats()
        self.transformer.reset_stats()
