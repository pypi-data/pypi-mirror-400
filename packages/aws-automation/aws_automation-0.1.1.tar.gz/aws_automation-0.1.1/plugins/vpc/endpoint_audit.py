"""
plugins/vpc/endpoint_audit.py - VPC Endpoint 미사용 분석

미사용 VPC Endpoint 탐지 (Interface Endpoint)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_endpoint_monthly_cost

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeVpcEndpoints",
        "cloudwatch:GetMetricStatistics",
    ],
}


class EndpointStatus(Enum):
    """VPC Endpoint 상태"""

    NORMAL = "normal"
    UNUSED = "unused"  # 트래픽 없음
    PENDING = "pending"  # 상태 pending


@dataclass
class VPCEndpointInfo:
    """VPC Endpoint 정보"""

    account_id: str
    account_name: str
    region: str
    endpoint_id: str
    endpoint_type: str  # Interface, Gateway, GatewayLoadBalancer
    service_name: str
    vpc_id: str
    state: str
    creation_time: datetime | None
    name: str = ""

    @property
    def is_interface(self) -> bool:
        return self.endpoint_type == "Interface"

    @property
    def monthly_cost(self) -> float:
        # Gateway Endpoints (S3, DynamoDB) are free
        if self.endpoint_type == "Gateway":
            return 0.0
        return get_endpoint_monthly_cost(self.region)


@dataclass
class EndpointFinding:
    """Endpoint 분석 결과"""

    endpoint: VPCEndpointInfo
    status: EndpointStatus
    recommendation: str


@dataclass
class EndpointAnalysisResult:
    """Endpoint 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    interface_count: int = 0
    gateway_count: int = 0
    unused_count: int = 0
    normal_count: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[EndpointFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_endpoints(session, account_id: str, account_name: str, region: str) -> list[VPCEndpointInfo]:
    """VPC Endpoints 수집"""
    ec2 = get_client(session, "ec2", region_name=region)
    endpoints = []

    paginator = ec2.get_paginator("describe_vpc_endpoints")
    for page in paginator.paginate():
        for ep in page.get("VpcEndpoints", []):
            # Name 태그 추출
            name = ""
            for tag in ep.get("Tags", []):
                if tag.get("Key") == "Name":
                    name = tag.get("Value", "")
                    break

            creation_time = ep.get("CreationTimestamp")
            if creation_time and not isinstance(creation_time, datetime):
                creation_time = None

            endpoints.append(
                VPCEndpointInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    endpoint_id=ep["VpcEndpointId"],
                    endpoint_type=ep.get("VpcEndpointType", "Unknown"),
                    service_name=ep.get("ServiceName", ""),
                    vpc_id=ep.get("VpcId", ""),
                    state=ep.get("State", ""),
                    creation_time=creation_time,
                    name=name,
                )
            )

    return endpoints


def check_endpoint_usage(session, region: str, endpoint_id: str, days: int = 7) -> bool:
    """
    CloudWatch 메트릭으로 Endpoint 사용량 확인
    BytesProcessed 또는 ActiveConnections 메트릭 확인
    """
    from botocore.exceptions import ClientError

    cloudwatch = get_client(session, "cloudwatch", region_name=region)

    end_time = datetime.now(timezone.utc)
    start_time = end_time - timedelta(days=days)

    try:
        # PrivateLink 메트릭 확인
        response = cloudwatch.get_metric_statistics(
            Namespace="AWS/PrivateLinkEndpoints",
            MetricName="BytesProcessed",
            Dimensions=[
                {"Name": "VPC Id", "Value": endpoint_id.split("-")[0]},  # Approximation
                {"Name": "Endpoint Type", "Value": "Interface"},
            ],
            StartTime=start_time,
            EndTime=end_time,
            Period=86400,  # 1 day
            Statistics=["Sum"],
        )

        datapoints = response.get("Datapoints", [])
        if datapoints:
            total_bytes = sum(dp.get("Sum", 0) for dp in datapoints)
            return bool(total_bytes > 0)

    except ClientError:
        pass

    # 메트릭이 없으면 사용 중으로 간주 (보수적 접근)
    return True


# =============================================================================
# 분석
# =============================================================================


def analyze_endpoints(
    endpoints: list[VPCEndpointInfo],
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> EndpointAnalysisResult:
    """VPC Endpoint 분석"""
    result = EndpointAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(endpoints),
    )

    for ep in endpoints:
        if ep.endpoint_type == "Gateway":
            result.gateway_count += 1
            # Gateway는 무료이므로 분석 제외
            result.normal_count += 1
            result.findings.append(
                EndpointFinding(
                    endpoint=ep,
                    status=EndpointStatus.NORMAL,
                    recommendation="Gateway Endpoint (무료)",
                )
            )
            continue

        result.interface_count += 1

        # Pending 상태 체크
        if ep.state.lower() in ("pending", "pendingacceptance"):
            result.findings.append(
                EndpointFinding(
                    endpoint=ep,
                    status=EndpointStatus.PENDING,
                    recommendation="Pending 상태 - 연결 확인 필요",
                )
            )
            continue

        # 사용량 체크는 시간이 오래 걸리므로 기본적으로 normal로 처리
        # 실제로는 CloudWatch 메트릭 기반 체크 필요
        # 여기서는 state가 available이 아닌 경우만 체크
        if ep.state.lower() != "available":
            result.unused_count += 1
            result.unused_monthly_cost += ep.monthly_cost
            result.findings.append(
                EndpointFinding(
                    endpoint=ep,
                    status=EndpointStatus.UNUSED,
                    recommendation=f"상태 비정상: {ep.state}",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            EndpointFinding(
                endpoint=ep,
                status=EndpointStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[EndpointAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "VPC Endpoint 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = ["Account", "Region", "전체", "Interface", "Gateway", "미사용", "월간 비용"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_count)
        ws.cell(row=row, column=4, value=r.interface_count)
        ws.cell(row=row, column=5, value=r.gateway_count)
        ws.cell(row=row, column=6, value=r.unused_count)
        ws.cell(row=row, column=7, value=f"${r.unused_monthly_cost:,.2f}")

    # 상세
    ws_detail = wb.create_sheet("Endpoints")
    detail_headers = [
        "Account",
        "Region",
        "Endpoint ID",
        "Name",
        "Type",
        "Service",
        "VPC",
        "State",
        "월간 비용",
        "상태",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            detail_row += 1
            ep = f.endpoint
            ws_detail.cell(row=detail_row, column=1, value=ep.account_name)
            ws_detail.cell(row=detail_row, column=2, value=ep.region)
            ws_detail.cell(row=detail_row, column=3, value=ep.endpoint_id)
            ws_detail.cell(row=detail_row, column=4, value=ep.name or "-")
            ws_detail.cell(row=detail_row, column=5, value=ep.endpoint_type)
            ws_detail.cell(row=detail_row, column=6, value=ep.service_name.split(".")[-1])
            ws_detail.cell(row=detail_row, column=7, value=ep.vpc_id)
            ws_detail.cell(row=detail_row, column=8, value=ep.state)
            ws_detail.cell(row=detail_row, column=9, value=f"${ep.monthly_cost:,.2f}")
            ws_detail.cell(row=detail_row, column=10, value=f.recommendation)

    # 열 너비
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"VPC_Endpoint_Audit_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EndpointAnalysisResult | None:
    """단일 계정/리전의 VPC Endpoint 수집 및 분석 (병렬 실행용)"""
    endpoints = collect_endpoints(session, account_id, account_name, region)
    if not endpoints:
        return None

    return analyze_endpoints(endpoints, session, account_id, account_name, region)


def run(ctx) -> None:
    """VPC Endpoint 미사용 분석"""
    console.print("[bold]VPC Endpoint 분석 시작...[/bold]\n")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ec2")

    # None 필터링
    results: list[EndpointAnalysisResult] = [r for r in result.get_data() if r is not None]

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_interface = sum(r.interface_count for r in results)
    total_gateway = sum(r.gateway_count for r in results)
    total_cost = sum(r.interface_count * get_endpoint_monthly_cost(r.region) for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"Interface Endpoint: {total_interface}개 (${total_cost:,.2f}/월)")
    console.print(f"Gateway Endpoint: {total_gateway}개 (무료)")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("endpoint-audit").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
