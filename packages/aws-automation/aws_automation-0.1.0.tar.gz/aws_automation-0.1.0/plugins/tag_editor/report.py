"""
plugins/tag_editor/report.py - MAP 태그 Excel 리포트 생성

분석 및 적용 결과를 Excel 파일로 출력
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .types import (
    MAP_TAG_KEY,
    MapTagAnalysisResult,
    MapTagApplyResult,
    TagOperationResult,
)

# =============================================================================
# 스타일 상수
# =============================================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _adjust_column_widths(wb: Workbook) -> None:
    """모든 시트의 열 너비 조정"""
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            col_idx = col[0].column
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"


# =============================================================================
# 분석 리포트
# =============================================================================


def generate_audit_report(
    results: list[MapTagAnalysisResult],
    output_dir: str,
    untagged_only: bool = False,
) -> str:
    """MAP 태그 분석 Excel 리포트 생성"""
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # ===== Summary 시트 =====
    ws = wb.create_sheet("Summary")
    ws["A1"] = "MAP 태그 분석 리포트"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"분석 태그: {MAP_TAG_KEY}"
    ws["A3"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # 전체 통계
    total_resources = sum(r.total_resources for r in results)
    total_tagged = sum(r.tagged_resources for r in results)
    total_untagged = sum(r.untagged_resources for r in results)
    overall_rate = (total_tagged / total_resources * 100) if total_resources > 0 else 0

    row = 5
    ws.cell(row=row, column=1, value="전체 현황").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="총 리소스")
    ws.cell(row=row, column=2, value=total_resources)
    row += 1
    ws.cell(row=row, column=1, value="태그됨")
    ws.cell(row=row, column=2, value=total_tagged)
    ws.cell(row=row, column=2).fill = GREEN_FILL
    row += 1
    ws.cell(row=row, column=1, value="미태그")
    ws.cell(row=row, column=2, value=total_untagged)
    if total_untagged > 0:
        ws.cell(row=row, column=2).fill = RED_FILL
    row += 1
    ws.cell(row=row, column=1, value="적용률")
    ws.cell(row=row, column=2, value=f"{overall_rate:.1f}%")

    # 계정/리전별 통계
    row += 2
    ws.cell(row=row, column=1, value="계정/리전별 현황").font = Font(bold=True, size=12)
    row += 1

    headers = ["Account", "Region", "전체", "태그됨", "미태그", "적용률"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = HEADER_FONT

    for r in results:
        row += 1
        rate = (r.tagged_resources / r.total_resources * 100) if r.total_resources > 0 else 0
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_resources)
        ws.cell(row=row, column=4, value=r.tagged_resources)
        ws.cell(row=row, column=5, value=r.untagged_resources)
        ws.cell(row=row, column=6, value=f"{rate:.1f}%")
        if r.untagged_resources > 0:
            ws.cell(row=row, column=5).fill = RED_FILL

    # ===== 리소스 타입별 시트 =====
    ws_type = wb.create_sheet("By Resource Type")

    # 타입별 통계 집계
    type_totals: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "tagged": 0})
    for r in results:
        for ts in r.type_stats:
            type_totals[ts.resource_type]["total"] += ts.total
            type_totals[ts.resource_type]["tagged"] += ts.tagged

    headers = ["리소스 타입", "전체", "태그됨", "미태그", "적용률"]
    for col, h in enumerate(headers, 1):
        ws_type.cell(row=1, column=col, value=h).fill = HEADER_FILL
        ws_type.cell(row=1, column=col).font = HEADER_FONT

    row = 1
    for res_type, counts in sorted(type_totals.items(), key=lambda x: x[1]["total"], reverse=True):
        row += 1
        total = counts["total"]
        tagged = counts["tagged"]
        untagged = total - tagged
        rate = (tagged / total * 100) if total > 0 else 0

        # 표시 이름
        parts = res_type.split(":")
        display = " ".join(p.capitalize() for p in parts)

        ws_type.cell(row=row, column=1, value=display)
        ws_type.cell(row=row, column=2, value=total)
        ws_type.cell(row=row, column=3, value=tagged)
        ws_type.cell(row=row, column=4, value=untagged)
        ws_type.cell(row=row, column=5, value=f"{rate:.1f}%")
        if untagged > 0:
            ws_type.cell(row=row, column=4).fill = RED_FILL

    # ===== 미태그 리소스 시트 =====
    ws_untagged = wb.create_sheet("Untagged Resources")
    headers = ["Account", "Region", "Type", "Resource ID", "Name", "ARN"]
    for col, h in enumerate(headers, 1):
        ws_untagged.cell(row=1, column=col, value=h).fill = HEADER_FILL
        ws_untagged.cell(row=1, column=col).font = HEADER_FONT

    row = 1
    for r in results:
        for res in r.resources:
            if not res.has_map_tag:
                row += 1
                # 표시 이름
                parts = res.resource_type.split(":")
                type_display = " ".join(p.capitalize() for p in parts)

                ws_untagged.cell(row=row, column=1, value=res.account_name)
                ws_untagged.cell(row=row, column=2, value=res.region)
                ws_untagged.cell(row=row, column=3, value=type_display)
                ws_untagged.cell(row=row, column=4, value=res.resource_id)
                ws_untagged.cell(row=row, column=5, value=res.name or "-")
                ws_untagged.cell(row=row, column=6, value=res.resource_arn)

    # ===== 태그된 리소스 시트 (옵션) =====
    if not untagged_only:
        ws_tagged = wb.create_sheet("Tagged Resources")
        headers = [
            "Account",
            "Region",
            "Type",
            "Resource ID",
            "Name",
            "MAP Tag Value",
            "ARN",
        ]
        for col, h in enumerate(headers, 1):
            ws_tagged.cell(row=1, column=col, value=h).fill = HEADER_FILL
            ws_tagged.cell(row=1, column=col).font = HEADER_FONT

        row = 1
        for r in results:
            for res in r.resources:
                if res.has_map_tag:
                    row += 1
                    parts = res.resource_type.split(":")
                    type_display = " ".join(p.capitalize() for p in parts)

                    ws_tagged.cell(row=row, column=1, value=res.account_name)
                    ws_tagged.cell(row=row, column=2, value=res.region)
                    ws_tagged.cell(row=row, column=3, value=type_display)
                    ws_tagged.cell(row=row, column=4, value=res.resource_id)
                    ws_tagged.cell(row=row, column=5, value=res.name or "-")
                    ws_tagged.cell(row=row, column=6, value=res.map_tag_value or "-")
                    ws_tagged.cell(row=row, column=7, value=res.resource_arn)

    # ===== 전체 리소스 시트 =====
    ws_all = wb.create_sheet("All Resources")
    headers = [
        "Account",
        "Region",
        "Type",
        "Resource ID",
        "Name",
        "MAP Tagged",
        "MAP Value",
        "ARN",
    ]
    for col, h in enumerate(headers, 1):
        ws_all.cell(row=1, column=col, value=h).fill = HEADER_FILL
        ws_all.cell(row=1, column=col).font = HEADER_FONT

    row = 1
    for r in results:
        for res in r.resources:
            row += 1
            parts = res.resource_type.split(":")
            type_display = " ".join(p.capitalize() for p in parts)

            ws_all.cell(row=row, column=1, value=res.account_name)
            ws_all.cell(row=row, column=2, value=res.region)
            ws_all.cell(row=row, column=3, value=type_display)
            ws_all.cell(row=row, column=4, value=res.resource_id)
            ws_all.cell(row=row, column=5, value=res.name or "-")

            tagged_cell = ws_all.cell(row=row, column=6, value="Yes" if res.has_map_tag else "No")
            if res.has_map_tag:
                tagged_cell.fill = GREEN_FILL
            else:
                tagged_cell.fill = RED_FILL

            ws_all.cell(row=row, column=7, value=res.map_tag_value or "-")
            ws_all.cell(row=row, column=8, value=res.resource_arn)

    # 열 너비 조정 및 저장
    _adjust_column_widths(wb)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"MAP_Tag_Audit_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 적용 리포트
# =============================================================================


def generate_apply_report(
    results: list[MapTagApplyResult],
    output_dir: str,
) -> str:
    """MAP 태그 적용 결과 Excel 리포트 생성"""
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # ===== Summary 시트 =====
    ws = wb.create_sheet("Summary")
    ws["A1"] = "MAP 태그 적용 결과 리포트"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # 전체 통계
    total_targeted = sum(r.total_targeted for r in results)
    total_success = sum(r.success_count for r in results)
    total_failed = sum(r.failed_count for r in results)
    total_skipped = sum(r.skipped_count for r in results)

    row = 4
    ws.cell(row=row, column=1, value="적용 결과").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="대상 리소스")
    ws.cell(row=row, column=2, value=total_targeted)
    row += 1
    ws.cell(row=row, column=1, value="성공")
    ws.cell(row=row, column=2, value=total_success)
    ws.cell(row=row, column=2).fill = GREEN_FILL
    row += 1
    ws.cell(row=row, column=1, value="실패")
    ws.cell(row=row, column=2, value=total_failed)
    if total_failed > 0:
        ws.cell(row=row, column=2).fill = RED_FILL
    row += 1
    ws.cell(row=row, column=1, value="스킵")
    ws.cell(row=row, column=2, value=total_skipped)

    # 계정/리전별 결과
    row += 2
    ws.cell(row=row, column=1, value="계정/리전별 결과").font = Font(bold=True, size=12)
    row += 1

    headers = ["Account", "Region", "태그 값", "대상", "성공", "실패", "스킵"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = HEADER_FONT

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.tag_value)
        ws.cell(row=row, column=4, value=r.total_targeted)
        ws.cell(row=row, column=5, value=r.success_count)
        ws.cell(row=row, column=6, value=r.failed_count)
        ws.cell(row=row, column=7, value=r.skipped_count)
        if r.failed_count > 0:
            ws.cell(row=row, column=6).fill = RED_FILL

    # ===== 상세 로그 시트 =====
    ws_log = wb.create_sheet("Operation Log")
    headers = [
        "Account",
        "Region",
        "Type",
        "Resource ID",
        "Name",
        "Operation",
        "Result",
        "Error",
        "Previous",
        "New",
    ]
    for col, h in enumerate(headers, 1):
        ws_log.cell(row=1, column=col, value=h).fill = HEADER_FILL
        ws_log.cell(row=1, column=col).font = HEADER_FONT

    row = 1
    for r in results:
        for log in r.operation_logs:
            row += 1
            parts = log.resource_type.split(":")
            type_display = " ".join(p.capitalize() for p in parts)

            ws_log.cell(row=row, column=1, value=r.account_name)
            ws_log.cell(row=row, column=2, value=r.region)
            ws_log.cell(row=row, column=3, value=type_display)
            ws_log.cell(row=row, column=4, value=log.resource_id)
            ws_log.cell(row=row, column=5, value=log.name or "-")
            ws_log.cell(row=row, column=6, value=log.operation)

            result_cell = ws_log.cell(row=row, column=7, value=log.result.value)
            if log.result == TagOperationResult.SUCCESS:
                result_cell.fill = GREEN_FILL
            elif log.result == TagOperationResult.FAILED:
                result_cell.fill = RED_FILL
            else:
                result_cell.fill = YELLOW_FILL

            ws_log.cell(row=row, column=8, value=log.error_message or "-")
            ws_log.cell(row=row, column=9, value=log.previous_value or "-")
            ws_log.cell(row=row, column=10, value=log.new_value or "-")

    # 열 너비 조정 및 저장
    _adjust_column_widths(wb)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"MAP_Tag_Apply_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath
