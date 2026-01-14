"""
测试稀疏数据处理（大量空行）
"""
import pytest
import base64
import openpyxl
import io
from free_mcp_excel.parser import ExcelParser
from free_mcp_excel.utils import excel_to_base64


@pytest.fixture
def parser():
    """创建解析器实例"""
    return ExcelParser()


@pytest.fixture
def sparse_file():
    """创建包含大量空行的测试文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "稀疏数据"
    
    # 第1行：标题
    ws["A1"] = "序号"
    ws["B1"] = "名称"
    ws["C1"] = "数值"
    
    # 第2-5行：第一组数据
    for i in range(2, 6):
        ws[f"A{i}"] = i - 1
        ws[f"B{i}"] = f"数据组1-项目{i-1}"
        ws[f"C{i}"] = (i - 1) * 100
    
    # 第6-20行：空行（不写入任何内容）
    
    # 第21-25行：第二组数据
    for i in range(21, 26):
        ws[f"A{i}"] = i - 16
        ws[f"B{i}"] = f"数据组2-项目{i-20}"
        ws[f"C{i}"] = (i - 16) * 100
    
    # 第26-50行：空行
    
    # 第51-55行：第三组数据
    for i in range(51, 56):
        ws[f"A{i}"] = i - 41
        ws[f"B{i}"] = f"数据组3-项目{i-50}"
        ws[f"C{i}"] = (i - 41) * 100
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
    wb.close()
    
    return file_base64


class TestSparseData:
    """稀疏数据测试类"""
    
    def test_read_sparse_data_skip_empty_rows(self, parser, sparse_file):
        """测试读取稀疏数据（跳过空行）"""
        result = parser.read_sheet_data(sparse_file, "稀疏数据", skip_empty_rows=True)
        
        assert result["status"] == "success"
        # 应该只返回有数据的行：4行第一组 + 5行第二组 + 5行第三组 = 14行（不含标题）
        assert result["data"]["row_count"] == 14
        assert len(result["data"]["data_rows"]) == 14
        # 验证第一行数据
        assert result["data"]["data_rows"][0] == [1, "数据组1-项目1", 100]
        # 验证第二组数据的第一行（跳过空行后应该是第5行）
        assert result["data"]["data_rows"][4] == [5, "数据组2-项目1", 500]
    
    def test_read_sparse_data_keep_empty_rows(self, parser, sparse_file):
        """测试读取稀疏数据（保留空行）"""
        result = parser.read_sheet_data(sparse_file, "稀疏数据", skip_empty_rows=False)
        
        assert result["status"] == "success"
        # openpyxl的max_row是55（第55行有数据），所以总行数应该是54行数据（不含标题）
        assert result["data"]["row_count"] == 54  # 55行数据（不含标题），索引0-53
        # 验证空行被表示为None
        # 第5行之后应该是空行（第6-20行）
        assert result["data"]["data_rows"][5] == [None, None, None]  # 第6行（索引5）
        # 注意：openpyxl的max_row可能不包含完全空白的行
        # 验证至少包含空行（第6行是空行）
        assert result["data"]["data_rows"][5] == [None, None, None]
        # 验证有数据的行被正确读取
        # 查找第二组数据（应该在索引20附近）
        data_rows = result["data"]["data_rows"]
        # 找到包含"数据组2"的行
        group2_rows = [i for i, row in enumerate(data_rows) if row and row[1] and "数据组2" in str(row[1])]
        assert len(group2_rows) > 0, "应该找到第二组数据"
        # 验证第二组数据的第一行
        assert data_rows[group2_rows[0]] == [5, "数据组2-项目1", 500]
    
    def test_read_sparse_data_range(self, parser, sparse_file):
        """测试范围读取稀疏数据"""
        # 读取A1:C30，包含第一组数据和空行区域
        result = parser.read_sheet_data(
            sparse_file, 
            "稀疏数据", 
            range="A1:C30",
            skip_empty_rows=True
        )
        
        assert result["status"] == "success"
        # 应该只返回有数据的行
        assert result["data"]["row_count"] == 9  # 4行第一组 + 5行第二组
        assert len(result["data"]["data_rows"]) == 9
    
    def test_read_cell_in_empty_row(self, parser, sparse_file):
        """测试读取空行中的单元格"""
        # 读取第10行（空行）的单元格
        result = parser.read_cell_data(sparse_file, "稀疏数据", "A10")
        
        assert result["status"] == "success"
        # 空单元格应该返回None
        assert result["data"]["cells"][0]["value"] is None
    
    def test_read_cell_in_data_row(self, parser, sparse_file):
        """测试读取有数据行中的单元格"""
        # 读取第21行（有数据）的单元格
        result = parser.read_cell_data(sparse_file, "稀疏数据", "A21")
        
        assert result["status"] == "success"
        assert result["data"]["cells"][0]["value"] == 5
    
    def test_read_large_sparse_range(self, parser, sparse_file):
        """测试读取大范围稀疏数据"""
        # 读取整个工作表
        result = parser.read_sheet_data(
            sparse_file,
            "稀疏数据",
            skip_empty_rows=True
        )
        
        assert result["status"] == "success"
        # 应该只返回有数据的行：4行第一组 + 5行第二组 + 5行第三组 = 14行
        assert result["data"]["row_count"] == 14
        # 验证所有数据组都被正确读取
        data_rows = result["data"]["data_rows"]
        # 第一组数据
        assert data_rows[0][0] == 1
        assert data_rows[3][0] == 4
        # 第二组数据
        assert data_rows[4][0] == 5
        assert data_rows[8][0] == 9
        # 第三组数据
        assert data_rows[9][0] == 10
        assert data_rows[13][0] == 14

