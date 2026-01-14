import json
from datetime import datetime
from pprint import pprint
from bson import ObjectId

from pollenisatorgui.core.components.apiclient import APIClient
from pollenisatorgui.pollenisator import consoleConnect
    
# Create a new worksheet using openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def testsslresults(apiclient):
    is_sncf: bool = False
    client_settings = apiclient.findInDb(apiclient.getCurrentPentest(), "settings", {"key":"client_name"}, multi=False)
    if client_settings is not None:
        is_sncf = "sncf" in client_settings.get("value", "").lower()
    tags_data = apiclient.findInDb(apiclient.getCurrentPentest(), "tags", { "tags.name": "SSL/TLS-flaws" }, multi=True)
    port_ids = []
    if tags_data is not None:
        for tag in tags_data:
            if tag.get("item_type","") == "ports":
                port_ids.append(tag.get("item_id",""))
    expectedVulns = ["SSLv2","SSLv3","TLS1","TLS1_1","TLS1_2","TLS1_3","FS","cipherlist_NULL","cipherlist_aNULL","cipherlist_EXPORT","cipherlist_DES+64Bit","cipherlist_128Bit","cipherlist_3DES","cipher_order","secure_renego","secure_client_renego","fallback_SCSV","heartbleed","ticketbleed","CCS","ROBOT","POODLE_SSL","CRIME_TLS","BREACH","SWEET32","FREAK","DROWN","LOGJAM","BAR_MITZVAH","LUCKY13","RC4","BEAST","BEAST_CBC_TLS1"]
    list_of_all_ports = apiclient.findInDb(apiclient.getCurrentPentest(), "ports", { "_id": { "$in": port_ids } }, multi=True)
    list_of_all_lines = []
    list_of_defects = []
    dict_line = {}
    for port in list_of_all_ports:
        list_of_defects = port.get("infos", {}).get("TestSSL", [])
        if not list_of_defects:
            continue
        
        # Add missing defects as "OK" status
        for defect_id in expectedVulns:
            if not any(defect.get("defect") == defect_id for defect in list_of_defects):
                list_of_defects.append({"defect": defect_id, "criticity": "OK", "details": ""})
        for vuln in list_of_defects:
            dict_line = {"ip": port.get("ip",""), "port": port.get("port",""), "domain": port.get("infos", {}).get("FQDN",""), "defect": vuln.get("defect",""), "criticity": vuln.get("criticity",""), "details": vuln.get("details","")}
            list_of_all_lines.append(dict_line)
    vulnDataList = []
    completeData = []
    newTLSDefect = set()
    notAfter = None
    notBefore = None
    lifespan = -1
    keySizeInfo = None
    TLSv1_0_or_1_1 = False
    already_added = False
    
    for vuln in list_of_all_lines:
        # Check if SSLv2 is enabled
        vuln["defect"] = vuln["defect"].strip()
        if vuln["defect"] == "SSLv2":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "SSLv2 Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SSLv2", "vulnerable": "yes", "state": "Enabled"})
                # newTLSDefect.add(json.dumps(["SSLv2 activé", "7"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SSLv2", "vulnerable": "no", "state": "Disabled"})
        
        # Check if SSLv3 is enabled
        elif vuln["defect"] == "SSLv3":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "SSLv3 Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SSLv3", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["SSLv3 activé", "7"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SSLv3", "vulnerable": "no", "state": "Disabled"})
        
        # Check if TLSv1.0 is enabled
        elif vuln["defect"] == "TLS1":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TLSv1.0 Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.0", "vulnerable": "yes", "state": "Enabled"})
                TLSv1_0_or_1_1 = True
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.0", "vulnerable": "no", "state": "Disabled"})
        
        # Check if TLSv1.1 is enabled
        elif vuln["defect"] == "TLS1_1":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TLSv1.1 Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.1", "vulnerable": "yes", "state": "Enabled"})
                TLSv1_0_or_1_1 = True
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.1", "vulnerable": "no", "state": "Disabled"})
        
        # Check if one of the TLSv1.0 or TLSv1.1 is enabled
        elif TLSv1_0_or_1_1 == True and already_added == False:
            newTLSDefect.add(json.dumps(["TLSv1.0 et TLSv1.1 activés", "21,66"]))
            already_added = True
        
        # Check if TLSv1.2 is disabled
        if vuln["defect"] == "TLS1_2":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TLSv1.2 Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.2", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["TLSv1.2 et/ou TLSv1.3 non-offert", "66"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.2", "vulnerable": "no", "state": "Enabled"})
        
        # Check if TLSv1.3 is disabled
        elif vuln["defect"] == "TLS1_3":
            if (vuln["criticity"] not in ["OK", "INFO"]) or ("not offered" in vuln["details"].lower()):
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TLSv1.3 Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.3", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["TLSv1.2 et/ou TLSv1.3 non-offert", "66"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLSv1.3", "vulnerable": "no", "state": "Enabled"})
        
        # Check if Perfect Forward Secrecy is disabled
        elif vuln["defect"] == "FS":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Perfect Forward Secrecy Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "PFS", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Pas de Perfect Forward Secrecy", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "PFS", "vulnerable": "no", "state": "Enabled"})
        
        # Check if NULL/eNULL Cipher is enabled
        elif vuln["defect"] == "cipherlist_NULL":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "NULL/eNULL Ciphersuites Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "NULL/eNULL Ciphersuites", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["Suites de chiffrement NULL/eNULL", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "NULL/eNULL Ciphersuites", "vulnerable": "no", "state": "Disabled"})
        
        # Check if NULL authentication is enabled
        elif vuln["defect"] == "cipherlist_aNULL":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "NULL Authentication Ciphersuites Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "aNULL Ciphersuites", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["Authentification NULL", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "aNULL Ciphersuites", "vulnerable": "no", "state": "Disabled"})
        
        # Check if Export Cipher is enabled
        elif vuln["defect"] == "cipherlist_EXPORT":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Export Ciphersuites Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "EXPORT Ciphersuites", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["Suites de chiffrement « Export »", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "EXPORT Ciphersuites", "vulnerable": "no", "state": "Disabled"})
        
        # Check if DES Ciphersuites have a key length < 64 bits
        elif vuln["defect"] == "cipherlist_DES+64Bit":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "DES Ciphersuites with keysize < 64 Bits", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "DES Ciphersuites with keysize < 64 Bits", "vulnerable": "yes", "state": "True"})
                # newTLSDefect.add(json.dumps(["Suites de chiffrement DES de clé < 64 bits", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "DES Ciphersuites with keysize < 64 Bits", "vulnerable": "no", "state": "False"})
        
        # Check if Ciphersuites have key length < 128 bits
        elif vuln["defect"] == "cipherlist_128Bit":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Ciphersuites with keysize < 128 Bits", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Ciphersuites with keysize < 128 Bits", "vulnerable": "yes", "state": "True"})
                newTLSDefect.add(json.dumps(["Suites de chiffrement de clé < 128 bits", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Ciphersuites with keysize < 128 Bits", "vulnerable": "no", "state": "False"})
        
        # Check if TripleDES Ciphersuites are enabled
        elif vuln["defect"] == "cipherlist_3DES":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TripleDES Ciphersuites Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TripleDES Ciphersuites", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["Suites de chiffrement utilisant TripleDES", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TripleDES Ciphersuites", "vulnerable": "no", "state": "Disabled"})
        
        # Check if AEAD is not available
        elif " " in vuln["defect"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "AEAD Encryption not available", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "AEAD Encryption", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Suites de chiffrement AEAD non disponibles", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "AEAD Encryption", "vulnerable": "no", "state": "Enabled"})
        
        # Check if Cipher ordering is disabled
        elif vuln["defect"] == "cipher_order":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Cipher Ordering Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Cipher Ordering", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Pas d'ordre dans les suites de chiffrement", "20"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Cipher Ordering", "vulnerable": "no", "state": "Enabled"})
        
        # Check if Secure renegotiation is disabled
        elif vuln["defect"] == "secure_renego":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Secure Renegociation Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Secure Renegotiation", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Renégociation sécurisée vulnérable (CVE-2009-3555)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Secure Renegotiation", "vulnerable": "no", "state": "Enabled"})
        
        # Check if Client-Initiated Renegotiation is enabled
        elif vuln["defect"] == "secure_client_renego":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Client-Initiated Renegociation Enabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Client-Initiated Renegotiation", "vulnerable": "yes", "state": "Enabled"})
                newTLSDefect.add(json.dumps(["Renégociation sécurisée initiée par le client", "23"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Client-Initiated Renegotiation", "vulnerable": "no", "state": "Disabled"})
        
        # Check if TLS_FALLBACK_SCSV is disabled
        elif vuln["defect"] == "fallback_SCSV":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "TLS_FALLBACK_SCSV Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLS_FALLBACK_SCSV", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Absence de support de TLS_FALLBACK_SCSV", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "TLS_FALLBACK_SCSV", "vulnerable": "no", "state": "Enabled"})
        
        # CERTIFICATES
        
        # Get certificate common name
        elif vuln["defect"] in ["cert_commonName", "cert_commonName <hostCert#1>"]:
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Common Name", "vulnerable": "no", "state": vuln["details"]})
        
        # Get certificate alternative names
        elif vuln["defect"] in ["cert_subjectAltName", "cert_subjecAltName <hostCert#1>"]:
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Alternative Name(s)", "vulnerable": "no", "state": vuln["details"]})
        
        # Get certificate Validity left
        elif vuln["defect"] in ["cert_expirationStatus", "cert_expirationStatus <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                if vuln["details"] == "expired":
                    newTLSDefect.add(json.dumps(["Certificat expiré", "6"]))
                    vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate " + vuln["details"], "criticity": vuln["criticity"]})
                    completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Validity Left", "vulnerable": "yes", "state": vuln["details"]})
                else:
                    vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate " + vuln["details"], "criticity": vuln["criticity"]})
                    completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Validity Left", "vulnerable": "yes", "state": vuln["details"]})
            else:
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Validity Left", "vulnerable": "no", "state": vuln["details"]})
        
        # Get notAfter date
        elif vuln["defect"] in ["cert_notAfter", "cert_notAfter <hostCert#1>"]:
            notAfter = datetime.strptime(vuln["details"], "%Y-%m-%d %H:%M") if vuln["details"] else None
        
        # Get notBefore date
        elif vuln["defect"] in ["cert_notBefore", "cert_notBefore <hostCert#1>"]:
            notBefore = datetime.strptime(vuln["details"], "%Y-%m-%d %H:%M") if vuln["details"] else None
        
        # Calculate certificate lifespan
        elif notAfter is not None and notBefore is not None and lifespan == -1:
            lifespan = abs((notAfter - notBefore).days)
            if lifespan > 398:
                newTLSDefect.add(json.dumps(["Durée de validité du certificat trop longue", "6"]))
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Total Lifespan", "vulnerable": "yes", "state": str(lifespan) + " days"})
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Total Lifespan", "vulnerable": "no", "state": str(lifespan) + " days"})
            lifespan = -1  # Reset lifespan for next certificate
            notAfter = None  # Reset notAfter for next certificate
            notBefore = None  # Reset notBefore for next certificate
        
        # Get certificate signature algorithm
        elif vuln["defect"] in ["cert_signatureAlgorithm", "cert_signatureAlgorithm <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate " + vuln["details"], "criticity": vuln["criticity"]})
                if "SHA1" in vuln["details"]:
                    newTLSDefect.add(json.dumps(["Certificat signé en SHA-1", "6"]))
                    completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Signature Algorithm", "vulnerable": "yes", "state": vuln["details"]})
                    continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Signature Algorithm", "vulnerable": "no", "state": vuln["details"]})
        
        # Get Certificate key size
        elif vuln["defect"] in ["cert_keySize", "cert_keySize <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate " + vuln["details"], "criticity": vuln["criticity"]})
                keySizeInfo = vuln["details"].split(" ")
                if (keySizeInfo[0] == "RSA" and int(keySizeInfo[1]) < 2048) or (keySizeInfo[0] == "EC" and int(keySizeInfo[1]) < 256):
                    newTLSDefect.add(json.dumps(["Taille de clé trop faible", "6"]))
                    completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Key Size", "vulnerable": "yes", "state": vuln["details"]})
                    continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Key Size", "vulnerable": "no", "state": vuln["details"]})
        
        # Get certificate issuer
        elif vuln["defect"] in ["cert_caIssuers", "cert_caIssuers <hostCert#1>"]:
            # Check if the issuer is DigiCert for SNCF only
            if is_sncf and "digicert" not in vuln["details"].lower():
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate issuer is not DigiCert (SNCF only)", "criticity": "INFO"})
                newTLSDefect.add(json.dumps(["Certificat émis par une autorité de certification non approuvée [SNCF uniquement, ne pas mentionner sinon]", "6"]))
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Issuer", "vulnerable": "yes", "state": vuln["details"]})
                continue
            # Check if the issuer is emmited by a private CA or self-signed (not for SNCF)
            if ("selfsigned" in vuln["details"].lower() or "private" in vuln["details"].lower()) and not is_sncf:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate issuer is self-signed or private", "criticity": "INFO"})
                newTLSDefect.add(json.dumps(["Certificat émis par une autorité de certification privée ou auto-signé", "6"]))
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Issuer", "vulnerable": "yes", "state": vuln["details"]})
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Issuer", "vulnerable": "no", "state": vuln["details"]})
        
        # Revocation list enabled ?
        elif vuln["defect"] in ["cert_crlDistributionPoints", "cert_crlDistributionPoints <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate Revocation List Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Revocation List", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Absence de liste de révocation", "6"]))
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Revocation List", "vulnerable": "no", "state": "Enabled"})
        
        # CA authorisation
        elif vuln["defect"] in ["DNS_CAArecord", "DNS_CAArecord <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "No CA Authorisation", "criticity": vuln["criticity"]})
                newTLSDefect.add(json.dumps(["Pas d'autorisation d'Autorités de Certification (CAA)", "141"]))
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CA Authorisation", "vulnerable": "yes", "state": "No CAA record found"})
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CA Authorisation", "vulnerable": "no", "state": vuln["details"]})
        
        # Transparency
        elif vuln["defect"] in ["certificate_transparency", "certificate_transparency <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Certificate Transparancy Disabled", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Transparency", "vulnerable": "yes", "state": "Disabled"})
                newTLSDefect.add(json.dumps(["Pas de Certificate Transparency", "6"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Transparency", "vulnerable": "no", "state": "Enabled"})
        
        # Chain of trust (SNCF is not concerned by this defect)
        elif vuln["defect"] in ["cert_chain_of_trust", "cert_chain_of_trust <hostCert#1>"]:
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Chain of Trust Incomplete", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Chain of Trust", "vulnerable": "yes", "state": "Incomplete"})
                if not is_sncf:
                    newTLSDefect.add(json.dumps(["Chaîne de confiance incomplète", "6"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Chain of Trust", "vulnerable": "no", "state": "Verified"})
        
        # VULNERABILITIES
        
        # Heartbleed
        elif vuln["defect"] == "heartbleed":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Heartbleed", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Heartbleed", "vulnerable": "yes", "state": "Vulnerable"})
                # newTLSDefect.add(json.dumps(["Vulnérabilité Heartbleed", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Heartbleed", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # Ticketbleed
        elif vuln["defect"] == "ticketbleed":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "Ticketbleed", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Ticketbleed", "vulnerable": "yes", "state": "Vulnerable"})
                # newTLSDefect.add(json.dumps(["Vulnérabilité Ticketbleed", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "Ticketbleed", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # CCS
        elif vuln["defect"] == "CCS":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "CCS Injection", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CCS Injection", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille CCS (CVE-2014-0224)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CCS Injection", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # ROBOT
        elif vuln["defect"] == "ROBOT":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "ROBOT", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "ROBOT", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à l'attaque ROBOT (CVE-2017-13099)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "ROBOT", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # POODLE
        elif vuln["defect"] == "POODLE_SSL":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "POODLE", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "POODLE", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille POODLE (CVE-2014-3566)", "7"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "POODLE", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # CRIME
        elif vuln["defect"] == "CRIME_TLS":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "CRIME", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CRIME", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille CRIME (CVE-2012-4929)", "26"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "CRIME", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # BREACH
        elif vuln["defect"] == "BREACH":
            if vuln["criticity"] not in ["OK", "INFO"]:
                if "test failed" not in vuln["details"].lower():
                    vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "BREACH", "criticity": vuln["criticity"]})
                    completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BREACH", "vulnerable": "yes", "state": "Vulnerable"})
                    newTLSDefect.add(json.dumps(["Vulnérabilité à BREACH", "25"]))
                    continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BREACH", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # SWEET32
        elif vuln["defect"] == "SWEET32":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "SWEET32", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SWEET32", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille SWEET32 (CVE-2016-2183)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "SWEET32", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # FREAK
        elif vuln["defect"] == "FREAK":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "FREAK", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "FREAK", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille FREAK (CVE-2015-0204)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "FREAK", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # DROWN
        elif vuln["defect"] == "DROWN":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "DROWN", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "DROWN", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille DROWN (CVE-2016-0800, CVE-2016-0703)", "7"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "DROWN", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # LOGJAM
        elif vuln["defect"] == "LOGJAM":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "LOGJAM", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "LOGJAM", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille LOGJAM (CVE-2015-4000)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "LOGJAM", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # BAR-MITZVAH
        elif vuln["defect"] == "BAR_MITZVAH":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "BAR-MITZVAH", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BAR-MITZVAH", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille BAR-MITZVAH (CVE-2015-2808)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BAR-MITZVAH", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # LUCKY13
        elif vuln["defect"] == "LUCKY13":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "LUCKY13", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "LUCKY13", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille LUCKY13 (CVE-2013-0169)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "LUCKY13", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # RC4
        elif vuln["defect"] == "RC4":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "RC4", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "RC4", "vulnerable": "yes", "state": "Vulnerable"})
                # newTLSDefect.add(json.dumps(["Vulnérabilité à la faille RC4", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "RC4", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # BEAST
        elif vuln["defect"] == "BEAST":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "BEAST", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BEAST", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille BEAST (CVE-2011-3389)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BEAST", "vulnerable": "no", "state": "Not Vulnerable"})
        
        # BEAST CBC TLS1.0
        elif vuln["defect"] == "BEAST_CBC_TLS1":
            if vuln["criticity"] not in ["OK", "INFO"]:
                vulnDataList.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "defect": "BEAST CBC TLS1.0", "criticity": vuln["criticity"]})
                completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BEAST CBC TLS1.0", "vulnerable": "yes", "state": "Vulnerable"})
                newTLSDefect.add(json.dumps(["Vulnérabilité à la faille BEAST (CVE-2011-3389)", "21"]))
                continue
            completeData.append({"ip": vuln["ip"], "port": vuln["port"], "domain": vuln["domain"], "about": "BEAST CBC TLS1.0", "vulnerable": "no", "state": "Not Vulnerable"})
    
        # Convert set to list and parse JSON strings back to arrays
    newTLSDefect = list(newTLSDefect)
    newTLSDefect = [json.loads(item) for item in newTLSDefect]
    
    return {"vulnTable": vulnDataList, "allDataForExcel": completeData, "newTLSDefectForReport": newTLSDefect}

def updateTLSDefect(defectTLS, new_defect, defect_to_keep, fixes_to_keep):
    defectTLS_description = defectTLS[0]["description"]
    if new_defect.get("description", "") == "":
        new_defect["description"] = defectTLS_description.split("\r\n\r\n")[0] + "\r\n\r\n"
    lignes = defectTLS_description.split("\r\n\r\n")
    title = ""
    for ligne in lignes:
        if ligne.startswith("#"):
            title = ligne.strip("#").strip()
            if title == defect_to_keep:
                new_defect["description"] += ligne + "\r\n\r\n"
                continue
        if title == defect_to_keep:
            new_defect["description"] += ligne + "\r\n\r\n"
            continue
    fixes_to_keep = fixes_to_keep.split(",")
    defectTLS_fixes = defectTLS[0]["fixes"]
    for fix in defectTLS_fixes:
        if str(fix["id"]) in fixes_to_keep and not fix in new_defect["fixes"]:
            new_defect["fixes"].append(fix)
    return new_defect
    
def generateDefect(apiclient, newTLSDefect):
    # Function to generate a defect in the report
    # This function should be implemented to handle the defect creation logic
    print("Generating defect with data:", newTLSDefect)
    defectTLS = apiclient.searchDefect("Défauts d'implémentation du TLS")[0]
    new_TLS_defect = {}
    new_TLS_defect["description"] = ""
    new_TLS_defect["fixes"] = []
    for d in newTLSDefect:
        new_TLS_defect = updateTLSDefect(defectTLS, new_TLS_defect, d[0], d[1])
    defectTLS[0]["description"] = new_TLS_defect["description"]
    defectTLS[0]["fixes"] = new_TLS_defect["fixes"]
    new_defect = defectTLS[0]
    all_defects = apiclient.getDefectTable()
    for d in all_defects:
        if d["title"] == new_defect["title"]:
            print("Defect already exists, updating it.")
            if "_id" in new_defect:
                del new_defect["_id"]
            apiclient.updateInDb(apiclient.getCurrentPentest(), "defects", {"_id":ObjectId(d["_id"])}, {"$set": new_defect})
            return {"status": "success", "data": new_defect}
    apiclient.insertInDb(apiclient.getCurrentPentest(), "defects", new_defect)
    return {"status": "success", "data": new_defect}

def generateExcel(excelData):
    Ips = []
    columnHeaders = ["IP", "Port", "Domain"]
    table = [["IP", "Port", "Domain"]]
    colorTable = [["no", "no", "no"]]
    for data in excelData:
        try:
            row_index = Ips.index(data.get("ip","")+"/"+data.get("port",""))
        except ValueError:
            row_index = len(Ips)
            Ips.append(data.get("ip","") + "/"+data.get("port",""))
            new_row = [""]*len(columnHeaders)
            new_row[0] = data.get("ip","")
            new_row[1] = data.get("port","")
            new_row[2] = data.get("domain","") or "N/A"
            table.append(new_row)
            colorTable.append(["no", "no", "no"])
        if data.get("about","") not in columnHeaders:
            columnHeaders.append(data.get("about",""))
            table[0].append(data.get("about",""))
        try:
            column_index = columnHeaders.index(data.get("about",""))
        except ValueError:
            column_index = -1
        if column_index == len(table[row_index + 1]):
            table[row_index + 1].append("")
        
        table[row_index + 1][column_index] = data.get("state","")
        if column_index == len(colorTable[row_index + 1]):
            colorTable[row_index + 1].append("")
        colorTable[row_index + 1][column_index] = data.get("vulnerable","")

    
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSSL Report"
    
    # Add data to worksheet
    for row_idx, row_data in enumerate(table):
        for col_idx, cell_data in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_data)
    
    # Color all the headers
    header_fill = PatternFill(start_color="95cd41", end_color="95cd41", fill_type="solid")
    for i in range(len(columnHeaders)):
        cell = ws.cell(row=1, column=i + 1)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
    
    # Specify column width and cell formatting
    for i, column_header in enumerate(columnHeaders):
        max_length = len(column_header)
        for j, row_data in enumerate(table[1:], 1):
            try:
                if len(str(row_data[i])) > max_length:
                    max_length = len(str(row_data[i]))
                
                cell = ws.cell(row=j + 1, column=i + 1)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # center text in cell
                if colorTable[j][i] == "yes":
                    # Set font color in red and bold
                    cell.font = Font(color="FF0000", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            except (IndexError, TypeError) as error:
                print("Index/Type error while processing table cell:", error)
        
        size = max_length * 1.2
        # Set column width
        column_letter = get_column_letter(i + 1)
        ws.column_dimensions[column_letter].width = size
    
    # Specify row height
    ws.row_dimensions[1].height = 15  # set the first row height (header row)
    for i in range(2, len(table) + 1):
        ws.row_dimensions[i].height = 50
    
    # Save the workbook
    wb.save("TestSSL_Report.xlsx")
    return "TestSSL_Report.xlsx"

    # Code to generate Excel file goes here

def main(apiclient, *args, **kwargs):
    # Example usage of the parseTestSSL function
    results = testsslresults(apiclient)
    if results:
        list_of_defects = results.get("vulnTable", [])
        print("Défauts trouvés :")
        print("\n".join(set([d["defect"] for d in list_of_defects])))
        response = input("Créer le défaut dans le rapport ? (y/N): ")
        if response.lower() == "y":
            results = generateDefect(apiclient, results.get("newTLSDefectForReport", []))
            if results["status"] == "success":
                print("Défaut créé avec succès ")
            else:
                print("Erreur dans la création du défaut.")
        response = input("Générer un fichier Excel d'annexe ? (y/N): ")
        if response.lower() == "y":
            # Code to generate Excel report goes here
            file_path = generateExcel(results.get("allDataForExcel", []))
            res = apiclient.uploadFile(file_path)
            if res:
                print("Fichier Excel TestSSL_Report.xlsx généré et uploadé avec succès.")
            else:
                print("Erreur lors de l'upload du fichier Excel.", res)
        return True, "Fichier Excel TestSSL_Report.xlsx généré et uploadé avec succès."    
    return False, "No results found or an error occurred."
# apiclient = APIClient.getInstance()  # Assuming APIClient is defined elsewhere
# apiclient.tryConnection()
# res = apiclient.tryAuth()
# if not res:
#     consoleConnect()
# if apiclient.isConnected() is False or apiclient.getCurrentPentest() == "":
#     print("You must be connected to a pentest to run this script.")

# main(apiclient)
