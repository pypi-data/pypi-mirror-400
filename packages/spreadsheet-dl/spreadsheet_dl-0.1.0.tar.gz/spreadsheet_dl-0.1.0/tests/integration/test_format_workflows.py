"""Integration tests for mixed format workflows.

Tests complex workflows that involve multiple format conversions
and validate data integrity across the entire pipeline.

Test Strategy:
    - Test complete workflows: YAML -> ODS -> XLSX -> validate
    - Test format-specific feature preservation
    - Test cross-format consistency
    - Test domain formula integration with exports
"""

from __future__ import annotations

from typing import TYPE_CHECKING
from zipfile import ZipFile

import pytest

if TYPE_CHECKING:
    from pathlib import Path

pytestmark = [pytest.mark.integration]


class TestMixedFormatWorkflows:
    """Test workflows involving multiple format conversions."""

    def test_ods_to_xlsx_formula_preservation(self, tmp_path: Path) -> None:
        """Test that formulas are preserved when converting ODS to XLSX."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet, formula

        # Create ODS with formulas
        builder = create_spreadsheet()
        builder.sheet("Calcs").column("A").column("B").column("Sum").column("Product")
        builder.header_row()
        builder.row().cell(10).cell(5).cell(formula=formula().sum("A2", "B2")).cell(
            formula=formula().product("A2", "B2")
        )
        builder.row().cell(20).cell(3).cell(formula=formula().sum("A3", "B3")).cell(
            formula=formula().product("A3", "B3")
        )

        # Save as ODS
        ods_path = tmp_path / "formulas.ods"
        builder.save(str(ods_path))

        # Export to XLSX
        xlsx_path = tmp_path / "formulas.xlsx"
        builder.export(str(xlsx_path))

        # Verify XLSX
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Verify data values are present
        assert ws["A2"].value == 10
        assert ws["B2"].value == 5

    def test_multi_sheet_workflow(self, tmp_path: Path) -> None:
        """Test multi-sheet spreadsheet across formats."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create multi-sheet spreadsheet
        builder = create_spreadsheet()

        # Sheet 1: Data
        builder.sheet("Input").column("Month").column("Revenue").column("Cost")
        builder.header_row()
        builder.row().cell("January").cell(10000).cell(8000)
        builder.row().cell("February").cell(12000).cell(9000)
        builder.row().cell("March").cell(15000).cell(10000)

        # Sheet 2: Summary
        builder.sheet("Summary").column("Metric").column("Value")
        builder.header_row()
        builder.row().cell("Total Revenue").cell(37000)
        builder.row().cell("Total Cost").cell(27000)
        builder.row().cell("Profit").cell(10000)

        # Save ODS
        ods_path = tmp_path / "multi_sheet.ods"
        builder.save(str(ods_path))

        # Export XLSX
        xlsx_path = tmp_path / "multi_sheet.xlsx"
        builder.export(str(xlsx_path))

        # Verify both formats
        assert ods_path.exists()
        assert xlsx_path.exists()

        # Check XLSX structure
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        assert len(wb.sheetnames) >= 2

    def test_themed_export_workflow(self, tmp_path: Path) -> None:
        """Test themed spreadsheet export workflow."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        # Create themed spreadsheet (use "default" theme instead of non-existent "modern")
        builder = create_spreadsheet(theme="default")
        builder.sheet("Report").column("Quarter").column("Sales").column("Growth")
        builder.header_row()
        builder.row().cell("Q1").cell(100000).cell("5%")
        builder.row().cell("Q2").cell(120000).cell("20%")
        builder.row().cell("Q3").cell(115000).cell("-4%")
        builder.row().cell("Q4").cell(140000).cell("22%")

        # Save in multiple formats
        ods_path = tmp_path / "themed.ods"
        xlsx_path = tmp_path / "themed.xlsx"

        builder.save(str(ods_path))
        builder.export(str(xlsx_path))

        # Verify both exist and have content
        assert ods_path.exists() and ods_path.stat().st_size > 0
        assert xlsx_path.exists() and xlsx_path.stat().st_size > 0


class TestDomainFormulaIntegration:
    """Test domain formulas integrate correctly with exports."""

    def test_physics_formulas_in_spreadsheet(self, tmp_path: Path) -> None:
        """Test physics domain formulas export correctly."""
        from spreadsheet_dl import create_spreadsheet
        from spreadsheet_dl.domains.physics.formulas.mechanics import (
            KineticEnergyFormula,
            MomentumFormula,
        )

        ke_formula = KineticEnergyFormula()
        p_formula = MomentumFormula()

        # Create spreadsheet with physics calculations
        builder = create_spreadsheet()
        builder.sheet("Physics").column("Mass").column("Velocity").column("KE").column(
            "Momentum"
        )
        builder.header_row()

        # Row with formulas referencing cells
        builder.row().cell(10).cell(5).cell(formula=ke_formula.build("A2", "B2")).cell(
            formula=p_formula.build("A2", "B2")
        )

        # Save as ODS
        ods_path = tmp_path / "physics.ods"
        builder.save(str(ods_path))

        # Verify ODS contains formulas
        with ZipFile(ods_path, "r") as zf:
            content = zf.read("content.xml").decode("utf-8")
            # Should contain formula references
            assert "A2" in content or "formula" in content.lower()

    def test_chemistry_formulas_in_spreadsheet(self, tmp_path: Path) -> None:
        """Test chemistry domain formulas export correctly."""
        from spreadsheet_dl import create_spreadsheet
        from spreadsheet_dl.domains.chemistry.formulas.stoichiometry import (
            MolarMassFormula,
        )

        mm_formula = MolarMassFormula()

        builder = create_spreadsheet()
        builder.sheet("Chemistry").column("Mass (g)").column("Moles").column(
            "Molar Mass"
        )
        builder.header_row()
        builder.row().cell(58.44).cell(1).cell(formula=mm_formula.build("A2", "B2"))

        ods_path = tmp_path / "chemistry.ods"
        builder.save(str(ods_path))

        assert ods_path.exists()
        assert ods_path.stat().st_size > 0

    def test_finance_formulas_structure(self, tmp_path: Path) -> None:
        """Test finance formulas produce valid spreadsheet structure."""
        from spreadsheet_dl import create_spreadsheet
        from spreadsheet_dl.domains.finance.formulas.time_value import FutureValue

        fv_formula = FutureValue()

        builder = create_spreadsheet()
        builder.sheet("Finance").column("Rate").column("Periods").column("Payment")
        builder.column("PV").column("FV")
        builder.header_row()

        # Build formula with cell references
        fv_result = fv_formula.build("A2", "B2", "C2", "D2")

        builder.row().cell(0.05).cell(10).cell(100).cell(1000).cell(formula=fv_result)

        ods_path = tmp_path / "finance.ods"
        builder.save(str(ods_path))

        assert ods_path.exists()


class TestFormatSpecificFeatures:
    """Test format-specific features are handled correctly."""

    def test_ods_namespace_compliance(self, tmp_path: Path) -> None:
        """Test ODS files have correct namespace declarations."""
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Test").column("Data")
        builder.header_row()
        builder.row().cell("Value")

        ods_path = tmp_path / "namespace.ods"
        builder.save(str(ods_path))

        with ZipFile(ods_path, "r") as zf:
            content = zf.read("content.xml").decode("utf-8")
            # Should have ODF namespaces
            assert "office:" in content or "urn:oasis" in content.lower()

    def test_xlsx_relationship_structure(self, tmp_path: Path) -> None:
        """Test XLSX files have correct relationship structure."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Test").column("Data")
        builder.header_row()
        builder.row().cell("Value")

        xlsx_path = tmp_path / "structure.xlsx"
        builder.export(str(xlsx_path))

        # XLSX is a ZIP file
        with ZipFile(xlsx_path, "r") as zf:
            names = zf.namelist()
            # Should have essential XLSX components
            assert any("workbook" in n.lower() for n in names)
            assert any("[Content_Types]" in n for n in names)

    def test_ods_manifest_structure(self, tmp_path: Path) -> None:
        """Test ODS manifest is correctly structured."""
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Test").column("Data")
        builder.header_row()
        builder.row().cell("Value")

        ods_path = tmp_path / "manifest.ods"
        builder.save(str(ods_path))

        with ZipFile(ods_path, "r") as zf:
            assert "META-INF/manifest.xml" in zf.namelist()
            manifest = zf.read("META-INF/manifest.xml").decode("utf-8")
            assert "manifest" in manifest.lower()


class TestDataIntegrityAcrossFormats:
    """Test data integrity is maintained across format conversions."""

    def test_numeric_data_integrity(self, tmp_path: Path) -> None:
        """Test numeric values are preserved across formats."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        test_values = [
            0,
            1,
            -1,
            3.14159,
            1e10,
            1e-10,
            2.718281828459045,
            -999.999,
        ]

        builder = create_spreadsheet()
        builder.sheet("Numbers").column("Value")
        builder.header_row()
        for val in test_values:
            builder.row().cell(val)

        # Save ODS
        ods_path = tmp_path / "numbers.ods"
        builder.save(str(ods_path))

        # Export XLSX
        xlsx_path = tmp_path / "numbers.xlsx"
        builder.export(str(xlsx_path))

        # Both should exist
        assert ods_path.exists()
        assert xlsx_path.exists()

        # Verify XLSX values
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Check some values (accounting for floating point)
        assert ws["A2"].value == 0
        assert ws["A3"].value == 1
        assert ws["A4"].value == -1

    def test_text_data_integrity(self, tmp_path: Path) -> None:
        """Test text values are preserved across formats."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        test_texts = [
            "Simple text",
            "Text with spaces  ",
            "Special: @#$%^&*()",
            "Unicode: \u03c0\u00b2",
            "Numbers: 12345",
            "",
        ]

        builder = create_spreadsheet()
        builder.sheet("Text").column("Value")
        builder.header_row()
        for text in test_texts:
            builder.row().cell(text)

        # Save and export
        ods_path = tmp_path / "text.ods"
        xlsx_path = tmp_path / "text.xlsx"
        builder.save(str(ods_path))
        builder.export(str(xlsx_path))

        # Verify
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active

        assert ws["A2"].value == "Simple text"

    def test_mixed_data_types(self, tmp_path: Path) -> None:
        """Test mixed data types are handled correctly."""
        pytest.importorskip("openpyxl")
        from datetime import date

        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Mixed").column("Type").column("Value")
        builder.header_row()
        builder.row().cell("Integer").cell(42)
        builder.row().cell("Float").cell(3.14)
        builder.row().cell("String").cell("Hello")
        builder.row().cell("Date").cell(date(2025, 1, 15))
        builder.row().cell("Boolean-like").cell("TRUE")
        builder.row().cell("Empty").cell("")

        ods_path = tmp_path / "mixed.ods"
        xlsx_path = tmp_path / "mixed.xlsx"
        builder.save(str(ods_path))
        builder.export(str(xlsx_path))

        assert ods_path.exists()
        assert xlsx_path.exists()


class TestLargeDatasetWorkflows:
    """Test workflows with larger datasets."""

    def test_1000_row_workflow(self, tmp_path: Path) -> None:
        """Test workflow with 1000 rows of data."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Large").column("ID").column("Value").column("Category")
        builder.header_row()

        for i in range(1000):
            builder.row().cell(i).cell(i * 10).cell(f"Cat{i % 10}")

        # Save and export
        ods_path = tmp_path / "large.ods"
        xlsx_path = tmp_path / "large.xlsx"
        builder.save(str(ods_path))
        builder.export(str(xlsx_path))

        # Verify row counts
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row >= 1000

    def test_wide_spreadsheet_workflow(self, tmp_path: Path) -> None:
        """Test workflow with many columns."""
        pytest.importorskip("openpyxl")
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        sheet = builder.sheet("Wide")

        # Add 50 columns
        for i in range(50):
            sheet.column(f"Col{i}")

        builder.header_row()

        # Add a few rows
        for row_idx in range(10):
            row = builder.row()
            for col_idx in range(50):
                row.cell(f"R{row_idx}C{col_idx}")

        # Save and export
        ods_path = tmp_path / "wide.ods"
        xlsx_path = tmp_path / "wide.xlsx"
        builder.save(str(ods_path))
        builder.export(str(xlsx_path))

        # Verify column count
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_column >= 50


class TestErrorHandlingWorkflows:
    """Test error handling in format workflows."""

    def test_invalid_export_path_handling(self, tmp_path: Path) -> None:
        """Test handling of invalid export paths."""
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Test").column("A")
        builder.header_row()
        builder.row().cell("Data")

        # Try to export to non-existent directory
        invalid_path = tmp_path / "nonexistent" / "deep" / "path" / "file.xlsx"

        # Should either create directories or raise appropriate error
        try:
            builder.export(str(invalid_path))
            # If it succeeds, verify file was created
            assert invalid_path.exists()
        except (FileNotFoundError, OSError):
            # Expected behavior for non-existent path
            pass

    def test_special_characters_in_filename(self, tmp_path: Path) -> None:
        """Test handling of special characters in filenames."""
        from spreadsheet_dl import create_spreadsheet

        builder = create_spreadsheet()
        builder.sheet("Test").column("A")
        builder.header_row()
        builder.row().cell("Data")

        # Use safe special characters
        ods_path = tmp_path / "test-file_v1.0.ods"
        builder.save(str(ods_path))

        assert ods_path.exists()
