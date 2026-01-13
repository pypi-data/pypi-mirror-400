"""Chart creation operations for ExceLLM MCP server.

Supports creating charts in Excel workbooks with dual-engine support:
- COM engine: Live Excel automation on Windows
- openpyxl engine: File-based cross-platform
"""

import logging
import os
from enum import Enum
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


class ChartType(str, Enum):
    """Supported chart types."""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"


# COM chart type constants (xlChartType enumeration)
XL_CHART_TYPES = {
    "line": 4,       # xlLine
    "bar": 57,       # xlColumnClustered
    "pie": 5,        # xlPie
    "scatter": -4169, # xlXYScatter
    "area": 1,       # xlArea
}


def create_chart_sync(
    workbook_name: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis_title: str = "",
    y_axis_title: str = "",
    width: float = 400,
    height: float = 300,
    style: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Create a chart in an Excel worksheet.
    
    Supports dual-engine operation:
    - COM engine for live Excel (Windows)
    - openpyxl engine for file-based operations
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet
        data_range: Source data range (e.g., "A1:D10")
        chart_type: Type of chart (line, bar, pie, scatter, area)
        target_cell: Cell where chart will be placed (top-left corner)
        title: Chart title
        x_axis_title: X-axis label
        y_axis_title: Y-axis label
        width: Chart width in points (default 400)
        height: Chart height in points (default 300)
        style: Optional style configuration dict
        
    Returns:
        Dictionary with operation result
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    from ..validators import validate_cell_reference, parse_range
    
    # Validate chart type
    chart_type_lower = chart_type.lower()
    if chart_type_lower not in XL_CHART_TYPES:
        return {
            "success": False,
            "error": f"Unsupported chart type: {chart_type}. "
                    f"Supported types: {', '.join(XL_CHART_TYPES.keys())}",
        }
    
    # Validate target cell
    if not validate_cell_reference(target_cell):
        return {
            "success": False,
            "error": f"Invalid target cell reference: {target_cell}",
        }
    
    # Engine Selection Logic
    use_com = True
    try:
        # Try to get COM app and check if workbook is open
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
    except Exception:
        # COM failed or workbook not open
        use_com = False
        
    if not use_com:
        # Fallback to openpyxl if it's a file
        if os.path.exists(workbook_name):
            return create_chart_openpyxl(
                workbook_name, sheet_name, data_range, chart_type,
                target_cell, title, x_axis_title, y_axis_title,
                width / 28.35, height / 28.35, style  # Convert points to cm (approx)
            )
        else:
            return {
                "success": False,
                "error": f"Workbook not found (checked open workbooks and file path): {workbook_name}"
            }

    try:
        worksheet = get_worksheet(workbook, sheet_name)
        
        # Get the data range
        try:
            source_range = worksheet.Range(data_range)
        except Exception as e:
            return {
                "success": False,
                "error": f"Invalid data range: {data_range}. {str(e)}",
            }
        
        # Get target cell position
        target = worksheet.Range(target_cell)
        left = target.Left
        top = target.Top
        
        # Create the chart
        chart_objects = worksheet.ChartObjects()
        chart_obj = chart_objects.Add(
            Left=left,
            Top=top,
            Width=width,
            Height=height,
        )
        chart = chart_obj.Chart
        
        # Set chart type
        chart.ChartType = XL_CHART_TYPES[chart_type_lower]
        
        # Set data source
        chart.SetSourceData(Source=source_range)
        
        # Set chart title
        if title:
            chart.HasTitle = True
            chart.ChartTitle.Text = title
        
        # Set axis titles (not applicable for pie charts)
        if chart_type_lower != "pie":
            if x_axis_title:
                try:
                    x_axis = chart.Axes(1)  # xlCategory = 1
                    x_axis.HasTitle = True
                    x_axis.AxisTitle.Text = x_axis_title
                except Exception:
                    pass  # Some chart types may not support axis titles
                    
            if y_axis_title:
                try:
                    y_axis = chart.Axes(2)  # xlValue = 2
                    y_axis.HasTitle = True
                    y_axis.AxisTitle.Text = y_axis_title
                except Exception:
                    pass
        
        # Apply style options if provided
        if style:
            _apply_chart_style(chart, style)
        
        logger.info(f"Created {chart_type} chart at {target_cell} in {sheet_name}")
        
        return {
            "success": True,
            "message": f"{chart_type.capitalize()} chart created successfully",
            "engine": "COM",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "data_range": data_range,
                "title": title,
                "size": {"width": width, "height": height},
            },
        }
        
    except Exception as e:
        logger.error(f"Failed to create chart: {e}")
        return {
            "success": False,
            "error": f"Failed to create chart: {str(e)}",
        }


def _apply_chart_style(chart, style: Dict[str, Any]) -> None:
    """Apply style options to a chart.
    
    Args:
        chart: COM chart object
        style: Style configuration dictionary
    """
    try:
        # Legend visibility
        if "show_legend" in style:
            chart.HasLegend = style["show_legend"]
            
        # Legend position
        if chart.HasLegend and "legend_position" in style:
            positions = {
                "bottom": -4107,  # xlLegendPositionBottom
                "top": -4160,     # xlLegendPositionTop
                "left": -4131,    # xlLegendPositionLeft
                "right": -4152,   # xlLegendPositionRight
            }
            pos = style["legend_position"].lower()
            if pos in positions:
                chart.Legend.Position = positions[pos]
        
        # Data labels
        if style.get("show_data_labels", False):
            try:
                for series in chart.SeriesCollection():
                    series.HasDataLabels = True
            except Exception:
                pass
                
        # Grid lines
        if "show_gridlines" in style:
            try:
                if chart.ChartType not in [5]:  # Not for pie charts
                    y_axis = chart.Axes(2)
                    y_axis.HasMajorGridlines = style["show_gridlines"]
            except Exception:
                pass
                
    except Exception as e:
        logger.warning(f"Failed to apply some chart styles: {e}")


def list_charts_sync(
    workbook_name: str,
    sheet_name: str,
) -> Dict[str, Any]:
    """List all charts in a worksheet.
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet
        
    Returns:
        Dictionary with list of charts
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    try:
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
        worksheet = get_worksheet(workbook, sheet_name)
        
        charts = []
        chart_objects = worksheet.ChartObjects()
        
        for i in range(1, chart_objects.Count + 1):
            chart_obj = chart_objects.Item(i)
            chart = chart_obj.Chart
            
            chart_info = {
                "index": i,
                "name": chart_obj.Name,
                "left": chart_obj.Left,
                "top": chart_obj.Top,
                "width": chart_obj.Width,
                "height": chart_obj.Height,
                "has_title": chart.HasTitle,
            }
            
            if chart.HasTitle:
                chart_info["title"] = chart.ChartTitle.Text
                
            charts.append(chart_info)
        
        return {
            "success": True,
            "charts": charts,
            "count": len(charts),
        }
        
    except Exception as e:
        logger.error(f"Failed to list charts: {e}")
        return {
            "success": False,
            "error": f"Failed to list charts: {str(e)}",
        }


def delete_chart_sync(
    workbook_name: str,
    sheet_name: str,
    chart_name: Optional[str] = None,
    chart_index: Optional[int] = None,
) -> Dict[str, Any]:
    """Delete a chart from a worksheet.
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet
        chart_name: Name of chart to delete
        chart_index: Index of chart to delete (1-based)
        
    Returns:
        Dictionary with operation result
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    if chart_name is None and chart_index is None:
        return {
            "success": False,
            "error": "Either chart_name or chart_index must be provided",
        }
    
    try:
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
        worksheet = get_worksheet(workbook, sheet_name)
        
        chart_objects = worksheet.ChartObjects()
        
        if chart_name:
            # Find by name
            chart_obj = chart_objects.Item(chart_name)
        else:
            # Find by index
            chart_obj = chart_objects.Item(chart_index)
        
        deleted_name = chart_obj.Name
        chart_obj.Delete()
        
        logger.info(f"Deleted chart '{deleted_name}' from {sheet_name}")
        
        return {
            "success": True,
            "message": f"Chart '{deleted_name}' deleted successfully",
        }
        
    except Exception as e:
        logger.error(f"Failed to delete chart: {e}")
        return {
            "success": False,
            "error": f"Failed to delete chart: {str(e)}",
        }


# ============================================================================
# openpyxl Engine - File-based Chart Creation (Cross-Platform)
# ============================================================================


def create_chart_openpyxl(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis_title: str = "",
    y_axis_title: str = "",
    width: float = 15,  # Width in cm for openpyxl
    height: float = 10,  # Height in cm for openpyxl
    style: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Create a chart in an Excel file using openpyxl (file-based).
    
    Cross-platform support - no Excel installation required.
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of worksheet
        data_range: Source data range (e.g., "A1:D10")
        chart_type: Type of chart (line, bar, pie, scatter, area)
        target_cell: Cell where chart will be placed
        title: Chart title
        x_axis_title: X-axis label
        y_axis_title: Y-axis label
        width: Chart width in cm (default 15)
        height: Chart height in cm (default 10)
        style: Optional style configuration
        
    Returns:
        Dictionary with operation result
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import (
            LineChart, BarChart, PieChart, ScatterChart, AreaChart,
            Reference
        )
        from openpyxl.utils import range_boundaries
    except ImportError:
        return {
            "success": False,
            "error": "openpyxl is required for file-based chart creation. Install with: pip install openpyxl",
        }
    
    chart_type_lower = chart_type.lower()
    
    # Map chart types to openpyxl chart classes
    chart_classes = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }
    
    if chart_type_lower not in chart_classes:
        return {
            "success": False,
            "error": f"Unsupported chart type: {chart_type}. Supported: {', '.join(chart_classes.keys())}",
        }
    
    try:
        # Load workbook
        wb = load_workbook(filepath)
        
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found in workbook",
            }
        
        ws = wb[sheet_name]
        
        # Parse data range (remove sheet name if present)
        if "!" in data_range:
            data_range = data_range.split("!")[-1]
            
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        
        # Create appropriate chart type
        ChartClass = chart_classes[chart_type_lower]
        chart = ChartClass()
        
        # Set chart title
        if title:
            chart.title = title
        
        # Set axis titles (not for pie charts)
        if chart_type_lower != "pie":
            if x_axis_title:
                chart.x_axis.title = x_axis_title
            if y_axis_title:
                chart.y_axis.title = y_axis_title
        
        # Set chart size
        chart.width = width
        chart.height = height
        
        # Create data references
        # Assume first row is categories (x-axis) and remaining are data series
        categories = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        
        for col in range(min_col + 1, max_col + 1):
            data = Reference(ws, min_col=col, min_row=min_row, max_row=max_row)
            
            if chart_type_lower == "scatter":
                # Scatter charts need XYSeries
                from openpyxl.chart import Series
                series = Series(data, xvalues=categories, title_from_data=True)
                chart.series.append(series)
            else:
                chart.add_data(data, titles_from_data=True)
        
        # Add categories for non-scatter charts
        if chart_type_lower != "scatter":
            chart.set_categories(categories)
        
        # Apply style options
        if style:
            if style.get("show_legend") is False:
                chart.legend = None
            if style.get("show_data_labels", False):
                chart.dataLabels = True
        
        # Add chart to worksheet
        ws.add_chart(chart, target_cell)
        
        # Save workbook
        wb.save(filepath)
        
        logger.info(f"Created {chart_type} chart at {target_cell} in {filepath}")
        
        return {
            "success": True,
            "message": f"{chart_type.capitalize()} chart created successfully",
            "engine": "openpyxl",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "data_range": data_range,
                "title": title,
                "size": {"width": width, "height": height},
            },
        }
        
    except Exception as e:
        logger.error(f"Failed to create chart with openpyxl: {e}")
        return {
            "success": False,
            "error": f"Failed to create chart: {str(e)}",
        }

