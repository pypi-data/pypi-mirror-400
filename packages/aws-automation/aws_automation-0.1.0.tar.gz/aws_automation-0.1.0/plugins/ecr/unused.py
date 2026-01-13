"""
plugins/ecr/unused.py - ECR 미사용 이미지 분석

오래된/미사용 ECR 이미지 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_ecr_storage_price

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ecr:DescribeRepositories",
        "ecr:GetLifecyclePolicy",
        "ecr:DescribeImages",
    ],
}

# 미사용 기준: 90일 이상 pull 없음
UNUSED_DAYS_THRESHOLD = 90


class ECRRepoStatus(Enum):
    """ECR 리포지토리 상태"""

    NORMAL = "normal"
    EMPTY = "empty"
    OLD_IMAGES = "old_images"
    NO_LIFECYCLE = "no_lifecycle"


@dataclass
class ECRRepoInfo:
    """ECR 리포지토리 정보"""

    account_id: str
    account_name: str
    region: str
    name: str
    arn: str
    uri: str
    created_at: datetime | None
    image_count: int = 0
    total_size_bytes: int = 0
    has_lifecycle_policy: bool = False
    old_image_count: int = 0
    old_images_size_bytes: int = 0

    @property
    def total_size_gb(self) -> float:
        return self.total_size_bytes / (1024**3)

    @property
    def old_images_size_gb(self) -> float:
        return self.old_images_size_bytes / (1024**3)

    @property
    def monthly_cost(self) -> float:
        return self.total_size_gb * get_ecr_storage_price(self.region)

    @property
    def old_images_monthly_cost(self) -> float:
        return self.old_images_size_gb * get_ecr_storage_price(self.region)


@dataclass
class ECRRepoFinding:
    """ECR 리포지토리 분석 결과"""

    repo: ECRRepoInfo
    status: ECRRepoStatus
    recommendation: str


@dataclass
class ECRAnalysisResult:
    """ECR 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_repos: int = 0
    empty_repos: int = 0
    repos_with_old_images: int = 0
    no_lifecycle_repos: int = 0
    total_images: int = 0
    old_images: int = 0
    total_size_gb: float = 0.0
    old_images_size_gb: float = 0.0
    old_images_monthly_cost: float = 0.0
    findings: list[ECRRepoFinding] = field(default_factory=list)


def collect_ecr_repos(session, account_id: str, account_name: str, region: str) -> list[ECRRepoInfo]:
    """ECR 리포지토리 수집"""
    from botocore.exceptions import ClientError

    ecr = get_client(session, "ecr", region_name=region)
    repos = []

    paginator = ecr.get_paginator("describe_repositories")
    for page in paginator.paginate():
        for repo in page.get("repositories", []):
            repo_info = ECRRepoInfo(
                account_id=account_id,
                account_name=account_name,
                region=region,
                name=repo.get("repositoryName", ""),
                arn=repo.get("repositoryArn", ""),
                uri=repo.get("repositoryUri", ""),
                created_at=repo.get("createdAt"),
            )

            # 라이프사이클 정책 확인
            try:
                ecr.get_lifecycle_policy(repositoryName=repo_info.name)
                repo_info.has_lifecycle_policy = True
            except ecr.exceptions.LifecyclePolicyNotFoundException:
                repo_info.has_lifecycle_policy = False
            except ClientError:
                pass

            # 이미지 정보 수집
            try:
                img_paginator = ecr.get_paginator("describe_images")
                now = datetime.now(timezone.utc)
                threshold_date = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

                for img_page in img_paginator.paginate(repositoryName=repo_info.name):
                    for img in img_page.get("imageDetails", []):
                        repo_info.image_count += 1
                        size = img.get("imageSizeInBytes", 0)
                        repo_info.total_size_bytes += size

                        pushed_at = img.get("imagePushedAt")
                        last_pull = img.get("lastRecordedPullTime")

                        # 마지막 pull이 90일 이상 전이거나 push 후 한번도 pull 안 된 경우
                        is_old = False
                        if (
                            last_pull
                            and last_pull < threshold_date
                            or not last_pull
                            and pushed_at
                            and pushed_at < threshold_date
                        ):
                            is_old = True

                        if is_old:
                            repo_info.old_image_count += 1
                            repo_info.old_images_size_bytes += size

            except ClientError:
                pass

            repos.append(repo_info)

    return repos


def analyze_ecr_repos(repos: list[ECRRepoInfo], account_id: str, account_name: str, region: str) -> ECRAnalysisResult:
    """ECR 리포지토리 분석"""
    result = ECRAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_repos=len(repos),
    )

    for repo in repos:
        result.total_images += repo.image_count
        result.old_images += repo.old_image_count
        result.total_size_gb += repo.total_size_gb
        result.old_images_size_gb += repo.old_images_size_gb
        result.old_images_monthly_cost += repo.old_images_monthly_cost

        # 빈 리포지토리
        if repo.image_count == 0:
            result.empty_repos += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.EMPTY,
                    recommendation="빈 리포지토리 - 삭제 검토",
                )
            )
            continue

        # 오래된 이미지가 있는 경우
        if repo.old_image_count > 0:
            result.repos_with_old_images += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.OLD_IMAGES,
                    recommendation=f"{repo.old_image_count}개 오래된 이미지 ({repo.old_images_size_gb:.2f} GB)",
                )
            )

            # 라이프사이클 정책 없는 경우 추가 경고
            if not repo.has_lifecycle_policy:
                result.no_lifecycle_repos += 1
            continue

        # 라이프사이클 정책 없는 경우
        if not repo.has_lifecycle_policy:
            result.no_lifecycle_repos += 1
            result.findings.append(
                ECRRepoFinding(
                    repo=repo,
                    status=ECRRepoStatus.NO_LIFECYCLE,
                    recommendation="라이프사이클 정책 없음 - 설정 권장",
                )
            )
            continue

        result.findings.append(
            ECRRepoFinding(
                repo=repo,
                status=ECRRepoStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[ECRAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    ws = wb.create_sheet("Summary")
    ws["A1"] = "ECR 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "Repos", "빈 Repo", "오래된 이미지", "총 크기", "낭비 비용"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_repos)
        ws.cell(row=row, column=4, value=r.empty_repos)
        ws.cell(row=row, column=5, value=r.old_images)
        ws.cell(row=row, column=6, value=f"{r.total_size_gb:.2f} GB")
        ws.cell(row=row, column=7, value=f"${r.old_images_monthly_cost:,.2f}")
        if r.empty_repos > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.old_images > 0:
            ws.cell(row=row, column=5).fill = yellow_fill

    ws_detail = wb.create_sheet("Repositories")
    detail_headers = [
        "Account",
        "Region",
        "Repository",
        "상태",
        "이미지수",
        "오래된",
        "크기",
        "낭비",
        "Lifecycle",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != ECRRepoStatus.NORMAL:
                detail_row += 1
                repo = f.repo
                ws_detail.cell(row=detail_row, column=1, value=repo.account_name)
                ws_detail.cell(row=detail_row, column=2, value=repo.region)
                ws_detail.cell(row=detail_row, column=3, value=repo.name)
                ws_detail.cell(row=detail_row, column=4, value=f.status.value)
                ws_detail.cell(row=detail_row, column=5, value=repo.image_count)
                ws_detail.cell(row=detail_row, column=6, value=repo.old_image_count)
                ws_detail.cell(row=detail_row, column=7, value=f"{repo.total_size_gb:.2f} GB")
                ws_detail.cell(
                    row=detail_row,
                    column=8,
                    value=f"${repo.old_images_monthly_cost:.2f}",
                )
                ws_detail.cell(
                    row=detail_row,
                    column=9,
                    value="있음" if repo.has_lifecycle_policy else "없음",
                )
                ws_detail.cell(row=detail_row, column=10, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"ECR_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ECRAnalysisResult | None:
    """단일 계정/리전의 ECR 리포지토리 수집 및 분석 (병렬 실행용)"""
    repos = collect_ecr_repos(session, account_id, account_name, region)
    if not repos:
        return None
    return analyze_ecr_repos(repos, account_id, account_name, region)


def run(ctx) -> None:
    """ECR 미사용 이미지 분석"""
    console.print("[bold]ECR 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="ecr")
    results: list[ECRAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_old = sum(r.old_images for r in results)
    total_cost = sum(r.old_images_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"오래된 이미지: [yellow]{total_old}개[/yellow] (${total_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("ecr-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
