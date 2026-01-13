#!/usr/bin/env python3
"""
Inventory management utilities for Dataform Checklist
"""

from typing import List, Optional
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

from .models import SqlxFile


def create_inventory_from_files(files: List[SqlxFile]) -> pd.DataFrame:
    """Create new inventory DataFrame from files"""
    data = []
    
    for idx, file in enumerate(files, start=1):
        data.append({
            'Index': idx,
            'Path/Folder/Directory': file.directory,
            'file_name': file.filename,
            'remark': file.sources,
            'status': 'New'
        })
    
    return pd.DataFrame(data)


def load_existing_inventory(excel_path: Path) -> Optional[pd.DataFrame]:
    """Load existing Excel inventory if it exists"""
    if not excel_path.exists():
        print("No existing file found. Creating new inventory.")
        return None
    
    try:
        print(f"Loading existing inventory from: {excel_path}")
        
        # Find the DataForm sheet with our data
        workbook = load_workbook(excel_path)
        target_sheet = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith('DataForm') or sheet_name.startswith('Dataform'):
                target_sheet = sheet_name
                break
        
        if not target_sheet:
            print("No DataForm sheet found in existing file. Creating new inventory.")
            return None
        
        worksheet = workbook[target_sheet]
        
        # Read data starting from row 9 (rows 1-8 are headers/metadata)
        data = []
        for row in worksheet.iter_rows(min_row=9, values_only=True):
            # Stop if we hit an empty row
            if all(cell is None or cell == '' for cell in row):
                break
            data.append(row)
        
        if not data:
            print("No valid existing inventory found. Creating new inventory.")
            return None
        
        # Create DataFrame with proper column names
        existing_df = pd.DataFrame(data, columns=['Index', 'Path/Folder/Directory', 'file_name', 'remark', 'status'])
        
        # Remove rows where all values are NaN
        existing_df = existing_df.dropna(how='all')
        
        print(f"Loaded {len(existing_df)} existing records")
        return existing_df
    except Exception as e:
        print(f"Warning: Could not load existing file: {e}")
        return None


def merge_inventories(new_inventory: pd.DataFrame, existing_inventory: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Merge new and existing inventories with status tracking"""
    if existing_inventory is None or len(existing_inventory) == 0:
        print("No existing inventory to merge. All files marked as 'New'")
        return new_inventory
    
    print("Merging inventories...")
    
    # Create unique keys for comparison
    new_inventory['_key'] = (new_inventory['Path/Folder/Directory'] + '|' + new_inventory['file_name'])
    existing_inventory['_key'] = (existing_inventory['Path/Folder/Directory'] + '|' + existing_inventory['file_name'])
    
    existing_keys = set(existing_inventory['_key'])
    new_keys = set(new_inventory['_key'])
    
    # Update status for new inventory
    new_count = 0
    existed_count = 0
    
    for idx, row in new_inventory.iterrows():
        if row['_key'] in existing_keys:
            new_inventory.at[idx, 'status'] = 'Existed'
            existed_count += 1
        else:
            new_inventory.at[idx, 'status'] = 'New'
            new_count += 1
    
    # Find deleted files
    deleted_files = []
    deleted_count = 0
    
    for idx, row in existing_inventory.iterrows():
        if row['_key'] not in new_keys:
            deleted_files.append({
                'Index': 0,  # Will be renumbered
                'Path/Folder/Directory': row['Path/Folder/Directory'],
                'file_name': row['file_name'],
                'remark': row.get('remark', ''),
                'status': 'Deleted',
                '_key': row['_key']
            })
            deleted_count += 1
    
    # Combine current files with deleted files
    merged = pd.concat([
        new_inventory,
        pd.DataFrame(deleted_files)
    ], ignore_index=True)
    
    # Renumber indices
    merged['Index'] = range(1, len(merged) + 1)
    
    # Remove temporary key column
    merged = merged.drop(columns=['_key'])
    
    print(f"Merge complete:")
    print(f"  - New files: {new_count}")
    print(f"  - Existed files: {existed_count}")
    print(f"  - Deleted files: {deleted_count}")
    
    return merged
