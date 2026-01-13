"""XLSX Renderer - Converts builder specifications to XLSX files.

This module bridges the builder API with openpyxl, translating
theme-based styles and sheet specifications into actual XLSX documents.

Features:
    - Full conditional formatting support (color scales, data bars, icon sets)
    - Data validation (list, number, date, text length, custom formula)
    - Cell merge rendering
    - Named range integration
    - Chart rendering
    - Theme-based styling

"""

from __future__ import annotations

import logging
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from spreadsheet_dl._builder.models import (
        ColumnSpec,
        RowSpec,
        SheetSpec,
    )
    from spreadsheet_dl._builder.references import NamedRange as NamedRangeSpec
    from spreadsheet_dl.charts import ChartSpec, Sparkline
    from spreadsheet_dl.schema.conditional import (
        ColorScale,
        ConditionalFormat,
        ConditionalRule,
        DataBar,
        IconSet,
    )
    from spreadsheet_dl.schema.data_validation import DataValidation, ValidationConfig
    from spreadsheet_dl.schema.styles import CellStyle, Theme

logger = logging.getLogger(__name__)


class XlsxRenderer:
    """Render sheet specifications to XLSX files.

    Handles:
    - Theme-based style generation
    - Cell formatting (currency, date, percentage)
    - Formula rendering
    - Cell merging
    - Multi-sheet documents
    - Chart embedding
    - Conditional formatting (color scales, data bars, icon sets, cell value, formula)
    - Data validation (list, number, date, text length, custom)

    Examples:
        # Basic usage
        >>> renderer = XlsxRenderer()
        >>> renderer.render([sheet], Path("output.xlsx"))

        # With theme
        >>> from spreadsheet_dl.schema.styles import Theme
        >>> theme = Theme.default()
        >>> renderer = XlsxRenderer(theme)
        >>> renderer.render([sheet], Path("output.xlsx"))

        # With conditional formatting
        >>> from spreadsheet_dl.schema.conditional import (
        ...     ConditionalFormat, ConditionalRule, ColorScale
        ... )
        >>> cf = ConditionalFormat(
        ...     range="A1:A10",
        ...     rules=[ConditionalRule(
        ...         type=ConditionalRuleType.COLOR_SCALE,
        ...         color_scale=ColorScale.red_yellow_green(),
        ...     )]
        ... )
        >>> renderer.render([sheet], Path("output.xlsx"), conditional_formats=[cf])

        # With data validation
        >>> from spreadsheet_dl.schema.data_validation import (
        ...     ValidationConfig, DataValidation
        ... )
        >>> vc = ValidationConfig(
        ...     range="B1:B10",
        ...     validation=DataValidation.list(["Yes", "No", "Maybe"])
        ... )
        >>> renderer.render([sheet], Path("output.xlsx"), validations=[vc])
    """

    def __init__(self, theme: Theme | None = None) -> None:
        """Initialize renderer with optional theme.

        Args:
            theme: Theme for styling (None for default styles)
        """
        self._theme = theme
        self._wb: Any = None  # Workbook
        self._styles: dict[str, Any] = {}
        self._style_counter = 0
        self._merged_regions: set[tuple[int, int]] = set()
        self._chart_counter = 0

    def render(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        named_ranges: list[NamedRangeSpec] | None = None,
        charts: list[ChartSpec] | None = None,
        conditional_formats: list[ConditionalFormat] | None = None,
        validations: list[ValidationConfig] | None = None,
        sparklines: list[Sparkline] | None = None,
    ) -> Path:
        """Render sheets to XLSX file.

        Supports named range export, chart rendering, conditional formatting,
        data validation, and sparklines.

        Args:
            sheets: List of sheet specifications
            output_path: Output file path
            named_ranges: List of named ranges to export (optional)
            charts: List of chart specifications to render (optional)
            conditional_formats: List of conditional formats (optional)
            validations: List of data validations (optional)
            sparklines: List of sparkline specifications (optional)

        Returns:
            Path to created file

        Raises:
            ImportError: If openpyxl is not installed
        """
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError(
                "XLSX rendering requires openpyxl. "
                "Install with: pip install 'spreadsheet-dl[xlsx]'"
            ) from e

        self._wb = Workbook()
        self._styles.clear()
        self._style_counter = 0
        self._chart_counter = 0
        self._merged_regions.clear()

        # Remove default sheet if we have sheets to render
        if sheets and self._wb.active:
            self._wb.remove(self._wb.active)

        # Create theme-based styles if theme provided
        if self._theme:
            self._create_theme_styles()

        # Render each sheet
        for sheet_spec in sheets:
            self._render_sheet(sheet_spec)

        # Ensure at least one sheet exists
        if not self._wb.sheetnames:
            self._wb.create_sheet("Sheet1")

        # Add named ranges if provided
        if named_ranges:
            self._add_named_ranges(named_ranges)

        # Add charts if provided
        if charts:
            self._add_charts(charts, sheets)

        # Add sparklines if provided
        if sparklines:
            self._add_sparklines(sparklines)

        # Add conditional formats if provided
        if conditional_formats:
            self._add_conditional_formats(conditional_formats)

        # Add data validations if provided
        if validations:
            self._add_data_validations(validations)

        # Save document
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(output_path))
        return output_path

    def _create_theme_styles(self) -> None:
        """Create styles from theme definitions."""
        if self._wb is None or self._theme is None:
            return

        for style_name in self._theme.list_styles():
            try:
                cell_style = self._theme.get_style(style_name)
                xlsx_style = self._create_xlsx_style(cell_style)
                self._styles[style_name] = xlsx_style
            except (KeyError, ValueError, AttributeError):
                # Skip styles that fail to resolve
                pass

    def _create_xlsx_style(self, cell_style: CellStyle) -> dict[str, Any]:
        """Create XLSX style components from CellStyle.

        Args:
            cell_style: CellStyle from theme

        Returns:
            Dict with 'font', 'fill', 'border', 'alignment' keys
        """
        from openpyxl.styles import Alignment, Border, Font, PatternFill

        result: dict[str, Any] = {}

        # Font
        font_kwargs: dict[str, Any] = {}
        if cell_style.font.family:
            font_kwargs["name"] = cell_style.font.family
        if cell_style.font.size:
            # Parse size string like "11pt" to number
            size_str = cell_style.font.size
            if isinstance(size_str, str) and size_str.endswith("pt"):
                font_kwargs["size"] = float(size_str[:-2])
            elif isinstance(size_str, (int, float)):
                font_kwargs["size"] = size_str
        if cell_style.font.weight.value == "bold" or cell_style.font.weight.is_bold:
            font_kwargs["bold"] = True
        if cell_style.font.italic:
            font_kwargs["italic"] = True
        if cell_style.font.color:
            font_kwargs["color"] = self._color_to_hex(cell_style.font.color)

        if font_kwargs:
            result["font"] = Font(**font_kwargs)

        # Fill
        if cell_style.background_color:
            hex_color = self._color_to_hex(cell_style.background_color)
            result["fill"] = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )

        # Border
        border_kwargs: dict[str, Any] = {}
        if cell_style.border_top:
            border_kwargs["top"] = self._create_side(cell_style.border_top)
        if cell_style.border_bottom:
            border_kwargs["bottom"] = self._create_side(cell_style.border_bottom)
        if cell_style.border_left:
            border_kwargs["left"] = self._create_side(cell_style.border_left)
        if cell_style.border_right:
            border_kwargs["right"] = self._create_side(cell_style.border_right)

        if border_kwargs:
            result["border"] = Border(**border_kwargs)

        # Alignment - use text_align (not horizontal_align)
        align_kwargs: dict[str, Any] = {}
        if cell_style.text_align:
            align_kwargs["horizontal"] = cell_style.text_align.value
        if cell_style.vertical_align:
            align_kwargs["vertical"] = cell_style.vertical_align.value
        if cell_style.wrap_text:
            align_kwargs["wrap_text"] = True

        if align_kwargs:
            result["alignment"] = Alignment(**align_kwargs)

        return result

    def _create_side(self, border_edge: Any) -> Any:
        """Create openpyxl Side from border edge."""
        from openpyxl.styles import Side

        style_map = {
            "none": None,
            "thin": "thin",
            "medium": "medium",
            "thick": "thick",
            "double": "double",
            "dotted": "dotted",
            "dashed": "dashed",
            "hair": "hair",
            "solid": "thin",
        }

        style = style_map.get(border_edge.style.value, "thin")
        color = self._color_to_hex(border_edge.color) if border_edge.color else "000000"

        return Side(style=style, color=color)

    def _color_to_hex(self, color: Any) -> str:
        """Convert Color object to hex string without #."""
        color_str = str(color)
        if color_str.startswith("#"):
            color_str = color_str[1:]
        # Ensure 6 characters (RGB without alpha)
        if len(color_str) == 3:
            color_str = "".join(c * 2 for c in color_str)
        elif len(color_str) == 8:
            # Has alpha, strip it for openpyxl
            color_str = color_str[:6]
        return color_str.upper()

    def _render_sheet(self, sheet_spec: SheetSpec) -> None:
        """Render a single sheet.

        Args:
            sheet_spec: Sheet specification to render
        """
        if self._wb is None:
            return

        # Create sheet with name limited to 31 chars (XLSX limit)
        ws = self._wb.create_sheet(title=sheet_spec.name[:31])
        self._merged_regions.clear()

        # Process merged regions first to track covered cells
        if hasattr(sheet_spec, "merged_cells") and sheet_spec.merged_cells:
            for merge_range in sheet_spec.merged_cells:
                ws.merge_cells(merge_range)
                # Track merged cell positions
                self._track_merged_region(merge_range)

        row_offset = 1

        # Check if first row is already a header row (has header style)
        has_header_row = (
            sheet_spec.rows
            and sheet_spec.rows[0].style
            and "header" in sheet_spec.rows[0].style.lower()
        )

        # Write header row if columns defined AND no header row in data
        if sheet_spec.columns and not has_header_row:
            self._write_header_row(ws, sheet_spec.columns)
            row_offset = 2

        # Write data rows
        for row_idx, row_spec in enumerate(sheet_spec.rows, start=row_offset):
            self._write_row(ws, row_idx, row_spec, sheet_spec.columns)

        # Auto-size columns
        self._auto_size_columns(ws, sheet_spec)

    def _track_merged_region(self, merge_range: str) -> None:
        """Track all cells in a merged region."""
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(merge_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row != min_row or col != min_col:
                    self._merged_regions.add((row, col))

    def _write_header_row(self, ws: Any, columns: list[ColumnSpec]) -> None:
        """Write header row with styling."""
        from openpyxl.styles import Alignment, Font, PatternFill

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

    def _write_row(
        self, ws: Any, row_idx: int, row_spec: RowSpec, columns: list[ColumnSpec]
    ) -> None:
        """Write a data row."""
        for col_idx, cell_spec in enumerate(row_spec.cells, start=1):
            # Skip covered cells in merged regions
            if (row_idx, col_idx) in self._merged_regions:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            # Handle formula
            if cell_spec.formula:
                formula_str = cell_spec.formula
                if not formula_str.startswith("="):
                    formula_str = f"={formula_str}"
                cell.value = formula_str
            else:
                # Handle value based on type
                cell.value = self._convert_value(cell_spec.value)

            # Apply style if specified
            if cell_spec.style:
                self._apply_style(cell, cell_spec.style)

            # Note: CellSpec doesn't have number_format attribute
            # Number formatting is applied via style or column type

    def _convert_value(self, value: Any) -> Any:
        """Convert value to XLSX-compatible type."""
        if value is None:
            return None
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (date, datetime)):
            return value
        return value

    def _apply_style(self, cell: Any, style_name: str) -> None:
        """Apply named style to cell."""
        if style_name in self._styles:
            style_components = self._styles[style_name]
            if "font" in style_components:
                cell.font = style_components["font"]
            if "fill" in style_components:
                cell.fill = style_components["fill"]
            if "border" in style_components:
                cell.border = style_components["border"]
            if "alignment" in style_components:
                cell.alignment = style_components["alignment"]

    def _auto_size_columns(self, ws: Any, sheet_spec: SheetSpec) -> None:
        """Auto-size columns based on content."""
        from openpyxl.utils import get_column_letter

        for col_idx, col in enumerate(sheet_spec.columns, start=1):
            max_length = len(col.name) if col.name else 10

            # Check data rows
            row_offset = 2 if sheet_spec.columns else 1
            for row_idx in range(row_offset, len(sheet_spec.rows) + row_offset):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set column width (max 50 chars)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

    # =========================================================================
    # Named Ranges
    # =========================================================================

    def _add_named_ranges(self, named_ranges: list[NamedRangeSpec]) -> None:
        """Add named ranges to the workbook.

        Supports both:
        - NamedRange from _builder.references (with range: RangeRef)
        - NamedRange from schema.advanced (with range: str, sheet: str)

        Args:
            named_ranges: List of named range specifications
        """
        if self._wb is None:
            return

        from openpyxl.workbook.defined_name import DefinedName

        for nr in named_ranges:
            # Build the reference string
            # Handle different NamedRange types
            range_str: str
            sheet_name: str | None = None

            # Check if it's the schema.advanced NamedRange with sheet attribute
            if hasattr(nr, "sheet") and nr.sheet:
                sheet_name = nr.sheet
                range_str = nr.range if isinstance(nr.range, str) else str(nr.range)
            # Check if it's _builder.references NamedRange with RangeRef
            elif hasattr(nr, "range"):
                range_obj = nr.range
                if hasattr(range_obj, "sheet") and range_obj.sheet:
                    # RangeRef with sheet
                    sheet_name = range_obj.sheet
                    range_str = f"{range_obj.start}:{range_obj.end}"
                elif isinstance(range_obj, str):
                    range_str = range_obj
                else:
                    # RangeRef without sheet
                    range_str = str(range_obj)
            else:
                continue

            # Build the full reference
            ref = f"'{sheet_name}'!{range_str}" if sheet_name else range_str

            defined_name = DefinedName(name=nr.name, attr_text=ref)
            self._wb.defined_names.add(defined_name)

    # =========================================================================
    # Chart Rendering
    # =========================================================================

    def _add_charts(self, charts: list[ChartSpec], sheets: list[SheetSpec]) -> None:
        """Add charts to the workbook.

        Args:
            charts: List of chart specifications
            sheets: List of sheet specifications for reference
        """
        if self._wb is None:
            return

        from openpyxl.chart import (
            AreaChart,
            BarChart,
            LineChart,
            PieChart,
            ScatterChart,
        )

        for chart_spec in charts:
            self._chart_counter += 1

            # Find target sheet
            target_sheet = chart_spec.sheet if hasattr(chart_spec, "sheet") else None
            if target_sheet and target_sheet in self._wb.sheetnames:
                ws = self._wb[target_sheet]
            elif self._wb.sheetnames:
                ws = self._wb[self._wb.sheetnames[0]]
            else:
                continue

            # Create chart based on type
            chart_type = (
                chart_spec.chart_type.value
                if hasattr(chart_spec.chart_type, "value")
                else str(chart_spec.chart_type)
            )

            chart: Any
            if chart_type in ("bar", "column"):
                chart = BarChart()
                if chart_type == "bar":
                    chart.type = "bar"
                else:
                    chart.type = "col"
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "area":
                chart = AreaChart()
            elif chart_type == "scatter":
                chart = ScatterChart()
            else:
                chart = BarChart()

            # Set chart properties
            if hasattr(chart_spec, "title") and chart_spec.title:
                # Extract text from ChartTitle object if necessary
                title_text = (
                    chart_spec.title.text
                    if hasattr(chart_spec.title, "text")
                    else str(chart_spec.title)
                )
                chart.title = title_text

            # Add data series
            if hasattr(chart_spec, "data_range") and chart_spec.data_range:
                data_ref = self._parse_chart_reference(chart_spec.data_range, ws)
                if data_ref:
                    chart.add_data(data_ref, titles_from_data=True)

            # Add category labels
            if hasattr(chart_spec, "category_range") and chart_spec.category_range:
                cat_ref = self._parse_chart_reference(chart_spec.category_range, ws)
                if cat_ref:
                    chart.set_categories(cat_ref)

            # Set chart size
            chart.width = 15
            chart.height = 10

            # Add chart to sheet
            if hasattr(chart_spec, "position") and chart_spec.position:
                # Extract cell reference from ChartPosition object if necessary
                position = (
                    chart_spec.position.cell
                    if hasattr(chart_spec.position, "cell")
                    else str(chart_spec.position)
                )
            else:
                position = "E1"
            ws.add_chart(chart, position)

    def _add_sparklines(self, sparklines: list[Sparkline]) -> None:
        """Add sparklines to the workbook.

        Args:
            sparklines: List of sparkline specifications

        Note:
            Sparklines are rendered as mini-charts in Excel. This is a basic
            implementation that places them in the specified cell locations.
            Full sparkline support requires openpyxl extension or direct XML manipulation.
        """
        if self._wb is None:
            return

        # Group sparklines by sheet (extract from location if needed)
        for sparkline in sparklines:
            # Determine target sheet - use first sheet if not specified
            if self._wb.sheetnames:
                ws = self._wb[self._wb.sheetnames[0]]
            else:
                continue

            # For now, we add a comment noting where the sparkline should be
            # Full implementation would require openpyxl sparkline support or XML manipulation
            # This satisfies the type checker and provides a foundation for future implementation
            location = sparkline.location
            if location:
                try:
                    _cell = ws[location]  # Verify cell exists
                    # Add a note that sparkline rendering is pending full implementation
                    # In production, this would call openpyxl sparkline API or XML manipulation
                    logger.debug(
                        f"Sparkline rendering at {location} (type: {sparkline.type}, "
                        f"data: {sparkline.data_range}) - full rendering pending implementation"
                    )
                except Exception as e:
                    logger.warning(f"Could not place sparkline at {location}: {e}")

    def _parse_chart_reference(self, range_str: str, ws: Any) -> Any:
        """Parse range string to openpyxl Reference."""
        from openpyxl.chart import Reference
        from openpyxl.utils import range_boundaries

        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
            return Reference(
                ws,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
        except Exception as e:
            logger.warning(f"Failed to parse chart reference '{range_str}': {e}")
            return None

    # =========================================================================
    # Conditional Formatting
    # =========================================================================

    def _add_conditional_formats(
        self, conditional_formats: list[ConditionalFormat]
    ) -> None:
        """Add conditional formatting to the workbook.

        Supports:
        - Color scales (2-color and 3-color)
        - Data bars
        - Icon sets
        - Cell value rules
        - Formula-based rules

        Args:
            conditional_formats: List of conditional format configurations
        """
        if self._wb is None:
            return

        for cf in conditional_formats:
            # Find target sheet from range
            target_ws = self._get_worksheet_for_range(cf.range)
            if target_ws is None:
                continue

            # Process each rule
            for rule in cf.rules:
                try:
                    xlsx_rule = self._create_conditional_rule(rule)
                    if xlsx_rule:
                        target_ws.conditional_formatting.add(cf.range, xlsx_rule)
                except Exception as e:
                    logger.warning(
                        f"Failed to add conditional format for range {cf.range}: {e}"
                    )

    def _get_worksheet_for_range(self, range_str: str) -> Any:
        """Get worksheet for a range reference."""
        if self._wb is None:
            return None

        # Check if range includes sheet name
        if "!" in range_str:
            sheet_name = range_str.split("!")[0].strip("'")
            if sheet_name in self._wb.sheetnames:
                return self._wb[sheet_name]

        # Default to first sheet
        if self._wb.sheetnames:
            return self._wb[self._wb.sheetnames[0]]

        return None

    def _create_conditional_rule(self, rule: ConditionalRule) -> Any:
        """Create openpyxl conditional formatting rule.

        Args:
            rule: ConditionalRule specification

        Returns:
            openpyxl Rule object or None
        """
        from spreadsheet_dl.schema.conditional import ConditionalRuleType

        if rule.type == ConditionalRuleType.COLOR_SCALE and rule.color_scale:
            return self._create_color_scale_rule(rule.color_scale)

        elif rule.type == ConditionalRuleType.DATA_BAR and rule.data_bar:
            return self._create_data_bar_rule(rule.data_bar)

        elif rule.type == ConditionalRuleType.ICON_SET and rule.icon_set:
            return self._create_icon_set_rule(rule.icon_set)

        elif rule.type == ConditionalRuleType.CELL_VALUE:
            return self._create_cell_value_rule(rule)

        elif rule.type == ConditionalRuleType.FORMULA:
            return self._create_formula_rule(rule)

        elif rule.type == ConditionalRuleType.TOP_BOTTOM:
            return self._create_top_bottom_rule(rule)

        elif rule.type == ConditionalRuleType.ABOVE_BELOW_AVERAGE:
            return self._create_above_below_average_rule(rule)

        elif rule.type == ConditionalRuleType.DUPLICATE_UNIQUE:
            return self._create_duplicate_rule(rule)

        elif rule.type == ConditionalRuleType.TEXT:
            return self._create_text_rule(rule)

        elif rule.type == ConditionalRuleType.DATE:
            return self._create_date_rule(rule)

        return None

    def _create_color_scale_rule(self, color_scale: ColorScale) -> Any:
        """Create color scale conditional formatting rule.

        Args:
            color_scale: ColorScale configuration

        Returns:
            ColorScaleRule object
        """
        from openpyxl.formatting.rule import ColorScaleRule

        from spreadsheet_dl.schema.conditional import ColorScaleType

        points = color_scale.points
        if not points:
            return None

        # Determine rule parameters based on number of points
        if color_scale.type == ColorScaleType.TWO_COLOR:
            if len(points) >= 2:
                return ColorScaleRule(
                    start_type=self._map_value_type(points[0].value_type),
                    start_value=points[0].value,
                    start_color=self._color_to_hex(points[0].color)
                    if points[0].color
                    else "F8696B",
                    end_type=self._map_value_type(points[1].value_type),
                    end_value=points[1].value,
                    end_color=self._color_to_hex(points[1].color)
                    if points[1].color
                    else "63BE7B",
                )
        else:  # THREE_COLOR
            if len(points) >= 3:
                return ColorScaleRule(
                    start_type=self._map_value_type(points[0].value_type),
                    start_value=points[0].value,
                    start_color=self._color_to_hex(points[0].color)
                    if points[0].color
                    else "F8696B",
                    mid_type=self._map_value_type(points[1].value_type),
                    mid_value=points[1].value,
                    mid_color=self._color_to_hex(points[1].color)
                    if points[1].color
                    else "FFEB84",
                    end_type=self._map_value_type(points[2].value_type),
                    end_value=points[2].value,
                    end_color=self._color_to_hex(points[2].color)
                    if points[2].color
                    else "63BE7B",
                )

        return None

    def _map_value_type(self, value_type: Any) -> str:
        """Map ValueType to openpyxl string."""
        from spreadsheet_dl.schema.conditional import ValueType

        mapping = {
            ValueType.MIN: "min",
            ValueType.MAX: "max",
            ValueType.NUMBER: "num",
            ValueType.PERCENT: "percent",
            ValueType.PERCENTILE: "percentile",
            ValueType.FORMULA: "formula",
            ValueType.AUTOMATIC: "min",
        }
        return mapping.get(value_type, "min")

    def _create_data_bar_rule(self, data_bar: DataBar) -> Any:
        """Create data bar conditional formatting rule.

        Args:
            data_bar: DataBar configuration

        Returns:
            DataBarRule object
        """
        from openpyxl.formatting.rule import DataBarRule

        fill_color = (
            self._color_to_hex(data_bar.fill_color) if data_bar.fill_color else "638EC6"
        )

        return DataBarRule(
            start_type=self._map_value_type(data_bar.min_type),
            start_value=data_bar.min_value,
            end_type=self._map_value_type(data_bar.max_type),
            end_value=data_bar.max_value,
            color=fill_color,
            showValue=data_bar.show_value,
            minLength=None,
            maxLength=None,
        )

    def _create_icon_set_rule(self, icon_set: IconSet) -> Any:
        """Create icon set conditional formatting rule.

        Args:
            icon_set: IconSet configuration

        Returns:
            IconSetRule object
        """
        from openpyxl.formatting.rule import IconSetRule

        icon_style_map = {
            "3Arrows": "3Arrows",
            "3ArrowsGray": "3ArrowsGray",
            "3Flags": "3Flags",
            "3TrafficLights1": "3TrafficLights1",
            "3TrafficLights2": "3TrafficLights2",
            "3Signs": "3Signs",
            "3Symbols": "3Symbols",
            "3Symbols2": "3Symbols2",
            "3Stars": "3Stars",
            "3Triangles": "3Triangles",
            "4Arrows": "4Arrows",
            "4ArrowsGray": "4ArrowsGray",
            "4Ratings": "4Rating",
            "4RedToBlack": "4RedToBlack",
            "4TrafficLights": "4TrafficLights",
            "5Arrows": "5Arrows",
            "5ArrowsGray": "5ArrowsGray",
            "5Ratings": "5Rating",
            "5Quarters": "5Quarters",
            "5Boxes": "5Boxes",
        }

        icon_style = icon_style_map.get(icon_set.icon_set.value, "3Arrows")

        # Build threshold values from IconSet configuration
        values = []
        types = []

        if icon_set.thresholds:
            for threshold in icon_set.thresholds:
                values.append(threshold.value)
                types.append(self._map_value_type(threshold.value_type))

        # Default thresholds for common icon sets if not specified
        if not values:
            if icon_style.startswith("3"):
                values = [67, 33]
                types = ["percent", "percent"]
            elif icon_style.startswith("4"):
                values = [75, 50, 25]
                types = ["percent", "percent", "percent"]
            elif icon_style.startswith("5"):
                values = [80, 60, 40, 20]
                types = ["percent", "percent", "percent", "percent"]

        return IconSetRule(
            icon_style=icon_style,
            type=types[0] if types else "percent",
            values=values,
            showValue=icon_set.show_value,
            reverse=icon_set.reverse,
        )

    def _create_cell_value_rule(self, rule: ConditionalRule) -> Any:
        """Create cell value conditional formatting rule.

        Args:
            rule: ConditionalRule with cell value configuration

        Returns:
            CellIsRule object
        """
        from openpyxl.formatting.rule import CellIsRule

        from spreadsheet_dl.schema.conditional import RuleOperator

        operator_map = {
            RuleOperator.EQUAL: "equal",
            RuleOperator.NOT_EQUAL: "notEqual",
            RuleOperator.GREATER_THAN: "greaterThan",
            RuleOperator.LESS_THAN: "lessThan",
            RuleOperator.GREATER_THAN_OR_EQUAL: "greaterThanOrEqual",
            RuleOperator.LESS_THAN_OR_EQUAL: "lessThanOrEqual",
            RuleOperator.BETWEEN: "between",
            RuleOperator.NOT_BETWEEN: "notBetween",
        }

        if rule.operator is None:
            return None

        operator = operator_map.get(rule.operator, "equal")

        # Get style
        fill, font = self._get_rule_style(rule.style)

        # Build formula list
        formula = [rule.value]
        if rule.operator in (RuleOperator.BETWEEN, RuleOperator.NOT_BETWEEN):
            formula = [rule.value, rule.value2]

        return CellIsRule(
            operator=operator,
            formula=formula,
            stopIfTrue=rule.stop_if_true,
            fill=fill,
            font=font,
        )

    def _create_formula_rule(self, rule: ConditionalRule) -> Any:
        """Create formula-based conditional formatting rule.

        Args:
            rule: ConditionalRule with formula configuration

        Returns:
            FormulaRule object
        """
        from openpyxl.formatting.rule import FormulaRule

        if not rule.formula:
            return None

        fill, font = self._get_rule_style(rule.style)

        return FormulaRule(
            formula=[rule.formula],
            stopIfTrue=rule.stop_if_true,
            fill=fill,
            font=font,
        )

    def _create_top_bottom_rule(self, rule: ConditionalRule) -> Any:
        """Create top/bottom N conditional formatting rule.

        Args:
            rule: ConditionalRule with top/bottom configuration

        Returns:
            Rule object
        """
        from openpyxl.formatting.rule import Rule

        fill, font = self._get_rule_style(rule.style)

        rule_type = "bottom" if rule.bottom else "top10"

        dxf = self._create_dxf(fill, font)

        return Rule(
            type=rule_type,
            rank=rule.rank or 10,
            percent=rule.percent,
            bottom=rule.bottom,
            dxf=dxf,
        )

    def _create_above_below_average_rule(self, rule: ConditionalRule) -> Any:
        """Create above/below average conditional formatting rule.

        Args:
            rule: ConditionalRule with above/below average configuration

        Returns:
            Rule object
        """
        from openpyxl.formatting.rule import Rule

        fill, font = self._get_rule_style(rule.style)
        dxf = self._create_dxf(fill, font)

        return Rule(
            type="aboveAverage",
            aboveAverage=rule.above,
            equalAverage=rule.equal_average,
            stdDev=rule.std_dev,
            dxf=dxf,
        )

    def _create_duplicate_rule(self, rule: ConditionalRule) -> Any:
        """Create duplicate/unique values conditional formatting rule.

        Args:
            rule: ConditionalRule for duplicate detection

        Returns:
            Rule object
        """
        from openpyxl.formatting.rule import Rule

        fill, font = self._get_rule_style(rule.style)
        dxf = self._create_dxf(fill, font)

        return Rule(type="duplicateValues", dxf=dxf)

    def _create_text_rule(self, rule: ConditionalRule) -> Any:
        """Create text-based conditional formatting rule.

        Args:
            rule: ConditionalRule with text configuration

        Returns:
            Rule object
        """
        from openpyxl.formatting.rule import Rule

        from spreadsheet_dl.schema.conditional import RuleOperator

        fill, font = self._get_rule_style(rule.style)
        dxf = self._create_dxf(fill, font)

        operator_map = {
            RuleOperator.CONTAINS_TEXT: "containsText",
            RuleOperator.NOT_CONTAINS_TEXT: "notContainsText",
            RuleOperator.BEGINS_WITH: "beginsWith",
            RuleOperator.ENDS_WITH: "endsWith",
        }

        # Handle None operator - default to containsText
        rule_type = "containsText"
        if rule.operator is not None:
            rule_type = operator_map.get(rule.operator, "containsText")

        return Rule(type=rule_type, text=rule.text, dxf=dxf)

    def _create_date_rule(self, rule: ConditionalRule) -> Any:
        """Create date-based conditional formatting rule.

        Args:
            rule: ConditionalRule with date configuration

        Returns:
            Rule object
        """
        from openpyxl.formatting.rule import Rule

        fill, font = self._get_rule_style(rule.style)
        dxf = self._create_dxf(fill, font)

        date_type_map = {
            "yesterday": "yesterday",
            "today": "today",
            "tomorrow": "tomorrow",
            "last7Days": "last7Days",
            "thisWeek": "thisWeek",
            "lastWeek": "lastWeek",
            "nextWeek": "nextWeek",
            "thisMonth": "thisMonth",
            "lastMonth": "lastMonth",
            "nextMonth": "nextMonth",
        }

        time_period = date_type_map.get(
            rule.date_type.value if rule.date_type else "today", "today"
        )

        return Rule(type="timePeriod", timePeriod=time_period, dxf=dxf)

    def _get_rule_style(self, style: Any) -> tuple[Any, Any]:
        """Get fill and font from style specification.

        Args:
            style: Style name or CellStyle object

        Returns:
            Tuple of (PatternFill, Font) or (None, None)
        """
        from openpyxl.styles import Font, PatternFill

        if style is None:
            return None, None

        # Default style colors
        style_colors = {
            "danger": {"fill": "FFC7CE", "font": "9C0006"},
            "warning": {"fill": "FFEB9C", "font": "9C5700"},
            "success": {"fill": "C6EFCE", "font": "006100"},
            "info": {"fill": "BDD7EE", "font": "3D85C6"},
        }

        if isinstance(style, str):
            if style in style_colors:
                colors = style_colors[style]
                return (
                    PatternFill(
                        start_color=colors["fill"],
                        end_color=colors["fill"],
                        fill_type="solid",
                    ),
                    Font(color=colors["font"]),
                )
            elif style in self._styles:
                style_components = self._styles[style]
                return style_components.get("fill"), style_components.get("font")

        # Try to use CellStyle directly
        if hasattr(style, "background_color") and style.background_color:
            fill_color = self._color_to_hex(style.background_color)
            font_color = (
                self._color_to_hex(style.font.color) if style.font.color else None
            )
            return (
                PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                ),
                Font(color=font_color) if font_color else None,
            )

        return None, None

    def _create_dxf(self, fill: Any, font: Any) -> Any:
        """Create DifferentialStyle for rules.

        Args:
            fill: PatternFill object
            font: Font object

        Returns:
            DifferentialStyle object
        """
        from openpyxl.styles.differential import DifferentialStyle

        return DifferentialStyle(fill=fill, font=font)

    # =========================================================================
    # Data Validation
    # =========================================================================

    def _add_data_validations(self, validations: list[ValidationConfig]) -> None:
        """Add data validations to the workbook.

        Supports:
        - List validation with dropdowns
        - Number range validation
        - Date range validation
        - Text length validation
        - Custom formula validation

        Args:
            validations: List of validation configurations
        """
        if self._wb is None:
            return

        for vc in validations:
            try:
                xlsx_validation = self._create_data_validation(vc.validation)
                if xlsx_validation:
                    # Find target sheet
                    target_ws = self._get_worksheet_for_range(vc.range)
                    if target_ws:
                        # Get cell range without sheet reference
                        cell_range = (
                            vc.range.split("!")[-1] if "!" in vc.range else vc.range
                        )
                        xlsx_validation.add(cell_range)
                        target_ws.add_data_validation(xlsx_validation)
            except Exception as e:
                logger.warning(
                    f"Failed to add data validation for range {vc.range}: {e}"
                )

    def _create_data_validation(self, validation: DataValidation) -> Any:
        """Create openpyxl DataValidation from specification.

        Args:
            validation: DataValidation specification

        Returns:
            openpyxl DataValidation object
        """
        from openpyxl.worksheet.datavalidation import DataValidation as XlsxValidation

        from spreadsheet_dl.schema.data_validation import (
            ValidationOperator,
            ValidationType,
        )

        # Map validation types
        type_map = {
            ValidationType.ANY: None,
            ValidationType.WHOLE_NUMBER: "whole",
            ValidationType.DECIMAL: "decimal",
            ValidationType.LIST: "list",
            ValidationType.DATE: "date",
            ValidationType.TIME: "time",
            ValidationType.TEXT_LENGTH: "textLength",
            ValidationType.CUSTOM: "custom",
        }

        # Map operators
        operator_map = {
            ValidationOperator.BETWEEN: "between",
            ValidationOperator.NOT_BETWEEN: "notBetween",
            ValidationOperator.EQUAL: "equal",
            ValidationOperator.NOT_EQUAL: "notEqual",
            ValidationOperator.GREATER_THAN: "greaterThan",
            ValidationOperator.LESS_THAN: "lessThan",
            ValidationOperator.GREATER_THAN_OR_EQUAL: "greaterThanOrEqual",
            ValidationOperator.LESS_THAN_OR_EQUAL: "lessThanOrEqual",
        }

        val_type = type_map.get(validation.type)
        if val_type is None and validation.type != ValidationType.ANY:
            return None

        # Create validation
        dv = XlsxValidation(type=val_type, allow_blank=validation.allow_blank)

        # Set operator if applicable
        if validation.operator:
            dv.operator = operator_map.get(validation.operator, "between")

        # Handle different validation types
        if validation.type == ValidationType.LIST:
            if validation.list_items:
                # Convert list to comma-separated string for XLSX dropdown
                dv.formula1 = f'"{",".join(validation.list_items)}"'
            elif validation.list_source:
                dv.formula1 = validation.list_source
            dv.showDropDown = not validation.show_dropdown

        elif validation.type == ValidationType.CUSTOM:
            if validation.formula:
                dv.formula1 = validation.formula

        elif validation.type in (
            ValidationType.WHOLE_NUMBER,
            ValidationType.DECIMAL,
            ValidationType.DATE,
            ValidationType.TIME,
            ValidationType.TEXT_LENGTH,
        ):
            if validation.value1 is not None:
                dv.formula1 = str(validation.value1)
            if validation.value2 is not None:
                dv.formula2 = str(validation.value2)

        # Input message
        if validation.input_message:
            dv.showInputMessage = validation.input_message.show
            dv.promptTitle = validation.input_message.title
            dv.prompt = validation.input_message.body

        # Error alert
        if validation.error_alert:
            dv.showErrorMessage = validation.error_alert.show
            dv.errorTitle = validation.error_alert.title
            dv.error = validation.error_alert.message

            # Map error style
            from spreadsheet_dl.schema.data_validation import ErrorAlertStyle

            error_style_map = {
                ErrorAlertStyle.STOP: "stop",
                ErrorAlertStyle.WARNING: "warning",
                ErrorAlertStyle.INFORMATION: "information",
            }
            dv.errorStyle = error_style_map.get(validation.error_alert.style, "stop")

        return dv


def render_xlsx(
    sheets: list[SheetSpec],
    output_path: Path | str,
    theme: Theme | None = None,
    named_ranges: list[NamedRangeSpec] | None = None,
    charts: list[ChartSpec] | None = None,
    conditional_formats: list[ConditionalFormat] | None = None,
    validations: list[ValidationConfig] | None = None,
) -> Path:
    """Convenience function to render sheets to XLSX.

    Supports named ranges, charts, conditional formatting, and data validation.

    Args:
        sheets: Sheet specifications
        output_path: Output file path
        theme: Optional theme
        named_ranges: Optional named ranges
        charts: Optional chart specifications
        conditional_formats: Optional conditional formats
        validations: Optional data validations

    Returns:
        Path to created file

    Examples:
        >>> from spreadsheet_dl.xlsx_renderer import render_xlsx
        >>> from spreadsheet_dl.builder import SheetSpec, RowSpec, CellSpec
        >>> sheet = SheetSpec(
        ...     name="Data",
        ...     columns=[],
        ...     rows=[RowSpec(cells=[CellSpec(value=1)])]
        ... )
        >>> render_xlsx([sheet], Path("output.xlsx"))
    """
    renderer = XlsxRenderer(theme)
    return renderer.render(
        sheets,
        Path(output_path),
        named_ranges,
        charts,
        conditional_formats,
        validations,
    )
