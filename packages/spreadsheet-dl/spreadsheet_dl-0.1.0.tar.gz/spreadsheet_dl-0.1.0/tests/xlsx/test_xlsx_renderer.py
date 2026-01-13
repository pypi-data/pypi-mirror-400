"""Comprehensive tests for XLSX renderer.

Tests:
    - Basic rendering functionality
    - Conditional formatting (color scales, data bars, icon sets, cell value, formula)
    - Data validation (list, number, date, text length, custom)
    - Named ranges
    - Charts
    - Cell merging
    - Theme-based styling
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec, SheetSpec

pytestmark = [pytest.mark.unit, pytest.mark.rendering]


# =============================================================================
# Basic Rendering Tests
# =============================================================================


class TestXlsxRendererBasic:
    """Test basic XLSX rendering functionality."""

    def test_render_empty_sheet(self, empty_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test rendering an empty sheet."""
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        output_path = tmp_path / "empty.xlsx"
        renderer = XlsxRenderer()
        result = renderer.render([empty_sheet], output_path)

        assert result == output_path
        assert output_path.exists()

    def test_render_simple_sheet(self, sample_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test rendering a simple sheet with data."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        output_path = tmp_path / "simple.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path)

        # Verify file contents
        wb = load_workbook(output_path)
        ws = wb.active

        # Check header row
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Age"
        assert ws.cell(1, 3).value == "Salary"

        # Check data rows
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 30
        assert ws.cell(2, 3).value == 75000.50

        wb.close()

    def test_render_multiple_sheets(self, tmp_path: Path) -> None:
        """Test rendering multiple sheets."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sheet1 = SheetSpec(
            name="Sheet1",
            columns=[ColumnSpec(name="A")],
            rows=[RowSpec(cells=[CellSpec(value=1)])],
        )
        sheet2 = SheetSpec(
            name="Sheet2",
            columns=[ColumnSpec(name="B")],
            rows=[RowSpec(cells=[CellSpec(value=2)])],
        )

        output_path = tmp_path / "multi.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet1, sheet2], output_path)

        wb = load_workbook(output_path)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames
        wb.close()

    def test_render_with_formulas(
        self, formula_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering sheets with formulas."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        output_path = tmp_path / "formulas.xlsx"
        renderer = XlsxRenderer()
        renderer.render([formula_sheet], output_path)

        wb = load_workbook(output_path, data_only=False)
        ws = wb.active
        assert ws.cell(2, 3).value == "=A2+B2"
        wb.close()

    def test_render_with_decimal_values(self, tmp_path: Path) -> None:
        """Test that Decimal values are converted to float."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sheet = SheetSpec(
            name="Decimals",
            columns=[ColumnSpec(name="Value")],
            rows=[RowSpec(cells=[CellSpec(value=Decimal("123.456"))])],
        )

        output_path = tmp_path / "decimals.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet], output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(2, 1).value == 123.456
        wb.close()

    def test_render_with_dates(self, tmp_path: Path) -> None:
        """Test rendering date values."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        test_date = date(2025, 6, 15)
        sheet = SheetSpec(
            name="Dates",
            columns=[ColumnSpec(name="Date")],
            rows=[RowSpec(cells=[CellSpec(value=test_date)])],
        )

        output_path = tmp_path / "dates.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet], output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        # openpyxl returns datetime for dates
        assert ws.cell(2, 1).value.date() == test_date
        wb.close()


# =============================================================================
# Conditional Formatting Tests
# =============================================================================


class TestXlsxConditionalFormatting:
    """Test conditional formatting support."""

    def test_color_scale_two_color(
        self, numeric_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test two-color scale conditional formatting."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ColorScale,
            ConditionalFormat,
            ConditionalRule,
            ConditionalRuleType,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule(
                    type=ConditionalRuleType.COLOR_SCALE,
                    color_scale=ColorScale.white_to_blue(),
                )
            ],
        )

        output_path = tmp_path / "color_scale_2.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_color_scale_three_color(
        self, numeric_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test three-color scale conditional formatting."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ColorScale,
            ConditionalFormat,
            ConditionalRule,
            ConditionalRuleType,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule(
                    type=ConditionalRuleType.COLOR_SCALE,
                    color_scale=ColorScale.red_yellow_green(),
                )
            ],
        )

        output_path = tmp_path / "color_scale_3.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_data_bar(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test data bar conditional formatting."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ConditionalFormat,
            ConditionalRule,
            ConditionalRuleType,
            DataBar,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule(
                    type=ConditionalRuleType.DATA_BAR,
                    data_bar=DataBar.default(),
                )
            ],
        )

        output_path = tmp_path / "data_bar.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_icon_set(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test icon set conditional formatting."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ConditionalFormat,
            ConditionalRule,
            ConditionalRuleType,
            IconSet,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="C2:C6",
            rules=[
                ConditionalRule(
                    type=ConditionalRuleType.ICON_SET,
                    icon_set=IconSet.three_arrows(),
                )
            ],
        )

        output_path = tmp_path / "icon_set.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_cell_value_rule(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test cell value conditional formatting rule."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ConditionalFormat,
            ConditionalRule,
            RuleOperator,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule.cell_value(
                    operator=RuleOperator.GREATER_THAN,
                    value=50,
                    style="success",
                )
            ],
        )

        output_path = tmp_path / "cell_value.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_formula_rule(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test formula-based conditional formatting rule."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import ConditionalFormat, ConditionalRule
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule.from_formula(
                    formula="$A2>50",
                    style="warning",
                )
            ],
        )

        output_path = tmp_path / "formula_rule.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_between_rule(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test between conditional formatting rule."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import ConditionalFormat, ConditionalRule
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule.between(
                    min_value=30,
                    max_value=70,
                    style="info",
                )
            ],
        )

        output_path = tmp_path / "between_rule.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_multiple_conditional_formats(
        self, numeric_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test multiple conditional formats on same range."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.conditional import (
            ConditionalFormat,
            ConditionalRule,
            RuleOperator,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        cf = ConditionalFormat(
            range="A2:A6",
            rules=[
                ConditionalRule.cell_value(
                    operator=RuleOperator.LESS_THAN,
                    value=30,
                    style="danger",
                    priority=1,
                ),
                ConditionalRule.cell_value(
                    operator=RuleOperator.GREATER_THAN,
                    value=70,
                    style="success",
                    priority=2,
                ),
            ],
        )

        output_path = tmp_path / "multi_cf.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, conditional_formats=[cf])

        wb = load_workbook(output_path)
        ws = wb.active
        # openpyxl stores rules per range - count total rules across all ranges
        total_rules = sum(
            len(rules) for rules in ws.conditional_formatting._cf_rules.values()
        )
        assert total_rules >= 2
        wb.close()


# =============================================================================
# Data Validation Tests
# =============================================================================


class TestXlsxDataValidation:
    """Test data validation support."""

    def test_list_validation(self, sample_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test list/dropdown validation."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.list(
                items=["Option A", "Option B", "Option C"],
            ),
        )

        output_path = tmp_path / "list_validation.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.data_validations.dataValidation) > 0
        wb.close()

    def test_number_validation(self, numeric_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test number range validation."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.decimal_between(
                min_value=0,
                max_value=100,
            ),
        )

        output_path = tmp_path / "number_validation.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.data_validations.dataValidation) > 0
        wb.close()

    def test_date_validation(self, tmp_path: Path) -> None:
        """Test date validation."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sheet = SheetSpec(
            name="Dates",
            columns=[ColumnSpec(name="Date")],
            rows=[RowSpec(cells=[CellSpec(value=date(2025, 1, 1))])],
        )

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.date_between(
                start_date=date(2025, 1, 1),
                end_date=date(2025, 12, 31),
            ),
        )

        output_path = tmp_path / "date_validation.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.data_validations.dataValidation) > 0
        wb.close()

    def test_text_length_validation(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test text length validation."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.text_max_length(max_length=50),
        )

        output_path = tmp_path / "text_length_validation.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.data_validations.dataValidation) > 0
        wb.close()

    def test_custom_formula_validation(
        self, numeric_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test custom formula validation."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.custom(
                formula="AND(A2>0, A2<100)",
            ),
        )

        output_path = tmp_path / "custom_validation.xlsx"
        renderer = XlsxRenderer()
        renderer.render([numeric_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws.data_validations.dataValidation) > 0
        wb.close()

    def test_validation_with_input_message(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test validation with input message."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            InputMessage,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.list(
                items=["Yes", "No"],
                input_message=InputMessage(
                    title="Select Option",
                    body="Please select Yes or No",
                ),
            ),
        )

        output_path = tmp_path / "validation_input_msg.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        dv = ws.data_validations.dataValidation[0]
        assert dv.promptTitle == "Select Option"
        assert dv.prompt == "Please select Yes or No"
        wb.close()

    def test_validation_with_error_alert(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test validation with error alert."""
        from openpyxl import load_workbook

        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ErrorAlert,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.positive_number(
                error_alert=ErrorAlert.stop(
                    title="Invalid Input",
                    message="Value must be positive",
                ),
            ),
        )

        output_path = tmp_path / "validation_error_alert.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, validations=[validation])

        wb = load_workbook(output_path)
        ws = wb.active
        dv = ws.data_validations.dataValidation[0]
        assert dv.errorTitle == "Invalid Input"
        assert dv.error == "Value must be positive"
        wb.close()


# =============================================================================
# Named Range Tests
# =============================================================================


class TestXlsxNamedRanges:
    """Test named range support."""

    def test_add_named_range(self, sample_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test adding a named range."""
        from openpyxl import load_workbook

        from spreadsheet_dl._builder.references import NamedRange, RangeRef
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        named_range = NamedRange(
            name="TestRange",
            range=RangeRef("A2", "C3", sheet="TestSheet"),
        )

        output_path = tmp_path / "named_range.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, named_ranges=[named_range])

        wb = load_workbook(output_path)
        assert "TestRange" in wb.defined_names
        wb.close()

    def test_multiple_named_ranges(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test adding multiple named ranges."""
        from openpyxl import load_workbook

        from spreadsheet_dl._builder.references import NamedRange, RangeRef
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        named_ranges = [
            NamedRange(name="Names", range=RangeRef("A2", "A3", sheet="TestSheet")),
            NamedRange(name="Ages", range=RangeRef("B2", "B3", sheet="TestSheet")),
            NamedRange(name="Salaries", range=RangeRef("C2", "C3", sheet="TestSheet")),
        ]

        output_path = tmp_path / "multi_named_range.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sample_sheet], output_path, named_ranges=named_ranges)

        wb = load_workbook(output_path)
        assert "Names" in wb.defined_names
        assert "Ages" in wb.defined_names
        assert "Salaries" in wb.defined_names
        wb.close()


# =============================================================================
# Convenience Function Tests
# =============================================================================


class TestRenderXlsxFunction:
    """Test the render_xlsx convenience function."""

    def test_render_xlsx_basic(self, sample_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test basic usage of render_xlsx function."""
        from spreadsheet_dl.xlsx_renderer import render_xlsx

        output_path = tmp_path / "convenience.xlsx"
        result = render_xlsx([sample_sheet], output_path)

        assert result == output_path
        assert output_path.exists()

    def test_render_xlsx_with_string_path(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test render_xlsx with string path."""
        from spreadsheet_dl.xlsx_renderer import render_xlsx

        output_path = str(tmp_path / "string_path.xlsx")
        result = render_xlsx([sample_sheet], output_path)

        assert result == Path(output_path)
        assert Path(output_path).exists()


# =============================================================================
# Edge Case Tests
# =============================================================================


class TestXlsxRendererEdgeCases:
    """Test edge cases and error handling."""

    def test_long_sheet_name_truncated(self, tmp_path: Path) -> None:
        """Test that sheet names longer than 31 chars are truncated."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        long_name = "A" * 50  # 50 characters
        sheet = SheetSpec(
            name=long_name,
            columns=[ColumnSpec(name="Col")],
            rows=[RowSpec(cells=[CellSpec(value=1)])],
        )

        output_path = tmp_path / "long_name.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet], output_path)

        wb = load_workbook(output_path)
        assert len(wb.sheetnames[0]) == 31
        wb.close()

    def test_special_characters_in_values(self, tmp_path: Path) -> None:
        """Test handling of special characters in cell values."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sheet = SheetSpec(
            name="Special",
            columns=[ColumnSpec(name="Data")],
            rows=[
                RowSpec(cells=[CellSpec(value="Hello & World")]),
                RowSpec(cells=[CellSpec(value="<test>")]),
                RowSpec(cells=[CellSpec(value='"quoted"')]),
            ],
        )

        output_path = tmp_path / "special.xlsx"
        renderer = XlsxRenderer()
        renderer.render([sheet], output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "Hello & World"
        assert ws.cell(3, 1).value == "<test>"
        assert ws.cell(4, 1).value == '"quoted"'
        wb.close()

    def test_empty_validation_list(
        self, sample_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test validation with empty list doesn't crash."""
        from spreadsheet_dl.schema.data_validation import (
            DataValidation,
            ValidationConfig,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        validation = ValidationConfig(
            range="A2:A10",
            validation=DataValidation.list(items=[]),  # Empty list
        )

        output_path = tmp_path / "empty_list_validation.xlsx"
        renderer = XlsxRenderer()
        # Should not raise an exception
        renderer.render([sample_sheet], output_path, validations=[validation])
        assert output_path.exists()

    def test_no_sheets_creates_default(self, tmp_path: Path) -> None:
        """Test that no sheets creates a default Sheet1."""
        from openpyxl import load_workbook

        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        output_path = tmp_path / "no_sheets.xlsx"
        renderer = XlsxRenderer()
        renderer.render([], output_path)

        wb = load_workbook(output_path)
        # Should have at least one sheet
        assert len(wb.sheetnames) >= 1
        wb.close()

    def test_color_conversion_short_hex(self, tmp_path: Path) -> None:
        """Test color conversion handles short hex codes."""
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        renderer = XlsxRenderer()

        # Test 3-character hex
        result = renderer._color_to_hex("#ABC")
        assert result == "AABBCC"

        # Test 6-character hex
        result = renderer._color_to_hex("#AABBCC")
        assert result == "AABBCC"

        # Test 8-character hex (with alpha)
        result = renderer._color_to_hex("#AABBCCDD")
        assert result == "AABBCC"
