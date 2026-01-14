"""
ExcelWriter测试用例
"""
import pytest
import base64
import openpyxl
import io
from free_mcp_excel.writer import ExcelWriter
from free_mcp_excel.utils import base64_to_excel


@pytest.fixture
def writer():
    """创建写入器实例"""
    return ExcelWriter()


@pytest.fixture
def test_file_xlsx():
    """创建测试用的.xlsx文件（Base64）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "测试"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_content = buffer.read()
    wb.close()
    
    return base64.b64encode(file_content).decode("utf-8")


class TestExcelWriter:
    """ExcelWriter测试类"""
    
    def test_create_workbook(self, writer):
        """测试创建工作簿"""
        result = writer.create_workbook("NewSheet")
        assert result["status"] == "success"
        assert "file" in result["data"]
        
        # 验证创建的文件
        file_content = base64_to_excel(result["data"]["file"])
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        assert "NewSheet" in wb.sheetnames
        wb.close()
    
    def test_create_sheet(self, writer, test_file_xlsx):
        """测试创建工作表"""
        result = writer.create_sheet(test_file_xlsx, "NewSheet")
        assert result["status"] == "success"
        assert "file" in result["data"]
        
        # 验证新工作表
        file_content = base64_to_excel(result["data"]["file"])
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        assert "NewSheet" in wb.sheetnames
        wb.close()
    
    def test_write_cell_data(self, writer, test_file_xlsx):
        """测试写入单元格数据"""
        result = writer.write_cell_data(
            test_file_xlsx,
            "Sheet",
            "B1",
            "新值"
        )
        assert result["status"] == "success"
        assert "file" in result["data"]
        
        # 验证写入的值
        file_content = base64_to_excel(result["data"]["file"])
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        assert ws["B1"].value == "新值"
        wb.close()
    
    def test_write_cell_formula(self, writer, test_file_xlsx):
        """测试写入单元格公式"""
        result = writer.write_cell_formula(
            test_file_xlsx,
            "Sheet",
            "A2",
            "=SUM(A1:A1)"
        )
        assert result["status"] == "success"
        assert "file" in result["data"]
    
    def test_write_range_data(self, writer, test_file_xlsx):
        """测试批量写入数据"""
        data = [
            ["列1", "列2"],
            ["值1", "值2"],
            ["值3", "值4"]
        ]
        result = writer.write_range_data(
            test_file_xlsx,
            "Sheet",
            "C1",
            data
        )
        assert result["status"] == "success"
        assert "file" in result["data"]
    
    def test_merge_cells(self, writer, test_file_xlsx):
        """测试合并单元格"""
        result = writer.merge_cells(
            test_file_xlsx,
            "Sheet",
            "A1:B1"
        )
        assert result["status"] == "success"
        assert "file" in result["data"]
    
    def test_unmerge_cells(self, writer, test_file_xlsx):
        """测试取消合并单元格"""
        # 先合并
        result1 = writer.merge_cells(test_file_xlsx, "Sheet", "A1:B1")
        # 再取消合并
        result2 = writer.unmerge_cells(
            result1["data"]["file"],
            "Sheet",
            "A1:B1"
        )
        assert result2["status"] == "success"

