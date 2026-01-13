"""
plugins/fn/provisioned.py - Provisioned Concurrency 분석

Lambda Provisioned Concurrency 최적화 분석:
- 실제 동시 실행 대비 PC 과다 설정 탐지
- 미사용 PC 탐지 (비용 낭비)
- PC 부족 탐지 (Throttle 발생)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_lambda_provisioned_monthly_cost

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListProvisionedConcurrencyConfigs",
        "lambda:GetFunctionConcurrency",
        "cloudwatch:GetMetricStatistics",
    ],
}


class PCStatus(Enum):
    """PC 상태"""

    UNUSED = "unused"  # PC 설정됐으나 미사용
    OVERSIZED = "oversized"  # 과다 설정
    OPTIMAL = "optimal"  # 적정
    UNDERSIZED = "undersized"  # 부족 (Throttle 발생)
    NO_PC = "no_pc"  # PC 미설정


@dataclass
class PCConfig:
    """Provisioned Concurrency 설정"""

    qualifier: str  # $LATEST, version, alias
    allocated: int
    available: int
    status: str  # InProgress, Ready, Failed


@dataclass
class LambdaPCInfo:
    """Lambda PC 정보"""

    function_name: str
    function_arn: str
    runtime: str
    memory_mb: int

    # PC 설정
    pc_configs: list[PCConfig] = field(default_factory=list)
    total_provisioned: int = 0
    reserved_concurrency: int | None = None

    # 메트릭
    invocations_30d: int = 0
    max_concurrent: int = 0
    avg_concurrent: float = 0.0
    throttles_30d: int = 0

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def has_pc(self) -> bool:
        return self.total_provisioned > 0

    @property
    def utilization_pct(self) -> float:
        """PC 활용률"""
        if self.total_provisioned == 0:
            return 0.0
        return min(100.0, self.max_concurrent / self.total_provisioned * 100)


@dataclass
class PCFinding:
    """PC 분석 결과"""

    function: LambdaPCInfo
    status: PCStatus
    recommendation: str
    recommended_pc: int | None = None
    monthly_waste: float = 0.0
    monthly_savings: float = 0.0


@dataclass
class PCAnalysisResult:
    """PC 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    functions_with_pc: int = 0
    unused_pc_count: int = 0
    oversized_pc_count: int = 0
    undersized_pc_count: int = 0
    total_pc_cost: float = 0.0
    potential_savings: float = 0.0
    findings: list[PCFinding] = field(default_factory=list)


# =============================================================================
# 수집
# =============================================================================


def collect_pc_info(
    session,
    account_id: str,
    account_name: str,
    region: str,
) -> list[LambdaPCInfo]:
    """Lambda PC 정보 수집"""
    from botocore.exceptions import ClientError

    functions = []

    try:
        lambda_client = get_client(session, "lambda", region_name=region)
        cloudwatch = get_client(session, "cloudwatch", region_name=region)

        end_time = datetime.now(timezone.utc)
        start_time = end_time - timedelta(days=30)

        paginator = lambda_client.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page.get("Functions", []):
                function_name = fn.get("FunctionName", "")

                info = LambdaPCInfo(
                    function_name=function_name,
                    function_arn=fn.get("FunctionArn", ""),
                    runtime=fn.get("Runtime", ""),
                    memory_mb=fn.get("MemorySize", 128),
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                )

                # PC 설정 조회
                try:
                    pc_response = lambda_client.list_provisioned_concurrency_configs(FunctionName=function_name)
                    for pc in pc_response.get("ProvisionedConcurrencyConfigs", []):
                        qualifier = (
                            pc.get("FunctionArn", "").split(":")[-1] if ":" in pc.get("FunctionArn", "") else "$LATEST"
                        )
                        config = PCConfig(
                            qualifier=qualifier,
                            allocated=pc.get("AllocatedProvisionedConcurrentExecutions", 0),
                            available=pc.get("AvailableProvisionedConcurrentExecutions", 0),
                            status=pc.get("Status", ""),
                        )
                        info.pc_configs.append(config)
                        info.total_provisioned += config.allocated
                except ClientError:
                    pass

                # Reserved Concurrency 조회
                try:
                    concurrency = lambda_client.get_function_concurrency(FunctionName=function_name)
                    info.reserved_concurrency = concurrency.get("ReservedConcurrentExecutions")
                except ClientError:
                    pass

                # CloudWatch 메트릭
                dimensions = [{"Name": "FunctionName", "Value": function_name}]

                # Invocations
                try:
                    inv_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="Invocations",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=86400 * 30,
                        Statistics=["Sum"],
                    )
                    for dp in inv_response.get("Datapoints", []):
                        info.invocations_30d += int(dp.get("Sum", 0))
                except ClientError:
                    pass

                # ConcurrentExecutions
                try:
                    conc_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="ConcurrentExecutions",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=3600,  # 1시간 단위
                        Statistics=["Maximum", "Average"],
                    )
                    for dp in conc_response.get("Datapoints", []):
                        max_val = int(dp.get("Maximum", 0))
                        if max_val > info.max_concurrent:
                            info.max_concurrent = max_val
                        info.avg_concurrent = max(info.avg_concurrent, dp.get("Average", 0))
                except ClientError:
                    pass

                # Throttles
                try:
                    throttle_response = cloudwatch.get_metric_statistics(
                        Namespace="AWS/Lambda",
                        MetricName="Throttles",
                        Dimensions=dimensions,
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=86400 * 30,
                        Statistics=["Sum"],
                    )
                    for dp in throttle_response.get("Datapoints", []):
                        info.throttles_30d += int(dp.get("Sum", 0))
                except ClientError:
                    pass

                # 비용 계산
                if info.total_provisioned > 0:
                    info.monthly_cost = get_lambda_provisioned_monthly_cost(
                        region=region,
                        memory_mb=info.memory_mb,
                        provisioned_concurrency=info.total_provisioned,
                    )

                functions.append(info)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return functions


# =============================================================================
# 분석
# =============================================================================


def analyze_pc(
    functions: list[LambdaPCInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> PCAnalysisResult:
    """PC 분석"""
    result = PCAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    for func in functions:
        finding = _analyze_single_pc(func, region)
        result.findings.append(finding)

        if func.has_pc:
            result.functions_with_pc += 1
            result.total_pc_cost += func.monthly_cost

        if finding.status == PCStatus.UNUSED:
            result.unused_pc_count += 1
            result.potential_savings += finding.monthly_waste
        elif finding.status == PCStatus.OVERSIZED:
            result.oversized_pc_count += 1
            result.potential_savings += finding.monthly_savings
        elif finding.status == PCStatus.UNDERSIZED:
            result.undersized_pc_count += 1

    return result


def _analyze_single_pc(func: LambdaPCInfo, region: str) -> PCFinding:
    """개별 함수 PC 분석"""

    # PC 미설정
    if not func.has_pc:
        # Throttle 발생했으면 PC 설정 권장
        if func.throttles_30d > 0:
            recommended = max(1, int(func.max_concurrent * 1.2))
            return PCFinding(
                function=func,
                status=PCStatus.UNDERSIZED,
                recommendation=f"Throttle {func.throttles_30d:,}회 발생 - PC {recommended}개 설정 권장",
                recommended_pc=recommended,
            )
        return PCFinding(
            function=func,
            status=PCStatus.NO_PC,
            recommendation="PC 미설정",
        )

    # 미사용 PC
    if func.invocations_30d == 0:
        return PCFinding(
            function=func,
            status=PCStatus.UNUSED,
            recommendation=f"미사용 함수에 PC {func.total_provisioned}개 설정됨 - PC 해제 권장",
            monthly_waste=func.monthly_cost,
        )

    # 활용률 분석
    utilization = func.utilization_pct

    # 과다 설정 (활용률 30% 미만)
    if utilization < 30 and func.max_concurrent < func.total_provisioned:
        recommended = max(1, int(func.max_concurrent * 1.3))  # 30% 여유
        current_cost = func.monthly_cost
        new_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=recommended,
        )
        savings = current_cost - new_cost

        return PCFinding(
            function=func,
            status=PCStatus.OVERSIZED,
            recommendation=f"PC 과다 설정 (활용률 {utilization:.0f}%) - {recommended}개로 축소 권장",
            recommended_pc=recommended,
            monthly_savings=savings,
        )

    # 부족 (Throttle 발생)
    if func.throttles_30d > 0:
        recommended = max(func.total_provisioned, int(func.max_concurrent * 1.2))
        return PCFinding(
            function=func,
            status=PCStatus.UNDERSIZED,
            recommendation=f"Throttle {func.throttles_30d:,}회 발생 - PC {recommended}개로 증가 권장",
            recommended_pc=recommended,
        )

    # 적정
    return PCFinding(
        function=func,
        status=PCStatus.OPTIMAL,
        recommendation=f"적정 (활용률 {utilization:.0f}%)",
    )


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[PCAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    status_fills = {
        PCStatus.UNUSED: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        PCStatus.OVERSIZED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        PCStatus.UNDERSIZED: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        PCStatus.OPTIMAL: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Lambda Provisioned Concurrency 분석"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = [
        "Account",
        "Region",
        "전체 함수",
        "PC 설정",
        "미사용 PC",
        "과다 설정",
        "부족",
        "PC 월 비용",
        "절감 가능",
    ]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_functions)
        ws.cell(row=row, column=4, value=r.functions_with_pc)
        ws.cell(row=row, column=5, value=r.unused_pc_count)
        ws.cell(row=row, column=6, value=r.oversized_pc_count)
        ws.cell(row=row, column=7, value=r.undersized_pc_count)
        ws.cell(row=row, column=8, value=f"${r.total_pc_cost:,.2f}")
        ws.cell(row=row, column=9, value=f"${r.potential_savings:,.2f}")

    # PC Functions
    ws_pc = wb.create_sheet("PC Functions")
    pc_headers = [
        "Account",
        "Region",
        "Function",
        "Runtime",
        "Memory",
        "PC 설정",
        "최대 동시성",
        "활용률",
        "Throttles",
        "월 비용",
        "상태",
        "권장 PC",
        "절감 가능",
        "권장 조치",
    ]
    for col, h in enumerate(pc_headers, 1):
        ws_pc.cell(row=1, column=col, value=h).fill = header_fill
        ws_pc.cell(row=1, column=col).font = header_font

    pc_row = 1
    for r in results:
        for f in r.findings:
            if f.function.has_pc or f.status == PCStatus.UNDERSIZED:
                pc_row += 1
                fn = f.function
                ws_pc.cell(row=pc_row, column=1, value=fn.account_name)
                ws_pc.cell(row=pc_row, column=2, value=fn.region)
                ws_pc.cell(row=pc_row, column=3, value=fn.function_name)
                ws_pc.cell(row=pc_row, column=4, value=fn.runtime)
                ws_pc.cell(row=pc_row, column=5, value=fn.memory_mb)
                ws_pc.cell(row=pc_row, column=6, value=fn.total_provisioned)
                ws_pc.cell(row=pc_row, column=7, value=fn.max_concurrent)
                ws_pc.cell(row=pc_row, column=8, value=f"{fn.utilization_pct:.0f}%")
                ws_pc.cell(row=pc_row, column=9, value=fn.throttles_30d)
                ws_pc.cell(row=pc_row, column=10, value=f"${fn.monthly_cost:,.2f}")
                ws_pc.cell(row=pc_row, column=11, value=f.status.value)
                ws_pc.cell(
                    row=pc_row,
                    column=12,
                    value=f.recommended_pc if f.recommended_pc else "-",
                )
                savings = f.monthly_waste + f.monthly_savings
                ws_pc.cell(
                    row=pc_row,
                    column=13,
                    value=f"${savings:,.2f}" if savings > 0 else "-",
                )
                ws_pc.cell(row=pc_row, column=14, value=f.recommendation)

                fill = status_fills.get(f.status)
                if fill:
                    ws_pc.cell(row=pc_row, column=11).fill = fill

    # 열 너비
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Lambda_PC_Analysis_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> PCAnalysisResult | None:
    """단일 계정/리전의 PC 정보 수집 및 분석 (병렬 실행용)"""
    functions = collect_pc_info(session, account_id, account_name, region)
    if not functions:
        return None
    return analyze_pc(functions, account_id, account_name, region)


def run(ctx) -> None:
    """PC 분석 실행"""
    console.print("[bold]Lambda Provisioned Concurrency 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")
    results: list[PCAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_with_pc = sum(r.functions_with_pc for r in results)
    total_unused = sum(r.unused_pc_count for r in results)
    total_oversized = sum(r.oversized_pc_count for r in results)
    total_undersized = sum(r.undersized_pc_count for r in results)
    total_cost = sum(r.total_pc_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"PC 설정된 함수: {total_with_pc}개")
    console.print(f"총 PC 비용: ${total_cost:,.2f}/월")

    if total_unused > 0:
        console.print(f"[red]미사용 PC: {total_unused}개[/red]")
    if total_oversized > 0:
        console.print(f"[yellow]과다 설정: {total_oversized}개[/yellow]")
    if total_undersized > 0:
        console.print(f"[yellow]부족 (Throttle): {total_undersized}개[/yellow]")
    if total_savings > 0:
        console.print(f"[green]절감 가능: ${total_savings:,.2f}/월[/green]")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("lambda-provisioned").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
