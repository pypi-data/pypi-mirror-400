"""Excel sheet generation."""

from typing import List, Optional, Tuple

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from ragraph.graph import Graph
from ragraph.node import Node

from raesl.doc import utils as doc_utils
from raesl.doc.lines import node_path, var_path
from raesl.doc.sections import get_var_table_row_elements
from raesl.excel import text, utils
from raesl.excel.defaults import OPTIONS


def add_components_sheet(wb: Workbook, components: List[Node]) -> Tuple[Worksheet, List[Node]]:
    """Add a components overview sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the components sheet to.
        components: List of component nodes.

    Returns:
        Components worksheet instance.
    """
    ws = wb.create_sheet("Components")

    # Handle headers.
    tags = utils.get_all_tags(components)
    headers = [
        "instance path",
        "component definition",
        "parent component",
        "comments",
    ] + tags
    ws.append(headers)

    # Handle content.
    def write_component_row(ws: Worksheet, tags: List[str], component: Node):
        """Write a component sheet row."""
        info = component.annotations.esl_info
        row = [
            node_path(component.name, italic=False, arrows=False, skip="world"),
            info["definition_name"],
            (
                node_path(component.parent.name, italic=False, arrows=False, skip="world")
                if component.parent
                else None
            ),
            utils.format_multiline(info["comments"]),
        ]
        for tag in tags:
            comments = info["tagged_comments"].get(tag, [])
            row.append(utils.format_multiline(comments))
        ws.append(row)

    for component in components:
        write_component_row(ws, tags, component)

    # Handle styling.
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    utils.make_table(
        ws,
        name="components",
        min_row=1,
        max_row=len(components) + 1,
        min_col=1,
        max_col=len(headers),
    )
    return ws, components


def add_goals_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a goal requirements sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the goals sheet to.
        graph: Graph to fetch goals from.
        components: Component nodes to fetch goals for.

    Returns:
        Goal requirements worksheet instance.
    """
    ws = wb.create_sheet("Goals")

    # Get requirements.
    requirements = utils.dedupe(
        [
            r
            for c in components
            for r in doc_utils.get_component_goals(c, graph, constraint=False, inherited=False)
        ]
        + [
            r
            for c in components
            for r in doc_utils.get_component_goals(c, graph, constraint=False, inherited=True)
        ]
        + [
            r
            for c in components
            for r in doc_utils.get_component_goals(c, graph, constraint=True, inherited=False)
        ]
        + [
            r
            for c in components
            for r in doc_utils.get_component_goals(c, graph, constraint=True, inherited=True)
        ]
    )

    # Handle headers.
    default_headers = [
        "instance path",
        "component definition",
        "form",
        "source component",
        "auxiliary",
        "verb",
        "variables",
        "preposition",
        "target component",
        "subclauses",
        "comments",
    ]
    tags = utils.get_all_tags(requirements)
    headers = default_headers + tags
    ws.append(headers)

    # Handle content.
    for requirement in requirements:
        info = requirement.annotations.esl_info
        body = info["body"]

        default_content = [
            node_path(requirement.name, italic=False, arrows=False, skip="world"),
            utils.parent_def(graph, requirement),
            info["form"],
            node_path(body["active"], italic=False, arrows=False, skip="world"),
            body["auxiliary"],
            body["verb"],
            ", ".join(
                var_path(
                    graph[var],
                    italic=False,
                    arrows=False,
                    skip="world",
                )
                for var in body["variables"]
            ),
            body["preposition"],
            node_path(body["passive"], italic=False, arrows=False, skip="world"),
            text.subclauses_text(requirement, graph, skip="world", spaces=0),
            utils.format_multiline(info["comments"]),
        ]
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        ws.append(default_content + tagged_content)

    # Handle styling.
    utils.make_table(
        ws,
        "goals",
        min_row=1,
        max_row=1 + len(requirements),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    return ws, requirements


def add_transformations_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a transformation requirements sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the transformations sheet to.
        graph: Graph to fetch transformation nodes from.
        components: Component nodes to fetch transformations for.

    Returns:
        Transformation requirements worksheet instance.
    """
    ws = wb.create_sheet("Transformations")

    # Select requirements.
    requirements = utils.dedupe(
        [
            r
            for c in components
            for r in doc_utils.get_component_transformations(c, graph, constraint=False)
        ]
        + [
            r
            for c in components
            for r in doc_utils.get_component_transformations(c, graph, constraint=True)
        ]
    )

    # Handle headers.
    default_headers = [
        "instance path",
        "component definition",
        "form",
        "auxiliary",
        "verb",
        "input_variables",
        "preposition",
        "output_variables",
        "subclauses",
        "comments",
    ]
    tags = utils.get_all_tags(requirements)
    headers = default_headers + tags
    ws.append(headers)

    for requirement in requirements:
        info = requirement.annotations.esl_info
        body = info["body"]

        default_content = [
            node_path(requirement.name, italic=False, arrows=False, skip="world"),
            utils.parent_def(graph, requirement),
            info["form"],
            body["auxiliary"],
            body["verb"],
            ", ".join(
                var_path(graph[var], italic=False, arrows=False, skip="world")
                for var in body["input_variables"]
            ),
            body["preposition"],
            ", ".join(
                var_path(graph[var], italic=False, arrows=False, skip="world")
                for var in body["output_variables"]
            ),
            text.subclauses_text(requirement, graph, skip="world", spaces=0),
            utils.format_multiline(info["comments"]),
        ]
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        ws.append(default_content + tagged_content)

    # Handle styling.
    utils.make_table(
        ws,
        "tranformations",
        min_row=1,
        max_row=1 + len(requirements),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)

    return ws, requirements


def add_designs_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a design requirements sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the design requirements sheet to.
        graph: Graph to fetch designrule nodes from.
        components: Component nodes to fetch design requirements for.

    Returns:
        Design requirements worksheet instance.
    """
    ws = wb.create_sheet("Design requirements")

    # Select requirements.
    requirements = utils.dedupe(
        [r for c in components for r in doc_utils.get_component_designs(c, graph, constraint=False)]
        + [
            r
            for c in components
            for r in doc_utils.get_component_designs(c, graph, constraint=True)
        ]
    )

    # Handle headers.
    default_headers = [
        "instance path",
        "component definition",
        "form",
        "subject",
        "auxiliary",
        "comparison",
        "bound",
        "subclauses",
        "comments",
    ]
    tags = utils.get_all_tags(requirements)
    headers = default_headers + tags
    ws.append(headers)
    rows = 1

    # Handle content.
    for requirement in requirements:
        info = requirement.annotations.esl_info
        for body in info["body"]:
            default_content = [
                node_path(requirement.name, italic=False, arrows=False, skip="world"),
                utils.parent_def(graph, requirement),
                info["form"],
                var_path(graph[body["subject"]], italic=False, arrows=False, skip="world"),
                "{}{}".format("EITHER " if len(info["body"]) > 1 else "", body["auxiliary"]),
                body["comparison"],
                "{} {}".format(body["bound"]["value"], body["bound"]["unit"]),
                text.subclauses_text(requirement, graph, skip="world", spaces=0),
                utils.format_multiline(info["comments"]),
            ]
            tagged_content = [
                utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
            ]
            ws.append(default_content + tagged_content)
            rows += 1

    # Handle styling.
    utils.make_table(
        ws,
        "designrequirements",
        min_row=1,
        max_row=rows,
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    return ws, requirements


def add_behaviors_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a behavior requirements sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the behavior requirements sheet to.
        graph: Graph to fetch behavior nodes from.
        components: Component nodes to fetch behavior requirements for.

    Returns:
        Behavior requirements worksheet instance.
    """
    ws = wb.create_sheet("Behavior requirements")

    # Select requirements.
    requirements = utils.dedupe(
        [
            r
            for c in components
            for r in doc_utils.get_component_behaviors(c, graph, constraint=False)
        ]
        + [
            r
            for c in components
            for r in doc_utils.get_component_behaviors(c, graph, constraint=True)
        ]
    )

    # Handle headers.
    default_headers = [
        "instance path",
        "component definition",
        "form",
        "case",
        "when",
        "then",
        "comments",
    ]
    tags = utils.get_all_tags(requirements)
    headers = default_headers + tags
    ws.append(headers)
    rows = 1

    # Handle content.
    for requirement in requirements:
        info = requirement.annotations.esl_info
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        if info["default"]:
            default_content = [
                node_path(requirement.name, italic=False, arrows=False, skip="world"),
                utils.parent_def(graph, requirement),
                info["form"],
                "default",
                "no other case applies",
                "\n".join(
                    text.designclause_text(clause["body"], graph, label=clause["name"])
                    for clause in info["default"]
                ),
                utils.format_multiline(info["comments"]),
            ]
            ws.append(default_content + tagged_content)
            rows += 1
        for case in info["cases"]:
            default_content = [
                node_path(requirement.name, italic=False, arrows=False, skip="world"),
                utils.parent_def(graph, requirement),
                info["form"],
                case["name"],
                "\n".join(
                    text.designclause_text(clause["body"], graph, label=clause["name"])
                    for clause in case["when_clauses"]
                ),
                "\n".join(
                    text.designclause_text(clause["body"], graph, label=clause["name"])
                    for clause in case["then_clauses"]
                ),
                utils.format_multiline(info["comments"]),
            ]
            ws.append(default_content + tagged_content)
            rows += 1

    # Handle styling.
    utils.make_table(
        ws,
        "behaviors",
        min_row=1,
        max_row=rows,
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    return ws, requirements


def add_needs_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a needs sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the needs sheet to.
        graph: Graph to fetch need nodes from.
        components: Component nodes to fetch needs for.

    Returns:
        Needs worksheet instance.
    """
    ws = wb.create_sheet("Needs")

    # Select requirements.
    requirements = utils.dedupe(
        r for c in components for r in doc_utils.get_component_needs(c, graph)
    )

    # Handle headers.
    default_headers = [
        "instance path",
        "component definition",
        "subject",
        "text",
    ]
    tags = utils.get_all_tags(requirements)
    headers = default_headers + tags
    ws.append(headers)

    # Handle content.
    for requirement in requirements:
        info = requirement.annotations.esl_info
        default_content = [
            requirement.name.split(".")[-1],
            utils.parent_def(graph, requirement),
            var_path(graph[info["subject"]], italic=False, arrows=False, skip="world"),
            info["text"],
            utils.format_multiline(info["comments"]),
        ]
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        ws.append(default_content + tagged_content)

    # Handle styling.
    utils.make_table(
        ws,
        "needs",
        min_row=1,
        max_row=1 + len(requirements),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    return ws, requirements


def add_variable_sheet(
    wb: Workbook, graph: Graph, components: List[Node]
) -> Tuple[Worksheet, List[Node]]:
    """Add a variable sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the needs sheet to.
        graph: Graph to fetch need nodes from.
        components: Component nodes to fetch needs for.

    Returns:
        Needs worksheet instance.
    """
    ws = wb.create_sheet("Variables")

    # Variables
    vrs = [n for n in graph.nodes if n.kind == "variable"]
    vrs = sorted(vrs, key=lambda x: x.name)

    # Handle headers.
    default_headers = ["instance path", "type", "domain", "units", "clarifaction"]
    tags = utils.get_all_tags(vrs)
    headers = default_headers + tags
    ws.append(headers)

    # Handle content.
    for var in vrs:
        info = var.annotations.esl_info
        default_content = list(get_var_table_row_elements(graph, var))
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        ws.append(default_content + tagged_content)

    # Handle styling.
    utils.make_table(
        ws,
        "variables",
        min_row=1,
        max_row=1 + len(vrs),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)
    return ws, vrs


def add_overview_sheet(
    wb: Workbook,
    graph: Graph,
    components: List[Node],
    goals: List[Node],
    transformations: List[Node],
    designs: List[Node],
    behaviors: List[Node],
    needs: List[Node],
) -> Worksheet:
    """Add an overview sheet to an Excel workbook.

    Arguments:
        wb: Excel workbook to add the needs sheet to.
        components: Component nodes.
        goals: Goal requirement nodes.
        transformations: Transformation requirement nodes.
        designs: Design requirement nodes.
        behaviors: Behavior requirement nodes.
        needs: Need nodes.

    Returns:
        Overview worksheet instance.
    """
    ws = wb.create_sheet("Overview", index=0)

    # Select requirements.
    requirements = goals + transformations + designs + behaviors + needs

    # Handle headers.
    default_headers = [
        "instance name",
        "specification text",
        "component path",
        "component definition",
        "kind",
        "form",
        "comments",
    ]
    tags = utils.get_all_tags(components + requirements)
    headers = default_headers + tags
    ws.append(headers)

    # Handle content.
    for requirement in requirements:
        info = requirement.annotations.esl_info
        skip = utils.parent_component(requirement, skip=None)
        default_content = [
            requirement.name.split(".")[-1],
            text.requirement_text(requirement, graph, skip=skip),
            utils.parent_component(requirement, skip="world"),
            utils.parent_def(graph, requirement),
            utils.requirement_kind(requirement),
            None if requirement.kind == "need" else info["form"],
            utils.format_multiline(info["comments"]),
        ]
        tagged_content = [
            utils.format_multiline(info["tagged_comments"].get(tag, [])) for tag in tags
        ]
        ws.append(default_content + tagged_content)

    utils.make_table(
        ws,
        "overview",
        min_row=1,
        max_row=1 + len(requirements),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS)

    return ws


def add_component_active_goals_sheet(
    wb: Workbook,
    graph: Graph,
    component: Node,
    flow_labels: Optional[List[str]] = None,
):
    ws_name = "Outgoing"
    ws = wb.create_sheet(ws_name)

    top_headers = ["instance path", "instance name"]
    ws.append(top_headers)
    split = component.name.rfind(".")
    path = component.name[:split]
    instance = component.name[split + 1 :]
    ws.append([path, instance])
    ws.append([])

    utils.make_table(
        ws,
        "outgoing-component-path",
        min_row=1,
        max_row=2,
        min_col=1,
        max_col=len(top_headers),
    )
    utils.apply_styling(ws, top_headers, defaults=OPTIONS, start_row=0, end_row=2)

    flows = set(flow_labels) if flow_labels else None

    functional_dependencies = utils.yield_functional_dependencies(graph, component, flows)

    function_specifications = set(
        [f for f in utils.yield_all_function_specifications(graph, functional_dependencies)]
    )

    active_goals = set(
        [(e, g) for (e, g) in utils.yield_active_functions(function_specifications, component.name)]
    )

    goalsdata = [
        data
        for (e, goal) in active_goals
        for data in utils.build_active_goal_data(graph, e, goal, function_specifications)
    ]

    headers = [
        "label",
        "target",
        "flows",
        "types",
        "subclause",
        "subject",
        "comparison",
        "bound",
        "unit",
    ]
    ws.append(headers)

    for data in goalsdata:
        ws.append([data.get(h, "") for h in headers])

    utils.make_table(
        ws,
        ws_name,
        min_row=4,
        max_row=4 + len(goalsdata),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS, start_row=2, end_row=None)

    return ws


def add_component_passive_goals_sheet(
    wb: Workbook,
    graph: Graph,
    component: Node,
    flow_labels: Optional[List[str]] = None,
):
    ws_name = "Incoming"
    ws = wb.create_sheet(ws_name)

    top_headers = ["instance path", "instance name"]
    ws.append(top_headers)
    split = component.name.rfind(".")
    path = component.name[:split]
    instance = component.name[split + 1 :]
    ws.append([path, instance])
    ws.append([])

    utils.make_table(
        ws,
        "incoming-component-path",
        min_row=1,
        max_row=2,
        min_col=1,
        max_col=len(top_headers),
    )
    utils.apply_styling(ws, top_headers, defaults=OPTIONS, start_row=0, end_row=2)

    flows = set(flow_labels) if flow_labels else None

    functional_dependencies = utils.yield_functional_dependencies(graph, component, flows)

    function_specifications = utils.yield_all_function_specifications(
        graph, functional_dependencies
    )

    passive_goals = set(
        [
            (e, g)
            for (e, g) in utils.yield_passive_functions(function_specifications, component.name)
        ]
    )

    goalsdata = [
        data
        for (e, goal) in passive_goals
        for data in utils.build_passive_goal_data(graph, e, goal, function_specifications)
    ]

    headers = [
        "label",
        "source",
        "flows",
        "types",
        "subclause",
        "subject",
        "comparison",
        "bound",
        "unit",
    ]
    ws.append(headers)

    for data in goalsdata:
        ws.append([data.get(h, "") for h in headers])

    utils.make_table(
        ws,
        ws_name,
        min_row=4,
        max_row=4 + len(goalsdata),
        min_col=1,
        max_col=len(headers),
    )
    utils.apply_styling(ws, headers, defaults=OPTIONS, start_row=2, end_row=None)

    return ws
