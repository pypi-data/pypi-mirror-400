from openpyxl import load_workbook, Workbook
import os

'''
@time: 2019-01-12
@auth: fs@file: ParseExcel.py
@IDE: pycharm
'''


class ReadExcel:
    def __init__(self, *filename):
        """获取api测试用例excel文件地址,当前是././mock_data/XXX.xlsx文件"""
        ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(os.getcwd())))
        self.excel_path = os.path.join(ROOT_DIR, *filename)
        print(f"读取excel文件路径：{self.excel_path}")
        if not os.path.exists(self.excel_path):  # 不存在则只退回一级
            ROOT_DIR = os.path.dirname(os.path.abspath(os.getcwd()))
            self.excel_path = os.sep.join([ROOT_DIR])
            for i in range(len(filename)):
                self.excel_path = os.sep.join([self.excel_path, filename[i]])
            if not os.path.exists(self.excel_path):  # 不存在则标识为根目录
                ROOT_DIR = os.path.abspath(os.getcwd())
                self.excel_path = os.sep.join([ROOT_DIR])
                for i in range(len(filename)):
                    self.excel_path = os.sep.join([self.excel_path, filename[i]])

    def read_excel_data(self, *sheet_name):
        """调用AnalysisExcel类的方法解析excel数据"""
        excel = AnalysisExcel(self.excel_path)
        print(self.excel_path)
        try:
            for i in range(len(sheet_name)):
                sheet = excel.get_sheet_by_name(sheet_name[i])
                test_data = excel.get_all_values_of_sheet(sheet)
        finally:
            excel.close()
        return test_data


class AnalysisExcel:
    def __init__(self, excel_path):
        """读取excel文件"""
        self.wk = load_workbook(excel_path)
        self.excelFile = excel_path

    def get_sheet_by_name(self, sheet_name):
        """获取excel的sheet对象"""
        sheet = self.wk[sheet_name]
        return sheet

    @staticmethod
    def get_row_num(sheet):
        """获取有效数据的最大行号"""
        return sheet.max_row

    @staticmethod
    def get_cols_num(sheet):
        """获取有效数据的最大列号"""
        return sheet.max_column

    def get_row_values(self, sheet, row_num):
        """获取某一行的数据"""
        max_row_num = self.get_cols_num(sheet)
        row_values = []
        for cols_num in range(1, max_row_num + 1):
            value = sheet.cell(row_num, cols_num).value
            if value is None:
                value = ''
            row_values.append(value)
        return tuple(row_values)

    def get_col_value(self, sheet, column_num):
        """获取某一列的数据"""
        max_row_num = self.get_row_num(sheet)
        column_values = []
        for row in range(2, max_row_num+1):
            value = sheet.cell(row, column_num).value
            if value is None:
                value = ''
            column_values.append(value)
        return tuple(column_values)

    def get_value_of_cell(self, sheet, row_num, column_num):
        value = sheet.cell(row_num, column_num).value
        if value is None:
            value = ''
        return value

    def get_all_values_of_sheet(self, sheet):
        """获取某一个sheet页的所有测试数据，返回一个元祖组成的列表"""
        max_row_num = self.get_row_num(sheet)
        column_num = self.get_cols_num(sheet)
        all_values = []
        for row in range(2, max_row_num + 1):
            row_values = []
            for column in range(1, column_num + 1):
                value = sheet.cell(row, column).value
                if value is None:
                    value = ''
                row_values.append(value)
            all_values.append(tuple(row_values))
        return all_values

    def close(self):
        self.wk.close()


class WriteExcel:
    def __init__(self, file_name):
        """获取api测试用例excel文件地址,当前是././mock_data/XXX.xlsx文件"""
        self.case_file_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
        self.case_file_dir = os.path.join(self.case_file_path + '\\', 'mock_data\\' + file_name)

    # 新建excel
    def creatwb(self):
        wb = Workbook()
        wb.save(filename=self.case_file_dir)
        print("新建Excel：" + self.case_file_dir + "成功")

    # 写入excel文件中 date 数据，date是list数据类型， fields 表头
    def savetoexcel(self, data, fields, sheetname):
        print("写入excel")
        wb = load_workbook(filename=self.case_file_dir)

        sheet = wb.active
        sheet.title = sheetname

        field = 1
        for field in range(1, len(fields) + 1):  # 写入表头
            sheet.cell(row=1, column=field, value=str(fields[field - 1]))

        row1 = 1
        col1 = 0
        for row1 in range(2, len(data) + 2):  # 写入数据
            for col1 in range(1, len(data[row1 - 2]) + 1):
                sheet.cell(row=row1, column=col1, value=str(data[row1 - 2][col1 - 1]))
                print("当前是第{}行，第{}列，数据是{}".format(row1, col1, str(data[row1 - 2][col1 - 1])))
        print(len(data))
        wb.save(filename=self.case_file_dir)


if __name__ == '__main__':
    # file_path = os.sep.join(['data', 'xcj', 'Itemexcutebid', '直接录入采购结果.xlsx'])  # 拼接xlsx完整路径
    rd = ReadExcel('data', 'xcj', '报表数据源1.xlsx').read_excel_data('refresh_put')
    # a = rd.read_excel_data('create')
    print(rd)
