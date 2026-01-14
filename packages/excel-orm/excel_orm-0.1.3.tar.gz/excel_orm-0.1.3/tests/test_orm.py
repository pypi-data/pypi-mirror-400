from __future__ import annotations

from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from excel_orm import Column, ExcelFile, PivotSheetSpec, text_column
from excel_orm.orm import SheetSpec


def test_generate_template_creates_sheet_and_titles_and_headers(tmp_path, excel_file):
    out = tmp_path / "template.xlsx"
    excel_file.generate_template(str(out))

    wb = load_workbook(out)
    assert "Cars" in wb.sheetnames
    ws = wb["Cars"]

    # Cars block: title row merged across A..C, title value "Cars"
    assert ws["A1"].value == "Cars"
    merged = [str(rng) for rng in ws.merged_cells.ranges]
    assert "A1:C1" in merged

    # Cars headers on row 2
    assert ws["A2"].value == "Make"
    assert ws["B2"].value == "Model"
    assert ws["C2"].value == "Year"

    # ManufacturingPlant block starts after 2-col gap: Cars ends at C, so start at F
    assert ws["F1"].value == "Manufacturing Plants"
    assert "F1:G1" in merged
    assert ws["F2"].value == "Factory Name"
    assert ws["G2"].value == "Location"


def test_load_data_end_to_end(tmp_path, excel_file):
    """
    Create a workbook that matches the template layout, fill a few rows,
    load it, and verify repositories.
    """
    out = tmp_path / "data.xlsx"
    excel_file.generate_template(str(out))

    # Fill data
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Cars"]

    # Cars data at row 3+
    ws["A3"].value = "Toyota"
    ws["B3"].value = "Camry"
    ws["C3"].value = 2020

    ws["A4"].value = "Honda"
    ws["B4"].value = "Civic"
    ws["C4"].value = "2019"  # string should parse to int

    # Plants data at row 3+ (F,G)
    ws["F3"].value = "Plant 1"
    ws["G3"].value = "NJ"

    ws["F4"].value = "Plant 2"
    ws["G4"].value = "PA"

    wb.save(out)

    # Load
    excel_file.load_data(str(out))

    cars = excel_file.cars.all()
    plants = excel_file.manufacturing_plants.all()

    assert len(cars) == 2
    assert cars[0].make == "Toyota"
    assert cars[0].model == "Camry"
    assert cars[0].year == 2020
    assert cars[1].year == 2019

    assert len(plants) == 2
    assert plants[0].name == "Plant 1"
    assert plants[0].location == "NJ"


def test_load_data_missing_sheet_raises(tmp_path, excel_file):
    # Create a workbook with a different sheet name
    from openpyxl import Workbook

    p = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "NotCars"
    wb.save(p)

    with pytest.raises(ValueError):
        excel_file.load_data(str(p))


def _as_date(x) -> date:
    """openpyxl may return date or datetime depending on formatting; normalize."""
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    raise TypeError(f"Expected date/datetime, got {type(x)}: {x!r}")


def test_generate_template_pivot_sheet_layout(tmp_path, pivot_excel_file, demand_model):
    out = tmp_path / "pivot_template.xlsx"
    pivot_excel_file.generate_template(str(out))

    wb = load_workbook(out)
    assert "Demand" in wb.sheetnames
    ws = wb["Demand"]

    # Title merged across A..D (A for row header + 3 pivot columns -> B,C,D)
    assert ws["A1"].value == "Demands"
    merged = [str(rng) for rng in ws.merged_cells.ranges]
    assert "A1:D1" in merged

    # Corner header for row keys (Region)
    assert ws["A2"].value in {"Region"}  # explicit

    # Pivot headers across B2..D2
    assert _as_date(ws["B2"].value) == date(2025, 6, 1)
    assert _as_date(ws["C2"].value) == date(2025, 7, 1)
    assert _as_date(ws["D2"].value) == date(2025, 8, 1)

    # Seeded row values appear in A3, A4
    assert ws["A3"].value == "NA"
    assert ws["A4"].value == "EU"


def test_load_data_pivot_end_to_end_skips_blank_cells(tmp_path, pivot_excel_file):
    """
    Fill a demand pivot matrix:
      rows: NA, EU
      cols: Jun, Jul, Aug
    Leave one cell blank and verify it is not emitted (include_blanks=False).
    """
    out = tmp_path / "pivot_data.xlsx"
    pivot_excel_file.generate_template(str(out))

    wb = load_workbook(out)
    ws = wb["Demand"]

    # NA row at row 3
    ws["B3"].value = 10  # NA, Jun
    ws["C3"].value = 20  # NA, Jul
    ws["D3"].value = None  # NA, Aug (blank -> should skip)

    # EU row at row 4
    ws["B4"].value = 5  # EU, Jun
    ws["C4"].value = 0  # EU, Jul (explicit 0 should be included)
    ws["D4"].value = 15  # EU, Aug

    wb.save(out)

    pivot_excel_file.load_data(str(out))

    # Repo name is plural snake_case of Demand -> demands
    rows = pivot_excel_file.demands.all()
    assert len(rows) == 5  # 6 cells minus 1 blank = 5

    # Compare as a set of tuples to avoid ordering assumptions
    got = {(r.region, r.dt, r.value) for r in rows}
    expected = {
        ("NA", date(2025, 6, 1), 10),
        ("NA", date(2025, 7, 1), 20),
        # ("NA", date(2025, 8, 1), ?) skipped
        ("EU", date(2025, 6, 1), 5),
        ("EU", date(2025, 7, 1), 0),
        ("EU", date(2025, 8, 1), 15),
    }
    assert got == expected


def test_load_data_pivot_stops_at_blank_region_row(tmp_path, demand_model):
    """
    If the region cell is blank, parsing should stop (contiguous-block rule for row labels).
    """
    pivot_values = [date(2025, 6, 1), date(2025, 7, 1)]
    spec = PivotSheetSpec(
        name="Demand",
        model=demand_model,
        pivot_field="dt",
        row_fields=["region"],
        value_field="value",
        pivot_values=pivot_values,
        row_values=None,  # user-entered regions
    )
    xf = ExcelFile(sheets=[spec])

    out = tmp_path / "stop_rule.xlsx"
    xf.generate_template(str(out))

    wb = load_workbook(out)
    ws = wb["Demand"]

    # Row 3 has region, row 4 is blank region -> stop; row 5 should not be read
    ws["A3"].value = "NA"
    ws["B3"].value = 1
    ws["C3"].value = 2

    ws["A4"].value = ""  # stop condition

    ws["A5"].value = "EU"
    ws["B5"].value = 9
    ws["C5"].value = 9

    wb.save(out)

    xf.load_data(str(out))
    rows = xf.demands.all()  # ty:ignore[unresolved-attribute]

    got = {(r.region, r.dt, r.value) for r in rows}
    assert got == {
        ("NA", date(2025, 6, 1), 1),
        ("NA", date(2025, 7, 1), 2),
    }


def test_load_data_missing_pivot_sheet_raises(tmp_path, pivot_excel_file):
    # Create a workbook with a different sheet name
    from openpyxl import Workbook

    p = tmp_path / "wrong.xlsx"
    wb = Workbook()
    wb.active.title = "NotDemand"
    wb.save(p)

    with pytest.raises(ValueError):
        pivot_excel_file.load_data(str(p))


def test_pivot_two_row_fields_template_and_parse_end_to_end(tmp_path, demand_two_pivot):
    """
    Two row_fields case:
      row_fields: [region, product]
      pivot cols: Jun, Jul
      row_values seeded with composite keys (NA/ABC, EU/ABC)
    Validate:
      - template layout places row headers in A2,B2
      - pivot headers start at C2,D2
      - seeded keys appear in A3,B3 and A4,B4
      - parsing emits correct objects with BOTH row fields populated
    """
    pivot_values = [date(2025, 6, 1), date(2025, 7, 1)]
    spec = PivotSheetSpec(
        name="Demand",
        model=demand_two_pivot,
        pivot_field="dt",
        row_fields=["region", "product"],
        value_field="value",
        pivot_values=pivot_values,
        row_values=[("NA", "ABC"), ("EU", "ABC")],
        row_header_col=1,
        include_blanks=False,
    )
    xf = ExcelFile(sheets=[spec])

    out = tmp_path / "pivot_two_row_fields.xlsx"
    xf.generate_template(str(out))

    wb = load_workbook(out)
    ws = wb["Demand"]

    # Title merged across A..D (A,B row headers + C,D pivot headers)
    assert ws["A1"].value == "Demands"
    merged = [str(rng) for rng in ws.merged_cells.ranges]
    assert "A1:D1" in merged

    # Row field headers in A2,B2
    assert ws["A2"].value == "Region"
    assert ws["B2"].value == "Product"

    # Pivot headers start at C2..D2
    assert _as_date(ws["C2"].value) == date(2025, 6, 1)
    assert _as_date(ws["D2"].value) == date(2025, 7, 1)

    # Seeded composite keys appear across A/B
    assert ws["A3"].value == "NA"
    assert ws["B3"].value == "ABC"
    assert ws["A4"].value == "EU"
    assert ws["B4"].value == "ABC"

    # Fill matrix values: (row 3: NA/ABC), (row 4: EU/ABC)
    ws["C3"].value = 10  # NA/ABC, Jun
    ws["D3"].value = 20  # NA/ABC, Jul
    ws["C4"].value = 5  # EU/ABC, Jun
    ws["D4"].value = 0  # EU/ABC, Jul

    wb.save(out)

    xf.load_data(str(out))
    rows = xf.demands.all()  # repo name plural of Demand

    # Expect 4 objects (2 rows x 2 pivots)
    assert len(rows) == 4

    got = {(r.region, r.product, r.dt, r.value) for r in rows}
    expected = {
        ("NA", "ABC", date(2025, 6, 1), 10),
        ("NA", "ABC", date(2025, 7, 1), 20),
        ("EU", "ABC", date(2025, 6, 1), 5),
        ("EU", "ABC", date(2025, 7, 1), 0),
    }
    assert got == expected


def test_duplicate_header_names_do_not_clash(tmp_path, models):
    Car, _ = models

    class ManufacturingPlant:
        name: Column[str] = text_column(header="Make", not_null=True)
        location: Column[str] = text_column(header="location")

    sheet = SheetSpec(name="Cars", models=[Car, ManufacturingPlant])
    file = ExcelFile(sheets=[sheet])
    out = tmp_path / "duplicate_headers.xlsx"
    file.generate_template(out)

    wb = load_workbook(out)
    ws = wb["Cars"]

    ws["A3"].value = "Acura"
    ws["B3"].value = "TSX"
    ws["C3"].value = 2005

    ws["F3"].value = "Duplicate"
    ws["G3"].value = "New Jersey"

    wb.save(out)

    file.load_data(out)

    cars = file.cars.all()
    plants = file.manufacturing_plants.all()

    assert plants[0].name == "Duplicate"
    assert plants[0].location == "New Jersey"
