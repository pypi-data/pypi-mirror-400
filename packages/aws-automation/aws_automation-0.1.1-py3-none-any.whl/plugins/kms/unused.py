"""
plugins/kms/unused.py - KMS 키 미사용 분석

미사용/비활성화 고객 관리 키 (CMK) 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_kms_key_price

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "kms:ListKeys",
        "kms:DescribeKey",
        "kms:ListAliases",
    ],
}


class KMSKeyStatus(Enum):
    """KMS 키 상태"""

    NORMAL = "normal"
    DISABLED = "disabled"
    PENDING_DELETE = "pending_delete"


@dataclass
class KMSKeyInfo:
    """KMS 키 정보"""

    account_id: str
    account_name: str
    region: str
    key_id: str
    arn: str
    description: str
    key_state: str
    key_manager: str  # AWS or CUSTOMER
    creation_date: datetime | None
    deletion_date: datetime | None = None
    alias: str = ""

    @property
    def is_customer_managed(self) -> bool:
        return self.key_manager == "CUSTOMER"

    @property
    def monthly_cost(self) -> float:
        return get_kms_key_price(self.region, self.key_manager)


@dataclass
class KMSKeyFinding:
    """KMS 키 분석 결과"""

    key: KMSKeyInfo
    status: KMSKeyStatus
    recommendation: str


@dataclass
class KMSKeyAnalysisResult:
    """KMS 키 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    customer_managed_count: int = 0
    aws_managed_count: int = 0
    disabled_count: int = 0
    pending_delete_count: int = 0
    normal_count: int = 0
    disabled_monthly_cost: float = 0.0
    findings: list[KMSKeyFinding] = field(default_factory=list)


def collect_kms_keys(session, account_id: str, account_name: str, region: str) -> list[KMSKeyInfo]:
    """KMS 키 수집"""
    from botocore.exceptions import ClientError

    kms = get_client(session, "kms", region_name=region)
    keys = []

    paginator = kms.get_paginator("list_keys")
    for page in paginator.paginate():
        for key in page.get("Keys", []):
            try:
                key_info = kms.describe_key(KeyId=key["KeyId"])["KeyMetadata"]

                alias = ""
                try:
                    aliases_resp = kms.list_aliases(KeyId=key["KeyId"])
                    for a in aliases_resp.get("Aliases", []):
                        if not a["AliasName"].startswith("alias/aws/"):
                            alias = a["AliasName"]
                            break
                except ClientError:
                    pass

                keys.append(
                    KMSKeyInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        key_id=key["KeyId"],
                        arn=key_info.get("Arn", ""),
                        description=key_info.get("Description", ""),
                        key_state=key_info.get("KeyState", ""),
                        key_manager=key_info.get("KeyManager", ""),
                        creation_date=key_info.get("CreationDate"),
                        deletion_date=key_info.get("DeletionDate"),
                        alias=alias,
                    )
                )
            except ClientError:
                continue

    return keys


def analyze_kms_keys(keys: list[KMSKeyInfo], account_id: str, account_name: str, region: str) -> KMSKeyAnalysisResult:
    """KMS 키 분석"""
    result = KMSKeyAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(keys),
    )

    for key in keys:
        if key.is_customer_managed:
            result.customer_managed_count += 1
        else:
            result.aws_managed_count += 1
            result.normal_count += 1
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.NORMAL,
                    recommendation="AWS 관리 키 (무료)",
                )
            )
            continue

        if key.key_state == "PendingDeletion":
            result.pending_delete_count += 1
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.PENDING_DELETE,
                    recommendation=f"삭제 예정: {key.deletion_date.strftime('%Y-%m-%d') if key.deletion_date else 'N/A'}",
                )
            )
            continue

        if key.key_state == "Disabled":
            result.disabled_count += 1
            result.disabled_monthly_cost += key.monthly_cost
            result.findings.append(
                KMSKeyFinding(
                    key=key,
                    status=KMSKeyStatus.DISABLED,
                    recommendation="비활성화됨 - 삭제 검토",
                )
            )
            continue

        result.normal_count += 1
        result.findings.append(
            KMSKeyFinding(
                key=key,
                status=KMSKeyStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[KMSKeyAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    ws = wb.create_sheet("Summary")
    ws["A1"] = "KMS 키 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "CMK", "AWS관리", "비활성화", "삭제예정", "비활성화 비용"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_count)
        ws.cell(row=row, column=4, value=r.customer_managed_count)
        ws.cell(row=row, column=5, value=r.aws_managed_count)
        ws.cell(row=row, column=6, value=r.disabled_count)
        ws.cell(row=row, column=7, value=r.pending_delete_count)
        ws.cell(row=row, column=8, value=f"${r.disabled_monthly_cost:,.2f}")
        if r.disabled_count > 0:
            ws.cell(row=row, column=6).fill = yellow_fill

    ws_detail = wb.create_sheet("KMS Keys")
    detail_headers = [
        "Account",
        "Region",
        "Key ID",
        "Alias",
        "Type",
        "상태",
        "월간 비용",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.key.is_customer_managed:
                detail_row += 1
                k = f.key
                ws_detail.cell(row=detail_row, column=1, value=k.account_name)
                ws_detail.cell(row=detail_row, column=2, value=k.region)
                ws_detail.cell(row=detail_row, column=3, value=k.key_id[:20] + "...")
                ws_detail.cell(row=detail_row, column=4, value=k.alias or "-")
                ws_detail.cell(row=detail_row, column=5, value=k.key_manager)
                ws_detail.cell(row=detail_row, column=6, value=k.key_state)
                ws_detail.cell(row=detail_row, column=7, value=f"${k.monthly_cost:,.2f}")
                ws_detail.cell(row=detail_row, column=8, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"KMS_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> KMSKeyAnalysisResult | None:
    """단일 계정/리전의 KMS 키 수집 및 분석 (병렬 실행용)"""
    keys = collect_kms_keys(session, account_id, account_name, region)
    if not keys:
        return None
    return analyze_kms_keys(keys, account_id, account_name, region)


def run(ctx) -> None:
    """KMS 키 미사용 분석"""
    console.print("[bold]KMS 키 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="kms")
    results: list[KMSKeyAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_cmk = sum(r.customer_managed_count for r in results)
    total_disabled = sum(r.disabled_count for r in results)
    disabled_cost = sum(r.disabled_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"CMK: {total_cmk}개 / 비활성화: [yellow]{total_disabled}개[/yellow] (${disabled_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("kms-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
