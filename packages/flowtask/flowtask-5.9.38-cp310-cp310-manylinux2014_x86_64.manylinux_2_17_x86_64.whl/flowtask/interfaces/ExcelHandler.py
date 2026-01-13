from pathlib import Path
from typing import Callable
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook as OpenpyxlWorkbook


class ExcelHandler:
    def pandas_to_excel(self, df: pd.DataFrame, file_path: Path, **kwargs):
        """
        Save a pandas DataFrame to an Excel file.

        :param df: DataFrame to save.
        :param file_path: Path to save the Excel file.
        :param kwargs: Additional keyword arguments for pandas' to_excel method.
        """
        df.to_excel(file_path, index=False, **kwargs)

    def edit_excel(self, file_path: Path, edit_func: Callable[[OpenpyxlWorkbook], None]):
        """
        Edit an existing Excel file using a provided function.

        :param file_path: Path to the Excel file.
        :param edit_func: Function that takes a Workbook object and edits it.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"The file {file_path} does not exist.")
        wb = load_workbook(file_path)
        edit_func(wb)
        wb.save(file_path)

    def append_data_to_excel(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str,
        start_row: int = None,
        start_col: int = 1,
        include_header: bool = False
    ):
        """
        Append data from a DataFrame to an existing Excel file.

        :param df: DataFrame containing data to append.
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to append data to.
        :param start_row: Starting row to append data (default is after last row).
        :param start_col: Starting column to append data.
        :param include_header: Whether to include the DataFrame's header.
        """
        if not file_path.exists():
            raise FileNotFoundError(f"The file {file_path} does not exist.")
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            start_row = 1
        else:
            ws = wb[sheet_name]
            if start_row is None:
                start_row = ws.max_row + 1

        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=include_header), start=start_row
        ):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(file_path)

    def create_new_excel_file(self, file_path: Path):
        """
        Create a new Excel file.

        :param file_path: Path where the new Excel file will be saved.
        """
        if file_path.exists():
            raise FileExistsError(f"The file {file_path} already exists.")
        wb = Workbook()
        wb.save(file_path)

    def create_excel_from_template(self, template_path: Path, new_file_path: Path):
        """
        Create a new Excel file from a template.

        :param template_path: Path to the Excel template.
        :param new_file_path: Path where the new Excel file will be saved.
        """
        if not template_path.exists():
            raise FileNotFoundError(f"The template {template_path} does not exist.")
        if new_file_path.exists():
            raise FileExistsError(f"The file {new_file_path} already exists.")
        wb = load_workbook(template_path)
        wb.save(new_file_path)
