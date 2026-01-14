"""
Excel解析核心类
提供Excel文件读取和解析功能
"""
import io
import base64
from typing import Dict, Any, Optional, List, Tuple
from datetime import datetime

import openpyxl
import pandas as pd

try:
    import xlrd
except ImportError:
    xlrd = None

from .utils import (
    base64_to_excel,
    detect_file_format,
    validate_file_size,
    format_error_response,
    format_success_response,
    parse_cell_address,
    parse_range_address,
    validate_cell_address,
    validate_range_address,
)


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化解析器
        
        Args:
            config: 配置字典，包含max_file_size_mb、skip_empty_rows等
        """
        self.config = config or {
            "max_file_size_mb": 100,
            "skip_empty_rows": True,
            "support_formats": [".xlsx", ".xls"]
        }
    
    def _load_workbook(self, file_content: bytes, file_format: str):
        """
        加载工作簿
        
        Args:
            file_content: 文件二进制数据
            file_format: 文件格式
            
        Returns:
            工作簿对象
        """
        if file_format == ".xlsx":
            return openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        else:  # .xls
            xls_workbook = xlrd.open_workbook(file_contents=file_content)
            return self._xls_to_xlsx_compatible(xls_workbook)
    
    def _xls_to_xlsx_compatible(self, xls_workbook) -> openpyxl.Workbook:
        """
        将xls格式转换为openpyxl兼容格式
        
        Args:
            xls_workbook: xlrd工作簿对象
            
        Returns:
            openpyxl工作簿对象
        """
        if xlrd is None:
            raise ImportError("xlrd未安装，无法处理.xls文件。请安装：pip install xlrd")
        
        compatible_wb = openpyxl.Workbook()
        compatible_wb.remove(compatible_wb.active)
        
        for sheet_name in xls_workbook.sheet_names():
            xls_sheet = xls_workbook.sheet_by_name(sheet_name)
            new_sheet = compatible_wb.create_sheet(title=sheet_name)
            
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    new_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        return compatible_wb
    
    def read_sheet_names(self, file_base64: str) -> Dict[str, Any]:
        """
        读取工作簿中所有工作表的名称列表
        
        Args:
            file_base64: Excel文件的Base64编码
            
        Returns:
            包含工作表名称列表的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        # 验证文件大小
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        # 检测文件格式
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        try:
            if file_format == ".xlsx":
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheets = workbook.sheetnames
                workbook.close()
            else:  # .xls
                if xlrd is None:
                    return format_error_response("xlrd未安装，无法读取.xls文件", "DEPENDENCY_MISSING")
                xls_workbook = xlrd.open_workbook(file_contents=file_content)
                sheets = xls_workbook.sheet_names()
            
            return format_success_response({
                "file_format": file_format,
                "sheet_names": sheets,
                "sheet_count": len(sheets)
            })
        except Exception as e:
            return format_error_response(f"获取工作表失败：{str(e)}", "PARSE_ERROR")
    
    def read_sheet_data(
        self,
        file_base64: str,
        sheet: Optional[str] = None,
        range: Optional[str] = None,
        skip_empty_rows: Optional[bool] = None,
        skip_empty_cols: bool = False
    ) -> Dict[str, Any]:
        """
        读取指定工作表的数据
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称，可选，默认第一个工作表
            range: 数据范围，可选，如"A1:B10"
            skip_empty_rows: 是否跳过空行，可选，默认使用配置
            skip_empty_cols: 是否跳过空列，默认False
            
        Returns:
            包含工作表数据的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        try:
            workbook = self._load_workbook(file_content, file_format)
            
            # 选择工作表
            if sheet and sheet in workbook.sheetnames:
                target_sheet = workbook[sheet]
            else:
                target_sheet = workbook.active
            
            # 处理合并单元格（仅.xlsx）
            if file_format == ".xlsx":
                merged_cells_info = []
                for merged_range in list(target_sheet.merged_cells.ranges):
                    top_left_cell = target_sheet[merged_range.coord.split(":")[0]]
                    fill_value = top_left_cell.value
                    merged_cells_info.append({
                        "range": merged_range.coord,
                        "value": fill_value
                    })
                    # 填充合并区域
                    for row in merged_range:
                        for cell in row:
                            target_sheet[cell.coordinate].value = fill_value
            else:
                merged_cells_info = []
            
            # 读取数据
            if range:
                # 解析范围
                try:
                    (start_row, start_col), (end_row, end_col) = parse_range_address(range)
                except ValueError as e:
                    workbook.close()
                    return format_error_response(str(e), "INVALID_RANGE")
                
                # 读取范围数据
                raw_data = []
                max_row = target_sheet.max_row
                max_col = target_sheet.max_column
                
                # 处理整行或整列的情况
                if end_row is None:  # 整列
                    actual_end_row = max_row - 1  # 转换为0-based
                else:
                    actual_end_row = min(end_row, max_row - 1)
                
                if end_col is None:  # 整行
                    actual_end_col = max_col - 1  # 转换为0-based
                else:
                    actual_end_col = min(end_col, max_col - 1)
                
                # 读取范围（转换为1-based索引，openpyxl使用1-based）
                # 注意：range函数是Python内置函数，不是参数range
                import builtins
                for row_idx_0based in builtins.range(start_row, actual_end_row + 1):
                    row_idx_1based = row_idx_0based + 1
                    row_data = []
                    for col_idx_0based in builtins.range(start_col, actual_end_col + 1):
                        col_idx_1based = col_idx_0based + 1
                        if row_idx_1based <= max_row and col_idx_1based <= max_col:
                            cell = target_sheet.cell(row=row_idx_1based, column=col_idx_1based)
                            row_data.append(cell.value)
                        else:
                            row_data.append(None)
                    raw_data.append(row_data)
            else:
                # 读取整个工作表
                raw_data = []
                # 获取实际的最大行和最大列（openpyxl是1-based）
                max_row_1based = target_sheet.max_row
                max_col_1based = target_sheet.max_column
                
                # 如果skip_empty_rows为False，需要读取到max_row以确保包含所有空行
                # 如果skip_empty_rows为True，可以使用iter_rows（只迭代有数据的行）
                skip_rows = skip_empty_rows if skip_empty_rows is not None else self.config.get("skip_empty_rows", True)
                
                if skip_rows:
                    # 只迭代有数据的行（更高效）
                    for row in target_sheet.iter_rows(values_only=True):
                        raw_data.append(list(row))
                else:
                    # 需要包含所有行，包括空行
                    # 从第1行到max_row，确保包含所有可能的空行
                    import builtins
                    for row_idx_1based in builtins.range(1, max_row_1based + 1):
                        row_data = []
                        for col_idx_1based in builtins.range(1, max_col_1based + 1):
                            cell = target_sheet.cell(row=row_idx_1based, column=col_idx_1based)
                            row_data.append(cell.value)
                        raw_data.append(row_data)
            
            # 处理空行和空列
            skip_rows = skip_empty_rows if skip_empty_rows is not None else self.config.get("skip_empty_rows", True)
            if skip_rows:
                raw_data = [row for row in raw_data if not all(cell is None for cell in row)]
            
            if skip_empty_cols:
                # 找出非空列
                non_empty_cols = set()
                for row in raw_data:
                    for col_idx, cell in enumerate(row):
                        if cell is not None:
                            non_empty_cols.add(col_idx)
                # 过滤空列
                raw_data = [[row[col_idx] for col_idx in sorted(non_empty_cols)] for row in raw_data]
            
            # 构造返回数据
            if not raw_data:
                return format_success_response({
                    "file_format": file_format,
                    "sheet_name": target_sheet.title,
                    "headers": [],
                    "data_rows": [],
                    "row_count": 0,
                    "column_count": 0,
                    "merged_cells": merged_cells_info
                })
            
            # 使用 pandas 处理数据（符合设计文档）
            # 先创建 DataFrame，然后处理 NaN 值
            data_frame = pd.DataFrame(raw_data)
            data_frame = data_frame.where(pd.notna(data_frame), None)
            
            # 智能识别表头：如果第一行是空行，找到第一个非空行作为表头
            header_row_idx = 0
            if not skip_rows:  # 只有在保留空行模式下才需要智能识别
                import builtins
                for idx in builtins.range(len(data_frame)):
                    row = data_frame.iloc[idx].tolist()
                    if not all(cell is None for cell in row):
                        header_row_idx = idx
                        break
            
            # 提取表头和数据行
            if len(data_frame) > 0:
                headers = data_frame.iloc[header_row_idx].tolist() if header_row_idx < len(data_frame) else []
                data_rows = data_frame.iloc[header_row_idx + 1:].values.tolist() if header_row_idx + 1 < len(data_frame) else []
            else:
                headers = []
                data_rows = []
            
            # 如果表头行本身是空行，则表头为空列表
            if headers and all(cell is None for cell in headers):
                headers = []
            
            workbook.close()
            
            return format_success_response({
                "file_format": file_format,
                "sheet_name": target_sheet.title,
                "headers": headers,
                "data_rows": data_rows,
                "row_count": len(data_rows),
                "column_count": len(headers),
                "merged_cells": merged_cells_info
            })
        except Exception as e:
            return format_error_response(f"读取工作表数据失败：{str(e)}", "PARSE_ERROR")
    
    def read_cell_data(
        self,
        file_base64: str,
        sheet: str,
        cell: str
    ) -> Dict[str, Any]:
        """
        读取单个或范围单元格数据
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            cell: 单元格地址或范围，如"A1"或"A1:B10"
            
        Returns:
            包含单元格数据的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        try:
            workbook = self._load_workbook(file_content, file_format)
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            cells_data = []
            
            # 判断是单个单元格还是范围
            if ":" in cell:
                # 范围
                if not validate_range_address(cell):
                    workbook.close()
                    return format_error_response(f"无效的范围地址：{cell}", "INVALID_RANGE")
                
                try:
                    (start_row, start_col), (end_row, end_col) = parse_range_address(cell)
                except ValueError as e:
                    workbook.close()
                    return format_error_response(str(e), "INVALID_RANGE")
                
                max_row = target_sheet.max_row
                max_col = target_sheet.max_column
                
                actual_end_row = end_row if end_row is not None else max_row
                actual_end_col = end_col if end_col is not None else max_col
                
                for row_idx in range(start_row + 1, actual_end_row + 2):
                    for col_idx in range(start_col + 1, actual_end_col + 2):
                        cell_obj = target_sheet.cell(row=row_idx, column=col_idx)
                        cells_data.append({
                            "address": cell_obj.coordinate,
                            "value": cell_obj.value,
                            "data_type": type(cell_obj.value).__name__ if cell_obj.value is not None else "None",
                            "format": str(cell_obj.number_format) if hasattr(cell_obj, 'number_format') else "General"
                        })
            else:
                # 单个单元格
                if not validate_cell_address(cell):
                    workbook.close()
                    return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
                
                try:
                    row_idx, col_idx = parse_cell_address(cell)
                except ValueError as e:
                    workbook.close()
                    return format_error_response(str(e), "INVALID_CELL")
                
                cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
                cells_data.append({
                    "address": cell_obj.coordinate,
                    "value": cell_obj.value,
                    "data_type": type(cell_obj.value).__name__ if cell_obj.value is not None else "None",
                    "format": str(cell_obj.number_format) if hasattr(cell_obj, 'number_format') else "General"
                })
            
            workbook.close()
            
            return format_success_response({
                "cells": cells_data
            })
        except Exception as e:
            return format_error_response(f"读取单元格数据失败：{str(e)}", "PARSE_ERROR")
    
    def read_cell_formula(
        self,
        file_base64: str,
        sheet: str,
        cell: str
    ) -> Dict[str, Any]:
        """
        读取单元格公式
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称
            cell: 单元格地址
            
        Returns:
            包含公式信息的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        if file_format == ".xls":
            return format_error_response(".xls格式不支持读取公式", "FORMULA_NOT_SUPPORTED")
        
        try:
            # 使用keep_vba=True以保留公式
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
            
            if sheet not in workbook.sheetnames:
                workbook.close()
                return format_error_response(f"工作表不存在：{sheet}", "SHEET_NOT_FOUND")
            
            target_sheet = workbook[sheet]
            
            if not validate_cell_address(cell):
                workbook.close()
                return format_error_response(f"无效的单元格地址：{cell}", "INVALID_CELL")
            
            try:
                row_idx, col_idx = parse_cell_address(cell)
            except ValueError as e:
                workbook.close()
                return format_error_response(str(e), "INVALID_CELL")
            
            cell_obj = target_sheet.cell(row=row_idx + 1, column=col_idx + 1)
            
            formula = cell_obj.value if isinstance(cell_obj.value, str) and cell_obj.value.startswith("=") else None
            calculated_value = None
            
            # 尝试读取已计算值
            if formula:
                workbook_calculated = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet_calculated = workbook_calculated[sheet]
                cell_calculated = sheet_calculated.cell(row=row_idx + 1, column=col_idx + 1)
                calculated_value = cell_calculated.value
                workbook_calculated.close()
            
            # 解析公式依赖（简单实现）
            dependencies = []
            if formula:
                import re
                # 匹配单元格引用，如A1, $A$1, Sheet1!A1等
                pattern = r'([A-Z]+\d+)'
                matches = re.findall(pattern, formula)
                dependencies = list(set(matches))
            
            workbook.close()
            
            return format_success_response({
                "cell": cell_obj.coordinate,
                "formula": formula,
                "calculated_value": calculated_value,
                "dependencies": dependencies
            })
        except Exception as e:
            return format_error_response(f"读取单元格公式失败：{str(e)}", "PARSE_ERROR")
    
    def read_merged_cells(
        self,
        file_base64: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        读取合并单元格信息
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称，可选，默认所有工作表
            
        Returns:
            包含合并单元格信息的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        if file_format == ".xls":
            # .xls格式的合并单元格处理较复杂，这里返回空
            return format_success_response({
                "merged_cells": []
            })
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            
            merged_cells_list = []
            
            if sheet and sheet in workbook.sheetnames:
                sheets_to_process = [workbook[sheet]]
            else:
                sheets_to_process = workbook.worksheets
            
            for target_sheet in sheets_to_process:
                for merged_range in target_sheet.merged_cells.ranges:
                    top_left_cell = target_sheet[merged_range.coord.split(":")[0]]
                    fill_value = top_left_cell.value
                    
                    # 解析范围坐标
                    start_cell, end_cell = merged_range.coord.split(":")
                    start_row, start_col = parse_cell_address(start_cell)
                    end_row, end_col = parse_cell_address(end_cell)
                    
                    merged_cells_list.append({
                        "range": merged_range.coord,
                        "value": fill_value,
                        "sheet": target_sheet.title,
                        "start_row": start_row + 1,  # 转换为1-based
                        "start_col": start_col + 1,
                        "end_row": end_row + 1,
                        "end_col": end_col + 1
                    })
            
            workbook.close()
            
            return format_success_response({
                "merged_cells": merged_cells_list
            })
        except Exception as e:
            return format_error_response(f"读取合并单元格失败：{str(e)}", "PARSE_ERROR")
    
    def read_chart_info(
        self,
        file_base64: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        读取图表信息
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称，可选，默认所有工作表
            
        Returns:
            包含图表信息的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            # .xls格式不支持图表读取
            return format_success_response({
                "charts": []
            })
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            
            charts_list = []
            
            if sheet and sheet in workbook.sheetnames:
                sheets_to_process = [workbook[sheet]]
            else:
                sheets_to_process = workbook.worksheets
            
            for target_sheet in sheets_to_process:
                # openpyxl中图表信息需要通过drawing对象获取
                if hasattr(target_sheet, '_charts') and target_sheet._charts:
                    for idx, chart in enumerate(target_sheet._charts):
                        chart_info = {
                            "id": f"chart_{idx}",
                            "name": getattr(chart, 'title', {}).get('name', f"Chart {idx + 1}") if hasattr(chart, 'title') else f"Chart {idx + 1}",
                            "type": type(chart).__name__,
                            "sheet": target_sheet.title
                        }
                        
                        # 尝试获取图表标题
                        if hasattr(chart, 'title') and chart.title:
                            if hasattr(chart.title, 'tx'):
                                chart_info["title"] = str(chart.title.tx)
                        
                        # 尝试获取数据源
                        if hasattr(chart, 'dataSources'):
                            chart_info["data_range"] = str(chart.dataSources)
                        
                        charts_list.append(chart_info)
            
            workbook.close()
            
            return format_success_response({
                "charts": charts_list
            })
        except Exception as e:
            return format_error_response(f"读取图表信息失败：{str(e)}", "PARSE_ERROR")
    
    def read_table_info(
        self,
        file_base64: str,
        sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        读取表格信息
        
        Args:
            file_base64: Excel文件的Base64编码
            sheet: 工作表名称，可选，默认所有工作表
            
        Returns:
            包含表格信息的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format or file_format == ".xls":
            # .xls格式不支持表格读取
            return format_success_response({
                "tables": []
            })
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            
            tables_list = []
            
            if sheet and sheet in workbook.sheetnames:
                sheets_to_process = [workbook[sheet]]
            else:
                sheets_to_process = workbook.worksheets
            
            for target_sheet in sheets_to_process:
                # openpyxl中表格信息通过tables属性获取
                if hasattr(workbook, 'defined_names') and workbook.defined_names:
                    # 查找表格定义
                    for name, definition in workbook.defined_names.items():
                        if 'table' in name.lower():
                            tables_list.append({
                                "name": name,
                                "sheet": target_sheet.title,
                                "range": str(definition)
                            })
            
            workbook.close()
            
            return format_success_response({
                "tables": tables_list
            })
        except Exception as e:
            return format_success_response({
                "tables": []
            })  # 表格读取失败时返回空列表，不报错
    
    def get_workbook_info(self, file_base64: str) -> Dict[str, Any]:
        """
        获取工作簿基本信息
        
        Args:
            file_base64: Excel文件的Base64编码
            
        Returns:
            包含工作簿信息的响应字典
        """
        try:
            file_content = base64_to_excel(file_base64)
        except ValueError as e:
            return format_error_response(str(e), "BASE64_DECODE_ERROR")
        
        is_valid, error_msg = validate_file_size(file_content)
        if not is_valid:
            return format_error_response(error_msg, "FILE_TOO_LARGE")
        
        file_format = detect_file_format(file_content)
        if not file_format:
            return format_error_response(
                f"不支持的格式，仅支持{','.join(self.config['support_formats'])}",
                "INVALID_FILE_FORMAT"
            )
        
        try:
            file_size_bytes = len(file_content)
            
            if file_format == ".xlsx":
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet_count = len(workbook.sheetnames)
                workbook.close()
            else:
                if xlrd is None:
                    return format_error_response("xlrd未安装，无法读取.xls文件", "DEPENDENCY_MISSING")
                xls_workbook = xlrd.open_workbook(file_contents=file_content)
                sheet_count = len(xls_workbook.sheet_names())
            
            return format_success_response({
                "file_format": file_format,
                "sheet_count": sheet_count,
                "file_size_bytes": file_size_bytes,
                "file_size_mb": round(file_size_bytes / 1024 / 1024, 2)
            })
        except Exception as e:
            return format_error_response(f"获取工作簿信息失败：{str(e)}", "PARSE_ERROR")

