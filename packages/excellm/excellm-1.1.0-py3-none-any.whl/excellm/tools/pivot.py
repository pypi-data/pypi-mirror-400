"""Pivot table operations for ExceLLM MCP server.

Supports creating pivot tables in Excel workbooks with dual-engine support:
- COM engine: Live Excel automation on Windows (native pivot tables)
- openpyxl engine: File-based cross-platform (summary tables)
"""

import logging
import os
import uuid
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


# Aggregation function constants for COM (xlConsolidationFunction)
XL_AGG_FUNCTIONS = {
    "sum": -4157,      # xlSum
    "count": -4112,    # xlCount
    "average": -4106,  # xlAverage
    "max": -4136,      # xlMax
    "min": -4139,      # xlMin
}


def create_pivot_table_sync(
    workbook_name: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "sum",
    target_sheet: Optional[str] = None,
    target_cell: str = "A1",
    table_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Create a pivot table from source data.
    
    Supports dual-engine operation:
    - COM engine for live Excel (Windows) - creates native pivot tables
    - openpyxl engine for file-based operations - creates summary tables
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet containing source data
        data_range: Source data range (e.g., "A1:D100")
        rows: Fields for row labels (column headers from data)
        values: Fields for values to aggregate
        columns: Optional fields for column labels
        agg_func: Aggregation function (sum, count, average, max, min)
        target_sheet: Sheet for pivot table (default: creates new sheet)
        target_cell: Cell for pivot table location (default: "A1")
        table_name: Optional name for the pivot table
        
    Returns:
        Dictionary with operation result
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    # Validate aggregation function
    agg_func_lower = agg_func.lower()
    if agg_func_lower not in XL_AGG_FUNCTIONS:
        return {
            "success": False,
            "error": f"Invalid aggregation function: {agg_func}. "
                    f"Supported: {', '.join(XL_AGG_FUNCTIONS.keys())}",
        }
    
    if not rows:
        return {
            "success": False,
            "error": "At least one row field is required",
        }
    
    if not values:
        return {
            "success": False,
            "error": "At least one value field is required",
        }
    
    try:
        # Engine Selection Logic
        use_com = True
        try:
            excel = get_excel_app()
            workbook = get_workbook(excel, workbook_name)
        except Exception:
            use_com = False
            
        if not use_com:
            # Fallback to openpyxl if it's a file
            if os.path.exists(workbook_name):
                return create_pivot_table_openpyxl(
                    workbook_name, sheet_name, data_range, rows, values,
                    columns, agg_func, target_sheet, target_cell, table_name
                )
            else:
                return {
                    "success": False,
                    "error": f"Workbook not found (checked open workbooks and file path): {workbook_name}"
                }

        source_sheet = get_worksheet(workbook, sheet_name)
        
        # Get the source data range
        try:
            source_range = source_sheet.Range(data_range)
        except Exception as e:
            return {
                "success": False,
                "error": f"Invalid data range: {data_range}. {str(e)}",
            }
        
        # Create or get target sheet
        if target_sheet:
            try:
                pivot_sheet = get_worksheet(workbook, target_sheet)
            except Exception:
                # Create new sheet
                pivot_sheet = workbook.Sheets.Add()
                pivot_sheet.Name = target_sheet
        else:
            # Create a new sheet with auto-generated name
            pivot_sheet_name = f"{sheet_name}_Pivot"
            # Check if sheet exists
            try:
                pivot_sheet = get_worksheet(workbook, pivot_sheet_name)
            except Exception:
                pivot_sheet = workbook.Sheets.Add()
                pivot_sheet.Name = pivot_sheet_name
                
            target_sheet = pivot_sheet_name
        
        # Get target cell position
        target_range = pivot_sheet.Range(target_cell)
        
        # Generate pivot table name if not provided
        if not table_name:
            table_name = f"PivotTable_{uuid.uuid4().hex[:8]}"
        
        # Create pivot cache
        pivot_cache = workbook.PivotCaches().Create(
            SourceType=1,  # xlDatabase
            SourceData=source_range,
        )
        
        # Create pivot table
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=target_range,
            TableName=table_name,
        )
        
        # Add row fields
        for i, field_name in enumerate(rows):
            try:
                field = pivot_table.PivotFields(field_name)
                field.Orientation = 1  # xlRowField
                field.Position = i + 1
            except Exception as e:
                logger.warning(f"Could not add row field '{field_name}': {e}")
        
        # Add column fields
        if columns:
            for i, field_name in enumerate(columns):
                try:
                    field = pivot_table.PivotFields(field_name)
                    field.Orientation = 2  # xlColumnField
                    field.Position = i + 1
                except Exception as e:
                    logger.warning(f"Could not add column field '{field_name}': {e}")
        
        # Add value fields
        for field_name in values:
            try:
                field = pivot_table.PivotFields(field_name)
                field.Orientation = 4  # xlDataField
                field.Function = XL_AGG_FUNCTIONS[agg_func_lower]
                # Set caption to include aggregation type
                field.Caption = f"{agg_func.capitalize()} of {field_name}"
            except Exception as e:
                logger.warning(f"Could not add value field '{field_name}': {e}")
        
        logger.info(f"Created pivot table '{table_name}' in {target_sheet}")
        
        return {
            "success": True,
            "message": "Pivot table created successfully",
            "engine": "COM",
            "details": {
                "table_name": table_name,
                "source_range": data_range,
                "target_sheet": target_sheet,
                "target_cell": target_cell,
                "rows": rows,
                "columns": columns or [],
                "values": values,
                "aggregation": agg_func,
            },
        }
        
    except Exception as e:
        logger.error(f"Failed to create pivot table: {e}")
        return {
            "success": False,
            "error": f"Failed to create pivot table: {str(e)}",
        }


def refresh_pivot_table_sync(
    workbook_name: str,
    sheet_name: str,
    table_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Refresh pivot table(s) to update with latest source data.
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet containing pivot table
        table_name: Name of specific pivot table (None = refresh all)
        
    Returns:
        Dictionary with operation result
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    try:
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
        worksheet = get_worksheet(workbook, sheet_name)
        
        refreshed = []
        
        if table_name:
            # Refresh specific pivot table
            try:
                pivot_table = worksheet.PivotTables(table_name)
                pivot_table.RefreshTable()
                refreshed.append(table_name)
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Pivot table '{table_name}' not found: {str(e)}",
                }
        else:
            # Refresh all pivot tables in the sheet
            for pt in worksheet.PivotTables():
                try:
                    pt.RefreshTable()
                    refreshed.append(pt.Name)
                except Exception as e:
                    logger.warning(f"Could not refresh pivot table: {e}")
        
        return {
            "success": True,
            "message": f"Refreshed {len(refreshed)} pivot table(s)",
            "refreshed": refreshed,
        }
        
    except Exception as e:
        logger.error(f"Failed to refresh pivot table: {e}")
        return {
            "success": False,
            "error": f"Failed to refresh pivot table: {str(e)}",
        }


def list_pivot_tables_sync(
    workbook_name: str,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """List all pivot tables in a workbook or specific sheet.
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Optional sheet name (None = all sheets)
        
    Returns:
        Dictionary with list of pivot tables
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    try:
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
        
        pivot_tables = []
        
        if sheet_name:
            sheets = [get_worksheet(workbook, sheet_name)]
        else:
            sheets = [workbook.Sheets(i) for i in range(1, workbook.Sheets.Count + 1)]
        
        for ws in sheets:
            try:
                for pt in ws.PivotTables():
                    pivot_info = {
                        "name": pt.Name,
                        "sheet": ws.Name,
                        "source_range": pt.SourceData if hasattr(pt, 'SourceData') else None,
                        "location": pt.TableRange2.Address if hasattr(pt, 'TableRange2') else None,
                    }
                    pivot_tables.append(pivot_info)
            except Exception:
                continue  # Sheet might not have pivot tables
        
        return {
            "success": True,
            "pivot_tables": pivot_tables,
            "count": len(pivot_tables),
        }
        
    except Exception as e:
        logger.error(f"Failed to list pivot tables: {e}")
        return {
            "success": False,
            "error": f"Failed to list pivot tables: {str(e)}",
        }


def delete_pivot_table_sync(
    workbook_name: str,
    sheet_name: str,
    table_name: str,
    keep_data: bool = False,
) -> Dict[str, Any]:
    """Delete a pivot table.
    
    Args:
        workbook_name: Name of open workbook
        sheet_name: Name of worksheet
        table_name: Name of pivot table to delete
        keep_data: If True, convert to values before deleting
        
    Returns:
        Dictionary with operation result
    """
    from ..core.connection import get_excel_app, get_workbook, get_worksheet
    
    try:
        excel = get_excel_app()
        workbook = get_workbook(excel, workbook_name)
        worksheet = get_worksheet(workbook, sheet_name)
        
        try:
            pivot_table = worksheet.PivotTables(table_name)
        except Exception:
            return {
                "success": False,
                "error": f"Pivot table '{table_name}' not found",
            }
        
        if keep_data:
            # Convert to values first
            try:
                table_range = pivot_table.TableRange2
                table_range.Copy()
                table_range.PasteSpecial(Paste=-4163)  # xlPasteValues
            except Exception as e:
                logger.warning(f"Could not convert to values: {e}")
        
        # Delete the pivot table
        pivot_table.TableRange2.Clear()
        
        logger.info(f"Deleted pivot table '{table_name}' from {sheet_name}")
        
        return {
            "success": True,
            "message": f"Pivot table '{table_name}' deleted successfully",
            "data_kept": keep_data,
        }
        
    except Exception as e:
        logger.error(f"Failed to delete pivot table: {e}")
        return {
            "success": False,
            "error": f"Failed to delete pivot table: {str(e)}",
        }


# ============================================================================
# openpyxl Engine - File-based Pivot Table Creation (Cross-Platform)
# ============================================================================


def create_pivot_table_openpyxl(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "sum",
    target_sheet: Optional[str] = None,
    target_cell: str = "A1",
    table_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Create a pivot table in an Excel file using openpyxl (file-based).
    
    LIMITATION: openpyxl can create the pivot table definition, but Excel
    performs the actual calculation when the file is opened. This means the 
    data won't be visible/calculated until opened in Excel.
    
    For immediate results without Excel, we can generate a "Summary Table" 
    instead (static values calculated in Python).
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of worksheet containing source data
        data_range: Source data range (e.g., "A1:D100")
        rows: Fields for row labels
        values: Fields for values
        columns: Optional fields for column labels
        agg_func: Aggregation function (sum, count, average, max, min)
        target_sheet: Sheet for pivot table
        
    Returns:
        Dictionary with operation result
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries, get_column_letter
    except ImportError:
        return {
            "success": False,
            "error": "openpyxl is required for file-based operations. Install with: pip install openpyxl",
        }
        
    # Since openpyxl pivot tables are complex and often require Excel to calc,
    # we'll implement a "Static Summary Table" which is more robust for LLMs to read back immediately.
    # This manually aggregates the data using Python.
    
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
        ws = wb[sheet_name]
        
        # Parse range (remove sheet name if present)
        if "!" in data_range:
            data_range = data_range.split("!")[-1]
            
        min_col, min_row, max_col, max_row = range_boundaries(data_range)
        
        # Read header row to map column names to indices
        headers = {}
        for col_idx in range(min_col, max_col + 1):
            cell_val = ws.cell(row=min_row, column=col_idx).value
            if cell_val:
                headers[str(cell_val)] = col_idx
                
        # Validate fields
        for f in rows + values + (columns or []):
            if f not in headers:
                return {"success": False, "error": f"Field '{f}' not found in headers"}
                
        # Create/Get target sheet
        if target_sheet:
            if target_sheet in wb.sheetnames:
                target_ws = wb[target_sheet]
                # Clear target area (simple clear)
                target_ws.delete_rows(1, target_ws.max_row + 1)
            else:
                target_ws = wb.create_sheet(target_sheet)
        else:
            target_sheet_name = f"{sheet_name}_Summary"
            if target_sheet_name in wb.sheetnames:
                target_ws = wb[target_sheet_name]
                target_ws.delete_rows(1, target_ws.max_row + 1)
            else:
                target_ws = wb.create_sheet(target_sheet_name)
            target_sheet = target_sheet_name
            
        # ---------------------------------------------------------
        # Perform Aggregation (Python-side Pivot)
        # ---------------------------------------------------------
        
        # Data structure: { (row_key_tuple, col_key_tuple): {val_field: [values]} }
        data_map = {}
        row_keys_set = set()
        col_keys_set = set()
        
        for r in range(min_row + 1, max_row + 1):
            # Build row key
            r_key = tuple(ws.cell(row=r, column=headers[field]).value for field in rows)
            
            # Build column key
            if columns:
                c_key = tuple(ws.cell(row=r, column=headers[field]).value for field in columns)
            else:
                c_key = ("Total",)
                
            row_keys_set.add(r_key)
            col_keys_set.add(c_key)
            
            combo_key = (r_key, c_key)
            if combo_key not in data_map:
                data_map[combo_key] = {v: [] for v in values}
                
            for v_field in values:
                val = ws.cell(row=r, column=headers[v_field]).value
                # Coerce to float
                try:
                    if val is not None:
                        val = float(val)
                    else:
                        val = 0.0
                except ValueError:
                    val = 0.0
                data_map[combo_key][v_field].append(val)
                
        # Sort keys
        sorted_row_keys = sorted(list(row_keys_set), key=lambda x: str(x))
        sorted_col_keys = sorted(list(col_keys_set), key=lambda x: str(x))
        
        # ---------------------------------------------------------
        # Write Result Table
        # ---------------------------------------------------------
        
        current_row = 1
        current_col = 1
        
        # Styles
        from openpyxl.styles import Font
        bold_font = Font(bold=True)
        
        # Write Headers
        # Row Labels
        for i, field in enumerate(rows):
            target_ws.cell(row=current_row, column=current_col + i, value=field).font = bold_font
            
        # Column Labels (if any)
        if columns:
            # Shift data start
            data_start_col = current_col + len(rows)
            for c_key in sorted_col_keys:
                # Join tuple parts
                col_label = " - ".join(str(k) for k in c_key)
                for v_field in values:
                    target_ws.cell(row=current_row, column=data_start_col, value=f"{col_label} | {v_field}").font = bold_font
                    data_start_col += 1
        else:
            # Just Value fields
            data_start_col = current_col + len(rows)
            for v_field in values:
                target_ws.cell(row=current_row, column=data_start_col, value=f"{agg_func.capitalize()} of {v_field}").font = bold_font
                data_start_col += 1
                
        current_row += 1
        
        # Helper for aggregation
        def calc_agg(vals, func):
            if not vals: return 0
            if func == "sum": return sum(vals)
            if func == "count": return len(vals)
            if func == "average": return sum(vals) / len(vals)
            if func == "max": return max(vals)
            if func == "min": return min(vals)
            return 0
            
        # Write Data

        for r_key in sorted_row_keys:
            # Write Row Labels
            for i, val in enumerate(r_key):
                target_ws.cell(row=current_row, column=1 + i, value=val)
                
            col_ptr = 1 + len(rows)
            
            # Write Data Values
            if columns:
                for c_key in sorted_col_keys:
                    stats = data_map.get((r_key, c_key))
                    for v_field in values:
                        if stats:
                            result = calc_agg(stats[v_field], agg_func)
                            target_ws.cell(row=current_row, column=col_ptr, value=result)
                        else:
                            target_ws.cell(row=current_row, column=col_ptr, value=0)
                        col_ptr += 1
            else:
                # No column groups
                stats = data_map.get((r_key, ("Total",)))
                for v_field in values:
                    if stats:
                        result = calc_agg(stats[v_field], agg_func)
                        target_ws.cell(row=current_row, column=col_ptr, value=result)
                    else:
                        target_ws.cell(row=current_row, column=col_ptr, value=0)
                    col_ptr += 1
            
            current_row += 1
            
        wb.save(filepath)
        
        return {
            "success": True,
            "message": "Pivot summary table created (Static Values)",
            "engine": "openpyxl (static)",
            "details": {
                "target_sheet": target_sheet,
                "rows_written": len(sorted_row_keys),
                "cols_written": col_ptr - 1
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to create pivot table with openpyxl: {e}")
        return {
            "success": False,
            "error": f"Failed to create pivot table: {str(e)}",
        }
