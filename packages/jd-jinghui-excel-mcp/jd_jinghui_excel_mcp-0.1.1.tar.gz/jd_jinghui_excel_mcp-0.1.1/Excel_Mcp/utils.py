from typing import List, Optional, Any, Dict, Union
import json
import os

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pandas as pd  # 可选：用于便捷的CSV/JSON导出
except Exception:
    pd = None


def describe_sheets(path: str) -> List[Dict[str, Any]]:
    """
    返回指定路径 Excel 工作簿中所有工作表的元数据。
    返回包含：name (表名), max_row (最大行), max_column (最大列)
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    result = []
    for name in wb.sheetnames:
        ws = wb[name]
        # 只读模式下 max_row 可能不准确，做个访问保护
        max_row = getattr(ws, "max_row", None)
        max_col = getattr(ws, "max_column", None)
        result.append({"name": name, "max_row": max_row, "max_column": max_col})
    wb.close()
    return result


def read_sheet(path: str, sheet: Optional[str] = None, start_row: int = 1, limit: Optional[int] = None) -> List[List[Any]]:
    """
    读取工作表中的值并以列表形式返回。

    参数:
    - `sheet`: 工作表名称，如果为 None 则使用当前激活的工作表
    - `start_row`: 开始读取的行号（从1开始）
    - `limit`: 限制返回的最大行数
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows: List[List[Any]] = []
    seen = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        rows.append(list(row))
        seen += 1
        if limit is not None and seen >= limit:
            break
    wb.close()
    return rows


def write_to_sheet(path: str, sheet: str, rows: List[List[Any]], create_if_missing: bool = True) -> None:
    """
    向工作表追加写入行数据。

    如果工作簿不存在将自动创建。
    `rows` 是一个二维列表，例如 [['a', 'b'], [1, 2]]。
    """
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()

    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        if create_if_missing:
            ws = wb.create_sheet(sheet)
        else:
            wb.close()
            raise ValueError(f"工作表 {sheet} 未在工作簿中找到，且 create_if_missing 为 False")

    # 追加行
    for r in rows:
        ws.append(r)

    wb.save(path)
    wb.close()


def copy_sheet(path: str, src: str, dest: str) -> None:
    """
    在同一个工作簿中复制现有的工作表，并重命名为 `dest`。
    """
    wb = load_workbook(path)
    if src not in wb.sheetnames:
        wb.close()
        raise ValueError(f"源工作表 {src} 未找到")
    source_ws = wb[src]
    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = dest
    wb.save(path)
    wb.close()


def create_table(path: str, sheet: str, header: List[str], rows: List[List[Any]]) -> None:
    """
    创建一个简单的表格（包含表头和数据行）。
    """
    write_to_sheet(path, sheet, [header] + rows, create_if_missing=True)


def format_range(path: str, sheet: str, start_cell: str, end_cell: str, bold: bool = False, bg_hex: Optional[str] = None, align_center: bool = False) -> None:
    """
    对矩形区域应用简单的格式（加粗/背景色/居中）。

    `start_cell` 和 `end_cell` 是 Excel 坐标，如 'A1' 和 'C3'。
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表 {sheet} 未找到")
    ws = wb[sheet]

    # 简单的坐标转换辅助函数 'A1' -> (row, col)
    def coord_to_rc(coord: str):
        letters = ''.join([c for c in coord if c.isalpha()])
        numbers = ''.join([c for c in coord if c.isdigit()])
        col = 0
        for ch in letters.upper():
            col = col * 26 + (ord(ch) - ord('A') + 1)
        row = int(numbers)
        return row, col

    r1, c1 = coord_to_rc(start_cell)
    r2, c2 = coord_to_rc(end_cell)
    
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if bold:
                cell.font = Font(bold=True)
            if bg_hex:
                color = bg_hex.replace('#', '')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if align_center:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)
    wb.close()


def export_sheet_to_json(path: str, sheet: Optional[str] = None, start_row: int = 1, limit: Optional[int] = None) -> str:
    """
    辅助函数：将工作表内容导出为 JSON 字符串。
    如果有 pandas 则使用 pandas，否则使用原生 json dump。
    """
    rows = read_sheet(path, sheet=sheet, start_row=start_row, limit=limit)
    if pd is not None and rows:
        # 假设第一行是表头
        header = rows[0]
        data = rows[1:]
        df = pd.DataFrame(data, columns=header)
        return df.to_json(orient="records", force_ascii=False)
    else:
        return json.dumps(rows, ensure_ascii=False)


# --- 以下是新增的常用功能 ---

def delete_sheet(path: str, sheet: str) -> None:
    """
    【新增】删除指定的工作表。
    """
    wb = load_workbook(path)
    if sheet in wb.sheetnames:
        del wb[sheet]
        wb.save(path)
    else:
        # 可以选择报错或者忽略，这里选择忽略
        pass
    wb.close()


def merge_cells(path: str, sheet: str, start_cell: str, end_cell: str) -> None:
    """
    【新增】合并单元格区域，例如 'A1' 到 'B2'。
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表 {sheet} 未找到")
    ws = wb[sheet]
    ws.merge_cells(f"{start_cell}:{end_cell}")
    wb.save(path)
    wb.close()


def write_formula(path: str, sheet: str, cell: str, formula: str) -> None:
    """
    【新增】向指定单元格写入公式，例如 '=SUM(A1:A10)'。
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表 {sheet} 未找到")
    ws = wb[sheet]
    ws[cell] = formula
    wb.save(path)
    wb.close()


def set_column_width(path: str, sheet: str, column: str, width: float) -> None:
    """
    【新增】设置列宽。
    
    参数:
    - column: 列字母，例如 'A', 'B'
    - width: 宽度数值，例如 20.0
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表 {sheet} 未找到")
    ws = wb[sheet]
    ws.column_dimensions[column].width = width
    wb.save(path)
    wb.close()


def add_auto_filter(path: str, sheet: str, cell_range: str) -> None:
    """
    【新增】给指定区域添加自动筛选（漏斗图标）。
    
    参数:
    - cell_range: 例如 "A1:C10"
    """
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"工作表 {sheet} 未找到")
    ws = wb[sheet]
    ws.auto_filter.ref = cell_range
    wb.save(path)
    wb.close()