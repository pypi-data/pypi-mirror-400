import ssl
import os
import re
import socks
import time
import socket
import certifi
from pathlib import Path
import json as Json
#pip install PySocks requests
import requests
from requests.exceptions import RequestException
#pip install lxml
from lxml import etree
#pip install openpyxl
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
#pip install pymysql
import pymysql
#pip install Pillow
import PIL
from requests.adapters import HTTPAdapter
from urllib3.util.ssl_ import create_urllib3_context
from . import Com
from . import Proxys

# 全局变量和函数
IsDebug=True
UserAgentMac="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.6 Safari/605.1.15"
UserAgentIPhone="Mozilla/5.0 (iPhone; CPU iPhone OS 15_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.6 Mobile/15E148 Safari/604.1"
UserAgent=UserAgentMac #请求客户端信息
IsProxy=False #是否使用代理
IsProxyLocal=False #是否使用本地抓包调试
ProxyLocalPemFile="" #本地抓包调试对应的pem证书文件
LocalIP=Com.GetLanIp() #本机局域网IP
ProxyConfigs=None #代理列表
ThreadCount=4 #线程数
Sessions=None #会话列表：会话列表数=代理列表数*线程数
SessionCallSpanTime=60 #会话调用间隔时间，单位秒
Http=requests.Session()

class CustomSSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = create_urllib3_context()
        kwargs['ssl_context'] = context
        context.load_default_certs()  # 加载系统默认证书
        return super().init_poolmanager(*args, **kwargs)

# 设置无代理配置
def GetNoneProxy():
    if IsProxyLocal:
        proxy={'http': f'http://{LocalIP}:8888', 'https': f'http://{LocalIP}:8888'}
        verify=ProxyLocalPemFile
    else:
        proxy=None
        verify=None
    return proxy,verify

# 获取代理列表
def GetProxys():
    proxys=[]
    # 无代理
    if(not IsProxy):return proxys
    # 获取可用代理
    if ProxyConfigs is None: ProxyConfigs=Proxys.GetProxyConfigs()
    for cfg in ProxyConfigs:
        if(cfg["ip"] is None): continue
        url = f"{cfg["protocol"]}://"+("" if cfg["user"] is None else f"{cfg["user"]}:{cfg["password"]}@")+cfg["addr"]
        proxy={
            'http': url,
            'https': url
        }
        proxys.append(proxy)
    return proxys

# 获取代理
def GetSessions():
    global Sessions
    if Sessions is not None: return Sessions
    Sessions=[]
    proxys=GetProxys()
    proxyCount=0 if proxys is None else len(proxys)
    for proxy in proxys:
        session={
            "errorCount":0,
            "proxy":proxy,
            "http":requests.Session(),
            "time":time.time()-SessionCallSpanTime,
            "verify":certifi.where()
        }
        Sessions.append(session)
    if(proxyCount<ThreadCount):
        proxy,verify=GetNoneProxy()
        for i in range(proxyCount,ThreadCount):
            session={
                "errorCount":0,
                "proxy":proxy,
                "http":requests.Session(),
                "time":time.time()-SessionCallSpanTime,
                "verify":verify
            }
            Sessions.append(session)
    return Sessions

# 获取会话
def GetSession():
    session=GetSessions().pop(0)
    spanTime=time.time()-session["time"]
    if spanTime<SessionCallSpanTime:
        time.sleep(SessionCallSpanTime-spanTime) #不足间隔时间则等待
    return session
# 设置会话
def SetSession(session):
    session["time"]=time.time()
    GetSessions().append(session)

# get请求
def Get(url,params=None,headers=None,encoding="utf-8"):
    session=GetSession()
    try:
        if(headers is None): headers={"User-Agent":UserAgent}
        res = session["http"].get(url,headers=headers, params=params, proxies=session["proxy"], verify=session["verify"])
        res.encoding=encoding
        SetSession(session)
    except RequestException as e:
        session["errorCount"]+=1
        SetSession(session)
        print("###请求网址: "+url+":::")
        print(e)
        return None

    return res.text

# get文件
def GetFile(url,headers=None):
    session=GetSession()
    try:
        if(headers is None): headers={"User-Agent":UserAgent}
        res = session["http"].get(url,headers=headers, proxies=session["proxy"],stream=True)
        SetSession(session)
    except RequestException as e:
        session["errorCount"]+=1
        SetSession(session)
        print("###请求网址: "+url+":::")
        print(e)
        return None
    return res


# 获取json字段，返回字段名和类型
# countPre预留字节数
def GetJsonFields(jsonObj,encoding="utf-8",countPre=10):
    fields={}
    if jsonObj is None: return fields # None则返回空
    if isinstance(jsonObj,list):
        if len(jsonObj)==0: return fields # 空列表则返回空
        else:isList=True
    else: isList=False
    if isList: obj= jsonObj[0] # 读取第一个元素
    for key in obj:
        field={"type":"varchar","size":countPre}
        fields[key]=field
        if isList:
            for it in jsonObj:
                text=it[key]
                if(text is None):continue
                countSrc= field["size"]
                count=len(text.encode(encoding))+countPre # 中文3字节，英文1字节，预留字节
                field["size"]=count if countSrc==None else (count if count>countSrc else countSrc)
        else:
            text=str(jsonObj[key])
            count=len(text.encode(encoding))+countPre # 中文3字节，英文1字节，预留字节
            field["size"]=count
    return fields

# 获取mysql表字段，返回字段名和类型
# countPre预留字节数
def GetMysqlFields(dbconfigOrConnect,tableName):
    funName="getMysqlFields: "
    fields={}
    if tableName is None: 
        print(funName+"表名不能为空")
        return fields
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return fields
    cursor = conn.cursor()
    try:
        sql = f"SHOW COLUMNS FROM {tableName}"
        cursor.execute(sql)
        results = cursor.fetchall()
        for row in results:
            field={"type":row[1],"size":None}
            typeMatch=re.match(r'(\w+)(\((\d+)\))?',row[1])
            if typeMatch:
                field["type"]=typeMatch.group(1)
                if typeMatch.group(3):
                    field["size"]=int(typeMatch.group(3))
            fields[row[0]]=field
    except Exception as e:
        print(e)
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)
        return fields

# 更新mysql表字段长度
def UpdateMysqlFieldsSize(dbconfigOrConnect,tableName,fields):
    funName="UpdateMysqlFieldsSize: "
    fieldsMysql=GetMysqlFields(dbconfigOrConnect,tableName)
    # 未返回字段时, 创建表
    if(fieldsMysql=={}):
        CreateMysqlTable(dbconfigOrConnect,tableName,fields)
        return fields
    # 连接数据库
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return
    cursor = conn.cursor()
    try:
        for key in fields:
            isExistMysql=key in fieldsMysql
            if(isExistMysql):
                typeMysql=fieldsMysql[key]["type"].lower()
                sizeMysql=fieldsMysql[key]["size"]
            else:
                typeMysql=None
                sizeMysql=None
            typeJson=fields[key]["type"].lower()
            sizeJson=fields[key]["size"]
            # 要更新字段值长度为None或者小于数据库中对应字段的长度，则跳过更新
            if sizeJson is None or (sizeMysql is not None and sizeMysql>=sizeJson):continue
            # 生成更新sql
            if key in fieldsMysql:
                # 扩展字段长度
                sql = f"ALTER TABLE {tableName} MODIFY COLUMN {key} {typeMysql}({sizeJson})"
            else:
                sql=f"ALTER TABLE {tableName} ADD COLUMN {key} {typeJson}({sizeJson})"
            print(funName+sql)
            cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(funName+str(e))
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 创建mysql表
def CreateMysqlTable(dbconfigOrConnect,tableName,fields):
    funName="CreateMysqlTable: "
    if tableName is None: 
        print(funName+"表名不能为空")
        return
    if fields is None or len(fields)==0:
        print(funName+"字段不能为空")
        return
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return
    cursor = conn.cursor()
    try:
        fieldStrs=["id int NOT NULL AUTO_INCREMENT PRIMARY KEY"]
        for key in fields:
            if(key.lower()=="id"):continue
            type=fields[key]["type"].lower()
            size=fields[key]["size"]
            if type in ["varchar","char"]:
                size=size if size is not None and size>0 else 255
                fieldStrs.append(f"{key} {type}({size})")
            elif type in ["int","bigint","smallint","tinyint","mediumint"]:
                fieldStrs.append(f"{key} {type}")
            elif type in ["float","double","decimal"]:
                fieldStrs.append(f"{key} {type}(10,2)")
            elif type in ["date","datetime","timestamp"]:
                fieldStrs.append(f"{key} {type}")
            elif type in ["text","tinytext","mediumtext","longtext"]:
                fieldStrs.append(f"{key} {type}")
            else:
                fieldStrs.append(f"{key} varchar(255)")
        fieldStr=",".join(fieldStrs)
        sql = f"CREATE TABLE IF NOT EXISTS {tableName} ({fieldStr}) ENGINE=InnoDB DEFAULT CHARSET=utf8"
        print(sql)
        cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(e)
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 获取mysql连接
def GetMysqlConnect(dbconfigOrConnect):
    if dbconfigOrConnect is None : return None
    if isinstance(dbconfigOrConnect,dict):
        dbconfigOrConnect["charset"]=dbconfigOrConnect.get("charset", 'utf8')
        conn = pymysql.connect(**dbconfigOrConnect)
    else:
        conn=dbconfigOrConnect
    return conn
# 关闭mysql连接:只有dbconfigOrConnect为dict时才关闭连接，否则跳过
def CloseMysqlConnect(conn,dbconfigOrConnect=None):
    if conn is None: return
    if isinstance(dbconfigOrConnect,dict):
        conn.close()

# 保存json到mysql数据库对应表中
def JsonToMysql(jsonObj, dbconfigOrConnect,tableName, isAppend=False):
    funName="JsonToMysql: "
    fieldsJson=GetJsonFields(jsonObj)
    conn = GetMysqlConnect(dbconfigOrConnect)
    fieldsMysql=UpdateMysqlFieldsSize(conn,tableName,fieldsJson)
    conn.commit()
    # 更新数据
    cursor = conn.cursor()
    data=[jsonObj] if isinstance(jsonObj,dict) else jsonObj
    try:
        if not isAppend:
            cursor.execute(f"DELETE FROM {tableName} where 1=1") # 清空数据
        i=0
        for item in data:
            keys = item.keys()
            values = tuple(item.values())
            sql = f"INSERT INTO {tableName} ({','.join(keys)}) VALUES {values}"
            cursor.execute(sql)  # 参数化查询更安全
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(funName+str(e))
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 从mysql表中分页读取记录并返回json对象
def MysqlToJson(dbconfigOrConnect,tableName,where=None,page=1,pageSize=100):
    funName="MysqlToJson: "
    if tableName is None: 
        print(funName+"表名不能为空")
        return []
    conn=GetMysqlConnect(dbconfigOrConnect)
    if conn is None: 
        print(funName+"数据库连接失败")
        return []
    cursor = conn.cursor(pymysql.cursors.DictCursor)
    try:
        offset=(page-1)*pageSize
        sql = f"SELECT * FROM {tableName} " + (f" WHERE {where} " if where is not None else "") + f" LIMIT {offset},{pageSize}"
        cursor.execute(sql)
        results = cursor.fetchall()
        return results
    except Exception as e:
        print(funName+str(e))
        return []
    finally:
        cursor.close()
        CloseMysqlConnect(conn,dbconfigOrConnect)

# 保存json到Excel文件中
def JsonToExcel(jsonObj,dirPath,name,isAppend=False):
    funName="jsonToExcel: "
    if not os.path.exists(dirPath):
        print(funName+"目录不存在")
        return
    data=[jsonObj] if isinstance(jsonObj,dict) else jsonObj
    if(len(data)==0):
        print(funName+"没有数据")
        return
    # 创建Excel工作簿和工作表
    excelFile=dirPath+name+".xlsx"
    if isAppend and os.path.exists(excelFile):
        wb=openpyxl.load_workbook(excelFile)
        ws = wb.active
        startRow=ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        startRow=2
        ws.title = name
    # 开始写数据
    p2e=openpyxl.utils.units.pixels_to_EMU
    cellWidth=120 # 单位: 像素
    cellHeight=50 # 单位: 像素
    cellRowHeight=GetRowHeightByPx(cellHeight) # 单位: 字符
    alg=openpyxl.styles.Alignment
    # 写入字段标题数据
    row=1
    col=1
    alignHeader=alg(horizontal='center', vertical='center')
    for key in data[0]:
        addrCol=openpyxl.utils.get_column_letter(col)
        addrRef=addrCol+str(row)
        cell=ws[addrRef]
        cell.value=key
        ws.column_dimensions[addrCol].width = GetColWidthByPx(cellWidth)
        ws.row_dimensions[row].height = cellRowHeight
        cell.alignment = alignHeader
        col+=1
    # 写入记录数据
    align=alg(horizontal='left', vertical='center')
    row=startRow
    for it in data:
        ws.row_dimensions[row].height = cellRowHeight
        col=1
        for key in it:
            val=it[key]
            addCol=openpyxl.utils.get_column_letter(col)
            addrRef=addCol+str(row)
            cell=ws[addrRef]
            if(val is None or val==''):
                cell.value = ""
                cell.alignment = align
            else:
                if(str(key).find("local_image_")==0):
                    cell.value = ""
                    img=GetExcelImage(ws,val,cellHeight,cellHeight,row,col)
                    ws.add_image(img)
                    cell.hyperlink = Hyperlink(ref=addrRef, target=f"file:///{val}")
                    cell.style="Hyperlink"
                    cell.alignment = align
                else:
                    if(str(key).find("local_file_")==0):
                        cell.value = "打开文件"
                        cell.hyperlink = Hyperlink(ref=addrRef,target=f"file:///{val}")
                        cell.style="Hyperlink"
                        cell.alignment = alignHeader
                    else:
                        cell.value = val
                        cell.alignment = align
            
            col+=1
        row+=1
    # 保存Excel文件
    wb.save(excelFile)
    print(funName+"转换完成,"+excelFile)
def GetExcelImage(ws,urlOrPath,cellWidth,cellHeight,row,col,endRow=None,endCol=None):
    endRow=row if endRow is None else endRow
    endCol=col if endCol is None else endCol
    img=openpyxl.drawing.image.Image(urlOrPath)
    if img.width>img.height:
        img.height=cellHeight * img.height / img.width
        img.width=cellWidth
    else:
        img.width=cellWidth * img.width / img.height
        img.height=cellHeight
    w, h = img.width, img.height
    x1, y1, w1, h1 = GetExcelAbsolute(ws,row,col)
    x2, y2, w2, h2 = GetExcelAbsolute(ws,endRow,endCol)
    x = (x2 + x1 - w1 - w) // 2
    y = (y2 + y1 - h1 - h) // 2
    p2e=openpyxl.utils.units.pixels_to_EMU
    pos = openpyxl.drawing.xdr.XDRPoint2D(p2e(x), p2e(y))
    size = openpyxl.drawing.xdr.XDRPositiveSize2D(p2e(w), p2e(h))
    img.anchor=img.anchor = AbsoluteAnchor(pos=pos,ext=size)
    return img
# 获取单元格的右下方绝对位置（单位：像素），及单元格的宽高
def GetExcelAbsolute(ws, row, col):
        x = 0
        y = 0
        gl=openpyxl.utils.get_column_letter
        # get_column_letter(int)把整数转换为Excel中的列索引
        col_letter = gl(col)
        # 获取每列的列宽
        width = ws.column_dimensions[col_letter].width
        # 计算第一列到目标列的总宽
        for i in range(col):
            col_letter = gl(i + 1)
            fcw = ws.column_dimensions[col_letter].width
            x += fcw
		# 如果Excel中高为默认值时，openpyxl却没有值为NoneValue，这一点我很奇怪。
        if not ws.row_dimensions[row].height:
            ws.row_dimensions[col].height = 13.5
            height = 13.5  # Excel默认列宽为13.5
        else:
            height = ws.row_dimensions[row].height
        # 计算第一行到目标行的总高
        for j in range(row):
            if not ws.row_dimensions[j + 1].height:
                ws.row_dimensions[j + 1].height = 13.5
                fch = 13.5
            else:
                fch = ws.row_dimensions[j + 1].height
            y += fch 
        # 把高单位转换为像素
        height = GetPxByRowHeight(height)
        # 把宽单位转换为像素
        width = GetPxByColWidth(width)
        x = GetPxByColWidth(x)
        y = GetPxByRowHeight(y)
        return x, y, width, height
def GetColWidthByPx(px):return px/7
def GetPxByColWidth(width):return width*7
def GetRowHeightByPx(px):return px*0.75
def GetPxByRowHeight(width):return width/0.75

# 保存json对文件
# page>=1表示分页，会创建name文件夹;0表示不分页,不会创建name文件夹,会替换同名文件;
def JsonToFile(jsonObj,dirPath,name,page,encoding="utf-8"):
    # 格式化json字符串
    rst=Json.dumps(jsonObj,indent=2,ensure_ascii=False)
    # 写文件
    fileName=GetJsonFileName(page,dirPath,name)
    open(fileName,"w",encoding=encoding).write(rst)
# 判断json文件是否存在
# page从1开始,0表示不分页
def ExistJsonFile(dirPath,name,page):
    filePath=GetJsonFileName(page,dirPath,name)
    return os.path.exists(filePath)

# Html保存到文件
# page>=1表示分页，会创建name文件夹;0表示不分页,不会创建name文件夹,会替换同名文件;
def SrcCodeToFile(data,dirPath,name,page,encoding="utf-8"):
    filename=GetSrcCodeFileName(page,dirPath,name)
    open(filename,"w",encoding=encoding).write(data)
# 判断html文件是否存在
# page从1开始,0表示不分页
def ExistSrcCodeFile(dirPath,name,page=0):
    filename=GetSrcCodeFileName(page,dirPath,name)
    return os.path.exists(filename)
# 读取html文件
# page从1开始,0表示不分页
def GetSrcCodeFile(dirPath,name,page=0,encoding="utf-8"):
    filename=GetSrcCodeFileName(page,dirPath,name)
    rst=open(filename,"r",encoding=encoding).read()
    return rst

# 获取Json目录路径
def GetSavePath(dirPath,saveName):
    return dirPath+saveName+os.sep
# 获取Json目录路径
def GetJsonDirPath(dirPath,name):
    return GetSavePath(dirPath,name)+"jsons"+os.sep
# 获取Json文件完整名称: name为None时,dirPath认为Json目录, 否则爬取的保存目录
def GetJsonFileName(page,dirPath,name=None):
    if(name==None): dir=dirPath
    else: dir=GetJsonDirPath(dirPath,name)
    filename=dir+str(page)+".json"
    return filename
# 获取源码目录路径
def GetSrcCodeDirPath(dirPath,name):
    return GetSavePath(dirPath,name)+"srccodes"+os.sep
# 获取源码文件完整名称: name为None时,dirPath认为Json目录, 否则爬取的保存目录
def GetSrcCodeFileName(page,dirPath,name=None):
    if(name==None): dir=dirPath
    else: dir=GetSrcCodeDirPath(dirPath,name)
    filename=dir+str(page)+".html"
    return filename
# 获取下载图片目录路径
def GetDownloadImageDirPath(dirPath,name):
    return GetSavePath(dirPath,name)+"images"+os.sep
# 获取下载图片目录路径
def GetDownloadFileDirPath(dirPath,name):
    return GetSavePath(dirPath,name)+"files"+os.sep

# 将目录下的json文件合并或分别保存到Excel文件
def JsonDirToExcel(dirPath,name,encoding="utf-8",isMerge=False):
    funName="jsonFileToExcelFile: "
    #读取json文件列表
    jsonDir=GetJsonDirPath(dirPath,name)
    if not os.path.exists(jsonDir):
        print(funName+"目录不存在"+jsonDir)
        return
    savePath=GetSavePath(dirPath,name)
    files=os.listdir(jsonDir)
    #转换Json数据
    allData=[]
    for file in files:
        if not file.endswith(".json"):continue
        filePath=jsonDir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        if isMerge:
            allData+=jsonObj
        else:
            # 分别保存
            JsonToExcel(jsonObj,savePath,file.replace(".json",""))
    # 合并保存
    if isMerge:
        if(len(allData)==0):
            print(funName+"没有数据")
            return
        JsonToExcel(allData,savePath,name)

# 将目录下的json文件合并保存到mysql
def JsonDirToMysql(dirPath, dbconfigOrConnect,name,encoding="utf-8"):
    funName="jsonDirToMysql: "
    jsonDir=GetJsonDirPath(dirPath,name)
    if not os.path.exists(jsonDir):
        print(funName+"目录不存在"+jsonDir)
        return
    files=os.listdir(jsonDir)
    allData=[]
    for file in files:
        if not file.endswith(".json"):continue
        filePath=jsonDir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        allData+=jsonObj
    if(len(allData)==0):
        print(funName+"没有数据")
        return
    JsonToMysql(allData,dbconfigOrConnect,name)

# 从mysql读取数据并保存到json文件
def MysqlToJsonDir(dbconfigOrConnect,dirPath,name,pageSize=100,encoding="utf-8"):
    funName="MysqlToJsonDir: "
    jsonDir=GetJsonDirPath(dirPath,name)
    CreatePathDirs(jsonDir)
    page=1
    while True:
        list=MysqlToJson(dbconfigOrConnect,name,None,page,pageSize)
        if(list is None or len(list)==0): break
        JsonToFile(list,dirPath,name,page)
        print(funName+"第"+str(page)+"页, 保存完毕")
        page+=1

# 从Excel文件读取数据并保存到json文件
def ExcelToJsonDir(dirPath,name,pageSize=100,encoding="utf-8"):
    funName="ExcelToJsonDir: "
    # Excel文件不存在，则跳过
    excelFile=dirPath+name+".xlsx"
    if not os.path.exists(excelFile):
        print(funName+"文件不存在"+excelFile)
        return
    # Excel文件没有数据，则跳过
    wb=openpyxl.load_workbook(excelFile)
    ws = wb.active
    maxRow=ws.max_row
    if maxRow<2:
        print(funName+"没有数据")
        return
    # 创建Json目录
    jsonDir=GetJsonDirPath(dirPath,name)
    CreatePathDirs(jsonDir)
    # 转换数据
    headers=[]
    for col in range(1,ws.max_column+1):
        headers.append(ws.cell(row=1,column=col).value)
    page=1
    isLoop=True
    while isLoop:
        list=[]
        startRow=(page-1)*pageSize+2
        endRow=startRow+pageSize-1
        if startRow>maxRow:
            isLoop=False
            continue
        if endRow>maxRow:endRow=maxRow
        for row in range(startRow,endRow+1):
            item={}
            for col in range(1,ws.max_column+1):
                key=headers[col-1]
                value=ws.cell(row=row,column=col).value
                item[key]=str(value).strip() if value is not None else ""
            list.append(item)
        JsonToFile(list,dirPath,name,page)
        print(funName+"第"+str(page)+"页, 保存完毕")
        page+=1

# 下载图片,返回None表示下载失败
def DownloadImage(url,dirPath,name):
    filename=dirPath+name
    if(os.path.exists(filename)): return filename
    try:
        img = PIL.Image.open(GetFile(url).raw)
        img.save(filename)
    except Exception as e:
        if IsDebug: print("DownloadImage: 下载失败,"+filename+ "\n" +str(e))
        return None
    return filename

# 下载文件,返回None表示下载失败
def DownloadFile(url,dirPath,name,cacheSize=8192):
    filename=dirPath+name
    if(os.path.exists(filename)): return filename
    try:
        with GetFile(url) as res:
            totalSize = int(res.headers.get('content-length', 0))
            with open(filename, 'wb') as file:
                for chunk in res.iter_content(chunk_size=cacheSize):
                    file.write(chunk)
    except Exception as e:
        if IsDebug: print("DownloadImage: 下载失败,"+filename+ "\n" +str(e))
        return None
    return filename

# 根据jsonObj列表，下载图片列表,返回完成数和总数
def DownloadImageByJson(jsonObj,fields,dirPath,startCount=0):
    funName="DownloadImageByJson: "
    data=[jsonObj] if isinstance(jsonObj,dict) else jsonObj
    doneCount=0
    count=startCount
    # 下载所有图片
    for field in fields:
        fun=fields.get(field,None)
        if(fun==None): continue
        for it in data:
            filedNewName="local_image_"+field
            it[filedNewName]=""
            val=it.get(field,None)
            if(val==None or val==""): continue
            count+=1
            filename=DownloadImage(val,dirPath,fun(it,count))
            if(filename==None): continue
            doneCount+=1
            it[filedNewName]=filename
    return doneCount,count-startCount

# 根据jsonObj列表，下载文件列表,返回完成数和总数
def DownloadFileByJson(jsonObj,fields,dirPath,startCount=0):
    funName="DownloadFileByJson: "
    data=[jsonObj] if isinstance(jsonObj,dict) else jsonObj
    # 保存目录
    doneCount=0
    count=startCount
    # 下载所有图片
    for field in fields:
        fun=fields.get(field,None)
        if(fun==None): continue
        for it in data:
            filedNewName="local_file_"+field
            it[filedNewName]=""
            val=it.get(field,None)
            if(val==None or val==""): continue
            count+=1
            filename=DownloadImage(val,dirPath,fun(it,count))
            if(filename==None): continue
            doneCount+=1
            it[filedNewName]=filename
    return doneCount,count-startCount

# 根据json文件夹中的json文件，下载图片列表,返回完成数和总数
def DownloadImageByJsonDir(dirPath,name,fieldFuncs,encoding="utf-8"):
    # 参数错误退出
    if(fieldFuncs is None or fieldFuncs=={}): return None
    funName="DownloadImageByJsonDir: "
    jsonDir=GetJsonDirPath(dirPath,name)
    if not os.path.exists(jsonDir):
        print(funName+"目录不存在"+jsonDir)
        return
    files=os.listdir(jsonDir)
    totalCount=0
    totalDoneCount=0
    downloadDir=GetDownloadImageDirPath(dirPath,name)
    CreatePathDirs(downloadDir)
    for file in files:
        if not file.endswith(".json"):continue
        filePath=jsonDir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        doneCount,count=DownloadImageByJson(jsonObj,fieldFuncs,downloadDir,totalCount)
        totalCount+=count
        totalDoneCount+=doneCount
        page=int(file.replace(".json",""))
        print(f",第{str(page)}页 下载{doneCount}/{count}个图片")
        JsonToFile(jsonObj,dirPath,name,page,encoding)
    return totalDoneCount,totalCount

# 根据json文件夹中的json文件，下载文件列表,返回完成数和总数
def DownloadFileByJsonDir(dirPath,name,fieldFuncs,encoding="utf-8"):
    # 参数错误退出
    if(fieldFuncs is None or fieldFuncs=={}):
        return None
    funName="DownloadFileByJsonDir: "
    jsonDir=GetJsonDirPath(dirPath,name)
    if not os.path.exists(jsonDir):
        print(funName+"目录不存在"+jsonDir)
        return
    files=os.listdir(jsonDir)
    totalCount=0
    totalDoneCount=0
    downloadDir=GetDownloadFileDirPath(dirPath,name)
    CreatePathDirs(downloadDir)
    for file in files:
        if not file.endswith(".json"):continue
        filePath=jsonDir+file
        jsonStr=open(filePath,"r",encoding=encoding).read()
        jsonObj=Json.loads(jsonStr)
        doneCount,count=DownloadFileByJson(jsonObj,fieldFuncs,downloadDir,totalCount)
        totalCount+=count
        totalDoneCount+=doneCount
        page=int(file.replace(".json",""))
        print(f",第{str(page)}页 下载{doneCount}/{count}个文件")
        JsonToFile(jsonObj,dirPath,name,page,encoding)
    return totalDoneCount,totalCount

# 创建路径下所有目录
def CreatePathDirs(path):
    pathabs = Path(path).absolute()
    pathabs.mkdir(parents=True, exist_ok=True)