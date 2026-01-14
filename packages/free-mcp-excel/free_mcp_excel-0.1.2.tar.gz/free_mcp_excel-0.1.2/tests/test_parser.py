"""
ExcelParser测试用例
"""
import pytest
import base64
import openpyxl
import io
from free_mcp_excel.parser import ExcelParser
from free_mcp_excel.utils import excel_to_base64


@pytest.fixture
def parser():
    """创建解析器实例"""
    return ExcelParser()


@pytest.fixture
def test_file_xlsx():
    """创建测试用的.xlsx文件（Base64）"""
    # 创建简单的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "列1"
    ws["B1"] = "列2"
    ws["A2"] = "值1"
    ws["B2"] = "值2"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_content = buffer.read()
    wb.close()
    
    return base64.b64encode(file_content).decode("utf-8")


class TestExcelParser:
    """ExcelParser测试类"""
    
    def test_read_sheet_names_xlsx(self, parser, test_file_xlsx):
        """测试读取.xlsx文件的工作表名称"""
        result = parser.read_sheet_names(test_file_xlsx)
        assert result["status"] == "success"
        assert "sheet_names" in result["data"]
        assert len(result["data"]["sheet_names"]) > 0
        assert result["data"]["file_format"] == ".xlsx"
    
    def test_read_sheet_names_invalid_format(self, parser):
        """测试无效格式文件"""
        invalid_file = base64.b64encode(b"invalid content").decode("utf-8")
        result = parser.read_sheet_names(invalid_file)
        assert result["status"] == "error"
        assert "INVALID_FILE_FORMAT" in result["error"]["code"]
    
    def test_read_sheet_data_full(self, parser, test_file_xlsx):
        """测试读取整个工作表数据"""
        result = parser.read_sheet_data(test_file_xlsx, sheet="TestSheet")
        assert result["status"] == "success"
        assert "headers" in result["data"]
        assert "data_rows" in result["data"]
        assert result["data"]["sheet_name"] == "TestSheet"
    
    def test_read_sheet_data_range(self, parser, test_file_xlsx):
        """测试读取指定范围数据"""
        result = parser.read_sheet_data(
            test_file_xlsx,
            sheet="TestSheet",
            range="A1:B2"
        )
        assert result["status"] == "success"
        assert len(result["data"]["data_rows"]) <= 1
    
    def test_read_cell_data_single(self, parser, test_file_xlsx):
        """测试读取单个单元格"""
        result = parser.read_cell_data(
            test_file_xlsx,
            sheet="TestSheet",
            cell="A1"
        )
        assert result["status"] == "success"
        assert "cells" in result["data"]
        assert len(result["data"]["cells"]) == 1
        assert result["data"]["cells"][0]["address"] == "A1"
    
    def test_read_cell_data_range(self, parser, test_file_xlsx):
        """测试读取单元格范围"""
        result = parser.read_cell_data(
            test_file_xlsx,
            sheet="TestSheet",
            cell="A1:B2"
        )
        assert result["status"] == "success"
        assert "cells" in result["data"]
        assert len(result["data"]["cells"]) >= 1
    
    def test_read_cell_formula(self, parser):
        """测试读取单元格公式"""
        # 创建包含公式的文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
        wb.close()
        
        result = parser.read_cell_formula(file_base64, "Sheet", "A3")
        assert result["status"] == "success"
        assert "formula" in result["data"]
    
    def test_read_merged_cells(self, parser):
        """测试读取合并单元格"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "合并值"
        ws.merge_cells("A1:B1")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
        wb.close()
        
        result = parser.read_merged_cells(file_base64)
        assert result["status"] == "success"
        assert "merged_cells" in result["data"]
    
    def test_get_workbook_info(self, parser, test_file_xlsx):
        """测试获取工作簿信息"""
        result = parser.get_workbook_info(test_file_xlsx)
        assert result["status"] == "success"
        assert "file_format" in result["data"]
        assert "sheet_count" in result["data"]
        assert "file_size_bytes" in result["data"]

