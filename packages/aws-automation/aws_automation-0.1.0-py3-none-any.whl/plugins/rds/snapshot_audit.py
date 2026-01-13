"""
plugins/rds/snapshot_audit.py - RDS Snapshot 미사용 분석

오래된 수동 RDS/Aurora 스냅샷 탐지

분석 기준:
- 수동 스냅샷만 (자동 스냅샷 제외)
- 14일 이상 경과

비용:
- RDS: $0.02/GB-월
- Aurora: $0.021/GB-월

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_rds_snapshot_monthly_cost

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "rds:DescribeDBSnapshots",
        "rds:DescribeDBClusterSnapshots",
    ],
}


# =============================================================================
# 상수
# =============================================================================

# 오래된 스냅샷 기준 (일)
OLD_SNAPSHOT_DAYS = 14


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    OLD = "old"  # 오래됨
    NORMAL = "normal"  # 최근


class SnapshotType(Enum):
    """스냅샷 유형"""

    RDS = "rds"
    AURORA = "aurora"


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class RDSSnapshotInfo:
    """RDS Snapshot 정보"""

    id: str
    db_identifier: str
    snapshot_type: SnapshotType
    engine: str
    engine_version: str
    status: str
    create_time: datetime | None
    allocated_storage_gb: int
    encrypted: bool
    arn: str
    tags: dict[str, str]

    # 메타
    account_id: str = ""
    account_name: str = ""
    region: str = ""

    # 비용
    monthly_cost: float = 0.0

    @property
    def age_days(self) -> int:
        """스냅샷 나이 (일)"""
        if not self.create_time:
            return 0
        now = datetime.now(timezone.utc)
        delta = now - self.create_time.replace(tzinfo=timezone.utc)
        return delta.days

    @property
    def is_old(self) -> bool:
        """오래된 스냅샷 여부"""
        return self.age_days >= OLD_SNAPSHOT_DAYS


@dataclass
class RDSSnapshotFinding:
    """Snapshot 분석 결과"""

    snapshot: RDSSnapshotInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class RDSSnapshotAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[RDSSnapshotFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    old_count: int = 0
    normal_count: int = 0

    # 용량/비용
    total_size_gb: int = 0
    old_size_gb: int = 0
    old_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_rds_snapshots(session, account_id: str, account_name: str, region: str) -> list[RDSSnapshotInfo]:
    """RDS 수동 스냅샷 수집 (RDS + Aurora)"""
    from botocore.exceptions import ClientError

    snapshots = []

    try:
        rds = get_client(session, "rds", region_name=region)

        # RDS 인스턴스 스냅샷 (수동)
        try:
            paginator = rds.get_paginator("describe_db_snapshots")
            for page in paginator.paginate(SnapshotType="manual"):
                for data in page.get("DBSnapshots", []):
                    if data.get("Status") != "available":
                        continue

                    snap = _parse_rds_snapshot(data, rds, account_id, account_name, region)
                    if snap:
                        snapshots.append(snap)
        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"[yellow]{account_name}/{region} RDS 스냅샷 수집 오류: {error_code}[/yellow]")

        # Aurora 클러스터 스냅샷 (수동)
        try:
            paginator = rds.get_paginator("describe_db_cluster_snapshots")
            for page in paginator.paginate(SnapshotType="manual"):
                for data in page.get("DBClusterSnapshots", []):
                    if data.get("Status") != "available":
                        continue

                    snap = _parse_aurora_snapshot(data, rds, account_id, account_name, region)
                    if snap:
                        snapshots.append(snap)
        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code", "Unknown")
            if not is_quiet():
                console.print(f"[yellow]{account_name}/{region} Aurora 스냅샷 수집 오류: {error_code}[/yellow]")

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"[yellow]{account_name}/{region} 수집 오류: {error_code}[/yellow]")

    return snapshots


def _parse_rds_snapshot(data: dict, rds, account_id: str, account_name: str, region: str) -> RDSSnapshotInfo | None:
    """RDS 스냅샷 파싱"""
    try:
        arn = data.get("DBSnapshotArn", "")
        allocated_storage = data.get("AllocatedStorage", 0)
        monthly_cost = get_rds_snapshot_monthly_cost(region, allocated_storage, is_aurora=False)

        # 태그 조회
        tags = {}
        from botocore.exceptions import ClientError

        try:
            tag_response = rds.list_tags_for_resource(ResourceName=arn)
            tags = {t["Key"]: t["Value"] for t in tag_response.get("TagList", []) if not t["Key"].startswith("aws:")}
        except ClientError:
            pass

        return RDSSnapshotInfo(
            id=data.get("DBSnapshotIdentifier", ""),
            db_identifier=data.get("DBInstanceIdentifier", ""),
            snapshot_type=SnapshotType.RDS,
            engine=data.get("Engine", ""),
            engine_version=data.get("EngineVersion", ""),
            status=data.get("Status", ""),
            create_time=data.get("SnapshotCreateTime"),
            allocated_storage_gb=allocated_storage,
            encrypted=data.get("Encrypted", False),
            arn=arn,
            tags=tags,
            account_id=account_id,
            account_name=account_name,
            region=region,
            monthly_cost=monthly_cost,
        )
    except Exception:
        return None


def _parse_aurora_snapshot(data: dict, rds, account_id: str, account_name: str, region: str) -> RDSSnapshotInfo | None:
    """Aurora 스냅샷 파싱"""
    try:
        arn = data.get("DBClusterSnapshotArn", "")
        allocated_storage = data.get("AllocatedStorage", 0)
        monthly_cost = get_rds_snapshot_monthly_cost(region, allocated_storage, is_aurora=True)

        # 태그 조회
        tags = {}
        from botocore.exceptions import ClientError

        try:
            tag_response = rds.list_tags_for_resource(ResourceName=arn)
            tags = {t["Key"]: t["Value"] for t in tag_response.get("TagList", []) if not t["Key"].startswith("aws:")}
        except ClientError:
            pass

        return RDSSnapshotInfo(
            id=data.get("DBClusterSnapshotIdentifier", ""),
            db_identifier=data.get("DBClusterIdentifier", ""),
            snapshot_type=SnapshotType.AURORA,
            engine=data.get("Engine", ""),
            engine_version=data.get("EngineVersion", ""),
            status=data.get("Status", ""),
            create_time=data.get("SnapshotCreateTime"),
            allocated_storage_gb=allocated_storage,
            encrypted=data.get("StorageEncrypted", False),
            arn=arn,
            tags=tags,
            account_id=account_id,
            account_name=account_name,
            region=region,
            monthly_cost=monthly_cost,
        )
    except Exception:
        return None


# =============================================================================
# 분석
# =============================================================================


def analyze_rds_snapshots(
    snapshots: list[RDSSnapshotInfo], account_id: str, account_name: str, region: str
) -> RDSSnapshotAnalysisResult:
    """RDS Snapshot 미사용 분석"""
    result = RDSSnapshotAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for snapshot in snapshots:
        finding = _analyze_single_snapshot(snapshot)
        result.findings.append(finding)
        result.total_size_gb += snapshot.allocated_storage_gb

        if finding.usage_status == UsageStatus.OLD:
            result.old_count += 1
            result.old_size_gb += snapshot.allocated_storage_gb
            result.old_monthly_cost += snapshot.monthly_cost
        else:
            result.normal_count += 1

    result.total_count = len(snapshots)
    return result


def _analyze_single_snapshot(snapshot: RDSSnapshotInfo) -> RDSSnapshotFinding:
    """개별 스냅샷 분석"""

    # 최근 생성
    if not snapshot.is_old:
        return RDSSnapshotFinding(
            snapshot=snapshot,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"최근 생성 ({snapshot.age_days}일)",
            recommendation="모니터링",
        )

    # 오래됨
    if snapshot.allocated_storage_gb >= 500:
        severity = Severity.HIGH
    elif snapshot.allocated_storage_gb >= 100:
        severity = Severity.MEDIUM
    else:
        severity = Severity.LOW

    type_str = "Aurora" if snapshot.snapshot_type == SnapshotType.AURORA else "RDS"
    return RDSSnapshotFinding(
        snapshot=snapshot,
        usage_status=UsageStatus.OLD,
        severity=severity,
        description=f"오래된 {type_str} 스냅샷 ({snapshot.age_days}일, {snapshot.allocated_storage_gb}GB, ${snapshot.monthly_cost:.2f}/월)",
        recommendation="필요 여부 검토 후 삭제",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[RDSSnapshotAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    status_fills = {
        UsageStatus.OLD: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "RDS Snapshot 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "old": sum(r.old_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "old_size": sum(r.old_size_gb for r in results),
        "old_cost": sum(r.old_monthly_cost for r in results),
    }

    stats = [
        ("항목", "값"),
        ("전체 스냅샷", totals["total"]),
        ("오래된 스냅샷", totals["old"]),
        ("최근 스냅샷", totals["normal"]),
        ("전체 용량 (GB)", totals["total_size"]),
        ("오래된 용량 (GB)", totals["old_size"]),
        ("오래된 월 비용 ($)", f"${totals['old_cost']:.2f}"),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "Snapshot ID",
        "DB Identifier",
        "Type",
        "Usage",
        "Engine",
        "Size (GB)",
        "Age (days)",
        "Monthly Cost ($)",
        "Encrypted",
        "Created",
        "Recommendation",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 오래된 것만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status == UsageStatus.OLD:
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.snapshot.monthly_cost, reverse=True)

    for f in all_findings:
        snap = f.snapshot
        ws2.append(
            [
                snap.account_name,
                snap.region,
                snap.id,
                snap.db_identifier,
                snap.snapshot_type.value.upper(),
                f.usage_status.value,
                f"{snap.engine} {snap.engine_version}",
                snap.allocated_storage_gb,
                snap.age_days,
                round(snap.monthly_cost, 2),
                "Yes" if snap.encrypted else "No",
                snap.create_time.strftime("%Y-%m-%d") if snap.create_time else "",
                f.recommendation,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=6).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 50)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"RDS_Snapshot_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> RDSSnapshotAnalysisResult | None:
    """단일 계정/리전의 RDS 스냅샷 수집 및 분석 (병렬 실행용)"""
    snapshots = collect_rds_snapshots(session, account_id, account_name, region)
    if not snapshots:
        return None
    return analyze_rds_snapshots(snapshots, account_id, account_name, region)


def run(ctx) -> None:
    """RDS Snapshot 미사용 분석 실행"""
    console.print("[bold]RDS Snapshot 미사용 분석 시작...[/bold]")
    console.print(f"  [dim]기준: {OLD_SNAPSHOT_DAYS}일 이상 오래된 수동 스냅샷[/dim]")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="rds")
    all_results: list[RDSSnapshotAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 RDS 스냅샷 없음[/yellow]")
        return

    # 요약
    totals = {
        "total": sum(r.total_count for r in all_results),
        "old": sum(r.old_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "old_size": sum(r.old_size_gb for r in all_results),
        "old_cost": sum(r.old_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 RDS 스냅샷: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["old"] > 0:
        console.print(
            f"  [yellow bold]오래됨: {totals['old']}개 ({totals['old_size']}GB, ${totals['old_cost']:.2f}/월)[/yellow bold]"
        )
    console.print(f"  [green]최근: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("rds-snapshot-audit").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
