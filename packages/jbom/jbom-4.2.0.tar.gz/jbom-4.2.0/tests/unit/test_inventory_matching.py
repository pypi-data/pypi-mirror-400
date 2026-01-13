#!/usr/bin/env python3
"""Unit tests for inventory matching and file loading."""
import unittest
import tempfile
import sys
from pathlib import Path

# Add src directory to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from jbom.common.types import Component
from jbom.processors.inventory_matcher import InventoryMatcher
from jbom import EXCEL_SUPPORT, NUMBERS_SUPPORT


class TestInventoryMatching(unittest.TestCase):
    """Test inventory matching logic.

    Validates:
    - Component to inventory matching by type, package, and value.
    - Priority-based ranking (lower Priority numbers rank higher).
    - No-match scenarios.
    """

    def setUp(self):
        # Create comprehensive test inventory with Priority column
        self.temp_inv = tempfile.NamedTemporaryFile(
            mode="w", delete=False, suffix=".csv"
        )
        csv_content = [
            (
                "IPN,Name,Category,Generic,Package,Value,Tolerance,LCSC,Manufacturer,"
                "MFGPN,Description,Datasheet,Priority"
            ),
            (
                "R001,330R 5%,RES,0603,0603,330R,5%,C25231,UNI-ROYAL,0603WAJ0331T5E,"
                "330Ω 5% 0603,https://example.com/r1,1"
            ),
            (
                "R002,10K 1%,RES,0603,0603,10K,1%,C25232,UNI-ROYAL,0603WAF1002T5E,"
                "10kΩ 1% 0603,https://example.com/r2,2"
            ),
            (
                "C001,100nF,CAP,0603,0603,100nF,10%,C14663,YAGEO,CC0603KRX7R9BB104,"
                "100nF X7R 0603,https://example.com/c1,1"
            ),
            (
                "L001,10uH,IND,0603,0603,10uH,20%,C1608,SUNLORD,SWPA3012S100MT,"
                "10µH 0603,https://example.com/l1,5"
            ),
        ]
        self.temp_inv.write("\n".join(csv_content))
        self.temp_inv.close()

        self.matcher = InventoryMatcher(Path(self.temp_inv.name))

    def tearDown(self):
        Path(self.temp_inv.name).unlink()

    def test_find_matches_resistor(self):
        """Test finding matches for resistors"""
        component = Component("R1", "Device:R", "330R", "PCM_SPCoast:0603-RES")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "330R")
        self.assertEqual(best_item.category, "RES")

    def test_find_matches_capacitor(self):
        """Test finding matches for capacitors"""
        component = Component("C1", "Device:C", "100nF", "PCM_SPCoast:0603-CAP")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "100nF")
        self.assertEqual(best_item.category, "CAP")

    def test_find_matches_inductor(self):
        """Test finding matches for inductors"""
        component = Component("L1", "Device:L", "10uH", "PCM_SPCoast:0603-IND")
        matches = self.matcher.find_matches(component)

        self.assertGreater(len(matches), 0)
        best_item, best_score, _ = matches[0]
        self.assertEqual(best_item.value, "10uH")
        self.assertEqual(best_item.category, "IND")

    def test_no_matches_found(self):
        """Test behavior when no matches are found"""
        component = Component(
            "U1", "MCU:Unknown", "ESP32-NONEXISTENT", "Package_QFP:LQFP-64"
        )
        matches = self.matcher.find_matches(component)

        self.assertEqual(len(matches), 0)

    def test_priority_ranking(self):
        """Test that Priority values are used for ranking (lower Priority = better)"""
        component = Component(
            "R1", "Device:R", "10K", "PCM_SPCoast:0603-RES"
        )  # Should match both R001 and R002
        matches = self.matcher.find_matches(component)

        if len(matches) > 1:
            # First match should have lower priority number (better)
            first_priority = matches[0][0].priority
            second_priority = matches[1][0].priority
            self.assertLessEqual(
                first_priority,
                second_priority,
                "Lower priority numbers should be ranked first",
            )


class TestSpreadsheetSupport(unittest.TestCase):
    """Test Excel and Numbers file support"""

    def setUp(self):
        # Create test data that matches expected inventory structure
        self.test_data = {
            "IPN": ["R001", "C001"],
            "Name": ["330R 5%", "100nF X7R"],
            "Category": ["RES", "CAP"],
            "Package": ["0603", "0603"],
            "Value": ["330R", "100nF"],
            "LCSC": ["C25231", "C14663"],
            "Manufacturer": ["UNI-ROYAL", "YAGEO"],
            "MFGPN": ["0603WAJ0331T5E", "CC0603KRX7R9BB104"],
            "Priority": [1, 1],
        }

    @unittest.skipUnless(EXCEL_SUPPORT, "openpyxl not available")
    def test_excel_file_loading(self):
        """Test loading inventory from Excel file"""
        import openpyxl

        # Create temporary Excel file
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel.close()

        try:
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            # Write headers
            for col, header in enumerate(self.test_data.keys(), 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row_idx in range(len(self.test_data["IPN"])):
                for col, key in enumerate(self.test_data.keys(), 1):
                    ws.cell(
                        row=row_idx + 2, column=col, value=self.test_data[key][row_idx]
                    )

            wb.save(temp_excel.name)
            wb.close()

            # Test loading
            matcher = InventoryMatcher(Path(temp_excel.name))

            # Verify data was loaded correctly
            self.assertEqual(len(matcher.inventory), 2)

            # Check first item
            r_item = next(
                (item for item in matcher.inventory if item.ipn == "R001"), None
            )
            self.assertIsNotNone(r_item)
            self.assertEqual(r_item.category, "RES")
            self.assertEqual(r_item.value, "330R")
            self.assertEqual(r_item.lcsc, "C25231")
            self.assertEqual(r_item.manufacturer, "UNI-ROYAL")

            # Check second item
            c_item = next(
                (item for item in matcher.inventory if item.ipn == "C001"), None
            )
            self.assertIsNotNone(c_item)
            self.assertEqual(c_item.category, "CAP")
            self.assertEqual(c_item.value, "100nF")
            self.assertEqual(c_item.lcsc, "C14663")

        finally:
            Path(temp_excel.name).unlink()

    def test_unsupported_file_format(self):
        """Test error handling for unsupported file formats"""
        temp_file = tempfile.NamedTemporaryFile(suffix=".txt", delete=False)
        temp_file.write(b"Some text content")
        temp_file.close()

        try:
            with self.assertRaises(ValueError) as context:
                InventoryMatcher(Path(temp_file.name))

            self.assertIn("Unsupported inventory file format", str(context.exception))
            self.assertIn(".txt", str(context.exception))

        finally:
            Path(temp_file.name).unlink()

    @unittest.skipIf(
        EXCEL_SUPPORT, "Testing Excel import error when openpyxl not available"
    )
    def test_excel_import_error(self):
        """Test error handling when openpyxl is not available"""
        temp_excel = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_excel.close()

        try:
            with self.assertRaises(ImportError) as context:
                InventoryMatcher(Path(temp_excel.name))

            self.assertIn("openpyxl package", str(context.exception))

        finally:
            Path(temp_excel.name).unlink()

    @unittest.skipIf(
        NUMBERS_SUPPORT,
        "Testing Numbers import error when numbers-parser not available",
    )
    def test_numbers_import_error(self):
        """Test error handling when numbers-parser is not available"""
        temp_numbers = tempfile.NamedTemporaryFile(suffix=".numbers", delete=False)
        temp_numbers.close()

        try:
            with self.assertRaises(ImportError) as context:
                InventoryMatcher(Path(temp_numbers.name))

            self.assertIn("numbers-parser package", str(context.exception))

        finally:
            Path(temp_numbers.name).unlink()

    def test_csv_still_works(self):
        """Test that CSV loading still works as before"""
        temp_csv = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False)

        # Write CSV content
        csv_content = [
            "IPN,Name,Category,Package,Value,LCSC,Manufacturer,MFGPN,Priority",
            "R001,330R 5%,RES,0603,330R,C25231,UNI-ROYAL,0603WAJ0331T5E,1",
            "C001,100nF X7R,CAP,0603,100nF,C14663,YAGEO,CC0603KRX7R9BB104,1",
        ]
        temp_csv.write("\n".join(csv_content))
        temp_csv.close()

        try:
            # Test loading
            matcher = InventoryMatcher(Path(temp_csv.name))

            # Verify data was loaded correctly
            self.assertEqual(len(matcher.inventory), 2)

            # Check that items are correct
            r_item = next(
                (item for item in matcher.inventory if item.ipn == "R001"), None
            )
            self.assertIsNotNone(r_item)
            self.assertEqual(r_item.value, "330R")

        finally:
            Path(temp_csv.name).unlink()


if __name__ == "__main__":
    unittest.main()
