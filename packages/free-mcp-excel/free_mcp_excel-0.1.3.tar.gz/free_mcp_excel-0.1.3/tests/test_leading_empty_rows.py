"""
测试从首行开始就有空行的情况
"""
import pytest
import base64
import openpyxl
import io
from free_mcp_excel.parser import ExcelParser
from free_mcp_excel.utils import excel_to_base64


@pytest.fixture
def parser():
    """创建解析器实例"""
    return ExcelParser()


@pytest.fixture
def leading_empty_rows_file():
    """创建从首行开始就有空行的测试文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "首行空行测试"
    
    # 第1-10行：空行（不写入任何内容）
    
    # 第11行：标题
    ws["A11"] = "序号"
    ws["B11"] = "名称"
    ws["C11"] = "数值"
    
    # 第12-15行：第一组数据
    for i in range(12, 16):
        ws[f"A{i}"] = i - 11
        ws[f"B{i}"] = f"项目{i-11}"
        ws[f"C{i}"] = (i - 11) * 100
    
    # 第16-20行：空行
    
    # 第21-25行：第二组数据
    for i in range(21, 26):
        ws[f"A{i}"] = i - 15
        ws[f"B{i}"] = f"项目{i-15}"
        ws[f"C{i}"] = (i - 15) * 100
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
    wb.close()
    
    return file_base64


class TestLeadingEmptyRows:
    """首行空行测试类"""
    
    def test_skip_empty_rows_mode(self, parser, leading_empty_rows_file):
        """测试跳过空行模式：应该正确识别表头"""
        result = parser.read_sheet_data(
            leading_empty_rows_file, "首行空行测试", skip_empty_rows=True
        )
        
        assert result["status"] == "success"
        # 应该正确识别第11行（第一个有数据的行）作为表头
        assert result["data"]["headers"] == ["序号", "名称", "数值"]
        # 应该返回9行数据（4行第一组 + 5行第二组）
        assert result["data"]["row_count"] == 9
        # 验证第一行数据
        assert result["data"]["data_rows"][0] == [1, "项目1", 100]
    
    def test_keep_empty_rows_mode(self, parser, leading_empty_rows_file):
        """测试保留空行模式：应该智能识别第一个非空行作为表头"""
        result = parser.read_sheet_data(
            leading_empty_rows_file, "首行空行测试", skip_empty_rows=False
        )
        
        assert result["status"] == "success"
        # 应该智能识别第11行（第一个非空行）作为表头，而不是第1行（空行）
        assert result["data"]["headers"] == ["序号", "名称", "数值"]
        # 数据行数：第12-15行（4行数据）+ 第16-20行（5行空行）+ 第21-25行（5行数据）= 14行
        assert result["data"]["row_count"] == 14
        # 验证第一行数据（第12行，表头后的第一行）
        assert result["data"]["data_rows"][0] == [1, "项目1", 100]
        # 验证包含空行（第16-20行是空行，在data_rows中的索引4-8）
        assert result["data"]["data_rows"][4] == [None, None, None]  # 第16行
        # 验证第二组数据的第一行（第21行）
        assert result["data"]["data_rows"][9] == [6, "项目6", 600]
    
    def test_all_empty_rows(self, parser):
        """测试所有行都是空行的情况"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "全空行"
        # 不写入任何数据
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
        wb.close()
        
        result = parser.read_sheet_data(file_base64, "全空行", skip_empty_rows=True)
        
        assert result["status"] == "success"
        assert result["data"]["row_count"] == 0
        assert result["data"]["headers"] == []
        assert result["data"]["data_rows"] == []
    
    def test_only_header_no_data(self, parser):
        """测试只有表头没有数据的情况"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "仅表头"
        
        # 第1-5行：空行
        # 第6行：表头
        ws["A6"] = "列1"
        ws["B6"] = "列2"
        ws["C6"] = "列3"
        # 第7行之后：空行
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_base64 = base64.b64encode(buffer.read()).decode("utf-8")
        wb.close()
        
        # 跳过空行模式
        result1 = parser.read_sheet_data(file_base64, "仅表头", skip_empty_rows=True)
        assert result1["status"] == "success"
        assert result1["data"]["headers"] == ["列1", "列2", "列3"]
        assert result1["data"]["row_count"] == 0
        assert result1["data"]["data_rows"] == []
        
        # 保留空行模式
        result2 = parser.read_sheet_data(file_base64, "仅表头", skip_empty_rows=False)
        assert result2["status"] == "success"
        assert result2["data"]["headers"] == ["列1", "列2", "列3"]
        assert result2["data"]["row_count"] == 0  # 只有表头，没有数据行

