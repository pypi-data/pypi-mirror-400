"""
pkg/io/excel/workbook.py - Excel Workbook 래퍼 클래스

일관된 Excel 출력을 위한 Workbook, Sheet, ColumnDef 클래스
및 워크시트 포맷팅 유틸리티 함수들

Note:
    이 모듈은 Lazy Import 패턴을 사용합니다.
    openpyxl 등 무거운 의존성을 실제 사용 시점에만 로드합니다.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal

if TYPE_CHECKING:
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.worksheet import Worksheet

# 문자열 상수와 함수만 모듈 레벨에서 import (openpyxl 필요 없음)
from .styles import (
    NUMBER_FORMAT_CURRENCY,
    NUMBER_FORMAT_INTEGER,
    NUMBER_FORMAT_PERCENT,
    NUMBER_FORMAT_TEXT,
    Styles,
    get_basic_header_style,
    get_center_alignment,
    get_data_font,
    get_header_fill,
    get_header_font,
    get_thin_border,
)


# Alignment 상수는 lazy import (사용 시점에 로드)
def _get_align_center() -> Alignment:
    from .styles import ALIGN_CENTER

    return ALIGN_CENTER


def _get_align_center_wrap() -> Alignment:
    from .styles import ALIGN_CENTER_WRAP

    return ALIGN_CENTER_WRAP


def _get_align_left() -> Alignment:
    from .styles import ALIGN_LEFT

    return ALIGN_LEFT


def _get_align_right_wrap() -> Alignment:
    from .styles import ALIGN_RIGHT_WRAP

    return ALIGN_RIGHT_WRAP


def _get_align_wrap() -> Alignment:
    from .styles import ALIGN_WRAP

    return ALIGN_WRAP


logger = logging.getLogger(__name__)


# Style 타입 정의
StyleType = Literal["data", "center", "wrap", "number", "currency", "percent", "text", "date"]


@dataclass
class ColumnDef:
    """컬럼 정의 클래스

    Attributes:
        header: 헤더 텍스트
        width: 컬럼 너비 (기본값: 15)
        style: 스타일 타입 ("data", "center", "wrap", "number",
               "currency", "percent")
    """

    header: str
    width: int = 15
    style: StyleType = "data"


@dataclass
class SummaryItem:
    """Summary 시트 항목 정의

    Attributes:
        label: 항목 레이블 (왼쪽)
        value: 항목 값 (오른쪽)
        is_header: 섹션 헤더 여부 (병합 + 강조)
        highlight: 강조 색상 ("danger", "warning", "success", None)
    """

    label: str
    value: Any = ""
    is_header: bool = False
    highlight: str | None = None


class SummarySheet:
    """Summary(분석 요약) 시트 래퍼 클래스

    Key-Value 형태의 요약 정보를 표시하는 시트.
    섹션 헤더, 항목, 강조 표시를 지원.

    Usage:
        summary = wb.new_summary_sheet("분석 요약")
        summary.add_title("ALB 로그 분석 보고서")
        summary.add_section("분석 정보")
        summary.add_item("계정 번호", "123456789012")
        summary.add_item("리전", "ap-northeast-2")
        summary.add_section("보안 정보")
        summary.add_item("탐지된 Abuse IP", "15개", highlight="danger")
    """

    def __init__(self, ws: Worksheet, workbook: Workbook):
        self._ws = ws
        self._workbook = workbook
        self._current_row = 1

        # 기본 컬럼 너비 설정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # 줌 비율 85% 설정
        ws.sheet_view.zoomScale = 85

    def add_title(self, title: str) -> SummarySheet:
        """제목 추가 (병합 + 큰 폰트)

        Args:
            title: 제목 텍스트

        Returns:
            self (메서드 체이닝)
        """
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = self._ws
        ws.merge_cells(f"A{self._current_row}:B{self._current_row}")
        cell = ws.cell(row=self._current_row, column=1, value=title)
        cell.font = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")
        cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_thin_border()
        # B열에도 테두리 적용
        ws.cell(row=self._current_row, column=2).border = get_thin_border()
        ws.row_dimensions[self._current_row].height = 40
        self._current_row += 2
        return self

    def add_section(self, section_name: str) -> SummarySheet:
        """섹션 헤더 추가 (병합 + 강조 배경)

        Args:
            section_name: 섹션 이름

        Returns:
            self (메서드 체이닝)
        """
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = self._ws
        ws.merge_cells(f"A{self._current_row}:B{self._current_row}")
        cell = ws.cell(row=self._current_row, column=1, value=section_name)
        cell.font = Font(name="맑은 고딕", size=12, bold=True, color="2F5597")
        cell.fill = PatternFill(start_color="EBF1FA", end_color="EBF1FA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_thin_border()
        ws.cell(row=self._current_row, column=2).border = get_thin_border()
        self._current_row += 1
        return self

    def add_item(
        self,
        label: str,
        value: Any,
        highlight: str | None = None,
        number_format: str | None = None,
    ) -> SummarySheet:
        """항목 추가 (레이블 + 값)

        Args:
            label: 항목 레이블
            value: 항목 값
            highlight: 강조 색상 ("danger", "warning", "success", None)
            number_format: 숫자 포맷 (예: "@" for text)

        Returns:
            self (메서드 체이닝)
        """
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = self._ws
        border = get_thin_border()
        label_font = Font(name="맑은 고딕", size=11, bold=True)
        value_font = Font(name="맑은 고딕", size=11)
        align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 레이블 셀
        label_cell = ws.cell(row=self._current_row, column=1, value=label)
        label_cell.font = label_font
        label_cell.alignment = align
        label_cell.border = border

        # 값 셀
        value_cell = ws.cell(row=self._current_row, column=2, value=value)
        value_cell.font = value_font
        value_cell.alignment = align
        value_cell.border = border

        if number_format:
            value_cell.number_format = number_format

        # 강조 색상 적용
        if highlight:
            highlight_colors = {
                "danger": "FFCCCC",
                "warning": "FFEB9C",
                "success": "C6EFCE",
                "info": "CCE5FF",
            }
            color = highlight_colors.get(highlight, "FFFFFF")
            value_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        self._current_row += 1
        return self

    def add_blank_row(self) -> SummarySheet:
        """빈 행 추가

        Returns:
            self (메서드 체이닝)
        """
        self._current_row += 1
        return self

    def add_list_section(
        self,
        section_name: str,
        items: list[tuple],
        max_items: int = 5,
    ) -> SummarySheet:
        """순위 리스트 섹션 추가 (예: Top 5 URL)

        Args:
            section_name: 섹션 이름
            items: (이름, 값) 튜플 리스트
            max_items: 최대 표시 개수

        Returns:
            self (메서드 체이닝)
        """
        from openpyxl.styles import Alignment, Font

        ws = self._ws
        border = get_thin_border()
        value_font = Font(name="맑은 고딕", size=11)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # 섹션 레이블
        label_cell = ws.cell(row=self._current_row, column=1, value=f"{section_name}:")
        label_cell.font = Font(name="맑은 고딕", size=11, bold=True)
        label_cell.alignment = align_left
        label_cell.border = border
        ws.cell(row=self._current_row, column=2).border = border
        self._current_row += 1

        if not items:
            cell = ws.cell(row=self._current_row, column=1, value="데이터 없음")
            cell.font = value_font
            cell.border = border
            ws.cell(row=self._current_row, column=2).border = border
            self._current_row += 1
        else:
            for i, (name, count) in enumerate(items[:max_items], 1):
                # 이름이 너무 길면 자르기
                display_name = name if len(str(name)) <= 50 else str(name)[:47] + "..."
                name_cell = ws.cell(row=self._current_row, column=1, value=f"{i}. {display_name}")
                name_cell.font = value_font
                name_cell.alignment = align_left
                name_cell.border = border

                # 값 포맷팅
                display_value = f"{count:,}" if isinstance(count, (int, float)) else str(count)
                count_cell = ws.cell(row=self._current_row, column=2, value=display_value)
                count_cell.font = value_font
                count_cell.alignment = align_right
                count_cell.border = border
                self._current_row += 1

        return self

    @property
    def current_row(self) -> int:
        """현재 행 번호"""
        return self._current_row


@dataclass
class Sheet:
    """시트 래퍼 클래스"""

    _ws: Worksheet
    _columns: list[ColumnDef]
    _current_row: int = 2  # 1은 헤더
    _workbook: Workbook | None = field(repr=False, default=None)

    def add_row(
        self,
        values: list[Any],
        style: dict | None = None,
    ) -> int:
        """데이터 행 추가

        Args:
            values: 각 컬럼 값 리스트
            style: 행 스타일 (Styles.warning(), Styles.danger() 등)

        Returns:
            추가된 행 번호
        """
        row_num = self._current_row

        for col_idx, value in enumerate(values, start=1):
            cell = self._ws.cell(row=row_num, column=col_idx, value=value)

            # 컬럼별 스타일 적용
            if col_idx <= len(self._columns):
                col_def = self._columns[col_idx - 1]
                self._apply_cell_style(cell, col_def)

            # 행 스타일 적용 (fill, font)
            if style:
                if style.get("fill"):
                    cell.fill = style["fill"]
                if style.get("font"):
                    cell.font = style["font"]
            else:
                cell.font = get_data_font()

            # 테두리 적용
            cell.border = get_thin_border()

        self._current_row += 1
        return row_num

    def add_summary_row(self, values: list[Any]) -> int:
        """요약 행 추가 (연한 노랑 배경, 볼드)

        Args:
            values: 각 컬럼 값 리스트

        Returns:
            추가된 행 번호
        """
        return self.add_row(values, style=Styles.summary())

    def _apply_cell_style(self, cell, col_def: ColumnDef) -> None:
        """컬럼 정의에 따른 셀 스타일 적용 (자동 줄바꿈 포함)"""
        style_map = {
            "data": (_get_align_wrap(), None),  # 줄바꿈 적용
            "center": (_get_align_center_wrap(), None),  # 줄바꿈 적용
            "wrap": (_get_align_wrap(), None),
            "number": (_get_align_right_wrap(), NUMBER_FORMAT_INTEGER),  # 줄바꿈 적용
            "currency": (_get_align_right_wrap(), NUMBER_FORMAT_CURRENCY),  # 줄바꿈 적용
            "percent": (_get_align_right_wrap(), NUMBER_FORMAT_PERCENT),  # 줄바꿈 적용
            "text": (_get_align_wrap(), NUMBER_FORMAT_TEXT),  # 텍스트 강제 + 줄바꿈
        }

        alignment, number_format = style_map.get(col_def.style, (_get_align_wrap(), None))
        cell.alignment = alignment

        if number_format and cell.value is not None:
            cell.number_format = number_format

    def finalize(self) -> None:
        """시트 마무리 작업 (자동 필터 적용)

        Workbook.save() 시 자동 호출됨
        """
        ws = self._ws
        # 자동 필터 적용 (데이터가 있을 때만)
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    @property
    def row_count(self) -> int:
        """현재 데이터 행 수 (헤더 제외)"""
        return self._current_row - 2


class Workbook:
    """Excel Workbook 래퍼 클래스

    Usage:
        wb = Workbook()

        columns = [
            ColumnDef(header="Volume ID", width=22, style="data"),
            ColumnDef(header="크기(GB)", width=10, style="number"),
            ColumnDef(header="상태", width=12, style="center"),
        ]

        sheet = wb.new_sheet(name="분석 결과", columns=columns)
        sheet.add_row(["vol-12345", 100, "available"])
        sheet.add_summary_row(["합계", 100, "-"])

        wb.save_as(output_dir, "unused_volumes", "ap-northeast-2")
    """

    def __init__(self):
        """Workbook 초기화"""
        from openpyxl import Workbook as _OpenpyxlWorkbook

        self._wb = _OpenpyxlWorkbook()
        # 기본 시트 제거
        if "Sheet" in self._wb.sheetnames:
            del self._wb["Sheet"]
        self._sheets: list[Sheet] = []

    @property
    def styles(self) -> type:
        """스타일 프리셋 접근"""
        return Styles

    def new_sheet(
        self,
        name: str,
        columns: list[ColumnDef],
    ) -> Sheet:
        """새 시트 생성

        Args:
            name: 시트 이름
            columns: 컬럼 정의 리스트

        Returns:
            Sheet 인스턴스
        """
        from openpyxl.utils import get_column_letter

        ws = self._wb.create_sheet(title=name)

        # 헤더 행 설정
        header_font = get_header_font()
        header_fill = get_header_fill()
        border = get_thin_border()

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = _get_align_center()
            cell.border = border

            # 컬럼 너비 설정
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_def.width

        # 헤더 행 고정
        ws.freeze_panes = "A2"

        # 줌 비율 85% 설정
        ws.sheet_view.zoomScale = 85

        sheet = Sheet(_ws=ws, _columns=columns, _workbook=self)
        self._sheets.append(sheet)
        return sheet

    def new_summary_sheet(
        self,
        name: str = "분석 요약",
        position: int = 0,
    ) -> SummarySheet:
        """Summary(분석 요약) 시트 생성

        Key-Value 형태의 요약 정보를 표시하는 시트.

        Args:
            name: 시트 이름
            position: 시트 위치 (0 = 맨 앞)

        Returns:
            SummarySheet 인스턴스
        """
        ws = self._wb.create_sheet(title=name, index=position)
        return SummarySheet(ws, self)

    @property
    def openpyxl_workbook(self):
        """내부 openpyxl Workbook 접근 (고급 사용)"""
        return self._wb

    def save(self, filepath: str | Path) -> Path:
        """워크북 저장

        Args:
            filepath: 저장 경로 (전체 파일 경로)

        Returns:
            저장된 파일 경로
        """
        # 모든 시트 마무리 작업 (자동 필터 등)
        for sheet in self._sheets:
            sheet.finalize()

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(path))
        return path

    def save_as(
        self,
        output_dir: str | Path,
        prefix: str,
        region: str | None = None,
        suffix: str | None = None,
    ) -> Path:
        """규칙에 따라 파일명 생성 후 저장

        파일명 형식: {prefix}_{region}_{날짜}_{suffix}.xlsx

        Args:
            output_dir: 출력 디렉토리
            prefix: 파일명 접두사
            region: 리전 (선택)
            suffix: 파일명 접미사 (선택)

        Returns:
            저장된 파일 경로
        """
        today = datetime.now().strftime("%Y%m%d")
        parts = [prefix]

        if region:
            parts.append(region)

        parts.append(today)

        if suffix:
            parts.append(suffix)

        filename = "_".join(parts) + ".xlsx"
        filepath = Path(output_dir) / filename

        return self.save(filepath)

    def close(self) -> None:
        """워크북 닫기"""
        self._wb.close()


# =============================================================================
# 워크시트 포맷팅 유틸리티 함수
# =============================================================================


def calculate_optimal_column_width(
    worksheet: Worksheet,
    column_letter: str,
    max_width: int = 50,
    min_width: int = 8,
) -> float:
    """컬럼 내용에 따른 최적 너비 계산

    Args:
        worksheet: 대상 워크시트
        column_letter: 컬럼 문자 (예: "A", "B")
        max_width: 최대 너비
        min_width: 최소 너비

    Returns:
        계산된 최적 너비
    """
    try:
        column = worksheet[column_letter]
        max_length = 0

        for cell in column:
            if cell.value:
                cell_text = str(cell.value)
                if "\n" in cell_text:
                    lines = cell_text.split("\n")
                    longest_line = max(lines, key=len)
                    cell_length = len(longest_line)
                else:
                    cell_length = len(cell_text)

                if cell_length > max_length:
                    max_length = cell_length

        optimal_width = max_length + 2
        return max(min(optimal_width, max_width), min_width)

    except Exception as e:
        logger.warning(f"컬럼 너비 계산 실패 ({column_letter}): {e}")
        return min_width


def calculate_optimal_row_height(
    worksheet: Worksheet,
    row_number: int,
    base_height: float = 15,
) -> float:
    """멀티라인 내용에 따른 최적 행 높이 계산

    Args:
        worksheet: 대상 워크시트
        row_number: 행 번호
        base_height: 기본 행 높이

    Returns:
        계산된 최적 높이
    """
    try:
        row = worksheet[row_number]
        max_lines = 1

        for cell in row:
            if cell.value:
                cell_text = str(cell.value)
                lines_count = cell_text.count("\n") + 1
                if lines_count > max_lines:
                    max_lines = lines_count

        return max(max_lines * base_height, base_height)

    except Exception as e:
        logger.warning(f"행 높이 계산 실패 (row {row_number}): {e}")
        return base_height


def apply_detail_sheet_formatting(
    worksheet: Worksheet,
    has_header: bool = True,
) -> None:
    """상세 시트용 포맷팅 적용

    - 자동 필터
    - 헤더 고정
    - 줌 85%
    - 테두리 및 정렬
    - 최적 컬럼 너비/행 높이

    Args:
        worksheet: 대상 워크시트
        has_header: 헤더 행 존재 여부
    """
    from openpyxl.utils import get_column_letter

    try:
        worksheet.auto_filter.ref = worksheet.dimensions
        if has_header:
            worksheet.freeze_panes = "A2"
        worksheet.sheet_view.zoomScale = 85

        thin_border = get_thin_border()
        center_alignment = get_center_alignment(wrap_text=True)

        for _row_idx, row in enumerate(worksheet.iter_rows(), 1):
            for _col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.alignment = center_alignment

        if has_header and worksheet.max_row > 0:
            header_style = get_basic_header_style()
            for cell in worksheet[1]:
                cell.font = header_style["font"]
                cell.fill = header_style["fill"]
                cell.alignment = header_style["alignment"]
                cell.border = header_style["border"]

        for col_num in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_num)
            optimal_width = calculate_optimal_column_width(worksheet, column_letter)
            worksheet.column_dimensions[column_letter].width = optimal_width

        for row_num in range(1, worksheet.max_row + 1):
            optimal_height = calculate_optimal_row_height(worksheet, row_num)
            worksheet.row_dimensions[row_num].height = optimal_height

        logger.info(f"상세 시트 포맷팅 적용: {worksheet.max_row}행, {worksheet.max_column}열")

    except Exception as e:
        logger.error(f"상세 시트 포맷팅 실패: {e}")


def apply_summary_formatting(worksheet: Worksheet) -> None:
    """요약/대시보드용 포맷팅 적용

    Args:
        worksheet: 대상 워크시트
    """
    try:
        worksheet.sheet_view.zoomScale = 90

        # 기본 컬럼 너비 설정
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 15
        worksheet.column_dimensions["D"].width = 12
        worksheet.column_dimensions["E"].width = 12
        worksheet.column_dimensions["F"].width = 15

        center_alignment = get_center_alignment(wrap_text=False)
        thin_border = get_thin_border()

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border

        for row_num in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 22

        logger.info(f"요약 시트 포맷팅 적용: {worksheet.max_row}행, {worksheet.max_column}열")

    except Exception as e:
        logger.error(f"요약 시트 포맷팅 실패: {e}")


def apply_worksheet_settings(
    worksheet: Worksheet,
    zoom_scale: int | None = 85,
    wrap_text: bool = True,
) -> None:
    """워크시트 기본 설정 적용

    Args:
        worksheet: 대상 워크시트
        zoom_scale: 줌 배율
        wrap_text: 텍스트 줄바꿈 여부
    """
    try:
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.zoomScale = zoom_scale

        for row in worksheet.iter_rows():
            for cell in row:
                alignment = cell.alignment.copy()
                alignment.wrap_text = wrap_text
                alignment.vertical = "center"
                cell.alignment = alignment

        for row_dim in worksheet.row_dimensions.values():
            row_dim.height = None

    except Exception as e:
        logger.error(f"워크시트 설정 적용 실패: {e}")


def save_to_csv(
    data: list[dict[str, Any]] | list[list[Any]],
    output_file: str,
    headers: list[str] | None = None,
    encoding: str = "utf-8-sig",
) -> None:
    """데이터를 CSV로 저장

    Args:
        data: 저장할 데이터 (딕셔너리 리스트 또는 리스트 리스트)
        output_file: 출력 파일 경로
        headers: 컬럼 헤더 (딕셔너리 리스트인 경우 자동 추출)
        encoding: 파일 인코딩 (기본값: utf-8-sig for Excel 호환)
    """
    try:
        with open(output_file, "w", newline="", encoding=encoding) as f:
            if not data:
                logger.warning(f"CSV 저장: 데이터가 비어있음 - {output_file}")
                return

            # 딕셔너리 리스트인 경우
            if isinstance(data[0], dict):
                dict_data: list[dict[str, Any]] = data  # type: ignore[assignment]
                fieldnames = headers or list(dict_data[0].keys())
                dict_writer = csv.DictWriter(f, fieldnames=fieldnames)
                dict_writer.writeheader()
                dict_writer.writerows(dict_data)
            # 리스트 리스트인 경우
            else:
                list_data: list[list[Any]] = data  # type: ignore[assignment]
                list_writer = csv.writer(f)
                if headers:
                    list_writer.writerow(headers)
                list_writer.writerows(list_data)

        logger.info(f"CSV 저장 완료: {output_file}")
    except Exception as e:
        logger.error(f"CSV 저장 실패: {e}")


def save_dict_list_to_excel(
    data: list[dict[str, Any]],
    output_file: str,
    sheet_name: str = "Sheet1",
    columns: list[str] | None = None,
) -> None:
    """딕셔너리 리스트를 Excel 파일로 저장 (pandas 대체용)

    pd.DataFrame(data).to_excel(output_file) 패턴을 대체합니다.

    Args:
        data: 딕셔너리 리스트 형태의 데이터
        output_file: 출력 Excel 파일 경로
        sheet_name: 시트 이름 (기본값: Sheet1)
        columns: 출력할 컬럼 목록 (순서 지정, None이면 데이터의 키 순서)

    Example:
        # 기존 pandas 코드:
        # df = pd.DataFrame(results)
        # df.to_excel("output.xlsx", index=False)

        # 새로운 코드:
        save_dict_list_to_excel(results, "output.xlsx")
    """
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.utils import get_column_letter

    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        wb.save(output_file)
        logger.warning(f"Excel 저장: 데이터가 비어있음 - {output_file}")
        return

    # 컬럼 순서 결정
    headers = columns or list(data[0].keys())

    # 헤더 작성
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_thin_border()

    # 데이터 작성
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = get_thin_border()

    # 컬럼 너비 자동 조정
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    wb.save(output_file)
    logger.info(f"Excel 저장 완료: {output_file}")


def add_sheet_from_dict_list(
    workbook: OpenpyxlWorkbook,
    data: list[dict[str, Any]],
    sheet_name: str,
    columns: list[str] | None = None,
) -> Worksheet:
    """기존 Workbook에 딕셔너리 리스트로 새 시트 추가 (pandas 대체용)

    pd.DataFrame(data).to_excel(writer, sheet_name=...) 패턴을 대체합니다.

    Args:
        workbook: openpyxl Workbook 객체
        data: 딕셔너리 리스트 형태의 데이터
        sheet_name: 시트 이름
        columns: 출력할 컬럼 목록 (순서 지정, None이면 데이터의 키 순서)

    Returns:
        생성된 Worksheet 객체

    Example:
        # 기존 pandas 코드:
        # with pd.ExcelWriter(path) as writer:
        #     df1.to_excel(writer, sheet_name="Sheet1", index=False)
        #     df2.to_excel(writer, sheet_name="Sheet2", index=False)

        # 새로운 코드:
        wb = openpyxl.Workbook()
        add_sheet_from_dict_list(wb, data1, "Sheet1")
        add_sheet_from_dict_list(wb, data2, "Sheet2")
        wb.save("output.xlsx")
    """
    from openpyxl.utils import get_column_letter

    ws = workbook.create_sheet(title=sheet_name)

    if not data:
        return ws

    # 컬럼 순서 결정
    headers = columns or list(data[0].keys())

    # 헤더 작성
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = get_header_font()
        cell.fill = get_header_fill()
        cell.alignment = get_center_alignment()
        cell.border = get_thin_border()

    # 데이터 작성
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = get_thin_border()

    # 컬럼 너비 자동 조정
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    return ws
