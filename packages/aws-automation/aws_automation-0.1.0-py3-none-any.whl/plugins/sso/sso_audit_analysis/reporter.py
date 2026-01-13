"""
SSO Excel Reporter - IAM Identity Center 감사 보고서 생성

시트 구성:
1. Summary: 전체 요약
2. Permission Sets: Permission Set 목록 및 위험도
3. Users: 사용자 목록 및 할당 현황
4. Groups: 그룹 목록 및 할당 현황
5. Admin Summary: 계정별 Admin 권한 현황
6. Issues: 전체 이슈 목록

Note:
    이 모듈은 Lazy Import 패턴을 사용합니다.
    openpyxl 등 무거운 의존성을 실제 사용 시점에만 로드합니다.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.styles import Alignment, Border, Font, PatternFill

from .analyzer import Severity, SSOAnalysisResult


class SSOExcelReporter:
    """SSO Excel 보고서 생성기"""

    # 스타일 캐시 (lazy initialization)
    _styles_initialized: bool = False
    _HEADER_FILL: PatternFill | None = None
    _HEADER_FONT: Font | None = None
    _THIN_BORDER: Border | None = None
    _CENTER_ALIGN: Alignment | None = None
    _LEFT_ALIGN: Alignment | None = None
    _SEVERITY_FILLS: dict[Severity, PatternFill] | None = None

    @classmethod
    def _init_styles(cls) -> None:
        """스타일 lazy 초기화"""
        if cls._styles_initialized:
            return

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        cls._HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cls._HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        cls._THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cls._CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cls._LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cls._SEVERITY_FILLS = {
            Severity.CRITICAL: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            Severity.HIGH: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            Severity.MEDIUM: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            Severity.LOW: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            Severity.INFO: PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        }
        cls._styles_initialized = True

    @property
    def HEADER_FILL(self) -> PatternFill:
        self._init_styles()
        return self._HEADER_FILL

    @property
    def HEADER_FONT(self) -> Font:
        self._init_styles()
        return self._HEADER_FONT

    @property
    def THIN_BORDER(self) -> Border:
        self._init_styles()
        return self._THIN_BORDER

    @property
    def CENTER_ALIGN(self) -> Alignment:
        self._init_styles()
        return self._CENTER_ALIGN

    @property
    def LEFT_ALIGN(self) -> Alignment:
        self._init_styles()
        return self._LEFT_ALIGN

    @property
    def SEVERITY_FILLS(self) -> dict[Severity, PatternFill]:
        self._init_styles()
        assert self._SEVERITY_FILLS is not None
        return self._SEVERITY_FILLS

    def __init__(
        self,
        results: list[SSOAnalysisResult],
        stats_list: list[dict[str, Any]],
    ):
        from openpyxl import Workbook as _Workbook

        self._init_styles()
        self.results = results
        self.stats_list = stats_list
        self.wb = _Workbook()

    def generate(self, output_dir: str) -> str:
        """Excel 보고서 생성"""
        # 기본 시트 제거
        self.wb.remove(self.wb.active)

        # 시트 생성
        self._create_summary_sheet()
        self._create_permission_sets_sheet()
        self._create_users_sheet()
        self._create_groups_sheet()
        self._create_admin_summary_sheet()
        self._create_issues_sheet()

        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"SSO_Audit_Report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        os.makedirs(output_dir, exist_ok=True)
        self.wb.save(filepath)

        return filepath

    def _apply_header_style(self, ws, row: int = 1) -> None:
        """헤더 스타일 적용"""
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN

    def _auto_column_width(self, ws, min_width: int = 10, max_width: int = 50) -> None:
        """열 너비 자동 조정"""
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self) -> None:
        """Summary 시트 생성"""
        ws = self.wb.create_sheet("Summary")

        # 제목
        ws["A1"] = "IAM Identity Center 보안 감사 보고서"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        ws["A2"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)

        # 전체 통계 (모든 결과 합산)
        totals = {
            "total_users": sum(s.get("total_users", 0) for s in self.stats_list),
            "total_groups": sum(s.get("total_groups", 0) for s in self.stats_list),
            "total_permission_sets": sum(s.get("total_permission_sets", 0) for s in self.stats_list),
            "users_with_admin": sum(s.get("users_with_admin", 0) for s in self.stats_list),
            "users_no_assignment": sum(s.get("users_no_assignment", 0) for s in self.stats_list),
            "admin_permission_sets": sum(s.get("admin_permission_sets", 0) for s in self.stats_list),
            "high_risk_permission_sets": sum(s.get("high_risk_permission_sets", 0) for s in self.stats_list),
            "empty_groups": sum(s.get("empty_groups", 0) for s in self.stats_list),
            "critical_issues": sum(s.get("critical_issues", 0) for s in self.stats_list),
            "high_issues": sum(s.get("high_issues", 0) for s in self.stats_list),
            "medium_issues": sum(s.get("medium_issues", 0) for s in self.stats_list),
            "low_issues": sum(s.get("low_issues", 0) for s in self.stats_list),
        }

        # 통계 테이블
        stats_data = [
            ("항목", "값", "설명"),
            ("전체 사용자", totals["total_users"], "Identity Center에 등록된 총 사용자 수"),
            ("전체 그룹", totals["total_groups"], "Identity Center에 등록된 총 그룹 수"),
            (
                "전체 Permission Sets",
                totals["total_permission_sets"],
                "정의된 Permission Set 수",
            ),
            ("", "", ""),
            (
                "Admin 권한 사용자",
                totals["users_with_admin"],
                "Admin Permission Set이 할당된 사용자",
            ),
            ("미할당 사용자", totals["users_no_assignment"], "어떤 계정에도 권한이 없는 사용자"),
            (
                "Admin Permission Sets",
                totals["admin_permission_sets"],
                "관리자 권한을 가진 Permission Set",
            ),
            (
                "위험 Permission Sets",
                totals["high_risk_permission_sets"],
                "위험 관리형 정책이 연결된 PS",
            ),
            ("빈 그룹", totals["empty_groups"], "멤버가 없는 그룹"),
            ("", "", ""),
            ("CRITICAL 이슈", totals["critical_issues"], "즉시 조치 필요"),
            ("HIGH 이슈", totals["high_issues"], "빠른 조치 권장"),
            ("MEDIUM 이슈", totals["medium_issues"], "검토 필요"),
            ("LOW 이슈", totals["low_issues"], "참고"),
        ]

        start_row = 4
        for i, (item, value, desc) in enumerate(stats_data):
            row = start_row + i
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=desc)

            # 첫 번째 행은 헤더
            if i == 0:
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.border = self.THIN_BORDER

            # 빈 행이 아닌 경우 테두리 적용
            elif item:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.THIN_BORDER

        self._auto_column_width(ws)

    def _create_permission_sets_sheet(self) -> None:
        """Permission Sets 시트 생성"""
        ws = self.wb.create_sheet("Permission Sets")

        headers = [
            "Permission Set Name",
            "Admin 권한",
            "할당 계정 수",
            "할당 계정",
            "AWS Managed Policies",
            "위험 정책",
            "Inline 위험 권한",
            "Risk Score",
            "Session Duration",
            "Description",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        for result in self.results:
            for ps_analysis in result.permission_set_analyses:
                ps = ps_analysis.permission_set

                row = [
                    ps.name,
                    "Yes" if ps.has_admin_access else "No",
                    len(ps.assigned_accounts),
                    ", ".join(ps.assigned_account_names[:5]) + ("..." if len(ps.assigned_account_names) > 5 else ""),
                    ", ".join(p.split("/")[-1] for p in ps.managed_policies),
                    ", ".join(p.split("/")[-1] for p in ps.high_risk_policies),
                    ", ".join(ps.dangerous_permissions[:5]) + ("..." if len(ps.dangerous_permissions) > 5 else ""),
                    ps_analysis.risk_score,
                    ps.session_duration,
                    ps.description[:100] if ps.description else "",
                ]
                ws.append(row)

                # Admin 권한 강조
                current_row = ws.max_row
                if ps.has_admin_access:
                    ws.cell(row=current_row, column=2).fill = PatternFill(
                        start_color="FFA500", end_color="FFA500", fill_type="solid"
                    )

                # Risk Score 색상
                risk_cell = ws.cell(row=current_row, column=8)
                if ps_analysis.risk_score >= 70:
                    risk_cell.fill = self.SEVERITY_FILLS[Severity.CRITICAL]
                elif ps_analysis.risk_score >= 40:
                    risk_cell.fill = self.SEVERITY_FILLS[Severity.HIGH]
                elif ps_analysis.risk_score >= 20:
                    risk_cell.fill = self.SEVERITY_FILLS[Severity.MEDIUM]

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"

    def _create_users_sheet(self) -> None:
        """Users 시트 생성"""
        ws = self.wb.create_sheet("Users")

        headers = [
            "Display Name",
            "Email",
            "Admin 권한",
            "Admin 계정",
            "할당 수",
            "할당 계정",
            "Risk Score",
            "User ID",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        for result in self.results:
            for user_analysis in result.user_analyses:
                user = user_analysis.user

                # 할당된 계정 목록
                assigned_accounts = list(set(a.get("account_name", "") for a in user.assignments))

                row = [
                    user.display_name or user.user_name,
                    user.email,
                    "Yes" if user.has_admin_access else "No",
                    ", ".join(user.admin_accounts[:3]) + ("..." if len(user.admin_accounts) > 3 else ""),
                    len(user.assignments),
                    ", ".join(assigned_accounts[:3]) + ("..." if len(assigned_accounts) > 3 else ""),
                    user_analysis.risk_score,
                    user.user_id,
                ]
                ws.append(row)

                # Admin 권한 강조
                current_row = ws.max_row
                if user.has_admin_access:
                    ws.cell(row=current_row, column=3).fill = PatternFill(
                        start_color="FFA500", end_color="FFA500", fill_type="solid"
                    )

                # 미할당 사용자 강조
                if not user.assignments:
                    ws.cell(row=current_row, column=5).fill = PatternFill(
                        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                    )

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"

    def _create_groups_sheet(self) -> None:
        """Groups 시트 생성"""
        ws = self.wb.create_sheet("Groups")

        headers = [
            "Group Name",
            "멤버 수",
            "Admin 권한",
            "할당 수",
            "할당 계정",
            "Description",
            "Group ID",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        for result in self.results:
            for group_analysis in result.group_analyses:
                group = group_analysis.group

                # 할당된 계정 목록
                assigned_accounts = list(set(a.get("account_name", "") for a in group.assignments))

                row = [
                    group.group_name,
                    group.member_count,
                    "Yes" if group.has_admin_access else "No",
                    len(group.assignments),
                    ", ".join(assigned_accounts[:3]) + ("..." if len(assigned_accounts) > 3 else ""),
                    group.description[:100] if group.description else "",
                    group.group_id,
                ]
                ws.append(row)

                # Admin 권한 강조
                current_row = ws.max_row
                if group.has_admin_access:
                    ws.cell(row=current_row, column=3).fill = PatternFill(
                        start_color="FFA500", end_color="FFA500", fill_type="solid"
                    )

                # 빈 그룹 강조
                if group.member_count == 0:
                    ws.cell(row=current_row, column=2).fill = PatternFill(
                        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                    )

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"

    def _create_admin_summary_sheet(self) -> None:
        """Admin Summary 시트 - 계정별 Admin 현황"""
        ws = self.wb.create_sheet("Admin Summary")

        headers = [
            "Account ID",
            "Account Name",
            "Admin Users",
            "Admin Groups",
            "Permission Sets",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        for result in self.results:
            for admin_summary in result.admin_account_summary:
                row = [
                    admin_summary.account_id,
                    admin_summary.account_name,
                    ", ".join(admin_summary.admin_users),
                    ", ".join(admin_summary.admin_groups),
                    ", ".join(admin_summary.permission_sets),
                ]
                ws.append(row)

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"

    def _create_issues_sheet(self) -> None:
        """Issues 시트 생성"""
        ws = self.wb.create_sheet("Issues")

        headers = [
            "Severity",
            "Issue Type",
            "Resource Type",
            "Resource Name",
            "Description",
            "Recommendation",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        # Severity 순서대로 정렬
        severity_order = {
            Severity.CRITICAL: 0,
            Severity.HIGH: 1,
            Severity.MEDIUM: 2,
            Severity.LOW: 3,
            Severity.INFO: 4,
        }

        all_issues = []
        for result in self.results:
            all_issues.extend(result.all_issues)

        sorted_issues = sorted(all_issues, key=lambda x: severity_order.get(x.severity, 5))

        for issue in sorted_issues:
            row = [
                issue.severity.value.upper(),
                issue.issue_type.value,
                issue.resource_type,
                issue.resource_name,
                issue.description,
                issue.recommendation,
            ]
            ws.append(row)

            # Severity 색상
            current_row = ws.max_row
            severity_fill = self.SEVERITY_FILLS.get(issue.severity)
            if severity_fill:
                ws.cell(row=current_row, column=1).fill = severity_fill

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"
