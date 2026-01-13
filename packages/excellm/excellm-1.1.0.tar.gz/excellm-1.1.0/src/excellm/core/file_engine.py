"""File-based Excel engine using openpyxl.

Cross-platform engine for working with Excel files without Excel running.
"""

from typing import Any, List, Dict, Optional, Tuple
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .engine import ExcelEngine
from .errors import ToolError, ErrorCodes


class FileEngine(ExcelEngine):
    """File-based Excel engine using openpyxl.
    
    Features:
    - Cross-platform (Windows, Mac, Linux)
    - No Excel required
    - Supports .xlsx, .xlsm, .xltx, .xltm
    
    Limitations:
    - No VBA execution
    - No screen capture
    - Limited formatting support
    - No live Excel integration
    """
    
    def __init__(self):
        """Initialize file engine."""
        if not OPENPYXL_AVAILABLE:
            raise ToolError(
                "openpyxl is not installed. Install it with: pip install openpyxl",
                code=ErrorCodes.EXCEL_NOT_RUNNING
            )
        
        self._workbook_cache: Dict[str, openpyxl.Workbook] = {}
    
    @property
    def engine_name(self) -> str:
        return "File"
    
    @property
    def supports_live_excel(self) -> bool:
        return False
    
    def _load_workbook(self, workbook_path: str) -> openpyxl.Workbook:
        """Load workbook from file, using cache if available."""
        abs_path = os.path.abspath(workbook_path)
        
        if abs_path in self._workbook_cache:
            return self._workbook_cache[abs_path]
        
        if not os.path.exists(abs_path):
            raise ToolError(
                f"File not found: {abs_path}",
                code=ErrorCodes.WORKBOOK_NOT_FOUND
            )
        
        try:
            wb = openpyxl.load_workbook(abs_path)
            self._workbook_cache[abs_path] = wb
            return wb
        except Exception as e:
            raise ToolError(
                f"Failed to open workbook: {str(e)}",
                code=ErrorCodes.WORKBOOK_NOT_FOUND
            )
    
    def _save_workbook(self, workbook_path: str):
        """Save workbook to file."""
        abs_path = os.path.abspath(workbook_path)
        
        if abs_path in self._workbook_cache:
            try:
                self._workbook_cache[abs_path].save(abs_path)
            except Exception as e:
                raise ToolError(
                    f"Failed to save workbook: {str(e)}",
                    code=ErrorCodes.WRITE_FAILED
                )
    
    def _parse_range(self, range_ref: str) -> Tuple[int, int, int, int]:
        """Parse Excel range reference to (start_row, start_col, end_row, end_col).
        
        Args:
            range_ref: Range like "A1:C5" or "B2"
            
        Returns:
            Tuple of (start_row, start_col, end_row, end_col) - 1-indexed
        """
        if ':' in range_ref:
            start, end = range_ref.split(':')
        else:
            start = end = range_ref
        
        # Parse start cell
        start_col = ''
        start_row = ''
        for char in start:
            if char.isalpha():
                start_col += char
            else:
                start_row += char
        
        start_col_idx = column_index_from_string(start_col)
        start_row_idx = int(start_row)
        
        # Parse end cell
        end_col = ''
        end_row = ''
        for char in end:
            if char.isalpha():
                end_col += char
            else:
                end_row += char
        
        end_col_idx = column_index_from_string(end_col)
        end_row_idx = int(end_row)
        
        return start_row_idx, start_col_idx, end_row_idx, end_col_idx
    
    def list_workbooks(self) -> List[Dict[str, Any]]:
        """List cached workbooks (file mode has no 'open' concept)."""
        workbooks = []
        
        for path, wb in self._workbook_cache.items():
            sheets = []
            for sheet in wb.worksheets:
                sheets.append({
                    "name": sheet.title,
                    "hidden": sheet.sheet_state != 'visible'
                })
            
            workbooks.append({
                "name": os.path.basename(path),
                "path": path,
                "sheets": sheets,
            })
        
        return workbooks
    
    def read_range(
        self,
        workbook_path: str,
        sheet_name: str,
        range_ref: str,
    ) -> List[List[Any]]:
        """Read range from file."""
        try:
            wb = self._load_workbook(workbook_path)
            
            if sheet_name not in wb.sheetnames:
                raise ToolError(
                    f"Sheet '{sheet_name}' not found in workbook.",
                    code=ErrorCodes.SHEET_NOT_FOUND
                )
            
            ws = wb[sheet_name]
            
            # Parse range
            start_row, start_col, end_row, end_col = self._parse_range(range_ref)
            
            # Read data
            data = []
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True
            ):
                data.append(list(row))
            
            return data if data else [[]]
            
        except ToolError:
            raise
        except Exception as e:
            raise ToolError(
                f"Failed to read range: {str(e)}",
                code=ErrorCodes.READ_FAILED
            )
    
    def write_range(
        self,
        workbook_path: str,
        sheet_name: str,
        range_ref: str,
        data: List[List[Any]],
    ) -> Dict[str, Any]:
        """Write range to file."""
        try:
            wb = self._load_workbook(workbook_path)
            
            if sheet_name not in wb.sheetnames:
                raise ToolError(
                    f"Sheet '{sheet_name}' not found in workbook.",
                    code=ErrorCodes.SHEET_NOT_FOUND
                )
            
            ws = wb[sheet_name]
            
            # Parse range
            start_row, start_col, _, _ = self._parse_range(range_ref)
            
            # Write data
            cells_written = 0
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    cell = ws.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                        value=value
                    )
                    cells_written += 1
            
            # Save workbook
            self._save_workbook(workbook_path)
            
            return {
                "success": True,
                "cells_written": cells_written,
                "range": range_ref,
            }
            
        except ToolError:
            raise
        except Exception as e:
            raise ToolError(
                f"Failed to write range: {str(e)}",
                code=ErrorCodes.WRITE_FAILED
            )
    
    def get_sheet_names(self, workbook_path: str) -> List[str]:
        """Get sheet names from file."""
        try:
            wb = self._load_workbook(workbook_path)
            return wb.sheetnames
            
        except ToolError:
            raise
        except Exception as e:
            raise ToolError(
                f"Failed to get sheet names: {str(e)}",
                code=ErrorCodes.WORKBOOK_NOT_FOUND
            )
    
    def create_sheet(
        self,
        workbook_path: str,
        sheet_name: str,
    ) -> Dict[str, Any]:
        """Create new sheet in file."""
        try:
            wb = self._load_workbook(workbook_path)
            
            if sheet_name in wb.sheetnames:
                raise ToolError(
                    f"Sheet '{sheet_name}' already exists.",
                    code=ErrorCodes.WRITE_FAILED
                )
            
            wb.create_sheet(title=sheet_name)
            self._save_workbook(workbook_path)
            
            return {
                "success": True,
                "sheet_name": sheet_name,
                "message": f"Created sheet '{sheet_name}'",
            }
            
        except ToolError:
            raise
        except Exception as e:
            raise ToolError(
                f"Failed to create sheet: {str(e)}",
                code=ErrorCodes.WRITE_FAILED
            )
    
    def delete_sheet(
        self,
        workbook_path: str,
        sheet_name: str,
    ) -> Dict[str, Any]:
        """Delete sheet from file."""
        try:
            wb = self._load_workbook(workbook_path)
            
            if sheet_name not in wb.sheetnames:
                raise ToolError(
                    f"Sheet '{sheet_name}' not found in workbook.",
                    code=ErrorCodes.SHEET_NOT_FOUND
                )
            
            if len(wb.sheetnames) == 1:
                raise ToolError(
                    "Cannot delete the only sheet in workbook.",
                    code=ErrorCodes.WRITE_FAILED
                )
            
            wb.remove(wb[sheet_name])
            self._save_workbook(workbook_path)
            
            return {
                "success": True,
                "sheet_name": sheet_name,
                "message": f"Deleted sheet '{sheet_name}'",
            }
            
        except ToolError:
            raise
        except Exception as e:
            raise ToolError(
                f"Failed to delete sheet: {str(e)}",
                code=ErrorCodes.WRITE_FAILED
            )
