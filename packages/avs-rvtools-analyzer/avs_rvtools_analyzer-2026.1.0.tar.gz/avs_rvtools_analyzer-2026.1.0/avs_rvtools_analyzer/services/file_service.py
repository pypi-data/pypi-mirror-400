"""
File handling service for RVTools analyzer.
"""

import io
import logging
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, Optional, Set
from zipfile import BadZipFile

import openpyxl
import pandas as pd
import xlrd
from fastapi import HTTPException, UploadFile

from ..config import FileConfig
from ..core.exceptions import (
    FileValidationError,
    ProtectedFileError,
    TemporaryFileError,
    UnsupportedFileFormatError,
)

logger = logging.getLogger(__name__)


class FileService:
    """Service for handling file operations including upload, validation, and cleanup."""

    def __init__(self, config: FileConfig):
        self.config = config
        self.temp_files = []  # Track temporary files for cleanup

    def validate_file(self, file: UploadFile) -> None:
        """
        Validate uploaded file.

        Args:
            file: The uploaded file to validate

        Raises:
            FileValidationError: If file validation fails
        """
        if not file.filename:
            raise FileValidationError("No file selected")

        if not self._is_allowed_file(file.filename):
            allowed_exts = list(self.config.allowed_extensions)
            raise UnsupportedFileFormatError(
                f"Invalid file format. Please upload a file with one of these extensions: {', '.join(f'.{ext}' for ext in allowed_exts)}",
                file_extension=(
                    file.filename.rsplit(".", 1)[-1].lower()
                    if "." in file.filename
                    else None
                ),
            )

    def _is_allowed_file(self, filename: str) -> bool:
        """
        Check if file extension is allowed.

        Args:
            filename: Name of the file to check

        Returns:
            True if file extension is allowed, False otherwise
        """
        if not filename:
            return False

        extension = filename.rsplit(".", 1)[-1].lower()
        return extension in self.config.allowed_extensions

    def _filter_powered_off_rows(self, data: list, headers: list) -> list:
        """
        Filter out rows where Powerstate column equals 'poweredOff'.

        Args:
            data: List of row dictionaries
            headers: List of column headers

        Returns:
            Filtered list of row dictionaries
        """
        # Look for Powerstate column (case-sensitive)
        powerstate_column = None
        for header in headers:
            if header == "Powerstate":
                powerstate_column = header
                break

        # If no Powerstate column found, return data unchanged
        if not powerstate_column:
            logger.debug("No 'Powerstate' column found, skipping power state filtering")
            return data

        # Filter out rows where Powerstate equals 'poweredOff' (case-sensitive)
        filtered_data = []
        filtered_count = 0
        for row in data:
            powerstate_value = row.get(powerstate_column)
            if powerstate_value != "poweredOff":
                filtered_data.append(row)
            else:
                filtered_count += 1

        logger.debug(f"Filtered out {filtered_count} powered off VMs from data")
        return filtered_data

    async def save_uploaded_file(self, file: UploadFile) -> Path:
        """
        Save uploaded file to temporary location.

        Args:
            file: The uploaded file to save

        Returns:
            Path to the saved temporary file

        Raises:
            TemporaryFileError: If file saving fails
        """
        try:
            # Create temporary file
            temp_file = tempfile.NamedTemporaryFile(
                delete=False, suffix=self.config.temp_file_suffix
            )

            # Read and write file content
            content = await file.read()
            temp_file.write(content)
            temp_file.close()

            # Track temp file for cleanup
            temp_file_path = Path(temp_file.name)
            self.temp_files.append(temp_file_path)

            logger.debug(f"Saved uploaded file to: {temp_file_path}")
            return temp_file_path

        except Exception as e:
            logger.error(f"Error saving uploaded file: {str(e)}")
            raise TemporaryFileError(f"Error saving file: {str(e)}", operation="save")

    async def load_excel_file_from_memory(self, file: UploadFile, filter_powered_off: bool = False) -> Dict[str, Any]:
        """
        Load Excel file directly from memory without saving to disk.

        This method provides true in-memory processing for enhanced security,
        ensuring that sensitive data never touches the file system.

        Args:
            file: The uploaded file to process
            filter_powered_off: If True, filter out rows where Powerstate column equals 'poweredOff'

        Returns:
            Dictionary containing sheet data

        Raises:
            FileValidationError: If file cannot be loaded or parsed
            ProtectedFileError: If file is password protected
        """
        try:
            logger.debug(f"Loading Excel file from memory: {file.filename}")

            # Read file content into memory
            content = await file.read()
            file_stream = io.BytesIO(content)

            # Try to load with openpyxl first (for .xlsx files)
            try:
                workbook = openpyxl.load_workbook(
                    file_stream, read_only=True, data_only=True
                )
                sheets = {}

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    data = []
                    headers = []

                    # Get headers from first row
                    first_row = next(
                        sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
                    )
                    if first_row:
                        headers = [
                            str(cell) if cell is not None else f"Column_{i}"
                            for i, cell in enumerate(first_row)
                        ]

                    # Get data from remaining rows
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            row_dict = {}
                            for i, cell in enumerate(row):
                                if i < len(headers):
                                    row_dict[headers[i]] = cell
                            data.append(row_dict)

                    # Apply power state filtering if requested
                    if filter_powered_off:
                        data = self._filter_powered_off_rows(data, headers)

                    sheets[sheet_name] = {
                        "headers": headers,
                        "data": data,
                        "row_count": len(data),
                    }

                workbook.close()
                logger.debug(f"Loaded Excel file from memory with {len(sheets)} sheets")
                return sheets

            except BadZipFile:
                # Reset stream position for xlrd attempt
                file_stream.seek(0)

                # Try with xlrd for .xls files
                try:
                    workbook = xlrd.open_workbook(file_contents=file_stream.read())
                    sheets = {}

                    for sheet_index in range(workbook.nsheets):
                        sheet = workbook.sheet_by_index(sheet_index)
                        sheet_name = sheet.name

                        data = []
                        headers = []

                        if sheet.nrows > 0:
                            # Get headers from first row
                            headers = [
                                str(sheet.cell_value(0, col))
                                for col in range(sheet.ncols)
                            ]

                            # Get data from remaining rows
                            for row in range(1, sheet.nrows):
                                row_dict = {}
                                for col in range(sheet.ncols):
                                    if col < len(headers):
                                        row_dict[headers[col]] = sheet.cell_value(
                                            row, col
                                        )
                                data.append(row_dict)

                        # Apply power state filtering if requested
                        if filter_powered_off:
                            data = self._filter_powered_off_rows(data, headers)

                        sheets[sheet_name] = {
                            "headers": headers,
                            "data": data,
                            "row_count": len(data),
                        }

                    logger.debug(
                        f"Loaded Excel file (xls) from memory with {len(sheets)} sheets"
                    )
                    return sheets

                except xlrd.XLRDError as e:
                    if "password" in str(e).lower() or "encrypted" in str(e).lower():
                        raise ProtectedFileError(
                            "File appears to be password protected"
                        )
                    raise FileValidationError(f"Error reading Excel file: {str(e)}")

        except (ProtectedFileError, FileValidationError):
            # Re-raise our custom exceptions
            raise
        except Exception as e:
            error_msg = str(e).lower()
            if (
                "password" in error_msg
                or "encrypted" in error_msg
                or "protected" in error_msg
            ):
                raise ProtectedFileError(
                    "File appears to be password protected or encrypted"
                )

            logger.error(f"Error loading Excel file from memory: {str(e)}")
            raise FileValidationError(
                f"Error loading Excel file from memory: {str(e)}", file_type="Excel"
            )
        finally:
            # Ensure file stream is closed and content is cleared from memory
            if "file_stream" in locals():
                file_stream.close()
            # Clear the content variable to free memory
            if "content" in locals():
                del content

    def load_excel_file(self, file_path: Path, filter_powered_off: bool = False) -> Dict[str, Any]:
        """
        Load Excel file and return sheets data.

        Args:
            file_path: Path to the Excel file
            filter_powered_off: If True, filter out rows where Powerstate column equals 'poweredOff'

        Returns:
            Dictionary containing sheet data

        Raises:
            FileValidationError: If file cannot be loaded or parsed
            ProtectedFileError: If file is password protected
        """
        try:
            # Check if file exists
            if not file_path.exists():
                raise FileValidationError(f"File not found: {file_path}")

            logger.debug(f"Loading Excel file: {file_path}")

            # Try to load with openpyxl first (for .xlsx files)
            try:
                workbook = openpyxl.load_workbook(
                    file_path, read_only=True, data_only=True
                )
                sheets = {}

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    data = []
                    headers = []

                    # Get headers from first row
                    first_row = next(
                        sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
                    )
                    if first_row:
                        headers = [
                            str(cell) if cell is not None else f"Column_{i}"
                            for i, cell in enumerate(first_row)
                        ]

                    # Get data from remaining rows
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            row_dict = {}
                            for i, cell in enumerate(row):
                                if i < len(headers):
                                    row_dict[headers[i]] = cell
                            data.append(row_dict)

                    # Apply power state filtering if requested
                    if filter_powered_off:
                        data = self._filter_powered_off_rows(data, headers)

                    sheets[sheet_name] = {
                        "headers": headers,
                        "data": data,
                        "row_count": len(data),
                    }

                workbook.close()
                logger.debug(f"Loaded Excel file with {len(sheets)} sheets")
                return sheets

            except BadZipFile:
                # Try with xlrd for .xls files
                try:
                    workbook = xlrd.open_workbook(file_path)
                    sheets = {}

                    for sheet_index in range(workbook.nsheets):
                        sheet = workbook.sheet_by_index(sheet_index)
                        sheet_name = sheet.name

                        data = []
                        headers = []

                        if sheet.nrows > 0:
                            # Get headers from first row
                            headers = [
                                str(sheet.cell_value(0, col))
                                for col in range(sheet.ncols)
                            ]

                            # Get data from remaining rows
                            for row in range(1, sheet.nrows):
                                row_dict = {}
                                for col in range(sheet.ncols):
                                    if col < len(headers):
                                        row_dict[headers[col]] = sheet.cell_value(
                                            row, col
                                        )
                                data.append(row_dict)

                        # Apply power state filtering if requested
                        if filter_powered_off:
                            data = self._filter_powered_off_rows(data, headers)

                        sheets[sheet_name] = {
                            "headers": headers,
                            "data": data,
                            "row_count": len(data),
                        }

                    logger.debug(f"Loaded Excel file (xls) with {len(sheets)} sheets")
                    return sheets

                except xlrd.XLRDError as e:
                    if "password" in str(e).lower() or "encrypted" in str(e).lower():
                        raise ProtectedFileError(
                            "File appears to be password protected"
                        )
                    raise FileValidationError(f"Error reading Excel file: {str(e)}")

        except (ProtectedFileError, FileValidationError):
            # Re-raise our custom exceptions
            raise
        except Exception as e:
            error_msg = str(e).lower()
            if (
                "password" in error_msg
                or "encrypted" in error_msg
                or "protected" in error_msg
            ):
                raise ProtectedFileError(
                    "File appears to be password protected or encrypted"
                )

            logger.error(f"Error loading Excel file {file_path}: {str(e)}")
            raise FileValidationError(
                f"Error loading Excel file: {str(e)}", file_type="Excel"
            )

    def cleanup_temp_file(self, file_path: Path) -> None:
        """
        Clean up a temporary file.

        Args:
            file_path: Path to the temporary file to clean up
        """
        try:
            if file_path.exists():
                os.unlink(file_path)
                logger.debug(f"Cleaned up temp file: {file_path}")

            # Remove from tracking list
            if file_path in self.temp_files:
                self.temp_files.remove(file_path)

        except Exception as e:
            logger.warning(f"Error cleaning up temp file {file_path}: {str(e)}")

    def cleanup(self) -> None:
        """
        Clean up temporary files.

        Raises:
            TemporaryFileError: If cleanup fails for critical operations
        """
        cleanup_errors = []

        for temp_file in self.temp_files:
            try:
                if temp_file.exists():
                    temp_file.unlink()
                    logger.debug(f"Cleaned up temporary file: {temp_file}")
            except Exception as e:
                error_msg = f"Error cleaning up {temp_file}: {str(e)}"
                logger.warning(error_msg)
                cleanup_errors.append(error_msg)

        # Clear the list
        self.temp_files.clear()

        # If we had errors, log them but don't raise exception unless critical
        if cleanup_errors:
            logger.warning(f"Cleanup completed with {len(cleanup_errors)} errors")
            # Only raise if all files failed to cleanup (indicating a serious issue)
            if len(cleanup_errors) == len(self.temp_files):
                raise TemporaryFileError(
                    f"Failed to cleanup temporary files: {'; '.join(cleanup_errors)}",
                    operation="cleanup",
                )

    def get_excel_sheets_data(self, excel_data: Dict[str, Any]) -> dict:
        """
        Extract all sheets data from Excel file data.

        Args:
            excel_data: Dictionary containing sheet data from load_excel_file

        Returns:
            Dictionary containing data from all sheets
        """
        sheets = {}
        try:
            for sheet_name, sheet_info in excel_data.items():
                # Extract just the data from each sheet
                sheets[sheet_name] = sheet_info.get("data", [])
            return sheets

        except Exception as e:
            logger.error(f"Error extracting sheets data: {str(e)}")
            raise HTTPException(
                status_code=500, detail=f"Error processing Excel sheets: {str(e)}"
            )

    def __del__(self):
        """Cleanup temp files when service is destroyed."""
        self.cleanup()
