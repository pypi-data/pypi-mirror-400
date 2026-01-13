"""Formula parsing and analysis utilities for spreadsheets."""

from typing import Dict, List, Set, Tuple, Optional, Any, Iterator
import re
import pandas as pd


class FormulaTokenizer:
    """Helper class to tokenize Excel formulas safely."""
    
    def tokenize(self, formula: str) -> List[str]:
        """Split formula into tokens while preserving structure."""
        # Simple regex to capture operators, references, strings, and numbers
        # This is more robust than simple splitting
        token_pattern = re.compile(r'([A-Za-z_][A-Za-z0-9_\.]*\(|\)|,|[+\-*/^&=<>]+|"[^"]*"|[0-9\.]+|[A-Za-z]+[0-9]+:[A-Za-z]+[0-9]+|[A-Za-z]+[0-9]+)')
        tokens = [t.strip() for t in token_pattern.findall(formula) if t.strip()]
        return tokens

    def extract_args(self, token_stream: Iterator[str]) -> List[str]:
        """Extract arguments from a function call, handling nested parenthesis."""
        args = []
        current_arg = []
        depth = 0
        
        for token in token_stream:
            if token == '(':
                depth += 1
                current_arg.append(token)
            elif token == ')':
                if depth == 0:
                    break
                depth -= 1
                current_arg.append(token)
            elif token == ',' and depth == 0:
                args.append(" ".join(current_arg))
                current_arg = []
            else:
                current_arg.append(token)
                
        if current_arg:
            args.append(" ".join(current_arg))
            
        return args


class FormulaParser:
    """
    Extracts, analyzes and simplifies Excel formulas from spreadsheets.
    Optimized for memory usage with streaming reads.
    """

    # Common Excel formula patterns
    CELL_REF_PATTERN = re.compile(r'([A-Z]+[0-9]+|[A-Z]+\:[A-Z]+|[0-9]+\:[0-9]+|[A-Z]+[0-9]+\:[A-Z]+[0-9]+)')
    
    def __init__(self):
        """Initialize the formula parser."""
        self.formula_map = {}  # Maps cell address to formula
        self.dependency_graph = {}  # Maps cell to its dependencies
        self.reverse_dependency = {}  # Maps cell to cells that depend on it
        self.tokenizer = FormulaTokenizer()
    
    def extract_formulas(self, excel_path: str, sheet_name: Optional[str] = None) -> Dict[str, str]:
        """
        Extract all formulas from an Excel file using Memory-Efficient Streaming.
        
        Args:
            excel_path: Path to the Excel file
            sheet_name: Optional specific sheet to parse (saves time)
            
        Returns:
            Dictionary mapping cell addresses to formulas
        """
        try:
            import openpyxl
            # read_only=True enables streaming (huge memory savings)
            workbook = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
            
            formulas = {}
            
            sheets_to_process = [sheet_name] if sheet_name else workbook.sheetnames
            
            for sheet in sheets_to_process:
                if sheet not in workbook.sheetnames:
                    continue
                    
                ws = workbook[sheet]
                # iter_rows in read_only mode returns raw cells efficiently
                for row_idx, row in enumerate(ws.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        # In read_only mode, cell.value is the formula string if present
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            # Convert 1-based indices to A1 notation manually or using util
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            coordinate = f"{col_letter}{row_idx}"
                            
                            cell_address = f"{sheet}!{coordinate}"
                            formulas[cell_address] = cell.value
            
            # Important: Close the workbook in read_only mode
            workbook.close()
            
            self.formula_map = formulas
            return formulas
            
        except ImportError:
            raise ImportError("openpyxl is required. Install with 'pip install openpyxl'")
    
    def build_dependency_graph(self) -> Dict[str, Set[str]]:
        """
        Build a graph of cell dependencies based on extracted formulas.
        """
        dependency_graph = {}
        reverse_dependency = {}
        
        for cell, formula in self.formula_map.items():
            references = self.extract_cell_references(formula)
            dependency_graph[cell] = set(references)
            
            for ref in references:
                if ref not in reverse_dependency:
                    reverse_dependency[ref] = set()
                reverse_dependency[ref].add(cell)
        
        self.dependency_graph = dependency_graph
        self.reverse_dependency = reverse_dependency
        return dependency_graph
    
    def extract_cell_references(self, formula: str) -> List[str]:
        """Extract all cell references from a formula."""
        if formula.startswith('='):
            formula = formula[1:]
        return self.CELL_REF_PATTERN.findall(formula)
    
    def simplify_formula(self, formula: str) -> str:
        """
        Generate a simplified explanation using robust tokenization.
        Handles nested functions better than Regex.
        """
        if not formula.startswith('='):
            return f"Static value: {formula}"
            
        formula_body = formula[1:]
        tokens = self.tokenizer.tokenize(formula_body)
        
        if not tokens:
            return "Empty Formula"

        main_func = tokens[0].upper()
        
        # Robust parsing for common functions
        if main_func == 'SUM(':
            args = self.tokenizer.extract_args(iter(tokens[1:]))
            return f"Sum of {', '.join(args)}"
            
        elif main_func == 'AVERAGE(':
            args = self.tokenizer.extract_args(iter(tokens[1:]))
            return f"Average of {', '.join(args)}"
            
        elif main_func == 'VLOOKUP(':
            args = self.tokenizer.extract_args(iter(tokens[1:]))
            if len(args) >= 3:
                return f"Lookup '{args[0]}' in {args[1]} (Column {args[2]})"
                
        elif main_func == 'IF(':
            args = self.tokenizer.extract_args(iter(tokens[1:]))
            if len(args) >= 3:
                return f"If {args[0]} is true, then {args[1]}, else {args[2]}"
        
        # Fallback for complex/unknown formulas
        return f"Formula: {formula}"

    def get_formula_impact(self, cell_address: str) -> Dict[str, Any]:
        """Analyze the impact of a formula cell."""
        if not self.dependency_graph:
            self.build_dependency_graph()
            
        dependencies = self.dependency_graph.get(cell_address, set())
        dependents = self.reverse_dependency.get(cell_address, set())
        formula = self.formula_map.get(cell_address, "")
        
        return {
            "cell": cell_address,
            "formula": formula,
            "simplified_explanation": self.simplify_formula(formula),
            "dependencies": list(dependencies),
            "dependents": list(dependents),
            "is_leaf": len(dependencies) == 0,
            "is_root": len(dependents) == 0,
            "dependency_depth": self._calculate_dependency_depth(cell_address)
        }
    
    def _calculate_dependency_depth(self, cell_address: str) -> int:
        if cell_address not in self.dependency_graph or not self.dependency_graph[cell_address]:
            return 0
        return 1 + max((self._calculate_dependency_depth(dep) for dep in self.dependency_graph[cell_address]), default=0)

    def encode_formulas_for_llm(self, formulas: Optional[Dict[str, str]] = None) -> str:
        """Generate LLM-friendly encoding."""
        if formulas is None:
            formulas = self.formula_map
            
        if not formulas:
            return "No formulas found"
            
        lines = ["## Spreadsheet Formulas"]
        
        # Group similar formulas to save tokens
        formula_groups: Dict[str, List[str]] = {}
        for cell, formula in formulas.items():
            if formula not in formula_groups:
                formula_groups[formula] = []
            formula_groups[formula].append(cell)
        
        for formula, cells in formula_groups.items():
            simplified = self.simplify_formula(formula)
            cell_list = ", ".join(cells[:5])
            if len(cells) > 5:
                cell_list += f" (+{len(cells)-5} more)"
                
            lines.append(f"- {simplified}")
            lines.append(f"  - Cells: {cell_list}")
            # Optional: Don't include raw formula if simplified is good enough
            lines.append(f"  - Raw: `{formula}`")
            lines.append("")
            
        return "\n".join(lines)


class FormulaDependencyAnalyzer:
    """Specialized analyzer for formula dependencies."""
    
    def __init__(self, formula_parser: Optional[FormulaParser] = None):
        self.parser = formula_parser or FormulaParser()
        
    # (Existing methods remain unchanged)
    def find_calculation_chains(self) -> List[List[str]]:
        # ... existing implementation ...
        if not self.parser.dependency_graph:
            raise ValueError("Dependency graph not built.")
        
        root_cells = {c for c in self.parser.reverse_dependency if not self.parser.reverse_dependency[c]}
        return [self._build_chain_from_cell(root) for root in root_cells]
    
    def _build_chain_from_cell(self, start_cell: str) -> List[str]:
        chain = [start_cell]
        current = start_cell
        while current in self.parser.dependency_graph and self.parser.dependency_graph[current]:
            next_cell = next(iter(self.parser.dependency_graph[current]))
            if next_cell in chain: break
            chain.append(next_cell)
            current = next_cell
        return chain
    
    def identify_critical_cells(self) -> List[str]:
        if not self.parser.reverse_dependency:
            raise ValueError("Dependency analysis not performed")
        importance = {cell: len(deps) for cell, deps in self.parser.reverse_dependency.items()}
        return sorted([c for c, count in importance.items() if count > 0], key=lambda x: importance[x], reverse=True)