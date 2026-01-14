import sys
import os
import subprocess
import datetime
import configparser
from pathlib import Path
from functools import reduce

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

# 给柱状图单个柱子上色用
try:
    from openpyxl.chart.series import DataPoint
except Exception:
    DataPoint = None  # 极少数版本差异，正常不会发生


# ======================
# 1. 配置管理
# ======================
class ConfigManager:
    def __init__(self, config_path="config.ini"):
        self.config_path = Path(config_path)
        self.config = configparser.ConfigParser()

    def load(self):
        if not self.config_path.exists():
            self._create_sample()
            print(f"⚠️ 未找到配置文件，已生成示例：{self.config_path.resolve()}")
            print("👉 请修改 config.ini 后重新运行程序")
            sys.exit(0)

        self.config.read(self.config_path, encoding="utf-8")

        try:
            self.names_xlsx = self.config["paths"]["names_xlsx"]
            self.homework_dir = self.config["paths"]["homework_dir"]
        except KeyError as e:
            raise KeyError(f"config.ini 中缺少必要配置项：{e}")

        return self

    def _create_sample(self):
        self.config["paths"] = {
            "names_xlsx": "2024数学类（周二上午）.xlsx",
            "homework_dir": "平时成绩"
        }
        with open(self.config_path, "w", encoding="utf-8") as f:
            self.config.write(f)

        # ✅ 跨平台打开（发布到 PyPI 后，macOS/Linux 不会因为 os.startfile 崩掉）
        self._open_file(self.config_path)

    @staticmethod
    def _open_file(path: Path) -> None:
        """
        跨平台打开文件：
        - Windows: os.startfile
        - macOS: open
        - Linux: xdg-open
        打不开也不影响主流程（静默失败）。
        """
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.run(["open", str(path)], check=False)
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except Exception:
            pass


# ======================
# 2. 作业扫描
# ======================
class HomeworkScanner:
    def __init__(self, homework_dir):
        self.root = Path(homework_dir)
        if not self.root.exists():
            raise FileNotFoundError(f"作业目录不存在：{self.root.resolve()}")

    def scan(self):
        homeworks = [p for p in self.root.iterdir() if p.is_dir()]
        dicts = map(self._dir_to_string, homeworks)
        return {k.name: v for d in dicts for k, v in d.items()}

    def _dir_to_string(self, root, level=0):
        lines = []
        indent = "    " * level
        lines.append(f"{indent}{root.name}/")

        for item in sorted(root.iterdir(), key=lambda x: (x.is_file(), x.name)):
            if item.is_dir():
                sub = self._dir_to_string(item, level + 1)
                lines.append(sub[item])
            else:
                lines.append(f"{'    ' * (level + 1)}{item.name}")

        return {root: "\n".join(lines)}


# ======================
# 3. 成绩计算
# ======================
class ScoreCalculator:
    def __init__(self, names_xlsx, homeworks_dict):
        self.names_xlsx = Path(names_xlsx)
        self.homeworks_dict = homeworks_dict

        if not self.names_xlsx.exists():
            raise FileNotFoundError(f"学生名单不存在：{self.names_xlsx.resolve()}")

    def load_students(self):
        df = pd.read_excel(
            self.names_xlsx,
            header=[1],
            skiprows=[0]  # 上师大特有格式
        )

        self.data = df[['学号', '姓名', '行政班']].copy()
        self.data['学号'] = self.data['学号'].astype(str).str.strip()
        self.data['姓名'] = self.data['姓名'].str.strip()
        return self

    def calculate(self):
        # 是否提交每次作业
        for k, v in self.homeworks_dict.items():
            self.data[k] = self.data['学号'].apply(lambda x: x in v)

        # 提交次数
        self.data['交作业次数'] = reduce(
            lambda x, y: x + self.data[y],
            self.homeworks_dict.keys(),
            0
        )

        # 提交率（0~1）
        self.data['提交率(%)'] = (
            self.data['交作业次数'] / len(self.homeworks_dict)
        ).round(4)

        # 平时成绩（封顶 100）
        self.data['平时成绩'] = (
            self.data['提交率(%)']
            .clip(upper=1)
            .mul(100)
            .round(2)
        )

        return self

    def export(self):
        output_file = self.names_xlsx.with_name(
            f"{self.names_xlsx.stem}-平时成绩-{datetime.date.today()}.xlsx"
        )

        # ========= 数据拆分 =========
        full_submit = self.data[self.data['平时成绩'] == 100].copy()
        not_full_submit = (
            self.data[self.data['平时成绩'] < 100]
            .sort_values('平时成绩')
            .copy()
        )

        # ========= 汇总表 =========
        summary = pd.DataFrame({
            '统计项': [
                '学生总人数',
                '100%提交人数',
                '未满100%提交人数',
                '平均提交率(%)'
            ],
            '数值': [
                len(self.data),
                len(full_submit),
                len(not_full_submit),
                round(self.data['平时成绩'].mean(), 2)
            ]
        })

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            summary.to_excel(writer, sheet_name='汇总', index=False)
            self.data.to_excel(writer, sheet_name='全部学生', index=False)
            full_submit.to_excel(writer, sheet_name='100%提交', index=False)
            not_full_submit.to_excel(writer, sheet_name='未满100%提交', index=False)

            wb = writer.book
            ws = wb['汇总']

            # ========= 柱状图（只画前三项：总人数/100%提交/未满100%） =========
            bar = BarChart()
            bar.title = "作业提交人数统计"
            bar.y_axis.title = "人数"
            bar.x_axis.title = "统计项"

            # 数据：B2:B4；分类：A2:A4
            bar_data = Reference(ws, min_col=2, min_row=2, max_row=4)
            bar_cats = Reference(ws, min_col=1, min_row=2, max_row=4)

            bar.add_data(bar_data, titles_from_data=False)
            bar.set_categories(bar_cats)
            bar.width = 18
            bar.height = 10
            bar.style = 10

            # ========= 给三根柱子分别上色 =========
            # 0: 学生总人数 -> 蓝色
            # 1: 100%提交人数 -> 绿色
            # 2: 未满100%提交人数 -> 红色
            if DataPoint is not None and bar.series:
                colors = ["4472C4", "70AD47", "C00000"]  # 蓝/绿/红
                s = bar.series[0]
                s.dPt = []
                for i, c in enumerate(colors):
                    dp = DataPoint(idx=i)
                    dp.graphicalProperties.solidFill = c
                    dp.graphicalProperties.line.solidFill = c
                    s.dPt.append(dp)

            ws.add_chart(bar, "D2")

            # ========= 未满100% 高亮 =========
            ws_risk = wb['未满100%提交']
            red_fill = PatternFill(
                start_color='FFF4CCCC',
                end_color='FFF4CCCC',
                fill_type='solid'
            )
            headers = [c.value for c in ws_risk[1]]
            score_idx = headers.index('平时成绩')
            for row in ws_risk.iter_rows(min_row=2):
                if row[score_idx].value < 100:
                    row[score_idx].fill = red_fill

        print(f"✅ 平时成绩统计（含汇总柱状图：总人数蓝/提交绿/未满红）已生成：\n{output_file.resolve()}")


# ======================
# 4. 主程序
# ======================
def main():
    config = ConfigManager().load()
    homeworks_dict = HomeworkScanner(config.homework_dir).scan()

    (
        ScoreCalculator(config.names_xlsx, homeworks_dict)
        .load_students()
        .calculate()
        .export()
    )


if __name__ == "__main__":
    main()

