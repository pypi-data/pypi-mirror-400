#!/usr/bin/env python3
"""
Excel formatting utilities for Dataform Checklist
"""

import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .models import StatusColors


def apply_cell_formatting(worksheet: Worksheet, inventory: pd.DataFrame, start_row: int = 9):
    """Apply Excel formatting and conditional styling"""
    colors = StatusColors()
    color_fills = {
        'New': PatternFill(start_color=colors.NEW, end_color=colors.NEW, fill_type='solid'),
        'Existed': PatternFill(start_color=colors.EXISTED, end_color=colors.EXISTED, fill_type='solid'),
        'Deleted': PatternFill(start_color=colors.DELETED, end_color=colors.DELETED, fill_type='solid')
    }
    
    # Apply conditional formatting to status column (column E)
    status_col = 5
    remark_col = 4
    
    for idx, row in enumerate(inventory.itertuples(), start=start_row):
        status = row.status
        if status in color_fills:
            worksheet.cell(row=idx, column=status_col).fill = color_fills[status]
        
        # Apply top alignment to all cells in the row
        for col_idx in range(1, 6):  # Columns A-E
            cell = worksheet.cell(row=idx, column=col_idx)
            # Enable text wrapping for remark column
            if col_idx == remark_col:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top')
    
    # Auto-adjust column widths for data columns
    _auto_adjust_column_widths(worksheet, inventory, start_row)


def _auto_adjust_column_widths(worksheet: Worksheet, inventory: pd.DataFrame, start_row: int):
    """Auto-adjust column widths based on content"""
    for col_idx in range(1, 6):  # Columns A-E
        max_length = 0
        column_letter = worksheet.cell(row=start_row, column=col_idx).column_letter
        
        for row_idx in range(start_row, start_row + len(inventory) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    # For cells with line breaks, take the longest line
                    lines = str(cell.value).split('\n')
                    max_line_length = max(len(line) for line in lines)
                    if max_line_length > max_length:
                        max_length = max_line_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        if adjusted_width > 10:  # Only adjust if meaningful
            worksheet.column_dimensions[column_letter].width = adjusted_width
