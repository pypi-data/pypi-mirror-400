import csv
import io
import json
from datetime import datetime
from typing import Any, Protocol, runtime_checkable

from llm_meter.cli.models import UsageExport


@runtime_checkable
class DataExporter(Protocol):
    """Protocol for data exporters to support Open/Closed Principle."""

    def export(self, data: list[UsageExport]) -> bytes:
        """Export data to bytes."""
        ...


class JSONExporter(DataExporter):
    def export(self, data: list[UsageExport]) -> bytes:
        content = json.dumps([row.model_dump(mode="json") for row in data], indent=2)
        return content.encode("utf-8")


class CSVExporter(DataExporter):
    def export(self, data: list[UsageExport]) -> bytes:
        if not data:
            return b""

        output_buffer = io.StringIO()
        keys = list(UsageExport.model_fields.keys())
        writer = csv.DictWriter(output_buffer, fieldnames=keys)
        writer.writeheader()
        for row in data:
            writer.writerow(row.model_dump(mode="json"))
        return output_buffer.getvalue().encode("utf-8")


class ExcelExporter(DataExporter):
    """
    Excel exporter using openpyxl.
    """

    def export(self, data: list[UsageExport]) -> bytes:
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("Excel export requires 'openpyxl'. Install it with 'pip install openpyxl'.") from None

        if not data:
            return b""

        wb = Workbook()
        ws = wb.active
        # Workbook() always creates at least one sheet, and .active is always set initially.
        # We ensure type safety for static analyzers.
        if ws is None:  # pragma: no cover
            ws = wb.create_sheet()

        # Write header
        keys = list(UsageExport.model_fields.keys())
        ws.append(keys)

        # Write data
        for row in data:
            dict_row = row.model_dump()
            row_values: list[Any] = []
            for key in keys:
                val = dict_row.get(key)
                if isinstance(val, datetime) and val.tzinfo is not None:
                    # Excel doesn't support timezones, so we convert to naive.
                    # We strip the timezone info.
                    val = val.replace(tzinfo=None)
                row_values.append(val)
            ws.append(row_values)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def get_exporter(format_name: str) -> DataExporter:
    """Factory to get the appropriate exporter based on format name."""
    exporters = {
        "json": JSONExporter(),
        "csv": CSVExporter(),
        "excel": ExcelExporter(),
        "xlsx": ExcelExporter(),
    }
    exporter = exporters.get(format_name.lower())
    if not exporter:
        raise ValueError(f"Unsupported format: {format_name}. Supported: {list(exporters.keys())}")
    return exporter
