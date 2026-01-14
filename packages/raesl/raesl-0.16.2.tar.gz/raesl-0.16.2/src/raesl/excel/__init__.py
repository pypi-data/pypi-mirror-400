"""Module to export as an Excel workbook."""

from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl.workbook import Workbook
from ragraph.graph import Graph

from raesl import logger
from raesl.doc.locales import register_locale
from raesl.excel import defaults, sheets
from raesl.utils import get_scoped_nodes

# Compile module is excluded during docs generation.
try:
    from raesl.compile import to_graph
except ImportError:
    pass


def convert(
    *paths: Union[str, Path],
    output: Optional[Union[str, Path]] = defaults.OUTPUT,
    scopes: Dict[str, Optional[int]] = defaults.SCOPES,
    language: str = "en",
) -> Workbook:
    """Write (part of) an ESL specification to a Excel workbook.

    Arguments:
        paths: ESL specification paths.
        output: Optional Workbook output path (will be overwritten without warning).
        scopes: Dictionary of component paths to relative depths of subcomponents to
            include as scopes for the generated output. Defaults to the complete tree.
        language: Output language.

    Returns:
        Excel workbook instance.
    """

    graph = to_graph(*paths)
    return overview(graph, output, scopes, language=language)


def overview(
    graph: Graph,
    output: Optional[Union[str, Path]] = defaults.OUTPUT,
    scopes: Dict[str, Optional[int]] = defaults.SCOPES,
    language: str = "en",
) -> Workbook:
    """Write (part of) an ESL specification to a Excel workbook.

    Arguments:
        graph: Compiled ESL graph.
        output: Optional Workbook output path (will be overwritten without warning).
        scopes: Dictionary of component paths to relative depths of subcomponents to
            include as scopes for the generated output. Defaults to the complete tree.
        language: Output language.

    Returns:
        Excel workbook instance.
    """
    # Create workbook, but delete default sheet.
    register_locale(language)

    wb = Workbook()
    wb.remove(wb["Sheet"])

    # Derive components from scopes.
    components = [node for node in get_scoped_nodes(graph, scopes) if node.kind == "component"]
    if not components:
        raise ValueError(f"No components found in selected scopes ('{scopes}'). Please reconsider.")

    # Add sheets.
    _, components = sheets.add_components_sheet(wb, components)
    _, goals = sheets.add_goals_sheet(wb, graph, components)
    _, transformations = sheets.add_transformations_sheet(wb, graph, components)
    _, designs = sheets.add_designs_sheet(wb, graph, components)
    _, behaviors = sheets.add_behaviors_sheet(wb, graph, components)
    _, needs = sheets.add_needs_sheet(wb, graph, components)
    _, variables = sheets.add_variable_sheet(wb, graph, components)

    sheets.add_overview_sheet(
        wb, graph, components, goals, transformations, designs, behaviors, needs
    )

    # Protect all sheets.
    for sheet in wb.sheetnames:
        wb[sheet].protection = defaults.SHEETPROTECTION

    # Write output if path is given.
    if output:
        wb.save(str(output))
    return wb


def component_overview(
    graph: Graph,
    component_path: str,
    flow_labels: Optional[List[str]] = None,
    output: Optional[Union[str, Path]] = None,
    language: str = "en",
) -> Workbook:
    """Create a component overview Excel sheet. Somewhat like a free body diagram for a single
    component. Meant to be more discrete but also compact than the complete overview Excel.

    Arguments:
        graph: Compiled ESL graph.
        component_path: ESL specification paths.
        flow_labels: Flow types to include in sheets.
        output: Optional Workbook output path (will be overwritten without warning).
        language: Output language.

    Returns:
        Excel workbook instance.
    """
    register_locale(language)

    # Create workbook, but delete default sheet.
    wb = Workbook()
    wb.remove(wb["Sheet"])

    try:
        if not component_path.startswith("world."):
            component_path = "world." + component_path
        component = graph[component_path]
    except KeyError:
        logger.error("Could not find component!")
        return

    sheets.add_component_active_goals_sheet(wb, graph, component, flow_labels)
    sheets.add_component_passive_goals_sheet(wb, graph, component, flow_labels)

    if output:
        wb.save(str(output))
    return wb
