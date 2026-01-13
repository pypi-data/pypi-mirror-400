import io
import re
from datetime import date, datetime, timedelta
from decimal import (
    ROUND_DOWN,
    ROUND_FLOOR,
    ROUND_HALF_UP,
    Decimal,
    getcontext,
    localcontext,
)
from pathlib import Path

import numpy as np
import pandas as pd
from dateutil import parser as date_parser
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import (
    DecimalField,
    ExpressionWrapper,
    F,
    IntegerField,
    Q,
    Sum,
    Value,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.encoding import escape_uri_path
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from fnschool import _, count_chinese_characters
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .forms import (
    CategoryForm,
    ConsumptionForm,
    IngredientForm,
    MealTypeForm,
    PurchasedIngredientsWorkBookForm,
)
from .models import (
    Category,
    Consumption,
    Ingredient,
    MealType,
    category_name_0,
    meal_type_name_0,
)

# Create your views here.

decimal_prec = getattr(settings, "DECIMAL_PREC", 2)

storage_date_header = (
    _("Storage Date"),
    _(
        'Formats like "YYYY.mm.dd", "YYYY/mm/dd", '
        + '"YYYY.mm.dd", "mm/dd", "mmdd", and "mm.dd" '
        + "are all acceptable. In short, FNSCHOOL wants to be "
        + "compatible with all the formats you like, but if "
        + "something goes wrong, you have to tell the "
        + "developers. Thank you!"
    ),
)
ingredient_name_header = (_("Ingredient Name"), _("Name of Ingredient"))
meal_type_header = (
    _("Meal Type"),
    _(
        "For example, breakfast, dinner, regular meals, "
        + "nutritious meals, etc., when generating a spreadsheet,"
        + " each meal category corresponds to a spread sheet. "
        + "If left blank, only one spreadsheet will be generated."
    ),
)

category_header = (
    _("Category"),
    _(
        "Usually there are seven categories: vegetables, meat, "
        + "grains, seasonings, eggs and milk, oils, and fruits."
    ),
)
quantity_header = (
    _("Quantity"),
    _(
        "To prevent the number of decimal places in the unit "
        + 'price from becoming too large, "quantity" is '
        + "only allowed to be an integer."
        + " If your unit price is a decimal, please use a "
        + "smaller quantity unit and then expand the "
        + "quantity to an integer."
    ),
)


total_price_header = (
    _("Total Price"),
    None,
)
quantity_unit_name_header = (
    _("Unit Name of Quantity"),
    None,
)
is_ignorable_header = (
    _("Is Ignorable"),
    _("As long as a cell has content, it will be considered " + 'as "yes".'),
)


date_patterns = [
    (r"\b\d{4}-\d{2}-\d{2}\b", "%Y-%m-%d"),
    (r"\b\d{4}/\d{2}/\d{2}\b", "%Y/%m/%d"),
    (r"\b\d{4}\.\d{2}\.\d{2}\b", "%Y.%m.%d"),
    (r"\b\d{8}\b", "%Y%m%d"),
]


def get_decimal_places(decimal_number):
    if not isinstance(decimal_number, Decimal):
        decimal_number = Decimal(str(decimal_number))
    tuple_rep = decimal_number.as_tuple()
    return abs(tuple_rep.exponent) if tuple_rep.exponent < 0 else 0


def split_price(total_price, quantity, prec=2):
    prec = prec
    prec_decimal = Decimal("0." + ("0" * prec))

    total_price0 = Decimal(str(total_price))
    quantity0 = Decimal(str(quantity))
    unit_price0 = total_price0 / quantity0

    unit_price_floor = unit_price0.quantize(prec_decimal, rounding=ROUND_FLOOR)

    if unit_price_floor == unit_price0:
        return [[total_price0, quantity0], [None, None]]

    total_price_floor = quantity0 * unit_price_floor
    total_price_diff = total_price0 - total_price_floor

    split_quantity = total_price_diff * Decimal(str(10**prec))

    unit_price0 = unit_price_floor
    quantity0 = quantity0 - split_quantity
    total_price0 = quantity0 * unit_price0

    unit_price1 = unit_price_floor + Decimal(str(1 / (10**prec)))
    quantity1 = split_quantity
    total_price1 = unit_price1 * quantity1

    return [[total_price0, quantity0], [total_price1, quantity1]]


@login_required
def create_consumptions(request, ingredient_id=None):
    date_start = request.GET.get(
        "storage_date_start", None
    ) or request.COOKIES.get("storage_date_start", None)
    if not date_start:
        ingredients = Ingredient.objects.annotate(
            total_consumed=Coalesce(
                Sum("consumptions__amount_used"), 0, output_field=IntegerField()
            )
        ).filter(
            Q(quantity__gt=F("total_consumed"))
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
            & Q(user=request.user)
        )
        date_start = ingredients.order_by("storage_date").first().storage_date

    else:
        date_start = date_parser.parse(date_start).date()

    date_end = request.GET.get("storage_date_end", None) or request.COOKIES.get(
        "storage_date_end", None
    )
    if not date_end:
        today = datetime.now().date()
        date_end = (
            today.replace(day=1) + relativedelta(months=2)
        ) - relativedelta(days=1)
    else:
        date_end = date_parser.parse(date_end).date()

    date_range = list(pd.date_range(start=date_start, end=date_end))
    date_range = [d.date() for d in date_range]

    if ingredient_id:
        ingredient = get_object_or_404(Ingredient, pk=ingredient_id)
        planned_consumptions = []
        consumptions = ingredient.consumptions.filter(
            Q(is_disabled=False)
        ).all()

        consumptions_len = len(consumptions)
        for c_index in range(consumptions_len):
            consumption = consumptions[c_index]
            for c0_index in range(c_index + 1, consumptions_len):
                consumption0 = consumptions[c0_index]
                if consumption.date_of_using == consumption0.date_of_using:
                    consumption.delete()

        for c in consumptions:
            if c.date_of_using < ingredient.storage_date:
                c.delete()

        consumption_dates = list(set([c.date_of_using for c in consumptions]))

        for per_day in date_range:
            consumption = None
            if per_day in consumption_dates:
                consumptions_per_day = [
                    c for c in consumptions if c.date_of_using == per_day
                ]
                consumption = consumptions_per_day[0]

            else:
                consumption = Consumption()
                consumption.ingredient = ingredient
                consumption.date_of_using = per_day

            if per_day < ingredient.storage_date:
                consumption.is_disabled = True

            planned_consumptions.append(consumption)

        form_list = []
        for c in planned_consumptions:
            form = ConsumptionForm(instance=c)
            form.fields["date_of_using"].label = ""
            form_list.append(form)

        return render(
            request,
            "canteen/consumption/_create.html",
            {"form_list": form_list},
        )

    queries = (
        Q(storage_date__lte=date_end)
        & Q(user=request.user)
        & Q(is_disabled=False)
        & Q(is_ignorable=False)
    )
    ingredients = Ingredient.objects

    sort_values = []
    for key, value in request.GET.items():
        if key.startswith("sort_") and value:
            key = key[5:]
            value = "" if value == "+" else "-"
            sort_values.append(value + key)
    ingredients = (
        ingredients.filter(queries).order_by(*sort_values)
        if sort_values
        else ingredients.filter(queries).order_by("storage_date")
    )

    ingredients = ingredients.prefetch_related("consumptions")
    ingredients = ingredients.distinct()

    if not ingredients:
        return render(
            request,
            "canteen/consumption/create.html",
            {"ingredients": ingredients, "date_range": date_range},
        )

    ingredients = [
        i
        for i in ingredients
        if (
            i.get_remaining_quantity(date_end) > 0
            or i.storage_date >= date_start
            or any(
                [
                    date_start <= c.date_of_using <= date_end
                    for c in i.consumptions.all()
                ]
            )
        )
    ]

    ingredients_pinned = []
    ingredients_unpinned = []
    categories_top = Category.objects.filter(
        Q(user=request.user)
        & Q(pin_to_consumptions_top=True)
        & Q(is_disabled=False)
    ).all()

    for i in ingredients:
        if i.category in categories_top:
            ingredients_pinned.append(i)
        else:
            ingredients_unpinned.append(i)

    ingredients = ingredients_pinned + ingredients_unpinned

    for ingredient in ingredients:
        ingredient.quantity_used = ingredient.get_consuming_quantity(
            (date_start + timedelta(days=-1))
        ) + sum(
            [
                c.amount_used
                for c in i.consumptions.all()
                if c.date_of_using > date_end
            ]
        )

    date_range_cp = date_range
    date_range_cp = [d.strftime("%Y-%m-%d") for d in date_range]
    meal_types = list(set([i.meal_type.name for i in ingredients]))
    ingredient_ids = [i.id for i in ingredients]
    months = list(set([d.strftime("%Y-%m") for d in date_range]))
    months = sorted(months, key=lambda d: int(d.split("-")[1]))
    return render(
        request,
        "canteen/consumption/create.html",
        {
            "ingredients": ingredients,
            "date_range": date_range_cp,
            "meal_types": meal_types,
            "ingredient_ids": ingredient_ids,
            "months": months,
            "storage_date_start": date_start.strftime("%Y-%m-%d"),
            "storage_date_end": date_end.strftime("%Y-%m-%d"),
        },
    )


@login_required
@require_POST
def new_consumption(request, consumption_id=None):
    form = None

    posted_date_of_using = date_parser.parse(request.POST.get("date_of_using"))
    posted_amount_used = request.POST.get("amount_used")

    if consumption_id:
        consumption = Consumption.objects.filter(
            Q(pk=consumption_id)
            & Q(ingredient__user=request.user)
            & Q(date_of_using=posted_date_of_using)
            & Q(is_disabled=False)
        ).first()
        if consumption:
            if Decimal(posted_amount_used).is_zero():
                consumption.delete()
                return HttpResponse("OK", status=201)

            form = ConsumptionForm(request.POST, instance=consumption)
        else:
            return HttpResponse("Accepted", status=202)

    else:
        ingredient_id = request.POST.get("ingredient")
        ingredient = Ingredient.objects.filter(
            Q(user=request.user) & Q(pk=ingredient_id)
        ).first()
        if not ingredient:
            return HttpResponse("Accepted", status=202)

        consumption = Consumption.objects.filter(
            Q(date_of_using=posted_date_of_using)
            & Q(ingredient__id=ingredient_id)
        ).first()

        if not consumption:
            consumption = Consumption()
        elif Decimal(posted_amount_used).is_zero():
            consumption.delete()
            return HttpResponse("OK", status=201)

        consumption.ingredient = ingredient
        form = ConsumptionForm(request.POST, instance=consumption)

    if form.is_valid() and not form.instance.is_disabled:
        consumption = form.save(commit=False)
        consumption.save()
        return HttpResponse("OK", status=201)

    return HttpResponse("Accepted", status=202)


@login_required()
def delete_ingredient(request, ingredient_id):
    ingredient = get_object_or_404(Ingredient, pk=ingredient_id)

    if request.method == "POST":
        if ingredient.user == request.user:
            ingredient.delete()
            return render(
                request,
                "canteen/close.html",
            )

    form = IngredientForm(instance=ingredient)
    return render(request, "canteen/ingredient/delete.html", {"form": form})


def edit_ingredient(request, ingredient_id):
    ingredient = get_object_or_404(Ingredient, pk=ingredient_id)

    if request.method == "POST":

        total_price = request.POST.get("total_price")
        quantity = request.POST.get("quantity")

        [total_price0, quantity0], [total_price1, quantity1] = split_price(
            total_price, quantity
        )

        form = IngredientForm(request.POST, instance=ingredient)
        if total_price1:
            unit_price_error_msg = _(
                "The unit pricei ({unit_price}) has more than {decimal_prec} decimal places and cannot be saved. Please modify it again."
            ).format(
                unit_price=(
                    Decimal(str(total_price)) / Decimal(str(float(quantity)))
                ).normalize(),
                decimal_prec=decimal_prec,
            )
            form.add_error("total_price", unit_price_error_msg)
            form.add_error("quantity", unit_price_error_msg)
            form.fields["total_price"].widget.attrs.update({"autofocus": ""})
            return render(
                request, "canteen/ingredient/update.html", {"form": form}
            )

        form.instance.user = request.user
        if form.is_valid():
            form.save()
            return render(
                request,
                "canteen/close.html",
            )
    else:
        form = IngredientForm(instance=ingredient)

    return render(request, "canteen/ingredient/update.html", {"form": form})


@login_required
def list_ingredients(request):
    search_query = request.GET.get("q", "")
    search_query_cp = search_query
    fields = [
        f
        for f in Ingredient._meta.fields
        if f.name in IngredientForm._meta.fields
    ]
    if search_query:
        queries = Q(user=request.user)

        search_query_dates = []

        for pattern, fmt in date_patterns:
            matches = re.findall(pattern, search_query)
            for match in matches:
                try:
                    date_obj = datetime.strptime(match, fmt).date()
                    search_query_dates.append(date_obj)
                    search_query = search_query.replace(match, "")

                except ValueError:
                    continue

        if len(search_query_dates) > 1:
            queries &= Q(storage_date__gte=min(search_query_dates))
            queries &= Q(storage_date__lte=max(search_query_dates))
        elif len(search_query_dates) == 1:
            queries &= Q(storage_date=search_query_dates[0])

        unit_names = Ingredient.objects.values("quantity_unit_name").distinct()
        unit_names = [
            c.get("quantity_unit_name")
            for c in unit_names
            if c.get("quantity_unit_name") in str(search_query)
        ]
        for unit_name in unit_names:
            queries &= Q(quantity_unit_name__icontains=unit_name)
            search_query = search_query.replace(unit_name, "")

        categories = Ingredient.objects.values("category__name").distinct()
        categories = [
            c.get("category__name")
            for c in categories
            if c.get("category__name") in search_query
        ]
        for category in categories:
            queries &= Q(category__name__icontains=category)
            search_query = search_query.replace(category, "")

        meal_types = Ingredient.objects.values("meal_type__name").distinct()
        meal_types = [
            m.get("meal_type__name")
            for m in meal_types
            if m.get("meal_type__name") in search_query
        ]
        for meal_type in meal_types:
            queries &= Q(meal_type__name__icontains=meal_type)
            search_query = search_query.replace(meal_type, "")

        names = re.split(r"\s+", search_query)
        name_queries = Q()
        for name in names:
            name_queries |= Q(name__icontains=name)
        queries &= name_queries

        ingredients = Ingredient.objects.filter(queries)

    else:
        ingredients = Ingredient.objects.filter(Q(user=request.user))

    orders = []
    for f in fields:
        sort_name = request.GET.get("sort_" + f.name, "")
        if sort_name and sort_name in "+-":
            sort_name = (
                sort_name[1:] if sort_name.startswith("+") else sort_name
            )
            sort_name += f.name
            orders.append(sort_name)
    if len(orders) < 1:
        ingredients = ingredients.order_by("storage_date", "category")
    else:
        ingredients = ingredients.order_by(*orders)

    page_size = request.GET.get("page_size", "")
    if not page_size:
        page_size = request.COOKIES.get("page_size", "")
    page_size = int(page_size) if str(page_size).isnumeric() else 10
    ingredients_len = len(ingredients)
    paginator = Paginator(ingredients, page_size)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)
    headers = [
        (f.name, request.GET.get("sort_" + f.name, ""), f.verbose_name)
        for f in fields
    ]

    meal_types = list(
        set([i.meal_type for i in ingredients if not i.meal_type.is_disabled])
    )
    total_price_title = ""
    total_price = Decimal("0.0")
    for meal_type in meal_types:
        sum_total_price = sum(
            ingredients.filter(Q(meal_type=meal_type)).values_list(
                "total_price", flat=True
            )
        )
        sum_total_price = Decimal(sum_total_price)
        total_price += sum_total_price
        total_price_title += "\n" + _("{meal_type}:{sum_total_price}").format(
            meal_type=meal_type.name,
            sum_total_price=sum_total_price.normalize(),
        )
    total_price_title = str(total_price.normalize()) + total_price_title

    context = {
        "page_obj": page_obj,
        "search_query": search_query_cp,
        "headers": headers,
        "page_size": page_size,
        "total_price_title": total_price_title,
        "ingredients_len": ingredients_len,
    }
    return render(request, "canteen/ingredient/list.html", context)


def create_ingredients(request):
    if request.method == "POST":
        form = PurchasedIngredientsWorkBookForm(request.POST, request.FILES)
        if form.is_valid():
            workbook_file = request.FILES["workbook_file"]

            if not workbook_file.name.endswith(".xlsx"):
                return HttpResponse(_('Please upload a file in "xlsx" format.'))

            df = pd.read_excel(workbook_file)
            saved_ingredients = (
                Ingredient.objects.filter(
                    Q(user=request.user) & Q(is_disabled=False)
                )
                .select_related("category")
                .only("name", "quantity_unit_name", "category__name")
            )

            new_ingredients = []
            for index, row in df.iterrows():
                category_name = row[category_header[0]]
                category_name = (
                    None if pd.isnull(category_name) else category_name
                )
                meal_type_name = row[meal_type_header[0]]
                meal_type_name = (
                    None if pd.isnull(meal_type_name) else meal_type_name
                )
                name = row[ingredient_name_header[0]]

                category = None

                if category_name:
                    named_category = Category.objects.filter(
                        Q(name=category_name) & Q(user=request.user)
                    ).first()

                    if named_category:
                        category = named_category
                    else:
                        new_category = Category.objects.create(
                            user=request.user,
                            name=category_name,
                        )
                        category = new_category
                else:
                    saved_categories = [
                        i.category
                        for i in saved_ingredients
                        if i.name == name and i.name != ""
                    ]
                    if saved_categories:
                        category = saved_categories[0]
                    else:
                        category_0 = Category.objects.filter(
                            Q(user=request.user) & Q(name="")
                        ).first()
                        if not category_0:
                            category_0 = Category.objects.create(
                                user=request.user,
                                name="",
                            )

                        category = category_0

                meal_type = MealType.objects.filter(
                    Q(name=meal_type_name or "") & Q(user=request.user)
                ).first()
                if not meal_type:
                    meal_type = MealType.objects.create(
                        user=request.user,
                        name="",
                    )

                storage_date = row[storage_date_header[0]]
                storage_date = date_parser.parse(str(storage_date))

                quantity_unit_name = row[quantity_unit_name_header[0]]
                if not quantity_unit_name:
                    if saved_ingredients:
                        quantity_unit_names = [
                            i.quantity_unit_name
                            for i in saved_ingredients
                            if i.name == name
                        ]
                        quantity_unit_name = (
                            quantity_unit_names[0]
                            if quantity_unit_names
                            else ""
                        )
                    else:
                        quantity_unit_name = (
                            quantity_unit_names[0]
                            if quantity_unit_names
                            else ""
                        )

                is_ignorable = not pd.isnull(row[is_ignorable_header[0]])

                total_price0 = row[total_price_header[0]]
                quantity0 = row[quantity_header[0]]

                [total_price0, quantity0], [total_price1, quantity1] = (
                    split_price(total_price0, quantity0)
                )

                if total_price1:
                    new_ingredients.append(
                        Ingredient(
                            user=request.user,
                            storage_date=storage_date,
                            name=name + _("(2)"),
                            meal_type=meal_type,
                            category=category,
                            quantity=quantity1,
                            total_price=total_price1,
                            quantity_unit_name=quantity_unit_name,
                            is_ignorable=is_ignorable,
                        )
                    )
                    name = name + _("(1)")

                new_ingredients.append(
                    Ingredient(
                        user=request.user,
                        storage_date=storage_date,
                        name=name,
                        meal_type=meal_type,
                        category=category,
                        quantity=quantity0,
                        total_price=total_price0,
                        quantity_unit_name=quantity_unit_name,
                        is_ignorable=is_ignorable,
                    )
                )
            Ingredient.objects.bulk_create(new_ingredients)

            return redirect("canteen:close_window")

    else:
        form = PurchasedIngredientsWorkBookForm()
    return render(request, "canteen/ingredient/create.html", {"form": form})


def get_template_workbook_of_purchased_ingredients(request):
    global storage_date_header, ingredient_name_header, meal_type_header
    global quantity_header, quantity_unit_name_header, total_price_header
    global is_ignorable_header
    headers = [
        storage_date_header,
        ingredient_name_header,
        meal_type_header,
        category_header,
        quantity_header,
        quantity_unit_name_header,
        total_price_header,
        is_ignorable_header,
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = _("Purchased Ingredients Sheet")

    for i, (h, c) in enumerate(headers):
        h_cell = ws.cell(1, i + 1)

        center_alignment = Alignment(horizontal="center", vertical="center")
        h_cell.alignment = center_alignment

        mono_font = Font(
            name="Mono",
            size=12,
        )
        h_cell.font = mono_font

        h_cell.value = h
        column_letter = get_column_letter(i + 1)
        hans_len = count_chinese_characters(h)
        hans_len = (hans_len + 2) if hans_len else hans_len
        ws.column_dimensions[column_letter].width = len(h) + hans_len + 2
        if c:
            h_cell.comment = Comment(c, _("the FNSCHOOL Authors"))
    ws.freeze_panes = "B2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type=(
            "application/"
            + "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    today = datetime.now().date()
    last_month = today + relativedelta(months=-1, day=1)
    filename = (
        _("Purchased Ingredients WorkBook ({0})").format(
            last_month.strftime("%Y%m")
        )
        + ".xlsx"
    )

    encoded_filename = escape_uri_path(filename)
    response["Content-Disposition"] = (
        f'attachment; filename="{encoded_filename}"'
    )

    return response


def close_window(request):
    return render(request, "canteen/close.html")


def generate_spreadsheet(request, month):
    from .workbook.generate import get_workbook_zip

    buffer = get_workbook_zip(request, month)
    filename = (
        _("Canteen Daybook WorkBook ({month}) of {affiliation}").format(
            month=month.replace("-", ""), affiliation=request.user.affiliation
        )
        + ".zip"
    )
    filename = escape_uri_path(filename)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


class MealTypeDeleteView(LoginRequiredMixin, DeleteView):
    model = MealType
    template_name = "canteen/meal_type/delete.html"
    success_url = reverse_lazy("canteen:close_window")
    context_object_name = "meal_type"

    def get_object(self, queryset=None):
        meal_type = super().get_object(queryset)
        if meal_type.user != self.request.user:
            raise Http404()
        return meal_type


class MealTypeUpdateView(LoginRequiredMixin, UpdateView):
    model = MealType
    template_name = "canteen/meal_type/update.html"
    success_url = reverse_lazy("canteen:close_window")
    form_class = MealTypeForm

    def get_object(self, queryset=None):
        meal_type = super().get_object(queryset)
        if meal_type.user != self.request.user:
            raise Http404()
        return meal_type

    def get_initial(self):
        return {"user": self.request.user, "created_at": datetime.now().date}


class MealTypeCreateView(LoginRequiredMixin, CreateView):
    model = MealType
    template_name = "canteen/meal_type/create.html"
    success_url = reverse_lazy("canteen:close_window")
    form_class = MealTypeForm

    def get_initial(self):
        return {"user": self.request.user, "created_at": datetime.now().date}

    def form_valid(self, form):
        meal_type_saved = MealType.objects.filter(
            Q(name=form.instance.name) & Q(user=self.request.user)
        ).first()
        if meal_type_saved:
            return redirect("canteen:close_window")
        form.instance.user = self.request.user
        form.instance.created_at = datetime.now().date()
        return super().form_valid(form)


class MealTypeListView(LoginRequiredMixin, ListView):
    model = MealType
    template_name = "canteen/meal_type/list.html"
    context_object_name = "meal_types"
    ordering = ["-created_at"]

    paginate_by = 10
    paginate_orphans = 2

    def get_paginate_by(self, queryset):
        page_size = self.request.GET.get("page_size")
        page_size = (
            page_size if page_size else self.request.COOKIES.get("page_size")
        )
        page_size = page_size if page_size else self.paginate_by
        return int(page_size)


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    model = Category
    template_name = "canteen/category/delete.html"
    success_url = reverse_lazy("canteen:close_window")
    context_object_name = "category"

    def get_object(self, queryset=None):
        category = super().get_object(queryset)
        if category.user != self.request.user:
            raise Http404()
        return category


class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = Category
    template_name = "canteen/category/update.html"
    success_url = reverse_lazy("canteen:close_window")
    form_class = CategoryForm

    def get_object(self, queryset=None):
        category = super().get_object(queryset)
        if category.user != self.request.user:
            raise Http404()
        return category

    def get_initial(self):
        return {"user": self.request.user, "created_at": datetime.now().date}


class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    template_name = "canteen/category/create.html"
    success_url = reverse_lazy("canteen:close_window")
    form_class = CategoryForm

    def get_initial(self):
        return {"user": self.request.user, "created_at": datetime.now().date}

    def form_valid(self, form):
        category_saved = Category.objects.filter(
            Q(name=form.instance.name) & Q(user=self.request.user)
        ).first()
        if category_saved:
            return redirect("canteen:close_window")
        form.instance.user = self.request.user
        form.instance.created_at = datetime.now().date()
        return super().form_valid(form)


class CategoryListView(LoginRequiredMixin, ListView):
    model = Category
    template_name = "canteen/category/list.html"
    context_object_name = "categories"

    paginate_by = 10
    paginate_orphans = 2

    def get_paginate_by(self, queryset):
        page_size = self.request.GET.get("page_size")
        page_size = (
            page_size if page_size else self.request.COOKIES.get("page_size")
        )
        page_size = page_size if page_size else self.paginate_by
        return int(page_size)


class IngredientCreateView(LoginRequiredMixin, CreateView):
    model = Ingredient
    template_name = "canteen/ingredient/create_one.html"
    success_url = reverse_lazy("canteen:close_window")
    form_class = IngredientForm

    def get_initial(self):
        return {
            "user": self.request.user,
            "storage_date": datetime.now().date(),
        }

    def form_valid(self, form):
        form.instance.user = self.request.user
        total_price = form.instance.total_price
        quantity = form.instance.quantity

        [total_price0, quantity0], [total_price1, quantity1] = split_price(
            total_price, quantity
        )

        if form.is_valid() and total_price1:
            Ingredient.objects.create(
                user=form.instance.user,
                storage_date=form.instance.storage_date,
                name=form.instance.name + _("(2)"),
                meal_type=form.instance.meal_type,
                category=form.instance.category,
                quantity=quantity1,
                total_price=total_price1,
                quantity_unit_name=form.instance.quantity_unit_name,
                is_ignorable=form.instance.is_ignorable,
            )
            form.instance.name = form.instance.name + _("(1)")

        form.instance.total_price = total_price0
        form.instance.quantity = quantity0

        return super().form_valid(form)


# The end.
