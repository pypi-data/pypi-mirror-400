from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, TypeVar

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .column import Column

M = TypeVar("M")


def _camel_to_snake(name: str) -> str:
    s1 = re.sub("(.)([A-Z][a-z]+)", r"\1_\2", name)
    s2 = re.sub("([a-z0-9])([A-Z])", r"\1_\2", s1)
    return s2.lower()


def _pluralize(s: str) -> str:
    # keep deliberately simple; can be swapped for inflect later
    if s.endswith("s"):
        return s
    if s.endswith("y") and len(s) >= 2 and s[-2] not in "aeiou":
        return s[:-1] + "ies"
    return s + "s"


def _repo_name_for_model(model: type[Any]) -> str:
    return _pluralize(_camel_to_snake(model.__name__))


def _display_name_for_model(model: type[Any]) -> str:
    # "manufacturing_plants" -> "Manufacturing Plants"
    return _repo_name_for_model(model).replace("_", " ").title()


def _get_model_columns(model: type[Any]) -> list[Column[Any]]:
    return list(getattr(model, "__columns__", []))


def _normalize_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _row_is_blank(values: list[Any]) -> bool:
    return all(_normalize_header(v) == "" for v in values)


def _instantiate_model[M](model: type[M]) -> M:
    obj = model.__new__(model)
    obj._values = {}
    # defaults
    for col in _get_model_columns(model):
        if col.name is None:
            raise RuntimeError("Column __set_name__ did not run.")
        obj._values[col.name] = col.spec.default
    return obj


class Repository(list[M]):
    def all(self) -> list[M]:
        return list(self)


@dataclass(frozen=True)
class SheetSpec:
    name: str
    models: list[type[Any]]

    title_row: int = 1
    header_row: int = 2
    data_start_row: int = 3

    template_table_gap: int = 2


@dataclass(frozen=True)
class PivotSheetSpec:
    name: str
    model: type[Any]  # single model only

    # field names on the model
    pivot_field: str
    row_fields: list[str]
    value_field: str

    # layout
    title_row: int = 1
    header_row: int = 2  # pivot headers
    row_header_col: int = 1
    data_start_row: int = 3

    # template: define the pivot column values (dates) to render across the top
    pivot_values: list[Any] | None = None  # e.g., list[date]; required for generation

    # optional: seed row keys (regions) on template
    row_values: list[Any] | None = None

    include_blanks: bool = False  # whether to load blank cells as data points

    @property
    def data_start_col(self) -> int:
        return len(self.row_fields) + 1


AnySheetSpec = PivotSheetSpec | SheetSpec


class ExcelFile:
    def __init__(self, *, sheets: list[AnySheetSpec]):
        self.sheets = sheets

        self._repos: dict[type[Any], Repository[Any]] = {}

        for sheet in sheets:
            models = [sheet.model] if isinstance(sheet, PivotSheetSpec) else sheet.models

            for model in models:
                repo_name = _repo_name_for_model(model)
                if hasattr(self, repo_name):
                    raise ValueError(
                        f"Duplicate repo name '{repo_name}' for model {model.__name__}"
                    )
                repo = Repository()
                self._repos[model] = repo
                setattr(self, repo_name, repo)

    def generate_template(self, filename: str) -> None:
        wb = Workbook()
        default_ws = wb.active
        if self.sheets:
            wb.remove(default_ws)

        for sheet in self.sheets:
            ws = wb.create_sheet(title=sheet.name)
            if isinstance(sheet, PivotSheetSpec):
                self._write_pivot_sheet_template(ws, sheet)
            else:
                self._write_sheet_template(ws, sheet)

        wb.save(filename)

    def _write_sheet_template(self, ws: Worksheet, spec: SheetSpec) -> None:
        current_col = 1  # 1-based index

        title_font = Font(bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")

        header_font = Font(bold=True)

        for model in spec.models:
            cols = _get_model_columns(model)
            headers = [c.spec.header or c.name for c in cols]
            width = len(headers)

            start_col = current_col
            end_col = current_col + width - 1

            # ---- merged title row ----
            ws.merge_cells(
                start_row=spec.title_row,
                start_column=start_col,
                end_row=spec.title_row,
                end_column=end_col,
            )
            title_cell = ws.cell(
                row=spec.title_row, column=start_col, value=_display_name_for_model(model)
            )
            title_cell.font = title_font
            title_cell.alignment = title_alignment

            for j, h in enumerate(headers):
                c = start_col + j
                cell = ws.cell(row=spec.header_row, column=c, value=h)
                cell.font = header_font

                col_letter = ws.cell(row=spec.header_row, column=c).column_letter
                ws.column_dimensions[col_letter].width = max(12, min(40, len(str(h)) + 4))

            current_col = end_col + 1 + spec.template_table_gap

    def load_data(self, filename: str) -> None:
        wb = load_workbook(filename=filename, data_only=True)
        for repo in self._repos.values():
            repo.clear()

        for sheet_spec in self.sheets:
            if sheet_spec.name not in wb.sheetnames:
                raise ValueError(f"Workbook missing sheet '{sheet_spec.name}'")
            ws = wb[sheet_spec.name]
            if isinstance(sheet_spec, PivotSheetSpec):
                self._parse_pivot_sheet(ws, sheet_spec)
            else:
                self._parse_sheet(ws, sheet_spec)

    def _parse_sheet(self, ws: Worksheet, spec: SheetSpec) -> None:
        for model in spec.models:
            found = self._find_header(ws, spec, model)
            if found is None:
                continue

            _, start_col = found
            cols = _get_model_columns(model)
            width = len(cols)

            repo: Repository[Any] = self._repos[model]

            r = spec.data_start_row
            while r <= ws.max_row:
                row_vals = [ws.cell(row=r, column=start_col + j).value for j in range(width)]
                if _row_is_blank(row_vals):
                    break

                # excludes (raw-value based)
                if any(
                    col.spec.excludes and row_vals[i] in col.spec.excludes
                    for i, col in enumerate(cols)
                ):
                    r += 1
                    continue

                obj = _instantiate_model(model)
                for i, col in enumerate(cols):
                    raw = row_vals[i]
                    parsed = col.parse_cell(raw)
                    setattr(obj, col.name, parsed)

                validate = getattr(obj, "validate", None)
                if callable(validate):
                    validate()

                repo.append(obj)
                r += 1

    def _find_header(
        self, ws: Worksheet, spec: SheetSpec, model: type[Any]
    ) -> tuple[int, int] | None:
        cols = _get_model_columns(model)
        expected = [_normalize_header(c.spec.header) for c in cols]
        if not expected:
            return None

        r = spec.header_row
        width = len(expected)
        max_c = ws.max_column or 0

        for start_col in range(1, max_c - width + 2):
            actual = [
                _normalize_header(ws.cell(row=r, column=start_col + j).value) for j in range(width)
            ]
            if actual == expected:
                return (r, start_col)

        return None

    def _write_pivot_sheet_template(self, ws: Worksheet, spec: PivotSheetSpec) -> None:
        if not spec.pivot_values:
            raise ValueError("PivotSheetSpec.pivot_values is required for template generation.")

        title_font = Font(bold=True)
        title_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True)

        # Title merged across the pivot header span
        end_col = spec.data_start_col + len(spec.pivot_values) - 1

        ws.merge_cells(
            start_row=spec.title_row,
            start_column=spec.row_header_col,
            end_row=spec.title_row,
            end_column=end_col,
        )
        tcell = ws.cell(spec.title_row, spec.row_header_col, _display_name_for_model(spec.model))
        tcell.font = title_font
        tcell.alignment = title_alignment

        # Row field headers (left block)
        for i, rf in enumerate(spec.row_fields):
            c = spec.row_header_col + i
            cell = ws.cell(spec.header_row, c, rf.title())
            cell.font = header_font
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = max(12, min(40, len(str(rf)) + 4))

        # Pivot headers across the top
        for j, pv in enumerate(spec.pivot_values):
            c = spec.data_start_col + j
            cell = ws.cell(spec.header_row, c, pv)
            cell.font = header_font
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = 14

        if spec.row_values:
            for i, rv in enumerate(spec.row_values):
                r = spec.data_start_row + i

                if len(spec.row_fields) == 1:
                    # Back-compat: rv is a single value
                    ws.cell(r, spec.row_header_col, rv)
                else:
                    # Expect rv to be a tuple/list matching row_fields
                    if not isinstance(rv, (tuple, list)) or len(rv) != len(spec.row_fields):
                        raise ValueError(
                            "For multi-row_fields, PivotSheetSpec.row_values must be a list of "
                            f"tuples/lists with length {len(spec.row_fields)}."
                        )
                    for k, part in enumerate(rv):
                        ws.cell(r, spec.row_header_col + k, part)

    def _parse_pivot_sheet(self, ws: Worksheet, spec: PivotSheetSpec) -> None:
        model = spec.model
        cols = {c.name: c for c in _get_model_columns(model)}  # Column descriptors by field name

        # Validate fields exist
        required = [spec.pivot_field, spec.value_field, *spec.row_fields]
        missing = [f for f in required if f not in cols]
        if missing:
            raise ValueError(
                f"{model.__name__} is missing Column field(s) required by PivotSheetSpec: {missing}"
            )

        pivot_col = cols[spec.pivot_field]
        val_col = cols[spec.value_field]
        row_cols = [cols[f] for f in spec.row_fields]

        # Determine pivot headers from sheet (or trust spec.pivot_values)
        pivot_headers: list[Any] = []
        j = 0
        while True:
            c = spec.data_start_col + j
            raw = ws.cell(spec.header_row, c).value
            if raw is None or str(raw).strip() == "":
                break
            pivot_headers.append(pivot_col.parse_cell(raw))
            j += 1

        if not pivot_headers:
            return

        repo: Repository[Any] = self._repos[model]

        r = spec.data_start_row
        while r <= ws.max_row:
            raw_parts = [
                ws.cell(r, spec.row_header_col + i).value for i in range(len(spec.row_fields))
            ]
            if _row_is_blank(raw_parts):
                break

            row_parts = [row_cols[i].parse_cell(raw_parts[i]) for i in range(len(row_cols))]

            for j, pivot_value in enumerate(pivot_headers):
                c = spec.data_start_col + j
                raw_val = ws.cell(r, c).value
                if not spec.include_blanks and (raw_val is None or raw_val == ""):
                    continue

                obj = _instantiate_model(model)

                for fname, parsed in zip(spec.row_fields, row_parts, strict=False):
                    setattr(obj, fname, parsed)
                setattr(obj, spec.pivot_field, pivot_value)
                setattr(obj, spec.value_field, val_col.parse_cell(raw_val))

                validate = getattr(obj, "validate", None)
                if callable(validate):
                    validate()

                repo.append(obj)

            r += 1
