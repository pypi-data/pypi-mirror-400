"""
FormulaCalculator测试用例
"""
import pytest
import base64
import openpyxl
import io
from free_mcp_excel.calculator import FormulaCalculator


@pytest.fixture
def calculator():
    """创建计算器实例"""
    return FormulaCalculator()


@pytest.fixture
def test_file_with_formula():
    """创建包含公式的测试文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_content = buffer.read()
    wb.close()
    
    return base64.b64encode(file_content).decode("utf-8")


class TestFormulaCalculator:
    """FormulaCalculator测试类"""
    
    def test_calc_cell_data_read_calculated(self, calculator, test_file_with_formula):
        """测试读取已计算值"""
        result = calculator.calc_cell_data(
            test_file_with_formula,
            "Sheet",
            "A3"
        )
        assert result["status"] == "success"
        assert "value" in result["data"]
        assert "formula" in result["data"]
    
    def test_calc_cell_data_direct_value(self, calculator, test_file_with_formula):
        """测试直接读取值（非公式单元格）"""
        result = calculator.calc_cell_data(
            test_file_with_formula,
            "Sheet",
            "A1"
        )
        assert result["status"] == "success"
        assert result["data"]["value"] == 10
    
    def test_calc_range_data(self, calculator, test_file_with_formula):
        """测试计算范围数据"""
        result = calculator.calc_range_data(
            test_file_with_formula,
            "Sheet",
            "A1:A3"
        )
        assert result["status"] == "success"
        assert "results" in result["data"]
        assert len(result["data"]["results"]) > 0

