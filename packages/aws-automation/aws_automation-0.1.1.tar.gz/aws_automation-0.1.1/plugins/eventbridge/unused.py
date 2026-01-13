"""
plugins/eventbridge/unused.py - EventBridge 미사용 규칙 분석

비활성화/미사용 규칙 탐지 (CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 미사용 기준: 7일간 트리거 0
UNUSED_DAYS_THRESHOLD = 7

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "events:ListEventBuses",
        "events:ListRules",
        "events:ListTargetsByRule",
        "cloudwatch:GetMetricStatistics",
    ],
}


class RuleStatus(Enum):
    """규칙 상태"""

    NORMAL = "normal"
    DISABLED = "disabled"
    NO_TARGETS = "no_targets"
    UNUSED = "unused"


@dataclass
class RuleInfo:
    """EventBridge 규칙 정보"""

    account_id: str
    account_name: str
    region: str
    rule_name: str
    rule_arn: str
    event_bus_name: str
    state: str
    schedule_expression: str
    event_pattern: str
    target_count: int
    # CloudWatch 지표
    invocations: float = 0.0
    failed_invocations: float = 0.0
    triggered_rules: float = 0.0


@dataclass
class RuleFinding:
    """규칙 분석 결과"""

    rule: RuleInfo
    status: RuleStatus
    recommendation: str


@dataclass
class EventBridgeAnalysisResult:
    """EventBridge 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_rules: int = 0
    disabled_rules: int = 0
    no_targets: int = 0
    unused_rules: int = 0
    normal_rules: int = 0
    findings: list[RuleFinding] = field(default_factory=list)


def collect_rules(session, account_id: str, account_name: str, region: str) -> list[RuleInfo]:
    """EventBridge 규칙 수집"""
    from botocore.exceptions import ClientError

    events = get_client(session, "events", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    rules = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    # 기본 이벤트 버스 + 커스텀 이벤트 버스
    event_buses = ["default"]
    try:
        resp = events.list_event_buses()
        for bus in resp.get("EventBuses", []):
            bus_name = bus.get("Name", "")
            if bus_name and bus_name != "default":
                event_buses.append(bus_name)
    except ClientError:
        pass

    for bus_name in event_buses:
        try:
            paginator = events.get_paginator("list_rules")
            paginate_kwargs = {}
            if bus_name != "default":
                paginate_kwargs["EventBusName"] = bus_name

            for page in paginator.paginate(**paginate_kwargs):
                for rule in page.get("Rules", []):
                    rule_name = rule.get("Name", "")
                    rule_arn = rule.get("Arn", "")

                    # 타겟 수 확인
                    target_count = 0
                    try:
                        targets_kwargs = {"Rule": rule_name}
                        if bus_name != "default":
                            targets_kwargs["EventBusName"] = bus_name
                        targets = events.list_targets_by_rule(**targets_kwargs)
                        target_count = len(targets.get("Targets", []))
                    except ClientError:
                        pass

                    info = RuleInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        rule_name=rule_name,
                        rule_arn=rule_arn,
                        event_bus_name=bus_name,
                        state=rule.get("State", ""),
                        schedule_expression=rule.get("ScheduleExpression", ""),
                        event_pattern=rule.get("EventPattern", "")[:100] if rule.get("EventPattern") else "",
                        target_count=target_count,
                    )

                    # CloudWatch 지표 조회 (활성화된 규칙만)
                    if info.state == "ENABLED":
                        try:
                            # TriggeredRules
                            trig_resp = cloudwatch.get_metric_statistics(
                                Namespace="AWS/Events",
                                MetricName="TriggeredRules",
                                Dimensions=[{"Name": "RuleName", "Value": rule_name}],
                                StartTime=start_time,
                                EndTime=now,
                                Period=86400,
                                Statistics=["Sum"],
                            )
                            if trig_resp.get("Datapoints"):
                                info.triggered_rules = sum(d["Sum"] for d in trig_resp["Datapoints"])

                            # Invocations
                            inv_resp = cloudwatch.get_metric_statistics(
                                Namespace="AWS/Events",
                                MetricName="Invocations",
                                Dimensions=[{"Name": "RuleName", "Value": rule_name}],
                                StartTime=start_time,
                                EndTime=now,
                                Period=86400,
                                Statistics=["Sum"],
                            )
                            if inv_resp.get("Datapoints"):
                                info.invocations = sum(d["Sum"] for d in inv_resp["Datapoints"])

                            # FailedInvocations
                            fail_resp = cloudwatch.get_metric_statistics(
                                Namespace="AWS/Events",
                                MetricName="FailedInvocations",
                                Dimensions=[{"Name": "RuleName", "Value": rule_name}],
                                StartTime=start_time,
                                EndTime=now,
                                Period=86400,
                                Statistics=["Sum"],
                            )
                            if fail_resp.get("Datapoints"):
                                info.failed_invocations = sum(d["Sum"] for d in fail_resp["Datapoints"])

                        except ClientError:
                            pass

                    rules.append(info)

        except ClientError:
            continue

    return rules


def analyze_rules(rules: list[RuleInfo], account_id: str, account_name: str, region: str) -> EventBridgeAnalysisResult:
    """EventBridge 규칙 분석"""
    result = EventBridgeAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_rules=len(rules),
    )

    for rule in rules:
        # 비활성화된 규칙
        if rule.state == "DISABLED":
            result.disabled_rules += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.DISABLED,
                    recommendation="비활성화됨 - 삭제 검토",
                )
            )
            continue

        # 타겟 없음
        if rule.target_count == 0:
            result.no_targets += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.NO_TARGETS,
                    recommendation="타겟 없음 - 삭제 검토",
                )
            )
            continue

        # 미사용 (트리거 없음)
        if rule.triggered_rules == 0 and rule.invocations == 0:
            result.unused_rules += 1
            result.findings.append(
                RuleFinding(
                    rule=rule,
                    status=RuleStatus.UNUSED,
                    recommendation="트리거 없음 - 사용 여부 확인",
                )
            )
            continue

        result.normal_rules += 1
        result.findings.append(
            RuleFinding(
                rule=rule,
                status=RuleStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EventBridgeAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "EventBridge 미사용 규칙 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "비활성화", "타겟없음", "미사용", "정상"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_rules)
        ws.cell(row=row, column=4, value=r.disabled_rules)
        ws.cell(row=row, column=5, value=r.no_targets)
        ws.cell(row=row, column=6, value=r.unused_rules)
        ws.cell(row=row, column=7, value=r.normal_rules)
        if r.disabled_rules > 0:
            ws.cell(row=row, column=4).fill = gray_fill
        if r.no_targets > 0:
            ws.cell(row=row, column=5).fill = red_fill
        if r.unused_rules > 0:
            ws.cell(row=row, column=6).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Rules")
    detail_headers = [
        "Account",
        "Region",
        "Rule Name",
        "Event Bus",
        "State",
        "Schedule",
        "Targets",
        "Triggers",
        "상태",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != RuleStatus.NORMAL:
                detail_row += 1
                rule = f.rule
                ws_detail.cell(row=detail_row, column=1, value=rule.account_name)
                ws_detail.cell(row=detail_row, column=2, value=rule.region)
                ws_detail.cell(row=detail_row, column=3, value=rule.rule_name)
                ws_detail.cell(row=detail_row, column=4, value=rule.event_bus_name)
                ws_detail.cell(row=detail_row, column=5, value=rule.state)
                ws_detail.cell(
                    row=detail_row,
                    column=6,
                    value=rule.schedule_expression or "-",
                )
                ws_detail.cell(row=detail_row, column=7, value=rule.target_count)
                ws_detail.cell(row=detail_row, column=8, value=int(rule.triggered_rules))
                ws_detail.cell(row=detail_row, column=9, value=f.status.value)
                ws_detail.cell(row=detail_row, column=10, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"EventBridge_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EventBridgeAnalysisResult | None:
    """단일 계정/리전의 EventBridge 규칙 수집 및 분석 (병렬 실행용)"""
    rules = collect_rules(session, account_id, account_name, region)
    if not rules:
        return None
    return analyze_rules(rules, account_id, account_name, region)


def run(ctx) -> None:
    """EventBridge 미사용 규칙 분석"""
    console.print("[bold]EventBridge 규칙 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="events")
    results: list[EventBridgeAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_disabled = sum(r.disabled_rules for r in results)
    total_no_targets = sum(r.no_targets for r in results)
    total_unused = sum(r.unused_rules for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"비활성화: {total_disabled}개 / "
        f"타겟없음: [red]{total_no_targets}개[/red] / "
        f"미사용: [yellow]{total_unused}개[/yellow]"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("eventbridge-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
