"""Format adapters for spreadsheet export/import.

Provides adapter interfaces for converting between SpreadsheetDL's
internal representation and various file formats.

**HTML Import:**
    HTML import is now fully supported with BeautifulSoup4 and lxml.
    Features include:
    - Parse HTML tables to SheetSpec
    - Handle <thead>, <tbody>, <tfoot>
    - Handle <th> vs <td> cells
    - Handle colspan/rowspan attributes
    - CSS selector filtering
    - Auto-detect data types
"""

from __future__ import annotations

import re
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, ClassVar

from spreadsheet_dl.builder import SheetSpec


class ExportFormat(Enum):
    """Supported export formats."""

    ODS = "ods"  # OpenDocument Spreadsheet (native)
    XLSX = "xlsx"  # Microsoft Excel (via openpyxl)
    CSV = "csv"  # Comma-Separated Values
    TSV = "tsv"  # Tab-Separated Values
    HTML = "html"  # HTML table
    JSON = "json"  # JSON data
    PDF = "pdf"  # PDF (future)


class ImportFormat(Enum):
    """Supported import formats."""

    ODS = "ods"  # OpenDocument Spreadsheet
    XLSX = "xlsx"  # Microsoft Excel
    CSV = "csv"  # Comma-Separated Values
    TSV = "tsv"  # Tab-Separated Values
    JSON = "json"  # JSON data


# Valid AdapterOptions fields for filtering kwargs
_ADAPTER_OPTIONS_FIELDS = frozenset(
    {
        "include_headers",
        "include_styles",
        "include_formulas",
        "include_charts",
        "encoding",
        "delimiter",
        "quote_char",
        "date_format",
        "decimal_places",
        "sheet_names",
    }
)


def _filter_adapter_options_kwargs(kwargs: dict[str, Any]) -> dict[str, Any]:
    """Filter kwargs to only include valid AdapterOptions fields.

    Args:
        kwargs: Dictionary of keyword arguments

    Returns:
        Filtered dictionary with only valid AdapterOptions fields
    """
    return {k: v for k, v in kwargs.items() if k in _ADAPTER_OPTIONS_FIELDS}


@dataclass
class AdapterOptions:
    """Configuration options for format adapters.

    Attributes:
        include_headers: Include column headers in export
        include_styles: Export style information
        include_formulas: Export formulas (vs computed values)
        include_charts: Export chart definitions
        encoding: Text encoding for CSV/TSV
        delimiter: Field delimiter for CSV/TSV
        quote_char: Quote character for CSV/TSV
        date_format: Date format string
        decimal_places: Number of decimal places for numbers
        sheet_names: Specific sheets to export (None for all)
    """

    include_headers: bool = True
    include_styles: bool = True
    include_formulas: bool = True
    include_charts: bool = True
    encoding: str = "utf-8"
    delimiter: str = ","
    quote_char: str = '"'
    date_format: str = "%Y-%m-%d"
    decimal_places: int = 2
    sheet_names: list[str] | None = None


@dataclass
class HTMLImportOptions(AdapterOptions):
    """Configuration options for HTML import.

    Attributes:
        table_selector: CSS selector for table elements (None for all tables)
        header_row: First row is header (auto-detect from <th> if None)
        skip_empty_rows: Skip rows with all empty cells
        trim_whitespace: Trim leading/trailing whitespace from cells
        detect_types: Auto-detect cell data types (int, float, date)
    """

    table_selector: str | None = None
    header_row: bool | None = None  # None = auto-detect
    skip_empty_rows: bool = True
    trim_whitespace: bool = True
    detect_types: bool = True


class FormatAdapter(ABC):
    """Abstract base class for format adapters.

    Subclasses implement specific format conversions.
    """

    @property
    @abstractmethod
    def format_name(self) -> str:
        """Return the adapter's format name."""
        ...  # pragma: no cover

    @property
    @abstractmethod
    def file_extension(self) -> str:
        """Return the file extension for this format."""
        ...  # pragma: no cover

    @abstractmethod
    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export sheets to file.

        Args:
            sheets: Sheet specifications to export
            output_path: Output file path
            options: Export options

        Returns:
            Path to created file
        """
        ...  # pragma: no cover

    @abstractmethod
    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import sheets from file.

        Args:
            input_path: Input file path
            options: Import options

        Returns:
            List of imported sheet specifications
        """
        ...  # pragma: no cover

    def load(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
        **kwargs: Any,
    ) -> list[SheetSpec]:
        """Load sheets from file (alias for import_file).

        This is an alias for import_file() to provide a simpler API
        for MCP tools and other consumers.

        Args:
            input_path: Input file path
            options: Import options
            **kwargs: Additional options (merged with options)

        Returns:
            List of imported sheet specifications
        """
        if kwargs:
            filtered_kwargs = _filter_adapter_options_kwargs(kwargs)
            if options is None:
                options = AdapterOptions(**filtered_kwargs)
            else:
                # Merge kwargs into a new options object
                opt_dict = {
                    "include_headers": options.include_headers,
                    "include_styles": options.include_styles,
                    "include_formulas": options.include_formulas,
                    "include_charts": options.include_charts,
                    "encoding": options.encoding,
                    "delimiter": options.delimiter,
                    "quote_char": options.quote_char,
                    "date_format": options.date_format,
                    "decimal_places": options.decimal_places,
                    "sheet_names": options.sheet_names,
                }
                opt_dict.update(filtered_kwargs)
                options = AdapterOptions(**opt_dict)  # type: ignore[arg-type]

        return self.import_file(input_path, options)

    def save(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
        **kwargs: Any,
    ) -> Path:
        """Save sheets to file (alias for export).

        This is an alias for export() to provide a simpler API
        for MCP tools and other consumers.

        Args:
            sheets: Sheet specifications to export
            output_path: Output file path
            options: Export options
            **kwargs: Additional options (merged with options)

        Returns:
            Path to created file
        """
        if kwargs:
            filtered_kwargs = _filter_adapter_options_kwargs(kwargs)
            if options is None:
                options = AdapterOptions(**filtered_kwargs)
            else:
                # Merge kwargs into a new options object
                opt_dict = {
                    "include_headers": options.include_headers,
                    "include_styles": options.include_styles,
                    "include_formulas": options.include_formulas,
                    "include_charts": options.include_charts,
                    "encoding": options.encoding,
                    "delimiter": options.delimiter,
                    "quote_char": options.quote_char,
                    "date_format": options.date_format,
                    "decimal_places": options.decimal_places,
                    "sheet_names": options.sheet_names,
                }
                opt_dict.update(filtered_kwargs)
                options = AdapterOptions(**opt_dict)  # type: ignore[arg-type]

        return self.export(sheets, output_path, options)


class OdsAdapter(FormatAdapter):
    """ODS format adapter (native format).

    Uses odfpy for ODS file operations.
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "OpenDocument Spreadsheet"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".ods"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to ODS format."""
        from spreadsheet_dl.renderer import render_sheets

        return render_sheets(sheets, output_path)

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from ODS format."""
        from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec, SheetSpec
        from spreadsheet_dl.streaming import StreamingReader

        sheets = []
        with StreamingReader(input_path) as reader:
            for sheet_name in reader.sheet_names():
                rows = []
                col_count = reader.column_count(sheet_name)
                columns = [ColumnSpec(name=f"Col{i + 1}") for i in range(col_count)]

                for streaming_row in reader.rows(sheet_name):
                    cells = []
                    for cell in streaming_row.cells:
                        cells.append(
                            CellSpec(
                                value=cell.value,
                                value_type=cell.value_type,
                                formula=cell.formula,
                                style=cell.style,
                            )
                        )
                    rows.append(RowSpec(cells=cells, style=streaming_row.style))

                sheets.append(SheetSpec(name=sheet_name, columns=columns, rows=rows))

        return sheets


class CsvAdapter(FormatAdapter):
    """CSV format adapter.

    Handles comma-separated values export/import.
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "Comma-Separated Values"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".csv"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to CSV format."""
        import csv

        options = options or AdapterOptions()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # CSV only supports single sheet - use first sheet or specified
        sheet = sheets[0] if sheets else None
        if options.sheet_names and sheets:
            for s in sheets:
                if s.name in options.sheet_names:
                    sheet = s
                    break

        if sheet is None:
            # Create empty file
            output_path.write_text("")
            return output_path

        with output_path.open("w", encoding=options.encoding, newline="") as f:
            writer = csv.writer(
                f,
                delimiter=options.delimiter,
                quotechar=options.quote_char,
            )

            # Write header row if present
            if options.include_headers and sheet.columns:
                writer.writerow([col.name for col in sheet.columns])

            # Write data rows
            for row in sheet.rows:
                values = []
                for cell in row.cells:
                    value = self._format_value(cell.value, options)
                    values.append(value)
                writer.writerow(values)

        return output_path

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from CSV format."""
        import csv

        from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec, SheetSpec

        options = options or AdapterOptions()
        input_path = Path(input_path)

        rows = []
        columns = []

        with input_path.open("r", encoding=options.encoding) as f:
            reader = csv.reader(
                f,
                delimiter=options.delimiter,
                quotechar=options.quote_char,
            )

            for row_idx, csv_row in enumerate(reader):
                if row_idx == 0 and options.include_headers:
                    # First row is headers
                    columns = [ColumnSpec(name=col) for col in csv_row]
                else:
                    cells = [CellSpec(value=val) for val in csv_row]
                    rows.append(RowSpec(cells=cells))

        # If no headers, create generic columns
        if not columns and rows:
            max_cols = max(len(row.cells) for row in rows)
            columns = [ColumnSpec(name=f"Column{i + 1}") for i in range(max_cols)]

        sheet_name = input_path.stem
        return [SheetSpec(name=sheet_name, columns=columns, rows=rows)]

    def _format_value(self, value: Any, options: AdapterOptions) -> str:
        """Format a cell value for CSV export."""
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime(options.date_format)
        if isinstance(value, (float, Decimal)):
            return f"{value:.{options.decimal_places}f}"
        return str(value)


class TsvAdapter(CsvAdapter):
    """TSV format adapter.

    Handles tab-separated values export/import.
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "Tab-Separated Values"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".tsv"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to TSV format."""
        options = options or AdapterOptions()
        options.delimiter = "\t"
        return super().export(sheets, output_path, options)

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from TSV format."""
        options = options or AdapterOptions()
        options.delimiter = "\t"
        return super().import_file(input_path, options)


class JsonAdapter(FormatAdapter):
    """JSON format adapter.

    Exports spreadsheet data as JSON for programmatic access.
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "JSON Data"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".json"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to JSON format."""
        from spreadsheet_dl.serialization import Serializer

        serializer = Serializer()
        return serializer.save_json(sheets, output_path)

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from JSON format."""
        from spreadsheet_dl.serialization import Serializer

        serializer = Serializer()
        data = serializer.load_json(input_path)

        if isinstance(data, list):
            return data
        return [data] if data else []


class HtmlAdapter(FormatAdapter):
    """HTML format adapter.

    Exports spreadsheet as HTML table(s).
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "HTML Table"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".html"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to HTML format."""
        options = options or AdapterOptions()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        html_parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            '<meta charset="utf-8">',
            "<title>Spreadsheet Export</title>",
            "<style>",
            "table { border-collapse: collapse; margin: 20px 0; }",
            "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "th { background-color: #4472C4; color: white; }",
            "tr:nth-child(even) { background-color: #f2f2f2; }",
            "</style>",
            "</head>",
            "<body>",
        ]

        for sheet in sheets:
            if options.sheet_names and sheet.name not in options.sheet_names:
                continue

            html_parts.append(f"<h2>{self._escape_html(sheet.name)}</h2>")
            html_parts.append("<table>")

            # Header row
            if options.include_headers and sheet.columns:
                html_parts.append("<thead><tr>")
                for col in sheet.columns:
                    html_parts.append(f"<th>{self._escape_html(col.name)}</th>")
                html_parts.append("</tr></thead>")

            # Data rows
            html_parts.append("<tbody>")
            for row in sheet.rows:
                html_parts.append("<tr>")
                for cell in row.cells:
                    value = self._format_value(cell.value, options)
                    html_parts.append(f"<td>{self._escape_html(value)}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody>")

            html_parts.append("</table>")

        html_parts.extend(["</body>", "</html>"])

        output_path.write_text("\n".join(html_parts), encoding="utf-8")
        return output_path

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from HTML format.

        Parses HTML tables using BeautifulSoup4 and converts them to SheetSpec.
        Handles thead/tbody/tfoot, th/td cells, colspan/rowspan attributes.

        Args:
            input_path: Path to HTML file
            options: Import options (HTMLImportOptions recommended)

        Returns:
            List of SheetSpec, one per table found

        Raises:
            ImportError: If beautifulsoup4 or lxml are not installed
            ValueError: If HTML file is invalid or contains no tables
        """
        try:
            from bs4 import BeautifulSoup
        except ImportError as e:
            raise ImportError(
                "HTML import requires beautifulsoup4 and lxml. "
                "Install with: pip install 'spreadsheet-dl[html]'"
            ) from e

        # Use HTMLImportOptions if provided, otherwise default AdapterOptions
        if options is None:
            options = HTMLImportOptions()
        elif not isinstance(options, HTMLImportOptions):
            # Convert AdapterOptions to HTMLImportOptions with defaults
            html_opts = HTMLImportOptions()
            # Copy common fields
            for field_name in [
                "include_headers",
                "encoding",
                "date_format",
                "sheet_names",
            ]:
                if hasattr(options, field_name):
                    setattr(html_opts, field_name, getattr(options, field_name))
            options = html_opts

        # Read and parse HTML
        html_content = Path(input_path).read_text(encoding=options.encoding)
        soup = BeautifulSoup(html_content, "lxml")

        # Find tables
        if options.table_selector:
            tables = soup.select(options.table_selector)
        else:
            tables = soup.find_all("table")

        if not tables:
            raise ValueError(
                f"No HTML tables found in {input_path}. "
                f"Selector: {options.table_selector or 'table'}"
            )

        sheets = []
        for idx, table in enumerate(tables):
            sheet_name = self._extract_table_name(table, idx)

            # Skip if not in sheet_names filter
            if options.sheet_names and sheet_name not in options.sheet_names:
                continue

            sheet = self._parse_table(table, sheet_name, options)
            if sheet:
                sheets.append(sheet)

        return sheets

    def _extract_table_name(self, table: Any, index: int) -> str:
        """Extract table name from heading or generate default.

        Args:
            table: BeautifulSoup table element
            index: Table index in document

        Returns:
            Sheet name
        """
        # Look for preceding h1-h6 heading
        prev = table.find_previous(["h1", "h2", "h3", "h4", "h5", "h6"])
        if prev and prev.get_text(strip=True):
            name = prev.get_text(strip=True)
            # Sanitize sheet name
            name = re.sub(r"[^\w\s-]", "", name)[:31]  # ODS 31 char limit
            return name or f"Table_{index + 1}"

        # Check for caption element
        caption = table.find("caption")
        if caption and caption.get_text(strip=True):
            name = caption.get_text(strip=True)
            name = re.sub(r"[^\w\s-]", "", name)[:31]
            return name or f"Table_{index + 1}"

        return f"Table_{index + 1}"

    def _parse_table(
        self, table: Any, sheet_name: str, options: HTMLImportOptions
    ) -> SheetSpec | None:
        """Parse HTML table to SheetSpec.

        Args:
            table: BeautifulSoup table element
            sheet_name: Name for the sheet
            options: Import options

        Returns:
            SheetSpec or None if table is empty
        """
        from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec

        # Extract all rows (from thead, tbody, tfoot)
        all_rows = []
        for section in ["thead", "tbody", "tfoot"]:
            section_elem = table.find(section)
            if section_elem:
                all_rows.extend(section_elem.find_all("tr"))

        # If no sections, get rows directly from table
        if not all_rows:
            all_rows = table.find_all("tr", recursive=False)

        if not all_rows:
            return None

        # Parse table to 2D grid, handling colspan/rowspan
        grid = self._build_cell_grid(all_rows, options)

        if not grid:
            return None

        # Detect headers
        has_header = self._detect_header_row(all_rows, options)

        # Build columns
        columns = []
        if has_header and grid:
            header_row = grid[0]
            for cell_value in header_row:
                col_name = str(cell_value) if cell_value else ""
                columns.append(ColumnSpec(name=col_name))
            # Remove header from grid
            grid = grid[1:]
        else:
            # Generate column names
            max_cols = max(len(row) for row in grid) if grid else 0
            for i in range(max_cols):
                columns.append(ColumnSpec(name=f"Column_{i + 1}"))

        # Build rows
        rows = []
        for row_data in grid:
            # Skip empty rows if requested
            if options.skip_empty_rows and all(
                cell is None or str(cell).strip() == "" for cell in row_data
            ):
                continue

            cells = []
            for cell_value in row_data:
                # Type detection
                typed_value = (
                    self._detect_type(cell_value, options)
                    if options.detect_types
                    else cell_value
                )
                cells.append(CellSpec(value=typed_value))

            if cells:
                rows.append(RowSpec(cells=cells))

        if not rows:
            return None

        return SheetSpec(name=sheet_name, columns=columns, rows=rows)

    def _build_cell_grid(
        self, rows: list[Any], options: HTMLImportOptions
    ) -> list[list[Any]]:
        """Build 2D cell grid, handling colspan/rowspan.

        Args:
            rows: List of BeautifulSoup <tr> elements
            options: Import options

        Returns:
            2D list of cell values
        """
        # First pass: determine grid dimensions
        max_cols = 0
        rowspan_tracker: dict[tuple[int, int], int] = {}  # (row, col) -> remaining span

        grid: list[list[Any]] = []

        for row_idx, row in enumerate(rows):
            cells = row.find_all(["td", "th"])
            grid.append([])

            col_idx = 0
            for cell in cells:
                # Skip columns occupied by previous rowspans
                while (row_idx, col_idx) in rowspan_tracker:
                    grid[row_idx].append(None)  # Placeholder for spanned cell
                    col_idx += 1

                # Get cell value
                cell_text = cell.get_text(strip=options.trim_whitespace)

                # Get colspan/rowspan
                colspan = int(cell.get("colspan", 1))
                rowspan = int(cell.get("rowspan", 1))

                # Add cell and colspan placeholders
                grid[row_idx].append(cell_text)
                for _ in range(colspan - 1):
                    col_idx += 1
                    grid[row_idx].append(None)

                # Track rowspan
                if rowspan > 1:
                    for r in range(1, rowspan):
                        for c in range(colspan):
                            rowspan_tracker[
                                (row_idx + r, col_idx - colspan + 1 + c)
                            ] = rowspan - r

                col_idx += 1

            max_cols = max(max_cols, len(grid[row_idx]))

        # Normalize row lengths
        for row in grid:
            while len(row) < max_cols:
                row.append(None)

        return grid

    def _detect_header_row(self, rows: list[Any], options: HTMLImportOptions) -> bool:
        """Detect if first row should be treated as header.

        Args:
            rows: List of BeautifulSoup <tr> elements
            options: Import options

        Returns:
            True if first row is header
        """
        if options.header_row is not None:
            return options.header_row

        # Auto-detect: check if first row is in <thead> or uses <th>
        if not rows:
            return False

        first_row = rows[0]

        # Check if in thead
        if first_row.find_parent("thead"):
            return True

        # Check if uses <th> cells
        th_cells = first_row.find_all("th")
        td_cells = first_row.find_all("td")

        # If row has more <th> than <td>, treat as header
        return len(th_cells) > len(td_cells)

    def _detect_type(self, value: Any, options: HTMLImportOptions) -> Any:
        """Detect and convert cell value type.

        Args:
            value: Raw cell value
            options: Import options

        Returns:
            Typed value (int, float, date, or str)
        """
        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        # Try integer
        try:
            if text.isdigit() or (text.startswith("-") and text[1:].isdigit()):
                return int(text)
        except ValueError:
            pass

        # Try float
        try:
            if "." in text or "e" in text.lower():
                return float(text)
        except ValueError:
            pass

        # Try date (common formats)
        date_patterns = [
            (r"^\d{4}-\d{2}-\d{2}$", "%Y-%m-%d"),  # ISO format
            (r"^\d{2}/\d{2}/\d{4}$", "%m/%d/%Y"),  # US format
            (r"^\d{2}-\d{2}-\d{4}$", "%m-%d-%Y"),  # US format with dashes
            (r"^\d{4}/\d{2}/\d{2}$", "%Y/%m/%d"),  # ISO with slashes
        ]

        for pattern, fmt in date_patterns:
            if re.match(pattern, text):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    pass

        # Return as string
        return text

    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        return (
            str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    def _format_value(self, value: Any, options: AdapterOptions) -> str:
        """Format a cell value for HTML export."""
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime(options.date_format)
        if isinstance(value, (float, Decimal)):
            return f"{value:.{options.decimal_places}f}"
        return str(value)


class XlsxAdapter(FormatAdapter):
    """XLSX format adapter.

    Handles Microsoft Excel (.xlsx) export/import using openpyxl.
    Requires openpyxl to be installed: pip install 'spreadsheet-dl[xlsx]'
    """

    @property
    def format_name(self) -> str:
        """Return format name."""
        return "Microsoft Excel"

    @property
    def file_extension(self) -> str:
        """Return file extension."""
        return ".xlsx"

    def export(
        self,
        sheets: list[SheetSpec],
        output_path: Path,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export to XLSX format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            raise ImportError(
                "XLSX export requires openpyxl. "
                "Install with: pip install 'spreadsheet-dl[xlsx]'"
            ) from e

        options = options or AdapterOptions()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        for sheet_spec in sheets:
            if options.sheet_names and sheet_spec.name not in options.sheet_names:
                continue

            ws = wb.create_sheet(title=sheet_spec.name[:31])  # XLSX 31 char limit

            row_offset = 1

            # Write header row if present
            if options.include_headers and sheet_spec.columns:
                for col_idx, col in enumerate(sheet_spec.columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col.name)
                    # Style header cells
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="4472C4", end_color="4472C4", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center")
                row_offset = 2

            # Write data rows
            for row_idx, row_spec in enumerate(sheet_spec.rows, start=row_offset):
                for col_idx, cell_spec in enumerate(row_spec.cells, start=1):
                    value = cell_spec.value
                    # Handle formula export
                    if options.include_formulas and cell_spec.formula:
                        value = cell_spec.formula
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-size columns based on content
            for col_idx, col in enumerate(sheet_spec.columns, start=1):
                max_length = len(col.name) if col.name else 10
                for row_idx in range(1, len(sheet_spec.rows) + row_offset):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                # Set column width (max 50 chars)
                from openpyxl.utils import get_column_letter

                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max_length + 2, 50
                )

        # Ensure at least one sheet exists
        if not wb.sheetnames:
            wb.create_sheet("Sheet1")

        wb.save(output_path)
        return output_path

    def import_file(
        self,
        input_path: Path,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import from XLSX format."""
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "XLSX import requires openpyxl. "
                "Install with: pip install 'spreadsheet-dl[xlsx]'"
            ) from e

        from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec

        options = options or AdapterOptions()
        input_path = Path(input_path)

        wb = load_workbook(input_path, data_only=not options.include_formulas)
        sheets = []

        for sheet_name in wb.sheetnames:
            if options.sheet_names and sheet_name not in options.sheet_names:
                continue

            ws = wb[sheet_name]
            rows = []
            columns = []

            # Determine dimensions
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            if max_row == 0 or max_col == 0:
                # Empty sheet
                sheets.append(SheetSpec(name=sheet_name, columns=[], rows=[]))
                continue

            # Extract header row if present
            start_row = 1
            if options.include_headers and max_row > 0:
                for col_idx in range(1, max_col + 1):
                    cell_value = ws.cell(row=1, column=col_idx).value
                    col_name = (
                        str(cell_value) if cell_value is not None else f"Col{col_idx}"
                    )
                    columns.append(ColumnSpec(name=col_name))
                start_row = 2
            else:
                # Generate generic columns
                for col_idx in range(1, max_col + 1):
                    columns.append(ColumnSpec(name=f"Column{col_idx}"))

            # Extract data rows
            for row_idx in range(start_row, max_row + 1):
                cells = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    formula = None

                    # Get formula if requested and available
                    if options.include_formulas and hasattr(cell, "value"):
                        # In data_only=False mode, formulas are stored
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if isinstance(cell_value, str) and cell_value.startswith("="):
                            formula = cell_value
                            value = None  # Formula takes precedence

                    cells.append(CellSpec(value=value, formula=formula))

                rows.append(RowSpec(cells=cells))

            sheets.append(SheetSpec(name=sheet_name, columns=columns, rows=rows))

        wb.close()
        return sheets


class AdapterRegistry:
    """Registry of available format adapters.

    Provides discovery and instantiation of format adapters.

    Examples:
        # Get adapter by format
        adapter = AdapterRegistry.get_adapter(ExportFormat.CSV)
        adapter.export(sheets, "output.csv")

        # List available formats
        formats = AdapterRegistry.list_formats()

        # Export to any format
        AdapterRegistry.export(sheets, "output.xlsx", ExportFormat.XLSX)
    """

    _adapters: ClassVar[dict[ExportFormat, type[FormatAdapter]]] = {
        ExportFormat.ODS: OdsAdapter,
        ExportFormat.XLSX: XlsxAdapter,
        ExportFormat.CSV: CsvAdapter,
        ExportFormat.TSV: TsvAdapter,
        ExportFormat.JSON: JsonAdapter,
        ExportFormat.HTML: HtmlAdapter,
    }

    @classmethod
    def get_adapter(cls, format: ExportFormat | str) -> FormatAdapter:
        """Get adapter instance for format.

        Args:
            format: Export format (enum or string like 'csv', 'xlsx')

        Returns:
            FormatAdapter instance

        Raises:
            ValueError: If format not supported
        """
        # Convert string to enum if needed
        if isinstance(format, str):
            try:
                format = ExportFormat(format.lower())
            except ValueError:
                raise ValueError(f"Unsupported format: {format}") from None

        adapter_class = cls._adapters.get(format)
        if adapter_class is None:
            raise ValueError(f"Unsupported format: {format}")
        return adapter_class()

    @classmethod
    def register_adapter(
        cls,
        format: ExportFormat,
        adapter_class: type[FormatAdapter],
    ) -> None:
        """Register a new adapter.

        Args:
            format: Format to register
            adapter_class: Adapter class
        """
        cls._adapters[format] = adapter_class

    @classmethod
    def list_formats(cls) -> list[ExportFormat]:
        """List available export formats."""
        return list(cls._adapters.keys())

    @classmethod
    def export(
        cls,
        sheets: list[SheetSpec],
        output_path: Path | str,
        format: ExportFormat | None = None,
        options: AdapterOptions | None = None,
    ) -> Path:
        """Export sheets to file.

        Args:
            sheets: Sheet specifications
            output_path: Output file path
            format: Export format (auto-detect from extension if None)
            options: Export options

        Returns:
            Path to created file
        """
        path = Path(output_path)

        if format is None:
            # Auto-detect from extension
            ext = path.suffix.lower()
            format_map = {
                ".ods": ExportFormat.ODS,
                ".xlsx": ExportFormat.XLSX,
                ".csv": ExportFormat.CSV,
                ".tsv": ExportFormat.TSV,
                ".json": ExportFormat.JSON,
                ".html": ExportFormat.HTML,
                ".htm": ExportFormat.HTML,
            }
            format = format_map.get(ext, ExportFormat.ODS)

        adapter = cls.get_adapter(format)
        return adapter.export(sheets, path, options)

    @classmethod
    def import_file(
        cls,
        input_path: Path | str,
        format: ImportFormat | None = None,
        options: AdapterOptions | None = None,
    ) -> list[SheetSpec]:
        """Import sheets from file.

        Args:
            input_path: Input file path
            format: Import format (auto-detect from extension if None)
            options: Import options

        Returns:
            List of imported sheet specifications
        """
        path = Path(input_path)

        if format is None:
            # Auto-detect from extension
            ext = path.suffix.lower()
            format_map = {
                ".ods": ExportFormat.ODS,
                ".xlsx": ExportFormat.XLSX,
                ".csv": ExportFormat.CSV,
                ".tsv": ExportFormat.TSV,
                ".json": ExportFormat.JSON,
            }
            export_format = format_map.get(ext, ExportFormat.ODS)
        else:
            export_format = ExportFormat(format.value)

        adapter = cls.get_adapter(export_format)
        return adapter.import_file(path, options)


# Convenience functions


def export_to(
    sheets: list[SheetSpec],
    output_path: Path | str,
    format: ExportFormat | str | None = None,
    **kwargs: Any,
) -> Path:
    """Export sheets to file.

    Convenience function for AdapterRegistry.export().

    Args:
        sheets: Sheet specifications
        output_path: Output file path
        format: Export format (string or enum)
        **kwargs: Additional options for AdapterOptions

    Returns:
        Path to created file
    """
    if isinstance(format, str):
        format = ExportFormat(format)

    filtered_kwargs = _filter_adapter_options_kwargs(kwargs)
    options = AdapterOptions(**filtered_kwargs) if filtered_kwargs else None
    return AdapterRegistry.export(sheets, output_path, format, options)


def import_from(
    input_path: Path | str,
    format: ImportFormat | str | None = None,
    **kwargs: Any,
) -> list[SheetSpec]:
    """Import sheets from file.

    Convenience function for AdapterRegistry.import_file().

    Args:
        input_path: Input file path
        format: Import format (string or enum)
        **kwargs: Additional options for AdapterOptions

    Returns:
        List of imported sheet specifications
    """
    if isinstance(format, str):
        format = ImportFormat(format)

    filtered_kwargs = _filter_adapter_options_kwargs(kwargs)
    options = AdapterOptions(**filtered_kwargs) if filtered_kwargs else None
    return AdapterRegistry.import_file(input_path, format, options)
