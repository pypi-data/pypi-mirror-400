"""
plugins/acm/unused.py - ACM 미사용 인증서 분석

미사용/만료 임박 인증서 탐지

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "acm:ListCertificates",
        "acm:DescribeCertificate",
    ],
}

# 만료 임박 기준: 30일 이내
EXPIRING_DAYS_THRESHOLD = 30


class CertStatus(Enum):
    """인증서 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    EXPIRING = "expiring"
    EXPIRED = "expired"
    PENDING = "pending"


@dataclass
class CertInfo:
    """ACM 인증서 정보"""

    account_id: str
    account_name: str
    region: str
    certificate_arn: str
    domain_name: str
    status: str
    cert_type: str  # AMAZON_ISSUED, IMPORTED
    key_algorithm: str
    in_use_by: list[str]
    not_before: datetime | None
    not_after: datetime | None
    renewal_eligibility: str

    @property
    def is_in_use(self) -> bool:
        return len(self.in_use_by) > 0

    @property
    def days_until_expiry(self) -> int | None:
        if self.not_after:
            now = datetime.now(timezone.utc)
            delta = self.not_after - now
            return delta.days
        return None


@dataclass
class CertFinding:
    """인증서 분석 결과"""

    cert: CertInfo
    status: CertStatus
    recommendation: str


@dataclass
class ACMAnalysisResult:
    """ACM 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_certs: int = 0
    unused_certs: int = 0
    expiring_certs: int = 0
    expired_certs: int = 0
    pending_certs: int = 0
    normal_certs: int = 0
    findings: list[CertFinding] = field(default_factory=list)


def collect_certificates(session, account_id: str, account_name: str, region: str) -> list[CertInfo]:
    """ACM 인증서 수집"""
    from botocore.exceptions import ClientError

    acm = get_client(session, "acm", region_name=region)
    certs = []

    try:
        paginator = acm.get_paginator("list_certificates")
        for page in paginator.paginate(
            Includes={
                "keyTypes": [
                    "RSA_1024",
                    "RSA_2048",
                    "RSA_3072",
                    "RSA_4096",
                    "EC_prime256v1",
                    "EC_secp384r1",
                    "EC_secp521r1",
                ]
            }
        ):
            for cert_summary in page.get("CertificateSummaryList", []):
                cert_arn = cert_summary.get("CertificateArn", "")

                try:
                    cert_detail = acm.describe_certificate(CertificateArn=cert_arn).get("Certificate", {})

                    info = CertInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        certificate_arn=cert_arn,
                        domain_name=cert_detail.get("DomainName", ""),
                        status=cert_detail.get("Status", ""),
                        cert_type=cert_detail.get("Type", ""),
                        key_algorithm=cert_detail.get("KeyAlgorithm", ""),
                        in_use_by=cert_detail.get("InUseBy", []),
                        not_before=cert_detail.get("NotBefore"),
                        not_after=cert_detail.get("NotAfter"),
                        renewal_eligibility=cert_detail.get("RenewalEligibility", ""),
                    )
                    certs.append(info)

                except ClientError:
                    continue

    except ClientError:
        pass

    return certs


def analyze_certificates(certs: list[CertInfo], account_id: str, account_name: str, region: str) -> ACMAnalysisResult:
    """ACM 인증서 분석"""
    result = ACMAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_certs=len(certs),
    )

    now = datetime.now(timezone.utc)

    for cert in certs:
        # 발급 대기 중
        if cert.status == "PENDING_VALIDATION":
            result.pending_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.PENDING,
                    recommendation="검증 대기 중 - 오래된 경우 삭제 검토",
                )
            )
            continue

        # 만료됨
        if cert.not_after and cert.not_after < now:
            result.expired_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.EXPIRED,
                    recommendation="만료됨 - 삭제 검토",
                )
            )
            continue

        # 만료 임박
        days_left = cert.days_until_expiry
        if days_left is not None and days_left <= EXPIRING_DAYS_THRESHOLD:
            result.expiring_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.EXPIRING,
                    recommendation=f"만료 임박 ({days_left}일 남음) - 갱신 필요",
                )
            )
            continue

        # 미사용
        if not cert.is_in_use:
            result.unused_certs += 1
            result.findings.append(
                CertFinding(
                    cert=cert,
                    status=CertStatus.UNUSED,
                    recommendation="미사용 - 삭제 검토",
                )
            )
            continue

        result.normal_certs += 1
        result.findings.append(
            CertFinding(
                cert=cert,
                status=CertStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[ACMAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "ACM 인증서 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "미사용", "만료임박", "만료됨", "대기중", "정상"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_certs)
        ws.cell(row=row, column=4, value=r.unused_certs)
        ws.cell(row=row, column=5, value=r.expiring_certs)
        ws.cell(row=row, column=6, value=r.expired_certs)
        ws.cell(row=row, column=7, value=r.pending_certs)
        ws.cell(row=row, column=8, value=r.normal_certs)
        if r.unused_certs > 0:
            ws.cell(row=row, column=4).fill = yellow_fill
        if r.expiring_certs > 0:
            ws.cell(row=row, column=5).fill = orange_fill
        if r.expired_certs > 0:
            ws.cell(row=row, column=6).fill = red_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Certificates")
    detail_headers = [
        "Account",
        "Region",
        "Domain",
        "Type",
        "Status",
        "Expiry",
        "Days Left",
        "In Use",
        "분석상태",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != CertStatus.NORMAL:
                detail_row += 1
                c = f.cert
                ws_detail.cell(row=detail_row, column=1, value=c.account_name)
                ws_detail.cell(row=detail_row, column=2, value=c.region)
                ws_detail.cell(row=detail_row, column=3, value=c.domain_name)
                ws_detail.cell(row=detail_row, column=4, value=c.cert_type)
                ws_detail.cell(row=detail_row, column=5, value=c.status)
                ws_detail.cell(
                    row=detail_row,
                    column=6,
                    value=c.not_after.strftime("%Y-%m-%d") if c.not_after else "-",
                )
                ws_detail.cell(
                    row=detail_row,
                    column=7,
                    value=c.days_until_expiry if c.days_until_expiry else "-",
                )
                ws_detail.cell(row=detail_row, column=8, value=len(c.in_use_by))
                ws_detail.cell(row=detail_row, column=9, value=f.status.value)
                ws_detail.cell(row=detail_row, column=10, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"ACM_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ACMAnalysisResult | None:
    """단일 계정/리전의 ACM 인증서 수집 및 분석 (병렬 실행용)"""
    certs = collect_certificates(session, account_id, account_name, region)
    if not certs:
        return None
    return analyze_certificates(certs, account_id, account_name, region)


def run(ctx) -> None:
    """ACM 미사용 인증서 분석"""
    console.print("[bold]ACM 인증서 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="acm")
    results: list[ACMAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_certs for r in results)
    total_expiring = sum(r.expiring_certs for r in results)
    total_expired = sum(r.expired_certs for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [yellow]{total_unused}개[/yellow] / "
        f"만료임박: [orange1]{total_expiring}개[/orange1] / "
        f"만료: [red]{total_expired}개[/red]"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("acm-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
