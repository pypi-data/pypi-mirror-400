"""Comprehensive tests for XLSX chart rendering.

Tests:
    - All chart types (column, bar, line, area, pie, scatter, combo)
    - Chart customization options
    - Sparklines
    - Chart positioning and sizing
    - Multi-series charts
    - Chart titles and legends
"""

from __future__ import annotations

from pathlib import Path

import pytest

from spreadsheet_dl.builder import CellSpec, ColumnSpec, RowSpec, SheetSpec

pytestmark = [pytest.mark.unit, pytest.mark.rendering, pytest.mark.charts]


# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def chart_data_sheet() -> SheetSpec:
    """Create a sheet with data suitable for charting."""
    return SheetSpec(
        name="ChartData",
        columns=[
            ColumnSpec(name="Month"),
            ColumnSpec(name="Sales"),
            ColumnSpec(name="Expenses"),
            ColumnSpec(name="Profit"),
        ],
        rows=[
            RowSpec(
                cells=[
                    CellSpec(value="Month"),
                    CellSpec(value="Sales"),
                    CellSpec(value="Expenses"),
                    CellSpec(value="Profit"),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="Jan"),
                    CellSpec(value=10000),
                    CellSpec(value=7500),
                    CellSpec(value=2500),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="Feb"),
                    CellSpec(value=12000),
                    CellSpec(value=8000),
                    CellSpec(value=4000),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="Mar"),
                    CellSpec(value=15000),
                    CellSpec(value=9000),
                    CellSpec(value=6000),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="Apr"),
                    CellSpec(value=14000),
                    CellSpec(value=8500),
                    CellSpec(value=5500),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="May"),
                    CellSpec(value=16000),
                    CellSpec(value=9500),
                    CellSpec(value=6500),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value="Jun"),
                    CellSpec(value=18000),
                    CellSpec(value=10000),
                    CellSpec(value=8000),
                ]
            ),
        ],
    )


@pytest.fixture
def scatter_data_sheet() -> SheetSpec:
    """Create a sheet with x-y data for scatter plots."""
    return SheetSpec(
        name="ScatterData",
        columns=[
            ColumnSpec(name="X"),
            ColumnSpec(name="Y1"),
            ColumnSpec(name="Y2"),
        ],
        rows=[
            RowSpec(
                cells=[
                    CellSpec(value="X"),
                    CellSpec(value="Y1"),
                    CellSpec(value="Y2"),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value=1),
                    CellSpec(value=2),
                    CellSpec(value=5),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value=2),
                    CellSpec(value=4),
                    CellSpec(value=8),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value=3),
                    CellSpec(value=5),
                    CellSpec(value=12),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value=4),
                    CellSpec(value=7),
                    CellSpec(value=15),
                ]
            ),
            RowSpec(
                cells=[
                    CellSpec(value=5),
                    CellSpec(value=8),
                    CellSpec(value=20),
                ]
            ),
        ],
    )


@pytest.fixture
def pie_data_sheet() -> SheetSpec:
    """Create a sheet with data for pie charts."""
    return SheetSpec(
        name="PieData",
        columns=[
            ColumnSpec(name="Category"),
            ColumnSpec(name="Value"),
        ],
        rows=[
            RowSpec(cells=[CellSpec(value="Category"), CellSpec(value="Value")]),
            RowSpec(cells=[CellSpec(value="Product A"), CellSpec(value=35)]),
            RowSpec(cells=[CellSpec(value="Product B"), CellSpec(value=25)]),
            RowSpec(cells=[CellSpec(value="Product C"), CellSpec(value=20)]),
            RowSpec(cells=[CellSpec(value="Product D"), CellSpec(value=15)]),
            RowSpec(cells=[CellSpec(value="Other"), CellSpec(value=5)]),
        ],
    )


# =============================================================================
# Column Chart Tests
# =============================================================================


class TestXlsxColumnCharts:
    """Test column chart rendering."""

    def test_simple_column_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a simple column chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.COLUMN,
            title="Monthly Sales",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
        )

        output_path = tmp_path / "column_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_stacked_column_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a stacked column chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.COLUMN_STACKED,
            title="Sales vs Expenses",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7"],
            ),
            position="E2",
        )

        output_path = tmp_path / "stacked_column_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_100_percent_stacked_column(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a 100% stacked column chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.COLUMN_PERCENT_STACKED,
            title="Category Distribution",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7"],
            ),
            position="E2",
        )

        output_path = tmp_path / "percent_stacked_column_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Bar Chart Tests
# =============================================================================


class TestXlsxBarCharts:
    """Test bar chart rendering."""

    def test_simple_bar_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a simple bar chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.BAR,
            title="Monthly Profit",
            data=DataRange(
                categories="A2:A7",
                values="D2:D7",
            ),
            position="E2",
        )

        output_path = tmp_path / "bar_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_clustered_bar_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a clustered bar chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.BAR_CLUSTERED,
            title="Sales and Expenses Comparison",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7"],
            ),
            position="E2",
        )

        output_path = tmp_path / "clustered_bar_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Line Chart Tests
# =============================================================================


class TestXlsxLineCharts:
    """Test line chart rendering."""

    def test_simple_line_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a simple line chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE,
            title="Sales Trend",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
        )

        output_path = tmp_path / "line_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_multi_series_line_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a line chart with multiple series."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE,
            title="Financial Trends",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7", "D2:D7"],
            ),
            position="E2",
            legend=True,
        )

        output_path = tmp_path / "multi_line_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_line_with_markers(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a line chart with markers."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE_MARKERS,
            title="Sales with Data Points",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
        )

        output_path = tmp_path / "line_markers_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Area Chart Tests
# =============================================================================


class TestXlsxAreaCharts:
    """Test area chart rendering."""

    def test_simple_area_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a simple area chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.AREA,
            title="Sales Area",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
        )

        output_path = tmp_path / "area_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_stacked_area_chart(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a stacked area chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.AREA_STACKED,
            title="Revenue Components",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7"],
            ),
            position="E2",
        )

        output_path = tmp_path / "stacked_area_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Pie Chart Tests
# =============================================================================


class TestXlsxPieCharts:
    """Test pie chart rendering."""

    def test_simple_pie_chart(self, pie_data_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test rendering a simple pie chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.PIE,
            title="Product Distribution",
            data=DataRange(
                categories="A2:A6",
                values="B2:B6",
            ),
            position="D2",
        )

        output_path = tmp_path / "pie_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([pie_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_pie_with_explosion(
        self, pie_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a pie chart with exploded slice."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.PIE_EXPLODED,
            title="Products (Exploded)",
            data=DataRange(
                categories="A2:A6",
                values="B2:B6",
            ),
            position="D2",
        )

        output_path = tmp_path / "exploded_pie_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([pie_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_doughnut_chart(self, pie_data_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test rendering a doughnut chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.DOUGHNUT,
            title="Product Market Share",
            data=DataRange(
                categories="A2:A6",
                values="B2:B6",
            ),
            position="D2",
        )

        output_path = tmp_path / "doughnut_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([pie_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Scatter Chart Tests
# =============================================================================


class TestXlsxScatterCharts:
    """Test scatter chart rendering."""

    def test_simple_scatter_chart(
        self, scatter_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a simple scatter chart."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.SCATTER,
            title="X-Y Relationship",
            data=DataRange(
                categories="A2:A6",
                values="B2:B6",
            ),
            position="E2",
        )

        output_path = tmp_path / "scatter_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([scatter_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_scatter_with_lines(
        self, scatter_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a scatter chart with connecting lines."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.SCATTER_LINES,
            title="X-Y with Lines",
            data=DataRange(
                categories="A2:A6",
                values="B2:B6",
            ),
            position="E2",
        )

        output_path = tmp_path / "scatter_lines_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([scatter_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Combo Chart Tests
# =============================================================================


class TestXlsxComboCharts:
    """Test combination chart rendering."""

    def test_column_line_combo(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a combo chart (column + line)."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange, SeriesSpec
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.COMBO,
            title="Sales vs Trend",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            series=[
                SeriesSpec(
                    name="Sales",
                    values="B2:B7",
                    chart_type=ChartType.COLUMN,
                ),
                SeriesSpec(
                    name="Profit",
                    values="D2:D7",
                    chart_type=ChartType.LINE,
                ),
            ],
            position="E2",
            legend=True,
        )

        output_path = tmp_path / "combo_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Sparkline Tests
# =============================================================================


class TestXlsxSparklines:
    """Test sparkline rendering."""

    def test_line_sparkline(self, chart_data_sheet: SheetSpec, tmp_path: Path) -> None:
        """Test rendering a line sparkline."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import SparklineSpec, SparklineType
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sparkline = SparklineSpec(
            type=SparklineType.LINE,
            data_range="B2:B7",
            location="F1",
        )

        output_path = tmp_path / "line_sparkline.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, sparklines=[sparkline])

        wb = load_workbook(output_path)
        # Sparklines are stored differently in openpyxl
        # Just verify the file was created successfully
        assert output_path.exists()
        wb.close()

    def test_column_sparkline(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a column sparkline."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import SparklineSpec, SparklineType
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sparkline = SparklineSpec(
            type=SparklineType.COLUMN,
            data_range="B2:B7",
            location="F2",
        )

        output_path = tmp_path / "column_sparkline.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, sparklines=[sparkline])

        wb = load_workbook(output_path)
        assert output_path.exists()
        wb.close()

    def test_winloss_sparkline(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering a win/loss sparkline."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import SparklineSpec, SparklineType
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sparkline = SparklineSpec(
            type=SparklineType.WIN_LOSS,
            data_range="D2:D7",
            location="F3",
        )

        output_path = tmp_path / "winloss_sparkline.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, sparklines=[sparkline])

        wb = load_workbook(output_path)
        assert output_path.exists()
        wb.close()

    def test_sparkline_with_markers(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test sparkline with high/low markers."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import SparklineMarkers, SparklineSpec, SparklineType
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        sparkline = SparklineSpec(
            type=SparklineType.LINE,
            data_range="B2:B7",
            location="F4",
            markers=SparklineMarkers(
                high=True,
                low=True,
                first=True,
                last=True,
            ),
        )

        output_path = tmp_path / "sparkline_markers.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, sparklines=[sparkline])

        wb = load_workbook(output_path)
        assert output_path.exists()
        wb.close()


# =============================================================================
# Chart Customization Tests
# =============================================================================


class TestXlsxChartCustomization:
    """Test chart customization options."""

    def test_chart_with_custom_size(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test chart with custom width and height."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.COLUMN,
            title="Custom Size Chart",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
            width=15,  # cm
            height=10,  # cm
        )

        output_path = tmp_path / "custom_size_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_chart_with_legend(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test chart with legend positioning."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import (
            ChartSpec,
            ChartType,
            DataRange,
            LegendPosition,
        )
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE,
            title="Chart with Legend",
            data=DataRange(
                categories="A2:A7",
                values=["B2:B7", "C2:C7"],
            ),
            position="E2",
            legend=True,
            legend_position=LegendPosition.BOTTOM,
        )

        output_path = tmp_path / "legend_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_chart_with_axis_titles(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test chart with custom axis titles."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import AxisSpec, ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE,
            title="Axis Titles Chart",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
            x_axis=AxisSpec(title="Month", format="text"),
            y_axis=AxisSpec(title="Amount ($)", format="number"),
        )

        output_path = tmp_path / "axis_titles_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()

    def test_chart_without_gridlines(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test chart with gridlines disabled."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        chart = ChartSpec(
            type=ChartType.LINE,
            title="No Gridlines",
            data=DataRange(
                categories="A2:A7",
                values="B2:B7",
            ),
            position="E2",
            show_gridlines=False,
        )

        output_path = tmp_path / "no_gridlines_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()


# =============================================================================
# Multiple Charts Tests
# =============================================================================


class TestXlsxMultipleCharts:
    """Test multiple charts on same sheet."""

    def test_multiple_charts_same_sheet(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test rendering multiple charts on one sheet."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        charts = [
            ChartSpec(
                type=ChartType.COLUMN,
                title="Sales",
                data=DataRange(categories="A2:A7", values="B2:B7"),
                position="F2",
            ),
            ChartSpec(
                type=ChartType.LINE,
                title="Profit Trend",
                data=DataRange(categories="A2:A7", values="D2:D7"),
                position="F20",
            ),
        ]

        output_path = tmp_path / "multiple_charts.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=charts)

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) == 2
        wb.close()


# =============================================================================
# Edge Cases
# =============================================================================


class TestXlsxChartEdgeCases:
    """Test chart edge cases and error handling."""

    def test_chart_with_empty_data(self, tmp_path: Path) -> None:
        """Test chart with empty data range doesn't crash."""
        from spreadsheet_dl.builder import SheetSpec
        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        empty_sheet = SheetSpec(
            name="Empty",
            columns=[],
            rows=[],
        )

        chart = ChartSpec(
            type=ChartType.COLUMN,
            title="Empty Data",
            data=DataRange(categories="A1:A1", values="B1:B1"),
            position="D2",
        )

        output_path = tmp_path / "empty_data_chart.xlsx"
        renderer = XlsxRenderer()
        # Should not raise an exception
        renderer.render([empty_sheet], output_path, charts=[chart])
        assert output_path.exists()

    def test_chart_long_title_truncated(
        self, chart_data_sheet: SheetSpec, tmp_path: Path
    ) -> None:
        """Test that very long chart titles are handled."""
        from openpyxl import load_workbook

        from spreadsheet_dl.charts import ChartSpec, ChartType, DataRange
        from spreadsheet_dl.xlsx_renderer import XlsxRenderer

        long_title = "A" * 200  # Very long title

        chart = ChartSpec(
            type=ChartType.COLUMN,
            title=long_title,
            data=DataRange(categories="A2:A7", values="B2:B7"),
            position="E2",
        )

        output_path = tmp_path / "long_title_chart.xlsx"
        renderer = XlsxRenderer()
        renderer.render([chart_data_sheet], output_path, charts=[chart])

        wb = load_workbook(output_path)
        ws = wb.active
        assert len(ws._charts) > 0
        wb.close()
