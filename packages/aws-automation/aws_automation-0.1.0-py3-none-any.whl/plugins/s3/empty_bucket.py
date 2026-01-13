"""
plugins/s3/empty_bucket.py - 빈 S3 버킷 탐지

객체가 없는 미사용 S3 버킷 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "s3:ListAllMyBuckets",
        "s3:GetBucketLocation",
        "s3:GetBucketVersioning",
        "s3:GetBucketLifecycleConfiguration",
        "s3:GetBucketLogging",
        "s3:GetBucketReplication",
        "s3:ListBucket",
    ],
}

# 미사용 기준: 90일 이상 접근 없음
UNUSED_DAYS_THRESHOLD = 90


class BucketStatus(Enum):
    """버킷 상태"""

    NORMAL = "normal"
    EMPTY = "empty"
    VERSIONING_ONLY = "versioning_only"
    SMALL = "small"


@dataclass
class BucketInfo:
    """S3 버킷 정보"""

    account_id: str
    account_name: str
    name: str
    region: str
    created_at: datetime | None
    object_count: int = 0
    total_size_bytes: int = 0
    versioning_enabled: bool = False
    has_lifecycle: bool = False
    has_logging: bool = False
    has_replication: bool = False
    encryption_type: str = "None"

    @property
    def total_size_mb(self) -> float:
        return self.total_size_bytes / (1024**2)

    @property
    def total_size_gb(self) -> float:
        return self.total_size_bytes / (1024**3)


@dataclass
class BucketFinding:
    """버킷 분석 결과"""

    bucket: BucketInfo
    status: BucketStatus
    recommendation: str


@dataclass
class S3AnalysisResult:
    """S3 분석 결과 집계"""

    account_id: str
    account_name: str
    total_buckets: int = 0
    empty_buckets: int = 0
    versioning_only_buckets: int = 0
    small_buckets: int = 0
    total_size_gb: float = 0.0
    findings: list[BucketFinding] = field(default_factory=list)


def get_bucket_region(s3_client, bucket_name: str) -> str:
    """버킷 리전 조회"""
    from botocore.exceptions import ClientError

    try:
        response = s3_client.get_bucket_location(Bucket=bucket_name)
        location = response.get("LocationConstraint")
        # None은 us-east-1을 의미
        return str(location) if location else "us-east-1"
    except ClientError:
        return "unknown"


def get_bucket_size_from_cloudwatch(cw_client, bucket_name: str) -> tuple:
    """CloudWatch에서 버킷 크기/객체수 조회 (더 효율적)"""
    from botocore.exceptions import ClientError

    try:
        now = datetime.now(timezone.utc)
        start_time = now - timedelta(days=2)

        # BucketSizeBytes 메트릭
        size_response = cw_client.get_metric_statistics(
            Namespace="AWS/S3",
            MetricName="BucketSizeBytes",
            Dimensions=[
                {"Name": "BucketName", "Value": bucket_name},
                {"Name": "StorageType", "Value": "StandardStorage"},
            ],
            StartTime=start_time,
            EndTime=now,
            Period=86400,
            Statistics=["Average"],
        )

        # NumberOfObjects 메트릭
        count_response = cw_client.get_metric_statistics(
            Namespace="AWS/S3",
            MetricName="NumberOfObjects",
            Dimensions=[
                {"Name": "BucketName", "Value": bucket_name},
                {"Name": "StorageType", "Value": "AllStorageTypes"},
            ],
            StartTime=start_time,
            EndTime=now,
            Period=86400,
            Statistics=["Average"],
        )

        size = 0
        count = 0

        if size_response.get("Datapoints"):
            size = int(size_response["Datapoints"][-1].get("Average", 0))

        if count_response.get("Datapoints"):
            count = int(count_response["Datapoints"][-1].get("Average", 0))

        return size, count, True

    except ClientError:
        return 0, 0, False


def collect_buckets(session, account_id: str, account_name: str) -> list[BucketInfo]:
    """S3 버킷 수집"""
    from botocore.exceptions import ClientError

    s3 = get_client(session, "s3")
    buckets = []

    try:
        response = s3.list_buckets()
    except ClientError:
        return []

    for bucket in response.get("Buckets", []):
        bucket_name = bucket.get("Name", "")
        region = get_bucket_region(s3, bucket_name)

        bucket_info = BucketInfo(
            account_id=account_id,
            account_name=account_name,
            name=bucket_name,
            region=region,
            created_at=bucket.get("CreationDate"),
        )

        # CloudWatch에서 크기/객체수 조회 시도
        try:
            cw = get_client(
                session,
                "cloudwatch",
                region_name=region if region != "unknown" else "us-east-1",
            )
            size, count, success = get_bucket_size_from_cloudwatch(cw, bucket_name)
            if success:
                bucket_info.total_size_bytes = size
                bucket_info.object_count = count
        except ClientError:
            pass

        # CloudWatch 메트릭이 없으면 list_objects_v2로 확인 (첫 1000개만)
        if bucket_info.object_count == 0:
            try:
                s3_regional = get_client(
                    session,
                    "s3",
                    region_name=region if region != "unknown" else "us-east-1",
                )
                obj_response = s3_regional.list_objects_v2(Bucket=bucket_name, MaxKeys=1)
                bucket_info.object_count = obj_response.get("KeyCount", 0)
            except ClientError:
                pass

        # 버킷 설정 확인
        try:
            s3.get_bucket_versioning(Bucket=bucket_name)
            versioning = s3.get_bucket_versioning(Bucket=bucket_name)
            bucket_info.versioning_enabled = versioning.get("Status") == "Enabled"
        except ClientError:
            pass

        try:
            s3.get_bucket_lifecycle_configuration(Bucket=bucket_name)
            bucket_info.has_lifecycle = True
        except ClientError as e:
            if e.response.get("Error", {}).get("Code") != "NoSuchLifecycleConfiguration":
                pass

        try:
            s3.get_bucket_logging(Bucket=bucket_name)
            logging = s3.get_bucket_logging(Bucket=bucket_name)
            bucket_info.has_logging = "LoggingEnabled" in logging
        except ClientError:
            pass

        try:
            encryption = s3.get_bucket_encryption(Bucket=bucket_name)
            rules = encryption.get("ServerSideEncryptionConfiguration", {}).get("Rules", [])
            if rules:
                sse = rules[0].get("ApplyServerSideEncryptionByDefault", {})
                bucket_info.encryption_type = sse.get("SSEAlgorithm", "None")
        except ClientError as e:
            if e.response.get("Error", {}).get("Code") != "ServerSideEncryptionConfigurationNotFoundError":
                pass

        buckets.append(bucket_info)

    return buckets


def analyze_buckets(buckets: list[BucketInfo], account_id: str, account_name: str) -> S3AnalysisResult:
    """버킷 분석"""
    result = S3AnalysisResult(
        account_id=account_id,
        account_name=account_name,
        total_buckets=len(buckets),
    )

    for bucket in buckets:
        result.total_size_gb += bucket.total_size_gb

        # 완전히 빈 버킷
        if bucket.object_count == 0 and bucket.total_size_bytes == 0:
            result.empty_buckets += 1
            result.findings.append(
                BucketFinding(
                    bucket=bucket,
                    status=BucketStatus.EMPTY,
                    recommendation="빈 버킷 - 삭제 검토",
                )
            )
            continue

        # 버전관리만 있는 경우 (삭제 마커 등)
        if bucket.object_count == 0 and bucket.total_size_bytes > 0:
            result.versioning_only_buckets += 1
            result.findings.append(
                BucketFinding(
                    bucket=bucket,
                    status=BucketStatus.VERSIONING_ONLY,
                    recommendation=f"버전 데이터만 존재 ({bucket.total_size_mb:.2f} MB)",
                )
            )
            continue

        # 매우 작은 버킷 (1MB 미만)
        if bucket.total_size_bytes < 1024 * 1024:
            result.small_buckets += 1
            result.findings.append(
                BucketFinding(
                    bucket=bucket,
                    status=BucketStatus.SMALL,
                    recommendation=f"매우 작음 ({bucket.object_count}개, {bucket.total_size_mb:.2f} MB)",
                )
            )
            continue

        result.findings.append(
            BucketFinding(
                bucket=bucket,
                status=BucketStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[S3AnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "S3 빈 버킷 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "전체 버킷", "빈 버킷", "버전만", "소형", "총 크기"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.total_buckets)
        ws.cell(row=row, column=3, value=r.empty_buckets)
        ws.cell(row=row, column=4, value=r.versioning_only_buckets)
        ws.cell(row=row, column=5, value=r.small_buckets)
        ws.cell(row=row, column=6, value=f"{r.total_size_gb:.2f} GB")
        if r.empty_buckets > 0:
            ws.cell(row=row, column=3).fill = red_fill
        if r.versioning_only_buckets > 0:
            ws.cell(row=row, column=4).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Buckets")
    detail_headers = [
        "Account",
        "Bucket",
        "Region",
        "상태",
        "객체수",
        "크기",
        "버전관리",
        "Lifecycle",
        "암호화",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != BucketStatus.NORMAL:
                detail_row += 1
                bucket = f.bucket
                ws_detail.cell(row=detail_row, column=1, value=bucket.account_name)
                ws_detail.cell(row=detail_row, column=2, value=bucket.name)
                ws_detail.cell(row=detail_row, column=3, value=bucket.region)
                ws_detail.cell(row=detail_row, column=4, value=f.status.value)
                ws_detail.cell(row=detail_row, column=5, value=bucket.object_count)
                ws_detail.cell(row=detail_row, column=6, value=f"{bucket.total_size_mb:.2f} MB")
                ws_detail.cell(
                    row=detail_row,
                    column=7,
                    value="Enabled" if bucket.versioning_enabled else "Disabled",
                )
                ws_detail.cell(
                    row=detail_row,
                    column=8,
                    value="있음" if bucket.has_lifecycle else "없음",
                )
                ws_detail.cell(row=detail_row, column=9, value=bucket.encryption_type)
                ws_detail.cell(row=detail_row, column=10, value=f.recommendation)

    # 컬럼 너비 자동 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"S3_EmptyBucket_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> S3AnalysisResult | None:
    """단일 계정의 S3 버킷 수집 및 분석 (병렬 실행용)

    S3는 글로벌 서비스이므로 region은 무시됩니다.
    parallel_collect의 중복 제거가 계정당 한 번만 실행되도록 보장합니다.
    """
    buckets = collect_buckets(session, account_id, account_name)
    if not buckets:
        return None
    return analyze_buckets(buckets, account_id, account_name)


def run(ctx) -> None:
    """빈 S3 버킷 분석"""
    console.print("[bold]S3 빈 버킷 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="s3")
    results: list[S3AnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_empty = sum(r.empty_buckets for r in results)
    total_versioning = sum(r.versioning_only_buckets for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"빈 버킷: [yellow]{total_empty}개[/yellow], 버전만: [yellow]{total_versioning}개[/yellow]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("s3-empty").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
