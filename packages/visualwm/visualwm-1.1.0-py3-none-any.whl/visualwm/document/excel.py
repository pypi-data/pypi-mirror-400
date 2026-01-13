"""
Excel文档水印处理器 - 支持.xlsx格式添加明水印
"""

import os
from pathlib import Path
from typing import Optional, Union, List
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.worksheet.header_footer import HeaderFooterItem

from visualwm.core.info import WatermarkInfo
from visualwm.core.style import WatermarkStyle, RiskLevel


class ExcelWatermark:
    """
    Excel文档水印处理器
    
    支持在Excel文档(.xlsx)中添加明水印，包括：
    - 页眉/页脚水印
    - 单元格背景水印
    - 批注水印
    - 首行/首列水印
    
    Features:
        - 支持.xlsx格式
        - 多工作表处理
        - 自定义样式
        - 批量处理
    """
    
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xlsm'}
    
    def __init__(self, default_style: Optional[WatermarkStyle] = None):
        """
        初始化Excel水印处理器
        
        Args:
            default_style: 默认水印样式
        """
        self.default_style = default_style or WatermarkStyle.default()
    
    def embed(
        self,
        input_path: Union[str, Path],
        output_path: Union[str, Path],
        info: WatermarkInfo,
        style: Optional[WatermarkStyle] = None,
        watermark_type: str = "header_footer",
        sheets: Optional[List[str]] = None
    ) -> str:
        """
        在Excel文档中嵌入水印
        
        Args:
            input_path: 输入文档路径
            output_path: 输出文档路径
            info: 水印信息
            style: 水印样式
            watermark_type: 水印类型 ("header_footer", "first_row", "comment", "background")
            sheets: 要处理的工作表名称列表，None表示所有工作表
            
        Returns:
            输出文件路径
        """
        input_path = Path(input_path)
        output_path = Path(output_path)
        style = style or self.default_style
        
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")
        
        # 打开工作簿
        wb = load_workbook(input_path)
        
        # 确定要处理的工作表
        target_sheets = sheets if sheets else wb.sheetnames
        
        # 为每个工作表添加水印
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                if watermark_type == "header_footer":
                    self._add_header_footer_watermark(ws, info, style)
                elif watermark_type == "first_row":
                    self._add_first_row_watermark(ws, info, style)
                elif watermark_type == "comment":
                    self._add_comment_watermark(ws, info, style)
                elif watermark_type == "background":
                    self._add_background_watermark(ws, info, style)
                else:
                    raise ValueError(f"不支持的水印类型: {watermark_type}")
        
        # 保存
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path)
    
    def embed_new(
        self,
        output_path: Union[str, Path],
        data: List[List[str]],
        info: WatermarkInfo,
        style: Optional[WatermarkStyle] = None,
        sheet_name: str = "Sheet1"
    ) -> str:
        """
        创建带水印的新Excel文档
        
        Args:
            output_path: 输出文档路径
            data: 表格数据（二维列表）
            info: 水印信息
            style: 水印样式
            sheet_name: 工作表名称
            
        Returns:
            输出文件路径
        """
        output_path = Path(output_path)
        style = style or self.default_style
        
        # 创建新工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 添加数据
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # 添加水印
        self._add_header_footer_watermark(ws, info, style)
        self._add_first_row_watermark(ws, info, style, insert_row=True)
        
        # 保存
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path)
    
    def _add_header_footer_watermark(
        self,
        ws,
        info: WatermarkInfo,
        style: WatermarkStyle
    ):
        """添加页眉页脚水印"""
        watermark_text = self._format_watermark_text(info, style)
        
        # 设置页眉
        ws.oddHeader.center.text = watermark_text
        ws.oddHeader.center.size = int(style.font_size * 0.6)
        ws.oddHeader.center.font = style.font_name
        
        # 设置页脚
        ws.oddFooter.center.text = watermark_text
        ws.oddFooter.center.size = int(style.font_size * 0.6)
        ws.oddFooter.center.font = style.font_name
        
        # 偶数页
        ws.evenHeader.center.text = watermark_text
        ws.evenFooter.center.text = watermark_text
    
    def _add_first_row_watermark(
        self,
        ws,
        info: WatermarkInfo,
        style: WatermarkStyle,
        insert_row: bool = False
    ):
        """添加首行水印"""
        watermark_text = self._format_watermark_text(info, style)
        
        if insert_row:
            ws.insert_rows(1)
        
        # 合并首行单元格
        max_col = ws.max_column if ws.max_column > 1 else 5
        
        # 写入水印
        cell = ws.cell(row=1, column=1, value=watermark_text)
        
        # 设置样式
        cell.font = Font(
            name=style.font_name,
            size=int(style.font_size * 0.7),
            bold=style.bold,
            italic=style.italic,
            color=f"{style.color[0]:02x}{style.color[1]:02x}{style.color[2]:02x}"
        )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并单元格
        if max_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        
        # 设置行高
        ws.row_dimensions[1].height = 25
    
    def _add_comment_watermark(
        self,
        ws,
        info: WatermarkInfo,
        style: WatermarkStyle
    ):
        """添加批注水印（在A1单元格）"""
        watermark_text = self._format_watermark_text(info, style)
        
        # 创建批注
        comment = Comment(
            f"【文档水印】\n{watermark_text}",
            "VisualWM"
        )
        comment.width = 300
        comment.height = 100
        
        # 添加到A1
        ws['A1'].comment = comment
    
    def _add_background_watermark(
        self,
        ws,
        info: WatermarkInfo,
        style: WatermarkStyle
    ):
        """添加背景水印（通过浅色填充实现提示效果）"""
        watermark_text = self._format_watermark_text(info, style)
        
        # 在工作表的显眼位置添加水印信息
        # 使用最后一列之后的位置
        max_col = ws.max_column + 2 if ws.max_column else 3
        max_row = ws.max_row if ws.max_row else 1
        
        # 添加水印列
        for row in range(1, max_row + 1, 10):  # 每10行添加一次
            cell = ws.cell(row=row, column=max_col, value=f"| {watermark_text}")
            cell.font = Font(
                size=8,
                color="CCCCCC"
            )
        
        # 设置列宽
        ws.column_dimensions[ws.cell(row=1, column=max_col).column_letter].width = 50
    
    def _format_watermark_text(self, info: WatermarkInfo, style: WatermarkStyle) -> str:
        """格式化水印文本"""
        text = info.to_string(separator=" | ", compact=True)
        
        if style.prefix:
            text = f"{style.prefix} {text}"
        if style.suffix:
            text = f"{text} {style.suffix}"
        
        return text
    
    def extract(
        self,
        excel_path: Union[str, Path]
    ) -> Optional[WatermarkInfo]:
        """
        从Excel文档中提取水印信息
        
        Args:
            excel_path: 文档路径
            
        Returns:
            提取的水印信息
        """
        wb = load_workbook(excel_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 从页眉提取
            header_text = ws.oddHeader.center.text
            if header_text and ("用户:" in header_text or "工号:" in header_text):
                return WatermarkInfo.from_string(header_text)
            
            # 从首行提取
            first_row_value = ws.cell(row=1, column=1).value
            if first_row_value and ("用户:" in str(first_row_value) or "工号:" in str(first_row_value)):
                return WatermarkInfo.from_string(str(first_row_value))
            
            # 从批注提取
            if ws['A1'].comment:
                comment_text = ws['A1'].comment.text
                if "用户:" in comment_text or "工号:" in comment_text:
                    # 提取水印部分
                    for line in comment_text.split('\n'):
                        if "用户:" in line or "工号:" in line:
                            return WatermarkInfo.from_string(line.strip())
        
        return None
    
    def batch_embed(
        self,
        input_dir: Union[str, Path],
        output_dir: Union[str, Path],
        info: WatermarkInfo,
        style: Optional[WatermarkStyle] = None,
        recursive: bool = False
    ) -> List[str]:
        """
        批量添加水印
        
        Args:
            input_dir: 输入目录
            output_dir: 输出目录
            info: 水印信息
            style: 水印样式
            recursive: 是否递归处理
            
        Returns:
            处理的文件列表
        """
        input_dir = Path(input_dir)
        output_dir = Path(output_dir)
        
        processed = []
        
        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        
        for input_file in input_dir.glob(pattern):
            if input_file.name.startswith('~$'):  # 跳过临时文件
                continue
            
            relative_path = input_file.relative_to(input_dir)
            output_file = output_dir / relative_path
            
            try:
                self.embed(input_file, output_file, info, style)
                processed.append(str(output_file))
            except Exception as e:
                print(f"处理失败 {input_file}: {e}")
        
        return processed
    
    def get_workbook_info(self, excel_path: Union[str, Path]) -> dict:
        """
        获取Excel工作簿信息
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            工作簿信息字典
        """
        wb = load_workbook(excel_path, read_only=True)
        
        info = {
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "sheets": {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info["sheets"][sheet_name] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }
        
        wb.close()
        return info
