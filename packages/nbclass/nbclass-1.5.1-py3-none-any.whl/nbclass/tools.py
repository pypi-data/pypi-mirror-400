# -*- coding: utf-8 -*-
"""
@ Created on 2024-09-04 15:37
---------
@summary: 
---------
@author: XiaoBai
"""
import calendar
import datetime
import inspect
import json
import random
import re
import stat
import string
import subprocess
import traceback
import uuid
from random import shuffle
from typing import Literal
from urllib import parse

import xlrd
import xlwt
from openpyxl import Workbook
from xlutils.copy import copy

from nbclass.common import *
from nbclass.decorators import *
from nbclass.email_sender import EmailSender
from nbclass.encrypt import *
from nbclass.geo_transform import *
from nbclass.typeshed import *


class MySubprocessPopen(subprocess.Popen):
    def __init__(self, *args, **kwargs):
        # 在调用父类（即 subprocess.Popen）的构造方法时，将 encoding 参数直接置为 UTF-8 编码格式
        kwargs['encoding'] = 'UTF-8'
        super().__init__(*args, **kwargs)


# 必须要在导入 PyExecJS 模块前，就将 subprocess.Popen 类重置为新的类
subprocess.Popen = MySubprocessPopen

import execjs


def add_batch_excel(datas: dict, filepath: str):
    """
    批量添加数据到excel表格
    data: {
            sheetName: [[sheetHeaders],[values]...]
        }
    filepath: 保存的文件路径
    """

    # 创建一个Workbook对象
    workbook = Workbook()

    # 遍历字典数据
    for sheet_name, sheet_data in datas.items():
        # 创建一个工作表
        sheet = workbook.create_sheet(title=sheet_name)

        # 写入表头
        header_row = sheet_data[0]
        for col_num, header in enumerate(header_row, 1):
            sheet.cell(row=1, column=col_num).value = header

        # 写入数据
        for row_num, row_data in enumerate(sheet_data[1:], 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value

    # 保存Excel文件
    workbook.save(filepath)


def b64encode(context, decode: bool = True) -> Union[str, bytes]:
    if isinstance(context, bytes):
        encoded_bytes = base64.b64encode(context)
    else:
        encoded_bytes = base64.b64encode(ensure_str(context).encode('utf-8'))

    return encoded_bytes.decode('utf-8') if decode else encoded_bytes


def b64decode(context: str, decode: bool = True) -> Union[str, bytes]:
    decoded_bytes = base64.b64decode(ensure_str(context))

    return decoded_bytes.decode('utf-8') if decode else decoded_bytes


def chmod(_path, mode: int = stat.S_IWRITE):
    """ 修改文件权限 """
    os.chmod(_path, os.stat(_path).st_mode | mode)


def delay_time(sleep_time=60):
    """
    @summary: 睡眠  默认1分钟
    ---------
    @param sleep_time: 以秒为单位
    ---------
    @result:
    """

    time.sleep(sleep_time)


def dirname(p: AnyStr, level: int = 0) -> AnyStr:
    """ 默认返回当前文件的绝对路径
    :param p: __file__ or 绝对文件路径
    :param level: 上几级父目录
    dirname(r"E:\nbclass\test.py", 1)  => E:\nbclass
    """
    abspath = os.path.abspath(p)

    if level <= 0:
        return abspath

    for _ in range(level):
        abspath = os.path.dirname(abspath)

    return abspath


def dict2query_string(params: dict, reverse: bool = True) -> str:
    """ 字典转查询参数 """
    sorted_params = sorted(params.items(), reverse=reverse)
    # 将排序后的键值对转换为查询参数格式
    return '&'.join([f"{key}={value}" for key, value in sorted_params])


def format_seconds(seconds):
    """
    @summary: 将秒转为时分秒
    ---------
    @param seconds:
    ---------
    @result: 2天3小时2分49秒
    """

    seconds = int(seconds + 0.5)  # 向上取整

    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    d, h = divmod(h, 24)

    times = ""
    if d:
        times += "{}天".format(d)
    if h:
        times += "{}小时".format(h)
    if m:
        times += "{}分".format(m)
    if s:
        times += "{}秒".format(s)

    return times


def ensure_str(text) -> str:
    if not text:
        return ""
    else:
        return str(text)


def ensure_int(text) -> int:
    if not text:
        return 0
    else:
        return int(text)


def ensure_float(text) -> float:
    if not text:
        return 0.0
    else:
        return float(text)


def email_warning(
        message,
        title,
        email_receiver=None,  # 收件人
        email_sender=None,  # 发件人
        email_password=None,  # 授权码
        email_smtpserver=None,  # 邮件服务器
        filepath: str = None
):
    # 为了加载最新的配置
    email_sender = email_sender
    email_password = email_password
    email_receiver = email_receiver
    email_smtpserver = email_smtpserver

    if not all([message, email_sender, email_password, email_receiver]):
        logger.error("message, email_sender, email_password, email_receiver都不能为空")
        return

    if isinstance(email_receiver, str):
        email_receiver = [email_receiver]

    with EmailSender(
            username=email_sender, password=email_password, smtp_server=email_smtpserver
    ) as email:
        return email.send(receivers=email_receiver, title=title, content=message, filepath=filepath)


def extract_branch_name(text):
    # 使用正则表达式匹配括号内的内容
    match = re.search(r'(.+?)\((.+?)\)', text)
    if match:
        # 提取括号外和括号内的内容
        outside = match.group(1).strip()
        inside = match.group(2).strip()
        return outside, inside
    else:
        # 如果没有匹配到，返回原始文本和空字符串
        return text, ''


def format_string_to_list(context):
    # 去除空格和换行符
    context = context.replace(' ', '').replace('\n', '')

    # 格式化为列表
    lst = list(context)

    return lst


def generate_random_string(length):
    # 生成包含大小写字母和数字的可选字符集
    choices = string.ascii_letters + string.digits
    # 从可选字符集中随机选择 length 个字符，组成一个字符串并返回
    return ''.join(random.choice(choices) for _ in range(length))


def get_datetime(
        current_time: bool = False,
        days: int = 0,
        hours: int = 0,
        minutes: int = 0,
        seconds: int = 0,
        microseconds: int = 0,
        milliseconds: int = 0,
        weeks: int = 0
) -> datetime.datetime:
    """
    current_time: True 表示现在的时间; False 凌晨零点的时间
    """
    current_date = datetime.datetime.now()
    timedelta = datetime.timedelta(
        days=days, hours=hours, minutes=minutes, seconds=seconds,
        microseconds=microseconds, milliseconds=milliseconds, weeks=weeks
    )

    if not current_time:
        # 创建凌晨零点的时间
        current_date = datetime.datetime.combine(current_date, datetime.datetime.min.time())

    current_date += timedelta

    return current_date


get_cache_expire = get_datetime


def get_cache_expire_at(now: bool = False) -> int:
    """
    now: is False 获取第二天凌晨零点的时间戳(秒级)
        is True 获取当天凌晨零点的时间戳(秒级)
    """
    current_date = datetime.datetime.now().date()
    if now:
        # 创建一个表示今天凌晨零点的 datetime 对象
        midnight = datetime.datetime.combine(current_date, datetime.datetime.min.time())

        # 将 datetime 对象转换为时间戳
        timestamp = midnight.timestamp()

    else:
        # 获取第二天日期
        next_day = current_date + datetime.timedelta(days=1)

        # 创建一个表示第二天凌晨零点的 datetime 对象
        midnight = datetime.datetime.combine(next_day, datetime.datetime.min.time())

        # 将 datetime 对象转换为时间戳
        timestamp = midnight.timestamp()
    return int(timestamp)


def get_cache_expire_time() -> int:
    """ 获取到凌晨零点的时间差 """
    # 获取当前时间
    now = datetime.datetime.now()

    # 计算距离第二天凌晨 12 点的时间差
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=1)
    expire_time = (midnight - now).seconds
    return expire_time


def get_cache_expire_at_week() -> int:
    """ 计算下一个周一的时间戳 """
    # 获取当前日期
    current_date = datetime.datetime.now()
    # 计算距离下一个周一的天数
    days_to_next_monday = (8 - current_date.isoweekday())
    # 计算下个周一的日期
    next_monday = current_date + datetime.timedelta(days=days_to_next_monday)
    # 设置为零点
    next_monday = next_monday.replace(hour=0, minute=0, second=0, microsecond=0)
    # 转换为时间戳
    timestamp = int(next_monday.timestamp())
    return timestamp


def get_max_days(year: int, month: int) -> int:
    """ 获取指定年月的最大天数 """
    return calendar.monthrange(year, month)[1]


def get_month_range(date_str=None, is_str: bool = False) -> tuple:
    """ 获取指定日期的月份范围， 默认当月 """
    if date_str:
        now = date_str
        if isinstance(now, str) and len(now) == 7:
            now = datetime.datetime.strptime(now, '%Y-%m').date()
        elif not isinstance(now, datetime.date):
            now = datetime.datetime.strptime(now, '%Y-%m-%d').date()
    else:
        now = datetime.date.today()  # 获取当前日期时间

    first_day = now.replace(day=1)  # 本月第一天
    # 上个月的结束日期
    last_day = first_day.replace(day=calendar.monthrange(first_day.year, first_day.month)[1])
    if is_str:
        return str(first_day), str(last_day)
    else:
        return first_day, last_day


def get_now_timestamp(msec: bool = True):
    if msec:
        return int(time.time() * 1000)
    else:
        return int(time.time())


def get_uuid():
    return uuid.uuid4()


def isalnum(arg: str) -> bool:
    """
    判断字符串是否为纯字母数字字符串
    """
    return arg.isalnum() and arg.isascii()


def isdir(s: Union[StrOrBytesPath, int]) -> bool:
    return os.path.isdir(s)


def isfile(s: Union[StrOrBytesPath, int]) -> bool:
    return os.path.isfile(s)


def is_base64(s: str):
    try:
        b64decode(s)
    except Exception as e:
        logger.warning(e)
        return False
    else:
        return True


@overload
def join_path(__path: StrPath, *paths: StrPath) -> str: ...


def join_path(__path: BytesPath, *paths: BytesPath) -> bytes:
    return os.path.join(__path, *paths)


def join_str(*args, sep: str = '') -> str:
    return f'{sep}'.join([ensure_str(x) for x in args])


def json_dumps(data, indent=4, ensure_ascii: bool = False, sort_keys=False, separators: tuple = None):
    """
    @summary: 序列化json
    ---------
    @param data: json格式的字符串或json对象
    @param indent: 格式化的缩进长度
    @param sort_keys: 是否排序
    @param ensure_ascii: 是否转为ascii编码
    @param separators: 是否转为最紧凑的json格式
    ---------
    @result: 格式化后的字符串
    """

    if isinstance(data, str):
        return data
    else:
        if indent == 0 or indent is None:
            separators = (',', ':')

        data = json.dumps(
            data,
            ensure_ascii=ensure_ascii,
            indent=indent,
            skipkeys=True,
            separators=separators,
            sort_keys=sort_keys,
            default=str,
        )
    return data


def json_loads(obj) -> Union[dict, list]:
    if isinstance(obj, (dict, list)):
        return obj

    return json.loads(obj)


def jsonp_to_json(jsonp):
    """jsonp转字典"""
    """ 方法一 """
    # start_index = jsonp.index('(') + 1
    # end_index = jsonp.rindex(')')
    # # callback_name = jsonp[:start_index - 1]
    # data_str = jsonp[start_index:end_index]
    # return json_loads(data_str)

    """ 方法二 """
    result = ''.join(re.findall(r'\w+[(](.*)[)]', jsonp, re.S))
    return json_loads(result)


def mkdir(path) -> None:
    if not os.path.exists(path):
        os.mkdir(path)


def math_ceil(x):
    """
    向上取整
    """
    return math.ceil(x)


def match_chinese(text):
    # 使用正则表达式匹配所有中文字符
    pattern = r'[\u4e00-\u9fff]+'
    matches = re.findall(pattern, text)
    return matches


def match_non_chinese(text):
    # 使用正则表达式匹配所有非中文字符
    pattern = r'[^\u4e00-\u9fff]+'
    matches = re.findall(pattern, text)
    return matches


def printer(*args, align: Literal["<", ">", "^"] = "<", width=10):
    """
    一个对齐打印方法
    printer("Hello", "World", "Python", align="<", width=10)
    printer("HE", "work", "py", align="<", width=10)
    输出:
        Hello     World     Python
        HE        work      py
    :param args: 需要打印的参数
    :param align: 对齐方式 左对齐: <, 右对齐: >, 居中对齐: ^, 默认左对齐
    :param width: 间隔距离
    :return: None
    """
    if align not in ['<', '>', '^']:
        align = "<"

    print(f''.join([f"{arg:{align}{width}}" for arg in args]))


def print_parent_caller(level: int = 1, types: bool = False):
    """
    打印函数父级调用者的装饰器
    Arg:
        level: 打印父级调用者 level=1, 打印祖父级调用者 level=2
        types:
    """

    def _print_parent_caller(func):
        def wrapper(*args, **kwargs):
            if types:
                frames = inspect.getouterframes(inspect.currentframe(), 2)
                if len(frames) > level:
                    parent_frame = frames[level]  # 直属父级调用者帧对象
                    parent_name = parent_frame[3]  # 直属父级调用者函数名
                    parent_line = parent_frame[2]  # 直属父级调用者行号
                    parent_file = parent_frame[1]  # 直属父级调用者文件路径
                    logger.info(f"File: {parent_file}, Line: {parent_line}, Function: {parent_name}")
                else:
                    logger.info("No parent caller found.")
            else:
                stack = traceback.extract_stack()
                for frame in stack[:-1]:
                    frame_file, frame_line, frame_func, _ = frame
                    logger.info(f"File: {frame_file}, Line: {frame_line}, Function: {frame_func}")

            return func(*args, **kwargs)

        return wrapper

    return _print_parent_caller


def path_exists(_path: str) -> bool:
    """ 查看文件路径是否存在 """
    return os.path.exists(_path)


def random_choices(datas: list, limit: int = 3) -> list:
    """ 随机从列表中取值 """
    return random.choices(datas, k=limit)


def random_random():
    return random.random()


def remove_file(path) -> None:
    """ 删除临时文件 """
    os.remove(path)


def remove_html(text: str) -> str:
    """ 正则去除HTML标签 """
    # regex = re.compile(r'<(?!b\b)[^>]*>')  # 保留b标签
    regex = re.compile(r'<[^>]+>')
    return regex.sub('', text)


def remove_non_chinese(text):
    """ 去除非汉字部分 """
    pattern = re.compile(r'[^\u4e00-\u9fa5]')
    return re.sub(pattern, '', text)


def remove_special_chars(chars: str, array: list, exclude: str = '') -> str:
    """
    移除特殊字符
    :param chars: 待清洗的文本
    :param array: 待移除的特殊文本列表
    :param exclude: 需保留的文本列表
    :return:
    """
    if not chars:
        return chars

    if not array:
        array = list("　 \r\n\t,，。·.．;；:：、！@$%*^`~=+&'\"|_-\\/")

    # 去除字符
    sb = []
    rm: bool = False
    for char in chars:
        if char in array and char not in exclude:
            rm = True
            continue
        sb.append(char)

    return ''.join(sb) if rm else chars


def replace_half_angle(s: str) -> str:
    """
    将字符串中的全角符号替换为半角符号
    """
    # 自定义特殊符号映射
    custom_map = {
        '《': '<',
        '》': '>',
        '【': '[',
        '】': ']',
        '“': '"',
        '”': '"',
        '‘': "'",
        '’': "'",
        '、': '\\',
        '，': ',',
        '。': '.',
        '——': '-',
        '…': '...',
        '￥': '¥',
    }
    result = []
    for char in s:
        if char in custom_map:
            result.append(custom_map[char])
        else:
            code = ord(char)
            # 全角空格
            if code == 0x3000:
                result.append(' ')
            # 全角英符号
            elif 0xFF01 <= code <= 0xFF5E:
                result.append(chr(code - 0xFEE0))
            else:
                result.append(char)

    return ''.join(result)


def save_excel(data, filename, headers, sheet: str = "基本详情", folder: str = "数据", encoding='utf-8'):
    """
    使用前，请先阅读代码
    :param data: 需要保存的data字典(有格式要求)
    :param headers: 表头
    :param sheet: excel文件的sheet表名
    :param folder: 需要保存的文件名称
    :param filename: 需要保存的文件名
    :param encoding: 文件编码
    :return:
    格式要求:
        data = {
        '基本详情': ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']
    }
    注意：这个模板程序会将data的key值作为excel的表名来判断，一样才会保存
    """
    os_mkdir_path = os.getcwd() + f'/{folder}/'
    # 判断这个路径是否存在，不存在就创建
    if not os.path.exists(os_mkdir_path):
        os.mkdir(os_mkdir_path)
    # 判断excel表格是否存在           工作簿文件名称
    os_excel_path = os_mkdir_path + filename
    if not os.path.exists(os_excel_path):
        # 不存在，创建工作簿(也就是创建excel表格)
        workbook = xlwt.Workbook(encoding=encoding)
        """工作簿中创建新的sheet表"""  # 设置表名
        worksheet1 = workbook.add_sheet(sheet, cell_overwrite_ok=True)
        """设置sheet表的表头"""
        sheet1_headers = headers
        # 将表头写入工作簿
        for header_num in range(0, len(sheet1_headers)):
            # 设置表格长度
            worksheet1.col(header_num).width = 2560 * 3
            # 写入            行, 列,           内容
            worksheet1.write(0, header_num, sheet1_headers[header_num])
        # 循环结束，代表表头写入完成，保存工作簿
        workbook.save(os_excel_path)
    # 判断工作簿是否存在
    if os.path.exists(os_excel_path):
        # 打开工作簿
        workbook = xlrd.open_workbook(os_excel_path)
        # 获取工作薄中所有表的个数
        sheets = workbook.sheet_names()
        for i in range(len(sheets)):
            for name in data.keys():
                worksheet = workbook.sheet_by_name(sheets[i])
                # 获取工作薄中所有表中的表名与数据名对比
                if worksheet.name == name:
                    # 获取表中已存在的行数
                    rows_old = worksheet.nrows
                    # 将xlrd对象拷贝转化为xlwt对象
                    new_workbook = copy(workbook)
                    # 获取转化后的工作薄中的第i张表
                    new_worksheet = new_workbook.get_sheet(i)
                    for num in range(0, len(data[name])):
                        new_worksheet.write(rows_old, num, data[name][num])
                    new_workbook.save(os_excel_path)


def save_excels(data, filename, headers, folder: str = "数据", encoding='utf-8'):
    """
    一个工作簿创建多张表
    """
    sheet_name = [i for i in data.keys()][0]
    os_mkdir_path = os.getcwd() + f'/{folder}/'

    if not os.path.exists(os_mkdir_path):
        os.mkdir(os_mkdir_path)

    os_excel_path = os_mkdir_path + filename

    if not os.path.exists(os_excel_path):
        # 不存在，创建工作簿(也就是创建excel表格)
        workbook = xlwt.Workbook(encoding=encoding)
        worksheet1 = workbook.add_sheet(sheet_name, cell_overwrite_ok=True)
        for header_num in range(0, len(headers)):
            worksheet1.col(header_num).width = 2560 * 3
            worksheet1.write(0, header_num, headers[header_num])

        workbook.save(os_excel_path)

    workbook = xlrd.open_workbook(os_excel_path)
    sheets_list = workbook.sheet_names()

    if sheet_name not in sheets_list:
        work = copy(workbook)
        sh = work.add_sheet(sheet_name)
        excel_headers_tuple = headers

        for head_num in range(0, len(excel_headers_tuple)):
            sh.col(head_num).width = 2560 * 3
            sh.write(0, head_num, excel_headers_tuple[head_num])

        work.save(os_excel_path)

    if os.path.exists(os_excel_path):
        workbook = xlrd.open_workbook(os_excel_path)
        sheets = workbook.sheet_names()

        for i in range(len(sheets)):
            for name in data.keys():
                worksheet = workbook.sheet_by_name(sheets[i])

                if worksheet.name == name:
                    rows_old = worksheet.nrows
                    new_workbook = copy(workbook)
                    new_worksheet = new_workbook.get_sheet(i)

                    for num in range(0, len(data[name])):
                        new_worksheet.write(rows_old, num, data[name][num])
                    new_workbook.save(os_excel_path)


def shuffle_str(s):
    """ 随机打乱字符串顺序 """
    str_list = list(s)
    # 调用random模块的shuffle函数打乱列表
    shuffle(str_list)
    return ''.join(str_list)


def str2datetime(date_str: str, _format: Union[str, DateFormat] = DateFormat.Y_M_D) -> datetime.datetime:
    """ 字符串格式化为datetime类型 """
    return datetime.datetime.strptime(date_str, _format)


def timestamp_to_date(timestamp, time_format="%Y-%m-%d %H:%M:%S"):
    """
    @summary:
    ---------
    @param timestamp: 将时间戳转化为日期
    @param time_format: 日期格式
    ---------
    @result: 返回日期
    """
    if timestamp is None:
        raise ValueError("timestamp is null")

    date = time.localtime(timestamp)
    return time.strftime(time_format, date)


def url_convert(url) -> tuple:
    """ 返回元组形式的url参数 """
    parsed_url = parse.urlparse(url)
    base_url = parsed_url.scheme + "://" + parsed_url.netloc + parsed_url.path
    params = parse.parse_qs(parsed_url.query)

    converted_params = []
    for key, value in params.items():
        if len(value) == 1:
            converted_params.append((key, value[0]))
        else:
            converted_params.append((key, value))

    return base_url, tuple(converted_params)


def url_parse(url: str) -> dict:
    """提取url中的params参数，返回item"""
    item = {}
    parsed_url = parse.urlparse(url)

    # 提取参数
    params = parse.parse_qs(parsed_url.query)

    # 输出参数值
    for key, value in params.items():
        item[key] = value.pop()

    return item


def url_quote(urls, plus: bool = False):
    """url编码"""
    if plus:
        return parse.quote_plus(urls)

    return parse.quote(urls)


def url_unquote(urls):
    """url解码"""
    return parse.unquote(urls)


def url_encode(url: str, params: dict):
    query_string = parse.urlencode(params)
    full_url = url + '?' + query_string
    return full_url


def with_save(file_path, content, mode="wb", encoding="utf-8") -> bool:
    try:
        if isinstance(content, str):
            mode = "w"

        if 'b' in mode:
            encoding = None

        if isinstance(content, (list, dict)):
            mode = 'w'
            content = json_dumps(content, indent=0)

        with open(file_path, mode, encoding=encoding) as fp:
            fp.write(content)
    except Exception as e:
        logger.error(e)
        return False
    else:
        logger.success(f"{file_path} 保存成功")
        return True


def with_open(file_path, mode="r", encoding="UTF-8", ignore: bool = False) -> AnyStr:
    if 'b' in mode:  # 二进制文件不允许编码
        encoding = None

    with open(file_path, mode=mode, encoding=encoding) as fp:
        file = fp.read()

    if ignore:
        return file.encode('gb2312', 'ignore').decode('gb2312')
    else:
        return file


class ExecuteJs:
    def __init__(self, js, func: str = None):
        if isfile(js):
            js = with_open(js)
        self.func = func
        self.execjs = execjs.compile(js)

    def call(self, *args, func: str = None):
        """
        @func: 要执行的js函数名
        args: js函数实参
        """
        return self.execjs.call(func or self.func, *args)
