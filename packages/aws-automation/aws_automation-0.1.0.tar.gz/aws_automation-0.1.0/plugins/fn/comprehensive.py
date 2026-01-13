"""
plugins/fn/comprehensive.py - Lambda 종합 분석 보고서

Lambda 함수 종합 분석:
- 런타임 EOL 분석
- 메모리 사용량 최적화
- 비용 분석
- 성능 지표 (에러율, Throttle)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import (
    get_lambda_monthly_cost,
    get_lambda_provisioned_monthly_cost,
)

from .common.collector import (
    LambdaFunctionInfo,
    collect_functions_with_metrics,
)
from .common.runtime_eol import (
    EOLStatus,
    get_recommended_upgrade,
    get_runtime_info,
)

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:ListTags",
        "lambda:ListProvisionedConcurrencyConfigs",
        "lambda:GetFunctionConcurrency",
        "cloudwatch:GetMetricStatistics",
    ],
}


class IssueType(Enum):
    """이슈 유형"""

    RUNTIME_EOL = "runtime_eol"
    MEMORY_OVERSIZED = "memory_oversized"
    MEMORY_UNDERSIZED = "memory_undersized"
    HIGH_ERROR_RATE = "high_error_rate"
    THROTTLED = "throttled"
    UNUSED = "unused"
    TIMEOUT_RISK = "timeout_risk"


class Severity(Enum):
    """심각도"""

    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class LambdaIssue:
    """Lambda 이슈"""

    issue_type: IssueType
    severity: Severity
    description: str
    recommendation: str
    potential_savings: float = 0.0


@dataclass
class LambdaComprehensiveResult:
    """Lambda 종합 분석 결과"""

    function: LambdaFunctionInfo
    issues: list[LambdaIssue] = field(default_factory=list)
    estimated_monthly_cost: float = 0.0
    memory_recommendation: int | None = None
    potential_savings: float = 0.0

    @property
    def has_critical_issues(self) -> bool:
        return any(i.severity == Severity.CRITICAL for i in self.issues)

    @property
    def has_high_issues(self) -> bool:
        return any(i.severity == Severity.HIGH for i in self.issues)

    @property
    def issue_count(self) -> int:
        return len(self.issues)


@dataclass
class ComprehensiveAnalysisResult:
    """종합 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_functions: int = 0
    functions_with_issues: int = 0
    runtime_eol_count: int = 0
    memory_issue_count: int = 0
    error_issue_count: int = 0
    total_monthly_cost: float = 0.0
    potential_savings: float = 0.0
    results: list[LambdaComprehensiveResult] = field(default_factory=list)


# =============================================================================
# 분석
# =============================================================================


def analyze_function_comprehensive(
    func: LambdaFunctionInfo,
    region: str,
    memory_stats: dict | None = None,
) -> LambdaComprehensiveResult:
    """Lambda 함수 종합 분석"""
    result = LambdaComprehensiveResult(function=func)
    metrics = func.metrics

    # 비용 계산
    if metrics and metrics.invocations > 0:
        result.estimated_monthly_cost = get_lambda_monthly_cost(
            region=region,
            invocations=metrics.invocations,
            avg_duration_ms=metrics.duration_avg_ms,
            memory_mb=func.memory_mb,
        )

    # PC 비용 추가
    if func.provisioned_concurrency > 0:
        pc_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=func.provisioned_concurrency,
        )
        result.estimated_monthly_cost += pc_cost

    # 1. 런타임 EOL 분석
    _analyze_runtime_eol(func, result)

    # 2. 메모리 분석
    _analyze_memory(func, result, memory_stats)

    # 3. 에러율 분석
    _analyze_errors(func, result)

    # 4. Throttle 분석
    _analyze_throttles(func, result)

    # 5. 미사용 분석
    _analyze_usage(func, result)

    # 6. Timeout 위험 분석
    _analyze_timeout_risk(func, result)

    # 총 잠재 절감액
    result.potential_savings = sum(i.potential_savings for i in result.issues)

    return result


def _analyze_runtime_eol(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """런타임 EOL 분석"""
    runtime_info = get_runtime_info(func.runtime)
    if not runtime_info:
        return

    status = runtime_info.status

    if status == EOLStatus.DEPRECATED:
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.CRITICAL,
                description=f"런타임 지원 종료됨: {runtime_info.name}",
                recommendation=f"{upgrade}로 업그레이드 필요",
            )
        )
    elif status == EOLStatus.CRITICAL:
        days = runtime_info.days_until_deprecation
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.CRITICAL,
                description=f"런타임 {days}일 내 지원 종료: {runtime_info.name}",
                recommendation=f"{upgrade}로 즉시 업그레이드 권장",
            )
        )
    elif status == EOLStatus.HIGH:
        days = runtime_info.days_until_deprecation
        upgrade = get_recommended_upgrade(func.runtime) or "최신 버전"
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.HIGH,
                description=f"런타임 {days}일 내 지원 종료: {runtime_info.name}",
                recommendation=f"{upgrade}로 업그레이드 계획 수립",
            )
        )
    elif status == EOLStatus.MEDIUM:
        days = runtime_info.days_until_deprecation
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.RUNTIME_EOL,
                severity=Severity.MEDIUM,
                description=f"런타임 {days}일 내 지원 종료 예정: {runtime_info.name}",
                recommendation="업그레이드 계획 수립 권장",
            )
        )


def _analyze_memory(
    func: LambdaFunctionInfo,
    result: LambdaComprehensiveResult,
    memory_stats: dict | None = None,
):
    """메모리 사용량 분석"""
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    # 메모리 통계가 있으면 사용 (CloudWatch Logs Insights 결과)
    if memory_stats:
        max_used = memory_stats.get("max_memory_used_mb", 0)
        memory_stats.get("avg_memory_used_mb", 0)

        if max_used > 0:
            utilization = max_used / func.memory_mb * 100

            # 과다 할당 (최대 사용량이 50% 미만)
            if utilization < 50 and func.memory_mb > 128:
                recommended = max(128, int(max_used * 1.5))  # 50% 여유
                recommended = (recommended // 64) * 64  # 64MB 단위로 올림
                if recommended < func.memory_mb:
                    # 절감액 계산
                    current_cost = get_lambda_monthly_cost(
                        region=result.function.region,
                        invocations=metrics.invocations,
                        avg_duration_ms=metrics.duration_avg_ms,
                        memory_mb=func.memory_mb,
                    )
                    new_cost = get_lambda_monthly_cost(
                        region=result.function.region,
                        invocations=metrics.invocations,
                        avg_duration_ms=metrics.duration_avg_ms,
                        memory_mb=recommended,
                    )
                    savings = current_cost - new_cost

                    result.memory_recommendation = recommended
                    result.issues.append(
                        LambdaIssue(
                            issue_type=IssueType.MEMORY_OVERSIZED,
                            severity=Severity.MEDIUM,
                            description=f"메모리 과다 할당 (사용률 {utilization:.0f}%, {max_used:.0f}MB/{func.memory_mb}MB)",
                            recommendation=f"{recommended}MB로 축소 권장",
                            potential_savings=savings,
                        )
                    )

            # 부족 (최대 사용량이 90% 이상)
            elif utilization >= 90:
                recommended = int(max_used * 1.3)  # 30% 여유
                recommended = ((recommended // 64) + 1) * 64  # 64MB 단위로 올림
                result.memory_recommendation = recommended
                result.issues.append(
                    LambdaIssue(
                        issue_type=IssueType.MEMORY_UNDERSIZED,
                        severity=Severity.HIGH,
                        description=f"메모리 부족 위험 (사용률 {utilization:.0f}%)",
                        recommendation=f"{recommended}MB로 증가 권장 (OOM 방지)",
                    )
                )

        return

    # 메모리 통계가 없으면 휴리스틱 사용
    # Duration이 매우 높고 메모리가 낮으면 메모리 부족 가능성
    if metrics.duration_max_ms > 10000 and func.memory_mb <= 256:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.MEMORY_UNDERSIZED,
                severity=Severity.LOW,
                description=f"실행 시간 {metrics.duration_max_ms:.0f}ms, 메모리 {func.memory_mb}MB",
                recommendation="메모리 증가 시 성능 개선 가능성 검토",
            )
        )


def _analyze_errors(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """에러율 분석"""
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    error_rate = metrics.errors / metrics.invocations * 100

    if error_rate >= 10:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.CRITICAL,
                description=f"높은 에러율: {error_rate:.1f}% ({metrics.errors:,}/{metrics.invocations:,})",
                recommendation="에러 원인 분석 및 수정 필요",
            )
        )
    elif error_rate >= 5:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.HIGH,
                description=f"에러율: {error_rate:.1f}% ({metrics.errors:,}/{metrics.invocations:,})",
                recommendation="에러 로그 분석 권장",
            )
        )
    elif error_rate >= 1:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.HIGH_ERROR_RATE,
                severity=Severity.MEDIUM,
                description=f"에러율: {error_rate:.1f}%",
                recommendation="에러 모니터링 권장",
            )
        )


def _analyze_throttles(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """Throttle 분석"""
    metrics = func.metrics
    if not metrics or metrics.throttles == 0:
        return

    throttle_rate = metrics.throttles / (metrics.invocations + metrics.throttles) * 100

    if throttle_rate >= 5:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.THROTTLED,
                severity=Severity.HIGH,
                description=f"Throttle 발생: {metrics.throttles:,}회 ({throttle_rate:.1f}%)",
                recommendation="Reserved Concurrency 또는 Provisioned Concurrency 설정 검토",
            )
        )
    elif metrics.throttles >= 100:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.THROTTLED,
                severity=Severity.MEDIUM,
                description=f"Throttle 발생: {metrics.throttles:,}회",
                recommendation="동시성 설정 검토 권장",
            )
        )


def _analyze_usage(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """미사용 분석"""
    metrics = func.metrics
    if not metrics:
        return

    if metrics.invocations == 0:
        # PC 설정된 미사용 함수
        if func.provisioned_concurrency > 0:
            pc_cost = get_lambda_provisioned_monthly_cost(
                region=func.region,
                memory_mb=func.memory_mb,
                provisioned_concurrency=func.provisioned_concurrency,
            )
            result.issues.append(
                LambdaIssue(
                    issue_type=IssueType.UNUSED,
                    severity=Severity.CRITICAL,
                    description=f"미사용 함수 (PC {func.provisioned_concurrency}개 설정됨)",
                    recommendation="PC 해제 또는 함수 삭제",
                    potential_savings=pc_cost,
                )
            )
        else:
            result.issues.append(
                LambdaIssue(
                    issue_type=IssueType.UNUSED,
                    severity=Severity.MEDIUM,
                    description="30일간 호출 없음",
                    recommendation="필요 여부 검토 후 삭제 고려",
                )
            )


def _analyze_timeout_risk(func: LambdaFunctionInfo, result: LambdaComprehensiveResult):
    """Timeout 위험 분석"""
    metrics = func.metrics
    if not metrics or metrics.invocations == 0:
        return

    timeout_ms = func.timeout_seconds * 1000
    max_duration = metrics.duration_max_ms

    # 최대 실행 시간이 Timeout의 80% 이상
    if max_duration >= timeout_ms * 0.8:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.TIMEOUT_RISK,
                severity=Severity.HIGH,
                description=f"Timeout 위험: 최대 {max_duration:.0f}ms (Timeout: {timeout_ms}ms)",
                recommendation=f"Timeout 증가 권장 (현재 {func.timeout_seconds}초)",
            )
        )
    elif max_duration >= timeout_ms * 0.6:
        result.issues.append(
            LambdaIssue(
                issue_type=IssueType.TIMEOUT_RISK,
                severity=Severity.LOW,
                description=f"Timeout 여유 부족: 최대 {max_duration:.0f}ms (Timeout: {timeout_ms}ms)",
                recommendation="Timeout 설정 검토",
            )
        )


def analyze_comprehensive(
    functions: list[LambdaFunctionInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> ComprehensiveAnalysisResult:
    """Lambda 함수들 종합 분석"""
    result = ComprehensiveAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_functions=len(functions),
    )

    for func in functions:
        comp_result = analyze_function_comprehensive(func, region)
        result.results.append(comp_result)

        result.total_monthly_cost += comp_result.estimated_monthly_cost
        result.potential_savings += comp_result.potential_savings

        if comp_result.issue_count > 0:
            result.functions_with_issues += 1

        for issue in comp_result.issues:
            if issue.issue_type == IssueType.RUNTIME_EOL:
                result.runtime_eol_count += 1
            elif issue.issue_type in (
                IssueType.MEMORY_OVERSIZED,
                IssueType.MEMORY_UNDERSIZED,
            ):
                result.memory_issue_count += 1
            elif issue.issue_type == IssueType.HIGH_ERROR_RATE:
                result.error_issue_count += 1

    return result


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[ComprehensiveAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    severity_fills = {
        Severity.CRITICAL: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        Severity.HIGH: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        Severity.MEDIUM: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        Severity.LOW: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Lambda 종합 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = [
        "Account",
        "Region",
        "전체 함수",
        "이슈 함수",
        "런타임 EOL",
        "메모리 이슈",
        "에러 이슈",
        "월 비용",
        "절감 가능",
    ]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_functions)
        ws.cell(row=row, column=4, value=r.functions_with_issues)
        ws.cell(row=row, column=5, value=r.runtime_eol_count)
        ws.cell(row=row, column=6, value=r.memory_issue_count)
        ws.cell(row=row, column=7, value=r.error_issue_count)
        ws.cell(row=row, column=8, value=f"${r.total_monthly_cost:,.2f}")
        ws.cell(row=row, column=9, value=f"${r.potential_savings:,.2f}")

    # 총계
    total_functions = sum(r.total_functions for r in results)
    total_issues = sum(r.functions_with_issues for r in results)
    total_cost = sum(r.total_monthly_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)

    row += 2
    ws.cell(row=row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_functions).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_issues).font = Font(bold=True)
    ws.cell(row=row, column=8, value=f"${total_cost:,.2f}").font = Font(bold=True)
    ws.cell(row=row, column=9, value=f"${total_savings:,.2f}").font = Font(bold=True, color="FF0000")

    # Issues
    ws_issues = wb.create_sheet("Issues")
    issue_headers = [
        "Account",
        "Region",
        "Function",
        "Runtime",
        "Memory",
        "이슈 유형",
        "심각도",
        "설명",
        "권장 조치",
        "절감 가능",
    ]
    for col, h in enumerate(issue_headers, 1):
        ws_issues.cell(row=1, column=col, value=h).fill = header_fill
        ws_issues.cell(row=1, column=col).font = header_font

    issue_row = 1
    for r in results:
        for comp in r.results:
            for issue in comp.issues:
                issue_row += 1
                fn = comp.function
                ws_issues.cell(row=issue_row, column=1, value=fn.account_name)
                ws_issues.cell(row=issue_row, column=2, value=fn.region)
                ws_issues.cell(row=issue_row, column=3, value=fn.function_name)
                ws_issues.cell(row=issue_row, column=4, value=fn.runtime)
                ws_issues.cell(row=issue_row, column=5, value=fn.memory_mb)
                ws_issues.cell(row=issue_row, column=6, value=issue.issue_type.value)
                ws_issues.cell(row=issue_row, column=7, value=issue.severity.value)
                ws_issues.cell(row=issue_row, column=8, value=issue.description)
                ws_issues.cell(row=issue_row, column=9, value=issue.recommendation)
                ws_issues.cell(
                    row=issue_row,
                    column=10,
                    value=f"${issue.potential_savings:,.2f}" if issue.potential_savings > 0 else "-",
                )

                fill = severity_fills.get(issue.severity)
                if fill:
                    ws_issues.cell(row=issue_row, column=7).fill = fill

    # All Functions
    ws_all = wb.create_sheet("All Functions")
    all_headers = [
        "Account",
        "Region",
        "Function",
        "Runtime",
        "Memory",
        "Timeout",
        "Invocations (30d)",
        "Avg Duration",
        "Errors",
        "Throttles",
        "이슈 수",
        "월 비용",
    ]
    for col, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col, value=h).fill = header_fill
        ws_all.cell(row=1, column=col).font = header_font

    all_row = 1
    for r in results:
        for comp in r.results:
            all_row += 1
            fn = comp.function
            metrics = fn.metrics

            ws_all.cell(row=all_row, column=1, value=fn.account_name)
            ws_all.cell(row=all_row, column=2, value=fn.region)
            ws_all.cell(row=all_row, column=3, value=fn.function_name)
            ws_all.cell(row=all_row, column=4, value=fn.runtime)
            ws_all.cell(row=all_row, column=5, value=fn.memory_mb)
            ws_all.cell(row=all_row, column=6, value=fn.timeout_seconds)
            ws_all.cell(row=all_row, column=7, value=metrics.invocations if metrics else 0)
            ws_all.cell(
                row=all_row,
                column=8,
                value=f"{metrics.duration_avg_ms:.1f}ms" if metrics else "-",
            )
            ws_all.cell(row=all_row, column=9, value=metrics.errors if metrics else 0)
            ws_all.cell(row=all_row, column=10, value=metrics.throttles if metrics else 0)
            ws_all.cell(row=all_row, column=11, value=comp.issue_count)
            ws_all.cell(row=all_row, column=12, value=f"${comp.estimated_monthly_cost:,.4f}")

    # 열 너비
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Lambda_Comprehensive_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> ComprehensiveAnalysisResult:
    """단일 계정/리전의 Lambda 종합 분석 (병렬 실행용)"""
    functions = collect_functions_with_metrics(session, account_id, account_name, region)
    return analyze_comprehensive(functions, account_id, account_name, region)


def run(ctx) -> None:
    """Lambda 종합 분석 실행"""
    console.print("[bold]Lambda 종합 분석 시작...[/bold]\n")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")

    results: list[ComprehensiveAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_functions = sum(r.total_functions for r in results)
    total_issues = sum(r.functions_with_issues for r in results)
    total_runtime = sum(r.runtime_eol_count for r in results)
    total_memory = sum(r.memory_issue_count for r in results)
    total_error = sum(r.error_issue_count for r in results)
    total_cost = sum(r.total_monthly_cost for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 Lambda 함수: {total_functions}개")
    console.print(f"이슈 함수: {total_issues}개")
    if total_runtime > 0:
        console.print(f"  [red]런타임 EOL: {total_runtime}개[/red]")
    if total_memory > 0:
        console.print(f"  [yellow]메모리 이슈: {total_memory}개[/yellow]")
    if total_error > 0:
        console.print(f"  [yellow]에러 이슈: {total_error}개[/yellow]")
    console.print(f"총 월간 비용: ${total_cost:,.2f}")
    if total_savings > 0:
        console.print(f"[green]절감 가능: ${total_savings:,.2f}[/green]")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("lambda-comprehensive").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
