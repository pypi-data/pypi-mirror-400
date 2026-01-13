"""
plugins/sns/unused.py - SNS 미사용 토픽 분석

유휴/미사용 SNS 토픽 탐지 (구독자 및 CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 미사용 기준: 7일간 메시지 발행 0
UNUSED_DAYS_THRESHOLD = 7

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "sns:ListTopics",
        "sns:ListSubscriptionsByTopic",
        "cloudwatch:GetMetricStatistics",
    ],
}


class TopicStatus(Enum):
    """토픽 상태"""

    NORMAL = "normal"
    NO_SUBSCRIBERS = "no_subscribers"
    NO_MESSAGES = "no_messages"
    UNUSED = "unused"


@dataclass
class SNSTopicInfo:
    """SNS 토픽 정보"""

    account_id: str
    account_name: str
    region: str
    topic_name: str
    topic_arn: str
    subscription_count: int
    # CloudWatch 지표
    messages_published: float = 0.0
    notifications_delivered: float = 0.0
    notifications_failed: float = 0.0


@dataclass
class TopicFinding:
    """토픽 분석 결과"""

    topic: SNSTopicInfo
    status: TopicStatus
    recommendation: str


@dataclass
class SNSAnalysisResult:
    """SNS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_topics: int = 0
    no_subscribers: int = 0
    no_messages: int = 0
    unused_topics: int = 0
    normal_topics: int = 0
    findings: list[TopicFinding] = field(default_factory=list)


def collect_sns_topics(session, account_id: str, account_name: str, region: str) -> list[SNSTopicInfo]:
    """SNS 토픽 수집"""
    from botocore.exceptions import ClientError

    sns = get_client(session, "sns", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    topics = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = sns.get_paginator("list_topics")
        for page in paginator.paginate():
            for topic in page.get("Topics", []):
                topic_arn = topic.get("TopicArn", "")
                topic_name = topic_arn.split(":")[-1]

                # 구독자 수 확인
                subscription_count = 0
                try:
                    sub_paginator = sns.get_paginator("list_subscriptions_by_topic")
                    for sub_page in sub_paginator.paginate(TopicArn=topic_arn):
                        subscription_count += len(sub_page.get("Subscriptions", []))
                except ClientError:
                    pass

                info = SNSTopicInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    topic_name=topic_name,
                    topic_arn=topic_arn,
                    subscription_count=subscription_count,
                )

                # CloudWatch 지표 조회
                try:
                    # NumberOfMessagesPublished
                    pub_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfMessagesPublished",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if pub_resp.get("Datapoints"):
                        info.messages_published = sum(d["Sum"] for d in pub_resp["Datapoints"])

                    # NumberOfNotificationsDelivered
                    del_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfNotificationsDelivered",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if del_resp.get("Datapoints"):
                        info.notifications_delivered = sum(d["Sum"] for d in del_resp["Datapoints"])

                    # NumberOfNotificationsFailed
                    fail_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/SNS",
                        MetricName="NumberOfNotificationsFailed",
                        Dimensions=[{"Name": "TopicName", "Value": topic_name}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if fail_resp.get("Datapoints"):
                        info.notifications_failed = sum(d["Sum"] for d in fail_resp["Datapoints"])

                except ClientError:
                    pass

                topics.append(info)

    except ClientError:
        pass

    return topics


def analyze_topics(topics: list[SNSTopicInfo], account_id: str, account_name: str, region: str) -> SNSAnalysisResult:
    """SNS 토픽 분석"""
    result = SNSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_topics=len(topics),
    )

    for topic in topics:
        # 구독자 없고 메시지도 없음 = 미사용
        if topic.subscription_count == 0 and topic.messages_published == 0:
            result.unused_topics += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.UNUSED,
                    recommendation="구독자 없음 + 메시지 없음 - 삭제 검토",
                )
            )
            continue

        # 구독자 없음
        if topic.subscription_count == 0:
            result.no_subscribers += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.NO_SUBSCRIBERS,
                    recommendation="구독자 없음 - 구독 추가 또는 삭제 검토",
                )
            )
            continue

        # 메시지 발행 없음
        if topic.messages_published == 0:
            result.no_messages += 1
            result.findings.append(
                TopicFinding(
                    topic=topic,
                    status=TopicStatus.NO_MESSAGES,
                    recommendation="메시지 발행 없음 - 사용 여부 확인",
                )
            )
            continue

        result.normal_topics += 1
        result.findings.append(
            TopicFinding(
                topic=topic,
                status=TopicStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SNSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "SNS 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "미사용", "구독자없음", "메시지없음", "정상"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_topics)
        ws.cell(row=row, column=4, value=r.unused_topics)
        ws.cell(row=row, column=5, value=r.no_subscribers)
        ws.cell(row=row, column=6, value=r.no_messages)
        ws.cell(row=row, column=7, value=r.normal_topics)
        if r.unused_topics > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.no_subscribers > 0:
            ws.cell(row=row, column=5).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Topics")
    detail_headers = [
        "Account",
        "Region",
        "Topic Name",
        "Subscribers",
        "상태",
        "Published",
        "Delivered",
        "Failed",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != TopicStatus.NORMAL:
                detail_row += 1
                t = f.topic
                ws_detail.cell(row=detail_row, column=1, value=t.account_name)
                ws_detail.cell(row=detail_row, column=2, value=t.region)
                ws_detail.cell(row=detail_row, column=3, value=t.topic_name)
                ws_detail.cell(row=detail_row, column=4, value=t.subscription_count)
                ws_detail.cell(row=detail_row, column=5, value=f.status.value)
                ws_detail.cell(row=detail_row, column=6, value=int(t.messages_published))
                ws_detail.cell(row=detail_row, column=7, value=int(t.notifications_delivered))
                ws_detail.cell(row=detail_row, column=8, value=int(t.notifications_failed))
                ws_detail.cell(row=detail_row, column=9, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"SNS_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SNSAnalysisResult | None:
    """단일 계정/리전의 SNS 토픽 수집 및 분석 (병렬 실행용)"""
    topics = collect_sns_topics(session, account_id, account_name, region)
    if not topics:
        return None
    return analyze_topics(topics, account_id, account_name, region)


def run(ctx) -> None:
    """SNS 미사용 토픽 분석"""
    console.print("[bold]SNS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="sns")
    results: list[SNSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.unused_topics for r in results)
    total_no_sub = sum(r.no_subscribers for r in results)
    total_no_msg = sum(r.no_messages for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(
        f"미사용: [red]{total_unused}개[/red] / "
        f"구독자없음: [yellow]{total_no_sub}개[/yellow] / "
        f"메시지없음: {total_no_msg}개"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("sns-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
