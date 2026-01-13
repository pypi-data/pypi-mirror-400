"""
plugins/ec2/ebs_audit.py - EBS 미사용 분석

미사용 EBS 볼륨 탐지 및 비용 절감 기회 식별

분석 기준:
- Status가 "available"인 볼륨 (아무 인스턴스에도 연결되지 않음)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any

from rich.console import Console

from core.parallel import get_client, is_quiet, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_ebs_monthly_cost

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "ec2:DescribeVolumes",
    ],
}


# =============================================================================
# 데이터 구조
# =============================================================================


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 미사용 (available 상태)
    NORMAL = "normal"  # 정상 사용 (in-use)
    PENDING = "pending"  # 확인 필요


class Severity(Enum):
    """심각도"""

    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"
    INFO = "info"


@dataclass
class EBSInfo:
    """EBS 볼륨 정보"""

    id: str
    name: str
    state: str
    volume_type: str
    size_gb: int
    iops: int
    throughput: int
    encrypted: bool
    kms_key_id: str
    availability_zone: str
    create_time: datetime
    snapshot_id: str
    attachments: list[dict[str, Any]]
    tags: dict[str, str]

    # 메타
    account_id: str
    account_name: str
    region: str

    # 비용
    monthly_cost: float = 0.0

    @property
    def is_attached(self) -> bool:
        """연결 여부"""
        return self.state == "in-use" and len(self.attachments) > 0

    @property
    def attached_instance_id(self) -> str:
        """연결된 인스턴스 ID"""
        if self.attachments:
            return str(self.attachments[0].get("InstanceId", ""))
        return ""


@dataclass
class EBSFinding:
    """EBS 분석 결과"""

    volume: EBSInfo
    usage_status: UsageStatus
    severity: Severity
    description: str
    recommendation: str


@dataclass
class EBSAnalysisResult:
    """분석 결과"""

    account_id: str
    account_name: str
    region: str
    findings: list[EBSFinding] = field(default_factory=list)

    # 통계
    total_count: int = 0
    unused_count: int = 0
    normal_count: int = 0
    pending_count: int = 0

    # 비용
    total_size_gb: int = 0
    unused_size_gb: int = 0
    unused_monthly_cost: float = 0.0


# =============================================================================
# 수집
# =============================================================================


def collect_ebs(session, account_id: str, account_name: str, region: str) -> list[EBSInfo]:
    """EBS 볼륨 목록 수집"""
    from botocore.exceptions import ClientError

    volumes = []

    try:
        ec2 = get_client(session, "ec2", region_name=region)
        paginator = ec2.get_paginator("describe_volumes")

        for page in paginator.paginate():
            for data in page.get("Volumes", []):
                # 태그 파싱
                tags = {
                    t.get("Key", ""): t.get("Value", "")
                    for t in data.get("Tags", [])
                    if not t.get("Key", "").startswith("aws:")
                }

                # 월간 비용 계산
                volume_type = data.get("VolumeType", "")
                size_gb = data.get("Size", 0)
                monthly_cost = get_ebs_monthly_cost(volume_type, size_gb, region)

                volume = EBSInfo(
                    id=data.get("VolumeId", ""),
                    name=tags.get("Name", ""),
                    state=data.get("State", ""),
                    volume_type=volume_type,
                    size_gb=size_gb,
                    iops=data.get("Iops", 0),
                    throughput=data.get("Throughput", 0),
                    encrypted=data.get("Encrypted", False),
                    kms_key_id=data.get("KmsKeyId", ""),
                    availability_zone=data.get("AvailabilityZone", ""),
                    create_time=data.get("CreateTime"),
                    snapshot_id=data.get("SnapshotId", ""),
                    attachments=data.get("Attachments", []),
                    tags=tags,
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    monthly_cost=monthly_cost,
                )
                volumes.append(volume)

    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "Unknown")
        if not is_quiet():
            console.print(f"    [yellow]{account_name}/{region} EBS 수집 오류: {error_code}[/yellow]")

    return volumes


# =============================================================================
# 분석
# =============================================================================


def analyze_ebs(volumes: list[EBSInfo], account_id: str, account_name: str, region: str) -> EBSAnalysisResult:
    """EBS 미사용 분석"""
    result = EBSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
    )

    for volume in volumes:
        finding = _analyze_single_volume(volume)
        result.findings.append(finding)
        result.total_size_gb += volume.size_gb

        if finding.usage_status == UsageStatus.UNUSED:
            result.unused_count += 1
            result.unused_size_gb += volume.size_gb
            result.unused_monthly_cost += volume.monthly_cost
        elif finding.usage_status == UsageStatus.NORMAL:
            result.normal_count += 1
        elif finding.usage_status == UsageStatus.PENDING:
            result.pending_count += 1

    result.total_count = len(volumes)
    return result


def _analyze_single_volume(volume: EBSInfo) -> EBSFinding:
    """개별 볼륨 분석"""

    # 1. 연결됨 = 정상
    if volume.is_attached:
        return EBSFinding(
            volume=volume,
            usage_status=UsageStatus.NORMAL,
            severity=Severity.INFO,
            description=f"사용 중 (인스턴스: {volume.attached_instance_id})",
            recommendation="정상 사용 중",
        )

    # 2. Available = 미사용
    if volume.state == "available":
        # 용량에 따른 심각도
        if volume.size_gb >= 500:
            severity = Severity.HIGH
        elif volume.size_gb >= 100:
            severity = Severity.MEDIUM
        else:
            severity = Severity.LOW

        cost_info = f"월 ${volume.monthly_cost:.2f}" if volume.monthly_cost > 0 else ""
        return EBSFinding(
            volume=volume,
            usage_status=UsageStatus.UNUSED,
            severity=severity,
            description=f"미사용 볼륨 ({volume.size_gb}GB, {volume.volume_type}) {cost_info}",
            recommendation="스냅샷 생성 후 삭제 검토",
        )

    # 3. 기타 (creating, deleting, error 등)
    return EBSFinding(
        volume=volume,
        usage_status=UsageStatus.PENDING,
        severity=Severity.INFO,
        description=f"상태: {volume.state}",
        recommendation="상태 안정화 대기",
    )


# =============================================================================
# Excel 보고서
# =============================================================================


def generate_report(results: list[EBSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        UsageStatus.PENDING: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    }

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "EBS 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    totals = {
        "total": sum(r.total_count for r in results),
        "unused": sum(r.unused_count for r in results),
        "normal": sum(r.normal_count for r in results),
        "pending": sum(r.pending_count for r in results),
        "total_size": sum(r.total_size_gb for r in results),
        "unused_size": sum(r.unused_size_gb for r in results),
        "unused_cost": sum(r.unused_monthly_cost for r in results),
    }

    stats = [
        ("항목", "값"),
        ("전체 볼륨", totals["total"]),
        ("미사용", totals["unused"]),
        ("정상 사용", totals["normal"]),
        ("확인 필요", totals["pending"]),
        ("전체 용량 (GB)", totals["total_size"]),
        ("미사용 용량 (GB)", totals["unused_size"]),
        ("미사용 월 비용 ($)", f"${totals['unused_cost']:.2f}"),
    ]

    for i, (item, value) in enumerate(stats):
        row = 4 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=value)
        if i == 0:
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).font = header_font

    # Findings
    ws2 = wb.create_sheet("Findings")
    headers = [
        "Account",
        "Region",
        "Volume ID",
        "Name",
        "State",
        "Usage",
        "Severity",
        "Type",
        "Size (GB)",
        "Monthly Cost ($)",
        "IOPS",
        "Encrypted",
        "AZ",
        "Created",
        "Description",
        "Recommendation",
    ]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # 미사용/확인필요만 표시
    all_findings = []
    for result in results:
        for f in result.findings:
            if f.usage_status in (UsageStatus.UNUSED, UsageStatus.PENDING):
                all_findings.append(f)

    # 비용순 정렬
    all_findings.sort(key=lambda x: x.volume.monthly_cost, reverse=True)

    for f in all_findings:
        vol = f.volume
        ws2.append(
            [
                vol.account_name,
                vol.region,
                vol.id,
                vol.name,
                vol.state,
                f.usage_status.value,
                f.severity.value,
                vol.volume_type,
                vol.size_gb,
                round(vol.monthly_cost, 2),
                vol.iops,
                "Yes" if vol.encrypted else "No",
                vol.availability_zone,
                vol.create_time.strftime("%Y-%m-%d") if vol.create_time else "",
                f.description,
                f.recommendation,
            ]
        )

        fill = status_fills.get(f.usage_status)
        if fill:
            ws2.cell(row=ws2.max_row, column=6).fill = fill

    # 열 너비
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 40)

    ws2.freeze_panes = "A2"

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"EBS_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 메인
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EBSAnalysisResult:
    """단일 계정/리전의 EBS 수집 및 분석 (병렬 실행용)"""
    volumes = collect_ebs(session, account_id, account_name, region)
    return analyze_ebs(volumes, account_id, account_name, region)


def run(ctx) -> None:
    """EBS 미사용 분석 실행 (병렬 처리)"""
    console.print("[bold]EBS 미사용 분석 시작...[/bold]")

    # 병렬 수집
    result = parallel_collect(
        ctx,
        _collect_and_analyze,
        max_workers=20,
        service="ec2",
    )

    # 결과 처리
    all_results: list[EBSAnalysisResult] = result.get_data()

    # 진행 상황 출력
    console.print(f"  [dim]수집 완료: 성공 {result.success_count}, 실패 {result.error_count}[/dim]")

    # 에러 요약
    if result.error_count > 0:
        console.print(f"\n[yellow]{result.get_error_summary()}[/yellow]")

    if not all_results:
        console.print("[yellow]분석할 EBS 없음[/yellow]")
        return

    # 개별 결과 요약
    for r in all_results:
        if r.unused_count > 0:
            cost_str = f" (${r.unused_monthly_cost:.2f}/월)" if r.unused_monthly_cost > 0 else ""
            console.print(
                f"  {r.account_name}/{r.region}: [red]미사용 {r.unused_count}개 ({r.unused_size_gb}GB){cost_str}[/red]"
            )
        elif r.pending_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [yellow]확인 필요 {r.pending_count}개[/yellow]")
        elif r.total_count > 0:
            console.print(f"  {r.account_name}/{r.region}: [green]정상 {r.normal_count}개[/green]")

    # 전체 통계
    totals = {
        "total": sum(r.total_count for r in all_results),
        "unused": sum(r.unused_count for r in all_results),
        "normal": sum(r.normal_count for r in all_results),
        "pending": sum(r.pending_count for r in all_results),
        "total_size": sum(r.total_size_gb for r in all_results),
        "unused_size": sum(r.unused_size_gb for r in all_results),
        "unused_cost": sum(r.unused_monthly_cost for r in all_results),
    }

    console.print(f"\n[bold]전체 EBS: {totals['total']}개 ({totals['total_size']}GB)[/bold]")
    if totals["unused"] > 0:
        console.print(
            f"  [red bold]미사용: {totals['unused']}개 ({totals['unused_size']}GB, ${totals['unused_cost']:.2f}/월)[/red bold]"
        )
    if totals["pending"] > 0:
        console.print(f"  [yellow]확인 필요: {totals['pending']}개[/yellow]")
    console.print(f"  [green]정상: {totals['normal']}개[/green]")

    # 보고서
    console.print("\n[cyan]Excel 보고서 생성 중...[/cyan]")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("ebs-audit").with_date().build()
    filepath = generate_report(all_results, output_path)

    console.print(f"[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
