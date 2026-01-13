"""
plugins/fn/unused.py - 미사용 Lambda 함수 분석

30일 이상 호출되지 않은 Lambda 함수 탐지

분석 기준:
- 30일간 Invocations 메트릭이 0인 함수
- Provisioned Concurrency가 설정된 미사용 함수 (비용 낭비)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import (
    get_lambda_monthly_cost,
    get_lambda_provisioned_monthly_cost,
)

from .common.collector import LambdaFunctionInfo, collect_functions_with_metrics

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "lambda:ListFunctions",
        "lambda:GetFunction",
        "lambda:ListProvisionedConcurrencyConfigs",
        "cloudwatch:GetMetricStatistics",
    ],
}


class UsageStatus(Enum):
    """사용 상태"""

    UNUSED = "unused"  # 30일간 호출 없음
    UNUSED_PROVISIONED = "unused_provisioned"  # 미사용 + PC 설정됨
    LOW_USAGE = "low_usage"  # 저사용 (월 100회 미만)
    NORMAL = "normal"  # 정상 사용


@dataclass
class LambdaFinding:
    """Lambda 분석 결과"""

    function: LambdaFunctionInfo
    status: UsageStatus
    recommendation: str
    monthly_waste: float = 0.0


@dataclass
class LambdaAnalysisResult:
    """Lambda 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unused_count: int = 0
    low_usage_count: int = 0
    normal_count: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[LambdaFinding] = field(default_factory=list)


# =============================================================================
# 분석
# =============================================================================


def analyze_functions(
    functions: list[LambdaFunctionInfo],
    account_id: str,
    account_name: str,
    region: str,
) -> LambdaAnalysisResult:
    """Lambda 함수 미사용 분석"""
    result = LambdaAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(functions),
    )

    for func in functions:
        finding = _analyze_single_function(func, region)
        result.findings.append(finding)

        if finding.status == UsageStatus.UNUSED or finding.status == UsageStatus.UNUSED_PROVISIONED:
            result.unused_count += 1
            result.unused_monthly_cost += finding.monthly_waste
        elif finding.status == UsageStatus.LOW_USAGE:
            result.low_usage_count += 1
        else:
            result.normal_count += 1

    return result


def _analyze_single_function(func: LambdaFunctionInfo, region: str) -> LambdaFinding:
    """개별 Lambda 함수 분석"""

    # 메트릭이 없으면 알 수 없음
    if func.metrics is None:
        return LambdaFinding(
            function=func,
            status=UsageStatus.NORMAL,
            recommendation="메트릭 조회 실패",
        )

    invocations = func.metrics.invocations

    # 미사용 (30일간 호출 없음)
    if invocations == 0:
        # Provisioned Concurrency가 있으면 더 심각
        if func.provisioned_concurrency > 0:
            waste = get_lambda_provisioned_monthly_cost(
                region=region,
                memory_mb=func.memory_mb,
                provisioned_concurrency=func.provisioned_concurrency,
            )
            return LambdaFinding(
                function=func,
                status=UsageStatus.UNUSED_PROVISIONED,
                recommendation=f"미사용 + PC {func.provisioned_concurrency}개 설정됨 - 즉시 삭제 또는 PC 해제 권장",
                monthly_waste=waste,
            )

        return LambdaFinding(
            function=func,
            status=UsageStatus.UNUSED,
            recommendation="30일간 호출 없음 - 삭제 또는 비활성화 검토",
            monthly_waste=0.0,  # 미호출 시 비용 없음 (PC 제외)
        )

    # 저사용 (월 100회 미만)
    if invocations < 100:
        return LambdaFinding(
            function=func,
            status=UsageStatus.LOW_USAGE,
            recommendation=f"저사용 (30일간 {invocations}회) - 통합 또는 삭제 검토",
        )

    # 정상 사용
    estimated_cost = get_lambda_monthly_cost(
        region=region,
        invocations=invocations,
        avg_duration_ms=func.metrics.duration_avg_ms,
        memory_mb=func.memory_mb,
    )

    # PC가 있으면 비용 추가
    if func.provisioned_concurrency > 0:
        pc_cost = get_lambda_provisioned_monthly_cost(
            region=region,
            memory_mb=func.memory_mb,
            provisioned_concurrency=func.provisioned_concurrency,
        )
        estimated_cost += pc_cost

    func.estimated_monthly_cost = estimated_cost

    return LambdaFinding(
        function=func,
        status=UsageStatus.NORMAL,
        recommendation=f"정상 사용 (30일간 {invocations:,}회)",
    )


# =============================================================================
# 보고서
# =============================================================================


def generate_report(results: list[LambdaAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Lambda 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    headers = ["Account", "Region", "전체", "미사용", "저사용", "정상", "월간 낭비"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_count)
        ws.cell(row=row, column=4, value=r.unused_count)
        ws.cell(row=row, column=5, value=r.low_usage_count)
        ws.cell(row=row, column=6, value=r.normal_count)
        ws.cell(row=row, column=7, value=f"${r.unused_monthly_cost:,.2f}")
        if r.unused_count > 0:
            ws.cell(row=row, column=4).fill = red_fill

    # 총계
    total_functions = sum(r.total_count for r in results)
    total_unused = sum(r.unused_count for r in results)
    total_waste = sum(r.unused_monthly_cost for r in results)

    row += 2
    ws.cell(row=row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=row, column=3, value=total_functions).font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_unused).font = Font(bold=True)
    ws.cell(row=row, column=7, value=f"${total_waste:,.2f}").font = Font(bold=True, color="FF0000")

    # Unused Functions
    ws_unused = wb.create_sheet("Unused")
    unused_headers = [
        "Account",
        "Region",
        "Function Name",
        "Runtime",
        "Memory (MB)",
        "Timeout (s)",
        "Code Size (MB)",
        "Last Modified",
        "Provisioned Concurrency",
        "상태",
        "월간 낭비",
        "권장 조치",
    ]
    for col, h in enumerate(unused_headers, 1):
        ws_unused.cell(row=1, column=col, value=h).fill = header_fill
        ws_unused.cell(row=1, column=col).font = header_font

    unused_row = 1
    for r in results:
        for f in r.findings:
            if f.status in (
                UsageStatus.UNUSED,
                UsageStatus.UNUSED_PROVISIONED,
                UsageStatus.LOW_USAGE,
            ):
                unused_row += 1
                fn = f.function
                ws_unused.cell(row=unused_row, column=1, value=fn.account_name)
                ws_unused.cell(row=unused_row, column=2, value=fn.region)
                ws_unused.cell(row=unused_row, column=3, value=fn.function_name)
                ws_unused.cell(row=unused_row, column=4, value=fn.runtime)
                ws_unused.cell(row=unused_row, column=5, value=fn.memory_mb)
                ws_unused.cell(row=unused_row, column=6, value=fn.timeout_seconds)
                ws_unused.cell(row=unused_row, column=7, value=round(fn.code_size_mb, 2))
                ws_unused.cell(
                    row=unused_row,
                    column=8,
                    value=fn.last_modified.strftime("%Y-%m-%d") if fn.last_modified else "-",
                )
                ws_unused.cell(row=unused_row, column=9, value=fn.provisioned_concurrency or "-")
                ws_unused.cell(row=unused_row, column=10, value=f.status.value)
                ws_unused.cell(
                    row=unused_row,
                    column=11,
                    value=f"${f.monthly_waste:,.2f}" if f.monthly_waste > 0 else "-",
                )
                ws_unused.cell(row=unused_row, column=12, value=f.recommendation)

                # 상태별 색상
                if f.status == UsageStatus.UNUSED_PROVISIONED or f.status == UsageStatus.UNUSED:
                    ws_unused.cell(row=unused_row, column=10).fill = red_fill
                elif f.status == UsageStatus.LOW_USAGE:
                    ws_unused.cell(row=unused_row, column=10).fill = yellow_fill

    # All Functions
    ws_all = wb.create_sheet("All Functions")
    all_headers = [
        "Account",
        "Region",
        "Function Name",
        "Runtime",
        "Memory (MB)",
        "Timeout (s)",
        "Code Size (MB)",
        "Invocations (30d)",
        "Avg Duration (ms)",
        "Errors",
        "Throttles",
        "PC",
        "Reserved",
        "상태",
        "추정 월 비용",
    ]
    for col, h in enumerate(all_headers, 1):
        ws_all.cell(row=1, column=col, value=h).fill = header_fill
        ws_all.cell(row=1, column=col).font = header_font

    all_row = 1
    for r in results:
        for f in r.findings:
            all_row += 1
            fn = f.function
            metrics = fn.metrics

            ws_all.cell(row=all_row, column=1, value=fn.account_name)
            ws_all.cell(row=all_row, column=2, value=fn.region)
            ws_all.cell(row=all_row, column=3, value=fn.function_name)
            ws_all.cell(row=all_row, column=4, value=fn.runtime)
            ws_all.cell(row=all_row, column=5, value=fn.memory_mb)
            ws_all.cell(row=all_row, column=6, value=fn.timeout_seconds)
            ws_all.cell(row=all_row, column=7, value=round(fn.code_size_mb, 2))
            ws_all.cell(row=all_row, column=8, value=metrics.invocations if metrics else 0)
            ws_all.cell(
                row=all_row,
                column=9,
                value=round(metrics.duration_avg_ms, 2) if metrics else 0,
            )
            ws_all.cell(row=all_row, column=10, value=metrics.errors if metrics else 0)
            ws_all.cell(row=all_row, column=11, value=metrics.throttles if metrics else 0)
            ws_all.cell(row=all_row, column=12, value=fn.provisioned_concurrency or "-")
            ws_all.cell(
                row=all_row,
                column=13,
                value=fn.reserved_concurrency if fn.reserved_concurrency is not None else "-",
            )
            ws_all.cell(row=all_row, column=14, value=f.status.value)
            ws_all.cell(
                row=all_row,
                column=15,
                value=f"${fn.estimated_monthly_cost:,.4f}" if fn.estimated_monthly_cost > 0 else "-",
            )

    # 열 너비
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        if sheet.title != "Summary":
            sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Lambda_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)

    return filepath


# =============================================================================
# 실행
# =============================================================================


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> LambdaAnalysisResult:
    """단일 계정/리전의 Lambda 수집 및 분석 (병렬 실행용)"""
    functions = collect_functions_with_metrics(session, account_id, account_name, region)
    return analyze_functions(functions, account_id, account_name, region)


def run(ctx) -> None:
    """Lambda 미사용 분석 실행"""
    console.print("[bold]Lambda 미사용 분석 시작...[/bold]\n")

    # 병렬 수집 및 분석
    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="lambda")

    results: list[LambdaAnalysisResult] = result.get_data()

    # 에러 출력
    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")
        console.print(f"[dim]{result.get_error_summary()}[/dim]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    # 요약
    total_functions = sum(r.total_count for r in results)
    total_unused = sum(r.unused_count for r in results)
    total_low = sum(r.low_usage_count for r in results)
    total_waste = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 Lambda 함수: {total_functions}개")
    if total_unused > 0:
        console.print(f"[red]미사용: {total_unused}개[/red]")
    if total_low > 0:
        console.print(f"[yellow]저사용: {total_low}개[/yellow]")
    if total_waste > 0:
        console.print(f"[red]월간 낭비 비용: ${total_waste:,.2f}[/red]")

    # 보고서
    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("lambda-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
