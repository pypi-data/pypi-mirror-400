import os
import requests
from google import genai
from dotenv import load_dotenv
from datetime import datetime

load_dotenv()

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
GOOGLE_MODEL = os.getenv("GOOGLE_API_MODEL", "gemini-2.5-flash")
OLLAMA_URL = os.getenv("OLLAMA_API_URL", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.2:latest")

google_client = None
if GOOGLE_API_KEY:
    try:
        google_client = genai.Client(api_key=GOOGLE_API_KEY)
        print("[FileAnalyzer] ✓ Google client initialized")
    except Exception as e:
        print(f"[FileAnalyzer] ✗ Google Init ERROR: {e}")


class FileAnalyzer:

    @staticmethod
    def extract_file_metadata(file_path: str, filename: str) -> dict:
        """Extract basic file metadata"""
        try:
            stat = os.stat(file_path)
            return {
                "filename": filename,
                "size_bytes": stat.st_size,
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "extension": os.path.splitext(filename)[1].lower()
            }
        except Exception as e:
            return {
                "filename": filename,
                "error": str(e)
            }

    @staticmethod
    def get_file_preview(file_path: str, max_chars: int = 1000) -> str:
        """Get a text preview of the file"""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            
            # Text files
            if ext in [".txt", ".md"]:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read(max_chars)
            
            # PDF - try to extract text
            if ext == ".pdf":
                try:
                    import PyPDF2
                    with open(file_path, "rb") as f:
                        reader = PyPDF2.PdfReader(f)
                        text = ""
                        for page in reader.pages[:3]:  # first 3 pages
                            text += page.extract_text()
                        return text[:max_chars]
                except:
                    return "[PDF file - text extraction not available]"
            
            # DOCX
            if ext == ".docx":
                try:
                    from docx import Document
                    doc = Document(file_path)
                    text = "\n".join([p.text for p in doc.paragraphs])
                    return text[:max_chars]
                except:
                    return "[DOCX file - text extraction not available]"
            
            # XLSX
            if ext == ".xlsx":
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    text = ""
                    for sheet in wb.sheetnames[:2]:  # first 2 sheets
                        ws = wb[sheet]
                        text += f"Sheet: {sheet}\n"
                        for row in list(ws.iter_rows(max_row=10, values_only=True)):
                            text += str(row) + "\n"
                    return text[:max_chars]
                except:
                    return "[XLSX file - text extraction not available]"
            
            # Images
            if ext in [".png", ".jpg", ".jpeg"]:
                return "[Image file - use Gemini Vision to extract content]"
            
            return "[File type not supported for preview]"
        
        except Exception as e:
            return f"[Error reading file: {str(e)}]"

    @staticmethod
    def analyze_with_google(file_path: str, file_content: str) -> dict:
        """Analyze file content with Google Gemini"""
        if google_client is None:
            return {
                "source": "google",
                "status": "error",
                "error": "Google API not initialized"
            }

        try:
            prompt = f"""
Analyze this document and return structured information:

1. **Summary** (3–5 sentences)
2. **Key Topics** (5–7 bullet points)
3. **Document Type** (e.g., finance, legal, HR, invoice, report, etc.)
4. **Confidence** (High/Medium/Low)
5. **Keywords** (up to 10 important words)

Document name: {os.path.basename(file_path)}

Content preview:
{file_content[:2000]}
"""

            response = google_client.models.generate_content(
                model=GOOGLE_MODEL,
                contents=prompt
            )

            return {
                "source": "google",
                "analysis": response.text if response else "No response",
                "status": "success"
            }

        except Exception as e:
            print(f"[Google Analysis Error] {e}")
            return {
                "source": "google",
                "status": "error",
                "error": str(e)
            }

    @staticmethod
    def analyze_with_ollama(file_path: str, file_content: str) -> dict:
        """Analyze file content with Ollama"""
        try:
            prompt = f"""
Analyze this document and provide:

1. Summary (3-5 sentences)
2. Key Topics (5-7 bullet points)
3. Document Type (finance, legal, HR, invoice, etc.)
4. Confidence (High/Medium/Low)
5. Keywords (10 important words)

Document name: {os.path.basename(file_path)}

Content:
{file_content[:2000]}
"""

            r = requests.post(
                f"{OLLAMA_URL}/api/generate",
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.3
                    }
                },
                timeout=60
            )

            if r.status_code == 200:
                return {
                    "source": "ollama",
                    "model": OLLAMA_MODEL,
                    "analysis": r.json().get("response", "No response"),
                    "status": "success"
                }

            return {
                "source": "ollama",
                "status": "error",
                "error": f"HTTP {r.status_code}: {r.text}"
            }

        except requests.exceptions.ConnectionError:
            return {
                "source": "ollama",
                "status": "error",
                "error": "Could not connect to Ollama. Make sure Ollama is running on http://localhost:11434"
            }
        except Exception as e:
            print(f"[Ollama Analysis Error] {e}")
            return {
                "source": "ollama",
                "status": "error",
                "error": str(e)
            }
    
    @staticmethod
    def analyze_image_with_google(filepath):
        """Analyze image using Google Gemini Vision - WITH DEBUG LOGGING"""
        print(f"\n{'='*60}")
        print(f"[IMAGE ANALYSIS] Starting analysis for: {filepath}")
        print(f"{'='*60}")
        
        try:
            # Import required libraries
            print("[1/6] Importing libraries...")
            import google.generativeai as genai
            from PIL import Image
            print("     ✓ Libraries imported")
            
            # Check API key
            print("[2/6] Checking API key...")
            google_api_key = os.getenv("GOOGLE_API_KEY", "")
            if not google_api_key:
                print("     ✗ API key not found")
                return {
                    "status": "error",
                    "error": "Google API key not configured",
                    "source": "google"
                }
            print(f"     ✓ API key found: {google_api_key[:10]}...{google_api_key[-4:]}")
            
            # Configure Gemini
            print("[3/6] Configuring Gemini...")
            genai.configure(api_key=google_api_key)
            print("     ✓ Gemini configured")
            
            # Open image
            print(f"[4/6] Opening image: {filepath}")
            if not os.path.exists(filepath):
                print(f"     ✗ File not found: {filepath}")
                return {
                    "status": "error",
                    "error": f"File not found: {filepath}",
                    "source": "google"
                }
            
            img = Image.open(filepath)
            print(f"     ✓ Image opened - Size: {img.size}, Mode: {img.mode}")
            
            # Create model
            print("[5/6] Creating Gemini Vision model...")
            model = genai.GenerativeModel('gemini-2.5-flash')
            print("     ✓ Model created")
            
            # Comprehensive analysis prompt
            prompt = """Analyze this image comprehensively and provide:

1. **Visual Description**: Describe what you see in detail (objects, people, colors, composition)
2. **Text Extraction**: Extract ALL visible text including signs, labels, documents, watermarks
3. **Context & Purpose**: What is this image about? What information does it convey?
4. **Key Information**: Identify key data points, numbers, dates, names
5. **Technical Details**: Note image quality, style, and any notable visual elements

Be thorough and extract as much information as possible."""
            
            # Generate analysis
            print("[6/6] Sending to Gemini for analysis...")
            response = model.generate_content([prompt, img])
            
            if response and response.text:
                print("     ✓ Analysis complete!")
                print(f"     Response length: {len(response.text)} characters")
                print(f"{'='*60}\n")
                return {
                    "status": "success",
                    "analysis": response.text,
                    "source": "google"
                }
            else:
                print("     ✗ No response from Gemini")
                print(f"{'='*60}\n")
                return {
                    "status": "error",
                    "error": "No response from Gemini Vision",
                    "source": "google"
                }
                
        except Exception as e:
            print(f"     ✗ ERROR: {str(e)}")
            import traceback
            print("\nFull traceback:")
            traceback.print_exc()
            print(f"{'='*60}\n")
            
            return {
                "status": "error",
                "error": f"Failed to analyze image: {str(e)}",
                "source": "google"
            }


# Test the class on import
if __name__ == "__main__":
    print("\n" + "="*60)
    print("TESTING FileAnalyzer.analyze_image_with_google()")
    print("="*60)
    
    # Create a test image
    from PIL import Image
    test_img = Image.new('RGB', (200, 100), color='red')
    test_path = "test_analyzer.jpg"
    test_img.save(test_path)
    print(f"Created test image: {test_path}")
    
    # Test the function
    result = FileAnalyzer.analyze_image_with_google(test_path)
    
    print("\n" + "="*60)
    print("RESULT:")
    print("="*60)
    print(f"Status: {result.get('status')}")
    if result.get('status') == 'success':
        print(f"Analysis: {result.get('analysis', '')[:200]}...")
    else:
        print(f"Error: {result.get('error')}")
    print("="*60)
    
    # Cleanup
    os.remove(test_path)
    print(f"\nCleaned up: {test_path}")