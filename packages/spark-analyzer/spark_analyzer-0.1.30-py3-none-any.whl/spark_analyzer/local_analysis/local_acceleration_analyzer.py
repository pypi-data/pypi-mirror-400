#!/usr/bin/env python3
"""
Local Spark acceleration analyzer.

This script takes a JSON file containing Spark diagnostics payload and generates
an acceleration analysis output in Excel format with multiple sheets.
"""

import argparse
import json
import os
import sys

from pathlib import Path
from typing import Dict, Any, List

# Import openpyxl for Excel formatting (will be imported in save_excel_output if needed)

script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from data_models import SparkDiagnosticsPayload
from metrics_analyzer import CostEstimatorMetricsAnalyzer, ProductConfig


def validate_input_data(data: Dict[str, Any]) -> None:
    required_fields = ['application_id', 'stages']
    
    for field in required_fields:
        if field not in data:
            raise ValueError(f"Missing required field: {field}")
    
    if not data['application_id']:
        raise ValueError("Application ID cannot be empty")
    
    if not data['stages']:
        raise ValueError("No stages found in metrics data")
    
    start_time = data.get('application_start_time_ms', 0)
    end_time = data.get('application_end_time_ms', 0)
    
    if end_time > 0 and end_time <= start_time:
        raise ValueError("Invalid application time range")


def load_input_file(input_path: str) -> Dict[str, Any]:
    try:
        with open(input_path, 'r') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {input_path}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in input file: {e}")


def save_output_file(output_path: str, analysis: Dict[str, Any]) -> None:
    try:
        save_excel_output(output_path, analysis)
    except Exception as e:
        raise RuntimeError(f"Failed to save output file: {e}")


def calculate_compute_waste_percentage(idle_core_ms: int, total_core_ms: int) -> float:
    if total_core_ms <= 0:
        return 0.0
    
    idle_core_mins = idle_core_ms / 60000
    total_core_mins = total_core_ms / 60000
    
    return round((idle_core_mins / total_core_mins) * 100, 1)


def save_excel_output(output_path: str, analysis: Dict[str, Any]) -> None:
    """Save analysis to Excel format with multiple sheets (App Level, Workflow Level, Stage Level)"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl import styles
    except ImportError:
        raise RuntimeError("Excel output requires pandas and openpyxl. Install with: pip install pandas openpyxl")
    
    wb = Workbook()
    
    wb.remove(wb.active)
    
    create_app_level_sheet(wb, analysis, styles)
    create_workflow_level_sheet(wb, analysis)
    create_stage_level_sheet(wb, analysis)
    
    wb.save(output_path)


def create_app_level_sheet(wb, analysis: Dict[str, Any], styles) -> None:
    ws = wb.create_sheet("App Level Summary")
    
    total_metrics = analysis.get('total_metrics', {})
    acceleration_summary = analysis.get('acceleration_summary', {})
    workflow_distribution = total_metrics.get('workflow_distribution', {})
    
    used_core_ms = total_metrics.get('used_core_ms', 0)
    used_core_mins = round(used_core_ms / 60000, 2)
    total_duration_ms = total_metrics.get('total_duration_ms', 0)
    total_duration_mins = round(total_duration_ms / 60000, 2)
    
    total_core_ms = total_metrics.get('total_core_ms', 0)
    idle_core_ms = total_metrics.get('idle_core_ms', 0)
    compute_waste_percentage = calculate_compute_waste_percentage(idle_core_ms, total_core_ms)
    
    acceleration_range = acceleration_summary.get('acceleration_range', '0% - 0%')
    
    utilization_percentage = 0.0
    if total_core_ms > 0:
        utilization_percentage = round(((total_core_ms - idle_core_ms) / total_core_ms) * 100, 1)
    
    stages = analysis.get('stages', [])
    longest_stage = None
    most_expensive_stage = None
    
    if stages:
        longest_stage = max(stages, key=lambda x: x.get('duration_ms', 0))
        most_expensive_stage = max(stages, key=lambda x: x.get('used_core_ms', 0))
    
    extract_percentage = round(workflow_distribution.get('extract_percentage', 0.0), 1)
    transform_percentage = round(workflow_distribution.get('transform_percentage', 0.0), 1)
    load_percentage = round(workflow_distribution.get('load_percentage', 0.0), 1)
    unknown_percentage = round(workflow_distribution.get('unknown_percentage', 0.0), 1)
    metadata_percentage = round(workflow_distribution.get('metadata_percentage', 0.0), 1)
    
    most_expensive_workflow = "unknown"
    max_percentage = 0
    for workflow_type in ['extract', 'transform', 'load', 'unknown', 'metadata']:
        percentage = workflow_distribution.get(f'{workflow_type}_percentage', 0)
        if percentage > max_percentage:
            max_percentage = percentage
            most_expensive_workflow = workflow_type
    
    current_row = 1
    ws.cell(row=current_row, column=1, value="YOUR SPARK APPLICATION INSIGHTS")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=14)
    current_row += 1
    
    summary_data = [
        ['Application ID', analysis.get('application_id', 'N/A'), 'Unique identifier for this Spark application'],
        ['Application Duration', f"{total_duration_mins} minutes", 'Total wall-clock time from start to completion'],
        ['Total Compute Minutes Used', f"{used_core_mins}", 'Sum of all compute minutes consumed across all stages'],
        ['Projected Onehouse Compute Minutes', calculate_projected_core_mins(used_core_ms, acceleration_range), 'Estimated compute time with Onehouse optimizations applied'],
        ['Resource Efficiency', f"{utilization_percentage}%", 'Percentage of allocated compute resources actually utilized (100% - waste %)'],
        ['Compute Waste', f"{compute_waste_percentage}% of allocated resources", 'Percentage of allocated compute resources that remained idle during execution']
    ]
    
    for metric, value, note in summary_data:
        ws.cell(row=current_row, column=1, value=metric)
        ws.cell(row=current_row, column=2, value=value)
        ws.cell(row=current_row, column=3, value=note)
        current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="PERFORMANCE ANALYSIS")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=14)
    current_row += 1
    
    if longest_stage:
        longest_duration_mins = round(longest_stage.get('duration_ms', 0) / 60000, 2)
        ws.cell(row=current_row, column=1, value=f"• Longest stage: Stage {longest_stage.get('stage_id', 'N/A')} ({longest_duration_mins} mins)")
        current_row += 1
    
    if most_expensive_stage:
        most_expensive_core_ms = most_expensive_stage.get('used_core_ms', 0)
        most_expensive_core_mins = round(most_expensive_core_ms / 60000, 2)
        ws.cell(row=current_row, column=1, value=f"• Most resource-intensive: Stage {most_expensive_stage.get('stage_id', 'N/A')} ({most_expensive_core_mins} core-minutes)")
        current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"• Workflow breakdown: Extract {extract_percentage}%, Transform {transform_percentage}%, Load {load_percentage}%, Unknown {unknown_percentage}%, Metadata {metadata_percentage}%")
    current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="OPTIMIZATION OPPORTUNITIES")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=14)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"• Compute waste: {compute_waste_percentage}% of allocated compute")
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"• Focus area: {most_expensive_workflow} operations consume most resources")
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"• Expected acceleration benefits from Onehouse optimization: {acceleration_range}")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=11, color="FB0064")
    current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="ADDITIONAL COST SAVINGS OPPORTUNITIES")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=14)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value=f"We also found additional opportunities to save compute costs:")
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"1) Your jobs suggest that {compute_waste_percentage}% of compute resources are idle")
    current_row += 1
    ws.cell(row=current_row, column=1, value="   Onehouse Compute Runtime can reduce wasteful idle resources by workload aware elastic auto-scaling.")
    ws.cell(row=current_row, column=1)
    current_row += 1
    ws.cell(row=current_row, column=1, value="   Combined with our optimized Spark platform and deep datalake storage expertise, Onehouse could make more holistic recommendations to lower your costs.")
    ws.cell(row=current_row, column=1)
    current_row += 1
    ws.cell(row=current_row, column=1, value="2) Speedup joins on large fact tables by 2-8x from Quanton's faster join algorithms")
    current_row += 1
    ws.cell(row=current_row, column=1, value="3) Improved read/write performance from scaling table metadata, fixing file-sizing or correcting over-partitioning.")
    current_row += 1
    ws.cell(row=current_row, column=1, value="To explore these additional savings, we recommend using the free LakeView tool from Onehouse or contact Onehouse support at support@onehouse.ai.")
    current_row += 1
    current_row += 1
    ws.cell(row=current_row, column=1, value="For more information, checkout these blogs:")
    current_row += 1
    ws.cell(row=current_row, column=1, value="(a) https://www.onehouse.ai/blog/introducing-onehouse-compute-runtime-to-accelerate-lakehouse-workloads-across-all-engines")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=11, color="0066CC")
    current_row += 1
    ws.cell(row=current_row, column=1, value="(b) https://www.onehouse.ai/blog/announcing-spark-and-sql-on-the-onehouse-compute-runtime-with-quanton")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=11, color="0066CC")
    current_row += 1
    
    current_row += 1
    ws.cell(row=current_row, column=1, value="RECOMMENDATIONS")
    ws.cell(row=current_row, column=1).font = styles.Font(bold=True, size=14)
    current_row += 1
    
    recommendations = [
        f"• Review resource allocation for {most_expensive_workflow} stages",
        f"• Optimize {most_expensive_workflow} operations for better performance",
        "• Consider adjusting executor/core counts based on actual usage patterns",
        f"• Explore Onehouse optimization for additional {acceleration_range} performance gains"
    ]
    
    for recommendation in recommendations:
        ws.cell(row=current_row, column=1, value=recommendation)
        current_row += 1
    
    ws.column_dimensions['A'].width = 140
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    
    for row in range(1, current_row):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            if cell.value:  # Only process cells with values
                cell.alignment = styles.Alignment(wrap_text=True)
                
                if cell.font.color and cell.font.color.rgb != "000000":  # If custom color is set, preserve it
                    continue
                    
                if row == 1 or any(keyword in str(cell.value) for keyword in ['INSIGHTS', 'ANALYSIS', 'OPPORTUNITIES', 'SAVINGS', 'BENEFITS', 'RECOMMENDATIONS']):
                    cell.font = styles.Font(bold=True, size=14)
                else:
                    cell.font = styles.Font(11)


def create_workflow_level_sheet(wb, analysis: Dict[str, Any]) -> None:
    """Create workflow-level summary sheet"""
    ws = wb.create_sheet("Workflow Level Summary")
    
    headers = ['Workflow Type', 'Total Stages', 'Compute Minutes Used', 'Projected Onehouse Compute Minutes', 'Reduction Range', 'Workflow Percentage']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    workflow_breakdown = analysis.get('workflow_breakdown', {})
    total_metrics = analysis.get('total_metrics', {})
    workflow_distribution = total_metrics.get('workflow_distribution', {})
    
    workflow_types = ['unknown', 'extract', 'transform', 'load', 'metadata']
    
    for row, workflow_type in enumerate(workflow_types, 2):
        workflow_data = workflow_breakdown.get(workflow_type, {})
        
        used_core_ms = workflow_data.get('used_core_ms', 0)
        used_core_mins = round(used_core_ms / 60000, 2)
        
        acceleration_range = workflow_data.get('acceleration_range', '0% - 0%')
        
        projected_core_mins_range = calculate_projected_core_mins(used_core_ms, acceleration_range)
        
        percentage_key = f"{workflow_type}_percentage"
        workflow_percentage = round(workflow_distribution.get(percentage_key, 0.0), 2)
        
        ws.cell(row=row, column=1, value=workflow_type)
        ws.cell(row=row, column=2, value=workflow_data.get('total_stages', 0))
        ws.cell(row=row, column=3, value=used_core_mins)
        ws.cell(row=row, column=4, value=projected_core_mins_range)
        ws.cell(row=row, column=5, value=acceleration_range)
        ws.cell(row=row, column=6, value=f"{workflow_percentage}%")


def create_stage_level_sheet(wb, analysis: Dict[str, Any]) -> None:
    """Create stage-level details sheet"""
    ws = wb.create_sheet("Stage Level Details")
    
    headers = ['Stage ID', 'Workflow Type', 'Workflow Subtype', 'Duration (mins)', 'Compute Minutes Used', 'Projected Onehouse Compute Minutes', 'Reduction Range']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    stages = analysis.get('stages', [])
    
    for row, stage in enumerate(stages, 2):
        ws.cell(row=row, column=1, value=stage.get('stage_id', 0))
        
        workflow_info = stage.get('workflow_info', {})
        workflow_type = workflow_info.get('type', 'WORKFLOW_TYPE_UNKNOWN')
        ws.cell(row=row, column=2, value=workflow_type)
        
        workflow_subtype = 'N/A'
        if workflow_info.get('load_type'):
            workflow_subtype = workflow_info.get('load_type')
        elif workflow_info.get('transform_type'):
            workflow_subtype = workflow_info.get('transform_type')
        elif workflow_info.get('extract_type'):
            workflow_subtype = workflow_info.get('extract_type')
        elif workflow_info.get('metadata_type'):
            workflow_subtype = workflow_info.get('metadata_type')
        
        ws.cell(row=row, column=3, value=workflow_subtype)
        
        duration_ms = stage.get('duration_ms', 0)
        duration_mins = round(duration_ms / 60000, 4) 
        ws.cell(row=row, column=4, value=duration_mins)
        
        used_core_ms = stage.get('used_core_ms', 0)
        used_core_mins = round(used_core_ms / 60000, 4)
        ws.cell(row=row, column=5, value=used_core_mins)
        
        acceleration = stage.get('acceleration', {})
        projected_range = acceleration.get('projected_work_ms_with_acceleration', '0 - 0 ms')
        projected_core_mins_range = convert_ms_range_to_mins(projected_range)
        ws.cell(row=row, column=6, value=projected_core_mins_range)
        
        acceleration_range = acceleration.get('acceleration_range', '0% - 0%')
        ws.cell(row=row, column=7, value=acceleration_range)


def calculate_projected_core_mins(used_core_ms: int, reduction_range: str) -> str:
    """Calculate projected compute minutes using used_core_ms and reduction range"""
    if used_core_ms <= 0:
        return "0.0 - 0.0 mins"
    
    try:
        parts = reduction_range.replace('%', '').split(' - ')
        if len(parts) == 2:
            min_percentage = float(parts[0].strip()) / 100.0
            max_percentage = float(parts[1].strip()) / 100.0
            
            if max_percentage >= 1.0:
                min_projected_ms = 0  
            else:
                max_factor = 1.0 / (1.0 - max_percentage)
                min_projected_ms = int(used_core_ms / max_factor) 
            
            if min_percentage >= 1.0:
                max_projected_ms = 0 
            else:
                min_factor = 1.0 / (1.0 - min_percentage)
                max_projected_ms = int(used_core_ms / min_factor) 
            
            min_projected_mins = round(min_projected_ms / 60000, 2)
            max_projected_mins = round(max_projected_ms / 60000, 2)
            
            return f"{min_projected_mins} - {max_projected_mins} mins"
        
        return "0.0 - 0.0 mins"
    except (ValueError, AttributeError):
        return "0.0 - 0.0 mins"


def convert_ms_range_to_mins(ms_range: str) -> str:
    """Convert a millisecond range string to minutes range string"""
    try:
        parts = ms_range.replace(' ms', '').split(' - ')
        if len(parts) == 2:
            min_ms = int(parts[0].strip())
            max_ms = int(parts[1].strip())
            
            min_mins = round(min_ms / 60000, 2)
            max_mins = round(max_ms / 60000, 2)
            
            return f"{min_mins} - {max_mins} mins"
        else:
            return "0.0 - 0.0 mins"
    except (ValueError, AttributeError):
        return "0.0 - 0.0 mins"


def detect_storage_format(stages: List[Dict[str, Any]]) -> str:
    """Detect the primary storage format from stages to determine reduction percentages"""
    from data_models import StorageFormat
    
    storage_formats = set()
    
    for stage in stages:
        workflow_info = stage.get('workflow_info', {})
        storage_format_str = workflow_info.get('storage_format', 'STORAGE_FORMAT_UNKNOWN')
        
        try:
            storage_format = StorageFormat[storage_format_str]
            storage_formats.add(storage_format)
        except KeyError:
            continue
    
    if StorageFormat.STORAGE_FORMAT_DELTA in storage_formats or StorageFormat.STORAGE_FORMAT_ICEBERG in storage_formats:
        return "DELTA_ICEBERG"
    elif StorageFormat.STORAGE_FORMAT_HUDI in storage_formats:
        return "HUDI"
    else:
        return "HUDI"


def create_product_config(storage_format: str = "HUDI") -> ProductConfig:
    """Create product config based on detected storage format"""
    if storage_format == "DELTA_ICEBERG":
        return ProductConfig(
            unknown_range="33% - 50%",
            extract_scan_range="40% - 60%",
            transform_operation_range="40% - 50%",
            transform_join_range="50% - 75%",
            load_upsert_range="60% - 90%", 
            load_insert_range="50% - 90%",
            load_overwrite_range="0% - 0%",
            load_indexing_range="40% - 50%",
            metadata_operation_range="0% - 0%" 
        )
    else:
        return ProductConfig(
            unknown_range="33% - 50%",
            extract_scan_range="40% - 60%",
            transform_operation_range="40% - 50%",
            transform_join_range="50% - 75%",
            load_upsert_range="50% - 75%",  
            load_insert_range="50% - 90%",
            load_overwrite_range="0% - 0%", 
            load_indexing_range="40% - 50%",
            metadata_operation_range="0% - 0%"
        )


def get_input_files(input_arg: str) -> List[str]:
    """Parse input argument to get list of input files"""
    import glob
    
    expanded_arg = os.path.expanduser(input_arg)
    
    if '*' in expanded_arg or '?' in expanded_arg:
        files = glob.glob(expanded_arg)
        if not files:
            raise ValueError(f"No files found matching pattern: {input_arg} (expanded to: {expanded_arg})")
        return sorted(files)
    
    if ',' in expanded_arg:
        files = [f.strip() for f in expanded_arg.split(',') if f.strip()]
        if not files:
            raise ValueError("No valid files found in comma-separated list")
        return files
    
    if not os.path.exists(expanded_arg):
        raise ValueError(f"Input file not found: {input_arg} (expanded to: {expanded_arg})")
    return [expanded_arg]


def get_output_files(input_files: List[str], output_arg: str = None, output_dir: str = None) -> List[str]:
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_files = []
        for input_file in input_files:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = os.path.join(output_dir, f"{base_name}_detailed_analysis.xlsx")
            output_files.append(output_file)
        return output_files
    
    elif output_arg:
        if ',' in output_arg:
            output_files = [f.strip() for f in output_arg.split(',') if f.strip()]
            if len(output_files) != len(input_files):
                raise ValueError(f"Number of output files ({len(output_files)}) doesn't match number of input files ({len(input_files)})")
            return output_files
        else:
            if len(input_files) > 1:
                raise ValueError("Single output file specified but multiple input files provided. Use --output-dir for batch processing.")
            return [output_arg]
    
    else:
        output_files = []
        for input_file in input_files:
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = os.path.join(os.path.dirname(input_file), f"{base_name}_detailed_analysis.xlsx")
            output_files.append(output_file)
        return output_files


def process_single_file(input_file: str, output_file: str, verbose: bool = False) -> bool:
    try:
        if verbose:
            print(f"�� Processing: {input_file}")
        
        input_data = load_input_file(input_file)
        
        if verbose:
            print(f"📋 Application ID: {input_data.get('application_id', 'N/A')}")
            print(f"📊 Number of stages: {len(input_data.get('stages', []))}")
        
        analysis_result = analyze_spark_diagnostics(input_data)
        
        save_output_file(output_file, analysis_result)
        
        if verbose:
            print(f"✅ Saved: {output_file}")
            print(f"📈 Output format: Excel detailed analysis")
        
        return True
        
    except Exception as e:
        print(f"❌ Error processing {input_file}: {e}", file=sys.stderr)
        return False


def analyze_spark_diagnostics(input_data: Dict[str, Any]) -> Dict[str, Any]:
    """Analyze Spark diagnostics and generate acceleration analysis"""
    validate_input_data(input_data)
    
    stages = input_data.get('stages', [])
    storage_format = detect_storage_format(stages)
    
    spark_payload = SparkDiagnosticsPayload.from_dict(input_data)
    
    product_config = create_product_config(storage_format)
    
    analyzer = CostEstimatorMetricsAnalyzer(spark_payload, product_config)
    acceleration_analysis = analyzer.generate_acceleration_analysis()
    
    return acceleration_analysis.to_dict()


def main():
    parser = argparse.ArgumentParser(
        description="Local Spark acceleration analyzer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single file processing (Excel output)
  python -m spark_analyzer.local_analysis.local_acceleration_analyzer \\
      --input spark_diagnostics.json \\
      --output detailed_analysis.xlsx

  # Batch processing with Excel output
  python -m spark_analyzer.local_analysis.local_acceleration_analyzer \\
      --input "~/Documents/new/*.json" \\
      --output-dir ~/Documents/new/outputs/
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Input JSON file(s) containing Spark diagnostics payload. Can be: single file, comma-separated list, or glob pattern'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Output file(s) for acceleration analysis. If not specified, output files will be created in the same directory as input files'
    )
    
    parser.add_argument(
        '--output-dir', '-d',
        help='Output directory for batch processing. Files will be named based on input filenames'
    )
    

    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output'
    )
    
    args = parser.parse_args()
    
    try:
        input_files = get_input_files(args.input)
        
        output_files = get_output_files(input_files, args.output, args.output_dir)
        
        if args.verbose:
            print(f"📋 Found {len(input_files)} input file(s)")
            print(f"📁 Will generate {len(output_files)} Excel output file(s)")
        
        successful = 0
        failed = 0
        
        for i, (input_file, output_file) in enumerate(zip(input_files, output_files)):
            if args.verbose:
                print(f"\n🔄 Processing file {i+1}/{len(input_files)}")
            
            if process_single_file(input_file, output_file, args.verbose):
                successful += 1
            else:
                failed += 1
        
        print(f"\n📊 Processing Summary:")
        print(f"✅ Successful: {successful}")
        print(f"❌ Failed: {failed}")
        print(f"📁 Total files processed: {len(input_files)}")
        print(f"📄 Output format: Excel")
        
        if failed > 0:
            sys.exit(1)
        
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main() 