"""
plugins/secretsmanager/unused.py - Secrets Manager 미사용 분석

미사용 시크릿 탐지 및 비용 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_secret_price

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "secretsmanager:ListSecrets",
    ],
}

# 미사용 기준: 90일 이상 액세스 없음
UNUSED_DAYS_THRESHOLD = 90


class SecretStatus(Enum):
    """시크릿 상태"""

    NORMAL = "normal"
    UNUSED = "unused"
    PENDING_DELETE = "pending_delete"


@dataclass
class SecretInfo:
    """Secrets Manager 시크릿 정보"""

    account_id: str
    account_name: str
    region: str
    arn: str
    name: str
    description: str
    created_date: datetime | None
    last_accessed_date: datetime | None
    last_changed_date: datetime | None
    rotation_enabled: bool
    deleted_date: datetime | None = None

    @property
    def monthly_cost(self) -> float:
        return get_secret_price(self.region)

    @property
    def days_since_access(self) -> int | None:
        if self.last_accessed_date:
            return (datetime.now(timezone.utc) - self.last_accessed_date).days
        return None


@dataclass
class SecretFinding:
    """시크릿 분석 결과"""

    secret: SecretInfo
    status: SecretStatus
    recommendation: str


@dataclass
class SecretAnalysisResult:
    """시크릿 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_count: int = 0
    unused_count: int = 0
    pending_delete_count: int = 0
    normal_count: int = 0
    total_monthly_cost: float = 0.0
    unused_monthly_cost: float = 0.0
    findings: list[SecretFinding] = field(default_factory=list)


def collect_secrets(session, account_id: str, account_name: str, region: str) -> list[SecretInfo]:
    """Secrets Manager 시크릿 수집"""
    sm = get_client(session, "secretsmanager", region_name=region)
    secrets = []

    paginator = sm.get_paginator("list_secrets")
    for page in paginator.paginate():
        for secret in page.get("SecretList", []):
            secrets.append(
                SecretInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    arn=secret.get("ARN", ""),
                    name=secret.get("Name", ""),
                    description=secret.get("Description", ""),
                    created_date=secret.get("CreatedDate"),
                    last_accessed_date=secret.get("LastAccessedDate"),
                    last_changed_date=secret.get("LastChangedDate"),
                    rotation_enabled=secret.get("RotationEnabled", False),
                    deleted_date=secret.get("DeletedDate"),
                )
            )

    return secrets


def analyze_secrets(secrets: list[SecretInfo], account_id: str, account_name: str, region: str) -> SecretAnalysisResult:
    """시크릿 분석"""
    result = SecretAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_count=len(secrets),
    )

    for secret in secrets:
        result.total_monthly_cost += secret.monthly_cost

        if secret.deleted_date:
            result.pending_delete_count += 1
            result.findings.append(
                SecretFinding(
                    secret=secret,
                    status=SecretStatus.PENDING_DELETE,
                    recommendation="삭제 예정됨",
                )
            )
            continue

        if secret.days_since_access and secret.days_since_access > UNUSED_DAYS_THRESHOLD:
            result.unused_count += 1
            result.unused_monthly_cost += secret.monthly_cost
            result.findings.append(
                SecretFinding(
                    secret=secret,
                    status=SecretStatus.UNUSED,
                    recommendation=f"{secret.days_since_access}일간 액세스 없음 - 삭제 검토",
                )
            )
            continue

        if secret.last_accessed_date is None and secret.created_date:
            age_days = (datetime.now(timezone.utc) - secret.created_date).days
            if age_days > UNUSED_DAYS_THRESHOLD:
                result.unused_count += 1
                result.unused_monthly_cost += secret.monthly_cost
                result.findings.append(
                    SecretFinding(
                        secret=secret,
                        status=SecretStatus.UNUSED,
                        recommendation=f"생성 후 {age_days}일간 액세스 없음",
                    )
                )
                continue

        result.normal_count += 1
        result.findings.append(
            SecretFinding(
                secret=secret,
                status=SecretStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[SecretAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    ws = wb.create_sheet("Summary")
    ws["A1"] = "Secrets Manager 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "Region", "전체", "미사용", "삭제예정", "총 비용", "낭비 비용"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_count)
        ws.cell(row=row, column=4, value=r.unused_count)
        ws.cell(row=row, column=5, value=r.pending_delete_count)
        ws.cell(row=row, column=6, value=f"${r.total_monthly_cost:,.2f}")
        ws.cell(row=row, column=7, value=f"${r.unused_monthly_cost:,.2f}")
        if r.unused_count > 0:
            ws.cell(row=row, column=4).fill = red_fill

    ws_detail = wb.create_sheet("Secrets")
    detail_headers = [
        "Account",
        "Region",
        "Name",
        "상태",
        "마지막 액세스",
        "Rotation",
        "월간 비용",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != SecretStatus.NORMAL:
                detail_row += 1
                s = f.secret
                ws_detail.cell(row=detail_row, column=1, value=s.account_name)
                ws_detail.cell(row=detail_row, column=2, value=s.region)
                ws_detail.cell(row=detail_row, column=3, value=s.name)
                ws_detail.cell(row=detail_row, column=4, value=f.status.value)
                ws_detail.cell(
                    row=detail_row,
                    column=5,
                    value=s.last_accessed_date.strftime("%Y-%m-%d") if s.last_accessed_date else "없음",
                )
                ws_detail.cell(row=detail_row, column=6, value="예" if s.rotation_enabled else "아니오")
                ws_detail.cell(row=detail_row, column=7, value=f"${s.monthly_cost:,.2f}")
                ws_detail.cell(row=detail_row, column=8, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Secrets_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> SecretAnalysisResult | None:
    """단일 계정/리전의 시크릿 수집 및 분석 (병렬 실행용)"""
    secrets = collect_secrets(session, account_id, account_name, region)
    if not secrets:
        return None
    return analyze_secrets(secrets, account_id, account_name, region)


def run(ctx) -> None:
    """Secrets Manager 미사용 분석"""
    console.print("[bold]Secrets Manager 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="secretsmanager")
    results: list[SecretAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total = sum(r.total_count for r in results)
    unused = sum(r.unused_count for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체: {total}개 / 미사용: [yellow]{unused}개[/yellow] (${unused_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("secrets-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
