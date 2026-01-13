"""
plugins/dynamodb/capacity_mode.py - DynamoDB 용량 모드 분석

Provisioned vs On-Demand 용량 모드 최적화 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing.dynamodb import (
    estimate_ondemand_cost,
    get_dynamodb_monthly_cost,
)

console = Console()

# 분석 기간 (일)
ANALYSIS_DAYS = 14

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "dynamodb:ListTables",
        "dynamodb:DescribeTable",
        "cloudwatch:GetMetricStatistics",
    ],
}


class CapacityRecommendation(Enum):
    """용량 모드 권장 사항"""

    KEEP_PROVISIONED = "keep_provisioned"
    SWITCH_TO_ONDEMAND = "switch_to_ondemand"
    SWITCH_TO_PROVISIONED = "switch_to_provisioned"
    REDUCE_CAPACITY = "reduce_capacity"
    INCREASE_CAPACITY = "increase_capacity"
    OPTIMAL = "optimal"


@dataclass
class TableCapacityInfo:
    """DynamoDB 테이블 용량 정보"""

    account_id: str
    account_name: str
    region: str
    table_name: str
    table_status: str
    billing_mode: str  # PROVISIONED or PAY_PER_REQUEST
    item_count: int
    size_bytes: int
    # Provisioned 설정
    provisioned_read: int = 0
    provisioned_write: int = 0
    last_increase_dt: datetime | None = None
    last_decrease_dt: datetime | None = None
    decreases_today: int = 0
    # CloudWatch 지표 (평균)
    avg_consumed_read: float = 0.0
    avg_consumed_write: float = 0.0
    max_consumed_read: float = 0.0
    max_consumed_write: float = 0.0
    throttled_read: float = 0.0
    throttled_write: float = 0.0
    created_at: datetime | None = None

    @property
    def size_mb(self) -> float:
        """테이블 크기 (MB)"""
        return self.size_bytes / (1024 * 1024)

    @property
    def read_utilization(self) -> float:
        """읽기 용량 사용률 (%)"""
        if self.provisioned_read <= 0:
            return 0
        return (self.avg_consumed_read / self.provisioned_read) * 100

    @property
    def write_utilization(self) -> float:
        """쓰기 용량 사용률 (%)"""
        if self.provisioned_write <= 0:
            return 0
        return (self.avg_consumed_write / self.provisioned_write) * 100

    @property
    def estimated_provisioned_cost(self) -> float:
        """Provisioned 모드 예상 월간 비용 (pricing 모듈 사용)"""
        storage_gb = self.size_bytes / (1024**3)
        return get_dynamodb_monthly_cost(
            region=self.region,
            billing_mode="PROVISIONED",
            rcu=self.provisioned_read,
            wcu=self.provisioned_write,
            storage_gb=storage_gb,
        )

    @property
    def estimated_ondemand_cost(self) -> float:
        """On-Demand 모드 예상 월간 비용 (pricing 모듈 사용)"""
        storage_gb = self.size_bytes / (1024**3)
        return estimate_ondemand_cost(
            region=self.region,
            avg_consumed_rcu=self.avg_consumed_read,
            avg_consumed_wcu=self.avg_consumed_write,
            storage_gb=storage_gb,
        )


@dataclass
class TableCapacityFinding:
    """테이블 용량 분석 결과"""

    table: TableCapacityInfo
    recommendation: CapacityRecommendation
    reason: str
    potential_savings: float = 0.0


@dataclass
class CapacityAnalysisResult:
    """용량 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_tables: int = 0
    provisioned_tables: int = 0
    ondemand_tables: int = 0
    optimization_candidates: int = 0
    potential_savings: float = 0.0
    findings: list[TableCapacityFinding] = field(default_factory=list)


def collect_capacity_info(session, account_id: str, account_name: str, region: str) -> list[TableCapacityInfo]:
    """DynamoDB 테이블 용량 정보 수집"""
    from botocore.exceptions import ClientError

    dynamodb = get_client(session, "dynamodb", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    tables = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=ANALYSIS_DAYS)

    try:
        paginator = dynamodb.get_paginator("list_tables")
        for page in paginator.paginate():
            for table_name in page.get("TableNames", []):
                try:
                    desc = dynamodb.describe_table(TableName=table_name)
                    t = desc.get("Table", {})

                    billing = t.get("BillingModeSummary", {})
                    billing_mode = billing.get("BillingMode", "PROVISIONED")

                    throughput = t.get("ProvisionedThroughput", {})

                    table = TableCapacityInfo(
                        account_id=account_id,
                        account_name=account_name,
                        region=region,
                        table_name=table_name,
                        table_status=t.get("TableStatus", ""),
                        billing_mode=billing_mode,
                        item_count=t.get("ItemCount", 0),
                        size_bytes=t.get("TableSizeBytes", 0),
                        provisioned_read=throughput.get("ReadCapacityUnits", 0),
                        provisioned_write=throughput.get("WriteCapacityUnits", 0),
                        last_increase_dt=throughput.get("LastIncreaseDateTime"),
                        last_decrease_dt=throughput.get("LastDecreaseDateTime"),
                        decreases_today=throughput.get("NumberOfDecreasesToday", 0),
                        created_at=t.get("CreationDateTime"),
                    )

                    # CloudWatch 지표 조회
                    try:
                        # ConsumedReadCapacityUnits
                        read_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedReadCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Average", "Maximum"],
                        )
                        if read_resp.get("Datapoints"):
                            table.avg_consumed_read = sum(d["Average"] for d in read_resp["Datapoints"]) / len(
                                read_resp["Datapoints"]
                            )
                            table.max_consumed_read = max(d["Maximum"] for d in read_resp["Datapoints"])

                        # ConsumedWriteCapacityUnits
                        write_resp = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ConsumedWriteCapacityUnits",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Average", "Maximum"],
                        )
                        if write_resp.get("Datapoints"):
                            table.avg_consumed_write = sum(d["Average"] for d in write_resp["Datapoints"]) / len(
                                write_resp["Datapoints"]
                            )
                            table.max_consumed_write = max(d["Maximum"] for d in write_resp["Datapoints"])

                        # ThrottledRequests (Read)
                        read_throttle = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="ReadThrottledRequests",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if read_throttle.get("Datapoints"):
                            table.throttled_read = sum(d["Sum"] for d in read_throttle["Datapoints"])

                        # ThrottledRequests (Write)
                        write_throttle = cloudwatch.get_metric_statistics(
                            Namespace="AWS/DynamoDB",
                            MetricName="WriteThrottledRequests",
                            Dimensions=[{"Name": "TableName", "Value": table_name}],
                            StartTime=start_time,
                            EndTime=now,
                            Period=86400,
                            Statistics=["Sum"],
                        )
                        if write_throttle.get("Datapoints"):
                            table.throttled_write = sum(d["Sum"] for d in write_throttle["Datapoints"])

                    except ClientError:
                        pass

                    tables.append(table)

                except ClientError:
                    continue

    except ClientError:
        pass

    return tables


def analyze_capacity(
    tables: list[TableCapacityInfo], account_id: str, account_name: str, region: str
) -> CapacityAnalysisResult:
    """DynamoDB 용량 모드 분석"""
    result = CapacityAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_tables=len(tables),
    )

    for table in tables:
        if table.billing_mode == "PAY_PER_REQUEST":
            result.ondemand_tables += 1

            # On-Demand → Provisioned 전환 검토
            # 사용량이 일정하고 예측 가능한 경우
            if table.avg_consumed_read > 0 or table.avg_consumed_write > 0:
                savings = table.estimated_ondemand_cost - table.estimated_provisioned_cost
                if savings > 0:
                    result.optimization_candidates += 1
                    result.potential_savings += savings
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.SWITCH_TO_PROVISIONED,
                            reason=f"Provisioned 전환 시 월 ${savings:.2f} 절감 가능",
                            potential_savings=savings,
                        )
                    )
                else:
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.OPTIMAL,
                            reason="On-Demand가 현재 사용 패턴에 적합",
                        )
                    )
            else:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.OPTIMAL,
                        reason="사용량 없음 (삭제 검토 권장)",
                    )
                )

        else:  # PROVISIONED
            result.provisioned_tables += 1

            # Throttling 발생 시 용량 증가 필요
            if table.throttled_read > 0 or table.throttled_write > 0:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.INCREASE_CAPACITY,
                        reason=f"Throttling 발생 (R:{table.throttled_read:.0f}, W:{table.throttled_write:.0f})",
                    )
                )
                continue

            # 사용률 분석
            read_util = table.read_utilization
            write_util = table.write_utilization

            # 매우 낮은 사용률 → On-Demand 전환 검토
            if read_util < 10 and write_util < 10:
                savings = table.estimated_provisioned_cost - table.estimated_ondemand_cost
                if savings > 0:
                    result.optimization_candidates += 1
                    result.potential_savings += savings
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.SWITCH_TO_ONDEMAND,
                            reason=f"저사용 (R:{read_util:.1f}%, W:{write_util:.1f}%) - On-Demand 전환 시 월 ${savings:.2f} 절감",
                            potential_savings=savings,
                        )
                    )
                else:
                    result.findings.append(
                        TableCapacityFinding(
                            table=table,
                            recommendation=CapacityRecommendation.REDUCE_CAPACITY,
                            reason=f"저사용 (R:{read_util:.1f}%, W:{write_util:.1f}%) - 용량 축소 검토",
                        )
                    )
            # 적정 사용률
            elif read_util < 70 and write_util < 70:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.OPTIMAL,
                        reason=f"적정 사용률 (R:{read_util:.1f}%, W:{write_util:.1f}%)",
                    )
                )
            # 높은 사용률
            else:
                result.findings.append(
                    TableCapacityFinding(
                        table=table,
                        recommendation=CapacityRecommendation.INCREASE_CAPACITY,
                        reason=f"높은 사용률 (R:{read_util:.1f}%, W:{write_util:.1f}%) - 용량 증가 검토",
                    )
                )

    return result


def generate_report(results: list[CapacityAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "DynamoDB 용량 모드 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Account",
        "Region",
        "전체",
        "Provisioned",
        "On-Demand",
        "최적화 대상",
        "예상 절감액",
    ]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_tables)
        ws.cell(row=row, column=4, value=r.provisioned_tables)
        ws.cell(row=row, column=5, value=r.ondemand_tables)
        ws.cell(row=row, column=6, value=r.optimization_candidates)
        ws.cell(row=row, column=7, value=f"${r.potential_savings:,.2f}")
        if r.potential_savings > 0:
            ws.cell(row=row, column=7).fill = green_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Tables")
    detail_headers = [
        "Account",
        "Region",
        "Table Name",
        "Billing Mode",
        "Prov. RCU",
        "Prov. WCU",
        "Avg R",
        "Avg W",
        "R Util%",
        "W Util%",
        "Prov 비용",
        "OD 비용",
        "권장 사항",
        "사유",
        "절감액",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            detail_row += 1
            t = f.table
            rec_labels = {
                CapacityRecommendation.SWITCH_TO_ONDEMAND: "On-Demand 전환",
                CapacityRecommendation.SWITCH_TO_PROVISIONED: "Provisioned 전환",
                CapacityRecommendation.REDUCE_CAPACITY: "용량 축소",
                CapacityRecommendation.INCREASE_CAPACITY: "용량 증가",
                CapacityRecommendation.OPTIMAL: "최적",
            }
            ws_detail.cell(row=detail_row, column=1, value=t.account_name)
            ws_detail.cell(row=detail_row, column=2, value=t.region)
            ws_detail.cell(row=detail_row, column=3, value=t.table_name)
            ws_detail.cell(row=detail_row, column=4, value=t.billing_mode)
            ws_detail.cell(row=detail_row, column=5, value=t.provisioned_read)
            ws_detail.cell(row=detail_row, column=6, value=t.provisioned_write)
            ws_detail.cell(row=detail_row, column=7, value=f"{t.avg_consumed_read:.1f}")
            ws_detail.cell(row=detail_row, column=8, value=f"{t.avg_consumed_write:.1f}")
            ws_detail.cell(row=detail_row, column=9, value=f"{t.read_utilization:.1f}")
            ws_detail.cell(row=detail_row, column=10, value=f"{t.write_utilization:.1f}")
            ws_detail.cell(
                row=detail_row,
                column=11,
                value=f"${t.estimated_provisioned_cost:.2f}",
            )
            ws_detail.cell(row=detail_row, column=12, value=f"${t.estimated_ondemand_cost:.2f}")
            ws_detail.cell(
                row=detail_row,
                column=13,
                value=rec_labels.get(f.recommendation, f.recommendation.value),
            )
            ws_detail.cell(row=detail_row, column=14, value=f.reason)
            ws_detail.cell(row=detail_row, column=15, value=f"${f.potential_savings:.2f}")

            # 색상 적용
            if f.recommendation in (
                CapacityRecommendation.SWITCH_TO_ONDEMAND,
                CapacityRecommendation.SWITCH_TO_PROVISIONED,
            ):
                ws_detail.cell(row=detail_row, column=13).fill = green_fill
            elif f.recommendation == CapacityRecommendation.INCREASE_CAPACITY:
                ws_detail.cell(row=detail_row, column=13).fill = red_fill
            elif f.recommendation == CapacityRecommendation.REDUCE_CAPACITY:
                ws_detail.cell(row=detail_row, column=13).fill = yellow_fill

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"DynamoDB_Capacity_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> CapacityAnalysisResult | None:
    """단일 계정/리전의 DynamoDB 용량 분석 (병렬 실행용)"""
    tables = collect_capacity_info(session, account_id, account_name, region)
    if not tables:
        return None
    return analyze_capacity(tables, account_id, account_name, region)


def run(ctx) -> None:
    """DynamoDB 용량 모드 분석"""
    console.print("[bold]DynamoDB 용량 모드 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="dynamodb")
    results: list[CapacityAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_tables = sum(r.total_tables for r in results)
    total_provisioned = sum(r.provisioned_tables for r in results)
    total_ondemand = sum(r.ondemand_tables for r in results)
    total_candidates = sum(r.optimization_candidates for r in results)
    total_savings = sum(r.potential_savings for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"전체 테이블: {total_tables}개")
    console.print(f"  Provisioned: {total_provisioned}개 / On-Demand: {total_ondemand}개")
    console.print(
        f"최적화 대상: [cyan]{total_candidates}개[/cyan] (예상 절감: [green]${total_savings:,.2f}/월[/green])"
    )

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("dynamodb-capacity").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
