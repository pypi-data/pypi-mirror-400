"""XLSX/XLSM file handler using openpyxl."""

# ============================================================================
# Imports
# ============================================================================

from pathlib import Path
from typing import BinaryIO

import numpy as np
import openpyxl
import pandas as pd

from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.parsing.base_handler import FileSource, FormatHandler, ParseOptions


# ============================================================================
# Config
# ============================================================================

EXCEL_ERRORS = [
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
    "#GETTING_DATA",
]


# ============================================================================
# Core
# ============================================================================

class XLSXHandler(FormatHandler):
    """Handler for XLSX and XLSM files."""

    def can_handle(self, format_type: str) -> bool:
        """Check if this handler can process the format."""
        return format_type in ("xlsx", "xlsm")

    def parse(
        self,
        file_source: FileSource,
        sheet: str | None,
        options: ParseOptions,
    ) -> pd.DataFrame:
        """Parse XLSX/XLSM file to DataFrame."""
        read_only = options.merge_strategy == "skip"

        # Reset buffer position if file-like object
        is_fileobj = hasattr(file_source, "read")
        if is_fileobj and hasattr(file_source, "seek"):
            file_source.seek(0)

        file_desc = "<stream>" if is_fileobj else str(file_source)

        try:
            wb = openpyxl.load_workbook(
                file_source,
                read_only  = read_only,
                data_only  = options.data_only,
                keep_links = False,
            )
        except PermissionError as e:
            raise FileError(
                f"Permission denied: {file_desc}",
                file_path  = file_desc,
                operation  = "open",
            ) from e
        except Exception as e:
            raise FormatError(
                f"Cannot open Excel file: {e}",
                file_path  = file_desc,
            ) from e

        try:
            if sheet:
                if sheet not in wb.sheetnames:
                    raise FormatError(
                        f"Sheet '{sheet}' not found",
                        file_path        = file_desc,
                        detected_format  = "xlsx",
                    )
                ws = wb[sheet]
            else:
                ws = wb.active

            if options.merge_strategy != "skip" and not read_only:
                self._handle_merged_cells(ws, options.merge_strategy)

            data = self._read_worksheet(ws, options)

            if not data:
                return pd.DataFrame()

            df = pd.DataFrame(data)

            if options.skip_rows > 0 and not options.cell_range:
                df = df.iloc[options.skip_rows:]

            if options.skip_footer > 0:
                df = df.iloc[:-options.skip_footer]

            if options.header_rows > 0 and len(df) >= options.header_rows:
                df, columns = self._generate_column_names(df, options.header_rows)
                df.columns  = columns
                df          = df.reset_index(drop=True)
            else:
                df.columns = [f"col_{i}" for i in range(len(df.columns))]

            df = self._clean_excel_data(df, options)

            return df

        finally:
            wb.close()

    def _read_worksheet(
        self,
        ws,
        options: ParseOptions,
    ) -> list[list]:
        """Read worksheet data respecting options."""
        data = []

        if options.cell_range:
            try:
                rows_iter = ws[options.cell_range]
                if not isinstance(rows_iter[0], tuple):
                    rows_iter = [rows_iter]

                for row in rows_iter:
                    if not isinstance(row, tuple):
                        row = [row]
                    row_values = [cell.value for cell in row]
                    data.append(row_values)
            except Exception as e:
                raise FormatError(
                    f"Invalid cell range: {options.cell_range}",
                    detected_format = "xlsx",
                ) from e
        else:
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                if options.ignore_hidden:
                    try:
                        if ws.row_dimensions[row_idx].hidden:
                            continue
                    except (KeyError, AttributeError):
                        pass

                row_values = []
                for cell in row:
                    if options.ignore_hidden:
                        try:
                            if ws.column_dimensions[cell.column_letter].hidden:
                                continue
                        except (KeyError, AttributeError):
                            pass

                    row_values.append(cell.value)

                data.append(row_values)

        return data

    def _handle_merged_cells(self, ws, strategy: str) -> None:
        """Handle merged cells according to strategy."""
        merged_ranges = list(ws.merged_cells.ranges)

        for merged_range in merged_ranges:
            top_left_value = ws.cell(
                merged_range.min_row,
                merged_range.min_col,
            ).value

            ws.unmerge_cells(str(merged_range))

            if strategy == "fill":
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        ws.cell(row, col).value = top_left_value

            elif strategy == "first_only":
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        if row == merged_range.min_row and col == merged_range.min_col:
                            continue
                        ws.cell(row, col).value = None

    def _clean_excel_data(
        self,
        df: pd.DataFrame,
        options: ParseOptions,
    ) -> pd.DataFrame:
        """Clean Excel-specific data issues."""
        # Replace Excel errors with NaN using map to avoid FutureWarning
        error_set = set(EXCEL_ERRORS)

        # Handle duplicate column names by using integer positions
        for idx in range(len(df.columns)):
            series = df.iloc[:, idx]
            if series.dtype == object:
                df.iloc[:, idx] = series.map(lambda x: np.nan if x in error_set else x)

        if options.na_values:
            na_set = set(options.na_values)
            for idx in range(len(df.columns)):
                series = df.iloc[:, idx]
                if series.dtype == object:
                    df.iloc[:, idx] = series.map(lambda x: np.nan if x in na_set else x)

        return df

    def get_sheet_names(self, file_source: FileSource) -> list[str]:
        """Get list of sheet names."""
        is_fileobj = hasattr(file_source, "read")
        if is_fileobj and hasattr(file_source, "seek"):
            file_source.seek(0)

        file_desc = "<stream>" if is_fileobj else str(file_source)

        try:
            wb     = openpyxl.load_workbook(file_source, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except PermissionError as e:
            raise FileError(
                f"Permission denied: {file_desc}",
                file_path  = file_desc,
                operation  = "get_sheets",
            ) from e
        except Exception as e:
            try:
                if is_fileobj and hasattr(file_source, "seek"):
                    file_source.seek(0)
                xl_file = pd.ExcelFile(file_source, engine="openpyxl")
                sheets  = xl_file.sheet_names
                xl_file.close()
                return sheets
            except Exception:
                raise FormatError(
                    f"Cannot read sheet names: {e}",
                    file_path  = file_desc,
                ) from e

    def validate(self, file_source: FileSource) -> tuple[bool, str | None]:
        """Validate that file can be parsed."""
        is_fileobj = hasattr(file_source, "read")
        if is_fileobj and hasattr(file_source, "seek"):
            file_source.seek(0)

        try:
            wb = openpyxl.load_workbook(file_source, read_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, str(e)
