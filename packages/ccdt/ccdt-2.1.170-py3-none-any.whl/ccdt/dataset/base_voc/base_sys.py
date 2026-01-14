# 计算机登录用户: jk
# 系统日期: 2023/10/31 14:13
# 项目名称: chipeak_cv_data_tool
# 开发者: zhanyong
import os.path
import json
from pathlib import Path
import string  # 直接使用 Python 的 string.ascii_uppercase 生成列字母 A 到 Z 的索引映射。这样可以确保无论 Excel 表格中是否有相应的列，列号映射始终有效。
import openpyxl
import xlrd
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from shapely.geometry import Polygon
from ccdt.dataset import *
from collections import Counter
import re
import cv2


class BaseSys(BaseLabelme):
    def __init__(self, *args, **kwargs):
        self.label_name = args[1].label_name  # 获取自定义模型标签名称
        self.background = args[1].background  # 获取是否保存json，默认false
        self.polygonVertex = args[1].polygonVertex  # 获取多边形坐标
        self.output_dir = args[1].output_dir  # 输出路径
        self.columns_fill = args[1].columns_fill  # excel表格列名称
        self.column_titles = args[1].column_titles  # excel表格列标题
        self.column_content_type = args[1].column_content_type  # excel表格某列填充内容条件判断
        self.excel_file = args[1].input_datasets
        self.fill_rules = args[1].fill_rules
        self.time = args[1].time  # 输入不同时间段
        # 在这里定义labelme数据结构格式初始化
        super(BaseSys, self).__init__(*args, **kwargs)

    def sys2labelme(self, args):
        sys_to_labelme = list()
        for dataset in self.datasets:
            # print(dataset)
            obj_path = Path(dataset.get("full_path"))
            labelme_file = obj_path.stem + '.json'
            # 拼接json路径，读取json文件，封装labelme_info对象
            # dataset.get("full_path")
            # original_json_path = os.path.join(obj_path.parent, json_file)
            # relative_path = Path('../', '00.images', dataset.get('image_file'))
            # image_dir = Path(dataset.get('image_dir'), "00.images")
            relative_path = os.path.join("../", "00.images", dataset.get("image_file"))  # 构建相对路径
            new_file_path = self.change_file_extension(dataset.get("full_path"), "json")
            with open(new_file_path, 'r', encoding='UTF-8') as labelme_fp:
                content = json.load(labelme_fp)
                if args.label_name == "channel_occupancy":
                    labelme_info = self.analysis_channel_occupancy_json(content, args)  # 通道占用数据处理
                    if not labelme_info.get("shapes"):
                        continue
                elif args.label_name == "channel_blockage":
                    if content.get("data").get("occupy") != '' or content.get("data").get("unoccupied") != '':
                        continue
                    labelme_info = self.analysis_channel_blockage_json(content, args)  # 通道阻塞数据处理
                    if not labelme_info.get("shapes"):  # 阻塞解除可以跳过处理
                        continue
                elif args.label_name == "blacked_out":
                    labelme_info = self.analysis_blacked_out_json(content, args)  # 通道占用数据处理
                    if not labelme_info.get("shapes"):
                        continue
                else:
                    labelme_info = self.analysis_json(content)  # 解析json格式内容，返回labelme格式
                # 更新labelme的字典结构内容
                labelme_info.update({'imagePath': relative_path})
                labelme_info.update({'imageHeight': dataset.get('image_height')})
                labelme_info.update({'imageWidth': dataset.get('image_width')})
                labelme_info.update({'md5Value': dataset.get('md5_value')})
                # 更新保存labelme的字典结构内容
                dataset.update({'relative_path': relative_path})
                dataset.update({'labelme_file': labelme_file})
                dataset.update({'labelme_info': labelme_info})
                dataset.update({'labelme_dir': os.path.join(dataset.get('image_dir'), "01.labelme")})
                dataset.update({'image_dir': os.path.join(dataset.get('image_dir'), "00.images")})
                new_json_path = os.path.join(dataset.get('output_dir'), dataset.get('labelme_dir'), dataset.get('labelme_file'))
                dataset.update({"json_path": new_json_path})
                dataset.update({'background': self.background})
            sys_to_labelme.append(dataset)
        self.save_labelme(sys_to_labelme, self.output_dir, None)  # self.output_dir为空字符串也是可以的

    @staticmethod
    def change_file_extension(file_path, new_extension):
        base = os.path.splitext(file_path)[0]  # 获取不带后缀的文件路径
        return f"{base}.{new_extension.lstrip('.')}"  # 确保新后缀无 `.` 冲突

    def newbie2labelme(self):
        """
        把香港newbie内部测试转换成labelme
        """
        for dataset in self.datasets:
            if dataset.get('labelme_info') is not None:
                if dataset.get('labelme_info').get("shapes") is not None:
                    dataset.get('labelme_info').update({'imagePath': dataset.get("relative_path")})
                    dataset.get('labelme_info').update({'imageWidth': dataset.get("image_width")})
                    dataset.get('labelme_info').update({'imageHeight': dataset.get("image_height")})
                    dataset.get('labelme_info').update({'md5Value': dataset.get("md5_value")})
        self.save_labelme(self.datasets, self.output_dir, None)

    def polygon2labelme(self):
        """
        输入多边形坐标，对labelme进行画框
        """
        for dataset in self.datasets:
            if dataset.get('labelme_info') is not None:
                if dataset.get('labelme_info').get("shapes") is not None:
                    points = self.polygonVertex
                    # 只处理带多边形框，并转labelme
                    shape = {"label": "polygon", "points": points, "group_id": None, "shape_type": "polygon", "flags": {}, 'text': None}
                    dataset.get('labelme_info').get("shapes").append(shape)
            else:
                labelme_data = dict(
                    version='4.5.9',
                    flags={},
                    shapes=[],
                    imagePath=None,
                    imageData=None,
                    imageHeight=None,
                    imageWidth=None,
                    md5Value=None
                )
                shapes = []
                points = self.polygonVertex
                # 只处理带多边形框，并转labelme
                shape = {"label": "polygon", "points": points, "group_id": None, "shape_type": "polygon", "flags": {}, 'text': None}
                shapes.append(shape)
                labelme_data.update({'shapes': shapes})
                labelme_data.update({'imagePath': dataset.get("relative_path")})
                labelme_data.update({'imageWidth': dataset.get("image_width")})
                labelme_data.update({'imageHeight': dataset.get("image_height")})
                labelme_data.update({'md5Value': dataset.get("md5_value")})
                dataset.update({'labelme_info': labelme_data})
                dataset.update({'background': True})
        self.save_labelme(self.datasets, self.output_dir, None)

    def analysis_json(self, content):
        # 定义labelme数据结构
        labelme_data = dict(
            version='4.5.14',
            flags={},
            shapes=[],
            imagePath=None,
            imageData=None,
            imageHeight=None,
            imageWidth=None,
            md5Value=None
        )
        for key, value in content.items():
            if key == "alarm_data":  # 目标检测框
                # 存在的隐患，alarm_data这个字典中，针对多个目标是什么结构目前不清楚，以下逻辑，有且只有一个目标框的情况
                lt_x = content.get("alarm_data").get("rectangle").get('lt_x')
                lt_y = content.get("alarm_data").get("rectangle").get('lt_y')
                rb_x = content.get("alarm_data").get("rectangle").get('rb_x')
                rb_y = content.get("alarm_data").get("rectangle").get('rb_y')
                message = content.get("alarm_data").get("message")
                confidence = str(content.get("alarm_data").get("confidence"))
                text = message + " 阈值" + confidence  # 车牌和阈值拼接
                points = [[lt_x, lt_y], [rb_x, rb_y]]  # 坐标点计算，目前使用左上角的点和右下角的点计算
                shape = {"label": self.label_name, "points": points, "group_id": None, "shape_type": "rectangle", "flags": {}, 'text': text}
                labelme_data.get('shapes').append(shape)
            if key == "vertex_data" and value is not None:  # 多边形坐标点，在多边形不为空的前提下进行
                for polygon_key, polygon_value in value.items():
                    polygon_points = list()
                    if len(polygon_value.get('polygon_vertex_list')) == 1:
                        for polygon_point in polygon_value.get('polygon_vertex_list')[0].get('polygon_vertex'):
                            point = list()
                            point.append(polygon_point.get('x'))
                            point.append(polygon_point.get('y'))
                            polygon_points.append(point)
                    else:
                        print("车牌关键点元素个数不对")
                        print(polygon_value.get('polygon_vertex_list'))
                        exit()
                    shape = {"label": "polygon", "points": polygon_points, "group_id": None, "shape_type": "polygon", "flags": {}, 'text': None}
                    labelme_data.get('shapes').append(shape)
        return labelme_data

    @staticmethod
    def analysis_channel_occupancy_json(content, args):
        # 定义labelme数据结构
        labelme_data = dict(
            version='4.5.14',
            flags={},
            shapes=[],
            imagePath=None,
            imageData=None,
            imageHeight=None,
            imageWidth=None,
            md5Value=None
        )
        alarm_polygon_points = list()
        for key, value in content.items():
            if key == "data":  # 报警多边形框，占用和占用解除，两种情况
                for polygon_key, polygon_value in value.items():
                    if (polygon_key == "occupy" and polygon_value != "") or (polygon_key == "unoccupied" and polygon_value != ""):
                        for polygon_point in polygon_value.values():
                            alarm_polygon_points.extend(polygon_point)
                        shape = {"label": "occupancy", "points": alarm_polygon_points, "group_id": None, "shape_type": "polygon", "flags": {},
                                 'text': content.get("data").get("alarm_message")}
                        labelme_data.get('shapes').append(shape)
            if key == "polygon_vertex_dict" and args.choose_more:  # 根据选择保存多边形框
                for polygon_key, polygon_value in value.items():
                    polygon_points = list()
                    for polygon_point in polygon_value.get('polygonVertexList')[0].get('polygonVertex'):
                        point = list()
                        point.append(polygon_point.get('x'))
                        point.append(polygon_point.get('y'))
                        polygon_points.append(point)
                    if Counter(map(tuple, alarm_polygon_points)) == Counter(map(tuple, polygon_points)):  # 针对重复的多边形框要进行去重
                        pass
                    else:
                        shape = {"label": polygon_key, "points": polygon_points, "group_id": None, "shape_type": "polygon", "flags": {}, 'text': None}
                        labelme_data.get('shapes').append(shape)
        return labelme_data

    @staticmethod
    def analysis_channel_blockage_json(content, args):
        # 定义labelme数据结构
        labelme_data = dict(
            version='4.5.14',
            flags={},
            shapes=[],
            imagePath=None,
            imageData=None,
            imageHeight=None,
            imageWidth=None,
            md5Value=None
        )
        for key, value in content.items():
            if key == "single_object_dict":
                for rectangular_key, rectangular_value in value.items():  # 车辆阻塞的时候，没有去判断阻塞的条件逻辑，alarmMapData
                    for objectRectangle_key, objectRectangle in rectangular_value.items():
                        for obj in objectRectangle:
                            if "alarmMapData" in obj:  # 时间复杂度：O(1)（哈希查找）
                                # print("键 'alarmMapData' 存在！")
                                ltx = obj.get("objectRectangle").get("ltX")
                                lty = obj.get("objectRectangle").get("ltY")
                                rbx = obj.get("objectRectangle").get("rbX")
                                rby = obj.get("objectRectangle").get("rbY")
                                polygon_points = [[ltx, lty], [rbx, rby]]
                                shape = {"label": "blockage", "points": polygon_points, "group_id": None, "shape_type": "rectangle", "flags": {},
                                         'text': content.get("data").get("alarm_message")}
                                labelme_data.get('shapes').append(shape)
                            else:
                                pass
            if key == "polygon_vertex_dict" and args.choose_polygon:
                for polygon_key, polygon_value in value.items():
                    polygon_points = list()
                    for polygon_point in polygon_value.get('polygonVertexList')[0].get('polygonVertex'):
                        point = list()
                        point.append(polygon_point.get('x'))
                        point.append(polygon_point.get('y'))
                        polygon_points.append(point)
                    shape = {"label": polygon_key, "points": polygon_points, "group_id": None, "shape_type": "polygon", "flags": {},
                             'text': content.get("data").get("alarm_message")}
                    labelme_data.get('shapes').append(shape)
        return labelme_data

    @staticmethod
    def analysis_blacked_out_json(content, args):
        # 定义labelme数据结构
        labelme_data = dict(
            version='4.5.14',
            flags={},
            shapes=[],
            imagePath=None,
            imageData=None,
            imageHeight=None,
            imageWidth=None,
            md5Value=None
        )
        alarm_polygon_points = list()
        alarm_message = ""  # 只取报警字段中的多边形坐标，获取更多数据的时候，就是车道阻塞报警的多边形坐标，否则就是车道占用时的多边形坐标
        for key, value in content.items():
            if key == "data":  # 报警多边形框，占用和占用解除，两种情况
                for polygon_key, polygon_value in value.items():
                    if (polygon_key == "occupy" and polygon_value != "") or (polygon_key == "unoccupied" and polygon_value != ""):
                        for polygon_point in polygon_value.values():
                            alarm_polygon_points.extend(polygon_point)
                        shape = {"label": "occupancy", "points": alarm_polygon_points, "group_id": None, "shape_type": "polygon", "flags": {},
                                 'text': content.get("data").get("alarm_message")}
                        labelme_data.get('shapes').append(shape)
                    if polygon_key == "alarm_message":  # 分离报警摄像头唯一标识字符串
                        alarm_message = polygon_value

            if key == "single_object_dict" and args.choose_more:  # 只是通道占用会跟矩形框无关，为了获取更多数据，才要保存更多矩形框，并把矩形框涂黑
                for rectangular_key, rectangular_value in value.items():  # 这里有很多矩形框，是如何判断这个矩形框是属于报警的矩形框还是非报警的矩形框呢
                    for objectRectangle_key, objectRectangle in rectangular_value.items():
                        for obj in objectRectangle:
                            if "alarmMapData" in obj:  # 时间复杂度：O(1)（哈希查找）
                                # print("键 'alarmMapData' 存在！")
                                pass
                            else:
                                # print("键 'alarmMapData' 不存在！")
                                ltx = obj.get("objectRectangle").get("ltX")
                                lty = obj.get("objectRectangle").get("ltY")
                                rbx = obj.get("objectRectangle").get("rbX")
                                rby = obj.get("objectRectangle").get("rbY")
                                polygon_points = [[ltx, lty], [rbx, rby]]
                                shape = {"label": "blacked_out", "points": polygon_points, "group_id": None, "shape_type": "rectangle", "flags": {},
                                         'text': content.get("data").get("alarm_message")}
                                labelme_data.get('shapes').append(shape)
            if key == "polygon_vertex_dict" and args.choose_more:  # 根据选择保存多边形框
                for polygon_key, polygon_value in value.items():
                    # 构造严格匹配的正则模式，确保唯一匹配
                    pattern = r'\b' + re.escape(polygon_key) + r'\b'
                    # (?<!\S)：确保 polygon_key 前面是空格、逗号、方括号等分隔符（或是行首）。
                    # re.escape(polygon_key)：确保 polygon_key 中的 -、_ 等字符不会干扰匹配。
                    # (?!\S)：确保 后面是空格、标点或行尾，防止 C-01M016_1 误匹配 C-01M016_123。
                    # pattern = r'(?<!\S)' + re.escape(polygon_key) + r'(?!\S)'  # 唯一匹配
                    if re.search(pattern, alarm_message):
                        # print(f"匹配成功: {polygon_key}")
                        polygon_points = list()
                        for polygon_point in polygon_value.get('polygonVertexList')[0].get('polygonVertex'):
                            point = list()
                            point.append(polygon_point.get('x'))
                            point.append(polygon_point.get('y'))
                            polygon_points.append(point)
                        shape = {"label": polygon_key, "points": polygon_points, "group_id": None, "shape_type": "polygon", "flags": {}, 'text': None}
                        labelme_data.get('shapes').append(shape)
                    else:
                        # print(f"未匹配: {polygon_key}")
                        pass
        return labelme_data

    def labelme2excel(self):
        """
        根据用户输入的列号自动填充Excel表格，兼容xlsx和xls。
        根据单元格大小自动调整图片尺寸，保持比例。
        Excel文件路径（支持xlsx和xls）
        """
        # 将单引号替换为双引号，并将反斜杠转义
        json_str = self.excel_file.replace("'", '"').replace("\\", "\\\\")
        # 解析 JSON 字符串
        json_data = json.loads(json_str)
        # 获取 input_txt_dir 的值
        input_excel_dir = json_data[0]['input_excel_dir']
        # 判断文件格式并转换xls为xlsx
        if input_excel_dir.endswith('.xls'):
            print("检测到xls文件，正在转换为xlsx...")
            input_excel_dir = self.convert_xls_to_xlsx(input_excel_dir)
        # 打开Excel文件
        wb = openpyxl.load_workbook(input_excel_dir)  # excel表格工作簿对象
        ws = wb.active  # 代表Sheet1
        # 映射列字母到数字（A -> 1, B -> 2, ...）
        # column_letter_to_index = {openpyxl.utils.get_column_letter(i): i for i in range(1, ws.max_column + 1)}
        # 使用列字母直接映射，而不依赖现有的最大列数，column_letter_to_index = {letter: idx + 1 for idx, letter in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZAAABAC")}
        column_letter_to_index = {letter: idx + 1 for idx, letter in enumerate(string.ascii_uppercase)}
        # 确保传入的标题数量与 columns_fill 长度一致
        if len(self.column_titles) != len(self.columns_fill):
            raise ValueError("列标题的数量必须与 columns_fill 的列数一致！")

        # 添加列标题到 start_row = 2 的这一行
        for idx, col in enumerate(self.columns_fill):
            col_idx = column_letter_to_index.get(col)  # 根据列字母获取列的索引
            if col_idx is None:
                print(f"列 {col} 不存在！")
                exit()
            # 将相应标题写入第二行
            ws.cell(row=2, column=col_idx).value = self.column_titles[idx]

        # 合并 ['A', 'B', 'C', 'D'] 单元格
        if len(self.columns_fill) >= 4:
            start_col = self.columns_fill[0]
            end_col = self.columns_fill[3]
            ws.merge_cells(f'{start_col}1:{end_col}1')  # 合并第1行的 A-D 单元格
            ws.cell(row=1, column=column_letter_to_index[start_col]).value = self.time
            ws.cell(row=1, column=column_letter_to_index[start_col]).alignment = openpyxl.styles.Alignment(horizontal='center')

        # 从第3行开始，逐行填充
        start_row = 3
        valid_rows = []
        # 需要给个进度
        for idx, dataset in enumerate(self.datasets):
            if self.background:  # 只插入背景图片
                # 遍历每一列
                for col in self.columns_fill:
                    col_idx = column_letter_to_index[col]  # 根据列字母获取列的索引
                    if col_idx is None:
                        print(f"列 {col} 不存在！")
                        exit()
                    content_type = self.column_content_type.get(col, None)  # 获取该列对应的内容类型
                    if content_type == 'image':  # 如果是图片
                        img_path = dataset.get("full_path")
                        if img_path:
                            # 调整图片到黄金比例尺寸
                            img = self.adjust_image_to_cell(img_path, ws, ws.cell(row=start_row + idx, column=col_idx), target_width=133, target_height=75)
                            ws.add_image(img, f'{col}{start_row + idx}')  # 插入图片
                    elif content_type == 'car_number':  # 如果是车牌号
                        for shape in dataset.get("labelme_info").get("shapes"):
                            if len(shape.get("points")) == 4 and shape.get("flags").get('other'):  # 人工筛选后勾选other=true时，保存车牌，即为确定事件的车牌号
                                car_number = shape.get("text")  # 获取车牌号，如果未提供，则默认 "AE3456"
                                ws[f'{col}{start_row + idx}'] = car_number  # 插入车牌号
                    elif content_type == 'event':
                        for shape in dataset.get("labelme_info").get("shapes"):
                            # 如果 'flags' 存在且不为空，执行逻辑
                            if shape.get("flags") and shape.get("flags").get('other') is False:
                                keys = list(shape.get('flags').keys())
                                result = '到'.join(keys)  # 将键列表通过 '到' 拼接成字符串
                                # ws[f'{col}{start_row + idx}'] = result  # 插入图片名称，暂时不写入事件
                    elif content_type == 'conclusion':
                        flags_sys = dataset.get("labelme_info").get("flags")
                        conclusion = list()
                        for key, value in flags_sys.items():
                            if value and key != '1':
                                conclusion.append(key)
                        ws[f'{col}{start_row + idx}'] = ', '.join(conclusion)  # 以逗号分隔多个值,插入勾选结论
            if self.fill_rules:  # 保存系统上报图片，存在重复代码需要合并，临时不解决，以后优化解决
                if dataset.get("background"):
                    # 遍历每一列
                    for col in self.columns_fill:
                        col_idx = column_letter_to_index[col]  # 根据列字母获取列的索引
                        if col_idx is None:
                            print(f"列 {col} 不存在！")
                            exit()
                        content_type = self.column_content_type.get(col, None)  # 获取该列对应的内容类型
                        if content_type == 'image':  # 如果是图片
                            img_path = dataset.get("full_path")
                            if img_path:
                                # 调整图片到黄金比例尺寸
                                img = self.adjust_image_to_cell(img_path, ws, ws.cell(row=start_row + idx, column=col_idx), target_width=133, target_height=75)
                                ws.add_image(img, f'{col}{start_row + idx}')  # 插入图片
                        elif content_type == 'car_number':  # 如果是车牌号
                            for shape in dataset.get("labelme_info").get("shapes"):
                                if len(shape.get("points")) == 4:
                                    car_number = shape.get("text")  # 获取车牌号，如果未提供，则默认 "AE3456"
                                    ws[f'{col}{start_row + idx}'] = car_number  # 插入车牌号
                        elif content_type == 'event':
                            for shape in dataset.get("labelme_info").get("shapes"):
                                # 如果 'flags' 存在且不为空，执行逻辑
                                if shape.get("flags"):
                                    keys = list(shape.get('flags').keys())
                                    result = '到'.join(keys)  # 将键列表通过 '到' 拼接成字符串
                                    ws[f'{col}{start_row + idx}'] = result  # 插入图片名称
                        elif content_type == 'conclusion':
                            flags_sys = dataset.get("labelme_info").get("flags")
                            conclusion = list()
                            for key, value in flags_sys.items():
                                if value and key != '1':
                                    conclusion.append(key)
                            ws[f'{col}{start_row + idx}'] = ', '.join(conclusion)  # 以逗号分隔多个值,插入勾选结论
                else:
                    pass
            else:  # 保存模型上报图片
                # 获取矩形框内部多边形对象和ROI外部多边形对象
                inner_polygon_points, outer_polygon_points = self.filter_polygon_shape(dataset)
                if len(inner_polygon_points) == 1 and len(outer_polygon_points) == 1:
                    # 把多边形标注，将坐标转换为 Shapely Polygon 对象
                    # Polygon.intersects()用于判断两个多边形是否相交。如果两个多边形有任何部分重叠，那么这个方法会返回 True。
                    # Polygon.within()从较小的多边形的角度来看，判断它是否在较大的多边形内部。
                    # Polygon.equals()判断两个多边形是否相等。
                    # Polygon.contains()从较大的多边形的角度来看，判断它是否包含较小的多边形。
                    inner_polygon = Polygon(inner_polygon_points[0].get('points'))
                    outer_polygon = Polygon(outer_polygon_points[0].get('points'))
                    if inner_polygon.within(outer_polygon):  # 满足车牌多边形和OUI画框多边形有值，且车牌在ROI画框范围内的情况下才进行excel表格行列的遍历，并写入图片和车牌号
                        condition_idx = idx
                        valid_rows.append(dataset)  # 保存有效行
                        # 遍历每一列
                        for col in self.columns_fill:
                            col_idx = column_letter_to_index[col]  # 根据列字母获取列的索引
                            if col_idx is None:
                                print(f"列 {col} 不存在！")
                                exit()
                            content_type = self.column_content_type.get(col, None)  # 获取该列对应的内容类型
                            if content_type == 'image':  # 如果是图片
                                img_path = dataset.get("full_path")
                                if img_path:
                                    # img = self.adjust_image_to_cell(img_path, ws, ws.cell(row=start_row + idx, column=col_idx))
                                    # 调整图片到黄金比例尺寸
                                    img = self.adjust_image_to_cell(img_path, ws, ws.cell(row=start_row + condition_idx, column=col_idx), target_width=133,
                                                                    target_height=75)
                                    ws.add_image(img, f'{col}{start_row + condition_idx}')  # 插入图片
                            elif content_type == 'car_number':  # 如果是车牌号
                                car_number = inner_polygon_points[0].get('text')  # 获取车牌号，如果未提供，则默认 "AE3456"
                                ws[f'{col}{start_row + condition_idx}'] = car_number  # 插入车牌号
                            elif content_type == 'event':
                                for shape in dataset.get("labelme_info").get("shapes"):
                                    # 如果 'flags' 存在且不为空，执行逻辑
                                    if shape.get("flags"):
                                        keys = list(shape.get('flags').keys())
                                        result = '到'.join(keys)  # 将键列表通过 '到' 拼接成字符串
                                        ws[f'{col}{start_row + idx}'] = result  # 插入图片名称
                            elif content_type == 'conclusion':
                                flags = dataset.get("labelme_info").get("flags")
                                conclusion = list()
                                for key, value in flags.items():
                                    if value and key != '1':
                                        conclusion.append(key)
                                ws[f'{col}{start_row + condition_idx}'] = ', '.join(conclusion)  # 以逗号分隔多个值,插入勾选结论
                else:
                    if len(inner_polygon_points) != 1:
                        print(f'该图片有两块车牌，或者没有车牌，请检查：{dataset.get("full_path")}')
                    if len(outer_polygon_points) != 1:
                        print(f'该图片存在多个多边形，或者为空，请检查：{dataset.get("full_path")}')

        # 删除空行,没有用
        # for row in range(ws.max_row, start_row - 1, -1):
        #     if row not in valid_rows:
        #         ws.delete_rows(row)
        # 保存修改后的文件
        self.save_excel(wb, input_excel_dir)
        if self.output_dir:  # 如果输出路径不为空则保存
            self.save_labelme(valid_rows, self.output_dir, None)

    @staticmethod
    def convert_xls_to_xlsx(xls_file):
        """将xls文件转换为xlsx"""
        xls_wb = xlrd.open_workbook(xls_file)
        xlsx_file = xls_file.replace('.xls', '.xlsx')

        # 创建一个新的xlsx工作簿
        xlsx_wb = openpyxl.Workbook()
        xlsx_ws = xlsx_wb.active

        xls_sheet = xls_wb.sheet_by_index(0)

        # 将xls内容写入新的xlsx工作簿
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_ws.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)

        # 保存为xlsx格式
        xlsx_wb.save(xlsx_file)
        print(f"Excel文件格式被更新，由.xls更新为.xlsx：{xlsx_file}")
        return xlsx_file

    @staticmethod
    def adjust_image_to_cell(img_path, ws, cell, target_width, target_height):

        """根据单元格大小自动调整图片尺寸，保持图片比例不失真"""
        # 使用 Pillow 来打开图片并获取尺寸
        pil_img = PILImage.open(img_path)
        img_width, img_height = pil_img.size

        # 计算图片的宽高比，确保图片按比例缩放，适应目标宽高，而不失真
        ratio = min(target_width / img_width, target_height / img_height)

        # 根据计算的比例，调整图片尺寸
        new_width = int(img_width * ratio)
        new_height = int(img_height * ratio)

        # 关闭 Pillow 打开的图片
        pil_img.close()

        # 使用 openpyxl 的 Image 插入图片，并设置调整后的尺寸
        img = OpenpyxlImage(img_path)
        img.width = new_width  # 设置调整后的宽度
        img.height = new_height  # 设置调整后的高度
        pixel_per_col_unit = target_width / 16.5  # Excel 单元格的 列宽度为 16.4
        pixel_per_row_unit = target_height / 55  # Excel 单元格的 行高度为 55
        # Excel列宽度的单位是字符单位，大约等于7像素，所以需要将目标宽度除以7进行转换
        ws.column_dimensions[cell.column_letter].width = target_width / pixel_per_col_unit  # 调整列宽度;这里的计算公式为：640/x=79.7，得到x=8.03
        ws.row_dimensions[cell.row].height = target_height / pixel_per_row_unit  # 调整行高度，单位为像素；这里的计算公式为：480/x=270，得到x=1.78

        return img

    @staticmethod
    def save_excel(wb, file_path):
        """
        保存 Excel 文件，如果文件被占用，提示用户关闭文件后再重试。
        """
        try:
            wb.save(file_path)
            print(f"Excel文件已成功保存：{file_path}")
        except PermissionError:
            print(f"无法保存文件：{file_path}，文件可能已被打开，请关闭文件后重试。")

    @staticmethod
    def filter_polygon_shape(dataset, outer_polygon_shape=None):
        """
        取出车牌多边形（4个点）和画框的ROI多边形（大于4个点）
        @param dataset:
        @param outer_polygon_shape:
        @return:
        """
        inner_polygon_shape = list()  # 矩形内多边形的标注对象
        outer_polygon_shape = list()  # ROI外部多边形的标注对象
        if dataset.get('background') is True:
            for shape in dataset.get('labelme_info').get('shapes'):
                if shape.get('shape_type') == 'polygon':
                    if len(shape.get("points")) == 4:
                        inner_polygon_shape.append(shape)
                    else:
                        outer_polygon_shape.append(shape)
        return inner_polygon_shape, outer_polygon_shape
