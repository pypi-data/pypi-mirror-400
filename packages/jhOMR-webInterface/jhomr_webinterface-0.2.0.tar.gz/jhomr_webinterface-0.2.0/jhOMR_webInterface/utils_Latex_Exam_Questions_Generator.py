


###########################################################################
# Latex and question processing related functions
###########################################################################

from fastapi import FastAPI, HTTPException
import subprocess
import os
import random
import string
import re
import numpy as np
import itertools
import shutil
from openpyxl import load_workbook
from pathlib import Path
import re
import html



from pydantic import BaseModel
from typing import List
from fastapi.responses import JSONResponse
###########################################################################

# # -------------------------------
# # Parse LaTeX questions
# # -------------------------------


def remove_comment_lines(text: str) -> str:
    """
    Remove full-line LaTeX comments that start with %
    but keep inline % symbols (like in math or text) intact.
    """
    lines = text.splitlines()
    cleaned_lines = [line for line in lines if not line.strip().startswith('%')]
    return "\n".join(cleaned_lines).strip()


def parse_all_latex_questions_from_question_bank(file_path):
    """Parse LaTeX questions (MCQ + True/False) from question bank."""
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Split all \question blocks
    questions_raw = re.split(r'\\question', content)
    parsed_mcq = []
    parsed_tf = []

    for q in questions_raw:
        q = q.strip()
        if not q:
            continue

        # Extract metadata inside square brackets
        meta_match = re.match(r'\[(.*?)\]', q)
        q_type = "mcq"  # default if not specified
        correct_answer = None
        mark = None

        if meta_match:
            metadata = meta_match.group(1)

            # Parse type, answer, and mark
            type_search = re.search(r'type=(\w+)', metadata)
            if type_search:
                q_type = type_search.group(1).lower()

            ans_search = re.search(r'answer=([A-Za-z])', metadata)
            if ans_search:
                correct_answer = ans_search.group(1)

            mark_search = re.search(r'mark=(\d+)', metadata)
            if mark_search:
                mark = int(mark_search.group(1))

            # Remove metadata from the question text
            q = re.sub(r'^\[.*?\]', '', q).strip()

        # === MCQ type ===
        if q_type == "mcq":
            parts = re.split(r'\\begin{choices}', q)
            if len(parts) < 2:
                continue  # skip malformed MCQs

            question_text = remove_comment_lines(parts[0].strip())
            choices_block = remove_comment_lines(parts[1])
    
            # Extract each \choice line
            choices = re.findall(r'\\choice\s+(.*)', choices_block)

            parsed_mcq.append({
                "type": q_type,
                "question": question_text,
                "choices": choices,
                "answer": correct_answer,
                "mark": mark
            })

        # === True/False type ===
        elif q_type == "tf":
            # 1. Cleanly extract only the question statement. 
            # Since TF questions don't use \begin{choices}, 
            # we take all the remaining content of 'q' as the statement.
            question_text = remove_comment_lines(q.strip())

            # 2. Add an explicit check to strip any remaining \end{choices} or other junk 
            # if the question bank is messy. (Defensive coding)
            question_text = re.sub(r'\\end\{choices\}', '', question_text).strip()
            
            parsed_tf.append({
                "type": q_type,
                "question": question_text,
                "choices": [], # Use an empty list to trigger the 'a) True \quad b) False' display
                "answer": correct_answer,
                "mark": mark
            })

    return parsed_mcq, parsed_tf


def generate_balanced_overlap_sets(selected_tf_set_1, num_additional_sets=2, num_trials=2000):
    # This function shuffles the true/false answers among multiple sets.
    # It does that in such a way that there will be minimum number of overlap (same answer for a given question number) between any two sets.
    
    # Extract the correct answers for Set-1 from the list of dictionaries
    Set_1_tf_correct_answers = [q['answer'] for q in selected_tf_set_1]
    
    n = len(Set_1_tf_correct_answers)

    # 🔹 Prevent infinite loop for trivial cases
    if n == 0:
        # No questions at all → return empty structures
        tf_questions_all_sets = [[] for _ in range(num_additional_sets + 1)]
        best_indices = [[] for _ in range(num_additional_sets)]
        tf_answer_overlap_between_sets = [
            (f"Set-{i}, Set-{j}", 0)
            for i in range(1, num_additional_sets + 2)
            for j in range(i + 1, num_additional_sets + 2)
        ]
        return tf_questions_all_sets, best_indices, tf_answer_overlap_between_sets

    elif n == 1:
        # One question only → replicate same across sets
        best_sets = [Set_1_tf_correct_answers.copy() for _ in range(num_additional_sets)]
        best_indices = [[0] for _ in range(num_additional_sets)]
        tf_answer_overlap_between_sets = [
            (f"Set-{i}, Set-{j}", 1)
            for i in range(1, num_additional_sets + 2)
            for j in range(i + 1, num_additional_sets + 2)
        ]
        tf_questions_all_sets = [selected_tf_set_1] + [
            [selected_tf_set_1[0]] for _ in range(num_additional_sets)
        ]
        return tf_questions_all_sets, best_indices, tf_answer_overlap_between_sets

    
    # -------------------------------------------------------------
    # 1. Helper function: Count matches (Overlap)
    # -------------------------------------------------------------
    def count_matches(a, b):
        return sum(i == j for i, j in zip(a, b))

    # -------------------------------------------------------------
    # 2. Nested function: Find the best permutation of answers
    # -------------------------------------------------------------
    def generate_best_shuffled_set():
        n = len(Set_1_tf_correct_answers)
        best_sets = []
        best_indices = []
        min_max_overlap = float('inf')
        
        for _ in range(num_trials):
            candidates = []
            indices_list = []
            
            for _ in range(num_additional_sets):
                permuted_indices = list(range(n))
                attempts = 0
                MAX_ATTEMPTS = 50  # Set a safe limit
                
                while attempts < MAX_ATTEMPTS:
                    random.shuffle(permuted_indices)
                    shuffled = [Set_1_tf_correct_answers[i] for i in permuted_indices]  
                    if shuffled != Set_1_tf_correct_answers:
                        break
                    attempts += 1
                if attempts == MAX_ATTEMPTS:
                    print(f"Warning: Could not find a unique shuffle after {MAX_ATTEMPTS} attempts. Accepting original set.")
                candidates.append(shuffled)
                indices_list.append(permuted_indices)
                
            all_sets = [Set_1_tf_correct_answers] + candidates
            pairwise_overlaps = []
            
            for i in range(len(all_sets)):
                for j in range(i + 1, len(all_sets)):
                    overlap = count_matches(all_sets[i], all_sets[j])
                    pairwise_overlaps.append(overlap)
                    
            max_overlap = max(pairwise_overlaps)
            
            if max_overlap < min_max_overlap:
                min_max_overlap = max_overlap
                best_sets = candidates
                best_indices = indices_list 
                
        # REMOVED: min_max_overlap from return
        return best_sets, best_indices
    
    # -------------------------------------------------------------
    # 3. Nested function: Compute all pairwise overlaps for the best sets
    # -------------------------------------------------------------
    def compute_overlap_between_sets(best_sets):
        tf_answer_overlap_between_sets = []
        
        for set_pair in itertools.combinations(range(num_additional_sets + 1), 2): 
            
            final_all_sets = [Set_1_tf_correct_answers] + best_sets

            index_a, index_b = set_pair
            set_a = final_all_sets[index_a]
            set_b = final_all_sets[index_b]
            
            overlap_count = count_matches(set_a, set_b)
            
            label_a = f'Set-{index_a + 1}'
            label_b = f'Set-{index_b + 1}'
            
            tf_answer_overlap_between_sets.append((f'{label_a}, {label_b}', overlap_count))
            
        # REMOVED: min_max_overlap from return
        return tf_answer_overlap_between_sets 

    # -------------------------------------------------------------
    # 4. NEW Nested function: Use indices to permute the question list
    # -------------------------------------------------------------
    def create_shuffled_question_sets(best_indices):
        shuffled_tf_sets = []
        
        for permutation_indices in best_indices:
            shuffled_set = [selected_tf_set_1[i] for i in permutation_indices]
            shuffled_tf_sets.append(shuffled_set)
            
        return shuffled_tf_sets


    # -------------------------------------------------------------
    # Main function execution flow:
    # -------------------------------------------------------------
    
    # Step A: Find the best answer permutations
    #print("trying to generate best shuffled set")
    best_sets_answers, best_indices = generate_best_shuffled_set() 
    
    #print("Computing overlap")
    # Step B: Calculate final overlaps using the best answer sets
    tf_answer_overlap_between_sets = compute_overlap_between_sets(best_sets_answers)
    #print("Creating shuffled question sets")
    # Step C: Create the final shuffled question lists (list of dictionaries)
    shuffled_tf_sets = create_shuffled_question_sets(best_indices)
    
    #print("combining with set-1")
    # Combine the Set-1 and the newly created sets (Set-2 and Set-3) in a single list
    tf_questions_all_sets = [selected_tf_set_1] + shuffled_tf_sets
    # Return all required results (removed min_max_overlap)
    return tf_questions_all_sets, best_indices, tf_answer_overlap_between_sets



def generate_shuffled_mcq_sets(selected_mcq_set_1, num_additional_sets=2):
    """
    Generates new MCQ sets by shuffling the order of the choices and
    **correctly updates the 'answer' key** to reflect the new position
    of the correct choice.
    """
    
    # Function to convert 0 -> 'a', 1 -> 'b', etc.
    def index_to_letter(index):
        # We use string.ascii_lowercase for guaranteed a-z
        return string.ascii_lowercase[index]

    mcq_questions_all_sets = [selected_mcq_set_1]
    
    # Total number of sets to generate (excluding Set-1)
    for _ in range(num_additional_sets):
        
        shuffled_set = []
        
        # Iterate through each question in the original Set-1 sequence
        for original_q in selected_mcq_set_1:
            
            new_q = original_q.copy()
            
            # --- 1. Find the content of the correct choice in Set-1 ---
            
            # Convert the answer letter ('a', 'b', etc.) back to an index (0, 1, etc.)
            original_answer_index = ord(original_q['answer']) - ord('a')
            
            # Get the content of the correct choice (e.g., 'Force' or '3 x 10^8 m/s')
            correct_choice_content = original_q['choices'][original_answer_index]
            
            # --- 2. Shuffle the choices ---
            
            choices_to_shuffle = new_q['choices'][:] # Use a copy for safe shuffling
            random.shuffle(choices_to_shuffle)
            new_q['choices'] = choices_to_shuffle
            
            # --- 3. Find the new index of the correct content ---
            
            # Find the new index of the correct content in the shuffled list
            new_answer_index = choices_to_shuffle.index(correct_choice_content)
            
            # --- 4. Update the 'answer' key ---
            
            # Convert the new index back to a letter
            new_q['answer'] = index_to_letter(new_answer_index)
            
            # 5. Add the newly shuffled question to the current set
            shuffled_set.append(new_q)
            
        mcq_questions_all_sets.append(shuffled_set)
        
    return mcq_questions_all_sets



# ###################################################################
# #Following function is responsible for processing the latex code
# ###################################################################



def sanitize_for_latex(text):
    """
    Cleans text for LaTeX safety:
    - Escapes only text-mode specials
    - Does NOT escape inside math mode (\(...\), $...$, \frac, etc.)
    - Wraps comparison symbols (> < =) in math mode if found outside math
    """
    if not text:
        return ""

    text = html.unescape(text).strip()

    # Split text into math and non-math parts
    # We'll detect math segments (anything between \( ... \), $...$, or \[ ... \])
    math_segments = []
    def math_replacer(match):
        math_segments.append(match.group(0))
        return f"@@MATH{len(math_segments)-1}@@"

    text = re.sub(r'(\\\([^\)]*\\\)|\\\[[^\]]*\\\]|\$[^$]*\$)', math_replacer, text)

    # Escape only outside math mode
    specials = {
        '&': r'\&', '%': r'\%', '$': r'\$', '#': r'\#', '_': r'\_',
        '{': r'\{', '}': r'\}', '~': r'\textasciitilde{}',
        '^': r'\textasciicircum{}',
        # NOTE: don't escape backslashes globally anymore!
    }
    for k, v in specials.items():
        text = text.replace(k, v)

    # Wrap comparison symbols (> < =) in math mode (outside math)
    def wrap_symbol(m):
        sym = m.group(0)
        return f'\\({sym}\\)'
    text = re.sub(r'(?<![\\$])([><=])(?!=)', wrap_symbol, text)

    # Restore original math segments
    for i, seg in enumerate(math_segments):
        text = text.replace(f"@@MATH{i}@@", seg)

    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text)

    return text




def create_question_latex_code_and_pdf(
    Department,
    exam_title,
    latex_template_file,
    Question_Name,
    selected_mcq_all_sets=[],
    selected_tf_all_sets=[],
    full_mark=0,
    exam_duration="0 min",
    tf_neg=50,
    mcq_neg=25,
    base_folder=None,  # ✅ new argument
):
    """
    Creates LaTeX exam sets and compiles them into PDFs.
    All output files will be stored under base_folder (passed from main.py).
    """

    # ✅ Default fallback if not provided
    if base_folder is None:
        base_folder = Path.home() / "Documents" / "WebCam_OMR"

    # ---------------------------
    # Utility: Format Question
    # ---------------------------
    def format_question_latex(question_data):
        question_raw = question_data.get("question", "")
        question = sanitize_for_latex(question_raw)

        q_type = question_data.get("type", "mcq").lower()
        choices = [sanitize_for_latex(c) for c in question_data.get("choices", [])]
        mark = question_data.get("mark", "")

        if q_type == "mcq":
            MARK_WIDTH = "0.25in"
            GAP_WIDTH = "0.1in"
            COLUMN_WIDTH = f"\\dimexpr\\linewidth-{MARK_WIDTH}-{GAP_WIDTH}\\relax"

            formatted_choices = " ".join(
                [
                    f"{chr(97+i)}) {choice}"
                    + ("" if i == len(choices) - 1 else " \\quad")
                    for i, choice in enumerate(choices)
                ]
            )

            latex = (
                "\\item \n"
                f"\\begin{{tabular}}[t]{{@{{}}p{{{COLUMN_WIDTH}}}@{{\\hspace{{{GAP_WIDTH}}}}}r@{{}}}}\n"
                f"{question}\n"
                "\\par\\vspace{1ex}\n"
                f"\\noindent {formatted_choices} & \n"
                f"\\makebox[{MARK_WIDTH}][r]{{\\textbf{{{mark}}}}}\n"
                "\\end{tabular}\n"
            )
        else:
            latex = "\\item " + question + "\n"

        return latex

    # ---------------------------
    # Utility: Create and Compile LaTeX
    # ---------------------------
    def create_latex_code_and_run(
        latex_template_file,
        output_tex_file,
        exam_title,
        mcq_image_path,
        selected_mcq,
        selected_tf,
        full_mark,
        exam_duration,
        tf_neg,
        mcq_neg,
    ):
        num_tf_questions = len(selected_tf)
        formatted_mcq = [format_question_latex(q) for q in selected_mcq]
        formatted_tf = [format_question_latex(q) for q in selected_tf]

        with open(latex_template_file, "r", encoding="utf-8") as f:
            latex_content = f.read()

        latex_content = latex_content.replace("%MCQ_PLACEHOLDER%", "".join(formatted_mcq))
        latex_content = latex_content.replace("%TF_PLACEHOLDER%", "".join(formatted_tf))
        latex_content = latex_content.replace("Exam Title Here", exam_title)
        latex_content = latex_content.replace("example-image.png", mcq_image_path.replace("\\", "/"))
        latex_content = latex_content.replace("{placeholder_full_mark}", str(full_mark))
        latex_content = latex_content.replace("{placeholder_time}", exam_duration)
        latex_content = latex_content.replace("tf_neg_placeholder", f"{tf_neg}\\%")
        latex_content = latex_content.replace("mcq_neg_placeholder", f"{mcq_neg}\\%")

        tf_total_value = num_tf_questions * 1
        latex_content = latex_content.replace(
            "(no of tf questions $ \\times $ 1 = value)",
            f"({num_tf_questions} $\\times$ 1 = {tf_total_value})",
        )

        with open(output_tex_file, "w", encoding="utf-8") as f:
            f.write(latex_content)

        print(f"✅ LaTeX file created: {output_tex_file}")

        try:
            tex_dir = os.path.dirname(output_tex_file)
            tex_filename = os.path.basename(output_tex_file)

            subprocess.run(
                ["pdflatex", "-interaction=nonstopmode", tex_filename],
                check=True,
                cwd=tex_dir,
            )
            print(f"✅ Successfully compiled {output_tex_file} and PDF is generated.")
        except subprocess.CalledProcessError:
            print("❌ Error: pdflatex compilation failed. Check LaTeX installation.")

    # ---------------------------
    # Prepare Output Folder
    # ---------------------------
    output_folder = base_folder / "Generated_Exams"
    output_folder.mkdir(parents=True, exist_ok=True)

    # Optional cleanup
    for ext in ("*.tex", "*.pdf", "*.aux", "*.log"):
        for file_path in output_folder.glob(ext):
            try:
                file_path.unlink()
            except PermissionError:
                print(f"❌ Cannot delete {file_path}, it may be in use.")

    # ---------------------------
    # Generate PDFs
    # ---------------------------
    output_pdf_file_names = []
    project_root = Path.cwd()
#    project_root = Path(__file__).parent.parent

    for index, (selected_mcq, selected_tf) in enumerate(
        zip(selected_mcq_all_sets, selected_tf_all_sets)
    ):
        output_tex_file = f"{Question_Name}_Set_{index+1}.tex"
        output_tex_path = output_folder / output_tex_file

        mcq_image_path = project_root / "static" / f"{Department}_Set_{index+1}.png"
        mcq_image_path_str = str(mcq_image_path.resolve())
        
        create_latex_code_and_run(
            latex_template_file,
            str(output_tex_path),
            exam_title,
            mcq_image_path_str,
            selected_mcq,
            selected_tf,
            full_mark,
            exam_duration,
            tf_neg,
            mcq_neg,
        )

        output_pdf_file_names.append(output_tex_file.replace(".tex", ".pdf"))

    return output_folder,output_pdf_file_names




def generate_set_wise_answers_and_marks(questions_all_sets,WA_neg_mark_percentage):
    answers_all_sets = []
    for question_set in questions_all_sets:    
        answers_for_current_set = []    
        selected_CA_marks=[]
        selected_WA_marks=[]
        for q in question_set:
            answers_for_current_set.append(q['answer'])     
            selected_CA_marks.append(float(q['mark']))
            selected_WA_marks.append(-float(np.round(float(q['mark'])*WA_neg_mark_percentage/100,2)))
            
        answers_all_sets.append(answers_for_current_set)
    return answers_all_sets,selected_CA_marks,selected_WA_marks



# ######################################################################
# ######################################################################



def create_correct_answer_excel(
    xlsx_template_file,
    output_xlsx_file,
    course_title,
    tf_answers_all_sets=[],
    selected_tf_CA_marks=[],
    selected_tf_WA_marks=[],
    mcq_answers_all_sets=[],
    selected_mcq_CA_marks=[],
    selected_mcq_WA_marks=[] ):
        
    def write_to_excel(list_of_values,excel_path,col='A',course_name=""):
        col_dict={'C':3,'D':4,'E':5,'F':6,'G':7,'J':10,'K':11,'L':12,'M':13,'N':14}
        col_num=col_dict[col]
        # Load the existing Excel workbook
        
        wb = load_workbook(excel_path)
        # Select the active sheet
        ws = wb.active  # Or specify: ws = wb["Sheet1"]


        # Start from row 2
        start_row = 2
        # Write values to a specific column
        for i, answer in enumerate(list_of_values):
            ws.cell(row=start_row + i, column=col_num, value=answer)

        if course_name != "":
            ws.cell(row=start_row, column=1, value=course_name)
        
        # Save workbook
        wb.save(excel_path)
        print("\nValues written to Excel column:", col)

    shutil.copyfile(xlsx_template_file, output_xlsx_file)
    
    write_to_excel(tf_answers_all_sets[0],output_xlsx_file,col='C')   
    write_to_excel(tf_answers_all_sets[1],output_xlsx_file,col='D')   
    write_to_excel(tf_answers_all_sets[2],output_xlsx_file,col='E')        
    write_to_excel(selected_tf_CA_marks,output_xlsx_file,col='F')
    write_to_excel(selected_tf_WA_marks,output_xlsx_file,col='G')
    
    write_to_excel(mcq_answers_all_sets[0],output_xlsx_file,col='J')
    write_to_excel(mcq_answers_all_sets[1],output_xlsx_file,col='K')
    write_to_excel(mcq_answers_all_sets[2],output_xlsx_file,col='L')  

    write_to_excel(selected_mcq_CA_marks,output_xlsx_file,col='M')
    write_to_excel(selected_mcq_WA_marks,output_xlsx_file,col='N',course_name=course_title)
    



# ######################################################################
# ######################################################################


# ###########################################################################
# # Webpage related functions
# ###########################################################################

# app = FastAPI()

async def load_question_bank(base_folder: Path, course_code: str):
    """
    Loads the LaTeX question bank file based on department and course code.
    The base_folder is provided by main.py.
    """
    question_bank_folder = base_folder / "Question_Bank"
    file_path = question_bank_folder / f"{course_code}.tex"

    # ✅ Ensure directory exists
    question_bank_folder.mkdir(parents=True, exist_ok=True)

    if not file_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Question bank file not found in {question_bank_folder}. "
                   f"Copy Question Bank (.tex file) in that directory."
        )

    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    return {"content": content}



# ------------------ Pydantic model for exam generation ------------------
class ExamGenerationRequest(BaseModel):
    department: str
    courseCode: str
    examType: str
    courseTitle: str
    fullMark: int
    duration: int
    mcqNeg: int
    tfNeg: int
    randomSelect: bool
    numTf: int
    numMcq: int
    selectedTfQuestions: List[int]
    selectedTfMarks: List[int]
    selectedMcqQuestions: List[int]
    selectedMcqMarks: List[int]



# ------------------ Exam generation handler ------------------
async def generate_exam(request_data: ExamGenerationRequest, base_folder: Path):
    """
    Handles the exam generation request.
    Generates exam sets and returns status messages.
    Paths are derived from base_folder provided by main.py.
    """
    import traceback
    messages = []
    print("Generate_exam called with base_folder=", base_folder)
    
    try:
        #--------------------------------------------#
        Department = request_data.department
        Course_Code = re.sub(r"\s+", "", request_data.courseCode)  # remove spaces
        Exam_type = request_data.examType
        Course_title = request_data.courseTitle
        full_mark = float(request_data.fullMark)
        exam_duration = str(request_data.duration) + " mins"
        mcq_neg_percentage = int(request_data.mcqNeg)
        tf_neg_percentage = int(request_data.tfNeg)
        
        Random_Selection = request_data.randomSelect
        num_tf_to_select = int(request_data.numTf)
        num_mcq_to_select = int(request_data.numMcq)
        Selected_TF_Questions = request_data.selectedTfQuestions
        Selected_MCQ_Questions = request_data.selectedMcqQuestions
        TF_Marks = request_data.selectedTfMarks
        MCQ_Marks = request_data.selectedMcqMarks

        #--------------------------------------------#
        if Course_title.strip():
            exam_title = f"{Exam_type} on {Course_Code}: {Course_title}"
        else:
            exam_title = f"{Exam_type} on {Course_Code}"

        Question_Name = f"{Exam_type}_{Course_Code}"
        question_bank_fileName = f"{Course_Code}.tex"

        # ✅ All paths now derived from base_folder
        question_bank_folder = base_folder / "Question_Bank"
        question_bank_folder.mkdir(parents=True, exist_ok=True)
        question_bank_file = question_bank_folder / question_bank_fileName
        
        generated_exams_dir = base_folder / "Generated_Exams"
        generated_exams_dir.mkdir(parents=True, exist_ok=True)


        # static assets (e.g. templates) can still come from your working directory
        latex_template_file = Path("static") / "latex_template.tex"
        correct_answers_template_file = Path("static") / "Answer_sheet_template.xlsx"
        correct_answer_excel_FN="Correct_Answers_"+Course_Code+".xlsx"
        correct_answers_output_file = generated_exams_dir / correct_answer_excel_FN
        #correct_answers_output_file = os.path.join(base_folder,"Generated_Exams",correct_answer_excel_FN)
        


        #--------------------------------------------#
        # Load questions
        all_mcq, all_tf = parse_all_latex_questions_from_question_bank(question_bank_file)
        

        #--------------------------------------------#
        # Selection logic
        if Random_Selection:
            if num_mcq_to_select > len(all_mcq):
                messages.append("You are trying to select more MCQ questions than are available in the question bank.")
            if num_tf_to_select > len(all_tf):
                messages.append("You are trying to select more T/F questions than are available in the question bank.")

            selected_mcq = random.sample(all_mcq, min(num_mcq_to_select, len(all_mcq)))
            selected_tf = random.sample(all_tf, min(num_tf_to_select, len(all_tf)))
        else:
            selected_mcq = [all_mcq[i - 1] for i in Selected_MCQ_Questions if 0 < i <= len(all_mcq)]
            selected_tf = [all_tf[i - 1] for i in Selected_TF_Questions if 0 < i <= len(all_tf)]

            for i, mark in enumerate(MCQ_Marks):
                selected_mcq[i]['mark'] = mark

        #--------------------------------------------#
        # Generation and reporting
        tf_questions_all_sets, best_indices, tf_answer_overlap_between_sets = generate_balanced_overlap_sets(selected_tf)
        
        
        mcq_questions_all_sets = generate_shuffled_mcq_sets(selected_mcq)
        tf_answers_all_sets, selected_tf_CA_marks, selected_tf_WA_marks = generate_set_wise_answers_and_marks(
            tf_questions_all_sets, tf_neg_percentage)
        mcq_answers_all_sets, selected_mcq_CA_marks, selected_mcq_WA_marks = generate_set_wise_answers_and_marks(
            mcq_questions_all_sets, mcq_neg_percentage)


        create_correct_answer_excel(correct_answers_template_file,correct_answers_output_file,Question_Name,
                                    tf_answers_all_sets,
                                    selected_tf_CA_marks,
                                    selected_tf_WA_marks,
                                    mcq_answers_all_sets,
                                    selected_mcq_CA_marks,
                                    selected_mcq_WA_marks)


        
        total_tf_marks = sum(selected_tf_CA_marks)
        total_mcq_marks = sum(selected_mcq_CA_marks)
        
        if total_tf_marks + total_mcq_marks != full_mark:
            messages.append("Total mark (mcq+tf) is different from input 'Marks'")
            messages.append("Updated the total marks")
            full_mark = total_tf_marks + total_mcq_marks

        # ✅ Pass the same base_folder for output too (so output paths are consistent)
        output_files_location,generated_pdfs  = create_question_latex_code_and_pdf(
            Department, exam_title, latex_template_file, Question_Name,
            mcq_questions_all_sets, tf_questions_all_sets,
            full_mark, exam_duration, tf_neg_percentage, mcq_neg_percentage,
            base_folder=base_folder
        )

        #--------------------------------------------#
        messages.append("Success!!")
        messages.append(f"{Question_Name}, generated")
        messages.append(f"Total MCQ: {len(selected_mcq)}, T/F: {len(selected_tf)}")

        return JSONResponse({
            "status": "success",
            "messages": messages,
            "exam_questions_folder_location": str(output_files_location),
            "generated_pdfs": generated_pdfs,
            "correct_answer_excel_FN":correct_answer_excel_FN
        })

    except Exception as e:
        print("Error during exam generation:", e)
        traceback.print_exc()
        messages.append(f"❌ Error: {str(e)}")
        return JSONResponse({"status": "error", "messages": messages})
