"""MessyWorkbook - Main entry point for parsing Excel files."""

# ============================================================================
# Imports
# ============================================================================

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from messy_xlsx.cache import get_structure_cache
from messy_xlsx.detection.format_detector import FormatDetector
from messy_xlsx.detection.structure_analyzer import StructureAnalyzer
from messy_xlsx.exceptions import FileError, FormatError
from messy_xlsx.formulas.config import FormulaConfig, FormulaEvaluationMode
from messy_xlsx.formulas.engine import FormulaEngine
from messy_xlsx.models import CellValue, SheetConfig, StructureInfo
from messy_xlsx.normalization.pipeline import NormalizationPipeline
from messy_xlsx.parsing.base_handler import ParseOptions
from messy_xlsx.parsing.handler_registry import HandlerRegistry
from messy_xlsx.sheet import MessySheet


# ============================================================================
# Core
# ============================================================================

class MessyWorkbook:
    """Main entry point for parsing Excel files."""

    def __init__(
        self,
        file_path: str | Path,
        sheet_config: SheetConfig | None = None,
        formula_config: FormulaConfig | None = None,
    ):
        """Open an Excel file for parsing."""
        self._file_path      = Path(file_path)
        self._sheet_config   = sheet_config or SheetConfig()
        self._formula_config = formula_config or FormulaConfig()

        self._detector      = FormatDetector()
        self._registry      = HandlerRegistry()
        self._analyzer      = StructureAnalyzer(get_structure_cache())
        self._formula_engine = FormulaEngine(self._formula_config)

        if not self._file_path.exists():
            raise FileError(
                f"File not found: {self._file_path}",
                file_path = str(self._file_path),
            )

        self._format_info = self._detector.detect(self._file_path)
        if self._format_info.format_type == "unknown":
            raise FormatError(
                f"Unknown file format: {self._file_path}",
                file_path = str(self._file_path),
            )

        self._sheet_names = self._registry.get_sheet_names(self._file_path)

        self._sheets: dict[str, MessySheet] = {}

        if self._formula_config.mode != FormulaEvaluationMode.DISABLED:
            if self._formula_engine.is_available:
                try:
                    self._formula_engine.load_workbook(self._file_path)
                except Exception as e:
                    pass

        self._wb: openpyxl.Workbook | None = None

    @property
    def file_path(self) -> Path:
        """Path to the Excel file."""
        return self._file_path

    @property
    def sheet_names(self) -> list[str]:
        """List of sheet names in the workbook."""
        return self._sheet_names.copy()

    @property
    def format_type(self) -> str:
        """Detected file format (xlsx, xls, csv, etc.)."""
        return self._format_info.format_type

    def get_sheet(self, name: str | None = None) -> MessySheet:
        """Get a sheet by name."""
        if name is None:
            name = self._sheet_names[0]

        if name not in self._sheet_names:
            raise FormatError(
                f"Sheet '{name}' not found",
                file_path = str(self._file_path),
            )

        if name not in self._sheets:
            self._sheets[name] = MessySheet(self, name)

        return self._sheets[name]

    def to_dataframe(
        self,
        sheet: str | None = None,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Convert a sheet to a pandas DataFrame."""
        sheet_name = sheet or self._sheet_names[0]
        return self._parse_sheet(sheet_name, config)

    def to_dataframes(
        self,
        config: SheetConfig | None = None,
    ) -> dict[str, pd.DataFrame]:
        """Convert all sheets to DataFrames."""
        result = {}
        for name in self._sheet_names:
            try:
                result[name] = self._parse_sheet(name, config)
            except Exception as e:
                pass
        return result

    def get_structure(self, sheet: str | None = None) -> StructureInfo:
        """Get detected structure for a sheet."""
        sheet_name = sheet or self._sheet_names[0]
        return self._analyze_structure(sheet_name)

    def get_cell(
        self,
        sheet: str,
        row: int,
        col: int,
    ) -> CellValue:
        """Get a single cell value."""
        self._ensure_workbook()

        ws   = self._wb[sheet]
        cell = ws.cell(row, col)

        cached_value = cell.value

        formula    = None
        is_formula = False
        if hasattr(cell, "data_type") and cell.data_type == "f":
            is_formula = True
            if hasattr(cell, "value") and isinstance(cell.value, str):
                if cell.value.startswith("="):
                    formula = cell.value

        if is_formula and self._formula_config.mode != FormulaEvaluationMode.DISABLED:
            try:
                cached_value = self._formula_engine.evaluate(
                    sheet, row, col, cached_value
                )
            except Exception:
                pass

        data_type = self._get_data_type(cached_value)

        is_merged = self._is_cell_merged(ws, row, col)

        is_hidden = self._is_cell_hidden(ws, row, col)

        return CellValue(
            value           = cached_value,
            formula         = formula,
            is_merged       = is_merged,
            is_hidden       = is_hidden,
            data_type       = data_type,
            original_format = cell.number_format if hasattr(cell, "number_format") else None,
        )

    def get_cell_by_ref(self, ref: str) -> CellValue:
        """Get a cell by A1-style reference."""
        from messy_xlsx.utils import cell_ref_to_coords

        sheet, row, col = cell_ref_to_coords(ref)
        sheet           = sheet or self._sheet_names[0]
        return self.get_cell(sheet, row, col)

    def _parse_sheet(
        self,
        sheet: str,
        config: SheetConfig | None = None,
    ) -> pd.DataFrame:
        """Parse a sheet to DataFrame with normalization."""
        config = config or self._sheet_config

        if config.auto_detect:
            structure        = self._analyze_structure(sheet)
            effective_config = self._apply_structure_detection(config, structure)
        else:
            effective_config = config

        parse_options = ParseOptions(
            skip_rows      = effective_config.skip_rows,
            header_rows    = effective_config.header_rows,
            skip_footer    = effective_config.skip_footer,
            merge_strategy = effective_config.merge_strategy,
            ignore_hidden  = not effective_config.include_hidden,
            cell_range     = effective_config.cell_range,
            data_only      = True,
        )

        df = self._registry.parse(
            self._file_path,
            sheet   = sheet,
            options = parse_options,
        )

        pipeline = NormalizationPipeline(
            decimal_separator   = None,
            thousands_separator = None,
        )

        type_hints = effective_config.type_hints.copy()

        df = pipeline.normalize(df, semantic_hints=type_hints)

        if effective_config.column_renames:
            df = df.rename(columns=effective_config.column_renames)

        return df

    def _analyze_structure(self, sheet: str) -> StructureInfo:
        """Analyze sheet structure."""
        return self._analyzer.analyze(self._file_path, sheet)

    def _apply_structure_detection(
        self,
        config: SheetConfig,
        structure: StructureInfo,
    ) -> SheetConfig:
        """Merge user config with detected structure."""
        return SheetConfig(
            skip_rows        = config.skip_rows if config.skip_rows > 0 else structure.suggested_skip_rows,
            header_rows      = config.header_rows,
            skip_footer      = config.skip_footer if config.skip_footer > 0 else structure.suggested_skip_footer,
            cell_range       = config.cell_range,
            column_renames   = config.column_renames,
            type_hints       = config.type_hints,
            auto_detect      = False,
            include_hidden   = config.include_hidden,
            merge_strategy   = config.merge_strategy,
            locale           = config.locale or structure.detected_locale,
            evaluate_formulas = config.evaluate_formulas,
            drop_regex       = config.drop_regex,
            drop_conditions  = config.drop_conditions,
        )

    def _ensure_workbook(self) -> None:
        """Ensure openpyxl workbook is loaded."""
        if self._wb is None:
            self._wb = openpyxl.load_workbook(
                self._file_path,
                read_only = True,
                data_only = True,
            )

    def _get_data_type(self, value: Any) -> str:
        """Determine data type string for a value."""
        if value is None:
            return "empty"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, str):
            if value.startswith("#") and value.endswith("!"):
                return "error"
            return "text"
        if hasattr(value, "date"):
            return "date"
        return "text"

    def _is_cell_merged(self, ws, row: int, col: int) -> bool:
        """Check if cell is part of a merged range."""
        try:
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    return True
        except Exception:
            pass
        return False

    def _is_cell_hidden(self, ws, row: int, col: int) -> bool:
        """Check if cell is in a hidden row or column."""
        try:
            if row in ws.row_dimensions and ws.row_dimensions[row].hidden:
                return True
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden:
                return True
        except Exception:
            pass
        return False

    def close(self) -> None:
        """Close the workbook and release resources."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    def __enter__(self) -> "MessyWorkbook":
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    def __repr__(self) -> str:
        return f"MessyWorkbook({self._file_path.name!r}, sheets={self._sheet_names})"
