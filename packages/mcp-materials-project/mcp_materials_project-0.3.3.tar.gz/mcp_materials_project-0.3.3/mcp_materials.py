"""
Materials Project MCP Server
Provides full material data access via mp-api with complete field extraction
"""

import json
import os
from typing import Optional, Any, List
from datetime import datetime

from fastmcp import FastMCP
from mp_api.client import MPRester
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

API_KEY = os.environ.get("MP_API_KEY")
if not API_KEY:
    raise ValueError("MP_API_KEY environment variable is required")

mcp = FastMCP("materials-project")


def serialize_object(obj: Any) -> Any:
    """Recursively serialize objects to JSON-compatible format"""
    if obj is None:
        return None
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (list, tuple)):
        return [serialize_object(item) for item in obj]
    if isinstance(obj, dict):
        return {str(k): serialize_object(v) for k, v in obj.items()}

    # Handle numpy arrays
    if hasattr(obj, 'tolist'):
        return obj.tolist()

    # Handle pymatgen Element objects - extract symbol directly
    if hasattr(obj, 'symbol') and hasattr(obj, 'Z'):
        # This is likely a pymatgen Element object
        return str(obj.symbol)

    # Handle pymatgen objects with as_dict
    if hasattr(obj, 'as_dict'):
        return serialize_object(obj.as_dict())

    # Handle objects with __dict__
    if hasattr(obj, '__dict__'):
        return serialize_object(obj.__dict__)

    return str(obj)


def flatten_dict(d: dict, parent_key: str = '', sep: str = '_') -> dict:
    """
    Flatten nested dictionary with underscore separator.
    Example: symmetry.crystal_system -> Symmetry_Crystal_System
    """
    items = []
    for k, v in d.items():
        # Capitalize first letter of each key segment
        formatted_key = k.replace('_', ' ').title().replace(' ', '_')
        new_key = f"{parent_key}{sep}{formatted_key}" if parent_key else formatted_key

        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        elif isinstance(v, list):
            if len(v) == 0:
                items.append((new_key, ""))
            elif isinstance(v[0], dict):
                # Serialize list of dicts as JSON string
                items.append((new_key, json.dumps(v, default=str)))
            elif isinstance(v[0], (list, tuple)):
                # Matrix data - format nicely
                matrix_str = '\n'.join([str(row) for row in v])
                items.append((new_key, matrix_str))
            else:
                # Simple list - join with comma
                items.append((new_key, ', '.join(str(x) for x in v)))
        else:
            items.append((new_key, v))
    return dict(items)


# Complete list of all available summary fields from Materials Project
SUMMARY_FIELDS = [
    # Basic info
    "material_id",
    "formula_pretty",
    "formula_anonymous",
    "chemsys",
    "composition",
    "composition_reduced",
    "elements",
    "nelements",
    "nsites",
    "volume",
    "density",
    "density_atomic",

    # Symmetry
    "symmetry",

    # Structure
    "structure",

    # Thermodynamics
    "energy_per_atom",
    "formation_energy_per_atom",
    "energy_above_hull",
    "is_stable",
    "equilibrium_reaction_energy_per_atom",
    "decomposes_to",

    # Electronic
    "band_gap",
    "cbm",
    "vbm",
    "efermi",
    "is_gap_direct",
    "is_metal",

    # Magnetism
    "is_magnetic",
    "ordering",
    "total_magnetization",
    "total_magnetization_normalized_vol",
    "total_magnetization_normalized_formula_units",
    "num_magnetic_sites",
    "num_unique_magnetic_sites",

    # Elasticity
    "bulk_modulus",
    "shear_modulus",
    "universal_anisotropy",
    "homogeneous_poisson",

    # Dielectric
    "e_total",
    "e_ionic",
    "e_electronic",
    "n",
    "e_ij_max",

    # Surface properties
    "weighted_surface_energy_EV_PER_ANG2",
    "weighted_surface_energy",
    "weighted_work_function",
    "surface_anisotropy",
    "shape_factor",
    "has_reconstructed",

    # Metadata
    "possible_species",
    "has_props",
    "theoretical",
    "database_IDs",
]


def format_structure_string(structure_dict: dict) -> str:
    """Format structure data into readable string"""
    if not structure_dict:
        return ""

    parts = []

    # Lattice info
    lattice = structure_dict.get('lattice', {})
    if lattice:
        a = lattice.get('a', 'N/A')
        b = lattice.get('b', 'N/A')
        c = lattice.get('c', 'N/A')
        alpha = lattice.get('alpha', 'N/A')
        beta = lattice.get('beta', 'N/A')
        gamma = lattice.get('gamma', 'N/A')
        vol = lattice.get('volume', 'N/A')
        parts.append(f"Lattice: a={a:.4f}, b={b:.4f}, c={c:.4f} Å")
        parts.append(f"Angles: α={alpha:.2f}°, β={beta:.2f}°, γ={gamma:.2f}°")
        parts.append(f"Volume: {vol:.4f} Å³")

    # Sites info
    sites = structure_dict.get('sites', [])
    if sites:
        parts.append(f"Sites: {len(sites)} atoms")
        site_summary = {}
        for site in sites[:20]:  # Limit for readability
            species = site.get('species', [{}])[0].get('element', 'Unknown')
            site_summary[species] = site_summary.get(species, 0) + 1
        parts.append(f"Composition: {site_summary}")

    return '\n'.join(parts)


def process_material_doc(doc: Any) -> dict:
    """Process a material document into flat dictionary with all fields"""
    doc_dict = serialize_object(doc)

    # Special handling for structure field
    structure_data = doc_dict.get('structure', {})
    structure_str = format_structure_string(structure_data)

    # Special handling for symmetry
    symmetry_data = doc_dict.get('symmetry', {})

    # Build result dict with organized fields
    result = {}

    # Core identifiers
    result['Material_ID'] = doc_dict.get('material_id', '')
    result['Formula'] = doc_dict.get('formula_pretty', '')
    result['Formula_Anonymous'] = doc_dict.get('formula_anonymous', '')
    result['Chemical_System'] = doc_dict.get('chemsys', '')

    # Handle elements - can be list of strings or list of Element objects
    elements = doc_dict.get('elements', [])
    if elements:
        # After serialize_object, Element objects should be converted to strings
        # But we still need to handle various formats
        element_symbols = []
        for e in elements:
            if isinstance(e, str):
                element_symbols.append(e)
            elif isinstance(e, dict):
                # Try to extract element symbol from dict
                if 'symbol' in e:
                    element_symbols.append(str(e['symbol']))
                elif 'element' in e:
                    element_symbols.append(str(e['element']))
                elif '_value_' in e:
                    element_symbols.append(str(e['_value_']))
                else:
                    # Fallback to string representation
                    element_symbols.append(str(e))
            else:
                element_symbols.append(str(e))
        result['Elements'] = ', '.join(element_symbols)
    else:
        result['Elements'] = ''

    result['N_Elements'] = doc_dict.get('nelements', '')
    result['N_Sites'] = doc_dict.get('nsites', '')

    # Thermodynamics
    result['Energy_Per_Atom_eV'] = doc_dict.get('energy_per_atom', '')
    result['Formation_Energy_eV_Atom'] = doc_dict.get('formation_energy_per_atom', '')
    result['Energy_Above_Hull_eV_Atom'] = doc_dict.get('energy_above_hull', '')
    result['Is_Stable'] = doc_dict.get('is_stable', '')
    result['Equilibrium_Reaction_Energy'] = doc_dict.get('equilibrium_reaction_energy_per_atom', '')
    decomposes = doc_dict.get('decomposes_to', [])
    result['Decomposes_To'] = json.dumps(decomposes, default=str) if decomposes else ''

    # Electronic
    result['Band_Gap_eV'] = doc_dict.get('band_gap', '')
    result['CBM_eV'] = doc_dict.get('cbm', '')
    result['VBM_eV'] = doc_dict.get('vbm', '')
    result['Fermi_Energy_eV'] = doc_dict.get('efermi', '')
    result['Is_Gap_Direct'] = doc_dict.get('is_gap_direct', '')
    result['Is_Metal'] = doc_dict.get('is_metal', '')

    # Magnetism
    result['Is_Magnetic'] = doc_dict.get('is_magnetic', '')
    result['Magnetic_Ordering'] = doc_dict.get('ordering', '')
    result['Total_Magnetization'] = doc_dict.get('total_magnetization', '')
    result['Magnetization_Per_Volume'] = doc_dict.get('total_magnetization_normalized_vol', '')
    result['Magnetization_Per_Formula'] = doc_dict.get('total_magnetization_normalized_formula_units', '')
    result['N_Magnetic_Sites'] = doc_dict.get('num_magnetic_sites', '')
    result['N_Unique_Magnetic_Sites'] = doc_dict.get('num_unique_magnetic_sites', '')

    # Elasticity
    bulk_mod = doc_dict.get('bulk_modulus', {})
    shear_mod = doc_dict.get('shear_modulus', {})
    if isinstance(bulk_mod, dict):
        result['Bulk_Modulus_VRH_GPa'] = bulk_mod.get('vrh', '')
        result['Bulk_Modulus_Voigt_GPa'] = bulk_mod.get('voigt', '')
        result['Bulk_Modulus_Reuss_GPa'] = bulk_mod.get('reuss', '')
    else:
        result['Bulk_Modulus_VRH_GPa'] = bulk_mod
    if isinstance(shear_mod, dict):
        result['Shear_Modulus_VRH_GPa'] = shear_mod.get('vrh', '')
        result['Shear_Modulus_Voigt_GPa'] = shear_mod.get('voigt', '')
        result['Shear_Modulus_Reuss_GPa'] = shear_mod.get('reuss', '')
    else:
        result['Shear_Modulus_VRH_GPa'] = shear_mod
    result['Universal_Anisotropy'] = doc_dict.get('universal_anisotropy', '')
    result['Poisson_Ratio'] = doc_dict.get('homogeneous_poisson', '')

    # Dielectric
    result['Dielectric_Total'] = doc_dict.get('e_total', '')
    result['Dielectric_Ionic'] = doc_dict.get('e_ionic', '')
    result['Dielectric_Electronic'] = doc_dict.get('e_electronic', '')
    result['Refractive_Index_n'] = doc_dict.get('n', '')
    result['Piezoelectric_Max'] = doc_dict.get('e_ij_max', '')

    # Physical properties
    result['Volume_A3'] = doc_dict.get('volume', '')
    result['Density_g_cm3'] = doc_dict.get('density', '')
    result['Density_Atomic'] = doc_dict.get('density_atomic', '')

    # Surface properties
    result['Surface_Energy_J_m2'] = doc_dict.get('weighted_surface_energy', '')
    result['Surface_Energy_eV_A2'] = doc_dict.get('weighted_surface_energy_EV_PER_ANG2', '')
    result['Work_Function_eV'] = doc_dict.get('weighted_work_function', '')
    result['Surface_Anisotropy'] = doc_dict.get('surface_anisotropy', '')
    result['Shape_Factor'] = doc_dict.get('shape_factor', '')
    result['Has_Reconstructed'] = doc_dict.get('has_reconstructed', '')

    # Symmetry
    if isinstance(symmetry_data, dict):
        result['Space_Group_Symbol'] = symmetry_data.get('symbol', '')
        result['Space_Group_Number'] = symmetry_data.get('number', '')
        result['Crystal_System'] = symmetry_data.get('crystal_system', '')
        result['Point_Group'] = symmetry_data.get('point_group', '')

    # Structure (formatted string)
    result['Structure_Details'] = structure_str

    # Metadata
    result['Possible_Species'] = ', '.join(doc_dict.get('possible_species', [])) if doc_dict.get('possible_species') else ''
    result['Has_Properties'] = json.dumps(doc_dict.get('has_props', []), default=str) if doc_dict.get('has_props') else ''
    result['Is_Theoretical'] = doc_dict.get('theoretical', '')

    db_ids = doc_dict.get('database_IDs', {})
    if isinstance(db_ids, dict):
        result['ICSD_IDs'] = ', '.join(str(x) for x in db_ids.get('icsd', [])) if db_ids.get('icsd') else ''
        result['COD_IDs'] = ', '.join(str(x) for x in db_ids.get('cod', [])) if db_ids.get('cod') else ''

    # Full raw data for reference
    result['Full_Properties'] = json.dumps(doc_dict, indent=2, default=str)

    return result


def _fetch_material_data_core(
    material_ids: Optional[str] = None,
    formula: Optional[str] = None,
    chemsys: Optional[str] = None,
    elements: Optional[str] = None,
    band_gap_min: Optional[float] = None,
    band_gap_max: Optional[float] = None,
    is_stable: Optional[bool] = None,
    is_metal: Optional[bool] = None,
    is_magnetic: Optional[bool] = None,
    num_results: int = 10
) -> dict:
    """
    Core function to fetch material data from Materials Project.
    Returns a dictionary with status and data.
    """
    try:
        with MPRester(API_KEY) as mpr:
            search_params = {"num_chunks": 1, "chunk_size": num_results}

            if material_ids:
                ids = [mid.strip() for mid in material_ids.split(",")]
                search_params["material_ids"] = ids

            if formula:
                search_params["formula"] = formula

            if chemsys:
                search_params["chemsys"] = chemsys

            if elements:
                elem_list = [e.strip() for e in elements.split(",")]
                search_params["elements"] = elem_list

            if band_gap_min is not None or band_gap_max is not None:
                bg_min = band_gap_min if band_gap_min is not None else 0
                bg_max = band_gap_max if band_gap_max is not None else 100
                search_params["band_gap"] = (bg_min, bg_max)

            if is_stable is not None:
                search_params["is_stable"] = is_stable

            if is_metal is not None:
                search_params["is_metal"] = is_metal

            if is_magnetic is not None:
                search_params["is_magnetic"] = is_magnetic

            search_params["fields"] = SUMMARY_FIELDS

            docs = mpr.materials.summary.search(**search_params)

            results = []
            for doc in docs:
                processed = process_material_doc(doc)
                results.append(processed)

            return {
                "status": "success",
                "count": len(results),
                "query_params": {k: str(v) for k, v in search_params.items() if k != "fields"},
                "data": results,
                "timestamp": datetime.now().isoformat()
            }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }


@mcp.tool
def fetch_full_material_data(
    material_ids: Optional[str] = None,
    formula: Optional[str] = None,
    chemsys: Optional[str] = None,
    elements: Optional[str] = None,
    band_gap_min: Optional[float] = None,
    band_gap_max: Optional[float] = None,
    is_stable: Optional[bool] = None,
    is_metal: Optional[bool] = None,
    is_magnetic: Optional[bool] = None,
    num_results: int = 10
) -> str:
    """
    Fetch full material data from Materials Project using summary.search.

    Args:
        material_ids: Comma-separated material IDs (e.g., "mp-149,mp-1234")
        formula: Chemical formula (e.g., "Si", "Fe2O3")
        chemsys: Chemical system (e.g., "Li-Fe-O")
        elements: Comma-separated elements to include (e.g., "Si,O")
        band_gap_min: Minimum band gap in eV
        band_gap_max: Maximum band gap in eV
        is_stable: Filter for thermodynamically stable materials
        is_metal: Filter for metallic materials
        is_magnetic: Filter for magnetic materials
        num_results: Maximum number of results (default 10)

    Returns:
        JSON string with full material data including thermodynamic, electronic,
        mechanical, magnetic, and symmetry properties.
    """
    result = _fetch_material_data_core(
        material_ids=material_ids,
        formula=formula,
        chemsys=chemsys,
        elements=elements,
        band_gap_min=band_gap_min,
        band_gap_max=band_gap_max,
        is_stable=is_stable,
        is_metal=is_metal,
        is_magnetic=is_magnetic,
        num_results=num_results
    )
    return json.dumps(result, indent=2, default=str)


@mcp.tool
def get_structure_details(material_id: str) -> str:
    """
    Get detailed crystal structure information for a specific material.

    Args:
        material_id: Materials Project ID (e.g., "mp-149")

    Returns:
        JSON string with lattice parameters, atomic sites, coordinates,
        space group, and other structural details.
    """
    try:
        with MPRester(API_KEY) as mpr:
            structure = mpr.get_structure_by_material_id(material_id)

            lattice = structure.lattice
            lattice_info = {
                "a": lattice.a,
                "b": lattice.b,
                "c": lattice.c,
                "alpha": lattice.alpha,
                "beta": lattice.beta,
                "gamma": lattice.gamma,
                "volume": lattice.volume,
                "matrix": lattice.matrix.tolist()
            }

            sites = []
            for site in structure.sites:
                site_info = {
                    "species": str(site.specie),
                    "coords_fractional": list(site.frac_coords),
                    "coords_cartesian": list(site.coords),
                    "properties": serialize_object(site.properties) if site.properties else {}
                }
                sites.append(site_info)

            analyzer_data = {}
            try:
                from pymatgen.symmetry.analyzer import SpacegroupAnalyzer
                sga = SpacegroupAnalyzer(structure)
                analyzer_data = {
                    "space_group_symbol": sga.get_space_group_symbol(),
                    "space_group_number": sga.get_space_group_number(),
                    "crystal_system": sga.get_crystal_system(),
                    "point_group": sga.get_point_group_symbol(),
                    "hall_symbol": sga.get_hall()
                }
            except Exception:
                pass

            output = {
                "status": "success",
                "material_id": material_id,
                "formula": structure.composition.reduced_formula,
                "lattice": lattice_info,
                "num_sites": len(sites),
                "sites": sites,
                "symmetry": analyzer_data,
                "timestamp": datetime.now().isoformat()
            }

            return json.dumps(output, indent=2, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "material_id": material_id,
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }, indent=2)


@mcp.tool
def search_materials_by_property(
    property_name: str,
    min_value: Optional[float] = None,
    max_value: Optional[float] = None,
    num_results: int = 20
) -> str:
    """
    Search materials by specific property ranges.

    Args:
        property_name: Property to search (band_gap, formation_energy_per_atom,
                      energy_above_hull, bulk_modulus, shear_modulus, density, volume)
        min_value: Minimum value for the property
        max_value: Maximum value for the property
        num_results: Maximum number of results

    Returns:
        JSON string with materials matching the property criteria.
    """
    valid_properties = {
        "band_gap": "band_gap",
        "formation_energy": "formation_energy_per_atom",
        "formation_energy_per_atom": "formation_energy_per_atom",
        "energy_above_hull": "energy_above_hull",
        "bulk_modulus": "bulk_modulus",
        "shear_modulus": "shear_modulus",
        "density": "density",
        "volume": "volume",
        "total_magnetization": "total_magnetization"
    }

    if property_name.lower() not in valid_properties:
        return json.dumps({
            "status": "error",
            "message": f"Invalid property. Valid options: {list(valid_properties.keys())}",
            "timestamp": datetime.now().isoformat()
        }, indent=2)

    prop_key = valid_properties[property_name.lower()]

    try:
        with MPRester(API_KEY) as mpr:
            search_params = {
                "num_chunks": 1,
                "chunk_size": num_results,
                "fields": SUMMARY_FIELDS
            }

            min_val = min_value if min_value is not None else -1e10
            max_val = max_value if max_value is not None else 1e10
            search_params[prop_key] = (min_val, max_val)

            docs = mpr.materials.summary.search(**search_params)

            results = []
            for doc in docs:
                processed = process_material_doc(doc)
                results.append(processed)

            output = {
                "status": "success",
                "count": len(results),
                "property_searched": property_name,
                "range": {"min": min_value, "max": max_value},
                "data": results,
                "timestamp": datetime.now().isoformat()
            }

            return json.dumps(output, indent=2, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }, indent=2)


@mcp.tool
def get_phase_diagram_info(chemsys: str) -> str:
    """
    Get phase diagram information for a chemical system.

    Args:
        chemsys: Chemical system (e.g., "Li-Fe-O", "Si-O")

    Returns:
        JSON string with phase diagram entries and stability information.
    """
    try:
        with MPRester(API_KEY) as mpr:
            entries = mpr.get_entries_in_chemsys(chemsys)

            entry_data = []
            for entry in entries:
                entry_info = {
                    "entry_id": str(entry.entry_id),
                    "composition": str(entry.composition.reduced_formula),
                    "energy": entry.energy,
                    "energy_per_atom": entry.energy_per_atom,
                    "correction": entry.correction,
                    "parameters": serialize_object(entry.parameters) if hasattr(entry, 'parameters') else {}
                }
                entry_data.append(entry_info)

            output = {
                "status": "success",
                "chemsys": chemsys,
                "num_entries": len(entry_data),
                "entries": entry_data,
                "timestamp": datetime.now().isoformat()
            }

            return json.dumps(output, indent=2, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "chemsys": chemsys,
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }, indent=2)


@mcp.tool
def compare_materials(material_ids: str) -> str:
    """
    Compare multiple materials side by side.

    Args:
        material_ids: Comma-separated material IDs (e.g., "mp-149,mp-1234,mp-5678")

    Returns:
        JSON string with comparison table of key properties.
    """
    try:
        ids = [mid.strip() for mid in material_ids.split(",")]

        with MPRester(API_KEY) as mpr:
            docs = mpr.materials.summary.search(
                material_ids=ids,
                fields=SUMMARY_FIELDS
            )

            comparison = []
            for doc in docs:
                processed = process_material_doc(doc)
                # Extract key comparison fields
                key_props = {
                    "Material_ID": processed.get("Material_ID"),
                    "Formula": processed.get("Formula"),
                    "Band_Gap_eV": processed.get("Band_Gap_eV"),
                    "Is_Metal": processed.get("Is_Metal"),
                    "Formation_Energy_eV_Atom": processed.get("Formation_Energy_eV_Atom"),
                    "Energy_Above_Hull_eV_Atom": processed.get("Energy_Above_Hull_eV_Atom"),
                    "Is_Stable": processed.get("Is_Stable"),
                    "Density_g_cm3": processed.get("Density_g_cm3"),
                    "Volume_A3": processed.get("Volume_A3"),
                    "N_Sites": processed.get("N_Sites"),
                    "Is_Magnetic": processed.get("Is_Magnetic"),
                    "Total_Magnetization": processed.get("Total_Magnetization"),
                    "Bulk_Modulus_VRH_GPa": processed.get("Bulk_Modulus_VRH_GPa"),
                    "Shear_Modulus_VRH_GPa": processed.get("Shear_Modulus_VRH_GPa"),
                    "Space_Group_Symbol": processed.get("Space_Group_Symbol"),
                    "Crystal_System": processed.get("Crystal_System"),
                }
                comparison.append(key_props)

            output = {
                "status": "success",
                "num_materials": len(comparison),
                "comparison": comparison,
                "timestamp": datetime.now().isoformat()
            }

            return json.dumps(output, indent=2, default=str)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }, indent=2)


def _style_excel_workbook(ws, df: pd.DataFrame):
    """Apply professional styling to Excel worksheet"""
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Column width settings
    wide_columns = ['Full_Properties', 'Structure_Details', 'Has_Properties', 'Decomposes_To']

    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)

        if any(wc.lower() in col_name.lower() for wc in wide_columns):
            ws.column_dimensions[col_letter].width = 70
        else:
            max_length = len(str(col_name))
            for cell in ws[col_letter][1:ws.max_row]:
                if cell.value:
                    cell_len = len(str(cell.value).split('\n')[0])
                    max_length = max(max_length, min(cell_len, 50))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)

    # Enable filter and freeze panes
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'B2'


def _create_comparison_excel(materials_data: List[dict], output_path: str):
    """Create horizontal comparison Excel format for multiple materials"""
    if not materials_data:
        return

    # Define key properties to include in comparison (prioritize most important fields)
    properties = [
        ('Material_ID', 'Material_ID'),
        ('Formula', 'Formula'),
        ('Band_Gap_eV', 'Band_Gap_eV'),
        ('Energy_Above_Hull_eV_Atom', 'Energy_Above_Hull_eV_Atom'),
        ('Is_Stable', 'Is_Stable'),
        ('Is_Metal', 'Is_Metal'),
        ('Is_Magnetic', 'Is_Magnetic'),
        ('Formation_Energy_eV_Atom', 'Formation_Energy_eV_Atom'),
        ('Density_g_cm3', 'Density_g_cm3'),
        ('Volume_A3', 'Volume_A3'),
        ('N_Sites', 'N_Sites'),
        ('N_Elements', 'N_Elements'),
        ('Elements', 'Elements'),
        ('Space_Group_Symbol', 'Space_Group_Symbol'),
        ('Space_Group_Number', 'Space_Group_Number'),
        ('Crystal_System', 'Crystal_System'),
        ('Point_Group', 'Point_Group'),
        ('CBM_eV', 'CBM_eV'),
        ('VBM_eV', 'VBM_eV'),
        ('Fermi_Energy_eV', 'Fermi_Energy_eV'),
        ('Is_Gap_Direct', 'Is_Gap_Direct'),
        ('Total_Magnetization', 'Total_Magnetization'),
        ('Magnetic_Ordering', 'Magnetic_Ordering'),
        ('Bulk_Modulus_VRH_GPa', 'Bulk_Modulus_VRH_GPa'),
        ('Shear_Modulus_VRH_GPa', 'Shear_Modulus_VRH_GPa'),
        ('Poisson_Ratio', 'Poisson_Ratio'),
        ('Dielectric_Total', 'Dielectric_Total'),
        ('Refractive_Index_n', 'Refractive_Index_n'),
        ('Surface_Energy_J_m2', 'Surface_Energy_J_m2'),
        ('Work_Function_eV', 'Work_Function_eV'),
    ]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials Comparison"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    property_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    property_font = Font(bold=True, size=11)
    property_alignment = Alignment(horizontal="left", vertical="center")

    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Write header row (material names)
    ws.cell(1, 1, "Property").fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = header_alignment
    ws.cell(1, 1).border = border

    for col_idx, mat in enumerate(materials_data, start=2):
        formula = mat.get('Formula', 'N/A')
        mat_id = mat.get('Material_ID', 'N/A')
        header_text = f"{formula} ({mat_id})"
        cell = ws.cell(1, col_idx, header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_idx, (prop_display, prop_key) in enumerate(properties, start=2):
        # Property name in first column
        cell = ws.cell(row_idx, 1, prop_display)
        cell.fill = property_fill
        cell.font = property_font
        cell.alignment = property_alignment
        cell.border = border

        # Material values in subsequent columns
        for col_idx, mat in enumerate(materials_data, start=2):
            value = mat.get(prop_key, None)

            # Format value - convert to string first to avoid MPID comparison issues
            if value is None:
                display_value = 'N/A'
            elif isinstance(value, bool):
                display_value = str(value)
            elif isinstance(value, (int, float)):
                if isinstance(value, float):
                    display_value = f"{value:.4g}"
                else:
                    display_value = str(value)
            else:
                # Convert to string first
                str_value = str(value)
                if str_value == '' or str_value == 'None':
                    display_value = 'N/A'
                else:
                    display_value = str_value

            cell = ws.cell(row_idx, col_idx, display_value)
            cell.alignment = data_alignment
            cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 30  # Property column
    for col_idx in range(2, len(materials_data) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25  # Material columns

    # Set row heights
    ws.row_dimensions[1].height = 30  # Header row
    for row_idx in range(2, len(properties) + 2):
        ws.row_dimensions[row_idx].height = 25

    # Freeze panes
    ws.freeze_panes = 'B2'

    # Save workbook
    wb.save(output_path)


@mcp.tool
def export_to_excel(
    material_ids: Optional[str] = None,
    formula: Optional[str] = None,
    chemsys: Optional[str] = None,
    elements: Optional[str] = None,
    band_gap_min: Optional[float] = None,
    band_gap_max: Optional[float] = None,
    is_stable: Optional[bool] = None,
    is_metal: Optional[bool] = None,
    is_magnetic: Optional[bool] = None,
    num_results: int = 10,
    output_filename: Optional[str] = None
) -> str:
    """
    Export material data to Excel file with professional formatting.

    Args:
        material_ids: Comma-separated material IDs (e.g., "mp-149,mp-1234")
        formula: Chemical formula (e.g., "Si", "Fe2O3")
        chemsys: Chemical system (e.g., "Li-Fe-O")
        elements: Comma-separated elements to include (e.g., "Si,O")
        band_gap_min: Minimum band gap in eV
        band_gap_max: Maximum band gap in eV
        is_stable: Filter for thermodynamically stable materials
        is_metal: Filter for metallic materials
        is_magnetic: Filter for magnetic materials
        num_results: Maximum number of results (default 10)
        output_filename: Custom output filename (without path, e.g., "my_materials.xlsx")

    Returns:
        JSON string with export status and file path
    """
    try:
        # Fetch data using core function (not the MCP tool)
        result_data = _fetch_material_data_core(
            material_ids=material_ids,
            formula=formula,
            chemsys=chemsys,
            elements=elements,
            band_gap_min=band_gap_min,
            band_gap_max=band_gap_max,
            is_stable=is_stable,
            is_metal=is_metal,
            is_magnetic=is_magnetic,
            num_results=num_results
        )

        if result_data.get("status") != "success":
            return json.dumps(result_data, indent=2)

        materials_data = result_data.get("data", [])

        if not materials_data:
            return json.dumps({
                "status": "error",
                "message": "No materials found matching the criteria",
                "timestamp": datetime.now().isoformat()
            }, indent=2)

        # Generate output path
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(os.getcwd(), "output")
        os.makedirs(output_dir, exist_ok=True)

        if output_filename:
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            output_path = os.path.join(output_dir, output_filename)
        else:
            # Auto-generate filename based on query
            if material_ids:
                first_id = material_ids.split(',')[0].strip()
                # Check if comparing multiple materials
                if ',' in material_ids and len(materials_data) >= 2:
                    filename = f"{first_id}_comparison_{timestamp}.xlsx"
                else:
                    filename = f"{first_id}_{timestamp}.xlsx"
            elif formula:
                filename = f"{formula}_{timestamp}.xlsx"
            else:
                filename = f"materials_export_{timestamp}.xlsx"
            output_path = os.path.join(output_dir, filename)

        # Always use horizontal comparison format
        _create_comparison_excel(materials_data, output_path)

        return json.dumps({
            "status": "success",
            "message": f"Exported {len(materials_data)} materials to Excel",
            "file_path": output_path,
            "num_materials": len(materials_data),
            "format": "horizontal_comparison",
            "timestamp": datetime.now().isoformat()
        }, indent=2)

    except Exception as e:
        return json.dumps({
            "status": "error",
            "message": str(e),
            "timestamp": datetime.now().isoformat()
        }, indent=2)


def main():
    """Entry point for the MCP server."""
    mcp.run()


if __name__ == "__main__":
    main()
