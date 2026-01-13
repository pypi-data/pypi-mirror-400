"""Status code sheets writer."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

from .base import BaseSheetWriter, get_column_letter
from .config import HEADERS, SHEET_NAMES, SheetConfig

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class StatusCodeSheetWriter(BaseSheetWriter):
    """Creates status code analysis sheets."""

    def write(self) -> None:
        """Create all status code sheets."""
        try:
            self._create_count_sheets()
            self._create_timestamp_sheets()
        except Exception as e:
            logger.error(f"상태 코드 시트 생성 중 오류: {e}")

    def _create_count_sheets(self) -> None:
        """Create count-based status code sheets."""
        count_sheets = {
            SHEET_NAMES.ELB_2XX: {
                "count": self.data.get("elb_2xx_count", 0),
                "full_logs": self.data.get("ELB 2xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.ELB_3XX: {
                "count": self.data.get("elb_3xx_count", 0),
                "full_logs": self.data.get("ELB 3xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.ELB_4XX_COUNT: {
                "count": self.data.get("elb_4xx_count", 0),
                "full_logs": self.data.get("ELB 4xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.ELB_5XX_COUNT: {
                "count": self.data.get("elb_5xx_count", 0),
                "full_logs": self.data.get("ELB 5xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.BACKEND_4XX_COUNT: {
                "count": self.data.get("backend_4xx_count", 0),
                "full_logs": self.data.get("Backend 4xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.BACKEND_5XX_COUNT: {
                "count": self.data.get("backend_5xx_count", 0),
                "full_logs": self.data.get("Backend 5xx Count", {}).get("full_logs", []),
            },
        }

        # Create sheets in severity order: 5xx -> 4xx -> 3xx -> 2xx
        ordered_names = [
            SHEET_NAMES.ELB_5XX_COUNT,
            SHEET_NAMES.BACKEND_5XX_COUNT,
            SHEET_NAMES.ELB_4XX_COUNT,
            SHEET_NAMES.BACKEND_4XX_COUNT,
            SHEET_NAMES.ELB_3XX,
            SHEET_NAMES.ELB_2XX,
        ]

        for sheet_name in ordered_names:
            sheet_data = count_sheets.get(sheet_name, {})
            self._create_count_sheet(sheet_name, sheet_data)

    def _create_timestamp_sheets(self) -> None:
        """Create timestamp-based status code sheets."""
        timestamp_sheets = {
            SHEET_NAMES.ELB_4XX_TS: {
                "count": self.data.get("elb_4xx_count", 0),
                "full_logs": self.data.get("ELB 4xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.ELB_5XX_TS: {
                "count": self.data.get("elb_5xx_count", 0),
                "full_logs": self.data.get("ELB 5xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.BACKEND_4XX_TS: {
                "count": self.data.get("backend_4xx_count", 0),
                "full_logs": self.data.get("Backend 4xx Count", {}).get("full_logs", []),
            },
            SHEET_NAMES.BACKEND_5XX_TS: {
                "count": self.data.get("backend_5xx_count", 0),
                "full_logs": self.data.get("Backend 5xx Count", {}).get("full_logs", []),
            },
        }

        # Create sheets in severity order
        ordered_names = [
            SHEET_NAMES.ELB_5XX_TS,
            SHEET_NAMES.BACKEND_5XX_TS,
            SHEET_NAMES.ELB_4XX_TS,
            SHEET_NAMES.BACKEND_4XX_TS,
        ]

        for sheet_name in ordered_names:
            sheet_data = timestamp_sheets.get(sheet_name, {})
            self._create_timestamp_sheet(sheet_name, sheet_data)

    def _create_count_sheet(
        self,
        sheet_name: str,
        status_data: dict[str, Any],
    ) -> None:
        """Create a count-aggregated status code sheet."""
        try:
            full_logs = status_data.get("full_logs") or []
            status_count_val = int(status_data.get("count", 0) or 0)

            if status_count_val == 0 and not full_logs:
                return

            ws = self.create_sheet(sheet_name)

            # Select headers based on sheet type
            is_3xx = "3xx" in sheet_name
            headers = list(HEADERS.STATUS_COUNT_3XX if is_3xx else HEADERS.STATUS_COUNT_BASE)

            self.write_header_row(ws, headers)

            # Aggregate by unique request pattern
            aggregated = self._aggregate_logs(full_logs, is_3xx)

            # Sort by count and limit for 2xx/3xx
            sorted_items = sorted(aggregated.items(), key=lambda x: x[1], reverse=True)
            if "2xx" in sheet_name or "3xx" in sheet_name:
                sorted_items = sorted_items[:100]

            # Write data
            rows_written = self._write_count_rows(ws, sorted_items, headers, is_3xx)

            self.finalize_sheet(ws, headers, rows_written)

        except Exception as e:
            logger.error(f"상태 코드 카운트 시트 생성 중 오류: {e}")

    def _aggregate_logs(
        self,
        logs: list[dict[str, Any]],
        include_redirect: bool = False,
    ) -> dict[tuple[Any, ...], int]:
        """Aggregate logs by unique request pattern."""
        counts: dict[tuple[Any, ...], int] = {}

        for log in logs:
            client_ip = log.get("client_ip", "N/A")
            request = log.get("request", "N/A")
            http_method = log.get("http_method", "").replace("-", "")
            elb_status = self.convert_status_code(log.get("elb_status_code", "N/A"))
            target_status = self.convert_status_code(log.get("target_status_code", "N/A"))
            target = log.get("target", "")
            target_field = "" if not target or target == "-" else target
            target_group = log.get("target_group_name", "") or ""

            key: tuple[Any, ...]
            if include_redirect:
                redirect_url = log.get("redirect_url", "").replace("-", "")
                key = (
                    client_ip,
                    target_field,
                    target_group,
                    http_method,
                    request,
                    redirect_url,
                    elb_status,
                    target_status,
                )
            else:
                key = (
                    client_ip,
                    target_field,
                    target_group,
                    http_method,
                    request,
                    elb_status,
                    target_status,
                )

            counts[key] = counts.get(key, 0) + 1

        return counts

    def _write_count_rows(
        self,
        ws: Worksheet,
        items: list[tuple[tuple, int]],
        headers: list[str],
        is_3xx: bool,
    ) -> int:
        """Write aggregated count rows."""
        border = self.styles.thin_border
        abuse_fill = self.styles.fill_abuse
        rows_written = 0

        for row_idx, (key_tuple, count) in enumerate(items, start=2):
            if is_3xx:
                (
                    client_ip,
                    target,
                    target_group,
                    method,
                    request,
                    redirect_url,
                    elb_status,
                    target_status,
                ) = key_tuple
            else:
                (
                    client_ip,
                    target,
                    target_group,
                    method,
                    request,
                    elb_status,
                    target_status,
                ) = key_tuple
                redirect_url = None

            is_abuse = client_ip in self.abuse_ip_set
            country = self.get_country_code(client_ip)

            # Build row values
            values = [count, client_ip, country, target, target_group, method, request]
            if is_3xx:
                values.append(redirect_url)
            values.extend([elb_status, target_status])

            # Write cells
            for col_idx, (header, value) in enumerate(zip(headers, values, strict=False), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                self.apply_cell_style(cell, header, value)

                if is_abuse:
                    cell.fill = abuse_fill

            rows_written += 1

        return rows_written

    def _create_timestamp_sheet(
        self,
        sheet_name: str,
        status_data: dict[str, Any],
    ) -> None:
        """Create a timestamp-based status code sheet."""
        try:
            full_logs = status_data.get("full_logs") or []
            status_count_val = int(status_data.get("count", 0) or 0)

            if status_count_val == 0 and not full_logs:
                return

            ws = self.create_sheet(sheet_name)

            headers = list(HEADERS.STATUS_TIMESTAMP)

            self.write_header_row(ws, headers)

            # Sort by timestamp
            sorted_logs = sorted(full_logs, key=self._safe_timestamp_key)

            # Write data
            rows_written = self._write_timestamp_rows(ws, sorted_logs, headers)

            self.finalize_sheet(ws, headers, rows_written)

        except Exception as e:
            logger.error(f"상태 코드 타임스탬프 시트 생성 중 오류: {e}")

    def _safe_timestamp_key(self, entry: dict[str, Any]) -> datetime:
        """Get timestamp for sorting."""
        ts = entry.get("timestamp")
        if isinstance(ts, datetime):
            return ts
        if isinstance(ts, str):
            try:
                return datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        return datetime.min

    def _write_timestamp_rows(
        self,
        ws: Worksheet,
        logs: list[dict[str, Any]],
        headers: list[str],
    ) -> int:
        """Write timestamp-ordered log rows."""
        border = self.styles.thin_border
        abuse_fill = self.styles.fill_abuse
        rows_written = 0

        for row_idx, log in enumerate(logs, start=2):
            client_ip = log.get("client_ip", "N/A")
            is_abuse = client_ip in self.abuse_ip_set

            timestamp_str = self.format_timestamp(log.get("timestamp"))
            country = self.get_country_code(client_ip)
            target = log.get("target", "")
            target_field = "" if not target or target == "-" else target
            target_group = log.get("target_group_name", "") or ""
            method = log.get("http_method", "")
            request = log.get("request", "N/A")
            user_agent = log.get("user_agent", "N/A")
            elb_status = self.convert_status_code(log.get("elb_status_code", "N/A"))
            target_status = self.convert_status_code(log.get("target_status_code", "N/A"))
            error_reason = log.get("error_reason", "-")
            if error_reason in (None, "-"):
                error_reason = ""

            values = [
                timestamp_str,
                client_ip,
                country,
                target_field,
                target_group,
                method,
                request,
                user_agent,
                elb_status,
                target_status,
                error_reason,
            ]

            for col_idx, (header, value) in enumerate(zip(headers, values, strict=False), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                self.apply_cell_style(cell, header, value)

                if is_abuse:
                    cell.fill = abuse_fill

            rows_written += 1

        return rows_written


class ClientStatusSheetWriter(BaseSheetWriter):
    """Creates client IP status code statistics sheet."""

    def write(self) -> None:
        """Create client status code statistics sheet."""
        try:
            client_status_stats = self.data.get("client_status_statistics", {})
            if not client_status_stats:
                logger.debug("Client 상태코드 통계 데이터가 없습니다.")
                return

            ws = self.create_sheet(SHEET_NAMES.CLIENT_STATUS)

            # Collect all status codes
            all_status_codes = set()
            for status_counts in client_status_stats.values():
                all_status_codes.update(status_counts.keys())

            # Sort status codes
            sorted_codes = sorted(
                all_status_codes,
                key=lambda x: (int(x) if x.isdigit() else float("inf"), x),
            )

            # Build headers
            headers = ["Client", "Country"] + sorted_codes + ["Count"]

            self.write_header_row(ws, headers)

            # Get IP country mapping
            ip_country_mapping = self.data.get("ip_country_mapping", {})

            # Calculate totals and sort
            client_totals = {ip: sum(counts.values()) for ip, counts in client_status_stats.items()}
            sorted_clients = sorted(client_totals.items(), key=lambda x: x[1], reverse=True)[
                : SheetConfig.TOP_CLIENT_LIMIT
            ]

            # Write data
            row = 2
            for client_ip, total_count in sorted_clients:
                status_counts = client_status_stats[client_ip]
                country = ip_country_mapping.get(client_ip, "UNKNOWN")

                # Client cell
                cell = ws.cell(row=row, column=1, value=client_ip)
                cell.alignment = self.styles.align_center
                cell.border = self.styles.thin_border

                # Country cell
                cell = ws.cell(row=row, column=2, value=country)
                cell.alignment = self.styles.align_center
                cell.border = self.styles.thin_border

                # Status code columns
                for col_idx, status_code in enumerate(sorted_codes, 3):
                    cell = ws.cell(row=row, column=col_idx, value=status_counts.get(status_code, 0))
                    cell.number_format = "#,##0"
                    cell.border = self.styles.thin_border

                # Total count cell
                cell = ws.cell(row=row, column=len(headers), value=total_count)
                cell.alignment = self.styles.align_right
                cell.number_format = "#,##0"
                cell.border = self.styles.thin_border

                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 10
            for col_idx in range(3, len(headers)):
                ws.column_dimensions[get_column_letter(col_idx)].width = 8
            ws.column_dimensions[get_column_letter(len(headers))].width = 10

            # Finalize
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE

        except Exception as e:
            logger.error(f"Client 상태코드 통계 시트 생성 중 오류: {e}")


class TargetStatusSheetWriter(BaseSheetWriter):
    """Creates target IP status code statistics sheet."""

    def write(self) -> None:
        """Create target status code statistics sheet."""
        try:
            target_status_stats = self.data.get("target_status_statistics", {})
            if not target_status_stats:
                logger.info("Target 상태코드 통계 데이터가 없습니다.")
                return

            ws = self.create_sheet(SHEET_NAMES.TARGET_STATUS)

            # Collect backend status codes only
            backend_codes = set()
            for status_counts in target_status_stats.values():
                for key in status_counts:
                    if key.startswith("Backend:"):
                        backend_codes.add(key.split("Backend:")[1])

            if not backend_codes:
                logger.warning("Backend 상태코드가 없습니다.")
                sorted_codes: list[str] = []
            else:
                sorted_codes = sorted(
                    backend_codes,
                    key=lambda x: (int(x) if x.isdigit() else float("inf"), x),
                )

            # Build headers
            headers = ["Target", "Target group name"] + sorted_codes + ["Count"]

            self.write_header_row(ws, headers)

            # Calculate backend totals
            target_totals = {}
            for target, status_counts in target_status_stats.items():
                backend_total = sum(count for key, count in status_counts.items() if key.startswith("Backend:"))
                if backend_total > 0:
                    target_totals[target] = backend_total

            sorted_targets = sorted(target_totals.items(), key=lambda x: x[1], reverse=True)

            # Write data
            row = 2
            for target_display, total_count in sorted_targets:
                status_counts = target_status_stats[target_display]

                # Parse target/group name
                parsed_target = target_display
                parsed_group = ""
                if "(" in target_display and target_display.endswith(")"):
                    open_idx = target_display.rfind("(")
                    parsed_group = target_display[:open_idx]
                    parsed_target = target_display[open_idx + 1 : -1]

                # Target cell
                cell = ws.cell(row=row, column=1, value=parsed_target)
                cell.alignment = self.styles.align_center
                cell.border = self.styles.thin_border

                # Group name cell
                cell = ws.cell(row=row, column=2, value=parsed_group)
                cell.alignment = self.styles.align_center
                cell.border = self.styles.thin_border

                # Backend status columns
                for col_idx, code in enumerate(sorted_codes, 3):
                    count = status_counts.get(f"Backend:{code}", 0)
                    cell = ws.cell(row=row, column=col_idx, value=count)
                    cell.number_format = "#,##0"
                    cell.border = self.styles.thin_border

                # Total count
                cell = ws.cell(row=row, column=len(headers), value=total_count)
                cell.alignment = self.styles.align_right
                cell.number_format = "#,##0"
                cell.border = self.styles.thin_border

                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 20
            for col_idx in range(3, len(headers)):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10

            # Finalize
            ws.freeze_panes = ws.cell(row=2, column=1)
            ws.sheet_view.zoomScale = SheetConfig.ZOOM_SCALE

        except Exception as e:
            logger.error(f"Target 상태코드 통계 시트 생성 중 오류: {e}")
