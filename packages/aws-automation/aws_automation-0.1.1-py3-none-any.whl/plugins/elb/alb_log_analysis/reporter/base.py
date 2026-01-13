"""Base sheet writer with performance optimizations."""

from __future__ import annotations

import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

from .config import (
    COLUMN_STYLE_MAP,
    COLUMN_WIDTH_MAP,
    STYLE_CENTER,
    STYLE_DATA,
    STYLE_DECIMAL,
    STYLE_DECIMAL3,
    STYLE_NUMBER,
    STYLE_STATUS,
    WRAP_TEXT_COLUMNS,
    SheetConfig,
)
from .styles import get_style_cache

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def get_column_letter(col_idx: int) -> str:
    """Get Excel column letter from index (1-based)."""
    from openpyxl.utils import get_column_letter as _gcl

    result: str = _gcl(col_idx)
    return result


class BaseSheetWriter:
    """Base class for sheet writers."""

    def __init__(
        self,
        workbook: Workbook,
        data: dict[str, Any],
        abuse_ip_set: set[str] | None = None,
    ) -> None:
        self.workbook = workbook
        self.data = data
        self.abuse_ip_set = abuse_ip_set or set()
        self.styles = get_style_cache()
        self.config = SheetConfig()

    def create_sheet(self, name: str) -> Worksheet:
        return self.workbook.create_sheet(name)

    def write_header_row(self, ws: Worksheet, headers: list[str] | tuple[str, ...], row: int = 1) -> None:
        style = self.styles.get_header_style()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = style["font"]
            cell.fill = style["fill"]
            cell.alignment = style["alignment"]
            cell.border = style["border"]

    def write_empty_message(self, ws: Worksheet, message: str, row: int, col_count: int) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
        cell = ws.cell(row=row, column=1, value=message)
        cell.alignment = self.styles.align_center
        cell.font = self.styles.font_italic_gray
        cell.border = self.styles.thin_border

    def finalize_sheet(
        self,
        ws: Worksheet,
        headers: list[str] | tuple[str, ...],
        data_count: int = 0,
    ) -> None:
        self._apply_column_widths(ws, headers)
        ws.row_dimensions[1].height = self.config.HEADER_ROW_HEIGHT
        if data_count > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_count + 1}"
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.sheet_view.zoomScale = self.config.ZOOM_SCALE

    def _apply_column_widths(self, ws: Worksheet, headers: list[str] | tuple[str, ...]) -> None:
        for col, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH_MAP.get(
                header, self.config.DEFAULT_COLUMN_WIDTH
            )

    def apply_cell_style(self, cell: Cell, header: str, value: Any) -> None:
        style_type = COLUMN_STYLE_MAP.get(header, STYLE_DATA)
        if style_type == STYLE_NUMBER:
            cell.alignment = self.styles.align_right
            cell.number_format = "#,##0"
        elif style_type == STYLE_DECIMAL:
            cell.alignment = self.styles.align_right
            if isinstance(value, (int, float)):
                cell.number_format = "0.00"
        elif style_type == STYLE_DECIMAL3:
            cell.alignment = self.styles.align_right
            if isinstance(value, (int, float)):
                cell.number_format = "0.000"
        elif style_type == STYLE_STATUS:
            cell.alignment = self.styles.align_right
            if isinstance(value, (int, float)):
                cell.number_format = "0"
        elif style_type == STYLE_CENTER:
            cell.alignment = self.styles.align_center
        else:
            cell.alignment = self.styles.align_left

    def write_data_rows(
        self,
        ws: Worksheet,
        data: list[dict[str, Any]],
        headers: list[str] | tuple[str, ...],
        start_row: int = 2,
        highlight_abuse: bool = True,
        client_column: str = "Client",
    ) -> int:
        if not data:
            return 0

        border = self.styles.thin_border
        abuse_fill = self.styles.fill_abuse
        headers_list = list(headers)
        batch_size = self.config.BATCH_SIZE

        for batch_start in range(0, len(data), batch_size):
            batch = data[batch_start : batch_start + batch_size]

            for offset, row_data in enumerate(batch):
                row_idx = start_row + batch_start + offset
                has_abuse = (
                    highlight_abuse
                    and self.abuse_ip_set
                    and client_column in headers_list
                    and row_data.get(client_column, "") in self.abuse_ip_set
                )

                for col, header in enumerate(headers_list, 1):
                    value = row_data.get(header, "")
                    if value is None:
                        value = ""
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = border
                    self.apply_cell_style(cell, header, value)
                    if has_abuse:
                        cell.fill = abuse_fill
                    if isinstance(value, bool):
                        cell.value = "Yes" if value else "No"

        return len(data)

    def apply_wrap_text(self, ws: Worksheet, headers: list[str] | tuple[str, ...]) -> None:
        from openpyxl.styles import Alignment

        for col, header in enumerate(list(headers), 1):
            if header not in WRAP_TEXT_COLUMNS:
                continue
            h_align = WRAP_TEXT_COLUMNS[header]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.alignment:
                    cell.alignment = cell.alignment.copy(wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)

    def convert_status_code(self, code: Any) -> int | str:
        if code is None or code == "" or code == "-":
            return ""
        if isinstance(code, int):
            return code
        if isinstance(code, str):
            code = code.strip()
            if code in ("", "-", "N/A"):
                return ""
            try:
                return int(float(code))
            except (ValueError, TypeError):
                return ""
        if isinstance(code, float):
            return int(code)
        return ""

    def format_bytes(self, size: int | float | str | None) -> str:
        try:
            size = float(size or 0)
            for unit in ("", "KB", "MB", "GB", "TB"):
                if size < 1024.0:
                    return f"{size:.2f} {unit}"
                size /= 1024.0
            return f"{size:.2f} PB"
        except (ValueError, TypeError):
            return "N/A"

    def format_timestamp(self, ts: Any) -> str:
        if isinstance(ts, datetime):
            return ts.strftime("%Y-%m-%d %H:%M:%S")
        return str(ts) if ts else "N/A"

    def get_country_code(self, client_ip: str) -> str:
        result: str = self.data.get("ip_country_mapping", {}).get(client_ip, "N/A")
        return result

    def get_matching_abuse_ips(self) -> set[str]:
        client_ips = set(self.data.get("client_ip_counts", {}).keys())
        abuse_list, _ = self.get_normalized_abuse_ips()
        return client_ips.intersection(abuse_list)

    def get_normalized_abuse_ips(self) -> tuple[list[str], dict[str, Any]]:
        excluded = {"abuse_ips", "abuse_ip_details", "timestamp"}

        def valid(ip: Any) -> bool:
            s = str(ip).strip() if ip else ""
            return bool(s) and not any(k in s for k in excluded)

        ips: list[str] = []
        details: dict[str, Any] = {}

        if self.data.get("abuse_ips_list"):
            ips = [str(ip).strip() for ip in self.data["abuse_ips_list"] if valid(ip)]
        elif self.data.get("abuse_ips"):
            abuse = self.data["abuse_ips"]
            if isinstance(abuse, (list, set)):
                ips = [str(ip).strip() for ip in abuse if valid(ip)]
            elif isinstance(abuse, dict):
                ips = [str(ip).strip() for ip in abuse if valid(ip)]
                details = abuse

        if self.data.get("abuse_ip_details"):
            details.update(self.data["abuse_ip_details"])

        return ips, details


class SummarySheetHelper:
    """Helper for summary sheet layout."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        self.styles = get_style_cache()
        self.row = 1
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

    def add_title(self, title: str) -> SummarySheetHelper:
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws.cell(row=self.row, column=1, value=title)
        cell.font = self.styles.title_font
        cell.fill = self.styles.title_fill
        cell.alignment = self.styles.align_center
        cell.border = self.styles.thin_border
        self.ws.cell(row=self.row, column=2).border = self.styles.thin_border
        self.ws.row_dimensions[self.row].height = 40
        self.row += 2
        return self

    def add_section(self, name: str) -> SummarySheetHelper:
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws.cell(row=self.row, column=1, value=name)
        cell.font = self.styles.header_font
        cell.fill = self.styles.header_fill
        cell.alignment = self.styles.align_center
        cell.border = self.styles.thin_border
        self.ws.cell(row=self.row, column=2).border = self.styles.thin_border
        self.row += 1
        return self

    def add_item(
        self,
        label: str,
        value: Any,
        highlight: str | None = None,
        number_format: str | None = None,
    ) -> SummarySheetHelper:
        label_cell = self.ws.cell(row=self.row, column=1, value=label)
        label_cell.font = self.styles.label_font
        label_cell.alignment = self.styles.align_left
        label_cell.border = self.styles.thin_border

        value_cell = self.ws.cell(row=self.row, column=2, value=value)
        value_cell.font = self.styles.value_font
        value_cell.alignment = self.styles.align_left
        value_cell.border = self.styles.thin_border

        if number_format:
            value_cell.number_format = number_format
        if highlight == "danger":
            value_cell.fill = self.styles.danger_fill
        elif highlight == "warning":
            value_cell.fill = self.styles.warning_fill

        self.row += 1
        return self

    def add_blank_row(self) -> SummarySheetHelper:
        self.row += 1
        return self

    def add_list_section(
        self,
        label: str,
        items: list[tuple[str, Any]],
        max_items: int = 5,
        suffix: str = "",
    ) -> SummarySheetHelper:
        label_cell = self.ws.cell(row=self.row, column=1, value=f"{label}:")
        label_cell.font = self.styles.label_font
        label_cell.alignment = self.styles.align_left
        label_cell.border = self.styles.thin_border
        self.ws.cell(row=self.row, column=2).border = self.styles.thin_border
        self.row += 1

        if not items:
            cell = self.ws.cell(row=self.row, column=1, value="데이터 없음")
            cell.font = self.styles.value_font
            cell.border = self.styles.thin_border
            self.ws.cell(row=self.row, column=2).border = self.styles.thin_border
            self.row += 1
        else:
            for i, (name, count) in enumerate(items[:max_items], 1):
                display = str(name)[:47] + "..." if len(str(name)) > 50 else str(name)
                name_cell = self.ws.cell(row=self.row, column=1, value=f"{i}. {display}")
                name_cell.font = self.styles.value_font
                name_cell.alignment = self.styles.align_left
                name_cell.border = self.styles.thin_border

                val = f"{count:,}{suffix}" if isinstance(count, (int, float)) else str(count)
                count_cell = self.ws.cell(row=self.row, column=2, value=val)
                count_cell.font = self.styles.value_font
                count_cell.alignment = self.styles.align_right
                count_cell.border = self.styles.thin_border
                self.row += 1

        return self
