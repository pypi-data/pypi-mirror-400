"""
plugins/efs/unused.py - EFS 미사용 파일시스템 분석

유휴/미사용 EFS 파일시스템 탐지 (마운트 타겟 및 CloudWatch 지표 기반)

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer

console = Console()

# 미사용 기준: 7일간 I/O 없음
UNUSED_DAYS_THRESHOLD = 7

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "elasticfilesystem:DescribeFileSystems",
        "elasticfilesystem:DescribeMountTargets",
        "cloudwatch:GetMetricStatistics",
    ],
}


class FileSystemStatus(Enum):
    """파일시스템 상태"""

    NORMAL = "normal"
    NO_MOUNT_TARGET = "no_mount_target"
    NO_IO = "no_io"
    EMPTY = "empty"


@dataclass
class EFSInfo:
    """EFS 파일시스템 정보"""

    account_id: str
    account_name: str
    region: str
    file_system_id: str
    name: str
    lifecycle_state: str
    performance_mode: str
    throughput_mode: str
    size_bytes: int
    mount_target_count: int
    created_at: datetime | None
    # CloudWatch 지표
    avg_client_connections: float = 0.0
    total_io_bytes: float = 0.0

    @property
    def size_gb(self) -> float:
        return self.size_bytes / (1024**3)

    @property
    def estimated_monthly_cost(self) -> float:
        """월간 비용 추정 (Standard 기준 $0.30/GB)"""
        return self.size_gb * 0.30


@dataclass
class EFSFinding:
    """EFS 분석 결과"""

    efs: EFSInfo
    status: FileSystemStatus
    recommendation: str


@dataclass
class EFSAnalysisResult:
    """EFS 분석 결과 집계"""

    account_id: str
    account_name: str
    region: str
    total_filesystems: int = 0
    no_mount_target: int = 0
    no_io: int = 0
    empty: int = 0
    normal: int = 0
    unused_monthly_cost: float = 0.0
    findings: list[EFSFinding] = field(default_factory=list)


def collect_efs_filesystems(session, account_id: str, account_name: str, region: str) -> list[EFSInfo]:
    """EFS 파일시스템 수집"""
    from botocore.exceptions import ClientError

    efs = get_client(session, "efs", region_name=region)
    cloudwatch = get_client(session, "cloudwatch", region_name=region)
    filesystems = []

    now = datetime.now(timezone.utc)
    start_time = now - timedelta(days=UNUSED_DAYS_THRESHOLD)

    try:
        paginator = efs.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page.get("FileSystems", []):
                fs_id = fs.get("FileSystemId", "")

                # 이름 태그 찾기
                name = ""
                for tag in fs.get("Tags", []):
                    if tag.get("Key") == "Name":
                        name = tag.get("Value", "")
                        break

                # 마운트 타겟 수 확인
                mount_target_count = 0
                try:
                    mt_resp = efs.describe_mount_targets(FileSystemId=fs_id)
                    mount_target_count = len(mt_resp.get("MountTargets", []))
                except ClientError:
                    pass

                info = EFSInfo(
                    account_id=account_id,
                    account_name=account_name,
                    region=region,
                    file_system_id=fs_id,
                    name=name,
                    lifecycle_state=fs.get("LifeCycleState", ""),
                    performance_mode=fs.get("PerformanceMode", ""),
                    throughput_mode=fs.get("ThroughputMode", ""),
                    size_bytes=fs.get("SizeInBytes", {}).get("Value", 0),
                    mount_target_count=mount_target_count,
                    created_at=fs.get("CreationTime"),
                )

                # CloudWatch 지표 조회
                try:
                    # ClientConnections
                    conn_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/EFS",
                        MetricName="ClientConnections",
                        Dimensions=[{"Name": "FileSystemId", "Value": fs_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Average"],
                    )
                    if conn_resp.get("Datapoints"):
                        info.avg_client_connections = sum(d["Average"] for d in conn_resp["Datapoints"]) / len(
                            conn_resp["Datapoints"]
                        )

                    # TotalIOBytes (읽기+쓰기)
                    io_resp = cloudwatch.get_metric_statistics(
                        Namespace="AWS/EFS",
                        MetricName="TotalIOBytes",
                        Dimensions=[{"Name": "FileSystemId", "Value": fs_id}],
                        StartTime=start_time,
                        EndTime=now,
                        Period=86400,
                        Statistics=["Sum"],
                    )
                    if io_resp.get("Datapoints"):
                        info.total_io_bytes = sum(d["Sum"] for d in io_resp["Datapoints"])

                except ClientError:
                    pass

                filesystems.append(info)
    except ClientError:
        pass

    return filesystems


def analyze_filesystems(
    filesystems: list[EFSInfo], account_id: str, account_name: str, region: str
) -> EFSAnalysisResult:
    """EFS 파일시스템 분석"""
    result = EFSAnalysisResult(
        account_id=account_id,
        account_name=account_name,
        region=region,
        total_filesystems=len(filesystems),
    )

    for fs in filesystems:
        # 마운트 타겟 없음
        if fs.mount_target_count == 0:
            result.no_mount_target += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_MOUNT_TARGET,
                    recommendation=f"마운트 타겟 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        # 빈 파일시스템 (거의 0 바이트)
        if fs.size_bytes < 1024 * 1024:  # 1MB 미만
            result.empty += 1
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.EMPTY,
                    recommendation="빈 파일시스템 - 삭제 검토",
                )
            )
            continue

        # I/O 없음
        if fs.total_io_bytes == 0 and fs.avg_client_connections == 0:
            result.no_io += 1
            result.unused_monthly_cost += fs.estimated_monthly_cost
            result.findings.append(
                EFSFinding(
                    efs=fs,
                    status=FileSystemStatus.NO_IO,
                    recommendation=f"I/O 없음 - 삭제 검토 (${fs.estimated_monthly_cost:.2f}/월)",
                )
            )
            continue

        result.normal += 1
        result.findings.append(
            EFSFinding(
                efs=fs,
                status=FileSystemStatus.NORMAL,
                recommendation="정상",
            )
        )

    return result


def generate_report(results: list[EFSAnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "EFS 미사용 분석 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Account",
        "Region",
        "전체",
        "마운트없음",
        "I/O없음",
        "빈FS",
        "정상",
        "미사용 비용",
    ]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.region)
        ws.cell(row=row, column=3, value=r.total_filesystems)
        ws.cell(row=row, column=4, value=r.no_mount_target)
        ws.cell(row=row, column=5, value=r.no_io)
        ws.cell(row=row, column=6, value=r.empty)
        ws.cell(row=row, column=7, value=r.normal)
        ws.cell(row=row, column=8, value=f"${r.unused_monthly_cost:,.2f}")
        if r.no_mount_target > 0:
            ws.cell(row=row, column=4).fill = red_fill
        if r.no_io > 0:
            ws.cell(row=row, column=5).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("FileSystems")
    detail_headers = [
        "Account",
        "Region",
        "ID",
        "Name",
        "Size",
        "Mount Targets",
        "Mode",
        "상태",
        "Avg Conn",
        "Total I/O",
        "월간 비용",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != FileSystemStatus.NORMAL:
                detail_row += 1
                fs = f.efs
                ws_detail.cell(row=detail_row, column=1, value=fs.account_name)
                ws_detail.cell(row=detail_row, column=2, value=fs.region)
                ws_detail.cell(row=detail_row, column=3, value=fs.file_system_id)
                ws_detail.cell(row=detail_row, column=4, value=fs.name or "-")
                ws_detail.cell(row=detail_row, column=5, value=f"{fs.size_gb:.2f} GB")
                ws_detail.cell(row=detail_row, column=6, value=fs.mount_target_count)
                ws_detail.cell(row=detail_row, column=7, value=fs.throughput_mode)
                ws_detail.cell(row=detail_row, column=8, value=f.status.value)
                ws_detail.cell(row=detail_row, column=9, value=f"{fs.avg_client_connections:.1f}")
                ws_detail.cell(
                    row=detail_row,
                    column=10,
                    value=f"{fs.total_io_bytes / (1024**2):.1f} MB",
                )
                ws_detail.cell(
                    row=detail_row,
                    column=11,
                    value=f"${fs.estimated_monthly_cost:.2f}",
                )
                ws_detail.cell(row=detail_row, column=12, value=f.recommendation)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"EFS_Unused_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> EFSAnalysisResult | None:
    """단일 계정/리전의 EFS 수집 및 분석 (병렬 실행용)"""
    filesystems = collect_efs_filesystems(session, account_id, account_name, region)
    if not filesystems:
        return None
    return analyze_filesystems(filesystems, account_id, account_name, region)


def run(ctx) -> None:
    """EFS 미사용 파일시스템 분석"""
    console.print("[bold]EFS 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="efs")
    results: list[EFSAnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.no_mount_target + r.no_io + r.empty for r in results)
    unused_cost = sum(r.unused_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용: [red]{total_unused}개[/red] (${unused_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("efs-unused").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
