"""
plugins/route53/empty_zone.py - 빈 Hosted Zone 탐지

레코드가 없는 미사용 Route53 Hosted Zone 분석

플러그인 규약:
    - run(ctx): 필수. 실행 함수.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum

from rich.console import Console

from core.parallel import get_client, parallel_collect
from core.tools.output import OutputPath, open_in_explorer
from plugins.cost.pricing import get_hosted_zone_price

console = Console()

# 필요한 AWS 권한 목록
REQUIRED_PERMISSIONS = {
    "read": [
        "route53:ListHostedZones",
        "route53:GetHostedZone",
        "route53:ListResourceRecordSets",
    ],
}


class ZoneStatus(Enum):
    """Hosted Zone 상태"""

    NORMAL = "normal"
    EMPTY = "empty"
    NS_SOA_ONLY = "ns_soa_only"


@dataclass
class HostedZoneInfo:
    """Hosted Zone 정보"""

    account_id: str
    account_name: str
    zone_id: str
    name: str
    is_private: bool
    record_count: int
    comment: str
    vpcs: list[str] = field(default_factory=list)
    has_real_records: bool = False

    @property
    def monthly_cost(self) -> float:
        # 첫 25개 기준 가격 (API에서 조회)
        return get_hosted_zone_price(zone_index=1)


@dataclass
class ZoneFinding:
    """Hosted Zone 분석 결과"""

    zone: HostedZoneInfo
    status: ZoneStatus
    recommendation: str


@dataclass
class Route53AnalysisResult:
    """Route53 분석 결과 집계"""

    account_id: str
    account_name: str
    total_zones: int = 0
    empty_zones: int = 0
    ns_soa_only_zones: int = 0
    private_zones: int = 0
    public_zones: int = 0
    wasted_monthly_cost: float = 0.0
    findings: list[ZoneFinding] = field(default_factory=list)


def collect_hosted_zones(session, account_id: str, account_name: str) -> list[HostedZoneInfo]:
    """Hosted Zone 수집 (글로벌 서비스)"""
    from botocore.exceptions import ClientError

    route53 = get_client(session, "route53")
    zones = []

    paginator = route53.get_paginator("list_hosted_zones")
    for page in paginator.paginate():
        for zone in page.get("HostedZones", []):
            zone_id = zone.get("Id", "").replace("/hostedzone/", "")

            zone_info = HostedZoneInfo(
                account_id=account_id,
                account_name=account_name,
                zone_id=zone_id,
                name=zone.get("Name", ""),
                is_private=zone.get("Config", {}).get("PrivateZone", False),
                record_count=zone.get("ResourceRecordSetCount", 0),
                comment=zone.get("Config", {}).get("Comment", ""),
            )

            # Private zone인 경우 연결된 VPC 조회
            if zone_info.is_private:
                try:
                    hz_detail = route53.get_hosted_zone(Id=zone_id)
                    vpcs = hz_detail.get("VPCs", [])
                    zone_info.vpcs = [f"{v.get('VPCRegion')}:{v.get('VPCId')}" for v in vpcs]
                except ClientError:
                    pass

            # 실제 레코드 존재 여부 확인 (NS, SOA 제외)
            try:
                records = route53.list_resource_record_sets(HostedZoneId=zone_id, MaxItems="100")
                for record in records.get("ResourceRecordSets", []):
                    record_type = record.get("Type", "")
                    if record_type not in ("NS", "SOA"):
                        zone_info.has_real_records = True
                        break
            except ClientError:
                pass

            zones.append(zone_info)

    return zones


def analyze_hosted_zones(zones: list[HostedZoneInfo], account_id: str, account_name: str) -> Route53AnalysisResult:
    """Hosted Zone 분석"""
    result = Route53AnalysisResult(
        account_id=account_id,
        account_name=account_name,
        total_zones=len(zones),
    )

    for zone in zones:
        if zone.is_private:
            result.private_zones += 1
        else:
            result.public_zones += 1

        # 완전히 빈 zone (레코드가 0개인 경우 - 드물지만)
        if zone.record_count == 0:
            result.empty_zones += 1
            result.wasted_monthly_cost += zone.monthly_cost
            result.findings.append(
                ZoneFinding(
                    zone=zone,
                    status=ZoneStatus.EMPTY,
                    recommendation="완전히 빈 Zone - 삭제 검토",
                )
            )
            continue

        # NS/SOA만 있는 경우 (실질적으로 미사용)
        if not zone.has_real_records:
            result.ns_soa_only_zones += 1
            result.wasted_monthly_cost += zone.monthly_cost
            result.findings.append(
                ZoneFinding(
                    zone=zone,
                    status=ZoneStatus.NS_SOA_ONLY,
                    recommendation="NS/SOA만 존재 - 사용 여부 확인",
                )
            )
            continue

        result.findings.append(
            ZoneFinding(
                zone=zone,
                status=ZoneStatus.NORMAL,
                recommendation="정상 (레코드 있음)",
            )
        )

    return result


def generate_report(results: list[Route53AnalysisResult], output_dir: str) -> str:
    """Excel 보고서 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")

    # Summary 시트
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Route53 빈 Hosted Zone 보고서"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Account", "전체 Zone", "빈 Zone", "NS/SOA만", "Public", "Private", "낭비 비용"]
    row = 3
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).fill = header_fill
        ws.cell(row=row, column=col).font = header_font

    for r in results:
        row += 1
        ws.cell(row=row, column=1, value=r.account_name)
        ws.cell(row=row, column=2, value=r.total_zones)
        ws.cell(row=row, column=3, value=r.empty_zones)
        ws.cell(row=row, column=4, value=r.ns_soa_only_zones)
        ws.cell(row=row, column=5, value=r.public_zones)
        ws.cell(row=row, column=6, value=r.private_zones)
        ws.cell(row=row, column=7, value=f"${r.wasted_monthly_cost:,.2f}")
        if r.empty_zones > 0:
            ws.cell(row=row, column=3).fill = red_fill
        if r.ns_soa_only_zones > 0:
            ws.cell(row=row, column=4).fill = yellow_fill

    # Detail 시트
    ws_detail = wb.create_sheet("Hosted Zones")
    detail_headers = [
        "Account",
        "Zone ID",
        "Domain",
        "Type",
        "레코드수",
        "상태",
        "Comment",
        "권장 조치",
    ]
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h).fill = header_fill
        ws_detail.cell(row=1, column=col).font = header_font

    detail_row = 1
    for r in results:
        for f in r.findings:
            if f.status != ZoneStatus.NORMAL:
                detail_row += 1
                zone = f.zone
                ws_detail.cell(row=detail_row, column=1, value=zone.account_name)
                ws_detail.cell(row=detail_row, column=2, value=zone.zone_id)
                ws_detail.cell(row=detail_row, column=3, value=zone.name)
                ws_detail.cell(
                    row=detail_row,
                    column=4,
                    value="Private" if zone.is_private else "Public",
                )
                ws_detail.cell(row=detail_row, column=5, value=zone.record_count)
                ws_detail.cell(row=detail_row, column=6, value=f.status.value)
                ws_detail.cell(row=detail_row, column=7, value=zone.comment)
                ws_detail.cell(row=detail_row, column=8, value=f.recommendation)

    # 컬럼 너비 자동 조정
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(c.value) if c.value else "") for c in col)  # type: ignore
            col_idx = col[0].column  # type: ignore
            if col_idx:
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
        sheet.freeze_panes = "A2"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"Route53_EmptyZone_{timestamp}.xlsx")
    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    return filepath


def _collect_and_analyze(session, account_id: str, account_name: str, region: str) -> Route53AnalysisResult | None:
    """단일 계정의 Hosted Zone 수집 및 분석 (병렬 실행용)

    Route53는 글로벌 서비스이므로 region은 무시됩니다.
    parallel_collect의 중복 제거가 계정당 한 번만 실행되도록 보장합니다.
    """
    zones = collect_hosted_zones(session, account_id, account_name)
    if not zones:
        return None
    return analyze_hosted_zones(zones, account_id, account_name)


def run(ctx) -> None:
    """빈 Hosted Zone 분석"""
    console.print("[bold]Route53 빈 Hosted Zone 분석 시작...[/bold]\n")

    result = parallel_collect(ctx, _collect_and_analyze, max_workers=20, service="route53")
    results: list[Route53AnalysisResult] = [r for r in result.get_data() if r is not None]

    if result.error_count > 0:
        console.print(f"[yellow]일부 오류 발생: {result.error_count}건[/yellow]")

    if not results:
        console.print("\n[yellow]분석 결과 없음[/yellow]")
        return

    total_unused = sum(r.empty_zones + r.ns_soa_only_zones for r in results)
    total_cost = sum(r.wasted_monthly_cost for r in results)

    console.print("\n[bold]종합 결과[/bold]")
    console.print(f"미사용 Hosted Zone: [yellow]{total_unused}개[/yellow] (${total_cost:,.2f}/월)")

    if hasattr(ctx, "is_sso_session") and ctx.is_sso_session() and ctx.accounts:
        identifier = ctx.accounts[0].id
    elif ctx.profile_name:
        identifier = ctx.profile_name
    else:
        identifier = "default"

    output_path = OutputPath(identifier).sub("route53-empty").with_date().build()
    filepath = generate_report(results, output_path)

    console.print(f"\n[bold green]완료![/bold green] {filepath}")
    open_in_explorer(output_path)
