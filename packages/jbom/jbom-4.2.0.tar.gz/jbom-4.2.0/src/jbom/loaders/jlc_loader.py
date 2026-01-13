"""
Loader for JLC Private Inventory Export (XLSX).
"""
from typing import List, Dict
from pathlib import Path

from jbom.common.types import InventoryItem, DEFAULT_PRIORITY

try:
    import openpyxl

    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class JLCPrivateInventoryLoader:
    """Loads inventory data from JLC 'My Parts Lib' export."""

    # Key fields in JLC export that identify it
    REQUIRED_JLC_HEADERS = ["JLCPCB Part #", "Category", "MFR Part #"]

    def __init__(self, inventory_path: Path):
        self.inventory_path = inventory_path
        self.inventory: List[InventoryItem] = []
        self.inventory_fields: List[str] = []

    def is_jlc_format(self, headers: List[str]) -> bool:
        """Check if headers match JLC export format."""
        # Check if at least some characteristic headers are present
        return "JLCPCB Part #" in headers and "Category" in headers

    def load(self) -> tuple[List[InventoryItem], List[str]]:
        """Load the JLC inventory."""
        if not EXCEL_SUPPORT:
            raise ImportError("JLC Loader requires openpyxl. pip install openpyxl")

        workbook = openpyxl.load_workbook(self.inventory_path, data_only=True)
        worksheet = workbook.active

        # JLC Export usually starts headers on row 1 or 2
        # Let's find the header row
        header_row = None
        for row_num in range(1, 5):
            cell_val = worksheet.cell(row_num, 1).value
            # Check for characteristic column
            if cell_val and (
                "JLCPCB Part #" in str(cell_val) or "Category" in str(cell_val)
            ):
                header_row = row_num
                break

        if not header_row:
            # Try column scan
            for row_num in range(1, 5):
                for col_num in range(1, 10):
                    cell_val = worksheet.cell(row_num, col_num).value
                    if cell_val and "JLCPCB Part #" in str(cell_val):
                        header_row = row_num
                        break
                if header_row:
                    break

        if not header_row:
            raise ValueError("Could not find 'JLCPCB Part #' header in JLC file.")

        # Extract headers
        headers = []
        for col in range(1, worksheet.max_column + 1):
            val = worksheet.cell(header_row, col).value
            if val:
                headers.append(str(val).strip())

        self.inventory_fields = headers

        # Load rows
        rows = []
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_data = {}
            has_data = False
            for idx, header in enumerate(headers):
                cell_val = worksheet.cell(row_num, idx + 1).value
                val_str = str(cell_val).strip() if cell_val is not None else ""
                row_data[header] = val_str
                if val_str:
                    has_data = True

            if has_data:
                rows.append(row_data)

        workbook.close()
        self._process_rows(rows)
        return self.inventory, self.inventory_fields

    def _process_rows(self, rows: List[Dict[str, str]]):
        """Convert JLC rows to InventoryItem."""
        for row in rows:
            # Map JLC fields to InventoryItem
            # JLCPCB Part # -> LCSC
            # Category -> Category
            # Description -> Description
            # MFR Part # -> MFGPN
            # Footprint -> Package
            # Unit Price($) -> (Price)

            lcsc = row.get("JLCPCB Part #", "")
            if not lcsc:
                continue

            item = InventoryItem(
                ipn=lcsc,  # JLC items use LCSC ID as their "Part Number"
                keywords=row.get("Description", ""),
                category=row.get("Category", ""),
                description=row.get("Description", ""),
                smd="",  # JLC doesn't explicitly have SMD flag column usually, infer?
                value="",  # No direct value column usually (it's in desc/params)
                type="",
                tolerance="",
                voltage="",
                amperage="",
                wattage="",
                lcsc=lcsc,
                manufacturer="",  # "Brand" not in this export format?
                mfgpn=row.get("MFR Part #", ""),
                datasheet="",
                package=row.get("Footprint", ""),
                fabricator="JLC",
                priority=DEFAULT_PRIORITY,
                source="JLC-Private",
                source_file=self.inventory_path,
                raw_data=row,
            )
            self.inventory.append(item)
