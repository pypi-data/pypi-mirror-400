import importlib
import io
import operator
from collections.abc import Callable
from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import List, Union, Optional, Dict, Type

import openpyxl
from fastapi import UploadFile
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import COLOR_INDEX, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from starlette._utils import is_async_callable
from starlette.responses import StreamingResponse, JSONResponse, FileResponse
from tortoise import Model
from pydantic import ValidationError

from fastgenerateapi import BaseModel, BaseView
from fastgenerateapi.api_view.mixin.dbmodel_mixin import DBModelMixin
from fastgenerateapi.api_view.mixin.response_mixin import ResponseMixin
from fastgenerateapi.api_view.mixin.tool_mixin import ToolMixin


class XlsxUtil:
    default_align = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=True,
        indent=0,
    )
    default_fill = PatternFill(start_color=Color(COLOR_INDEX[44]), end_color=Color(COLOR_INDEX[44]), fill_type='solid')
    thin_border = Border(top=Side('thin'),left=Side('thin'),right=Side('thin'),bottom=Side('thin'))

    @classmethod
    def write_headers(cls, sh: Worksheet, xlsx_headers: List[str]) -> List[int]:
        """
        写入第一行信息
        :return:
        """
        col_width_list = []
        sh.row_dimensions[1].height = 26
        for col, header in enumerate(xlsx_headers, 1):
            sh.cell(1, col).value = header
            sh.cell(1, col).alignment = XlsxUtil.default_align
            sh.cell(1, col).fill = XlsxUtil.default_fill
            sh.cell(1, col).alignment = XlsxUtil.default_align
            sh.cell(1, col).border = cls.thin_border
            col_width_list.append(len(header.encode('gb18030')))

        return col_width_list

    @staticmethod
    def write_content(model_list: List[Model]):
        """
        填写内容部分
        :return:
        """
        # 跳过标题，从第二行开始写入
        for row, model in enumerate(model_list, 2):
            ...

        return

    @staticmethod
    def adaptive_format(sh: Worksheet, col_max_len_list: List[int], height_num: int):
        """
        自适应宽度
        :return:
        """
        # 设置自适应列宽
        for i, col_max_len in enumerate(col_max_len_list, 1):
            # 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
            max_width = col_max_len + 4
            if max_width > 256:
                max_width = 256
            sh.column_dimensions[get_column_letter(i)].width = max_width
        for y in range(2, height_num + 2):
            sh.row_dimensions[y].height = 18

        return

    @staticmethod
    async def export_xlsx(
            xlsx_headers: List[str],
            model_list: List[Model],
            fields: List[str],
            index: Optional[bool] = False,
            model_handler: Callable = None,
            model_handler_list: List[Callable] = None,
            fields_handler: dict = None,
            filename: Optional[str] = '导出文件.xlsx',
            file_save_path: Optional[str] = None,
            sheet: Optional[str] = None,
            model_class: Optional[Model] = None,
            *args,
            **kwargs,
    ) -> StreamingResponse:
        """
        导出excel文件
        :param xlsx_headers: 第一行表头
        :param model_list: 模型列表
        :param fields: 字段列表
        :param index: 是否导出序号字段
        :param model_handler: 模型列表的处理方法
        :param model_handler_list: 单个模型处理方法
        :param fields_handler: 字段处理方法
        :param filename: 导出文件名
        :param file_save_path: 指定文件路径保存，不返回
        :param sheet: sheet名称
        :param model_class: 当无title时，使用表名称
        :return:
        """
        wb = openpyxl.Workbook()
        col_width_list = []

        def write(sh, row, col, value):
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            sh.cell(row, col).value = value
            sh.cell(row, col).alignment = XlsxUtil.default_align
            if col_width_list[col - 1] < len(str(value).encode('gb18030')):
                col_width_list[col - 1] = len(str(value).encode('gb18030'))

        start_row = 1
        try:
            sh = wb.active
            sh.title = sheet or (f'{model_class._meta.table_description}' if model_class else "sheet1")

            col_width_list = XlsxUtil.write_headers(sh, xlsx_headers)

            if model_handler:
                if is_async_callable(model_handler):
                    model_list = await model_handler(model_list, *args, **kwargs)
                else:
                    model_list = model_handler(model_list, *args, **kwargs)

            for row, model in enumerate(model_list, start_row + 1):
                if model_handler_list:
                    for model_handler in model_handler_list:
                        if is_async_callable(model_handler):
                            model = await model_handler(model, *args, **kwargs)
                        else:
                            model = model_handler(model, *args, **kwargs)
                start_col = 1
                if index:
                    write(sh, row, start_col, str(row-start_row))
                    start_col += 1
                for col, field in enumerate(fields, start_col):
                    info = getattr(model, field, "")
                    handler = fields_handler.get(field)
                    if handler and hasattr(handler, "__call__"):
                        if is_async_callable(handler):
                            info = await handler(info, *args, **kwargs)
                        else:
                            info = handler(info, *args, **kwargs)
                    write(sh, row, col, info)

            XlsxUtil.adaptive_format(sh, col_width_list, len(xlsx_headers))
        finally:
            if file_save_path:
                wb.save(file_save_path)
                return ResponseMixin.success(msg="请求成功")
            bytes_io = io.BytesIO()
            wb.save(bytes_io)
            bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, filename=filename, is_xlsx=True)

    @staticmethod
    async def excel_model(
            xlsx_headers: List[str] = None,
            sheet: Optional[str] = None,
            model_class: Optional[Model] = None,
            filename: Optional[str] = '导入模板.xlsx',
            excel_model_path: Optional[str] = None,
            modules: str = "openpyxl",
    ) -> Union[FileResponse, StreamingResponse]:
        """
        导出excel模板
        :param xlsx_headers: 第一行表头
        :param sheet: sheet名称
        :param model_class: 当无title时，使用表名称
        :param filename: 导出文件名
        :param excel_model_path: 使用已存在文件直接导出
        :param modules: 暂只支持 openpyxl
        :return:
        """
        if excel_model_path:
            return FileResponse(
                path=excel_model_path,
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=UTF-8",
            )

        limit_modules = ["openpyxl", "xlsxwriter"]
        if modules not in limit_modules:
            return ResponseMixin.error(msg=f"export xlsx modules only import {'、'.join(limit_modules)}")
        try:
            wb = importlib.import_module(modules).Workbook()
        except Exception:
            return ResponseMixin.error(msg=f"please pip install {modules}")
        try:
            sh = wb.active
            sh.title = sheet or (f'{model_class._meta.table_description}' if model_class else "sheet1")

            col_width_list = XlsxUtil.write_headers(sh, xlsx_headers)

            XlsxUtil.adaptive_format(sh, col_width_list, len(xlsx_headers))
        finally:
            bytes_io = io.BytesIO()
            wb.save(bytes_io)
            bytes_io.seek(0)

        return ResponseMixin.stream(bytes_io, filename=filename, is_xlsx=True)

    @staticmethod
    async def excel_row_data_dict(row_list, fields: List[str]) -> dict:
        row_data = {}
        for i in range(min(len(row_list), len(fields))):
            row_value = row_list[i].value
            if row_value not in [None, "", " "]:
                if isinstance(row_value, int):
                    row_value = str(row_value)
                row_data[fields[i]] = row_value
            else:
                row_data[fields[i]] = None
        if  any(row_data.values()):
            return row_data
        return {}

    @staticmethod
    async def get_import_xlsx_data(
            file: UploadFile,
            xlsx_headers: List[str],
            fields: List[str],
            sheet: Optional[str] = None,
    ) -> List[dict]:
        if not file:
            return ResponseMixin.error(msg=f"请先选择合适的文件")
        file_data_list = []
        with NamedTemporaryFile() as tmp2:
            tmp2.write(await file.read())
            wb = openpyxl.load_workbook(tmp2, read_only=True, data_only=True)
            try:
                if sheet:
                    ws = wb[sheet]
                else:
                    ws = wb.active

                header_row = ws[1]
                header_list = []
                for msg in header_row:
                    header_list.append(str(msg.value).replace(" ", ''))

                if len(header_list) != len(xlsx_headers):
                    return ResponseMixin.error(msg="文件首行长度校验错误")
                if not operator.eq(header_list, xlsx_headers):
                    return ResponseMixin.error(msg="文件首行内容校验错误")

                for row in range(2, ws.max_row + 1):
                    row_data = await XlsxUtil.excel_row_data_dict(ws[row], fields)
                    if not row_data:
                        continue
                    row_data["_row_index"] = row
                    file_data_list.append(row_data)
            finally:
                wb.close()

        return file_data_list

    @classmethod
    async def import_xlsx(
            cls,
            file: UploadFile,
            xlsx_headers: List[str],
            fields: List[str],
            model_class: Optional[Type[Model]],
            create_schema: Optional[Type[BaseModel]],
            fields_handler: Callable = None,
            field_handler_list: List[Callable] = None,
            # storage_path: Union[str, Path],
            sheet: Optional[str] = None,
            *args, **kwargs,
    ) -> JSONResponse:
        """
        导入只能用于一张表
        :param file:
        :param xlsx_headers: 请求文件首行文字校验
        :param fields: 值对应的字段
        :param model_class: 用于创建数据
        :param create_schema: 用于校验必填字段
        :param fields_handler: 总数据处理方法
        :param field_handler_list: 每行数据处理方法
        :param sheet: sheet名称
        :return:
        """
        file_data_list = await cls.get_import_xlsx_data(file, xlsx_headers, fields, sheet)
        if not file_data_list:
            return ResponseMixin.error(msg="导入数据不能为空")
        create_list = []
        if fields_handler:
            if is_async_callable(fields_handler):
                file_data_list = await fields_handler(file_data_list, *args, **kwargs)
            else:
                file_data_list = fields_handler(file_data_list, *args, **kwargs)
        for file_data in file_data_list:
            try:
                create_dict = create_schema(**file_data).model_dump(exclude_unset=True)
                if field_handler_list:
                    for field_handler in field_handler_list:
                        if is_async_callable(field_handler):
                            create_dict = await field_handler(create_dict, *args, **kwargs)
                        else:
                            create_dict = field_handler(create_dict, *args, **kwargs)
                create_obj = model_class(**create_dict)
            except ValidationError as e:
                alise_dict = ToolMixin.get_schema_alise_to_name(create_schema)
                _error_field = e.errors()[0].get('loc')[0]
                error_field = alise_dict.get(_error_field, _error_field)
                description = DBModelMixin.get_field_description(model_class, error_field)
                if not file_data.get(error_field):
                    return ResponseMixin.error(msg=f"第{file_data['_row_index']}行【{description}】不能为空")
                return ResponseMixin.error(msg=f"第{file_data['_row_index']}行【{description}】填写错误")
            create_list.append(create_obj)
        await model_class.bulk_create(create_list)

        return ResponseMixin.success(msg='创建成功')

