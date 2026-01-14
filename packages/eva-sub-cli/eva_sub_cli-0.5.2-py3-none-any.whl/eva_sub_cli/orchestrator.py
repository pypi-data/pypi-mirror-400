#!/usr/bin/env python
import csv
import os
import re
from collections import defaultdict

import requests
from ebi_eva_common_pyutils.config import WritableConfig
from ebi_eva_common_pyutils.logger import logging_config
from openpyxl.reader.excel import load_workbook
from packaging import version
from requests import HTTPError
from retry import retry

import eva_sub_cli
from eva_sub_cli import MINIMUM_METADATA_XLSX_TEMPLATE_VERSION
from eva_sub_cli import SUB_CLI_CONFIG_FILE, __version__
from eva_sub_cli.exceptions.invalid_file_type_exception import InvalidFileTypeError
from eva_sub_cli.exceptions.metadata_template_version_exception import MetadataTemplateVersionException, \
    MetadataTemplateVersionNotFoundException
from eva_sub_cli.exceptions.submission_not_found_exception import SubmissionNotFoundException
from eva_sub_cli.exceptions.submission_status_exception import SubmissionStatusException
from eva_sub_cli.file_utils import is_vcf_file
from eva_sub_cli.metadata import EvaMetadataJson
from eva_sub_cli.submission_ws import SubmissionWSClient
from eva_sub_cli.submit import StudySubmitter, SUB_CLI_CONFIG_KEY_SUBMISSION_ID, \
    SUB_CLI_CONFIG_KEY_SUBMISSION_UPLOAD_URL
from eva_sub_cli.utils import get_project_title_from_ena
from eva_sub_cli.validators.docker_validator import DockerValidator
from eva_sub_cli.validators.native_validator import NativeValidator
from eva_sub_cli.validators.validator import READY_FOR_SUBMISSION_TO_EVA, ALL_VALIDATION_TASKS

VALIDATE = 'validate'
SUBMIT = 'submit'
DOCKER = 'docker'
NATIVE = 'native'

logger = logging_config.get_logger(__name__)


def get_vcf_files(mapping_file):
    vcf_files = []
    with open(mapping_file) as open_file:
        reader = csv.DictReader(open_file, delimiter=',')
        for row in reader:
            vcf_files.append(row['vcf'])
    return vcf_files


def remove_non_vcf_files_from_metadata(metadata_json, metadata_xlsx):
    if metadata_json:
        metadata = EvaMetadataJson(metadata_json)
        if any(not is_vcf_file(f['fileName']) for f in metadata.files):
            vcf_files = [f for f in metadata.files if is_vcf_file(f['fileName'])]
            metadata.set_files(vcf_files)
            metadata.write(metadata_json)
            logger.warning(f"Some files mentioned in the metadata file's ({metadata_json}) files section are not VCF files and have been removed.")
    elif metadata_xlsx:
        workbook = load_workbook(metadata_xlsx)
        files_sheet = workbook['Files']
        files_headers = {cell.value: cell.column - 1 for cell in files_sheet[1]}

        rows_to_delete = []
        for i, row in enumerate(files_sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue
            if not is_vcf_file(row[files_headers['File Name']]):
                rows_to_delete.append(i)

        if rows_to_delete:
            # Delete from bottom to top so that we don't invalidate the row indices
            for row_idx in reversed(rows_to_delete):
                files_sheet.delete_rows(row_idx, 1)

            workbook.save(metadata_xlsx)
            logger.warning(f"Some files mentioned in the metadata xlsx's ({metadata_xlsx}) Files sheet are not VCF files and have been removed.")


def get_project_title_and_create_vcf_files_mapping(submission_dir, metadata_json, metadata_xlsx, metadata_xlsx_version):
    """
    Get project title and mapping between VCF files and reference FASTA files, from two sources: metadata JSON file
    or metadata XLSX file.

    :param submission_dir: Directory where mapping file will be saved
    :param metadata_json: Metadata JSON from command line, if present
    :param metadata_xlsx: Metadata XLSX from command line, if present
    :param metadata_xlsx_version: Version of metadata XLSX
    :return: Project title and path to the mapping file
    """
    mapping_file = os.path.join(submission_dir, 'vcf_mapping_file.csv')
    with open(mapping_file, 'w') as open_file:
        writer = csv.writer(open_file, delimiter=',')
        writer.writerow(['vcf', 'fasta', 'report'])

        vcf_files_mapping = []
        if metadata_json:
            project_title, vcf_files_mapping = get_project_and_vcf_fasta_mapping_from_metadata_json(metadata_json)
        elif metadata_xlsx:
            project_title, vcf_files_mapping = get_project_and_vcf_fasta_mapping_from_metadata_xlsx(metadata_xlsx, metadata_xlsx_version)

        # Filter out non-vcf files
        vcf_files_mapping = [(vcf, fasta, report) for vcf, fasta, report in vcf_files_mapping if is_vcf_file(vcf)]
        validate_vcf_mapping(vcf_files_mapping)
        for mapping in vcf_files_mapping:
            writer.writerow(mapping)

    return project_title, mapping_file


def validate_vcf_mapping(vcf_mapping):
    """
    Validate that VCF files and FASTA files in the mapping are present and FASTA files are not compressed.

    :param vcf_mapping: iterable of triples (VCF file path, reference FASTA path, optional assembly report path)
    :return:
    """
    for vcf_file, fasta_file, report_file in vcf_mapping:
        if not (vcf_file and os.path.isfile(vcf_file)):
            raise FileNotFoundError(f'The variant file {vcf_file} does not exist, please check the file path.')
        if not (fasta_file and os.path.isfile(fasta_file)):
            raise FileNotFoundError(f'The reference fasta {fasta_file} does not exist, please check the file path.')
        if fasta_file.lower().endswith('gz'):
            raise InvalidFileTypeError(f'The reference fasta {fasta_file} is compressed, please uncompress the file.')
        if report_file and not os.path.isfile(report_file):
            raise FileNotFoundError(f'The assembly report file {report_file} does not exist, please check the file '
                                    f'path.')


def get_project_and_vcf_fasta_mapping_from_metadata_json(metadata_json):
    metadata = EvaMetadataJson(metadata_json)

    project_title = metadata.project.get('title')
    if not project_title:
        project_accession = metadata.project.get('projectAccession')
        if project_accession:
            project_title = get_project_title_from_ena(project_accession)

    vcf_fasta_report_mapping = []

    analysis_alias_dict = defaultdict(dict)
    for analysis in metadata.analyses:
        analysis_alias_dict[analysis['analysisAlias']]['referenceFasta'] = analysis['referenceFasta']
        analysis_alias_dict[analysis['analysisAlias']]['assemblyReport'] = analysis['assemblyReport'] \
            if 'assemblyReport' in analysis else ''

    for file_dict in metadata.resolved_files:
        reference_fasta = analysis_alias_dict[file_dict['analysisAlias']]['referenceFasta']
        assembly_report = analysis_alias_dict[file_dict['analysisAlias']]['assemblyReport']
        vcf_fasta_report_mapping.append([os.path.abspath(file_dict['fileName']),
                                         os.path.abspath(reference_fasta),
                                         os.path.abspath(assembly_report) if assembly_report else ''])

    return project_title, vcf_fasta_report_mapping


def get_sub_cli_version():
    if version.parse(eva_sub_cli.__version__).is_devrelease:
        version_values = [int(v) for v in version.parse(eva_sub_cli.__version__).base_version.split('.')]
        major = version_values[0]
        minor = version_values[1] if len(version_values) > 1 else 0
        patch = version_values[2] if len(version_values) > 2 else 0
        if patch > 0:
            patch -= 1
        elif minor > 0:
            minor -= 1
            patch = 0
        elif major > 0:
            major -= 1
            minor = 0
            patch = 0
        return f"{major}.{minor}.{patch}"
    else:
        return version.parse(eva_sub_cli.__version__).base_version


@retry(exceptions=(HTTPError,), tries=3, delay=2, backoff=1.2, jitter=(1, 3))
def get_sub_cli_github_tags():
    url = f"https://api.github.com/repos/EBIvariation/eva-sub-cli/tags"
    response = requests.get(url)
    if response.status_code == 200:
        tags = [tag["name"][1:] for tag in response.json()]
        return tags
    else:
        return []


def get_metadata_xlsx_template_link():
    sub_cli_version = get_sub_cli_version()
    sub_cli_tags = get_sub_cli_github_tags()
    if sub_cli_version in sub_cli_tags:
        return f'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/refs/tags/v{sub_cli_version}/eva_sub_cli/etc/EVA_Submission_template.xlsx'
    else:
        return 'https://raw.githubusercontent.com/EBIvariation/eva-sub-cli/main/eva_sub_cli/etc/EVA_Submission_template.xlsx'


def verify_and_get_metadata_xlsx_version(metadata_xlsx, min_req_version):
    workbook = load_workbook(metadata_xlsx)
    instructions_sheet = workbook['PLEASE READ FIRST']
    xlsx_sheet_version_value = instructions_sheet[3][0].value
    match = re.search(r'(\d+\.\d+\.\d+)', '' if xlsx_sheet_version_value is None else xlsx_sheet_version_value)
    xlsx_version = match.group(1) if match else None
    if xlsx_version:
        if version.parse(xlsx_version) < version.parse(min_req_version):
            raise MetadataTemplateVersionException(
                f"Metadata template version {xlsx_version} is lower than min required {min_req_version}. "
                f"Please download the correct template from EVA github project {get_metadata_xlsx_template_link()}")
    else:
        raise MetadataTemplateVersionNotFoundException(
            f"No version information found in metadata xlsx sheet {metadata_xlsx}. "
            f"Please download the correct template from EVA github project {get_metadata_xlsx_template_link()}")

    return xlsx_version


def get_project_and_vcf_fasta_mapping_from_metadata_xlsx(metadata_xlsx, metadata_xlsx_version):
    workbook = load_workbook(metadata_xlsx)

    project_sheet = workbook['Project']
    project_headers = {}

    if version.parse(metadata_xlsx_version) < version.parse('3.0.0'):
        for cell in project_sheet[1]:
            project_headers[cell.value] = cell.column
        project_title = project_sheet.cell(row=2, column=project_headers['Project Title']).value
    else:
        for cell in project_sheet[3]:
            project_headers[cell.value] = cell.column
        project_title = project_sheet.cell(row=4, column=project_headers['Project Title']).value
        if not project_title:
            project_accession = project_sheet.cell(row=4, column=project_headers['Project Accession']).value
            if project_accession:
                project_title = get_project_title_from_ena(project_accession)

    vcf_fasta_report_mapping = []
    analysis_alias_sheet = workbook['Analysis']
    analysis_headers = {}
    for cell in analysis_alias_sheet[1]:
        analysis_headers[cell.value] = cell.column - 1

    analysis_alias_dict = {}
    for row in analysis_alias_sheet.iter_rows(min_row=2, values_only=True):
        analysis_alias = row[analysis_headers['Analysis Alias']]
        reference_fasta = row[analysis_headers['Reference Fasta Path']]
        analysis_alias_dict[analysis_alias] = reference_fasta

    files_sheet = workbook['Files']
    files_headers = {}
    for cell in files_sheet[1]:
        files_headers[cell.value] = cell.column - 1

    for row in files_sheet.iter_rows(min_row=2, values_only=True):
        file_name = row[files_headers['File Name']]
        if file_name:
            file_name = os.path.abspath(file_name)
        analysis_alias = row[files_headers['Analysis Alias']]
        reference_fasta = analysis_alias_dict[analysis_alias]
        if reference_fasta:
            reference_fasta = os.path.abspath(reference_fasta)
        if file_name and reference_fasta:
            vcf_fasta_report_mapping.append([file_name, reference_fasta, ''])

    return project_title, vcf_fasta_report_mapping


def check_validation_required(tasks, sub_config, username=None, password=None):
    # Validation is mandatory so if submit is requested then VALIDATE must have run before or be requested as well
    if SUBMIT not in tasks:
        return False
    if not sub_config.get(READY_FOR_SUBMISSION_TO_EVA, False):
        return True
    # If we are working with an existing submission check its status to see if it was submitted and failed before.
    submission_id = sub_config.get(SUB_CLI_CONFIG_KEY_SUBMISSION_ID, None)
    if submission_id:
        try:
            submission_status = SubmissionWSClient(username, password).get_submission_status(submission_id)
            if submission_status == 'FAILED':
                # Reset the submission_id which will force the creation of a new one
                sub_config.set(SUB_CLI_CONFIG_KEY_SUBMISSION_ID, value=None)
                sub_config.set(SUB_CLI_CONFIG_KEY_SUBMISSION_UPLOAD_URL, value=None)
                return True
            else:
                return False
        except requests.HTTPError as ex:
            if ex.response.status_code == 404:
                logger.error(
                    f'Submission with id {submission_id} could not be found: '
                    f'status code: {ex.response.status_code} response: {ex.response.text}')
                raise SubmissionNotFoundException(f'Submission with id {submission_id} could not be found')
            else:
                logger.error(f'Error occurred while getting status of the submission with Id {submission_id}: '
                             f'status code: {ex.response.status_code} response: {ex.response.text}')
                raise SubmissionStatusException(f'Error occurred while getting status of the submission '
                                                f'with Id {submission_id}')

    logger.debug(f'submission id not found in config. This might be the first time user is submitting')
    return False


def orchestrate_process(submission_dir, metadata_json, metadata_xlsx,
                        tasks, executor, validation_tasks=ALL_VALIDATION_TASKS, username=None, password=None,
                        shallow_validation=False, nextflow_config=None, **kwargs):
    # load config
    config_file_path = os.path.join(submission_dir, SUB_CLI_CONFIG_FILE)
    sub_config = WritableConfig(config_file_path, version=__version__)

    metadata_file = metadata_json or metadata_xlsx
    if not os.path.exists(os.path.abspath(metadata_file)):
        raise FileNotFoundError(f'The provided metadata file {os.path.abspath(metadata_file)} does not exist')

    if metadata_json:
        metadata_json = os.path.abspath(metadata_json)
    metadata_xlsx_version = None
    if metadata_xlsx:
        metadata_xlsx = os.path.abspath(metadata_xlsx)
        # check metadata xlsx version is not lower than the required min metadata template version
        metadata_xlsx_version = verify_and_get_metadata_xlsx_version(metadata_xlsx, MINIMUM_METADATA_XLSX_TEMPLATE_VERSION)

    # remove non vcf files from metadata
    remove_non_vcf_files_from_metadata(metadata_json, metadata_xlsx)

    # Get the provided Project Title and VCF files mapping (VCF, Fasta and Report)
    project_title, vcf_files_mapping = get_project_title_and_create_vcf_files_mapping(
        submission_dir, metadata_json, metadata_xlsx, metadata_xlsx_version
    )
    vcf_files = get_vcf_files(vcf_files_mapping)

    if VALIDATE not in tasks and check_validation_required(tasks, sub_config, username, password):
        tasks.append(VALIDATE)

    if VALIDATE in tasks:
        if executor == DOCKER:
            validator = DockerValidator(vcf_files_mapping, submission_dir, project_title, metadata_json, metadata_xlsx,
                                        metadata_xlsx_version, validation_tasks=validation_tasks,
                                        shallow_validation=shallow_validation, submission_config=sub_config)
        # default to native execution
        else:
            validator = NativeValidator(vcf_files_mapping, submission_dir, project_title, metadata_json, metadata_xlsx,
                                        metadata_xlsx_version, validation_tasks=validation_tasks,
                                        shallow_validation=shallow_validation, submission_config=sub_config,
                                        nextflow_config=nextflow_config)
        with validator:
            validator.validate_and_report()
            if not metadata_json:
                metadata_json = os.path.join(validator.output_dir, 'metadata.json')
            sub_config.set('metadata_json', value=metadata_json)
            sub_config.set('vcf_files', value=vcf_files)

    if SUBMIT in tasks:
        with StudySubmitter(submission_dir, submission_config=sub_config, username=username,
                            password=password) as submitter:
            submitter.submit()
