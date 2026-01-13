"""Multi-sheet workbook handling and cross-sheet reference management."""

from typing import Dict, List, Tuple, Any, Optional, Set
import pandas as pd
import os
import re


class WorkbookManager:
    """
    Manages multi-sheet workbooks and cross-sheet references.
    
    This class provides utilities to:
    1. Load and process entire Excel workbooks with multiple sheets
    2. Handle cross-sheet references and relationships
    3. Compress entire workbooks
    4. Identify inter-sheet relationships
    """
    
    def __init__(self) -> None:
        """Initialize the workbook manager."""
        self.sheets: Dict[str, pd.DataFrame] = {}
        self.sheet_metadata: Dict[str, Dict[str, Any]] = {}
        self.cross_references: Dict[str, Set[str]] = {}
        
    def load_workbook(self, excel_path: str) -> Dict[str, pd.DataFrame]:
        """
        Load all sheets from an Excel workbook.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            Dictionary mapping sheet names to dataframes
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
            
        # Store path for later use
        self._last_loaded_path = excel_path
            
        # Load all sheets into a dict of dataframes
        sheet_dict = pd.read_excel(excel_path, sheet_name=None)
        self.sheets = sheet_dict  # type: ignore
        
        # Initialize metadata for each sheet
        for sheet_name in sheet_dict:
            self.sheet_metadata[sheet_name] = {
                "shape": sheet_dict[sheet_name].shape,
                "non_empty_cells": (~sheet_dict[sheet_name].isna()).sum().sum(),
                "has_header_row": self._detect_header_row(sheet_dict[sheet_name])
            }
        
        return sheet_dict
    
    def _detect_header_row(self, df: pd.DataFrame) -> bool:
        """
        Detect if the first row is likely a header row.
        
        Args:
            df: Dataframe to analyze
            
        Returns:
            Boolean indicating if first row appears to be a header
        """
        # Simple heuristic: if first row has different data types than rows below
        if df.shape[0] <= 1:
            return False
            
        first_row_types = [type(val) for val in df.iloc[0] if pd.notna(val)]
        rest_types = []
        
        # Check next few rows for types
        for i in range(1, min(4, df.shape[0])):
            rest_types.extend([type(val) for val in df.iloc[i] if pd.notna(val)])
            
        # If first row has mostly strings and rest has numbers, likely a header
        first_row_string_ratio = sum(1 for t in first_row_types if t == str) / len(first_row_types) if first_row_types else 0
        rest_string_ratio = sum(1 for t in rest_types if t == str) / len(rest_types) if rest_types else 0
        
        return first_row_string_ratio > 0.7 and first_row_string_ratio > rest_string_ratio * 1.5
    
    def detect_cross_sheet_references(self) -> Dict[str, Set[str]]:
        """
        Detect references between sheets using formula analysis.
        
        Returns:
            Dictionary mapping sheet names to sets of referenced sheets
        """
        if not self.sheets:
            raise ValueError("No workbook loaded. Call load_workbook first.")
            
        try:
            import openpyxl
            
            # We need to reopen the workbook to access formulas
            excel_path = getattr(self, "_last_loaded_path", None)
            if not excel_path or not os.path.exists(excel_path):
                raise ValueError("Cannot detect cross-references: original Excel file path not available")
                
            workbook = openpyxl.load_workbook(excel_path, data_only=False)
            
            # Pattern to detect sheet references in formulas
            sheet_ref_pattern = re.compile(r'([\'"]?)([^!\'"\[\]]+)(?:\1)!')
            
            cross_refs = {}
            
            # Check each sheet for references to other sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                references = set()
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            # Look for sheet references in the formula
                            matches = sheet_ref_pattern.findall(cell.value)
                            for _, ref_sheet in matches:
                                if ref_sheet in workbook.sheetnames and ref_sheet != sheet_name:
                                    references.add(ref_sheet)
                
                cross_refs[sheet_name] = references
                
            self.cross_references = cross_refs
            return cross_refs
            
        except ImportError:
            raise ImportError("openpyxl is required for cross-reference detection")
    
    def get_sheet_relationship_graph(self) -> Dict[str, Any]:
        """
        Generate a directed graph of sheet relationships.
        
        Returns:
            Dictionary with graph representation of sheet relationships
        """
        if not self.cross_references:
            self.detect_cross_sheet_references()
            
        # Create nodes for each sheet
        nodes = []
        for sheet_name, metadata in self.sheet_metadata.items():
            nodes.append({
                "id": sheet_name,
                "name": sheet_name,
                "rows": metadata["shape"][0],
                "columns": metadata["shape"][1],
                "nonEmptyCells": metadata["non_empty_cells"]
            })
            
        # Create edges for references
        edges = []
        for source, targets in self.cross_references.items():
            for target in targets:
                edges.append({
                    "source": source,
                    "target": target
                })
                
        return {
            "nodes": nodes,
            "edges": edges
        }
    
    def compress_workbook(self, compressor: Any) -> Dict[str, Any]:
        """
        Compress all sheets in the workbook.
        
        Args:
            compressor: SheetCompressor instance to use for compression
            
        Returns:
            Dictionary with compression results for each sheet
        """
        if not self.sheets:
            raise ValueError("No workbook loaded. Call load_workbook first.")
            
        compression_results = {}
        
        for sheet_name, df in self.sheets.items():
            # Compress each sheet
            compression_results[sheet_name] = compressor.compress(df)
            
        # Calculate overall statistics
        total_original_cells = sum(df.shape[0] * df.shape[1] for df in self.sheets.values())
        total_compressed_cells = sum(
            result["compressed_data"].shape[0] * result["compressed_data"].shape[1] 
            for result in compression_results.values()
        )
        
        overall_ratio = total_original_cells / total_compressed_cells if total_compressed_cells > 0 else 1
        
        # Add workbook-level summary
        compression_results["__workbook_summary__"] = {
            "total_sheets": len(self.sheets),
            "total_original_cells": total_original_cells,
            "total_compressed_cells": total_compressed_cells,
            "overall_compression_ratio": overall_ratio,
            "sheet_relationships": self.get_sheet_relationship_graph() if self.cross_references else None
        }
        
        return compression_results
    
    def encode_workbook_for_llm(self, compression_results: Dict[str, Any]) -> str:
        """
        Generate LLM-ready encoding of the entire workbook.
        
        Args:
            compression_results: Output from compress_workbook
            
        Returns:
            LLM-ready text representation of the workbook
        """
        if "__workbook_summary__" not in compression_results:
            raise ValueError("Invalid compression results. Use output from compress_workbook.")
            
        lines = []
        
        # Add workbook summary
        summary = compression_results["__workbook_summary__"]
        lines.append(f"# Workbook Analysis (Compressed {summary['overall_compression_ratio']:.1f}x)")
        lines.append(f"Contains {summary['total_sheets']} sheets with {summary['total_original_cells']} total cells")
        lines.append("")
        
        # Add sheet relationship information if available
        if summary.get("sheet_relationships"):
            lines.append("## Sheet Relationships")
            relationships = summary["sheet_relationships"]
            
            for edge in relationships["edges"]:
                lines.append(f"- '{edge['source']}' references '{edge['target']}'")
            
            lines.append("")
        
        # Add each sheet's compression data
        for sheet_name, result in compression_results.items():
            if sheet_name == "__workbook_summary__":
                continue
                
            lines.append(f"## Sheet: {sheet_name} (Compressed {result['compression_ratio']:.1f}x)")
            
            # Add inverted index data
            if "inverted_index" in result:
                lines.append("### Values (value|addresses):")
                for value, addresses in list(result["inverted_index"].items())[:15]:  # Limit to first 15
                    addr_str = ",".join(addresses[:5])
                    if len(addresses) > 5:
                        addr_str += f" (+{len(addresses)-5} more)"
                    lines.append(f"{value}|{addr_str}")
            
            lines.append("")
        
        return "\n".join(lines)
    
    def get_sheet_importance_ranking(self) -> List[Tuple[str, float]]:
        """
        Rank sheets by their importance in the workbook.
        
        Returns:
            List of (sheet_name, importance_score) tuples, sorted by importance
        """
        # Try to get cross-references, but continue with empty refs if not available
        if not hasattr(self, 'cross_references') or not self.cross_references:
            try:
                self.detect_cross_sheet_references()
            except ValueError:
                # If we can't detect cross references, continue with empty dict
                self.cross_references = {}
                
        # Calculate in-degree (how many sheets reference this sheet)
        in_degree = {sheet: 0 for sheet in self.sheets}
        for sheet, refs in self.cross_references.items():
            for ref in refs:
                if ref in in_degree:
                    in_degree[ref] += 1
        
        # Calculate sheet complexity score
        complexity = {}
        for sheet, df in self.sheets.items():
            non_empty = (~df.isna()).sum().sum()
            total = df.shape[0] * df.shape[1]
            density = non_empty / total if total > 0 else 0
            
            # Importance based on size, density and references
            complexity[sheet] = (
                0.4 * non_empty +  # Size factor
                0.3 * density +    # Density factor
                0.3 * in_degree.get(sheet, 0)  # Reference factor
            )
            
        # Sort by importance score
        ranked_sheets = sorted(complexity.items(), key=lambda x: x[1], reverse=True)
        return ranked_sheets
