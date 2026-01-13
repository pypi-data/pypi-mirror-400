#!/usr/bin/env python3
"""
Export Dataform Inventory to Excel
Generates deployment checklist from Dataform repository .sqlx files
"""

import sys
import shutil
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook

from .file_utils import find_excel_file, extract_version_from_filename, find_template, scan_dataform_files
from .inventory_utils import create_inventory_from_files, load_existing_inventory as load_inventory, merge_inventories
from .excel_utils import apply_cell_formatting


class DataformInventoryExporter:
    """Export Dataform .sqlx files inventory to Excel with status tracking"""
    
    def __init__(self, repository_path: str, output_path: Optional[str] = None, 
                 repository_name: Optional[str] = None, version: Optional[str] = None):
        self.repository_path = Path(repository_path)
        self.output_path = Path(output_path) if output_path else self.repository_path
        self.repository_name = repository_name or self.repository_path.name
        
        # Search for existing file by repository name
        self.excel_path = find_excel_file(self.output_path, self.repository_name, version)
        self.version = version or extract_version_from_filename(self.excel_path)
        self.excel_filename = self.excel_path.name
        
        # Find template file
        self.template_path = find_template()
    
    def export_to_excel(self, inventory: pd.DataFrame):
        """Export inventory to Excel with formatting using template"""
        print(f"Using template: {self.template_path}")
        print(f"Exporting to Excel: {self.excel_path}")
        
        # Copy template to output location
        shutil.copy2(self.template_path, self.excel_path)
        
        # Load the workbook
        workbook = load_workbook(self.excel_path)
        
        # Find sheet that starts with "DataForm"
        target_sheet = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith('DataForm') or sheet_name.startswith('Dataform'):
                target_sheet = sheet_name
                break
        
        if not target_sheet:
            # Create new sheet if no DataForm sheet exists
            target_sheet = f'DataForm {self.version}'
            workbook.create_sheet(target_sheet)
        
        worksheet = workbook[target_sheet]
        
        # Write data starting from A9 (template has headers at row 7-8)
        start_row = 9
        start_col = 1
        
        # Write data rows directly (no headers, template already has them)
        for row_idx, row_data in enumerate(inventory.itertuples(index=False), start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply formatting using utility function
        apply_cell_formatting(worksheet, inventory, start_row)
        
        # Save workbook
        workbook.save(self.excel_path)
        
        print("Export completed successfully!")
        print(f"File saved to: {self.excel_path}")
    
    def run(self):
        """Main execution method"""
        print("\n========================================")
        print("  Dataform Deployment Checklist Export")
        print("========================================\n")
        
        # Validate repository path
        if not self.repository_path.exists():
            print(f"Error: Repository path not found: {self.repository_path}")
            sys.exit(1)
        
        # Ensure output directory exists
        self.output_path.mkdir(parents=True, exist_ok=True)
        
        # Scan for .sqlx and dependencies.js files
        files = scan_dataform_files(self.repository_path)
        
        if not files:
            print("Warning: No .sqlx files found in the repository!")
            return
        
        # Create new inventory
        new_inventory = create_inventory_from_files(files)
        
        # Load existing inventory
        existing_inventory = load_inventory(self.excel_path)
        
        # Merge inventories with status tracking
        final_inventory = merge_inventories(new_inventory, existing_inventory)
        
        # Export to Excel
        self.export_to_excel(final_inventory)
        
        print("\n========================================")
        print("Summary:")
        print(f"  Total files: {len(final_inventory)}")
        print(f"  Output file: {self.excel_filename}")
        print(f"  Location: {self.output_path}")
        print("========================================\n")
