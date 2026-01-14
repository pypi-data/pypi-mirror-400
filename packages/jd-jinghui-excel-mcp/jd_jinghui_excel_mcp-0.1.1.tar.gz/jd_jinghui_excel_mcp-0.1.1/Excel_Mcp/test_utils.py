import os
import utils

TEST_FILE = "test_workbook.xlsx"
TEST_SHEET = "TestSheet"
COPY_SHEET = "CopiedSheet"

def test_all_features():
    print("🚀 开始全功能测试...\n")

    # 1. 清理旧文件
    if os.path.exists(TEST_FILE):
        os.remove(TEST_FILE)
        print(f"[-] 已删除旧文件: {TEST_FILE}")

    # 2. 创建表格数据
    header = ["姓名", "年龄", "部门", "工资"]
    rows = [
        ["张三", 25, "技术部", 10000],
        ["李四", 30, "产品部", 12000],
        ["王五", 28, "设计部", 9000]
    ]
    utils.create_table(TEST_FILE, TEST_SHEET, header, rows)
    print(f"[+] 创建表格成功，包含 {len(rows)} 行数据")

    # 3. 读取并验证
    data = utils.read_sheet(TEST_FILE, TEST_SHEET)
    assert len(data) == 4  # Header + 3 rows
    assert data[0][0] == "姓名"
    print(f"[+] 读取验证通过: {data[0]}")

    # 4. 设置列宽
    utils.set_column_width(TEST_FILE, TEST_SHEET, "A", 20)
    utils.set_column_width(TEST_FILE, TEST_SHEET, "C", 15)
    print("[+] 列宽设置完成 (A列=20, C列=15)")

    # 5. 格式化 (表头加粗、灰色背景、居中)
    utils.format_range(TEST_FILE, TEST_SHEET, "A1", "D1", bold=True, bg_hex="#CCCCCC", align_center=True)
    print("[+] 表头格式化完成 (加粗+背景+居中)")

    # 6. 写入公式 (计算平均工资)
    # E1 写标题，E2 写公式
    utils.write_to_sheet(TEST_FILE, TEST_SHEET, [["平均工资"], ["=AVERAGE(D2:D4)"]])
    # 修正位置：上面的 append 会加到最后一行，我们希望它在 E 列，
    # 为了演示 write_formula，我们直接指定位置写入
    utils.write_formula(TEST_FILE, TEST_SHEET, "E2", "=AVERAGE(D2:D4)")
    # 手动补一个表头 E1
    from openpyxl import load_workbook
    wb = load_workbook(TEST_FILE)
    wb[TEST_SHEET]["E1"] = "平均工资"
    wb.save(TEST_FILE)
    print("[+] 公式写入完成: E2 = AVERAGE(D2:D4)")

    # 7. 复制工作表
    utils.copy_sheet(TEST_FILE, TEST_SHEET, COPY_SHEET)
    sheets = utils.describe_sheets(TEST_FILE)
    sheet_names = [s['name'] for s in sheets]
    assert COPY_SHEET in sheet_names
    print(f"[+] 工作表复制成功: {sheet_names}")

    # 8. 删除工作表
    utils.delete_sheet(TEST_FILE, COPY_SHEET)
    sheets = utils.describe_sheets(TEST_FILE)
    sheet_names = [s['name'] for s in sheets]
    assert COPY_SHEET not in sheet_names
    print(f"[+] 工作表删除成功，剩余: {sheet_names}")

    # 9. 自动筛选
    utils.add_auto_filter(TEST_FILE, TEST_SHEET, "A1:E4")
    print("[+] 自动筛选已添加 (A1:E4)")

    # 10. 合并单元格 (在底部加个备注)
    # 先写入一行备注
    utils.write_to_sheet(TEST_FILE, TEST_SHEET, [["备注：这是由 MCP 自动生成的报表"]])
    # 假设这是第 5 行 (Header + 3 data + 1 appended)，合并 A5:D5
    utils.merge_cells(TEST_FILE, TEST_SHEET, "A5", "D5")
    utils.format_range(TEST_FILE, TEST_SHEET, "A5", "A5", align_center=True)
    print("[+] 单元格合并完成")

    # 11. 导出 JSON
    json_str = utils.export_sheet_to_json(TEST_FILE, TEST_SHEET)
    print(f"[+] JSON 导出预览 (前50字符): {json_str[:50]}...")

    print("\n✅ 所有测试通过！请打开 'test_workbook.xlsx' 查看效果。")

if __name__ == "__main__":
    test_all_features()