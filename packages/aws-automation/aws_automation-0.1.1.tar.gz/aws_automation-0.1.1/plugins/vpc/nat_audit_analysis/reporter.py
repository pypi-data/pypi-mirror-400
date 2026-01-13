"""
NAT Gateway Excel 보고서 생성기

시트 구성:
1. Summary - 전체 요약 (미사용, 저사용, 비용)
2. Findings - 상세 분석 결과
3. All NAT Gateways - 전체 NAT Gateway 목록

Note:
    이 모듈은 Lazy Import 패턴을 사용합니다.
    openpyxl 등 무거운 의존성을 실제 사용 시점에만 로드합니다.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.styles import Alignment, Border, Font, PatternFill

from .analyzer import NATAnalysisResult, Severity, UsageStatus


class NATExcelReporter:
    """NAT Gateway Excel 보고서 생성기"""

    # 스타일 캐시 (lazy initialization)
    _styles_initialized: bool = False
    _HEADER_FILL: PatternFill | None = None
    _HEADER_FONT: Font | None = None
    _THIN_BORDER: Border | None = None
    _CENTER_ALIGN: Alignment | None = None
    _STATUS_FILLS: dict[UsageStatus, PatternFill] | None = None
    _SEVERITY_FILLS: dict[Severity, PatternFill] | None = None

    @classmethod
    def _init_styles(cls) -> None:
        """스타일 lazy 초기화"""
        if cls._styles_initialized:
            return

        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        cls._HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cls._HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        cls._THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cls._CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cls._STATUS_FILLS = {
            UsageStatus.UNUSED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            UsageStatus.LOW_USAGE: PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
            UsageStatus.NORMAL: PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
            UsageStatus.PENDING: PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid"),
            UsageStatus.UNKNOWN: PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid"),
        }
        cls._SEVERITY_FILLS = {
            Severity.CRITICAL: PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
            Severity.HIGH: PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
            Severity.MEDIUM: PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid"),
            Severity.LOW: PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid"),
            Severity.INFO: PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        }
        cls._styles_initialized = True

    @property
    def HEADER_FILL(self) -> PatternFill:
        self._init_styles()
        return self._HEADER_FILL

    @property
    def HEADER_FONT(self) -> Font:
        self._init_styles()
        return self._HEADER_FONT

    @property
    def THIN_BORDER(self) -> Border:
        self._init_styles()
        return self._THIN_BORDER

    @property
    def CENTER_ALIGN(self) -> Alignment:
        self._init_styles()
        return self._CENTER_ALIGN

    @property
    def STATUS_FILLS(self) -> dict[UsageStatus, PatternFill]:
        self._init_styles()
        assert self._STATUS_FILLS is not None
        return self._STATUS_FILLS

    @property
    def SEVERITY_FILLS(self) -> dict[Severity, PatternFill]:
        self._init_styles()
        assert self._SEVERITY_FILLS is not None
        return self._SEVERITY_FILLS

    def __init__(
        self,
        results: list[NATAnalysisResult],
        stats_list: list[dict[str, Any]],
    ):
        from openpyxl import Workbook as _Workbook

        self._init_styles()
        self.results = results
        self.stats_list = stats_list
        self.wb = _Workbook()

    def generate(self, output_dir: str) -> str:
        """Excel 보고서 생성"""
        # 기본 시트 제거
        self.wb.remove(self.wb.active)

        # 시트 생성
        self._create_summary_sheet()
        self._create_findings_sheet()
        self._create_all_nat_sheet()

        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"NAT_Gateway_Audit_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        os.makedirs(output_dir, exist_ok=True)
        self.wb.save(filepath)

        return filepath

    def _apply_header_style(self, ws, row: int = 1) -> None:
        """헤더 스타일 적용"""
        for cell in ws[row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN

    def _auto_column_width(self, ws, min_width: int = 10, max_width: int = 40) -> None:
        """열 너비 자동 조정"""
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self) -> None:
        """Summary 시트 생성"""
        ws = self.wb.create_sheet("Summary")

        # 제목
        ws["A1"] = "NAT Gateway 미사용 분석 보고서"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        ws["A2"] = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True)

        # 전체 통계
        totals = {
            "total_nat_count": sum(s.get("total_nat_count", 0) for s in self.stats_list),
            "unused_count": sum(s.get("unused_count", 0) for s in self.stats_list),
            "low_usage_count": sum(s.get("low_usage_count", 0) for s in self.stats_list),
            "normal_count": sum(s.get("normal_count", 0) for s in self.stats_list),
            "pending_count": sum(s.get("pending_count", 0) for s in self.stats_list),
            "total_monthly_cost": sum(s.get("total_monthly_cost", 0) for s in self.stats_list),
            "total_monthly_waste": sum(s.get("total_monthly_waste", 0) for s in self.stats_list),
            "total_annual_savings": sum(s.get("total_annual_savings", 0) for s in self.stats_list),
        }

        # 핵심 지표
        key_metrics = [
            ("", "", ""),
            ("핵심 지표", "값", "설명"),
            ("전체 NAT Gateway", totals["total_nat_count"], "분석 대상 총 개수"),
            ("미사용 (삭제 권장)", totals["unused_count"], "14일간 트래픽 0"),
            ("저사용 (검토 필요)", totals["low_usage_count"], "일평균 1GB 미만"),
            ("정상 사용", totals["normal_count"], "정상적으로 사용 중"),
            ("대기 중", totals["pending_count"], "생성 중이거나 7일 미만"),
            ("", "", ""),
            ("월간 총 비용", f"${totals['total_monthly_cost']:,.2f}", "모든 NAT Gateway 비용"),
            ("월간 낭비 추정", f"${totals['total_monthly_waste']:,.2f}", "미사용+저사용 낭비"),
            ("연간 절감 가능액", f"${totals['total_annual_savings']:,.2f}", "삭제 시 절감 가능"),
        ]

        start_row = 4
        for i, (item, value, desc) in enumerate(key_metrics):
            row = start_row + i
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=desc)

            if item == "핵심 지표":
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.border = self.THIN_BORDER
            elif item:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.THIN_BORDER

        # 계정별 요약
        account_start = start_row + len(key_metrics) + 2
        ws.cell(row=account_start, column=1, value="계정/리전별 현황").font = Font(bold=True)

        headers = [
            "Account",
            "Region",
            "Total",
            "Unused",
            "Low Usage",
            "Normal",
            "Monthly Cost",
            "Monthly Waste",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=account_start + 1, column=col, value=header)
        self._apply_header_style(ws, account_start + 1)

        for i, stats in enumerate(self.stats_list):
            row = account_start + 2 + i
            ws.cell(row=row, column=1, value=stats.get("account_name", ""))
            ws.cell(row=row, column=2, value=stats.get("region", ""))
            ws.cell(row=row, column=3, value=stats.get("total_nat_count", 0))
            ws.cell(row=row, column=4, value=stats.get("unused_count", 0))
            ws.cell(row=row, column=5, value=stats.get("low_usage_count", 0))
            ws.cell(row=row, column=6, value=stats.get("normal_count", 0))
            ws.cell(row=row, column=7, value=f"${stats.get('total_monthly_cost', 0):,.2f}")
            ws.cell(row=row, column=8, value=f"${stats.get('total_monthly_waste', 0):,.2f}")

            # 미사용이 있으면 강조
            if stats.get("unused_count", 0) > 0:
                ws.cell(row=row, column=4).fill = self.STATUS_FILLS[UsageStatus.UNUSED]

        self._auto_column_width(ws)

    def _create_findings_sheet(self) -> None:
        """Findings 시트 - 조치 필요한 항목만"""
        ws = self.wb.create_sheet("Findings")

        headers = [
            "Account",
            "Region",
            "NAT Gateway ID",
            "Name",
            "Status",
            "Severity",
            "Confidence",
            "Description",
            "Monthly Waste",
            "Annual Savings",
            "Recommendation",
            "VPC ID",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        # 심각도 순으로 정렬
        severity_order = {
            Severity.CRITICAL: 0,
            Severity.HIGH: 1,
            Severity.MEDIUM: 2,
            Severity.LOW: 3,
            Severity.INFO: 4,
        }

        all_findings = []
        for result in self.results:
            for finding in result.findings:
                # 정상 사용은 제외
                if finding.usage_status != UsageStatus.NORMAL:
                    all_findings.append(finding)

        sorted_findings = sorted(all_findings, key=lambda x: severity_order.get(x.severity, 5))

        for finding in sorted_findings:
            nat = finding.nat
            row = [
                nat.account_name,
                nat.region,
                nat.nat_gateway_id,
                nat.name,
                finding.usage_status.value,
                finding.severity.value.upper(),
                finding.confidence.value,
                finding.description,
                f"${finding.monthly_waste:,.2f}",
                f"${finding.annual_savings:,.2f}",
                finding.recommendation,
                nat.vpc_id,
            ]
            ws.append(row)

            # 상태/심각도 색상
            current_row = ws.max_row
            status_fill = self.STATUS_FILLS.get(finding.usage_status)
            if status_fill:
                ws.cell(row=current_row, column=5).fill = status_fill

            severity_fill = self.SEVERITY_FILLS.get(finding.severity)
            if severity_fill:
                ws.cell(row=current_row, column=6).fill = severity_fill

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"

    def _create_all_nat_sheet(self) -> None:
        """All NAT Gateways 시트 - 전체 목록"""
        ws = self.wb.create_sheet("All NAT Gateways")

        headers = [
            "Account",
            "Region",
            "NAT Gateway ID",
            "Name",
            "VPC ID",
            "Subnet ID",
            "State",
            "Public IP",
            "Type",
            "Age (Days)",
            "Bytes Out (14d)",
            "Days with Traffic",
            "Monthly Cost",
            "Usage Status",
            "Tags",
        ]
        ws.append(headers)
        self._apply_header_style(ws)

        for result in self.results:
            for finding in result.findings:
                nat = finding.nat

                # 바이트를 읽기 쉽게 변환
                bytes_out = nat.bytes_out_total
                if bytes_out >= 1024**3:
                    bytes_str = f"{bytes_out / (1024**3):.2f} GB"
                elif bytes_out >= 1024**2:
                    bytes_str = f"{bytes_out / (1024**2):.2f} MB"
                elif bytes_out >= 1024:
                    bytes_str = f"{bytes_out / 1024:.2f} KB"
                else:
                    bytes_str = f"{bytes_out:.0f} B"

                # 태그 문자열
                tags_str = ", ".join(f"{k}={v}" for k, v in nat.tags.items() if k != "Name")

                row = [
                    nat.account_name,
                    nat.region,
                    nat.nat_gateway_id,
                    nat.name,
                    nat.vpc_id,
                    nat.subnet_id,
                    nat.state,
                    nat.public_ip,
                    nat.connectivity_type,
                    nat.age_days,
                    bytes_str,
                    nat.days_with_traffic,
                    f"${nat.total_monthly_cost:,.2f}",
                    finding.usage_status.value,
                    tags_str[:100],  # 최대 100자
                ]
                ws.append(row)

                # 상태 색상
                current_row = ws.max_row
                status_fill = self.STATUS_FILLS.get(finding.usage_status)
                if status_fill:
                    ws.cell(row=current_row, column=14).fill = status_fill

        self._auto_column_width(ws)
        ws.freeze_panes = "A2"
